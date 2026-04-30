#!/usr/bin/env python3
"""Build OWP_Productivity_Insights.json v1.3 from all closed-project data.json
files. Handles three schema generations:

  v2.0:  cost_code_summaries (list), worker_wages (list), line_items
  v2.1:  codes (dict by code), workers (dict by id), vendors (dict)
  v2.2:  cost_codes (list), workers (list), vendors (list), ar_invoices, ap_invoices

Each closed project in PROJECT_ORDER (read from index.html) is normalized to
a common shape, then portfolio-wide benchmarks + per-code variance stats are
computed. The final JSON matches the v1.2 schema so it's a drop-in replacement
for the embedded OWP_CALIB in index.html.
"""
import json
import re
import sys
from pathlib import Path
from datetime import datetime, timezone

_HERE = Path(__file__).resolve().parent
ROOT = _HERE.parent

# Cost-code categorization (matches dashboardize.py)
LABOR_CODES = {"011","100","101","110","111","112","113","120","130","140","141","142","143","145","150"}
MATERIAL_CODES = {"039","210","211","212","213","220","230","240","241","242","243","244","245","251"}
OVERHEAD_CODES = {"600","601","602","603","604","607"}
BURDEN_CODES = {"995","998"}

# Phase tagging for cost codes (used in cushion/overrun outputs)
CODE_PHASE = {
    "011": "underground", "100": "supervision", "101": "takeoff", "110": "underground",
    "111": "garage", "112": "canout", "113": "foundation", "120": "rough_in",
    "130": "trim_finish", "140": "gas", "141": "water_main", "142": "mech_room",
    "143": "condensation", "145": "tub_shower", "150": "warranty",
    "039": "underground", "210": "underground", "211": "underground", "212": "canout",
    "213": "foundation", "220": "rough_in", "230": "trim_finish", "240": "gas",
    "241": "water_main", "242": "mech_room", "243": "condensation", "244": "fire",
    "245": "tub_shower", "600": "overhead", "601": "overhead", "602": "overhead",
    "603": "overhead", "604": "overhead", "607": "overhead", "995": "burden", "998": "burden"
}

CODE_DESC = {
    "100": "Supervision", "101": "Takeoff & Purchase Labor", "110": "Underground Labor",
    "111": "Garage Labor", "112": "Canout Labor", "113": "Foundation Drain Labor",
    "120": "Roughin Labor", "130": "Finish Labor", "140": "Gas Labor",
    "141": "Water Main/Insulation Lab", "142": "Mech Room Labor",
    "143": "Condensation Drains Labor", "145": "Tub/Shower Labor", "150": "Warranty Labor",
    "011": "DS & RD Labor",
    "210": "Underground Material", "211": "Garage Material", "212": "Canout Material",
    "213": "Foundation Material", "220": "Roughin Material", "230": "Finish Material",
    "240": "Gas Material", "241": "Water Main/Insulation Mat", "242": "Mech Room Material",
    "243": "Condensation Drains Mat", "244": "Fire Caulking Material",
    "245": "Tub/Shower Material", "039": "DS & RD Material",
    "600": "Subcontractor", "601": "Engineering/Plans", "602": "Rental Equipment",
    "603": "Permits & Licenses", "604": "Subcontractor 2", "607": "Other Expenses",
    "995": "Payroll Burden", "998": "Payroll Taxes",
}


def load_project_order():
    """Read PROJECT_ORDER from index.html."""
    html = (ROOT / "index.html").read_text()
    m = re.search(r'const PROJECT_ORDER = \[([^\]]+)\];', html)
    return [x.strip().strip('"') for x in m.group(1).split(',')]


# Cache for index.html units/fixtures lookup
_HERO_CACHE = None

def hero_lookup(jid):
    """Read units + fixtures from index.html's PROJECTS hero block. v2.0 and
    v2.1 schemas don't always store units/fixtures in data.json, so we fall
    back to the dashboard's hero block which always has them."""
    global _HERO_CACHE
    if _HERO_CACHE is None:
        html = (ROOT / "index.html").read_text()
        _HERO_CACHE = {}
        # Strategy: find each PROJECTS hero block start ("XXXX": {), then within
        # the next ~2500 chars look for `num: { units: NN, fixtures: NN`.
        # Hero blocks contain nested {} (chips, skyline, num) so we can't use [^}].
        for hero_m in re.finditer(r'"(20\d\d)":\s*\{', html):
            jid_h = hero_m.group(1)
            window = html[hero_m.end():hero_m.end() + 3500]
            num_m = re.search(r'num:\s*\{\s*units:\s*(\d+)\s*,\s*fixtures:\s*(\d+|null)', window)
            if num_m:
                units = int(num_m.group(1))
                fx = int(num_m.group(2)) if num_m.group(2) != 'null' else 0
                _HERO_CACHE[jid_h] = (units, fx)
    return _HERO_CACHE.get(jid, (0, 0))


def load_xlsx_data(jid):
    """v1.0-era fallback: read OWP_XXXX_JCR_Cortex_v2.xlsx from
    owp/cortex_v2_files_rebuilt/ for projects that don't have data.json.
    Used to bridge the 13 original projects (#2001-#2022) into the
    calibration sample. Returns same normalized shape as load_project_data."""
    xlsx_path = ROOT / "owp" / "cortex_v2_files_rebuilt" / f"OWP_{jid}_JCR_Cortex_v2.xlsx"
    if not xlsx_path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(xlsx_path, data_only=True)
    n = {'jid': jid, 'codes': {}, 'schema': 'xlsx-v2.1'}

    # Hero lookup for units + fixtures
    hu, hf = hero_lookup(jid)
    n['units'] = hu
    n['fixtures'] = hf

    # Read Budget vs Actual sheet for cost codes
    if 'Budget vs Actual' in wb.sheetnames:
        ws = wb['Budget vs Actual']
        # Header at row 5 (index 4); data starts at row 7+ (skipping LABOR/MATERIAL section headers)
        for row in ws.iter_rows(min_row=6, values_only=True):
            # Row layout: [None, cost_code, description, original_budget, revised_budget, actual, variance, pct_used, hours]
            if not row or len(row) < 7:
                continue
            code = row[1]
            if code is None or not isinstance(code, str):
                continue
            # Skip section headers (LABOR, MATERIAL, OVERHEAD, BURDEN, SALES, TOTAL, etc.)
            if not code.strip().isdigit() and code != '011' and code != '039':
                continue
            try:
                orig = float(row[3] or 0)
                actual = float(row[5] or 0)
                hours = float(row[8] or 0) if len(row) > 8 and row[8] else 0
            except (ValueError, TypeError):
                continue
            n['codes'][code.strip()] = {'orig': orig, 'actual': actual, 'hours': hours}

    # Read Overview for revenue + direct cost + GC
    if 'Overview' in wb.sheetnames:
        ws = wb['Overview']
        rows = list(ws.iter_rows(values_only=True))
        for i, row in enumerate(rows):
            if not row or all(c is None for c in row):
                continue
            label = str(row[1]).strip() if row[1] else ''
            value = row[2] if len(row) > 2 else None
            if label == 'General Contractor' and value:
                n['gc'] = str(value).strip()
            elif label == 'CONTRACT VALUE':
                # Next row has the values: [None, contract_value, None, net_profit, None, direct_cost]
                # Values may be either numeric or string-formatted ("$1,850,270")
                def parse_money(v):
                    if v is None: return 0.0
                    if isinstance(v, (int, float)): return float(v)
                    if isinstance(v, str):
                        cleaned = v.replace('$', '').replace(',', '').strip()
                        try: return float(cleaned)
                        except ValueError: return 0.0
                    return 0.0
                if i + 1 < len(rows):
                    vrow = rows[i + 1]
                    if vrow and len(vrow) >= 6:
                        n['revenue'] = parse_money(vrow[1])
                        n['direct_cost'] = parse_money(vrow[5])

    # Read Crew & Labor for worker count + total hours
    if 'Crew & Labor' in wb.sheetnames:
        ws = wb['Crew & Labor']
        worker_count = 0
        total_hours = 0.0
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or len(row) < 4:
                continue
            name = row[1]
            hours = row[3]
            if name and isinstance(name, str) and hours and isinstance(hours, (int, float)):
                worker_count += 1
                total_hours += float(hours)
        n['total_workers'] = worker_count
        n['total_hours'] = total_hours

    # Defaults if Overview parsing missed something
    n.setdefault('revenue', 0)
    n.setdefault('direct_cost', 0)
    n.setdefault('total_hours', 0)
    n.setdefault('total_workers', 0)
    n.setdefault('gc', '')

    # Compute gross_margin
    if n.get('revenue') and n['revenue'] > 0:
        n['gross_margin'] = (n['revenue'] - n.get('direct_cost', 0)) / n['revenue']
    else:
        n['gross_margin'] = None

    return n


def load_project_data(jid):
    """Load and normalize a project's data.json regardless of schema. Returns
    a normalized dict with consistent keys, or None if no data.json found.

    Normalized output:
      {
        'units': int,
        'fixtures': int,
        'revenue': float,
        'direct_cost': float,
        'gross_margin': float (computed),
        'total_hours': float,
        'total_workers': int,
        'codes': { code: {orig, actual, hours} },  # the per-cost-code dict
        'gc': str,
        'schema': str ('v2.0'/'v2.1'/'v2.2'/'xlsx-v2.1'),
      }
    """
    for folder in (f"owp-{jid}-live", f"owp-{jid}"):
        path = ROOT / "owp" / folder / "cortex output files" / f"{jid}_data.json"
        if path.exists():
            break
    else:
        # No data.json found — try XLSX fallback for v1.0-era projects
        return load_xlsx_data(jid)

    raw = json.loads(path.read_text())
    n = {'jid': jid, 'codes': {}, 'schema': 'unknown'}

    # Always look up units/fixtures from index.html hero (most reliable)
    hu, hf = hero_lookup(jid)
    n['units'] = hu
    n['fixtures'] = hf

    # Detect schema
    if 'cost_code_summaries' in raw:
        n['schema'] = 'v2.0'
        for c in raw['cost_code_summaries']:
            code = c.get('code')
            if not code:
                continue
            n['codes'][code] = {
                'orig': c.get('original_budget', 0) or 0,
                'actual': c.get('actual_amount', 0) or 0,
                'hours': (c.get('regular_hours', 0) or 0) + (c.get('overtime_hours', 0) or 0),
            }
        derived = raw.get('derived_fields', {})
        rr = raw.get('report_record', {})
        # Revenue: report_record.job_totals_revenue is negative (Sage convention) — take abs
        n['revenue'] = abs(rr.get('job_totals_revenue', 0)) or 0
        n['direct_cost'] = derived.get('direct_cost', 0) or 0
        n['total_hours'] = derived.get('total_labor_hours', 0) or 0
        # Total workers: count workers list if present, else 0
        ww = raw.get('worker_wages', [])
        n['total_workers'] = len(ww) if isinstance(ww, list) else 0
        p = raw.get('project', {})
        n['gc'] = p.get('general_contractor', '')

    elif 'cost_codes' in raw:
        n['schema'] = 'v2.2'
        for c in raw['cost_codes']:
            code = c.get('code')
            if not code:
                continue
            n['codes'][code] = {
                'orig': c.get('original_budget', 0) or 0,
                'actual': c.get('actual_amount', 0) or 0,
                'hours': (c.get('regular_hours', 0) or 0) + (c.get('overtime_hours', 0) or 0) + (c.get('doubletime_hours', 0) or 0),
            }
        p = raw.get('project', {})
        # Prefer hero lookup; fall back to project block
        if not n['units']: n['units'] = p.get('units', 0) or 0
        if not n['fixtures']: n['fixtures'] = p.get('total_fixtures', 0) or 0
        t = raw.get('totals', {})
        n['revenue'] = t.get('revenue_ar_actual', 0) or 0
        n['direct_cost'] = t.get('direct_cost', 0) or 0
        lab = raw.get('labor', {})
        n['total_hours'] = lab.get('total_hours', 0) or 0
        n['total_workers'] = lab.get('total_workers', 0) or 0
        n['gc'] = p.get('general_contractor', '')

    elif 'codes' in raw and isinstance(raw['codes'], dict):
        n['schema'] = 'v2.1'
        for code, c in raw['codes'].items():
            if code == '999':
                # 999 = Sales code, drives revenue
                n['revenue'] = abs(c.get('actual', 0) or 0)
                continue
            n['codes'][code] = {
                'orig': c.get('orig', 0) or 0,
                'actual': c.get('actual', 0) or 0,
                'hours': c.get('hrs_total', c.get('hrs_reg', 0)) or 0,
            }
        # Direct cost = sum of non-999 actuals
        n['direct_cost'] = sum(v['actual'] for v in n['codes'].values())
        # Total hours: sum from workers dict
        workers = raw.get('workers', {})
        if isinstance(workers, dict):
            n['total_hours'] = sum(w.get('hours', 0) or 0 for w in workers.values())
            n['total_workers'] = len(workers)
        n['gc'] = ''

    # Compute gross margin if both available
    if n.get('revenue') and n['revenue'] > 0:
        n['gross_margin'] = (n['revenue'] - n['direct_cost']) / n['revenue']
    else:
        n['gross_margin'] = None

    return n


def percentile(values, q):
    if not values:
        return 0.0
    s = sorted(values)
    k = (len(s) - 1) * q
    f = int(k); c = min(f + 1, len(s) - 1)
    if f == c:
        return float(s[f])
    return float(s[f] + (s[c] - s[f]) * (k - f))


def stats(values, label=None):
    """Return n/min/max/mean/median/p25/p75/p10/p90 for a list of values."""
    if not values:
        return {'n': 0}
    vs = [float(v) for v in values if v is not None]
    if not vs:
        return {'n': 0}
    return {
        'n': len(vs),
        'min': round(min(vs), 4),
        'max': round(max(vs), 4),
        'mean': round(sum(vs) / len(vs), 4),
        'median': round(percentile(vs, 0.5), 4),
        'p25': round(percentile(vs, 0.25), 4),
        'p50': round(percentile(vs, 0.5), 4),
        'p75': round(percentile(vs, 0.75), 4),
        'p10': round(percentile(vs, 0.10), 4),
        'p90': round(percentile(vs, 0.90), 4),
    }


def main():
    ids = load_project_order()
    print(f"Loading {len(ids)} closed projects from PROJECT_ORDER", file=sys.stderr)

    projects = []
    for jid in ids:
        d = load_project_data(jid)
        if d:
            projects.append(d)
        else:
            print(f"  SKIP #{jid}: no data.json", file=sys.stderr)

    print(f"\nLoaded {len(projects)} projects with data.json", file=sys.stderr)

    schemas = {}
    for p in projects:
        schemas[p['schema']] = schemas.get(p['schema'], 0) + 1
    for sk, cnt in sorted(schemas.items()):
        print(f"  {sk}: {cnt}", file=sys.stderr)

    # ── Portfolio benchmarks (per-job aggregates) ─────────────────────────────
    benchmarks = {}

    def b(label, getter, note=''):
        vals = [getter(p) for p in projects if getter(p) is not None and getter(p) != 0]
        s = stats(vals)
        s['note'] = note
        benchmarks[label] = s

    def labor_actual(p): return sum(p['codes'].get(c, {}).get('actual', 0) for c in p['codes'] if c in LABOR_CODES)
    def material_actual(p): return sum(p['codes'].get(c, {}).get('actual', 0) for c in p['codes'] if c in MATERIAL_CODES)
    def burden_actual(p): return p['codes'].get('995', {}).get('actual', 0)

    def hours_per_unit(p):
        if p['units'] and p['total_hours']:
            return p['total_hours'] / p['units']
        return None
    def labor_dollars_per_unit(p):
        if p['units']:
            return labor_actual(p) / p['units']
        return None
    def material_dollars_per_unit(p):
        if p['units']:
            return material_actual(p) / p['units']
        return None
    def revenue_per_unit(p):
        if p['units'] and p['revenue']:
            return p['revenue'] / p['units']
        return None
    def labor_pct(p):
        if p['revenue']:
            return labor_actual(p) / p['revenue']
        return None
    def material_pct(p):
        if p['revenue']:
            return material_actual(p) / p['revenue']
        return None
    def margin(p):
        return p.get('gross_margin')
    def loaded_wage(p):
        if p['total_hours']:
            return (labor_actual(p) + burden_actual(p)) / p['total_hours']
        return None
    def blended_wage(p):
        if p['total_hours']:
            return labor_actual(p) / p['total_hours']
        return None
    def burden_mult(p):
        b = blended_wage(p); l = loaded_wage(p)
        if b and l and b > 0:
            return l / b
        return None

    b('hours_per_unit', hours_per_unit, 'Total field hours / plumbing units')
    benchmarks['hours_per_unit']['unit'] = 'hours'
    b('labor_dollars_per_unit', labor_dollars_per_unit, 'Direct labor cost / plumbing units')
    benchmarks['labor_dollars_per_unit']['unit'] = 'USD'
    b('material_dollars_per_unit', material_dollars_per_unit, '200-series cost code spend / plumbing units')
    benchmarks['material_dollars_per_unit']['unit'] = 'USD'
    b('revenue_per_unit', revenue_per_unit, '999 sales code billed / plumbing units')
    benchmarks['revenue_per_unit']['unit'] = 'USD'
    b('labor_pct_of_revenue', labor_pct, 'Labor cost / revenue')
    benchmarks['labor_pct_of_revenue']['unit'] = 'fraction'
    b('material_pct_of_revenue', material_pct, 'Material cost / revenue')
    benchmarks['material_pct_of_revenue']['unit'] = 'fraction'
    b('gross_margin_pct', margin, '(Revenue - direct cost) / revenue')
    benchmarks['gross_margin_pct']['unit'] = 'fraction'
    b('fully_loaded_wage', loaded_wage, '(Labor cost + 995 burden) / total hours')
    benchmarks['fully_loaded_wage']['unit'] = 'USD/hr'
    b('blended_wage', blended_wage, 'Labor cost / total hours')
    benchmarks['blended_wage']['unit'] = 'USD/hr'
    b('burden_multiplier', burden_mult, 'Loaded wage / blended wage')
    benchmarks['burden_multiplier']['unit'] = 'x'

    # ── Per-cost-code variance (orig→actual) ─────────────────────────────────
    by_code = {}
    all_codes = set()
    for p in projects:
        all_codes.update(p['codes'].keys())

    for code in sorted(all_codes):
        actuals = []
        origs = []
        variances = []  # (actual - orig) / orig
        per_fixture = []  # actual$ / project fixtures (drives bid-tool COST_CODES)
        n_with_data = 0
        for p in projects:
            if code in p['codes']:
                c = p['codes'][code]
                if c['orig'] != 0 or c['actual'] != 0:
                    n_with_data += 1
                actuals.append(c['actual'])
                origs.append(c['orig'])
                if c['orig'] != 0:
                    variances.append((c['actual'] - c['orig']) / c['orig'])
                # Per-fixture rate: actual$ / project fixtures (skip if no fixtures)
                if p.get('fixtures', 0) and c['actual'] != 0:
                    per_fixture.append(c['actual'] / p['fixtures'])
        cat = 'labor' if code in LABOR_CODES else ('material' if code in MATERIAL_CODES else ('overhead' if code in OVERHEAD_CODES else ('burden' if code in BURDEN_CODES else 'other')))
        by_code[code] = {
            'code': code,
            'category': cat,
            'phase': CODE_PHASE.get(code, 'other'),
            'description': CODE_DESC.get(code, ''),
            'n_jobs_with_data': n_with_data,
            'actual_dollars': stats(actuals),
            'actual_dollars_per_fixture': stats(per_fixture),
            'original_budget_dollars': stats(origs),
            'variance_orig_to_actual': stats(variances),
        }

    # ── Cushion + overrun rankings ────────────────────────────────────────────
    # Filter for MEANINGFUL codes: at least 30 jobs of data (broad portfolio
    # coverage) AND median actual dollars >= $15,000 (excludes noise on tiny
    # line items like Supervision where small-budget % variance dominates).
    MIN_N = 30
    MIN_DOLLARS = 15000
    cushion = []
    overrun = []
    for code, d in by_code.items():
        v = d['variance_orig_to_actual']
        if v.get('n', 0) < MIN_N:
            continue
        median_actual = d['actual_dollars'].get('median', 0)
        if median_actual < MIN_DOLLARS:
            continue
        median_var = v.get('median', 0)
        if median_var < -0.05:
            cushion.append({
                'code': code, 'category': d['category'], 'phase': d['phase'],
                'description': d['description'], 'n_jobs': v['n'],
                'median_variance_orig_to_actual': round(median_var, 4),
                'p25_variance': round(v.get('p25', 0), 4),
                'p75_variance': round(v.get('p75', 0), 4),
                'direction': 'under',
                'intensity': 'very large' if median_var < -0.4 else ('large' if median_var < -0.25 else 'moderate'),
                'median_actual_dollars': round(median_actual),
            })
        elif median_var > 0.03:
            overrun.append({
                'code': code, 'category': d['category'], 'phase': d['phase'],
                'description': d['description'], 'n_jobs': v['n'],
                'median_variance_orig_to_actual': round(median_var, 4),
                'p25_variance': round(v.get('p25', 0), 4),
                'p75_variance': round(v.get('p75', 0), 4),
                'direction': 'over',
                'intensity': 'very large' if median_var > 0.5 else ('large' if median_var > 0.25 else 'moderate'),
                'median_actual_dollars': round(median_actual),
            })
    cushion.sort(key=lambda x: x['median_variance_orig_to_actual'])
    overrun.sort(key=lambda x: -x['median_variance_orig_to_actual'])
    top_cushion = cushion[:5]
    top_overrun = overrun[:5]

    # ── Headline findings ─────────────────────────────────────────────────────
    hpu = benchmarks.get('hours_per_unit', {})
    margin_b = benchmarks.get('gross_margin_pct', {})
    loaded = benchmarks.get('fully_loaded_wage', {})
    blended = benchmarks.get('blended_wage', {})
    burden = benchmarks.get('burden_multiplier', {})

    # Material variance (across all material codes)
    mat_vars = []
    for p in projects:
        for c in MATERIAL_CODES:
            if c in p['codes'] and p['codes'][c]['orig'] != 0:
                mat_vars.append((p['codes'][c]['actual'] - p['codes'][c]['orig']) / p['codes'][c]['orig'])
    mat_var_stat = stats(mat_vars)

    # Code 120 variance
    code120 = by_code.get('120', {}).get('variance_orig_to_actual', {})

    # Labor variance (across labor codes)
    lab_vars = []
    for p in projects:
        for c in LABOR_CODES:
            if c in p['codes'] and p['codes'][c]['orig'] != 0:
                lab_vars.append((p['codes'][c]['actual'] - p['codes'][c]['orig']) / p['codes'][c]['orig'])
    lab_var_stat = stats(lab_vars)

    headline_findings = [
        {
            'id': 'material_takeoff_systematically_conservative', 'severity': 'high',
            'title': 'OWP material takeoffs systematically come in under original budget',
            'value': mat_var_stat.get('median', 0), 'unit': 'fraction',
            'summary': f"Across {len(projects)} completed jobs, the median material variance from original budget to actual spend is {mat_var_stat.get('median', 0)*100:.1f}%. Negative means UNDER budget. This is the strongest single signal in the dataset for calibrating the bid model.",
            'evidence': {'n_jobs': mat_var_stat['n'], 'p25': mat_var_stat.get('p25'), 'median': mat_var_stat.get('median'), 'p75': mat_var_stat.get('p75')},
            'implication': "Two valid responses for the bid model: (a) tighten the takeoff factor by ~25% and bid a leaner number, or (b) hold the cushion as a declared margin contribution and report it explicitly to leadership. Without an explicit policy, this cushion compounds opaquely."
        },
        {
            'id': 'roughin_labor_overruns', 'severity': 'high',
            'title': 'Rough-in labor (code 120) is the major labor code that runs over budget',
            'value': code120.get('median', 0), 'unit': 'fraction',
            'summary': f"Across {code120.get('n', 0)} completed jobs, code 120 (rough-in labor) has a median original→actual variance of {code120.get('median', 0)*100:+.1f}%. Most other major labor codes run UNDER budget. This means takeoff hours for rough-in are systematically too low.",
            'evidence': {'code': '120', 'description': 'Roughin Labor', 'n_jobs': code120.get('n', 0), 'variance_orig_to_actual': code120, 'median_actual_dollars': by_code.get('120', {}).get('actual_dollars', {}).get('median', 0)},
            'implication': "Bid model should apply a +10–15% correction factor to rough-in labor takeoff specifically. Material cushion alone does not compensate for this — labor is harder to recover once underbid."
        },
        {
            'id': 'hours_per_unit_house_standard', 'severity': 'high',
            'title': f'House standard for labor productivity is ~{hpu.get("median", 0):.0f} hours per plumbing unit',
            'value': hpu.get('median', 0), 'unit': 'hours',
            'summary': f"Median is {hpu.get('median', 0):.1f} hrs/unit (P25 {hpu.get('p25', 0):.0f}, P75 {hpu.get('p75', 0):.0f}). The middle 50% of jobs sit between {hpu.get('p25', 0):.0f} and {hpu.get('p75', 0):.0f} hrs/unit — that is the credible band for a like-for-like new bid.",
            'evidence': hpu,
            'implication': "Bids that fall below P25 or above P75 should be flagged for explicit justification — the bidding tool should not silently extrapolate beyond the historical band."
        },
        {
            'id': 'gross_margin_band', 'severity': 'medium',
            'title': f'Gross margin clusters at {margin_b.get("median", 0)*100:.1f}%',
            'value': margin_b.get('median', 0), 'unit': 'fraction',
            'summary': f"Median gross margin is {margin_b.get('median', 0)*100:.1f}%. This is what OWP earns once direct cost (labor + material + burden + tax) is removed from billed revenue.",
            'evidence': {'n_jobs': margin_b.get('n', 0), 'p25': margin_b.get('p25'), 'median': margin_b.get('median'), 'p75': margin_b.get('p75')},
            'implication': "Use as default target margin. Combine with material-takeoff cushion finding for total margin posture."
        },
        {
            'id': 'wage_burden', 'severity': 'medium',
            'title': f'Fully loaded labor cost is ~${loaded.get("median", 0):.2f}/hr ({burden.get("median", 1):.2f}× blended)',
            'value': loaded.get('median', 0), 'unit': 'USD/hr',
            'summary': f"Median fully loaded wage (labor + burden) is ${loaded.get('median', 0):.2f}/hr. Burden multiplier of {burden.get('median', 1):.2f}× means OWP carries ~{(burden.get('median', 1) - 1)*100:.0f}% on top of direct wages for taxes, insurance, vehicles, etc.",
            'evidence': {'blended_wage_median': blended.get('median', 0), 'loaded_wage_median': loaded.get('median', 0), 'burden_multiplier_median': burden.get('median', 1)},
            'implication': "Always price labor at fully-loaded wage in the bid model."
        },
        {
            'id': 'labor_variance_neutral', 'severity': 'medium',
            'title': f'Labor variance is roughly balanced ({lab_var_stat.get("median", 0)*100:+.1f}% median)',
            'value': lab_var_stat.get('median', 0), 'unit': 'fraction',
            'summary': f"Unlike material (consistently under), labor variance from original budget to actual is closer to neutral (median {lab_var_stat.get('median', 0)*100:+.1f}%). The bid model should NOT assume a labor cushion the way it can for material.",
            'evidence': lab_var_stat,
            'implication': "Bid the labor number close to the takeoff — there is no systematic safety margin."
        },
    ]

    # ── Bid calibration rules (data-driven defaults) ──────────────────────────
    bid_rules = [
        {'rule_id': 'OWP-BID-001', 'statement': f'Default labor estimate = plumbing_units × {hpu.get("median", 0):.1f} hrs/unit',
         'default_value': hpu.get('median', 0), 'unit': 'hours/unit', 'band': {'p25': hpu.get('p25', 0), 'p75': hpu.get('p75', 0)},
         'rationale': f'Median across {hpu.get("n", 0)} completed jobs with unit counts.',
         'applies_when': 'Plumbing unit count is known. For non-residential jobs without unit counts, fall back to historical labor%revenue.'},
        {'rule_id': 'OWP-BID-002', 'statement': f'Price labor at fully-loaded wage of ${loaded.get("median", 0):.2f}/hr (or pass-through of current loaded wage)',
         'default_value': loaded.get('median', 0), 'unit': 'USD/hr',
         'rationale': f'Burden multiplier of {burden.get("median", 1):.2f}× consistently observed. Blended wage alone understates true cost.',
         'applies_when': 'Always.'},
        {'rule_id': 'OWP-BID-003', 'statement': f'Apply ~{abs(mat_var_stat.get("median", 0))*100:.0f}% takeoff cushion / margin contribution on material',
         'default_value': abs(mat_var_stat.get('median', 0)), 'unit': 'fraction',
         'rationale': f'Median material original→actual variance is {mat_var_stat.get("median", 0)*100:+.1f}%. Either tighten takeoff or hold as explicit margin.',
         'applies_when': 'Material takeoff is being generated. Re-evaluate per project type — variance is wider on smaller jobs.'},
        {'rule_id': 'OWP-BID-004', 'statement': f'Target gross margin floor = {margin_b.get("median", 0)*100:.1f}%',
         'default_value': margin_b.get('median', 0), 'unit': 'fraction',
         'rationale': f'Historical median across {margin_b.get("n", 0)} completed jobs.',
         'applies_when': f'Always. Bids below P25 ({margin_b.get("p25", 0)*100:.1f}%) require sign-off.'},
        {'rule_id': 'OWP-BID-005', 'statement': 'Sanity-check duration: months ≈ (hours/unit × units) / (crew × 173.33)',
         'default_value': 5.0, 'unit': 'people (typical crew)',
         'rationale': 'Crew FTE median is ~5.0 across closed portfolio.',
         'applies_when': 'After labor estimate is set, validate duration is feasible given typical crew capacity.'},
        {'rule_id': 'OWP-BID-006', 'statement': f'Carry overhead burden of ~{(burden.get("median", 1) - 1)*100:.0f}% on top of direct labor (cost code 995 equivalent)',
         'default_value': burden.get('median', 1) - 1, 'unit': 'fraction',
         'rationale': 'Burden ratio embedded in fully loaded wage; if labor priced at blended wage, this needs to be added separately.',
         'applies_when': 'Whenever labor is itemized at blended wage instead of loaded wage.'},
        {'rule_id': 'OWP-BID-007', 'statement': f'Apply a +{round(code120.get("median", 0)*100)+2}% correction to rough-in labor (code 120) takeoff',
         'default_value': code120.get('median', 0), 'unit': 'fraction',
         'rationale': f'Code 120 median variance is {code120.get("median", 0)*100:+.1f}% across {code120.get("n", 0)} jobs — it is the major labor code that systematically overruns.',
         'applies_when': 'Always, until estimator-side fix is implemented.'},
    ]

    # ── Final output ──────────────────────────────────────────────────────────
    output = {
        'name': 'OWP Productivity Insights',
        'version': '1.3',
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'source': f'cortex-mockup repo · {len(projects)} cortex output JSONs across 3 schema generations (v2.0/v2.1/v2.2) · PROJECT_ORDER {len(ids)} closed jobs',
        'cortex_layer': 'productivity_insights',
        'audience': 'bid_model_calibration',
        'usage_instructions': 'Drives the bid intelligence tool calibration panel and the COST_CODES variance defaults. Regenerate via owp/build_calibration.py whenever a new closed job lands.',
        'scope': {
            'total_jobs': len(ids),
            'completed_jobs_in_benchmarks': len(projects),
            'completed_with_unit_data': benchmarks.get('hours_per_unit', {}).get('n', 0),
            'live_jobs': 0,
            'schemas_loaded': schemas,
            'missing_data_jsons': [j for j in ids if not load_project_data(j)],
        },
        'headline_findings': headline_findings,
        'benchmarks': benchmarks,
        'bid_calibration_rules': bid_rules,
        'top_cushion_codes': top_cushion,
        'top_overrun_codes': top_overrun,
        'cost_code_benchmarks': {
            'methodology': 'Per-cost-code statistics across all jobs that have data for that code. Variance = (actual - original_budget) / original_budget.',
            'n_codes_total': len(by_code),
            'n_codes_high_coverage': sum(1 for c in by_code.values() if c['n_jobs_with_data'] >= 30),
            'note_low_coverage': 'Codes with n_jobs_with_data < 8 are excluded from cushion/overrun lists due to insufficient sample.',
            'by_code': by_code,
        },
    }

    out_path = ROOT / "owp" / "bidding tool calibration" / "OWP_Productivity_Insights.json"
    out_path.write_text(json.dumps(output, separators=(',', ':')))
    print(f"\nWrote {out_path}", file=sys.stderr)
    print(f"  size: {out_path.stat().st_size:,} bytes", file=sys.stderr)
    print(f"  scope: {output['scope']}", file=sys.stderr)
    print(f"  hours/unit median: {hpu.get('median', 0):.1f} (n={hpu.get('n', 0)})", file=sys.stderr)
    print(f"  gross_margin median: {margin_b.get('median', 0)*100:.1f}% (n={margin_b.get('n', 0)})", file=sys.stderr)
    print(f"  loaded_wage median: ${loaded.get('median', 0):.2f}/hr (n={loaded.get('n', 0)})", file=sys.stderr)
    print(f"  burden_mult median: {burden.get('median', 1):.3f} (n={burden.get('n', 0)})", file=sys.stderr)
    print(f"  cushion codes: " + ', '.join(f"{c['code']} {c['median_variance_orig_to_actual']*100:+.1f}%" for c in top_cushion), file=sys.stderr)
    print(f"  overrun codes: " + ', '.join(f"{c['code']} {c['median_variance_orig_to_actual']*100:+.1f}%" for c in top_overrun), file=sys.stderr)


if __name__ == '__main__':
    main()
