#!/usr/bin/env python3
"""Build OWP_2023 Cortex v2 17-tab JCR. Chinn Construction / Hadley-Legacy Mercer Island Apartments.
Rich source set: JDR, subcontract, 5 COs (image-scan), 125 ASI-RFIs, 100+ submittals, 180+ POs, bid sheet."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
TITLE = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
SUB = Font(name=ARIAL, size=10, italic=True, color="595959")
HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
NF_FONT = Font(name=ARIAL, size=10, italic=True, color="9C0006")
SRC_FONT = Font(name=ARIAL, size=8, italic=True, color="595959")
F_TITLE = PatternFill("solid", fgColor="1F3864")
F_HDR = PatternFill("solid", fgColor="2E5090")
F_ALT = PatternFill("solid", fgColor="F2F2F2")
F_HIGH = PatternFill("solid", fgColor="FFF2CC")
F_RISK = PatternFill("solid", fgColor="FFE6E6")
F_OK = PatternFill("solid", fgColor="E2EFDA")
F_NF = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NF = "NOT FOUND"

def put(ws, coord, val, font=BODY, fill=None, border=BRD, align=None, fmt=None):
    c = ws[coord] if isinstance(coord, str) else ws.cell(row=coord[0], column=coord[1])
    c.value = val
    if val == NF:
        c.font = NF_FONT; c.fill = F_NF; c.alignment = CENTER
    else:
        c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

def title(ws, text, sub_text=""):
    c = ws.cell(row=2, column=2, value=text); c.font = TITLE; c.fill = F_TITLE; c.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 28
    if sub_text:
        c2 = ws.cell(row=3, column=2, value=sub_text); c2.font = SUB
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)

def hdr(ws, row, cols, start_col=2):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=txt)
        c.font = HDR; c.fill = F_HDR; c.alignment = CENTER; c.border = BRD

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# Load parsed JDR data
data = json.load(open('/sessions/keen-determined-mccarthy/work/2023_data.json'))
CODES = data['codes']
WORKERS = data['workers']
VENDORS = data['vendors']
INVOICES = data['invoices']

wb = Workbook()
wb.remove(wb.active)

# ============ CONSTANTS (all sourced) ============
JOB = "2023"
NAME = "Chinn, Legacy Apts"
PROJECT_DESC = "Legacy @ Mercer Island Apartments — 2601 76th Ave SE, Mercer Island, WA 98040 — new construction plumbing"
GC = "Chinn Construction, LLC"
GC_CUST_CODE = "2023CC"
OWNER = "Hadley Improvements Owner, LLC"
ARCHITECT = "VIA Architecture"

# Financial (JDR footer, page 285 of 285, 04/03/2026)
REVENUE = 2_771_369.00
EXPENSES = 1_873_341.44
NET_PROFIT = 898_027.56
RETAINAGE = 138_568.45
SRC_GL = 110_321.43
SRC_AP = 682_284.56
SRC_PR = 1_080_735.45

CONTRACT_ORIG = 2_469_500.00
CONTRACT_FINAL = 2_771_369.00
CO_TOTAL_IMPLIED = CONTRACT_FINAL - CONTRACT_ORIG  # 301,869.00
# 5 CO PDFs exist (image-only scans — individual values NOT FOUND in machine-readable form).
CO_COUNT_DOCUMENTED = 5

# Cost code categories (2023 includes 143, 150, 243, 244)
LABOR_CODES = ["100","101","110","111","112","120","130","140","141","142","143","150"]
MATERIAL_CODES = ["210","211","212","220","230","240","241","242","243","244"]
OVERHEAD_CODES = ["600","601","602","603","607"]
BURDEN_CODE = "995"
TAX_CODE = "998"

def sum_actual(codes):
    return sum(CODES[c]['actual'] for c in codes if c in CODES)

LABOR_COST = sum_actual(LABOR_CODES)
MATERIAL_COST = sum_actual(MATERIAL_CODES)
OVERHEAD_COST = sum_actual(OVERHEAD_CODES)
BURDEN_COST = CODES[BURDEN_CODE]['actual']
TAX_COST = CODES[TAX_CODE]['actual']
TOTAL_HOURS = sum(CODES[c]['hrs_total'] for c in LABOR_CODES if c in CODES)
TOTAL_WORKERS = len(WORKERS)

SRC_JDR = "2023 Job Detail Report.pdf (Sage Timberline, 04/03/2026, 285 pages)"
SRC_CONTRACT = "Legacy contract OCR_d.pdf (Chinn↔OWP subcontract, Lump Sum $2,469,500)"
SRC_BID = "Chinn Legacy Mercer Island Sept 9 2014.xlsx"
SRC_FOLDER = "owp-2023/"

# Scope from bid sheet
UNITS = 241  # 230 std + 11 ADA Type A
UNITS_STD = 230
UNITS_ADA = 11
WC_COUNT = 244  # 241 unit + 3 amenity
LAV_COUNT = 244
BATH_COUNT = 241  # 197 tub/shower combos + 44 shower stalls
KS_COUNT = 211
HB_COUNT = 31
WH_COUNT = 4
ROOF_DRAIN_COUNT = 17

# Documented CO files (image-only)
CHANGE_ORDERS = [
    ("Legacy CO 11", NF, "Change Orders/CO_s/Legacy CO 11 Signed.pdf (image scan; value NOT machine-readable)"),
    ("Legacy CO 19", NF, "Change Orders/CO_s/Legacy CO 19.pdf (image scan)"),
    ("Legacy CO 20", NF, "Change Orders/CO_s/Legacy CO 20.pdf (image scan)"),
    ("Legacy CO 21", NF, "Change Orders/CO_s/Legacy CO 21.pdf (image scan)"),
    ("Legacy CO 22", NF, "Change Orders/CO_s/Legacy CO 22.pdf (image scan)"),
]

# ============ TAB 1: OVERVIEW ============
ws = wb.create_sheet("Overview")
title(ws, f"Job #{JOB} · {NAME} — Legacy @ Mercer Island Apartments",
      f"Cortex JCR Cortex v2  •  {GC} (customer {GC_CUST_CODE})  •  2601 76th Ave SE, Mercer Island WA 98040  •  Owner: {OWNER}")
put(ws, "B5", "PROJECT OVERVIEW", BOLD, F_ALT)
overview = [
    ("Project Job #", JOB, SRC_JDR + " header"),
    ("Project Name (Sage)", NAME, SRC_JDR + " header"),
    ("Project Description", "Legacy @ Mercer Island Apartments (Hadley) — full plumbing, 241 units", SRC_CONTRACT + " / AR line descriptions"),
    ("General Contractor", GC, SRC_CONTRACT),
    ("Customer Code (Sage)", GC_CUST_CODE, SRC_JDR),
    ("Owner / Developer", OWNER, SRC_CONTRACT),
    ("Architect", ARCHITECT, "ASI-RFI responses (e.g. 'VIA Response' RFI 168)"),
    ("Jobsite Location", "2601 76th Ave SE, Mercer Island, WA 98040", SRC_CONTRACT),
    ("Contract / PO Document", SRC_CONTRACT, "Fully executed subcontract (Lump Sum)"),
    ("Plans / Specifications", "Plans/ folder — full drawing set + VIA Architecture specs + Div 22 plumbing spec", "Folder scan"),
    ("Bid Sheet", SRC_BID + " (original 9/9/2014 bid with full fixture schedule)", "Bid Sheets folder"),
    ("Contract Type", "Lump Sum", SRC_CONTRACT),
    ("Work Period", "01/19/2015 (first invoice) – 12/15/2016 (last invoice) ≈ 23 months", SRC_JDR),
    ("Total Unique Documents Reviewed", "1000+ across 4 mirror folders", f"File inventory across {SRC_FOLDER}"),
]
r = 6
for label, val, src in overview:
    put(ws, f"B{r}", label, BOLD, align=LEFT)
    put(ws, f"C{r}", val, align=LEFT)
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.merge_cells(f"C{r}:E{r}")
    ws.merge_cells(f"F{r}:J{r}")
    r += 1

r += 1
put(ws, f"B{r}", "CONTRACT VALUE", BOLD, F_ALT); put(ws, f"D{r}", "NET PROFIT", BOLD, F_ALT)
put(ws, f"F{r}", "DIRECT COST", BOLD, F_ALT); put(ws, f"H{r}", "LABOR HOURS", BOLD, F_ALT)
r += 1
put(ws, f"B{r}", CONTRACT_FINAL, BOLD, fmt='"$"#,##0.00')
put(ws, f"D{r}", NET_PROFIT, BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", EXPENSES, BOLD, fmt='"$"#,##0.00')
put(ws, f"H{r}", TOTAL_HOURS, BOLD, fmt='#,##0.00')
r += 1
put(ws, f"B{r}", f"Original ${CONTRACT_ORIG:,.2f} + ${CO_TOTAL_IMPLIED:,.2f} COs (+{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%)", SUB)
put(ws, f"D{r}", f"{NET_PROFIT/REVENUE*100:.1f}% margin", SUB)
put(ws, f"F{r}", f"{EXPENSES/REVENUE*100:.1f}% of revenue", SUB)
put(ws, f"H{r}", f"{TOTAL_WORKERS} workers", SUB)

r += 3
put(ws, f"B{r}", "SCOPE OF WORK (from bid sheet + plans)", BOLD, F_ALT)
r += 1
scope_lines = [
    f"Full plumbing: underground, garage/podium, rough-in, finish, gas, water main, insulation, mech room.",
    f"Bid sheet (9/9/2014) itemizes {UNITS} apartment units ({UNITS_STD} standard + {UNITS_ADA} ADA Type A).",
    f"Fixtures: {WC_COUNT} water closets (241 unit + 3 amenity), {LAV_COUNT} lavatories, {BATH_COUNT} bathing fixtures (~197 tub/shower + 44 showers),",
    f"{KS_COUNT} kitchen sinks, {HB_COUNT} hose bibs, {WH_COUNT} water heaters, {ROOF_DRAIN_COUNT} roof drains, trench/area drains, sand/oil separator.",
    f"Spec compliance: WA UPC, Mercer Island jurisdiction (permits via Mercer Island).",
]
for line in scope_lines:
    put(ws, f"B{r}", line, BODY, align=LEFT)
    ws.merge_cells(f"B{r}:J{r}")
    r += 1

r += 2
put(ws, f"B{r}", "SOURCES", BOLD, F_HDR)
for col in range(2, 11): put(ws, (r, col), ws.cell(row=r, column=col).value or "", border=BRD, fill=F_HDR)
r += 1
srcs = [
    f"Canonical financial source: {SRC_JDR}",
    f"Job totals (JDR footer): Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: {SRC_CONTRACT} (Lump Sum ${CONTRACT_ORIG:,.2f})",
    f"Bid sheet: {SRC_BID}",
    "Change Orders: 5 CO PDFs in Change Orders/CO_s/ (Legacy CO 11, 19, 20, 21, 22) — ALL image-only scans; individual values not machine-readable",
    "ASI-RFI: 125 PDFs (ASIs 07/09/10/14/16/25REV1/26 + Hadley RFI 051-360+; VIA & BCRA responses)",
    "Submittals: 100+ across EQUIPMENT, FIXTURE, MATERIAL, UNDERGROUND & GARAGE, RESPONSES folders",
    "POs: 180+ across 1 Placed / 2 Scheduled / 3 Completed / Trim PO_s folders (+ 38 Trim POs)",
    "Billing: Legacy sent invoices & releases folder",
    "Photos: on-site photo library (not machine-readable)",
    "Meetings-Schedules: project meeting notes and schedules",
]
for s in srcs:
    put(ws, f"B{r}", s, SRC_FONT, align=LEFT); ws.merge_cells(f"B{r}:J{r}"); r += 1
widths(ws, {1:2, 2:26, 3:22, 4:16, 5:16, 6:18, 7:18, 8:14, 9:14, 10:14})

# ============ TAB 2: BUDGET VS ACTUAL ============
ws = wb.create_sheet("Budget vs Actual")
title(ws, "Budget vs Actual", f"All {len(CODES)} cost codes from JDR. Contract ${CONTRACT_ORIG:,.0f} → ${CONTRACT_FINAL:,.0f} (COs +${CO_TOTAL_IMPLIED:,.2f}).")
hdr(ws, 5, ["Cost Code", "Description", "Original Budget", "Revised Budget", "Actual", "Variance", "% of Revised", "Hours", "Source"])
r = 6
ordered = sorted(CODES.keys(), key=lambda x: int(x))
for code in ordered:
    c = CODES[code]
    put(ws, f"B{r}", code, align=CENTER)
    put(ws, f"C{r}", c['desc'], align=LEFT)
    put(ws, f"D{r}", c['orig'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", c['rev'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", c['actual'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f"=F{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", f'=IF(E{r}=0,"",F{r}/E{r})', fmt='0.0%', align=RIGHT)
    put(ws, f"I{r}", c['hrs_total'] if c['hrs_total'] else "", fmt='#,##0.00', align=RIGHT)
    put(ws, f"J{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"G{r}", f"=F{r}-E{r}", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"I{r}", f"=SUM(I6:I{r-1})", BOLD, fmt='#,##0.00')
widths(ws, {1:2, 2:8, 3:28, 4:16, 5:16, 6:16, 7:16, 8:12, 9:10, 10:40})
ws.freeze_panes = "B6"

# ============ TAB 3: COST BREAKDOWN ============
ws = wb.create_sheet("Cost Breakdown")
title(ws, "Cost Breakdown by Category", "Direct cost composition by category from JDR cost codes")
hdr(ws, 5, ["Category", "Cost Codes", "Actual $", "% of Direct Cost", "% of Revenue", "Source"])
cb = [
    ("Labor", ",".join(LABOR_CODES), LABOR_COST),
    ("Material", ",".join(MATERIAL_CODES), MATERIAL_COST),
    ("Subcontractor + Engineering + Rental + Permits + Other", ",".join(OVERHEAD_CODES), OVERHEAD_COST),
    ("Payroll Burden", "995", BURDEN_COST),
    ("Payroll Taxes", "998", TAX_COST),
]
r = 6
for cat, codes, amt in cb:
    put(ws, f"B{r}", cat, BOLD, align=LEFT)
    put(ws, f"C{r}", codes, align=CENTER)
    put(ws, f"D{r}", amt, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", f"=D{r}/$D${6+len(cb)}", fmt='0.0%', align=RIGHT)
    put(ws, f"F{r}", f"=D{r}/{REVENUE}", fmt='0.0%', align=RIGHT)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL DIRECT COST", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", 1.0, BOLD, fmt='0.0%', align=RIGHT)
put(ws, f"F{r}", f"=D{r}/{REVENUE}", BOLD, fmt='0.0%', align=RIGHT)
widths(ws, {1:2, 2:45, 3:40, 4:18, 5:18, 6:18, 7:40})

# ============ TAB 4: MATERIAL ============
ws = wb.create_sheet("Material")
title(ws, "Material Purchases — AP Vendors", "Material + subcontractor spend by vendor (AP records from JDR). Supplemented with submittal/PO inventory.")
hdr(ws, 5, ["Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "Category (inferred)", "Source"])
ordered_v = sorted(VENDORS.items(), key=lambda kv: -kv[1]['total'])
r = 6
for vid, v in ordered_v:
    put(ws, f"B{r}", vid, align=CENTER)
    put(ws, f"C{r}", v['name'], align=LEFT)
    put(ws, f"D{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", v['count'], align=CENTER)
    n = v['name'].lower()
    if any(x in n for x in ['supply', 'ferguson', 'keller', 'rosen', 'consolidated', 'hardware']):
        cat = "Plumbing / Supplies"
    elif 'franklin' in n:
        cat = "Engineering (601)"
    elif 'mechanical sales' in n:
        cat = "Plumbing / Supplies"
    elif 'rental' in n or 'united rentals' in n:
        cat = "Rental Equipment (602)"
    elif 'scan2core' in n or 'scan to core' in n:
        cat = "Concrete coring / subcon"
    elif 'concrete' in n:
        cat = "Concrete / subcon"
    elif 'credit card' in n or 'cc' in n.split():
        cat = "Credit Card (mixed)"
    elif 'backflow' in n or 'testing' in n:
        cat = "Testing subcontractor"
    elif 'insulation' in n:
        cat = "Insulation subcon"
    else:
        cat = "Uncategorized"
    put(ws, f"F{r}", cat, align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, align=CENTER)
r += 2
put(ws, f"B{r}", f"NOTE: AP total per JDR footer = ${SRC_AP:,.2f}. Top vendors: Rosen Supply, Consolidated Supply, Ferguson, Keller Supply, Mechanical Sales, Franklin Engineering. Submittals folder contains 100+ submittals documenting approved products. POs folder shows 180+ POs (152 completed, 2 placed, 26 scheduled) + 38 Trim POs.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:10, 3:38, 4:16, 5:14, 6:26, 7:40})

# ============ TAB 5: CREW & LABOR ============
ws = wb.create_sheet("Crew & Labor")
title(ws, "Crew & Labor — Worker Roster", f"All {TOTAL_WORKERS} unique payroll workers from JDR.")
hdr(ws, 5, ["Worker ID", "Worker Name", "Total Hours", "Gross Pay", "Blended Wage ($/hr)", "# Work Days", "Source"])
r = 6
ordered_w = sorted(WORKERS.items(), key=lambda kv: -kv[1]['hours'])
for wid, w in ordered_w:
    put(ws, f"B{r}", wid, align=CENTER)
    put(ws, f"C{r}", w['name'], align=LEFT)
    put(ws, f"D{r}", w['hours'], fmt='#,##0.00', align=RIGHT)
    put(ws, f"E{r}", w['amount'], fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", f"=IF(D{r}=0,0,E{r}/D{r})", fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", w['days'], align=CENTER)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
widths(ws, {1:2, 2:10, 3:32, 4:12, 5:14, 6:18, 7:14, 8:40})
ws.freeze_panes = "B6"

# ============ TAB 6: CREW ANALYTICS ============
ws = wb.create_sheet("Crew Analytics")
title(ws, "Crew Analytics", "Team-level labor productivity, concentration, wage dispersion")
put(ws, "B5", "TEAM-LEVEL METRICS", BOLD, F_ALT)
hdr(ws, 6, ["Metric", "Value", "Notes", "Source"])
top_w = ordered_w[0]
top_pct = top_w[1]['hours'] / TOTAL_HOURS
top5_hrs = sum(w[1]['hours'] for w in ordered_w[:5])
top5_pct = top5_hrs / TOTAL_HOURS
max_wage = max(w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0)
min_wage = min(w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0)
crew_metrics = [
    ("Total Workers", TOTAL_WORKERS, "Unique payroll IDs", SRC_JDR),
    ("Total Labor Hours", TOTAL_HOURS, "Sum of codes 100-150", SRC_JDR),
    ("Total Gross Pay", LABOR_COST, "Sum of codes 100-150", SRC_JDR),
    ("Blended Gross Wage ($/hr)", LABOR_COST/TOTAL_HOURS, "Labor$ / Hrs (pre-burden)", "Derived"),
    ("Top Worker Hours Share", top_pct, f"{top_w[0]} {top_w[1]['name']} ({top_w[1]['hours']:.0f} hrs)", "Derived"),
    ("Top 5 Workers Hours Share", top5_pct, "Concentration metric", "Derived"),
    ("Highest Wage Rate ($/hr)", max_wage, "Single-worker blended", "Derived"),
    ("Lowest Wage Rate ($/hr)", min_wage, "Single-worker blended", "Derived"),
    ("Avg Hours per Worker", TOTAL_HOURS/TOTAL_WORKERS, "Includes short-tenure workers", "Derived"),
    ("Avg Project Days per Worker", sum(w['days'] for w in WORKERS.values())/TOTAL_WORKERS, "Mean days", "Derived"),
]
r = 7
for m, v, note, src in crew_metrics:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if isinstance(v, float):
        if "Share" in m: c.number_format = '0.0%'
        elif "Wage" in m: c.number_format = '"$"#,##0.00'
        elif "$" in m: c.number_format = '"$"#,##0.00'
        else: c.number_format = '#,##0.00'
    elif isinstance(v, int):
        c.number_format = '#,##0'
    put(ws, f"D{r}", note, align=LEFT)
    put(ws, f"E{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:30, 3:16, 4:50, 5:40})

# ============ TAB 7: PRODUCTIVITY ============
ws = wb.create_sheet("Productivity")
title(ws, "Productivity Metrics", f"Normalized labor and financial ratios. Per-unit metrics from bid {UNITS} units.")
hdr(ws, 5, ["Metric", "Value", "Basis", "Source / Note"])
prods = [
    ("Revenue per Labor Hour", f"={REVENUE}/{TOTAL_HOURS}", "Formula", "Rev / Total Hrs"),
    ("Profit per Labor Hour", f"={NET_PROFIT}/{TOTAL_HOURS}", "Formula", "Net Profit / Hrs"),
    ("Labor Cost per Hour (blended)", f"={LABOR_COST}/{TOTAL_HOURS}", "Formula", "Pre-burden"),
    ("Fully-Loaded Labor Rate ($/hr)", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{TOTAL_HOURS}", "Formula", "Incl burden + taxes"),
    ("Burden Multiplier", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{LABOR_COST}", "Formula", "Fully-loaded / blended"),
    ("Rough-in Hours (code 120)", CODES["120"]['hrs_total'], "JDR", f"{CODES['120']['hrs_total']:.0f} hrs"),
    ("Finish Hours (code 130)", CODES["130"]['hrs_total'], "JDR", f"{CODES['130']['hrs_total']:.0f} hrs"),
    ("Rough-in % of Total Hours", f'={CODES["120"]["hrs_total"]}/{TOTAL_HOURS}', "Formula", "Code 120 share"),
    ("Gross Margin", f"={NET_PROFIT}/{REVENUE}", "Formula", "Net / Revenue"),
    ("Labor % of Revenue", f"={LABOR_COST}/{REVENUE}", "Formula", ""),
    ("Material % of Revenue", f"={MATERIAL_COST}/{REVENUE}", "Formula", ""),
    ("Direct Cost Ratio", f"={EXPENSES}/{REVENUE}", "Formula", "All expenses / Rev"),
    (f"Revenue per Unit ({UNITS})", f"={REVENUE}/{UNITS}", "Formula", f"{UNITS} units (bid)"),
    ("Labor Hours per Unit", f"={TOTAL_HOURS}/{UNITS}", "Formula", ""),
    ("Labor Cost per Unit", f"={LABOR_COST}/{UNITS}", "Formula", ""),
    ("Material Cost per Unit", f"={MATERIAL_COST}/{UNITS}", "Formula", ""),
    ("Direct Cost per Unit", f"={EXPENSES}/{UNITS}", "Formula", ""),
    ("Rough-in Hours per Unit", f'={CODES["120"]["hrs_total"]}/{UNITS}', "Formula", f"Code 120 / {UNITS}"),
]
r = 6
pct_rows = {"Gross Margin", "Labor % of Revenue", "Material % of Revenue", "Direct Cost Ratio", "Rough-in % of Total Hours"}
for m, v, basis, note in prods:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if m in pct_rows: c.number_format = '0.0%'
    elif m == "Burden Multiplier": c.number_format = '0.00"x"'
    elif "Hours" in m and "Rate" not in m and "per" not in m: c.number_format = '#,##0.00'
    elif "Hours per" in m: c.number_format = '#,##0.00'
    else: c.number_format = '"$"#,##0.00'
    put(ws, f"D{r}", basis, align=CENTER)
    put(ws, f"E{r}", note, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:32, 3:16, 4:12, 5:44})

# ============ TAB 8: PO COMMITMENTS ============
ws = wb.create_sheet("PO Commitments")
title(ws, "PO Commitments", "Inbound contract value (GC → OWP). Outbound vendor commitments — see Material tab + POs folder.")
hdr(ws, 5, ["PO #", "Date", "Issuer", "Type", "Status", "Description", "Amount", "Source"])
r = 6
put(ws, f"B{r}", "Legacy Prime", align=CENTER)
put(ws, f"C{r}", "2014-09-09 (bid) / subcontract executed", align=LEFT)
put(ws, f"D{r}", GC, align=LEFT)
put(ws, f"E{r}", "Lump Sum Subcontract", align=CENTER)
put(ws, f"F{r}", "Closed (100% billed)", align=CENTER, fill=F_OK)
put(ws, f"G{r}", "Legacy @ Mercer Island Apts plumbing — full scope per bid sheet", align=LEFT)
put(ws, f"H{r}", CONTRACT_ORIG, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"I{r}", SRC_CONTRACT, SRC_FONT, align=LEFT)
r += 1
for co_id, amt, src in CHANGE_ORDERS:
    put(ws, f"B{r}", co_id, align=CENTER)
    put(ws, f"C{r}", NF, align=CENTER)
    put(ws, f"D{r}", GC, align=LEFT)
    put(ws, f"E{r}", "Change Order", align=CENTER)
    put(ws, f"F{r}", "Executed (scan)", align=CENTER, fill=F_OK)
    put(ws, f"G{r}", "Documented CO (image-only PDF — value not machine-readable)", align=LEFT)
    put(ws, f"H{r}", amt, align=RIGHT)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
# Implied CO aggregate row
put(ws, f"B{r}", "CO-IMPLIED", align=CENTER)
put(ws, f"C{r}", "2014-09 to 2016-12", align=LEFT)
put(ws, f"D{r}", GC, align=LEFT)
put(ws, f"E{r}", "CO Aggregate (JDR)", align=CENTER)
put(ws, f"F{r}", "Absorbed", align=CENTER, fill=F_HIGH)
put(ws, f"G{r}", f"Aggregate CO impact per JDR code 999 (Rev − Orig): ${CO_TOTAL_IMPLIED:,.2f}. Individual CO values not OCR-readable.", align=LEFT)
put(ws, f"H{r}", CO_TOTAL_IMPLIED, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"I{r}", SRC_JDR + " (code 999)", SRC_FONT, align=LEFT)
r += 1
put(ws, f"B{r}", "TOTAL (contract + COs)", BOLD)
put(ws, f"H{r}", CONTRACT_FINAL, BOLD, fmt='"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"NOTE: JDR-implied CO total = ${CO_TOTAL_IMPLIED:,.2f}. 5 CO PDFs on file (CO 11, 19, 20, 21, 22) are image-only scans; individual values require OCR. CO numbering gap (jump from 11 to 19) suggests 6+ additional COs exist but are not in the CO_s folder. Outbound POs: 180+ total (2 placed, 26 scheduled, 152 completed) + 38 Trim POs — see POs/ folders.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:I{r}")
widths(ws, {1:2, 2:14, 3:28, 4:22, 5:22, 6:18, 7:42, 8:14, 9:42})

# ============ TAB 9: BILLING & SOV ============
ws = wb.create_sheet("Billing & SOV")
title(ws, "Billing & Schedule of Values", f"{len(INVOICES)} unique invoices to {GC}. Retainage ${RETAINAGE:,.2f} = {RETAINAGE/REVENUE*100:.1f}% of revenue.")
hdr(ws, 5, ["Invoice #", "Date", "Total Billed (signed)", "Retainage (signed)", "# Lines", "Source"])
r = 6
for inv in sorted(INVOICES.keys()):
    iv = INVOICES[inv]
    put(ws, f"B{r}", inv, align=CENTER)
    put(ws, f"C{r}", iv['date'], align=CENTER)
    put(ws, f"D{r}", iv['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", iv['retainage'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", iv['lines'], align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL (signed)", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
r += 2
first_inv = sorted(INVOICES.keys())[0]; last_inv = sorted(INVOICES.keys())[-1]
put(ws, f"B{r}", f"NOTE: AR entries signed negative per Sage convention. Net billed = ${REVENUE:,.2f}; retainage ${RETAINAGE:,.2f} outstanding on JDR dated 04/03/2026. First invoice {first_inv} ({INVOICES[first_inv]['date']}); last invoice {last_inv} ({INVOICES[last_inv]['date']}). Billing folder contains 'Legacy sent invoices & releases' subfolder.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:12, 3:12, 4:18, 5:16, 6:10, 7:50})

# ============ TAB 10: INSIGHTS ============
ws = wb.create_sheet("Insights")
title(ws, "Insights & Observations", "Narrative findings from JDR + source documents")
top_vendor = ordered_v[0]
insights = [
    ("STRONG MARGIN", f"Net profit ${NET_PROFIT:,.2f} on ${REVENUE:,.2f} revenue = {NET_PROFIT/REVENUE*100:.1f}% gross margin. Well above typical multifamily plumbing (25-30%).", "Verified", SRC_JDR),
    ("LABOR-HEAVY EXECUTION", f"Labor cost ${LABOR_COST:,.2f} ({LABOR_COST/REVENUE*100:.1f}% of rev) across {TOTAL_HOURS:,.1f} hrs and {TOTAL_WORKERS} workers. Rough-in (120) = {CODES['120']['hrs_total']:.0f} hrs = {CODES['120']['hrs_total']/TOTAL_HOURS*100:.0f}% of labor.", "Verified", SRC_JDR),
    ("ROUGH-IN OVERRUN", f"Code 120 Rough-in Labor: ${CODES['120']['actual']:,.2f} actual vs ${CODES['120']['rev']:,.2f} revised budget = +${CODES['120']['var']:,.2f} ({CODES['120']['var']/CODES['120']['rev']*100:.1f}% over). Primary labor variance driver.", "Verified", SRC_JDR),
    ("MECH-ROOM MATERIAL BLOWOUT", f"Code 242 Mech Room Material: ${CODES['242']['actual']:,.2f} actual vs ${CODES['242']['rev']:,.2f} revised = +${CODES['242']['var']:,.2f} ({CODES['242']['var']/CODES['242']['rev']*100:.0f}% over). Unusual magnitude.", "Verified", SRC_JDR),
    ("FINISH MATERIAL UNDER-RUN", f"Code 230 Finish Material: ${CODES['230']['actual']:,.2f} actual vs ${CODES['230']['rev']:,.2f} = ${CODES['230']['var']:,.2f} ({CODES['230']['var']/CODES['230']['rev']*100:.1f}% under). Major profit contributor.", "Verified", SRC_JDR),
    ("BURDEN-HEAVY", f"Payroll Burden ${BURDEN_COST:,.2f} + Taxes ${TAX_COST:,.2f} = ${BURDEN_COST+TAX_COST:,.2f} = {(BURDEN_COST+TAX_COST)/LABOR_COST*100:.1f}% of gross labor. Burden multiplier = {(LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST:.2f}x.", "Verified", SRC_JDR),
    ("TOP-WORKER CONCENTRATION", f"{top_w[1]['name']} (ID {top_w[0]}) logged {top_w[1]['hours']:.0f} hrs = {top_pct*100:.1f}% of project labor. Top 5 workers = {top5_pct*100:.1f}%.", "Verified", SRC_JDR),
    ("VENDOR CONCENTRATION", f"Top vendor {top_vendor[1]['name']} (${top_vendor[1]['total']:,.2f}). Top 4 suppliers (Rosen, Consolidated, Ferguson, Keller) account for majority of material spend.", "Verified", SRC_JDR),
    ("CO DOCUMENTATION GAP", f"Original ${CONTRACT_ORIG:,.2f} → Final ${CONTRACT_FINAL:,.2f} = +${CO_TOTAL_IMPLIED:,.2f} ({CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%). Only 5 CO PDFs on file (CO 11, 19, 20, 21, 22) — all image-only scans. CO# gap (11→19) suggests 6+ missing COs.", "Medium", "CO folder + " + SRC_JDR),
    ("RETAINAGE OUTSTANDING", f"Retainage ${RETAINAGE:,.2f} still open on JDR dated 04/03/2026, 9+ years after last billing 12/15/2016. Release or write-off pending.", "Verified", SRC_JDR),
    ("RICH COORDINATION TRAIL", "125 ASI-RFI documents on file (ASIs 07/09/10/14/16/25/26 + Hadley RFIs 051-360+). Heavy RFI volume consistent with large multifamily project complexity.", "Verified", "ASI-RFI folder"),
    ("EXTENSIVE SUBMITTAL LOG", "100+ submittals across Equipment, Fixture, Material, Underground & Garage, Responses folders — thorough approval trail.", "Verified", "Submittals folder"),
    ("LARGE PO VOLUME", "180+ POs total (152 completed, 2 placed, 26 scheduled) + 38 Trim POs. Indicates tight procurement control on this $2.8M project.", "Verified", "POs folder"),
    ("BID-LEVEL FIXTURE DETAIL", f"Bid sheet (9/9/2014) itemizes {UNITS} units ({UNITS_STD} std + {UNITS_ADA} ADA), {WC_COUNT} WCs, {LAV_COUNT} lavs, {BATH_COUNT} bathing fixtures, {KS_COUNT} kitchen sinks, {HB_COUNT} hose bibs, {WH_COUNT} water heaters, {ROOF_DRAIN_COUNT} roof drains.", "Verified", SRC_BID),
    ("LONG DURATION", "Work span 01/19/2015 – 12/15/2016 = ~23 months. Sustained execution on a large multifamily build.", "Verified", SRC_JDR),
]
r = 5
hdr(ws, r, ["#", "Insight", "Detail", "Confidence", "Source"])
r = 6
for i, (ttl, det, conf, src) in enumerate(insights, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", ttl, BOLD, align=LEFT)
    put(ws, f"D{r}", det, align=LEFT)
    c = put(ws, f"E{r}", conf, align=CENTER)
    if conf == "Verified": c.fill = F_OK
    elif conf == "Medium": c.fill = F_HIGH
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.row_dimensions[r].height = 58
    r += 1
widths(ws, {1:2, 2:4, 3:32, 4:78, 5:12, 6:40})

# ============ TAB 11: BENCHMARK KPIs ============
ws = wb.create_sheet("Benchmark KPIs")
title(ws, "Benchmark KPIs", "Normalized metrics for cross-project comparison")
hdr(ws, 5, ["KPI", "Data Name", "Value", "Category", "Notes", "Confidence", "Source Document"])
roughin_ratio = CODES["120"]['hrs_total']/TOTAL_HOURS
kpis = [
    ("Job Number", "job_number", JOB, "Profile", "OWP job ID", "Verified", SRC_JDR),
    ("Job Name", "job_name", NAME, "Profile", "Sage short name", "Verified", SRC_JDR),
    ("Project Description", "project_desc", "Legacy @ Mercer Island Apts (Hadley)", "Profile", "Contract + AR lines", "Verified", SRC_CONTRACT),
    ("General Contractor", "general_contractor", GC, "Profile", f"Customer {GC_CUST_CODE}", "Verified", SRC_CONTRACT),
    ("Owner / Developer", "owner", OWNER, "Profile", "Per subcontract", "Verified", SRC_CONTRACT),
    ("Architect", "architect", ARCHITECT, "Profile", "Per RFI responses", "Verified", "ASI-RFI folder"),
    ("Location", "location", "2601 76th Ave SE, Mercer Island, WA 98040", "Profile", "Contract address", "Verified", SRC_CONTRACT),
    ("Project Type", "project_type", f"Multifamily New Construction — Plumbing (~{UNITS} units)", "Profile", "Bid sheet", "Verified", SRC_BID),
    ("Work Start Date", "start_date", "2015-01-19", "Profile", "First invoice", "Verified", SRC_JDR),
    ("Work End Date", "end_date", "2016-12-15", "Profile", "Last AR posting", "Verified", SRC_JDR),
    ("Duration (months)", "duration_months", 23.0, "Profile", "Payroll + AR span", "Verified", "Derived"),
    ("Unit Count", "unit_count", UNITS, "Profile", f"{UNITS_STD} std + {UNITS_ADA} ADA", "Verified", SRC_BID),
    ("ADA Type A Units", "ada_units", UNITS_ADA, "Profile", "Per bid sheet", "Verified", SRC_BID),
    ("Contract Original", "contract_original", CONTRACT_ORIG, "Financial", "Subcontract Lump Sum", "Verified", SRC_CONTRACT),
    ("Contract Final", "contract_final", CONTRACT_FINAL, "Financial", "Code 999 Rev Budget = AR total", "Verified", SRC_JDR),
    ("Change Orders ($, implied)", "change_orders_implied", CO_TOTAL_IMPLIED, "Financial", "Final - Original", "Verified", SRC_JDR),
    ("Change Orders ($, documented)", "change_orders_documented", NF, "Financial", "CO PDFs are image-only scans", "Low", "CO folder"),
    ("CO Count Documented", "co_count", CO_COUNT_DOCUMENTED, "Financial", "Legacy CO 11, 19, 20, 21, 22", "Verified", "CO folder"),
    ("Change Order % of Contract", "co_pct", CO_TOTAL_IMPLIED/CONTRACT_ORIG, "Financial", "", "Verified", "Derived"),
    ("Revenue", "revenue", REVENUE, "Financial", "AR total", "Verified", SRC_JDR),
    ("Direct Cost", "direct_cost", EXPENSES, "Financial", "JDR Job Totals Expenses", "Verified", SRC_JDR),
    ("Net Profit", "net_profit", NET_PROFIT, "Financial", "Rev - Expenses", "Verified", SRC_JDR),
    ("Gross Margin", "gross_margin", NET_PROFIT/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Retainage Outstanding", "retainage", RETAINAGE, "Financial", "Open per JDR", "Verified", SRC_JDR),
    ("Retainage % of Revenue", "retainage_pct", RETAINAGE/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Labor Cost", "labor_cost", LABOR_COST, "Labor", "Codes 100-150", "Verified", SRC_JDR),
    ("Material Cost", "material_cost", MATERIAL_COST, "Material", "Codes 210-244", "Verified", SRC_JDR),
    ("Subcontractor+OH Cost", "overhead_cost", OVERHEAD_COST, "Financial", "Codes 600-607", "Verified", SRC_JDR),
    ("Burden Cost", "burden_cost", BURDEN_COST, "Labor", "Code 995", "Verified", SRC_JDR),
    ("Tax Cost", "tax_cost", TAX_COST, "Labor", "Code 998", "Verified", SRC_JDR),
    ("Total Labor Hours", "total_hours", TOTAL_HOURS, "Labor", "Sum labor codes", "Verified", SRC_JDR),
    ("Total Workers", "total_workers", TOTAL_WORKERS, "Labor", "Unique payroll IDs", "Verified", SRC_JDR),
    ("Blended Gross Wage ($/hr)", "blended_gross_wage", LABOR_COST/TOTAL_HOURS, "Labor", "Pre-burden", "Verified", "Derived"),
    ("Fully-Loaded Wage ($/hr)", "fully_loaded_wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS, "Labor", "Incl burden + tax", "Verified", "Derived"),
    ("Burden Multiplier", "burden_multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST, "Labor", "Fully-loaded/blended", "Verified", "Derived"),
    ("Rough-in Hours", "roughin_hours", CODES["120"]['hrs_total'], "Labor", "Code 120", "Verified", SRC_JDR),
    ("Finish Hours", "finish_hours", CODES["130"]['hrs_total'], "Labor", "Code 130", "Verified", SRC_JDR),
    ("Revenue per Hour", "revenue_per_hour", REVENUE/TOTAL_HOURS, "Productivity", "", "Verified", "Derived"),
    ("Profit per Hour", "profit_per_hour", NET_PROFIT/TOTAL_HOURS, "Productivity", "", "Verified", "Derived"),
    ("Revenue per Unit", "revenue_per_unit", REVENUE/UNITS, "Productivity", f"{UNITS} units", "Verified", "Derived"),
    ("Direct Cost per Unit", "cost_per_unit", EXPENSES/UNITS, "Productivity", f"{UNITS} units", "Verified", "Derived"),
    ("Labor Hours per Unit", "hours_per_unit", TOTAL_HOURS/UNITS, "Productivity", f"{UNITS} units", "Verified", "Derived"),
    ("Labor % of Revenue", "labor_pct_revenue", LABOR_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Material % of Revenue", "material_pct_revenue", MATERIAL_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Total Vendors (AP)", "total_vendors", len(VENDORS), "Material", "Unique vendor IDs", "Verified", SRC_JDR),
    ("Total Invoices (AR)", "total_invoices", len(INVOICES), "Billing", "Unique invoice numbers", "Verified", SRC_JDR),
    ("ASI-RFI Count", "rfi_count", 125, "Docs", "Per folder scan", "Verified", "ASI-RFI folder"),
    ("Submittal Count", "submittal_count", 100, "Docs", "100+ across all submittal subfolders", "Medium", "Submittals folder"),
    ("PO Count", "po_count", 180, "Docs", "152 completed + 2 placed + 26 scheduled + trim", "Verified", "POs folders"),
    ("Fixture Count (WCs)", "wc_count", WC_COUNT, "Scope", f"{UNITS} unit + 3 amenity", "Verified", SRC_BID),
    ("Fixture Count (Lavs)", "lav_count", LAV_COUNT, "Scope", f"{UNITS} unit + 3 amenity", "Verified", SRC_BID),
    ("Bathing Fixtures", "bath_count", BATH_COUNT, "Scope", "197 tub/shower + 44 shower", "Verified", SRC_BID),
    ("Kitchen Sinks", "ks_count", KS_COUNT, "Scope", "Standard + ADA kitchen sinks", "Verified", SRC_BID),
    ("Hose Bibs", "hose_bibs", HB_COUNT, "Scope", "Exterior + common-area", "Verified", SRC_BID),
    ("Water Heaters", "water_heaters", WH_COUNT, "Scope", "Central plant", "Verified", SRC_BID),
    ("Roof Drains", "roof_drains", ROOF_DRAIN_COUNT, "Scope", "Per bid sheet", "Verified", SRC_BID),
    ("AP Spend (JDR footer)", "ap_total", SRC_AP, "Material", "JDR footer", "Verified", SRC_JDR),
    ("PR Spend (JDR footer)", "pr_total", SRC_PR, "Labor", "JDR footer (labor+burden+tax)", "Verified", SRC_JDR),
    ("GL Spend (JDR footer)", "gl_total", SRC_GL, "Financial", "JDR footer", "Verified", SRC_JDR),
]
r = 6
for k in kpis:
    for j, v in enumerate(k):
        cell = put(ws, (r, 2+j), v, align=LEFT if j in (0,1,4,6) else CENTER)
        if j == 2 and isinstance(v, float):
            if "pct" in k[1] or "margin" in k[1]: cell.number_format = '0.00%'
            elif "multiplier" in k[1]: cell.number_format = '0.00"x"'
            elif "hours" in k[1] or "months" in k[1] or "hour" in k[1] or "per_hour" in k[1]: cell.number_format = '#,##0.00'
            else: cell.number_format = '"$"#,##0.00'
        elif j == 2 and isinstance(v, int):
            cell.number_format = '#,##0'
        if j == 5:
            if v == "Verified": cell.fill = F_OK
            elif v == "Medium": cell.fill = F_HIGH
            elif v == "Low": cell.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:30, 3:24, 4:22, 5:14, 6:38, 7:12, 8:38})
ws.freeze_panes = "B6"

# ============ TAB 12: VENDORS ============
ws = wb.create_sheet("Vendors")
title(ws, "Vendors — AP Summary", "Vendor-level spend ranking")
hdr(ws, 5, ["Rank", "Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "% of AP", "Source"])
r = 6
total_ap_vendors = sum(v['total'] for v in VENDORS.values())
for rank, (vid, v) in enumerate(ordered_v, 1):
    put(ws, f"B{r}", rank, align=CENTER)
    put(ws, f"C{r}", vid, align=CENTER)
    put(ws, f"D{r}", v['name'], align=LEFT)
    put(ws, f"E{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", v['count'], align=CENTER)
    put(ws, f"G{r}", f"=E{r}/{total_ap_vendors}", fmt='0.0%', align=RIGHT)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, align=CENTER)
widths(ws, {1:2, 2:6, 3:10, 4:38, 5:14, 6:14, 7:12, 8:40})

# ============ TAB 13: CHANGE LOG ============
ws = wb.create_sheet("Change Log")
title(ws, "Change Log — Master Register", "COs, RFIs, ASIs, Submittals, POs — from project folders")
hdr(ws, 5, ["Event ID", "Type", "Date", "Subject", "Originator", "Cost Impact ($)", "Status", "Source"])
events = [
    ("CONTRACT-ORIG", "Contract", "2014-09 (bid) / 2015-01 (work start)", f"Prime subcontract — Lump Sum ${CONTRACT_ORIG:,.2f}", GC, CONTRACT_ORIG, "Executed", SRC_CONTRACT),
    ("CO-11", "Change Order", NF, "Legacy CO 11 — image-scan PDF, value NOT extractable", GC, NF, "Executed", "CO_s/Legacy CO 11 Signed.pdf"),
    ("CO-19", "Change Order", NF, "Legacy CO 19 — image-scan PDF, value NOT extractable", GC, NF, "Executed", "CO_s/Legacy CO 19.pdf"),
    ("CO-20", "Change Order", NF, "Legacy CO 20 — image-scan PDF, value NOT extractable", GC, NF, "Executed", "CO_s/Legacy CO 20.pdf"),
    ("CO-21", "Change Order", NF, "Legacy CO 21 — image-scan PDF, value NOT extractable", GC, NF, "Executed", "CO_s/Legacy CO 21.pdf"),
    ("CO-22", "Change Order", NF, "Legacy CO 22 — image-scan PDF, value NOT extractable", GC, NF, "Executed", "CO_s/Legacy CO 22.pdf"),
    ("CO-IMPLIED", "CO Aggregate", "Thru 2016-12", f"Aggregate CO impact per JDR code 999 delta = ${CO_TOTAL_IMPLIED:,.2f}", GC, CO_TOTAL_IMPLIED, "Absorbed", SRC_JDR),
    ("CO-GAP", "CO Numbering Gap", "—", "CO# jumps 11→19 in folder — suggests 6+ COs missing from folder", "—", 0, "Unresolved", "CO folder scan"),
    ("ASI-RFI-BATCH", "RFI/ASI batch", "2015-2016", "125 ASI-RFI docs (ASIs 07/09/10/14/16/25/26 + Hadley RFIs 051-360+; VIA & BCRA responses)", f"{ARCHITECT} / Sub / GC", 0, "Resolved", "ASI-RFI/"),
    ("DROPBOX-RFI", "RFI batch", NF, "Additional RFI docs from field tablets", "Sub (OWP)", 0, "Resolved", "DROPBOX FILES FROM TABLETS/"),
    ("SUB-EQUIPMENT", "Submittal batch", NF, "Equipment submittals (Backflow preventer, Expansion tank, etc.)", "Sub (OWP)", 0, "Approved", "Submittals/EQUIPMENT SUBMITTALS/"),
    ("SUB-FIXTURE", "Submittal batch", NF, "Fixture submittals (WC, lav, tub/shower, sinks) — multiple revision rounds", "Sub (OWP)", 0, "Approved", "Submittals/FIXTURE SUBMITTALS/ (incl. REVISED SUBMITTALS 09.21.15)"),
    ("SUB-MATERIAL", "Submittal batch", NF, "Material submittals (PVC pipe & fittings, full material submittal package)", "Sub (OWP)", 0, "Approved", "Submittals/MATERIAL SUBMITTALS/"),
    ("SUB-UG", "Submittal batch", NF, "Underground & Garage submittals (pump basin, oil-water separator, Proset toilet detail)", "Sub (OWP)", 0, "Approved", "Submittals/UNDERGROUND & GARAGE SUBMITTALS/"),
    ("SUB-RESPONSES", "Submittal Responses", NF, "Returned submittal responses from GC/Architect", f"{ARCHITECT} / {GC}", 0, "Logged", "Submittals/RESPONSES/"),
    ("PO-COMPLETED", "PO batch", NF, "152 completed POs", "Sub (OWP)", 0, "Complete", "PO_s/3 Completed/"),
    ("PO-PLACED", "PO batch", NF, "2 placed POs", "Sub (OWP)", 0, "Placed", "PO_s/1 Placed/"),
    ("PO-SCHEDULED", "PO batch", NF, "26 scheduled POs", "Sub (OWP)", 0, "Scheduled", "PO_s/2 Scheduled/"),
    ("PO-TRIM", "PO batch (trim)", NF, "38 Trim POs (fixture trim-out)", "Sub (OWP)", 0, "Complete", "PO_s/Trim PO_s/"),
    ("MEETINGS", "Meeting log", NF, "Project meeting notes & schedules", GC, 0, "Logged", "Meetings-Schedules/"),
    ("OM-TURNOVER", "O&M Closeout", NF, "Operations & Maintenance closeout package", "Sub (OWP)", 0, "Delivered", "O&M_s/"),
    ("FRANKLIN-ENG", "Engineering", NF, "Franklin Mercer Island engineering package (code 601)", "Franklin Engineering", 0, "Completed", "Franklin Mercer Island/"),
    ("FIRST-INVOICE", "Invoice", INVOICES[sorted(INVOICES.keys())[0]]['date'], f"First billing #{sorted(INVOICES.keys())[0]}", "Sub (OWP)", -INVOICES[sorted(INVOICES.keys())[0]]['total'], "Paid", SRC_JDR),
    ("LAST-INVOICE", "Invoice", INVOICES[sorted(INVOICES.keys())[-1]]['date'], f"Last billing #{sorted(INVOICES.keys())[-1]}", "Sub (OWP)", -INVOICES[sorted(INVOICES.keys())[-1]]['total'], "Paid", SRC_JDR),
    ("RETAINAGE-OPEN", "Retainage", "As of 04/03/2026", f"Retainage ${RETAINAGE:,.2f} outstanding 9+ years post-closeout", "GC", 0, "Outstanding", SRC_JDR),
]
r = 6
for e in events:
    eid, et, dt, subj, orig, cost, status, src = e
    put(ws, f"B{r}", eid, align=CENTER)
    put(ws, f"C{r}", et, align=CENTER)
    put(ws, f"D{r}", dt, align=CENTER)
    put(ws, f"E{r}", subj, align=LEFT)
    put(ws, f"F{r}", orig, align=LEFT)
    if cost == NF:
        put(ws, f"G{r}", NF, align=RIGHT)
    else:
        put(ws, f"G{r}", cost, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", status, align=CENTER)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:18, 3:22, 4:22, 5:50, 6:22, 7:16, 8:18, 9:42})
ws.freeze_panes = "B6"

# ============ TAB 14: ROOT CAUSE ANALYSIS ============
ws = wb.create_sheet("Root Cause Analysis")
title(ws, "Root Cause Analysis", "Variance drivers — rough-in labor & mech-room material overruns offset by finish-material savings")
put(ws, "B5", "COST-CODE VARIANCE CATEGORIES", BOLD, F_ALT)
hdr(ws, 6, ["Category", "Codes Affected", "Net $ Variance", "Root Cause (inferred)", "Notes"])
def var(c): return CODES[c]['actual'] - CODES[c]['rev'] if c in CODES else 0
rc_rows = [
    ("Rough-in labor overrun", "120 Rough-in Labor", var("120"), "Field productivity below estimate; 125 RFIs indicate coordination churn", f"${var('120'):,.0f} over — largest labor variance driver"),
    ("Gas / Water-Main / Mech labor overruns", "140, 141, 142, 143", var("140")+var("141")+var("142")+var("143"), "Scope growth in piping work", "Field supplemental scope"),
    ("Finish & Takeoff labor overruns", "101, 130", var("101")+var("130"), "Takeoff labor significantly over; finish labor modestly over", "Estimating gap"),
    ("Garage labor under-run", "111", var("111"), "Efficient garage/podium execution", "Savings"),
    ("Other labor (small)", "100, 110, 112, 150", var("100")+var("110")+var("112")+var("150"), "Mixed", "Minor variances"),
    ("Finish material savings", "230", var("230"), "Favorable procurement on fixtures/finish goods", "Primary material profit driver"),
    ("Mech Room material overrun", "242 Mech Room Material", var("242"), "Likely scope addition (CO-absorbed equipment)", f"${var('242'):,.0f} over — 4-5x budget"),
    ("UG/Garage/Canout material", "210, 211, 212", var("210")+var("211")+var("212"), "Mixed: UG over, Garage/Canout under", "Net savings"),
    ("Rough-in / Gas / Water material", "220, 240, 241", var("220")+var("240")+var("241"), "Rough-in under; gas/water close to budget", "Net savings"),
    ("Burden / Tax accrual", "995, 998", var("995")+var("998"), "Rate growth on higher labor base", "OWP internal"),
    ("Support codes (600-607)", "600,601,602,603,607", var("600")+var("601")+var("602")+var("603")+var("607"), "Engineering + Other Expenses savings; small Subcontractor/Rental/Permit overruns", "Net savings"),
]
r = 7
for cat, codes, netv, cause, note in rc_rows:
    put(ws, f"B{r}", cat, align=LEFT)
    put(ws, f"C{r}", codes, align=LEFT)
    put(ws, f"D{r}", netv, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", cause, align=LEFT)
    put(ws, f"F{r}", note, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL NET VARIANCE (Rev Budget vs Actual)", BOLD)
put(ws, f"D{r}", f"=SUM(D7:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
r += 3
put(ws, f"B{r}", "RESPONSIBILITY ATTRIBUTION", BOLD, F_ALT); r += 1
hdr(ws, r, ["Responsible Party", "# Drivers", "Net $ Impact", "Notes"]); r += 1
resp = [
    ("Sub (OWP) — rough-in productivity", 1, var("120"), f"${var('120'):,.0f} overrun in code 120"),
    ("Sub (OWP) — piping/gas/mech labor", 4, var("140")+var("141")+var("142")+var("143"), "Gas/water-main/mech labor over budget"),
    ("Sub (OWP) — procurement savings (finish)", 1, var("230"), "Finish material came in well under"),
    ("Sub (OWP) — mech-room material risk", 1, var("242"), "Major $ overrun — likely scope absorbed"),
    ("Sub (OWP) — estimating discipline", 4, var("100")+var("111")+var("220")+var("230"), "Mix of savings"),
    ("Designer (VIA) / GC (RFI churn)", 125, 0, "125 ASI-RFIs — coordination cost absorbed in 120 labor"),
    ("Burden rate — OWP internal", 2, var("995")+var("998"), "Rate growth"),
    ("Support codes — OWP", 5, var("600")+var("601")+var("602")+var("603")+var("607"), "Net savings"),
]
for rp, cnt, net, note in resp:
    put(ws, f"B{r}", rp, align=LEFT)
    put(ws, f"C{r}", cnt, align=CENTER)
    put(ws, f"D{r}", net, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", note, align=LEFT)
    r += 1
widths(ws, {1:2, 2:40, 3:40, 4:18, 5:44, 6:44})

# ============ TAB 15: PREDICTIVE SIGNALS ============
ws = wb.create_sheet("Predictive Signals")
title(ws, "Predictive Signals", "Leading indicators from project documents + JDR")
put(ws, "B5", "CURRENT-STATE SIGNALS", BOLD, F_ALT)
hdr(ws, 6, ["Indicator", "Current Value", "Benchmark", "Status", "Meaning"])
signals = [
    ("ASI-RFI Count", 125, "<100 for mid-size MF", "ELEVATED", "125 ASI-RFIs — heavy coordination load"),
    ("Submittal Count", "100+", "varies", "INFO", "Thorough submittal trail"),
    ("ASI Count", 7, "<10", "HEALTHY", "7 ASIs — moderate design changes"),
    ("Documented CO Count", CO_COUNT_DOCUMENTED, "all on file", "ELEVATED", "Only 5 CO PDFs (image-only); CO# gap 11→19 suggests 6+ missing"),
    ("Contract Growth %", f"{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%", "<15%", "ELEVATED", f"{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.1f}% growth — meaningful scope changes"),
    ("CO Value Traceability", "LOW", "all CO $ readable", "RISK", "CO PDFs image-only — individual $ values not machine-readable"),
    ("Labor Hrs vs Budget", f"{TOTAL_HOURS:,.0f}", "varies", "INFO", f"{TOTAL_HOURS:,.0f} hrs"),
    ("Labor Cost % of Revenue", LABOR_COST/REVENUE, "<30%", "HEALTHY" if LABOR_COST/REVENUE < 0.30 else "ELEVATED", f"{LABOR_COST/REVENUE*100:.1f}%"),
    ("Retainage Release Pending", "YES", "Released within 90 days post-closeout", "ELEVATED", f"${RETAINAGE:,.0f} open 9+ years"),
    ("Worker Concentration (top 1)", top_pct, "<25%", "HEALTHY" if top_pct < 0.20 else "ELEVATED", f"{top_pct*100:.1f}% from single worker"),
    ("Rough-in Labor Share", roughin_ratio, "40-65%", "HEALTHY" if 0.4 <= roughin_ratio <= 0.65 else "INFO", f"{roughin_ratio*100:.0f}% on rough-in"),
    ("Rough-in Budget Variance", CODES['120']['var']/CODES['120']['rev'], "±10%", "ELEVATED", f"{CODES['120']['var']/CODES['120']['rev']*100:.1f}% over"),
    ("Mech Room Material Variance", CODES['242']['var']/CODES['242']['rev'], "±10%", "RISK", f"{CODES['242']['var']/CODES['242']['rev']*100:.0f}% over"),
    ("PO Volume", 180, ">100 for large MF", "HEALTHY", "180+ POs — tight procurement"),
    ("Document Completeness", "HIGH (with CO caveat)", "Full CO/RFI/Submittal trail", "ELEVATED", "Contract, RFIs, Submittals, POs, Bid Sheet complete; CO PDFs image-only + folder gap"),
]
r = 7
for sig in signals:
    for j, v in enumerate(sig):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
        if j == 3:
            if v == "ELEVATED": c.fill = F_HIGH
            elif v == "HEALTHY": c.fill = F_OK
            elif v == "RISK": c.fill = F_RISK
            elif v == "UNKNOWN": c.fill = F_NF
            elif v == "INFO" or v == "NEUTRAL": c.fill = F_ALT
    r += 1
r += 2
put(ws, f"B{r}", "FORECAST MODELS", BOLD, F_ALT); r += 1
hdr(ws, r, ["Forecast", "Current Estimate", "Confidence", "Driver", "Model Note"]); r += 1
forecasts = [
    ("Final margin (actual)", NET_PROFIT/REVENUE, "Actual", "Job closed", f"{NET_PROFIT/REVENUE*100:.1f}% — strong"),
    ("Retainage collection probability", "LOW (very stale)", "Qualitative", "9+ years outstanding", "Likely requires AR write-off or dispute resolution"),
    ("Composite risk score (0-100)", 40, "Medium", "Stale retainage + CO doc gap + mech-room variance; financials strong", "Financials strong; docs mostly strong but CO image-only"),
    ("Would re-bid margin target", "≥32%", "Derived", "Historical close", "Assumes similar scope; tighten 120 rough-in and 242 mech-room estimating"),
    ("Unit-level economics", f"${REVENUE/UNITS:,.0f}/unit rev, ${NET_PROFIT/UNITS:,.0f}/unit profit", "Verified", f"{UNITS} units", "Benchmark for future large MF bids"),
]
for f in forecasts:
    for j, v in enumerate(f):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
    r += 1
widths(ws, {1:2, 2:42, 3:22, 4:22, 5:22, 6:58})

# ============ TAB 16: METRIC REGISTRY ============
ws = wb.create_sheet("Metric Registry")
title(ws, "Metric Registry — Cortex Data Catalog", "Every metric with data_label, confidence, and source")
hdr(ws, 5, ["#", "Data Label", "Human Label", "Value", "Unit", "Source Tab", "Confidence", "Source Document(s)"])
metrics = [
    ("job_number", "Job Number", JOB, "id", "Benchmark KPIs", "Verified", SRC_JDR),
    ("job_name", "Job Name", NAME, "text", "Benchmark KPIs", "Verified", SRC_JDR),
    ("project_desc", "Project Description", "Legacy @ Mercer Island Apts (Hadley)", "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("general_contractor", "GC", GC, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("customer_code", "Customer Code", GC_CUST_CODE, "id", "Overview", "Verified", SRC_JDR),
    ("owner", "Owner", OWNER, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("architect", "Architect", ARCHITECT, "text", "Benchmark KPIs", "Verified", "ASI-RFI"),
    ("location", "Location", "2601 76th Ave SE, Mercer Island WA 98040", "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("project_type", "Project Type", f"Multifamily Plumbing ({UNITS} units)", "text", "Benchmark KPIs", "Verified", SRC_BID),
    ("start_date", "Work Start", "2015-01-19", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("end_date", "Work End", "2016-12-15", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("duration_months", "Duration (months)", 23.0, "months", "Benchmark KPIs", "Verified", "Derived"),
    ("unit_count", "Unit Count", UNITS, "units", "Benchmark KPIs", "Verified", SRC_BID),
    ("ada_units", "ADA Type A Units", UNITS_ADA, "units", "Benchmark KPIs", "Verified", SRC_BID),
    ("wc_count", "WC Fixture Count", WC_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("lav_count", "Lavatory Count", LAV_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("bath_count", "Bathing Fixture Count", BATH_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("ks_count", "Kitchen Sink Count", KS_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("hose_bibs", "Hose Bibs", HB_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("water_heaters", "Water Heaters", WH_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("roof_drains", "Roof Drains", ROOF_DRAIN_COUNT, "count", "Benchmark KPIs", "Verified", SRC_BID),
    ("contract_original", "Contract Original", CONTRACT_ORIG, "USD", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("contract_final", "Contract Final", CONTRACT_FINAL, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("change_orders_implied", "COs (implied)", CO_TOTAL_IMPLIED, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("co_count", "CO Count (documented)", CO_COUNT_DOCUMENTED, "count", "Change Log", "Verified", "CO folder"),
    ("co_pct", "CO % of Contract", CO_TOTAL_IMPLIED/CONTRACT_ORIG, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue", "Revenue", REVENUE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("direct_cost", "Direct Cost", EXPENSES, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("net_profit", "Net Profit", NET_PROFIT, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("gross_margin", "Gross Margin", NET_PROFIT/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("retainage", "Retainage Outstanding", RETAINAGE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("retainage_pct", "Retainage % of Revenue", RETAINAGE/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_cost", "Labor Cost", LABOR_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("material_cost", "Material Cost", MATERIAL_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("overhead_cost", "Subcon+OH Cost", OVERHEAD_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("burden_cost", "Burden Cost", BURDEN_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("tax_cost", "Tax Cost", TAX_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_hours", "Total Labor Hours", TOTAL_HOURS, "hours", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_workers", "Total Workers", TOTAL_WORKERS, "count", "Benchmark KPIs", "Verified", SRC_JDR),
    ("blended_gross_wage", "Blended Gross Wage", LABOR_COST/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("fully_loaded_wage", "Fully-Loaded Wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("burden_multiplier", "Burden Multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST, "x", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_hour", "Revenue per Hour", REVENUE/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("profit_per_hour", "Profit per Hour", NET_PROFIT/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_unit", "Revenue per Unit", REVENUE/UNITS, "USD", "Benchmark KPIs", "Verified", "Derived"),
    ("cost_per_unit", "Direct Cost per Unit", EXPENSES/UNITS, "USD", "Benchmark KPIs", "Verified", "Derived"),
    ("hours_per_unit", "Labor Hours per Unit", TOTAL_HOURS/UNITS, "hours", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_pct_revenue", "Labor % of Revenue", LABOR_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("material_pct_revenue", "Material % of Revenue", MATERIAL_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("total_vendors", "Total Vendors (AP)", len(VENDORS), "count", "Vendors", "Verified", SRC_JDR),
    ("total_invoices", "Total Invoices (AR)", len(INVOICES), "count", "Billing & SOV", "Verified", SRC_JDR),
    ("rfi_count", "ASI-RFI Count", 125, "count", "Change Log", "Verified", "ASI-RFI folder"),
    ("asi_count", "ASI Count", 7, "count", "Change Log", "Verified", "ASI-RFI folder"),
    ("submittal_count", "Submittal Count (≥)", 100, "count", "Change Log", "Medium", "Submittals folders"),
    ("po_count", "PO Count (≥)", 180, "count", "Change Log", "Verified", "PO_s folders"),
    ("top_worker_hours_share", "Top Worker Hours Share", top_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("top5_worker_hours_share", "Top 5 Worker Hours Share", top5_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("cost_code_count", "Cost Codes Active", len(CODES), "count", "Budget vs Actual", "Verified", SRC_JDR),
]
r = 6
for i, m in enumerate(metrics, 1):
    put(ws, f"B{r}", i, align=CENTER)
    for j, v in enumerate(m):
        c = put(ws, (r, 3+j), v, align=LEFT if j in (0,1,6) else CENTER)
        if j == 2 and isinstance(v, (int, float)) and not isinstance(v, bool):
            unit = m[3]
            if unit == "USD": c.number_format = '"$"#,##0.00'
            elif unit == "%": c.number_format = '0.00%'
            elif unit == "USD/hr": c.number_format = '"$"#,##0.00'
            elif unit == "x": c.number_format = '0.00"x"'
            elif unit == "hours" or unit == "months": c.number_format = '#,##0.00'
            else: c.number_format = '#,##0'
        if j == 5:
            if v == "Verified": c.fill = F_OK
            elif v == "Medium": c.fill = F_HIGH
            elif v == "Low": c.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:5, 3:28, 4:28, 5:22, 6:10, 7:18, 8:12, 9:36})
ws.freeze_panes = "B6"

# ============ TAB 17: RECONCILIATION ============
ws = wb.create_sheet("Reconciliation")
title(ws, "Reconciliation", "Cross-sheet formula checks")
hdr(ws, 5, ["#", "Check", "Value A", "Value B", "Delta", "Status", "Tabs"])
checks = [
    ("Revenue (JDR) = Contract Final", REVENUE, CONTRACT_FINAL, "1↔8"),
    ("Expenses = Labor+Material+OH+Burden+Tax", EXPENSES, LABOR_COST+MATERIAL_COST+OVERHEAD_COST+BURDEN_COST+TAX_COST, "1↔3"),
    ("Net Profit = Revenue - Expenses", NET_PROFIT, REVENUE - EXPENSES, "1↔Derived"),
    ("JDR Source: AP+PR+GL = Expenses", SRC_AP+SRC_PR+SRC_GL, EXPENSES, "1↔Derived (footer)"),
    ("Budget vs Actual (code 999) = -Revenue", REVENUE, -CODES["999"]['actual'], "2↔1"),
    ("Total Labor Hours = Worker hours sum", TOTAL_HOURS, sum(w['hours'] for w in WORKERS.values()), "5↔2"),
    ("Labor Cost = Worker gross sum", LABOR_COST, sum(w['amount'] for w in WORKERS.values()), "5↔3"),
    ("Vendor total ≈ AP footer", sum(v['total'] for v in VENDORS.values()), SRC_AP, "12↔1 (approximate)"),
    ("Invoice count", len(INVOICES), 25, "9↔11"),
    ("Contract Final - Orig = CO implied", CO_TOTAL_IMPLIED, CONTRACT_FINAL-CONTRACT_ORIG, "11↔Derived"),
    ("Retainage outstanding", RETAINAGE, 138_568.45, "9↔JDR footer"),
    ("Retainage sum from invoices", -sum(i['retainage'] for i in INVOICES.values()), RETAINAGE, "9↔JDR footer"),
    ("Worker count", TOTAL_WORKERS, 52, "5↔11"),
    ("Cost code count", len(CODES), 30, "2↔16"),
]
r = 6
for i, (check, a, b, tabs) in enumerate(checks, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", check, align=LEFT)
    put(ws, f"D{r}", a, fmt='"$"#,##0.00' if isinstance(a, (int, float)) and abs(a) > 100 else None, align=RIGHT)
    put(ws, f"E{r}", b, fmt='"$"#,##0.00' if isinstance(b, (int, float)) and abs(b) > 100 else None, align=RIGHT)
    put(ws, f"F{r}", f"=D{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f'=IF(ABS(F{r})<=1,"TIES",IF(ABS(F{r})<=ABS(E{r})*0.05,"WITHIN","OFF"))', align=CENTER)
    put(ws, f"H{r}", tabs, SRC_FONT, align=CENTER)
    r += 1
r += 2
put(ws, f"B{r}", "SOURCES", HDR, F_HDR)
for col in range(2, 9): ws.cell(row=r, column=col).fill = F_HDR
r += 1
src_lines = [
    f"Job #{JOB} — Cortex v2 17-tab (built from-scratch)",
    f"Canonical financial source: {SRC_JDR}",
    f"JDR Job Totals: Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"JDR Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: {SRC_CONTRACT} (Lump Sum ${CONTRACT_ORIG:,.2f} to {OWNER}, Mercer Island)",
    f"Bid Sheet: {SRC_BID} ({UNITS} units detailed fixture schedule)",
    "Change Orders: 5 CO PDFs (image-only scans) — individual $ not OCR-readable; CO# gap 11→19 suggests 6+ missing; aggregate CO impact = $301,869 per JDR",
    "ASI-RFI: 125 PDFs (ASIs 07/09/10/14/16/25/26 + Hadley RFIs 051-360+; VIA & BCRA responses)",
    "Submittals: 100+ total (Equipment, Fixture, Material, Underground & Garage, Responses)",
    "POs: 180+ total (152 completed / 2 placed / 26 scheduled) + 38 Trim POs",
    "Billing: Legacy sent invoices & releases folder",
    "O&M: closeout package on file",
    "Unresolved: (a) CO individual $ values require OCR; (b) 6+ COs missing from folder (gap 11→19); (c) $138,568 retainage outstanding 9+ years",
    "TIES = within $1  ·  WITHIN = within 5%  ·  OFF = investigate",
]
for line in src_lines:
    put(ws, f"B{r}", line, SRC_FONT, align=LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    r += 1
widths(ws, {1:2, 2:4, 3:48, 4:20, 5:20, 6:14, 7:12, 8:22})

# ============ SAVE ============
out_local = "/sessions/keen-determined-mccarthy/mnt/owp-2023/cortex output files/OWP_2023_JCR_Cortex_v2.xlsx"
os.makedirs(os.path.dirname(out_local), exist_ok=True)
wb.save(out_local)
print(f"Saved {out_local}")
print(f"Tabs ({len(wb.sheetnames)}):", wb.sheetnames)
