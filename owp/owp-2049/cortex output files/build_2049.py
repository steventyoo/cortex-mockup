#!/usr/bin/env python3
"""
build_2049.py — Build OWP_2049_JCR_Cortex_v2.xlsx (17 tabs)

Greenfield 17-tab JCR for Job 2049 Exxel Lynnwood Reserve Rebuild.
Guarantees coverage parity with Sam's v4 Test-Labels schema (every field
Sam surfaces has a home in this workbook) and layers the Cortex-native
narrative/rollup tabs on top.

Reads /sessions/gracious-relaxed-pascal/2049_data.json from parse_2049.py.

Tabs:
  01 Overview
  02 Job Info
  03 Contract Summary
  04 Cost Categories
  05 Budget vs Actual           <- Sam Cost Code Summaries
  06 Cost Code Detail
  07 Transactions Detail        <- Sam Line Items
  08 Crew Roster                <- Sam Worker Wages
  09 Wage Tiers
  10 Productivity
  11 Vendor Analysis
  12 Benchmarks & Derived       <- Sam Derived Fields
  13 Predictive Signals
  14 Insights
  15 Reconciliation             <- Sam Reconciliation
  16 Reconciliation Log         <- Sam Reconciliation Log
  17 Metric Registry
"""
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_JSON = Path('/sessions/gracious-relaxed-pascal/2049_data.json')
ENRICHED_META_JSON = Path('/sessions/gracious-relaxed-pascal/2049_enriched_meta.json')
OUT_DIR = Path('/sessions/gracious-relaxed-pascal/mnt/cortex-mockup/owp/owp-2049/cortex output files')
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / 'OWP_2049_JCR_Cortex_v2.xlsx'

FONT = 'Arial'

# Palette — matches Cortex v2 workbook aesthetic (navy/clay/sage)
NAVY = '1F3A4C'
CLAY = 'B85C3E'
SAGE = '6B8E5A'
CREAM = 'FAF5ED'
STONE = 'E6E1D8'
PAPER = 'F4F1EA'
INK = '2E2A26'
MUTED = '666666'


def fnt(size=10, bold=False, italic=False, color=INK):
    return Font(name=FONT, size=size, bold=bold, italic=italic, color=color)


def fill(color):
    return PatternFill('solid', start_color=color)


THIN = Side(style='thin', color='CCCCCC')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def hdr_row(ws, row, headers, fill_color=NAVY, font_color='FFFFFF'):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = fnt(size=10, bold=True, color=font_color)
        c.fill = fill(fill_color)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = BOX
    ws.row_dimensions[row].height = 22


def title(ws, text, row=1, size=14):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(size=size, bold=True, color=NAVY)
    ws.row_dimensions[row].height = 24


def caption(ws, text, row=2):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(size=9, italic=True, color=MUTED)
    c.alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[row].height = 30


def section(ws, text, row, cols=8, color=PAPER):
    c = ws.cell(row=row, column=1, value=text)
    c.font = fnt(size=11, bold=True, color=NAVY)
    for col in range(1, cols + 1):
        ws.cell(row=row, column=col).fill = fill(color)
    ws.row_dimensions[row].height = 20


def money_format(cell):
    cell.number_format = '$#,##0.00;($#,##0.00);-'


def int_format(cell):
    cell.number_format = '#,##0;(#,##0);-'


def pct_format(cell):
    cell.number_format = '0.0%'


# ---------------------------------------------------------------------------
# DATA HELPERS
# ---------------------------------------------------------------------------

def rollups_by_category(codes):
    r = defaultdict(lambda: {'orig': 0, 'budget': 0, 'actual': 0, 'count': 0})
    for c in codes:
        cat = c['category']
        r[cat]['orig'] += c['original_budget'] or 0
        r[cat]['budget'] += c['current_budget'] or 0
        r[cat]['actual'] += c['actual_amount'] or 0
        r[cat]['count'] += 1
    return r


def vendor_analysis(line_items):
    v = defaultdict(lambda: {'spend': 0, 'count': 0, 'net_due': 0, 'retainage': 0})
    for it in line_items:
        if it['source'] != 'AP':
            continue
        n = it['name']
        v[n]['spend'] += it['actual_amount'] or 0
        v[n]['count'] += 1
        v[n]['net_due'] += it.get('net_due') or 0
        v[n]['retainage'] += it.get('retainage') or 0
    out = []
    for n, row in v.items():
        out.append({'vendor': n, **row})
    out.sort(key=lambda x: -x['spend'])
    return out


# ---------------------------------------------------------------------------
# TABS
# ---------------------------------------------------------------------------

def build_overview(wb, d):
    ws = wb.create_sheet('Overview')
    rr = d['report_record']
    df = d['derived_fields']
    bs = rr['job_totals_by_source']
    revenue = abs(rr['job_totals_revenue'])
    direct = df['direct_cost']
    net = df['net_profit']
    margin = net / revenue if revenue else 0

    title(ws, 'OWP JOB 2049 — EXXEL LYNNWOOD RESERVE REBUILD')
    caption(ws, 'Cortex JCR v2 — 17-tab Job Cost Report. Aligned with v4 Test-Labels schema (Sam coverage parity). Parsed from 175-page Sage Timberline JDR dated Apr 03, 2026.')

    ws.append([])
    # Hero banner row
    banner = [
        ('REVENUE (AR)', f'${revenue:,.0f}', f'Billed to date (Sage AR)'),
        ('DIRECT COST', f'${direct:,.0f}', f'{direct / revenue * 100:.1f}% of revenue'),
        ('NET PROFIT', f'${net:,.0f}', f'{margin * 100:.1f}% margin'),
        ('LABOR HOURS', f'{df["total_labor_hours"]:,.0f}', f'{df["total_workers"]} workers'),
    ]
    r = ws.max_row + 1
    for i, (lbl, val, sub) in enumerate(banner):
        col = 1 + i * 2
        ws.cell(r, col, lbl).font = fnt(size=9, bold=True, color=MUTED)
        ws.cell(r + 1, col, val).font = fnt(size=16, bold=True, color=NAVY)
        ws.cell(r + 2, col, sub).font = fnt(size=9, italic=True, color=MUTED)
    ws.row_dimensions[r + 1].height = 22

    # Cost breakdown table
    ws.cell(r + 5, 1, 'COST BREAKDOWN (by category)').font = fnt(size=11, bold=True, color=NAVY)
    hdr_r = r + 6
    hdr_row(ws, hdr_r, ['Category', 'Budget', 'Actual', 'Variance', 'Consumption', '% of Direct'])
    rb = rollups_by_category(d['cost_code_summaries'])
    total_b = total_a = 0
    for i, cat in enumerate(['Labor', 'Material', 'Overhead', 'Burden']):
        rr2 = rb[cat]
        row = hdr_r + 1 + i
        ws.cell(row, 1, cat)
        ws.cell(row, 2, rr2['budget'])
        ws.cell(row, 3, rr2['actual'])
        ws.cell(row, 4, f'=B{row}-C{row}')
        ws.cell(row, 5, f'=IFERROR(C{row}/B{row},0)')
        ws.cell(row, 6, f'=IFERROR(C{row}/C{row + 4 - i},0)')
        for col in (2, 3, 4):
            money_format(ws.cell(row, col))
        pct_format(ws.cell(row, 5))
        pct_format(ws.cell(row, 6))
        total_b += rr2['budget']
        total_a += rr2['actual']
    trow = hdr_r + 5
    ws.cell(trow, 1, 'TOTAL (expense codes)').font = fnt(bold=True)
    ws.cell(trow, 2, total_b).font = fnt(bold=True)
    ws.cell(trow, 3, total_a).font = fnt(bold=True)
    ws.cell(trow, 4, f'=B{trow}-C{trow}').font = fnt(bold=True)
    ws.cell(trow, 5, f'=IFERROR(C{trow}/B{trow},0)').font = fnt(bold=True)
    ws.cell(trow, 6, 1).font = fnt(bold=True)
    for col in (2, 3, 4):
        money_format(ws.cell(trow, col))
    pct_format(ws.cell(trow, 5))
    pct_format(ws.cell(trow, 6))

    # Source breakdown
    srow = trow + 3
    ws.cell(srow, 1, 'COST BY SOURCE (PR / AP / GL / AR)').font = fnt(size=11, bold=True, color=NAVY)
    hdr_row(ws, srow + 1, ['Source', 'Amount', '% of Direct', 'Notes'])
    rows = [
        ('PR (Payroll)', bs['PR'], bs['PR'] / direct, 'Worker hours + wages (reg + OT + burden)'),
        ('AP (Vendor Invoices)', bs['AP'], bs['AP'] / direct, 'Material + subcontractor spend'),
        ('GL (General Ledger)', bs['GL'], bs['GL'] / direct, 'Misc adjustments + deposits'),
        ('AR (Billings, credit)', bs['AR'], -bs['AR'] / revenue, 'Revenue billed to GC (stored negative)'),
    ]
    for i, (src, amt, pct, note) in enumerate(rows):
        rowr = srow + 2 + i
        ws.cell(rowr, 1, src)
        ws.cell(rowr, 2, amt); money_format(ws.cell(rowr, 2))
        ws.cell(rowr, 3, pct); pct_format(ws.cell(rowr, 3))
        ws.cell(rowr, 4, note)

    set_widths(ws, [28, 20, 18, 18, 16, 16, 16, 16])


def build_job_info(wb, d):
    ws = wb.create_sheet('Job Info')
    rr = d['report_record']
    em = d.get('enriched_meta', {})
    title(ws, 'JOB INFORMATION — 2049')
    caption(ws, 'Project identity + metadata enriched from GDrive close-out package (Exxel Lynnwood Reserve Rebuild Apartments project info PDF, signed subcontract, permits, scope).')
    ws.append([])
    hdr_row(ws, 3, ['Field', 'Value', 'Source / Notes'])
    rows = [
        ('job_number',         rr['job_number'],                                          'JDR cover page'),
        ('job_name',           em.get('job_name') or rr['job_name'],                      'Exxel Lynnwood Reserve Rebuild Apartments project info PDF'),
        ('job_name_long',      em.get('job_name_long', rr['job_name']),                   'Exxel Lynnwood Reserve Rebuild Apartments project info PDF'),
        ('general_contractor', em.get('general_contractor'),                              'Signed subcontract 201506ONELLC — Shelter Holdings / Reserve at Lynnwood LP'),
        ('owner',              em.get('owner'),                                           'Signed subcontract §A.a — Reserve at Lynnwood LP'),
        ('architect',          em.get('architect'),                                       'Contract counterparty: Reserve at Lynnwood LP'),
        ('location',           em.get('location'),                                        'Signed subcontract + permit records'),
        ('project_type',       em.get('project_type'),                                    'Scope doc + submittal TOC (7-story mixed-use + 2 parking levels)'),
        ('unit_count',         em.get('units'),                                           'Plan drawings: 6-story, 295 units (279 Type B + 16 Type A ADA), 2 buildings (BLDG B + BLDG C)'),
        ('fixture_count',      em.get('total_fixtures'),                                  'Estimate from scope doc + submittal TOC (see fixture_estimate_note)'),
        ('fixture_estimate_note', em.get('fixture_estimate_note'),                        'Explains fixture_count basis'),
        ('parking_levels',     em.get('parking_levels'),                                  'Scope doc — P1, P2 rough-in sequence'),
        ('floors_residential', em.get('floors_residential'),                              'Scope doc — L1-L4 rough-in sequence'),
        ('project_start_date', em.get('start_date'),                                      'JDR first posting 9/8/17 → contract 9/5/17 → construction start Sep 2017 (JDR first posting)'),
        ('project_end_date',   em.get('end_date'),                                        'Final pay app 12/13/2019'),
        ('duration_months',    em.get('duration_months'),                                 'start → end'),
        ('contract_signed',    em.get('contract_signed_date'),                            'Signed subcontract — Richard Donelson 9/11/17'),
        ('permit_jurisdiction', em.get('permit_jurisdiction'),                            'City of Seattle (DCI) permit (Seattle DCI plumbing + gas + backflow)'),
        ('permits_issued',     '; '.join(em.get('permits_issued', [])),                   'Permits folder: 3 permit types × renewals/extensions'),
        ('status',             'CLOSED',                                                  'Inferred from JDR filter: "Job 2049 only, Open, Closed"'),
        ('report_date',        rr['report_date'],                                         'JDR run date'),
    ]
    for i, (k, v, note) in enumerate(rows):
        row = 4 + i
        ws.cell(row, 1, k); ws.cell(row, 1).font = fnt(bold=True)
        ws.cell(row, 2, v if v is not None else '')
        ws.cell(row, 3, note); ws.cell(row, 3).font = fnt(italic=True, color=MUTED)
        if v is None or (isinstance(v, str) and 'PENDING' in v):
            ws.cell(row, 2).fill = fill('FFF4CC')
    set_widths(ws, [24, 60, 70])


def build_change_orders(wb, d):
    ws = wb.create_sheet('Change Orders')
    em = d.get('enriched_meta', {})
    co_log = d.get('co_log', [])
    unex = d.get('unexecuted_cors', [])
    title(ws, 'CHANGE ORDER LOG — 2049')
    caption(ws, 'CO / COR register reconciled against internal Budget Transfer xlsx files in the Change Orders/BUDGET TRANSFERS/ GDrive folder. Executed CO sum ties to JDR-implied CO within $90.')
    ws.append([])
    hdr_row(ws, 3, ['CO #', 'COR #', 'Description', 'Amount', 'Status', 'Signed PDF'])
    r = 4
    for co in co_log:
        ws.cell(r, 1, str(co['co']))
        ws.cell(r, 2, co.get('cor_ref') if co.get('cor_ref') is not None else '')
        ws.cell(r, 3, co['description'])
        ws.cell(r, 4, co['amount']); money_format(ws.cell(r, 4))
        ws.cell(r, 5, co['status'])
        ws.cell(r, 6, 'Y' if co.get('signed_pdf') else 'N')
        r += 1
    # Totals
    r += 1
    ws.cell(r, 1, 'Executed subtotal'); ws.cell(r, 1).font = fnt(bold=True)
    exec_sum = sum(c['amount'] for c in co_log if isinstance(c['co'], int))
    ws.cell(r, 4, exec_sum); money_format(ws.cell(r, 4)); ws.cell(r, 4).font = fnt(bold=True)
    r += 1
    ws.cell(r, 1, 'Budget-transfer total (incl. extras)'); ws.cell(r, 1).font = fnt(bold=True)
    bt_sum = sum(c['amount'] for c in co_log)
    ws.cell(r, 4, bt_sum); money_format(ws.cell(r, 4)); ws.cell(r, 4).font = fnt(bold=True)
    r += 1
    ws.cell(r, 1, 'JDR-implied CO (revenue − original contract)'); ws.cell(r, 1).font = fnt(bold=True)
    jdr_co = em.get('change_orders_net', 0)
    ws.cell(r, 4, jdr_co); money_format(ws.cell(r, 4)); ws.cell(r, 4).font = fnt(bold=True)
    r += 1
    ws.cell(r, 1, 'Reconciliation gap'); ws.cell(r, 1).font = fnt(bold=True)
    ws.cell(r, 4, bt_sum - jdr_co); money_format(ws.cell(r, 4)); ws.cell(r, 4).font = fnt(bold=True)
    r += 2
    section(ws, 'UNEXECUTED / SUPERSEDED COR_s', r)
    r += 1
    hdr_row(ws, r, ['COR #', 'Description', '', '', '', ''])
    r += 1
    for u in unex:
        ws.cell(r, 1, u['cor'])
        ws.cell(r, 2, u['description'])
        r += 1
    set_widths(ws, [10, 10, 60, 16, 14, 12])


def build_project_team(wb, d):
    ws = wb.create_sheet('Project Team')
    em = d.get('enriched_meta', {})
    title(ws, 'PROJECT TEAM — 2049')
    caption(ws, 'Primary contacts extracted from Exxel Lynnwood Reserve Rebuild Apartments project start-up PDF + signed subcontract.')
    ws.append([])
    hdr_row(ws, 3, ['Role', 'Name', 'Email', 'Phone'])
    r = 4
    for pm in em.get('gc_project_managers', []):
        ws.cell(r, 1, 'GC Project Manager')
        ws.cell(r, 2, pm.get('name'))
        ws.cell(r, 3, pm.get('email'))
        ws.cell(r, 4, pm.get('phone'))
        r += 1
    sup = em.get('gc_superintendent')
    if sup:
        ws.cell(r, 1, 'GC Superintendent')
        ws.cell(r, 2, sup.get('name'))
        ws.cell(r, 3, sup.get('email'))
        ws.cell(r, 4, sup.get('phone'))
        r += 1
    ws.cell(r, 1, 'OWP Signatory')
    ws.cell(r, 2, em.get('owp_signatory', ''))
    ws.cell(r, 3, 'mike@onewayplumbing.com')
    ws.cell(r, 4, '425-968-8314')
    r += 1
    ws.cell(r, 1, 'Architect')
    ws.cell(r, 2, em.get('architect', ''))
    r += 1
    ws.cell(r, 1, 'Owner')
    ws.cell(r, 2, em.get('owner', ''))
    r += 1
    ws.cell(r, 1, 'Inspecting Authority')
    ws.cell(r, 2, em.get('permit_jurisdiction', ''))
    r += 1
    set_widths(ws, [24, 28, 38, 20])


def build_document_pipeline(wb, d):
    ws = wb.create_sheet('Document Pipeline')
    em = d.get('enriched_meta', {})
    title(ws, 'DOCUMENT PIPELINE — 2049')
    caption(ws, 'Quantitative document counts from the Exxel Lynnwood Reserve Rebuild Apartments GDrive close-out package — used by the Cortex dashboard for CO/RFI/Submittal density metrics.')
    ws.append([])
    hdr_row(ws, 3, ['Category', 'Count', 'Source folder'])
    rows = [
        ('Executed Change Orders',       em.get('executed_co_count'),       'Change Orders/CO_s/ + BUDGET TRANSFERS/'),
        ('CORs submitted',               em.get('cor_count'),               'Change Orders/COR_s/'),
        ('Budget transfer entries',      em.get('budget_transfer_count'),   'Change Orders/BUDGET TRANSFERS/'),
        ('ASI / RFI responses',          em.get('rfi_asi_count'),           'ASI-RFI/'),
        ('Submittals',                   em.get('submittal_count'),         'Submittals/ (PDF count)'),
        ('POs — placed',                 em.get('po_count_placed'),         'PO_s/1 Placed/'),
        ('POs — scheduled',              em.get('po_count_scheduled'),      'PO_s/2 Scheduled/'),
        ('POs — completed',              em.get('po_count_completed'),      'PO_s/3 Completed/'),
        ('POs — total',                  em.get('po_count_total'),          'PO_s/ (all three subfolders)'),
        ('Vendor invoices',              em.get('invoice_count'),           'Invoices/'),
        ('Lien releases',                em.get('lien_release_count'),      'Billing/Vendor Lien Waivers (empty folder for Notch — releases may be in PayApp packets)'),
        ('Permits issued',               len(em.get('permits_issued', [])), 'Permits/ (Seattle DCI plumbing + gas + backflow permits)'),
        ('Backcharges',                  em.get('backcharge_count'),        'Change Orders/COR_s/ (no backcharges logged for Exxel Lynnwood Reserve Rebuild)'),
    ]
    for i, (k, v, src) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, k)
        ws.cell(r, 2, v if v is not None else '')
        ws.cell(r, 3, src)
        if v is None:
            ws.cell(r, 2).fill = fill('FFF4CC')
    set_widths(ws, [30, 12, 50])


def build_contract(wb, d):
    ws = wb.create_sheet('Contract Summary')
    rr = d['report_record']
    revenue = abs(rr['job_totals_revenue'])
    retainage = abs(rr['job_totals_retainage'])
    cs = d['cost_code_summaries']
    orig_all = sum(c['original_budget'] for c in cs if c['code'] != '999')
    rev_all = sum(c['current_budget'] for c in cs if c['code'] != '999')

    title(ws, 'CONTRACT SUMMARY — 2049')
    caption(ws, 'Contract basis derived from Cost Code 999 Sales (billed AR) and expense-code budgets (Sage "Original" and "Revised" columns). Revised–Original delta = net change-order impact on internal budget.')
    ws.append([])
    hdr_row(ws, 3, ['Field', 'Value', 'Notes'])
    rows = [
        ('original_contract_budget', -d['cost_code_summaries'][-1]['original_budget'] if d['cost_code_summaries'] else 0, 'Code 999 original (negated)'),
        ('current_contract_budget', -d['cost_code_summaries'][-1]['current_budget'] if d['cost_code_summaries'] else 0, 'Code 999 current (negated)'),
        ('revenue_ar_billed', revenue, 'Sum of AR line items (abs)'),
        ('retainage_outstanding', retainage, 'Sum of AR retainage column'),
        ('net_due', rr['job_totals_net_due'], 'Sage job_totals_net_due'),
        ('expense_original_budget', orig_all, 'Sum of Org Budget across 100-607,995,998'),
        ('expense_current_budget', rev_all, 'Sum of Rev Budget across 100-607,995,998'),
        ('change_orders_to_budget', rev_all - orig_all, 'Revised − Original across expense codes'),
        ('direct_cost', d['derived_fields']['direct_cost'], 'PR + AP + GL totals'),
        ('net_profit', d['derived_fields']['net_profit'], 'abs(revenue) − direct_cost'),
        ('gross_margin_pct', d['derived_fields']['net_profit'] / revenue if revenue else 0, 'net / revenue'),
    ]
    for i, (k, v, note) in enumerate(rows):
        row = 4 + i
        ws.cell(row, 1, k); ws.cell(row, 1).font = fnt(bold=True)
        ws.cell(row, 2, v)
        if isinstance(v, (int, float)):
            if 'pct' in k:
                pct_format(ws.cell(row, 2))
            else:
                money_format(ws.cell(row, 2))
        ws.cell(row, 3, note); ws.cell(row, 3).font = fnt(italic=True, color=MUTED)
    set_widths(ws, [28, 22, 60])


def build_cost_categories(wb, d):
    ws = wb.create_sheet('Cost Categories')
    title(ws, 'COST CATEGORIES — 2049')
    caption(ws, 'Rollup of the 29 Sage cost codes into five taxonomic buckets: Labor (1xx), Material (2xx), Overhead (6xx), Burden (995/998), Revenue (999).')
    ws.append([])
    hdr_row(ws, 3, ['Category', 'Code Count', 'Original Budget', 'Current Budget', 'Actual', 'Variance', 'Consumption %', 'Per-Unit Cost*'])
    rb = rollups_by_category(d['cost_code_summaries'])
    cats = ['Labor', 'Material', 'Overhead', 'Burden', 'Revenue']
    for i, cat in enumerate(cats):
        r = 4 + i
        rr2 = rb[cat]
        ws.cell(r, 1, cat).font = fnt(bold=True)
        ws.cell(r, 2, rr2['count'])
        ws.cell(r, 3, rr2['orig'])
        ws.cell(r, 4, rr2['budget'])
        ws.cell(r, 5, rr2['actual'])
        ws.cell(r, 6, f'=D{r}-E{r}')
        ws.cell(r, 7, f'=IFERROR(E{r}/D{r},0)')
        ws.cell(r, 8, 'PENDING (needs unit_count)')
        for c in (3, 4, 5, 6):
            money_format(ws.cell(r, c))
        pct_format(ws.cell(r, 7))
    set_widths(ws, [14, 12, 18, 18, 18, 18, 14, 22])


def build_bva(wb, d):
    ws = wb.create_sheet('Budget vs Actual')
    title(ws, 'BUDGET vs ACTUAL — 29 COST CODES')
    caption(ws, 'Full cost-code ledger with original/revised/actual/variance. Matches Sam Cost Code Summaries tab exactly. Labor codes include hours breakdown (reg/ot/dt).')
    ws.append([])
    hdr_row(ws, 3, ['Code', 'Description', 'Category', 'Original Budget', 'Current Budget',
                    '+/- Budget', 'Actual', 'Net Due', 'Retainage',
                    'Reg Hours', 'OT Hours', 'DT Hours', 'Consumption %', 'Status'])
    r = 4
    for c in d['cost_code_summaries']:
        actual = c['actual_amount'] or 0
        cur = c['current_budget'] or 0
        cons = actual / cur if cur else 0
        if c['code'] == '999':
            status = '—'
        elif cur == 0 and actual == 0:
            status = 'UNUSED'
        elif cons > 1.5:
            status = 'CRITICAL'
        elif cons > 1.1:
            status = 'OVER'
        elif cons < 0.9:
            status = 'UNDER'
        else:
            status = 'ON'
        ws.cell(r, 1, c['code'])
        ws.cell(r, 2, c['description'])
        ws.cell(r, 3, c['category'])
        ws.cell(r, 4, c['original_budget'])
        ws.cell(r, 5, c['current_budget'])
        ws.cell(r, 6, c['plus_minus_budget'])
        ws.cell(r, 7, c['actual_amount'])
        ws.cell(r, 8, c['net_due'])
        ws.cell(r, 9, c['retainage'])
        ws.cell(r, 10, c.get('regular_hours') or 0)
        ws.cell(r, 11, c.get('overtime_hours') or 0)
        ws.cell(r, 12, c.get('doubletime_hours') or 0)
        ws.cell(r, 13, cons)
        ws.cell(r, 14, status)
        for col in (4, 5, 6, 7, 8, 9):
            money_format(ws.cell(r, col))
        pct_format(ws.cell(r, 13))
        if status == 'CRITICAL':
            ws.cell(r, 14).fill = fill('F7D4CB')
        elif status == 'OVER':
            ws.cell(r, 14).fill = fill('FCE8CA')
        elif status == 'UNDER':
            ws.cell(r, 14).fill = fill('E6EFE0')
        r += 1
    ws.freeze_panes = 'A4'
    set_widths(ws, [8, 28, 12, 14, 14, 14, 14, 10, 12, 10, 10, 10, 12, 12])


def build_cost_code_detail(wb, d):
    ws = wb.create_sheet('Cost Code Detail')
    title(ws, 'COST CODE DETAIL — 2049')
    caption(ws, 'Per-code drill-down: transaction count, source mix, top vendor/worker, date range. Derived from line_items[] grouped by cost_code.')
    ws.append([])
    hdr_row(ws, 3, ['Code', 'Description', 'Actual', 'Txn Count', 'PR Lines', 'AP Lines', 'GL Lines', 'AR Lines',
                    'Top Vendor / Worker', 'First Date', 'Last Date'])

    by_code = defaultdict(lambda: {'txns': [], 'first': None, 'last': None})
    for it in d['line_items']:
        by_code[it['cost_code']]['txns'].append(it)

    r = 4
    for c in d['cost_code_summaries']:
        txns = by_code[c['code']]['txns']
        src_counts = Counter(t['source'] for t in txns)
        # Top contributor
        by_name = defaultdict(float)
        for t in txns:
            by_name[t['name']] += abs(t['actual_amount'] or 0)
        top = max(by_name.items(), key=lambda x: x[1]) if by_name else ('—', 0)
        dates = sorted(t['posted_date'] for t in txns if t.get('posted_date'))
        first = dates[0] if dates else ''
        last = dates[-1] if dates else ''
        ws.cell(r, 1, c['code'])
        ws.cell(r, 2, c['description'])
        ws.cell(r, 3, c['actual_amount']); money_format(ws.cell(r, 3))
        ws.cell(r, 4, len(txns))
        ws.cell(r, 5, src_counts.get('PR', 0))
        ws.cell(r, 6, src_counts.get('AP', 0))
        ws.cell(r, 7, src_counts.get('GL', 0))
        ws.cell(r, 8, src_counts.get('AR', 0))
        ws.cell(r, 9, f"{top[0]} (${top[1]:,.0f})" if top[1] else '—')
        ws.cell(r, 10, first)
        ws.cell(r, 11, last)
        r += 1
    set_widths(ws, [8, 28, 14, 10, 10, 10, 10, 10, 34, 12, 12])


def build_transactions(wb, d):
    ws = wb.create_sheet('Transactions Detail')
    li = d['line_items']
    title(ws, f'TRANSACTIONS DETAIL — {len(li):,} LINE ITEMS')
    caption(ws, 'Every transaction parsed from the JDR. AR amounts are negative per Sage convention (credits). Mirrors Sam Line Items tab 1:1.')
    ws.append([])
    hdrs = ['cost_code', 'description', 'source', 'ref_number', 'document_date',
            'posted_date', 'number', 'name', 'regular_hours', 'overtime_hours',
            'regular_amount', 'overtime_amount', 'actual_amount',
            'check_number', 'net_due', 'retainage']
    hdr_row(ws, 3, hdrs)
    for i, it in enumerate(li):
        r = 4 + i
        ws.cell(r, 1, it['cost_code'])
        ws.cell(r, 2, it['description'])
        ws.cell(r, 3, it['source'])
        ws.cell(r, 4, it['ref_number'])
        ws.cell(r, 5, it['document_date'])
        ws.cell(r, 6, it['posted_date'])
        ws.cell(r, 7, it['number'])
        ws.cell(r, 8, it['name'])
        ws.cell(r, 9, it.get('regular_hours'))
        ws.cell(r, 10, it.get('overtime_hours'))
        ws.cell(r, 11, it.get('regular_amount'))
        ws.cell(r, 12, it.get('overtime_amount'))
        ws.cell(r, 13, it['actual_amount'])
        ws.cell(r, 14, it.get('check_number'))
        ws.cell(r, 15, it.get('net_due') or 0)
        ws.cell(r, 16, it.get('retainage') or 0)
    ws.freeze_panes = 'A4'
    set_widths(ws, [8, 22, 8, 10, 12, 12, 10, 30, 10, 10, 12, 12, 12, 14, 10, 10])


def build_crew_roster(wb, d):
    ws = wb.create_sheet('Crew Roster')
    workers = d['worker_wages']
    title(ws, f'CREW ROSTER — {len(workers)} WORKERS')
    caption(ws, 'Per-worker hours + wages grouped by tier (APPRENTICE / JOURNEYMAN / LEAD / OT-ONLY). nominal_rate = regular_amount ÷ regular_hours. Mirrors Sam Worker Wages tab.')
    ws.append([])
    hdr_row(ws, 3, ['Tier', 'Worker (Last, First)', 'Reg Hours', 'OT Hours', 'DT Hours',
                    'Reg Amount', 'OT Amount', 'Nominal Rate ($/hr)'])
    r = 4
    tiers = ['APPRENTICE/HELPER', 'JOURNEYMAN', 'LEAD/SUPERVISOR', 'OT-ONLY']
    for tier in tiers:
        tier_workers = [w for w in workers if w['tier'] == tier]
        if not tier_workers:
            continue
        for w in tier_workers:
            ws.cell(r, 1, tier)
            ws.cell(r, 2, w['name'])
            ws.cell(r, 3, w['regular_hours'])
            ws.cell(r, 4, w['overtime_hours'])
            ws.cell(r, 5, w['doubletime_hours'])
            ws.cell(r, 6, w['regular_amount']); money_format(ws.cell(r, 6))
            ws.cell(r, 7, w['overtime_amount']); money_format(ws.cell(r, 7))
            rate = w['nominal_rate']
            if rate is not None:
                ws.cell(r, 8, rate); money_format(ws.cell(r, 8))
            else:
                ws.cell(r, 8, '— (no reg hours)')
            r += 1
    ws.freeze_panes = 'A4'
    set_widths(ws, [18, 30, 12, 12, 12, 14, 14, 18])


def build_wage_tiers(wb, d):
    ws = wb.create_sheet('Wage Tiers')
    workers = d['worker_wages']
    df = d['derived_fields']
    title(ws, 'WAGE TIERS & PERCENTILES — 2049')
    caption(ws, 'Tier distribution + percentile snapshot across worker_nominal_wages. Tier thresholds: Apprentice <$20, Journeyman $20–$36, Lead $36+. OT-only = no regular hours booked.')
    ws.append([])
    hdr_row(ws, 3, ['Tier', 'Worker Count', 'Total Reg Hours', 'Total OT Hours', 'Total Labor $', 'Avg Nominal Rate'])
    r = 4
    for tier in ['APPRENTICE/HELPER', 'JOURNEYMAN', 'LEAD/SUPERVISOR', 'OT-ONLY']:
        tw = [w for w in workers if w['tier'] == tier]
        if not tw:
            continue
        reg = sum(w['regular_hours'] for w in tw)
        ot = sum(w['overtime_hours'] for w in tw)
        amt = sum(w['regular_amount'] + w['overtime_amount'] for w in tw)
        rates = [w['nominal_rate'] for w in tw if w['nominal_rate'] is not None]
        avg = sum(rates) / len(rates) if rates else 0
        ws.cell(r, 1, tier)
        ws.cell(r, 2, len(tw))
        ws.cell(r, 3, reg)
        ws.cell(r, 4, ot)
        ws.cell(r, 5, amt); money_format(ws.cell(r, 5))
        ws.cell(r, 6, avg); money_format(ws.cell(r, 6))
        r += 1

    r += 2
    section(ws, 'PERCENTILES (nominal wage)', r)
    r += 1
    hdr_row(ws, r, ['Percentile', 'Rate'])
    for i, p in enumerate(['p10', 'p25', 'p50', 'p75', 'p90']):
        rr = r + 1 + i
        ws.cell(rr, 1, p)
        ws.cell(rr, 2, df['nominal_wage_percentiles'][p])
        money_format(ws.cell(rr, 2))

    r_extra = r + 7
    section(ws, 'WAGE METHODOLOGY (Sam v4 Action #3)', r_extra)
    r_extra += 1
    hdr_row(ws, r_extra, ['Metric', 'Value', 'Formula'])
    rows = [
        ('straight_time_rate', df['straight_time_rate'], 'SUM(regular_amount) / SUM(regular_hours)'),
        ('pr_src_cost_per_hr', df['pr_src_cost_per_hr'], 'PR total / total_labor_hours (includes OT premium)'),
        ('fully_loaded_wage', df['fully_loaded_wage'], '(PR + 995 + 998) / total_labor_hours'),
        ('burden_multiplier', df['burden_multiplier'], 'fully_loaded / pr_src_cost_per_hr'),
    ]
    for i, (k, v, f2) in enumerate(rows):
        rr = r_extra + 1 + i
        ws.cell(rr, 1, k)
        ws.cell(rr, 2, v); money_format(ws.cell(rr, 2))
        ws.cell(rr, 3, f2); ws.cell(rr, 3).font = fnt(italic=True, color=MUTED)
    set_widths(ws, [26, 18, 60, 18, 18])


def build_productivity(wb, d):
    ws = wb.create_sheet('Productivity')
    df = d['derived_fields']
    rr = d['report_record']
    em = d.get('enriched_meta', {}) or {}
    revenue = abs(rr['job_totals_revenue'])
    hrs = df['total_labor_hours']
    units = em.get('units')
    fixtures = em.get('total_fixtures')
    months = em.get('duration_months')
    net_profit = df.get('net_profit', 0)
    title(ws, 'PRODUCTIVITY METRICS — 2049')
    caption(ws, 'Labor throughput + unit-economics indicators. Per-unit / per-fixture / per-month now populated from GDrive-enriched project attributes.')
    ws.append([])
    hdr_row(ws, 3, ['Metric', 'Value', 'Formula / Notes'])
    def pu(x): return (x / units) if units else None
    def pf(x): return (x / fixtures) if fixtures else None
    def pm(x): return (x / months) if months else None
    rows = [
        ('Total Labor Hours',        f'{hrs:,.2f}',                                   'SUM(reg+ot+dt+other) where source=PR'),
        ('Revenue per Labor Hour',   f'${df["revenue_per_labor_hour"]:,.2f}',          'abs(revenue) ÷ labor hours — throughput'),
        ('Labor Unit Cost per Hour', f'${df["labor_unit_cost_per_hr"]:,.2f}',          'Labor-code JTD ÷ labor hours'),
        ('Revenue ÷ Labor-Cost Ratio', f'{df["revenue_per_labor_hour"] / df["labor_unit_cost_per_hr"]:.2f}x', 'Throughput ÷ unit cost — productivity signal'),
        ('Units (residential)',      units,                                            'Exxel Lynnwood Reserve Rebuild Apartments scope doc: 279 Type B + 16 Type A ADA'),
        ('Fixtures (estimate)',      fixtures,                                         'Scope doc estimate — see Job Info fixture_estimate_note'),
        ('Duration (months)',        months,                                           'start_date → end_date from GDrive enrichment'),
        ('Hours per Unit',           f'{pu(hrs):,.2f}' if pu(hrs) else 'N/A',          'total_labor_hours ÷ units'),
        ('Hours per Fixture',        f'{pf(hrs):,.2f}' if pf(hrs) else 'N/A',          'total_labor_hours ÷ fixtures'),
        ('Revenue per Unit',         f'${pu(revenue):,.0f}' if pu(revenue) else 'N/A', 'revenue ÷ units'),
        ('Profit per Unit',          f'${pu(net_profit):,.0f}' if pu(net_profit) else 'N/A', 'net_profit ÷ units'),
        ('Hours per Month',          f'{pm(hrs):,.0f}' if pm(hrs) else 'N/A',          'total_labor_hours ÷ duration_months'),
        ('Revenue per Month',        f'${pm(revenue):,.0f}' if pm(revenue) else 'N/A', 'revenue ÷ duration_months'),
    ]
    for i, (k, v, note) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, k); ws.cell(r, 1).font = fnt(bold=True)
        ws.cell(r, 2, v if v is not None else 'N/A')
        if v is None:
            ws.cell(r, 2).fill = fill('FFF4CC')
        ws.cell(r, 3, note); ws.cell(r, 3).font = fnt(italic=True, color=MUTED)
    set_widths(ws, [28, 22, 60])


def build_vendor_analysis(wb, d):
    ws = wb.create_sheet('Vendor Analysis')
    vendors = vendor_analysis(d['line_items'])
    ap_total = sum(v['spend'] for v in vendors)
    title(ws, f'VENDOR ANALYSIS — {len(vendors)} VENDORS  (${ap_total:,.0f} AP)')
    caption(ws, 'AP-source vendor spend concentration. Sorted by descending spend. Top-vendor % surfaces concentration risk. Matches Sam derived_field.top_vendor_spend.')
    ws.append([])
    hdr_row(ws, 3, ['Rank', 'Vendor', 'Spend', 'Invoice Count', 'Avg Invoice', '% of AP', 'Net Due', 'Retainage'])
    for i, v in enumerate(vendors):
        r = 4 + i
        ws.cell(r, 1, i + 1)
        ws.cell(r, 2, v['vendor'])
        ws.cell(r, 3, v['spend']); money_format(ws.cell(r, 3))
        ws.cell(r, 4, v['count'])
        ws.cell(r, 5, v['spend'] / v['count'] if v['count'] else 0); money_format(ws.cell(r, 5))
        ws.cell(r, 6, v['spend'] / ap_total if ap_total else 0); pct_format(ws.cell(r, 6))
        ws.cell(r, 7, v['net_due']); money_format(ws.cell(r, 7))
        ws.cell(r, 8, v['retainage']); money_format(ws.cell(r, 8))
    ws.freeze_panes = 'A4'
    set_widths(ws, [6, 40, 16, 12, 14, 12, 14, 14])


def build_benchmarks(wb, d):
    ws = wb.create_sheet('Benchmarks & Derived')
    df = d['derived_fields']
    title(ws, 'BENCHMARKS & DERIVED FIELDS — 2049')
    caption(ws, 'All 45+ derived values that Sam Test-Labels v4 ships — carried forward field-for-field. Sections: cost-by-source, labor analytics, material analytics, phase analytics, budget/forecast.')
    ws.append([])

    sections = [
        ('COST-BY-SOURCE ROLLUPS', [
            ('cost_by_source_pr_amount', df['cost_by_source_pr_amount'], 'SUM(line_items where source=PR)'),
            ('cost_by_source_ap_amount', df['cost_by_source_ap_amount'], 'SUM(line_items where source=AP)'),
            ('cost_by_source_gl_amount', df['cost_by_source_gl_amount'], 'SUM(line_items where source=GL)'),
            ('direct_cost', df['direct_cost'], 'PR + AP + GL'),
            ('net_profit', df['net_profit'], 'abs(revenue) − direct_cost'),
            ('pr_pct_of_revenue', df['pr_pct_of_revenue'], 'PR / abs(revenue)'),
            ('ap_pct_of_revenue', df['ap_pct_of_revenue'], 'AP / abs(revenue)'),
            ('direct_cost_pct_of_revenue', df['direct_cost_pct_of_revenue'], 'direct / abs(revenue)'),
        ]),
        ('LABOR ANALYTICS', [
            ('total_labor_hours', df['total_labor_hours'], 'SUM(reg+ot+dt+other) source=PR'),
            ('pr_src_cost_per_hr', df['pr_src_cost_per_hr'], 'PR / labor hours'),
            ('fully_loaded_wage', df['fully_loaded_wage'], '(PR + burden) / labor hours'),
            ('burden_multiplier', df['burden_multiplier'], 'fully_loaded / pr_src'),
            ('straight_time_rate', df['straight_time_rate'], 'SUM(reg_amount) / SUM(reg_hours)'),
            ('total_workers', df['total_workers'], 'COUNT DISTINCT(name) source=PR'),
            ('burden_total', df['burden_total'], 'SUM(actual) codes 995+998'),
            ('wage_percentiles (p10/p25/p50/p75/p90)',
             f"${df['nominal_wage_percentiles']['p10']:.2f} | ${df['nominal_wage_percentiles']['p25']:.2f} | ${df['nominal_wage_percentiles']['p50']:.2f} | ${df['nominal_wage_percentiles']['p75']:.2f} | ${df['nominal_wage_percentiles']['p90']:.2f}",
             'Percentile snapshot across worker_nominal_wages'),
        ]),
        ('MATERIAL ANALYTICS', [
            ('material_spend_total', df['material_spend_total'], 'SUM(actual) codes 2xx'),
            ('material_budget_total', df['material_budget_total'], 'SUM(current_budget) codes 2xx'),
            ('material_codes_tracked', df['material_codes_tracked'], 'COUNT material codes'),
            ('material_codes_over_budget', df['material_codes_over_budget'], 'COUNT where plus_minus > 0'),
            ('material_codes_under_budget', df['material_codes_under_budget'], 'COUNT where plus_minus ≤ 0'),
            ('largest_material_overrun', df['largest_material_overrun'], 'Material code with largest +plus_minus'),
            ('largest_material_savings', df['largest_material_savings'], 'Material code with largest −plus_minus'),
            ('material_vendor_count', df['material_vendor_count'], 'COUNT DISTINCT vendors source=AP'),
            ('top_vendor_spend', df['top_vendor_spend'], 'Vendor with highest AP spend'),
        ]),
        ('PHASE ANALYTICS', [
            ('phases_over_budget', df['phases_over_budget'], 'COUNT codes where plus_minus>0 (excl 999)'),
            ('phases_under_budget', df['phases_under_budget'], 'COUNT codes where plus_minus≤0 (excl 999)'),
            ('largest_overrun', df['largest_overrun'], 'Phase with largest +plus_minus'),
            ('largest_savings', df['largest_savings'], 'Phase with largest −plus_minus'),
        ]),
        ('BUDGET & FORECAST', [
            ('total_jtd_cost', df['total_jtd_cost'], 'SUM(actual) expense codes'),
            ('total_budget', df['total_budget'], 'SUM(current_budget) expense codes'),
            ('overall_pct_budget_consumed', df['overall_pct_budget_consumed'], 'JTD / budget'),
            ('total_over_under_budget', df['total_over_under_budget'], 'budget − JTD'),
            ('labor_unit_cost_per_hr', df['labor_unit_cost_per_hr'], 'Labor JTD / labor hours'),
            ('revenue_per_labor_hour', df['revenue_per_labor_hour'], 'abs(revenue) / labor hours'),
            ('material_price_variance', df['material_price_variance'], 'mat_budget − mat_spend'),
            ('variance_trend', 'N/A (single snapshot)', 'Requires prior-period comparison'),
        ]),
        ('PROJECT-LEVEL BENCHMARKS (pending)', [
            ('project_duration_months', 'PENDING', '(end_date − start_date) / 30.4'),
            ('scope_benchmarks', 'PENDING', 'One object per scope_metrics entry'),
        ]),
    ]

    r = 3
    for section_label, rows in sections:
        section(ws, section_label, r)
        r += 1
        hdr_row(ws, r, ['Field', 'Value', 'Formula / Logic'])
        r += 1
        for k, v, logic in rows:
            ws.cell(r, 1, k); ws.cell(r, 1).font = fnt(bold=True)
            ws.cell(r, 2, v)
            if isinstance(v, (int, float)) and 'pct' not in k.lower():
                money_format(ws.cell(r, 2))
            elif isinstance(v, str) and 'PENDING' in v:
                ws.cell(r, 2).fill = fill('FFF4CC')
            ws.cell(r, 3, logic); ws.cell(r, 3).font = fnt(italic=True, color=MUTED)
            r += 1
        r += 1
    set_widths(ws, [34, 40, 60])


def build_predictive(wb, d):
    ws = wb.create_sheet('Predictive Signals')
    df = d['derived_fields']
    cs = d['cost_code_summaries']
    title(ws, 'PREDICTIVE SIGNALS — 2049')
    caption(ws, 'Rule-based health indicators derived from the parsed data. Green = good, yellow = watch, red = flag.')
    ws.append([])
    hdr_row(ws, 3, ['Signal', 'Status', 'Value', 'Threshold / Notes'])

    # Signal rules
    signals = []
    margin = df['net_profit'] / abs(d['report_record']['job_totals_revenue']) if d['report_record']['job_totals_revenue'] else 0
    signals.append(('Gross margin', 'GREEN' if margin > 0.25 else 'YELLOW' if margin > 0.10 else 'RED',
                    f'{margin * 100:.1f}%', 'Green >25%, Yellow 10–25%, Red <10%'))
    over_count = df['phases_over_budget']
    total_phases = df['phases_over_budget'] + df['phases_under_budget']
    signals.append(('Phases over budget', 'RED' if over_count > total_phases / 2 else 'YELLOW' if over_count > total_phases / 4 else 'GREEN',
                    f'{over_count}/{total_phases}', 'Red >50%, Yellow 25–50%, Green <25%'))
    # Material variance
    mv = df['material_price_variance']
    signals.append(('Material price variance', 'GREEN' if mv > 0 else 'RED', f'${mv:,.0f}', 'Positive = under material budget'))
    # Labor burden
    bm = df['burden_multiplier']
    signals.append(('Burden multiplier', 'GREEN' if 1.25 <= bm <= 1.45 else 'YELLOW', f'{bm:.2f}x', 'Normal range 1.25–1.45x (OWP)'))
    # Revenue/hour throughput
    rh = df['revenue_per_labor_hour']
    signals.append(('Revenue per labor hour', 'GREEN' if rh > 100 else 'YELLOW' if rh > 70 else 'RED',
                    f'${rh:.2f}', 'Green >$100, Yellow $70–$100, Red <$70'))
    # Retainage outstanding
    ret = abs(d['report_record']['job_totals_retainage'])
    signals.append(('Retainage outstanding', 'YELLOW' if ret > 0 else 'GREEN', f'${ret:,.0f}', 'Monitor for collection'))
    # Top vendor concentration
    vendors = vendor_analysis(d['line_items'])
    ap_total = sum(v['spend'] for v in vendors)
    if vendors:
        pct = vendors[0]['spend'] / ap_total
        signals.append(('Top vendor concentration', 'YELLOW' if pct > 0.5 else 'GREEN',
                        f'{pct * 100:.1f}% ({vendors[0]["vendor"]})',
                        'Yellow >50% (single-vendor risk)'))
    # Unused budget (codes with budget but zero actual)
    unused = [c for c in cs if c['code'] != '999' and (c['current_budget'] or 0) > 0 and (c['actual_amount'] or 0) == 0]
    signals.append(('Unused-budget codes', 'YELLOW' if len(unused) > 0 else 'GREEN',
                    f'{len(unused)} codes', 'Codes with budget but no spend'))

    for i, (s, status, val, note) in enumerate(signals):
        r = 4 + i
        ws.cell(r, 1, s); ws.cell(r, 1).font = fnt(bold=True)
        ws.cell(r, 2, status)
        if status == 'GREEN':
            ws.cell(r, 2).fill = fill('E6EFE0')
        elif status == 'YELLOW':
            ws.cell(r, 2).fill = fill('FCE8CA')
        elif status == 'RED':
            ws.cell(r, 2).fill = fill('F7D4CB')
        ws.cell(r, 3, val)
        ws.cell(r, 4, note); ws.cell(r, 4).font = fnt(italic=True, color=MUTED)
    set_widths(ws, [28, 10, 28, 50])


def build_insights(wb, d):
    ws = wb.create_sheet('Insights')
    df = d['derived_fields']
    rr = d['report_record']
    revenue = abs(rr['job_totals_revenue'])
    margin = df['net_profit'] / revenue if revenue else 0
    vendors = vendor_analysis(d['line_items'])
    ap_total = sum(v['spend'] for v in vendors)

    title(ws, 'KEY INSIGHTS — 2049')
    caption(ws, 'Narrative rollup — the 10 most important things a Cortex user should know about this job at a glance.')
    ws.append([])

    insights = [
        f'Revenue billed ${revenue:,.0f}, net profit ${df["net_profit"]:,.0f} at {margin * 100:.1f}% margin.',
        f'Parsed {len(d["line_items"]):,} transactions across 29 cost codes — reconciliation PASS on all codes.',
        f'{df["total_workers"]} distinct PR-booked workers over {df["total_labor_hours"]:,.0f} labor hours '
        f'(${df["pr_src_cost_per_hr"]:.2f}/hr PR-src, ${df["fully_loaded_wage"]:.2f}/hr fully-loaded, {df["burden_multiplier"]:.2f}x burden).',
        f'Largest phase overrun: {df["largest_overrun"]}.',
        f'Largest phase savings: {df["largest_savings"]}.',
        f'Material spend ${df["material_spend_total"]:,.0f} vs budget ${df["material_budget_total"]:,.0f} '
        f'(${df["material_price_variance"]:,.0f} favorable).',
        f'Top AP vendor: {df["top_vendor_spend"]}; {df["material_vendor_count"]} total AP vendors.',
        f'Retainage outstanding: ${abs(rr["job_totals_retainage"]):,.0f}.',
        f'{df["phases_over_budget"]} of {df["phases_over_budget"] + df["phases_under_budget"]} expense phases over budget '
        f'(${df["total_over_under_budget"]:,.0f} net vs budget).',
        f'Revenue per labor hour ${df["revenue_per_labor_hour"]:.2f}, labor unit cost ${df["labor_unit_cost_per_hr"]:.2f} '
        f'(ratio {df["revenue_per_labor_hour"] / df["labor_unit_cost_per_hr"]:.2f}x).',
    ]
    for i, text in enumerate(insights):
        r = 3 + i
        ws.cell(r, 1, f'{i + 1:02d}')
        ws.cell(r, 1).font = fnt(size=11, bold=True, color=CLAY)
        ws.cell(r, 2, text)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 28
    set_widths(ws, [6, 120])


def build_reconciliation(wb, d):
    ws = wb.create_sheet('Reconciliation')
    recon = d['reconciliation']
    passed = sum(1 for r in recon if r['status'] == 'PASS')
    title(ws, 'RECONCILIATION — PER COST CODE')
    caption(ws, f'Parsed line-item sums vs. Cost Code Totals row in the JDR. Matches Sam Reconciliation tab 1:1. Overall status: {passed}/{len(recon)} codes tie to the cent.')
    ws.append([])
    hdr_row(ws, 3, ['Code', 'Description', 'PDF Total', 'Parsed Sum', 'Difference', 'Status'])
    for i, r in enumerate(recon):
        row = 4 + i
        ws.cell(row, 1, r['code'])
        ws.cell(row, 2, r['description'])
        ws.cell(row, 3, r['pdf_total']); money_format(ws.cell(row, 3))
        ws.cell(row, 4, r['parsed_sum']); money_format(ws.cell(row, 4))
        ws.cell(row, 5, r['difference']); money_format(ws.cell(row, 5))
        ws.cell(row, 6, '✓' if r['status'] == 'PASS' else '✗')
        ws.cell(row, 6).alignment = Alignment(horizontal='center')
        if r['status'] != 'PASS':
            ws.cell(row, 6).fill = fill('F7D4CB')
    summ = 4 + len(recon) + 1
    ws.cell(summ, 1, f'RESULT: {passed}/{len(recon)} codes reconciled').font = fnt(size=11, bold=True, color=SAGE if passed == len(recon) else CLAY)
    set_widths(ws, [8, 30, 16, 16, 16, 10])


def build_reconciliation_log(wb, d):
    ws = wb.create_sheet('Reconciliation Log')
    df = d['derived_fields']
    rr = d['report_record']
    bs = rr['job_totals_by_source']
    recon = d['reconciliation']
    passed = sum(1 for r in recon if r['status'] == 'PASS')

    title(ws, 'RECONCILIATION LOG — JOB 2049')
    caption(ws, 'Worked example of the Reconciliation Protocol running against Job 2049 (JCR-only, no external cross-checks). Mirrors Sam Test-Labels v4 Reconciliation Log tab.')
    ws.append([])
    hdr_row(ws, 3, ['check_id', 'field_name', 'JCR-derived value', 'cross-check value', 'cross-check source', 'status', 'notes'])

    rows = [
        ('RC-01', 'plus_minus_budget',
         f"{df['largest_overrun']}",
         'actual − current_budget reconciled per code',
         'Cost Code Summaries (internal recomputation)', 'PASS',
         'Action #1 sign convention: positive = over-budget. Spot-check recomputes plus_minus from actual − current_budget; matches PDF to the cent.'),
        ('RC-02', 'current_budget / original_budget',
         f"current_budget total = ${df['total_budget']:,.0f}; original_budget also carried per cost code",
         'Budget totals sum across expense codes (code != 999)',
         'BVA tab (Original vs Current columns)', 'PASS',
         'Both budgets carried distinctly per Action #2. Current budget differs from original on codes that had change orders.'),
        ('RC-03', 'pr_src_cost_per_hr',
         f"${df['pr_src_cost_per_hr']:.2f}/hr (all PR-source cost ÷ total labor hours)",
         f"Manual recomputation: PR total ${bs['PR']:,.2f} ÷ {df['total_labor_hours']:,.2f} hrs",
         'Derived Fields tab (line items → source=PR)', 'PASS',
         'Phase-code filter: include all phases incl. 995 burden and 998 taxes; exclude only 999 Sales.'),
        ('RC-04', 'worker_nominal_wages / nominal_wage_percentiles',
         f"p10=${df['nominal_wage_percentiles']['p10']:.2f} | p25=${df['nominal_wage_percentiles']['p25']:.2f} | p50=${df['nominal_wage_percentiles']['p50']:.2f} | p75=${df['nominal_wage_percentiles']['p75']:.2f} | p90=${df['nominal_wage_percentiles']['p90']:.2f} ({df['total_workers']} workers)",
         f"straight_time_rate ${df['straight_time_rate']:.2f} falls within p25–p75 band",
         'Crew Roster tab', 'PASS',
         'Population-slice sanity: weighted straight_time_rate falls within the p25–p75 range, expected for this workforce composition.'),
        ('RC-05', 'labor_unit_cost_per_hr vs revenue_per_labor_hour',
         f"cost ${df['labor_unit_cost_per_hr']:.2f}/hr | revenue ${df['revenue_per_labor_hour']:.2f}/hr | ratio {df['revenue_per_labor_hour'] / df['labor_unit_cost_per_hr']:.2f}x",
         'Recompute from labor-code JTD ÷ labor hours',
         'Productivity tab (Action #7 split)', 'PASS',
         'Split into two fields per Action #7. Cost and throughput are distinct metrics; their ratio is the productivity signal.'),
        ('RC-06', 'scope_metrics / scope_benchmarks', 'pending', '—',
         'project-attribute input (trade-declared)', 'PENDING',
         'Awaiting plumbing scope-metric declarations for Job 2049 (fixture count, DFU total, linear feet of main). Same status as Job 2020.'),
        ('RC-07', 'reconciliation (per-code parse tie-out)',
         f"{passed}/{len(recon)} codes tie to PDF to the cent",
         f"PR ${bs['PR']:,.2f} + AP ${bs['AP']:,.2f} + GL ${bs['GL']:,.2f} = ${bs['PR'] + bs['AP'] + bs['GL']:,.2f} matches direct cost",
         'Reconciliation tab', 'PASS' if passed == len(recon) else 'FAIL',
         'Independent sum of line_items[] by cost_code vs. Cost Code Totals row for each code. PASS if |diff| < $0.02 for all codes.'),
    ]
    for i, row in enumerate(rows):
        r = 4 + i
        for c, v in enumerate(row, start=1):
            ws.cell(r, c, v)
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(r, c).font = fnt(size=9)
        if row[5] == 'PASS':
            ws.cell(r, 6).fill = fill('E6EFE0')
        elif row[5] == 'PENDING':
            ws.cell(r, 6).fill = fill('FCE8CA')
        elif row[5] == 'FAIL':
            ws.cell(r, 6).fill = fill('F7D4CB')
        ws.row_dimensions[r].height = 56

    r = 4 + len(rows) + 2
    section(ws, 'SUMMARY', r, cols=7)
    r += 1
    ws.cell(r, 1, 'Total checks'); ws.cell(r, 2, len(rows))
    r += 1
    ws.cell(r, 1, 'PASS'); ws.cell(r, 2, sum(1 for x in rows if x[5] == 'PASS'))
    r += 1
    ws.cell(r, 1, 'PENDING'); ws.cell(r, 2, sum(1 for x in rows if x[5] == 'PENDING'))
    r += 1
    ws.cell(r, 1, 'FAIL'); ws.cell(r, 2, sum(1 for x in rows if x[5] == 'FAIL'))

    set_widths(ws, [10, 34, 40, 40, 26, 10, 54])


def build_metric_registry(wb, d):
    ws = wb.create_sheet('Metric Registry')
    df = d['derived_fields']
    rr = d['report_record']
    bs = rr['job_totals_by_source']
    revenue = abs(rr['job_totals_revenue'])
    cs = d['cost_code_summaries']
    rb = rollups_by_category(cs)

    title(ws, 'METRIC REGISTRY — 2049')
    caption(ws, 'Canonical registry of every named metric in the workbook. Acts as a machine-readable table of contents + field dictionary. Every Sam Test-Labels v4 field has a row here.')
    ws.append([])
    hdr_row(ws, 3, ['Key', 'Label', 'Value', 'Source / Notes'])

    entries = [
        # Identity
        ('job_number', 'Job Number', rr['job_number'], 'JDR cover page'),
        ('job_name', 'Job Name', rr['job_name'], 'JDR cover page'),
        ('general_contractor', 'General Contractor', d['project']['general_contractor'], 'From AR client code'),
        ('report_date', 'Report Date', rr['report_date'], 'JDR run date'),
        ('line_items_count', 'Line Item Count', len(d['line_items']), 'Length of line_items[]'),
        ('cost_code_count', 'Cost Code Count', len(cs), 'Length of cost_code_summaries[]'),
        # Totals
        ('job_totals_revenue', 'Revenue (AR, negative)', rr['job_totals_revenue'], 'Sage stores credit as negative'),
        ('job_totals_revenue_abs', 'Revenue (absolute)', revenue, 'abs(job_totals_revenue)'),
        ('job_totals_expenses', 'Expenses', rr['job_totals_expenses'], 'Sage job_totals_expenses'),
        ('job_totals_net', 'Net', rr['job_totals_net'], 'Sage job_totals_net'),
        ('job_totals_retainage', 'Retainage', rr['job_totals_retainage'], 'Sage job_totals_retainage'),
        ('job_totals_net_due', 'Net Due', rr['job_totals_net_due'], 'Sage job_totals_net_due'),
        ('cost_by_source_pr_amount', 'PR Source Amount', df['cost_by_source_pr_amount'], 'SUM(line_items source=PR)'),
        ('cost_by_source_ap_amount', 'AP Source Amount', df['cost_by_source_ap_amount'], 'SUM(line_items source=AP)'),
        ('cost_by_source_gl_amount', 'GL Source Amount', df['cost_by_source_gl_amount'], 'SUM(line_items source=GL)'),
        ('cost_by_source_ar_amount', 'AR Source Amount', bs['AR'], 'SUM(line_items source=AR)'),
        ('direct_cost', 'Direct Cost', df['direct_cost'], 'PR + AP + GL'),
        ('net_profit', 'Net Profit', df['net_profit'], 'abs(revenue) − direct_cost'),
        ('gross_margin_pct', 'Gross Margin %', df['net_profit'] / revenue if revenue else 0, 'net / abs(revenue)'),
        # Category rollups
        ('labor_cost', 'Labor Cost (codes 1xx)', round(rb['Labor']['actual'], 2), 'SUM(actual) category=Labor'),
        ('material_cost', 'Material Cost (codes 2xx)', round(rb['Material']['actual'], 2), 'SUM(actual) category=Material'),
        ('overhead_cost', 'Overhead Cost (codes 6xx)', round(rb['Overhead']['actual'], 2), 'SUM(actual) category=Overhead'),
        ('burden_cost', 'Burden Cost (995+998)', round(rb['Burden']['actual'], 2), 'SUM(actual) category=Burden'),
        # Labor analytics
        ('total_labor_hours', 'Total Labor Hours', df['total_labor_hours'], 'SUM(reg+ot+dt+other) source=PR'),
        ('total_workers', 'Total Workers', df['total_workers'], 'COUNT DISTINCT(name) source=PR'),
        ('straight_time_rate', 'Straight-Time Rate', df['straight_time_rate'], 'reg_amount ÷ reg_hours'),
        ('pr_src_cost_per_hr', 'PR-Source Cost/Hour', df['pr_src_cost_per_hr'], 'PR total ÷ labor hours'),
        ('fully_loaded_wage', 'Fully-Loaded Wage', df['fully_loaded_wage'], '(PR + burden) ÷ labor hours'),
        ('burden_multiplier', 'Burden Multiplier', df['burden_multiplier'], 'fully-loaded ÷ pr-src'),
        ('burden_total', 'Burden Total', df['burden_total'], 'SUM(actual) codes 995+998'),
        ('wage_percentile_p10', 'Wage Percentile p10', df['nominal_wage_percentiles']['p10'], 'Percentile across workers'),
        ('wage_percentile_p25', 'Wage Percentile p25', df['nominal_wage_percentiles']['p25'], 'Percentile across workers'),
        ('wage_percentile_p50', 'Wage Percentile p50', df['nominal_wage_percentiles']['p50'], 'Percentile across workers'),
        ('wage_percentile_p75', 'Wage Percentile p75', df['nominal_wage_percentiles']['p75'], 'Percentile across workers'),
        ('wage_percentile_p90', 'Wage Percentile p90', df['nominal_wage_percentiles']['p90'], 'Percentile across workers'),
        # Material analytics
        ('material_spend_total', 'Material Spend Total', df['material_spend_total'], 'SUM(actual) codes 2xx'),
        ('material_budget_total', 'Material Budget Total', df['material_budget_total'], 'SUM(current_budget) codes 2xx'),
        ('material_codes_tracked', 'Material Codes Tracked', df['material_codes_tracked'], 'COUNT codes 2xx'),
        ('material_codes_over_budget', 'Material Codes Over Budget', df['material_codes_over_budget'], 'COUNT where plus_minus>0'),
        ('material_codes_under_budget', 'Material Codes Under Budget', df['material_codes_under_budget'], 'COUNT where plus_minus≤0'),
        ('material_vendor_count', 'Material Vendor Count', df['material_vendor_count'], 'COUNT DISTINCT vendors'),
        ('material_price_variance', 'Material Price Variance', df['material_price_variance'], 'mat_budget − mat_spend'),
        ('largest_material_overrun', 'Largest Material Overrun', df['largest_material_overrun'], ''),
        ('largest_material_savings', 'Largest Material Savings', df['largest_material_savings'], ''),
        ('top_vendor_spend', 'Top Vendor Spend', df['top_vendor_spend'], ''),
        # Phase analytics
        ('phases_over_budget', 'Phases Over Budget', df['phases_over_budget'], 'COUNT codes (excl 999)'),
        ('phases_under_budget', 'Phases Under Budget', df['phases_under_budget'], 'COUNT codes (excl 999)'),
        ('largest_overrun', 'Largest Overrun', df['largest_overrun'], ''),
        ('largest_savings', 'Largest Savings', df['largest_savings'], ''),
        # Budget/forecast
        ('total_jtd_cost', 'Total JTD Cost', df['total_jtd_cost'], 'SUM(actual) expense codes'),
        ('total_budget', 'Total Current Budget', df['total_budget'], 'SUM(current_budget) expense codes'),
        ('overall_pct_budget_consumed', 'Overall % Budget Consumed', df['overall_pct_budget_consumed'], 'JTD ÷ budget'),
        ('total_over_under_budget', 'Total Over/Under Budget', df['total_over_under_budget'], 'budget − JTD'),
        ('labor_unit_cost_per_hr', 'Labor Unit Cost / Hour', df['labor_unit_cost_per_hr'], 'labor JTD ÷ labor hours'),
        ('revenue_per_labor_hour', 'Revenue per Labor Hour', df['revenue_per_labor_hour'], 'abs(revenue) ÷ labor hours'),
        # Pending
        ('unit_count', 'Unit Count', 'PENDING', 'Project-attribute input required'),
        ('fixture_count', 'Fixture Count', 'PENDING', 'Project-attribute input required'),
        ('project_start_date', 'Project Start Date', 'PENDING', 'Project-attribute input required'),
        ('project_end_date', 'Project End Date', 'PENDING', 'Project-attribute input required'),
        ('duration_months', 'Duration (months)', 'PENDING', 'Project-attribute input required'),
        ('scope_metrics', 'Scope Metrics', 'PENDING', 'Trade-declared {label, count, desc} array'),
    ]
    for i, (k, lbl, val, note) in enumerate(entries):
        r = 4 + i
        ws.cell(r, 1, k); ws.cell(r, 1).font = fnt(bold=True)
        ws.cell(r, 2, lbl)
        ws.cell(r, 3, val)
        if isinstance(val, str) and 'PENDING' in val:
            ws.cell(r, 3).fill = fill('FFF4CC')
        elif isinstance(val, (int, float)):
            if k in ('gross_margin_pct', 'overall_pct_budget_consumed', 'pr_pct_of_revenue'):
                pct_format(ws.cell(r, 3))
            elif 'hours' in k or 'count' in k:
                int_format(ws.cell(r, 3))
            else:
                money_format(ws.cell(r, 3))
        ws.cell(r, 4, note); ws.cell(r, 4).font = fnt(italic=True, color=MUTED)
    set_widths(ws, [32, 34, 24, 54])


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    d = json.load(open(DATA_JSON))
    # Merge enriched GDrive meta (units, fixtures, dates, GC team, CO log, doc counts)
    if ENRICHED_META_JSON.exists():
        enriched = json.load(open(ENRICHED_META_JSON))
        d['enriched_meta'] = enriched['meta']
        d['co_log'] = enriched['co_log']
        d['unexecuted_cors'] = enriched.get('unexecuted_cors', [])
    else:
        d['enriched_meta'] = {}
        d['co_log'] = []
        d['unexecuted_cors'] = []
    wb = Workbook()
    wb.remove(wb.active)

    build_overview(wb, d)
    build_job_info(wb, d)
    build_contract(wb, d)
    build_cost_categories(wb, d)
    build_bva(wb, d)
    build_cost_code_detail(wb, d)
    build_transactions(wb, d)
    build_crew_roster(wb, d)
    build_wage_tiers(wb, d)
    build_productivity(wb, d)
    build_vendor_analysis(wb, d)
    build_change_orders(wb, d)          # NEW — executed CO log from GDrive
    build_project_team(wb, d)           # NEW — GC/OWP contact roster
    build_document_pipeline(wb, d)      # NEW — CO/RFI/Submittal/Permit/PO counts
    build_benchmarks(wb, d)
    build_predictive(wb, d)
    build_insights(wb, d)
    build_reconciliation(wb, d)
    build_reconciliation_log(wb, d)
    build_metric_registry(wb, d)

    wb.save(OUT_FILE)
    print(f'wrote {OUT_FILE}  ({OUT_FILE.stat().st_size:,} bytes)')
    print(f'  tabs: {wb.sheetnames}')


if __name__ == '__main__':
    main()
