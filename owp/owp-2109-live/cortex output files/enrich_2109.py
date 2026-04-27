#!/usr/bin/env python3
"""
enrich_2109.py — Enrich OWP_2109_JCR_Cortex_v2.xlsx with key players + populate
sparse tabs.  Mirrors enrich_2020.py but adapted to the 17-numbered-tab schema
used by live projects.

Inputs:
  • 2109_data.json            — Sage JDR parse (rich, ~3.7MB)
  • 2109_dashboard_arrays.json — 16 arrays mirroring index.html PROJECTS['2109']

Outputs (in-place):
  • OWP_2109_JCR_Cortex_v2.xlsx — Job Info tab rebuilt with sectioned key-players
    + project metadata; Overview tab gains a Project Team block; status text
    corrected from the cancelled-template default to "ACTIVE · RI phase".

Project context (2026-04-27):
  Marysville Phase 2 · Intracorp · 246 units · ACTIVE (Roughin in progress)
  $3,570,012 billed to date · $3,718,851 final contract (revised)
  Foreman: Garrett / Mitch · No GDrive folder; team data from index.html
  PROJECT_TEAMS['2109'] + Sage JDR identity block.
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2109_data.json"
DASH_FILE  = SCRIPT_DIR / "2109_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2109_JCR_Cortex_v2.xlsx"

# ============================================================================
# Project-team metadata
# ============================================================================
META_2109 = {
    "job_id":          "2109",
    "name":            "Marysville Phase 2",
    "short_name":      "Marysville Ph2",
    "gc":              "Intracorp",
    "gc_pm":           "TBD (Intracorp) — primary contact pending",
    "gc_sup":          "TBD",
    "gc_pe":           "TBD",
    "owp_ri_foreman":  "Garrett Wilson / Mitchell Gerard",
    "owp_trim_foreman":"TBD (Trim phase not yet started — schedule Q3 2026)",
    "owp_signatory":   "Richard Donelson",
    "owp_estimator":   "Jeffrey S. Gerard / Jordan E. Gerard",
    "owner":           "Intracorp (owner-GC)",
    "developer":       "Intracorp",
    "architect":       "TBD",
    "structural":      "TBD",
    "mep_engineer":    "Franklin Engineering (per AP — $22.3k design fee)",
    "civil":           "TBD",
    "landscape":       "TBD",
    "site_address":    "Marysville, WA",
    "permit":          "TBD",
    "insurance":       "Standard (COI) — non-Wrap",
    "lien_position":   "Standard subcontract — no priority docs filed",
    "warranty":        "Standard 1-yr plumbing (post-completion)",
    "delivery_route":  "TBD",
    "contract_doc":    "Subcontract executed (revised contract $3,718,851)",
    "gdrive_status":   "GDrive folder not yet wired to local mount. Team data from "
                        "index.html PROJECT_TEAMS['2109'] + JDR identity block.",
    "status_text":     "ACTIVE · Roughin phase",
    "status_severity": "HEALTHY",
    "ar_billed":       3570012,
    "direct_cost":     2683501,
    "net_profit":      886511,
    "retention":       153880,
    "contract_orig":   3700937,
    "contract_final":  3718851,
    "co_net":          17914,
    "units":           246,
    "fixtures":        "TBD (fixture schedule not yet finalized in this dataset)",
    "project_type":    "Multi-family apartments — Phase 2",
    "schedule":        "Apr 2026 → Feb 2028 (est. ~22 months) · Roughin in progress",
    "expected_start":  "2026-04 (active)",
    "expected_finish": "2028-02 (estimated)",
    "duration_months": 22,
    "workers":         30,
    "hours":           377,
    "top_vendor":      "Pacific Plumbing Supply Co., LLC ($500,869 / 169 invoices / 46.7% of AP)",
    "ap_total":        1071980,
    "ap_vendor_count": 13,
}

# ============================================================================
# Style helpers (same as enrich_2108)
# ============================================================================
INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
ACTIVE_FONT = Font(name="Arial", size=10, color="155724", bold=True)

HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
ROW_ALT     = PatternFill("solid", start_color="FAFAFA")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
ACTIVE_FILL = PatternFill("solid", start_color="D4EDDA")
WARN_FILL   = PatternFill("solid", start_color="FFF3CD")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")


def clear_range(ws, r1, r2, c1=1, c2=10):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).border = Border()


# ============================================================================
# JOB INFO tab — full rebuild
# ============================================================================
def rebuild_job_info(ws, dash):
    clear_range(ws, 1, ws.max_row + 5, 1, 6)

    ws["A1"] = f"JOB #{META_2109['job_id']} · PROJECT INFORMATION & KEY PLAYERS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Active project — Roughin phase. Team data sourced from "
                "index.html PROJECT_TEAMS['2109'] + Sage JDR identity block + "
                "AP vendor concentration analysis. GDrive folder not yet mounted; "
                "design-team contacts will be filled in once accessed.")
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    forecast_margin = (META_2109["contract_final"] - META_2109["direct_cost"]) / max(META_2109["contract_final"], 1)
    pct_complete = META_2109["ar_billed"] / max(META_2109["contract_final"], 1)

    sections = [
        ("IDENTITY", [
            ("Job Number",    META_2109["job_id"]),
            ("Job Name",      META_2109["name"]),
            ("Project Type",  META_2109["project_type"]),
            ("Site Address",  META_2109["site_address"]),
            ("Permit",        META_2109["permit"]),
            ("Status",        META_2109["status_text"]),
        ]),
        ("SCHEDULE", [
            ("Project Start",     META_2109["expected_start"]),
            ("Expected End",      META_2109["expected_finish"]),
            ("Duration (est.)",   f"{META_2109['duration_months']} months"),
            ("Schedule Note",     META_2109["schedule"]),
        ]),
        ("CONTRACT & FINANCIALS", [
            ("Original Contract", f"${META_2109['contract_orig']:,}"),
            ("Net Change Orders", f"${META_2109['co_net']:,}"),
            ("Final Contract",    f"${META_2109['contract_final']:,}"),
            ("AR Billed to Date", f"${META_2109['ar_billed']:,}"),
            ("% Complete (billed/revised)", f"{pct_complete*100:.1f}%"),
            ("Direct Cost",       f"${META_2109['direct_cost']:,}"),
            ("Forecast Net Profit", f"${META_2109['net_profit']:,}"),
            ("Forecast Margin",   f"{forecast_margin*100:.1f}%"),
            ("Retainage Held",    f"${META_2109['retention']:,}"),
            ("Insurance",         META_2109["insurance"]),
            ("Lien Position",     META_2109["lien_position"]),
            ("Warranty",          META_2109["warranty"]),
            ("Contract on File",  META_2109["contract_doc"]),
        ]),
        ("SCOPE", [
            ("Plumbing Units",  META_2109["units"]),
            ("Total Fixtures",  META_2109["fixtures"]),
            ("Floors",          "TBD"),
            ("Fixture Counts",  "TBD"),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR", [
            ("General Contractor",  META_2109["gc"]),
            ("GC Project Manager",  META_2109["gc_pm"]),
            ("GC Superintendent",   META_2109["gc_sup"]),
            ("GC Project Engineer", META_2109["gc_pe"]),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2109["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2109["owp_trim_foreman"]),
            ("OWP Estimator (takeoff)", META_2109["owp_estimator"]),
            ("OWP Signatory",       META_2109["owp_signatory"]),
            ("Active Crew",         f"{META_2109['workers']} workers · {META_2109['hours']} hrs logged YTD"),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record", META_2109["owner"]),
            ("Developer",       META_2109["developer"]),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",          META_2109["architect"]),
            ("Structural Engineer", META_2109["structural"]),
            ("MEP / Plumbing Engineer", META_2109["mep_engineer"]),
            ("Civil Engineer",     META_2109["civil"]),
            ("Landscape",          META_2109["landscape"]),
        ]),
        ("AP VENDOR PROFILE (CURRENT)", [
            ("Top Vendor",       META_2109["top_vendor"]),
            ("Total AP Spend",   f"${META_2109['ap_total']:,}"),
            ("Active Vendors",   f"{META_2109['ap_vendor_count']} unique vendors"),
            ("Concentration Note", "Pacific Plumbing 46.7% + Consolidated 28.5% + Rosen 9.7% = "
                                   "Top 3 dominate at 84.9% of AP. Standard for OWP active jobs."),
        ]),
        ("DOCUMENT META (current)", [
            ("Pay Apps (filed)",   "Tracked in JDR (no separate filing log mounted)"),
            ("Executed COs",       "Net +$17,914 booked (CO detail not yet itemized in this dataset)"),
            ("CORs",               "TBD — pending log extraction"),
            ("RFIs",               "TBD"),
            ("Submittals",         "TBD"),
            ("POs",                "TBD"),
            ("Permit Count",       "TBD"),
            ("Notes",              "Project is in active Roughin. Documents flow has begun but the "
                                   "GDrive folder for this job hasn't been wired to the dashboard "
                                   "yet — once mounted, the document counts will populate."),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",            "2109 Job Detail Report (Sage Timberline · Apr 3 2026 run)"),
            ("Parsed Data",        "2109_data.json (~3.7MB — full Sage extract)"),
            ("Dashboard Arrays",   "2109_dashboard_arrays.json (16 arrays sourced from index.html PROJECTS['2109'])"),
            ("GDrive Folder",      "Not yet mounted to local environment"),
            ("GDrive Status",      META_2109["gdrive_status"]),
        ]),
        ("RISK FLAGS / ACTIONS", [
            ("New GC (Intracorp)",  "OWP has no closed-job history with Intracorp. "
                                    "Pull AR aging report monthly. Watch payment cadence + "
                                    "change-order velocity."),
            ("Engineering overrun (601)", "Code 601 Engineering/Plans — actual $22,320 vs revised "
                                          "$5,850 (+281.5%). All booked to Franklin Engineering. "
                                          "Likely an early scope-add. Confirm CO coverage before trim."),
            ("Trim manpower forecast", "Roughin in progress. Trim phase typically starts at ~60% "
                                       "schedule completion. Coordinate with manpower allocation "
                                       "matrix for Q3 2026 crew availability."),
            ("Retention aging",     f"${META_2109['retention']:,} held · 4.3% of billed. Track "
                                    "release schedule once project nears closeout."),
        ]),
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        for label, val in fields:
            ws.cell(row=r, column=1, value=label).font = INK_BOLD
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_v.font = INK
            cell_v.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=1).border = BORDER
            cell_v.border = BORDER
            if section_name.startswith("PROJECT TEAM") or section_name.startswith("RISK") or section_name.startswith("AP VENDOR"):
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v.fill = TEAM_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28


# ============================================================================
# OVERVIEW tab — fix CANCELLED status; add Project Team block
# ============================================================================
def patch_overview_team_block(ws):
    for r in range(1, 12):
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "CANCELLED" in v:
                ws.cell(row=r, column=c).value = v.replace(
                    "CANCELLED", META_2109["status_text"])
                ws.cell(row=r, column=c).font = ACTIVE_FONT

    start = 40
    clear_range(ws, start, start + 30, 1, 8)

    ws.cell(row=start, column=1, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)

    ws.cell(row=start+1, column=1, value=(
        "Active project · Roughin phase. Team data from index.html PROJECT_TEAMS['2109'] "
        "+ JDR + AP vendor concentration."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=7)

    rows = [
        ("General Contractor",     META_2109["gc"]),
        ("OWP Roughin Foreman",    META_2109["owp_ri_foreman"]),
        ("OWP Estimator",          META_2109["owp_estimator"]),
        ("OWP Signatory",          META_2109["owp_signatory"]),
        ("Owner",                  META_2109["owner"]),
        ("MEP / Plumbing Engineer", META_2109["mep_engineer"]),
        ("Site Address",           META_2109["site_address"]),
        ("Insurance",              META_2109["insurance"]),
        ("Status",                 META_2109["status_text"]),
        ("Original Contract",      f"${META_2109['contract_orig']:,}"),
        ("Final Contract (rev)",   f"${META_2109['contract_final']:,}"),
        ("AR Billed",              f"${META_2109['ar_billed']:,}"),
        ("Retention Held",         f"${META_2109['retention']:,}"),
        ("Forecast Margin",        f"{(META_2109['contract_final']-META_2109['direct_cost'])/META_2109['contract_final']*100:.1f}%"),
        ("Active Crew",            f"{META_2109['workers']} workers · {META_2109['hours']} hrs"),
    ]
    r = start + 3
    ws.cell(row=r, column=1, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=1, value=role).font = INK_BOLD
        ws.cell(row=r, column=2, value=val).font = INK
        ws.cell(row=r, column=1).fill = TEAM_FILL
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1


# ============================================================================
# PREDICTIVE SIGNALS tab — expand
# ============================================================================
def patch_predictive_signals(ws, dash):
    signals = dash.get("predictiveSignals", [])
    clear_range(ws, 1, 30, 1, 6)
    ws["A1"] = f"Predictive Signals — Job #{META_2109['job_id']} (active · roughin)"
    ws["A1"].font = TITLE_FONT

    headers = ["#", "Signal", "Current Value", "Threshold", "Severity"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    extra = [
        ["Project status",            META_2109["status_text"], "ACTIVE",          "HEALTHY"],
        ["% Complete (billed/revised)", f"{META_2109['ar_billed']/META_2109['contract_final']*100:.1f}%", "Reference", "INFO"],
        ["GC closed-job history",     "0 jobs with Intracorp",  ">=1",                "WATCH"],
        ["Engineering overrun (601)", "+281.5%",                 "<+50%",              "CRIT"],
        ["Top-vendor share",          "46.7% (Pacific Plumbing)", "<50%",              "INFO"],
        ["Active crew size",          f"{META_2109['workers']} workers", "Reference",  "INFO"],
        ["Trim phase coordination",   "Q3 2026 forecast",        "Plan early",          "INFO"],
    ]
    rows = [s if isinstance(s, list) else list(s) for s in signals] + extra
    r = 4
    for i, s in enumerate(rows, 1):
        if not isinstance(s, list) or len(s) < 4: continue
        signal, val, thr, sev = s[0], s[1], s[2], s[3]
        ws.cell(row=r, column=1, value=i).font = INK
        ws.cell(row=r, column=2, value=signal).font = INK
        ws.cell(row=r, column=3, value=val).font = INK
        ws.cell(row=r, column=4, value=thr).font = INK
        sev_cell = ws.cell(row=r, column=5, value=sev)
        sev_cell.font = INK_BOLD
        if sev in ("WARN", "WATCH"): sev_cell.fill = WARN_FILL
        elif sev in ("CRIT", "CRITICAL"): sev_cell.fill = PatternFill("solid", start_color="F8D7DA")
        elif sev == "HEALTHY": sev_cell.fill = ACTIVE_FILL
        elif sev == "INFO": sev_cell.fill = PatternFill("solid", start_color="D1ECF1")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
        r += 1

    widths = {1: 5, 2: 38, 3: 22, 4: 24, 5: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# Patch CHANGE LOG (workbook log)
# ============================================================================
def patch_change_log(ws):
    last_row = ws.max_row
    r = last_row + 1
    ws.cell(row=r, column=1, value="2026-04-27").font = INK
    ws.cell(row=r, column=2, value="v1.1 · enriched").font = INK_BOLD
    ws.cell(row=r, column=3, value=(
        "Enriched Job Info tab with 12-section sectioned layout (identity / schedule / "
        "contract & financials / scope / project team — 4 sub-sections / AP vendor "
        "profile / document meta / data sources / risk flags + actions). Corrected "
        "status from 'CANCELLED' (template default) to 'ACTIVE · Roughin phase'. "
        "Predictive Signals expanded with active-project signals."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP


# ============================================================================
# Main
# ============================================================================
def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)
    data = json.loads(JSON_FILE.read_text()) if JSON_FILE.exists() else {}
    dash = json.loads(DASH_FILE.read_text())

    print("→ Rebuilding Job Info tab (sectioned layout)...")
    rebuild_job_info(wb["02 Job Info"], dash)

    print("→ Patching Overview status + adding Project Team block...")
    patch_overview_team_block(wb["01 Overview"])

    print("→ Expanding Predictive Signals tab...")
    patch_predictive_signals(wb["14 Predictive Signals"], dash)

    print("→ Logging enrichment pass to Change Log...")
    patch_change_log(wb["17 Change Log"])

    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
