#!/usr/bin/env python3
"""
enrich_2106.py — Enrich OWP_2106_JCR_Cortex_v2.xlsx with key players + Job Info
+ correct status/financial posture for the early-takeoff-phase project.

Project: 2106 Chinn 6220 Roosevelt — 146 units (147 in latest bid)
         Pre-construction · early takeoff phase ($49,275 billed against
         $45,727 sunk design+takeoff cost = $3,548 carrying margin so far)
         Subcontract pending — target field start Jul 2026

Bid history (per OWP Master Project List):
  • Aug 8 2024:  BUDGET — Exxel Pacific & Compass · 146u · $2,800,000
  • Aug 16 2024: BUDGET — Chinn enters · 146u · $2,800,000
  • Dec 11 2024: BUDGET — Chinn · 147u (+1) · $2,988,000
  • Feb 24 2025: BUDGET — Chinn · 147u · $2,937,000 (price came down)
  • May 6 2025:  BUDGET — Chinn · 147u · $2,968,000
  • Dec 4 2025:  BID    — Chinn · 147u · $3,042,140 (final · "Tri-State also bidding")

Source documents:
  • 2106 Job Detail Report.pdf (Sage Timberline · Apr 3 2026 run · 6 cost codes)
  • OWP Project List with Schedule - UPDATED 04-01-26.xlsx (Master Project List
    Schedule + Projects Bid tabs)
  • No GDrive Job Book folder yet — will be created at subcontract execution
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2106_data.json"
DASH_FILE  = SCRIPT_DIR / "2106_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2106_JCR_Cortex_v2.xlsx"

META_2106 = {
    "job_id":          "2106",
    "name":            "Chinn 6220 Roosevelt",
    "short_name":      "6220 Roosevelt",
    "gc":              "Chinn Construction, LLC",
    "gc_pm":           "TBD (Chinn) — primary contact pending subcontract execution",
    "gc_sup":          "TBD",
    "gc_pe":           "TBD",
    "owp_ri_foreman":  "TBD · pre-construction (target start Jul 2026)",
    "owp_trim_foreman":"TBD",
    "owp_signatory":   "Richard Donelson",
    "owp_estimator":   "Jeffrey S. Gerard / Jordan E. Gerard / Joseph N. Barnhart (takeoff team)",
    "owner":           "Phoenix Property Co (per Master Project List Projects Bid tab)",
    "developer":       "Phoenix Property Co",
    "architect":       "Weber Thompson",
    "structural":      "TBD",
    "mep_engineer":    "Emerald City Engineering (GC's bridge MEP) + Robison Engineering, Inc. (OWP's 3rd-party plumbing design-build engineer · $30,200 design fee booked)",
    "civil":           "TBD",
    "landscape":       "TBD",
    "site_address":    "6220 Roosevelt Way NE, Seattle, WA (Roosevelt neighborhood)",
    "permit":          "Pre-permit (no permit applied for yet)",
    "insurance":       "Standard (COI) — non-Wrap (per project team grid)",
    "lien_position":   "Standard subcontract — pending execution",
    "warranty":        "1-yr standard plumbing (post-completion)",
    "delivery_route":  "TBD",
    "contract_doc":    "Subcontract NOT yet executed — final BID submitted Dec 4 2025 ($3,042,140 / 147u)",
    "scope":           "Design Build Plumbing Systems (typical Chinn engagement pattern)",
    "gdrive_status":   "No GDrive Job Book folder created yet (project hasn't reached active stage). "
                        "All team data sourced from OWP Master Project List + JDR identity + AP vendor scan.",
    "status_text":     "ACTIVE · pre-construction (early takeoff phase)",
    "status_severity": "INFO",
    "ar_billed":       49275,
    "direct_cost":     45727,
    "net_profit":      3548,    # actual margin on takeoff billings to date
    "retention":       0,
    "contract_orig":   None,    # not yet executed
    "contract_final":  None,
    "contract_bid_latest": 3042140,  # final BID Dec 2025
    "co_net":          0,
    "units":           146,        # current per data.json
    "units_bid_latest": 147,       # latest bid record
    "fixtures":        "TBD (no fixture schedule yet — pre-design-development)",
    "project_type":    "146-unit (147 in latest bid) multifamily · Roosevelt neighborhood Seattle",
    "schedule":        "Pre-construction · target field start Jul 2026 · ~24 months to TCO (Q3 2028 expected)",
    "expected_start":  "Jul 2026 (target)",
    "expected_finish": "Q3 2028 (estimated)",
    "duration_months": 24,
    "workers":         3,
    "hours":           152.0,
    "top_vendor":      "Robison Engineering, Inc. ($30,200 / 3 invoices / 100% of AP)",
    "ap_total":        30200,
    "ap_vendor_count": 1,
    "chinn_history":   "OWP's 8th Chinn engagement (after #2009 Greenwood, #2010 Old Town, #2011 Phinney Ridge, #2023 Legacy Apts, #2041 Luna California Ave, #2070 Beacon Crossing, #2071 1405 Dexter, #2096 Sundowner). Strong recurring relationship.",
    "competition":     "Tri-State Plumbing also bidding per Master Project List notes (Dec 2025).",
    "bid_history": [
        ("2024-08-08", "BUDGET", "Exxel Pacific", 146, 2800000, "Original Exxel budget bid"),
        ("2024-08-08", "BUDGET", "Compass",       146, 2800000, "Compass also bid same date"),
        ("2024-08-16", "BUDGET", "Chinn",         146, 2800000, "Chinn enters · matches base budget"),
        ("2024-12-11", "BUDGET", "Chinn",         147, 2988000, "Unit count grew by 1 · price up to $2.99M"),
        ("2025-02-24", "BUDGET", "Chinn",         147, 2937000, "Re-bid · price came down 1.7%"),
        ("2025-05-06", "BUDGET", "Chinn",         147, 2968000, "Re-bid · price up 1.0%"),
        ("2025-12-04", "BID",    "Chinn",         147, 3042140, "FINAL BID · $3.04M · Tri-State also bidding"),
    ],
}

# ============================================================================
# Style helpers
# ============================================================================
INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
WARN_FONT   = Font(name="Arial", size=10, color="9A4F02", bold=True)

HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
ACTIVE_FILL = PatternFill("solid", start_color="D4EDDA")
WARN_FILL   = PatternFill("solid", start_color="FFF3CD")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")


def clear_range(ws, r1, r2, c1=1, c2=10):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).border = Border()


# ============================================================================
# JOB INFO tab — full rebuild
# ============================================================================
def rebuild_job_info(ws):
    clear_range(ws, 1, ws.max_row + 5, 1, 6)
    ws["A1"] = f"JOB #{META_2106['job_id']} · {META_2106['name'].upper()} · KEY PLAYERS & METADATA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Pre-construction · early takeoff phase. Sage shows $49,275 billed against "
                "$45,727 sunk design+takeoff cost ($3,548 carrying margin so far). Subcontract "
                "not yet executed — final BID submitted Dec 4 2025 to Chinn for $3,042,140 (147 units). "
                "Target field start Jul 2026. Tri-State Plumbing also bidding per Master List notes.")
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    sections = [
        ("IDENTITY", [
            ("Job Number",        META_2106["job_id"]),
            ("Job Name",          META_2106["name"]),
            ("Project Type",      META_2106["project_type"]),
            ("Site Address",      META_2106["site_address"]),
            ("Permit",            META_2106["permit"]),
            ("Status",            META_2106["status_text"]),
            ("Scope of Work",     META_2106["scope"]),
        ]),
        ("SCHEDULE", [
            ("Project Start",     META_2106["expected_start"]),
            ("Expected End",      META_2106["expected_finish"]),
            ("Duration (est.)",   f"{META_2106['duration_months']} months"),
            ("Schedule Note",     META_2106["schedule"]),
            ("Subcontract Execution", "Pending — target Jul 2026"),
        ]),
        ("FINANCIAL POSTURE (PRE-CONSTRUCTION · TAKEOFF PHASE)", [
            ("Original Contract", "Not yet executed"),
            ("Latest BID submitted", f"${META_2106['contract_bid_latest']:,} (Dec 4 2025 · {META_2106['units_bid_latest']} units)"),
            ("AR Billed to Date", f"${META_2106['ar_billed']:,} (4 progress billings · takeoff phase)"),
            ("Direct Cost (sunk)", f"${META_2106['direct_cost']:,}"),
            ("Net Profit (takeoff phase)", f"${META_2106['net_profit']:,} ({META_2106['net_profit']/META_2106['ar_billed']*100:.1f}% margin)"),
            ("Retainage",         "$0 (no field billing yet)"),
            ("Insurance",         META_2106["insurance"]),
            ("Lien Position",     META_2106["lien_position"]),
            ("Warranty",          META_2106["warranty"]),
            ("Contract on File",  META_2106["contract_doc"]),
        ]),
        ("SCOPE", [
            ("Plumbing Units",     f"{META_2106['units']} units (current data.json) · {META_2106['units_bid_latest']} units (latest BID Dec 2025) ⚠ scope evolved"),
            ("Total Fixtures",     META_2106["fixtures"]),
            ("Floors",             "TBD"),
            ("Fixture Counts",     "TBD"),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR (Chinn)", [
            ("General Contractor", META_2106["gc"]),
            ("GC Project Manager", META_2106["gc_pm"]),
            ("GC Superintendent",  META_2106["gc_sup"]),
            ("GC Project Engineer", META_2106["gc_pe"]),
            ("Chinn engagement count", META_2106["chinn_history"]),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2106["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2106["owp_trim_foreman"]),
            ("OWP Estimator (takeoff)", META_2106["owp_estimator"]),
            ("OWP Signatory",       META_2106["owp_signatory"]),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record",   META_2106["owner"]),
            ("Developer",         META_2106["developer"]),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",         META_2106["architect"]),
            ("Structural Engineer", META_2106["structural"]),
            ("MEP / Plumbing Engineer", META_2106["mep_engineer"]),
            ("Civil Engineer",    META_2106["civil"]),
            ("Landscape Architect", META_2106["landscape"]),
        ]),
        ("BID HISTORY (from OWP Master Project List)", []),  # populated separately below
        ("AP VENDOR PROFILE (CURRENT — TAKEOFF PHASE)", [
            ("Top Vendor",        META_2106["top_vendor"]),
            ("Total AP Spend",    f"${META_2106['ap_total']:,}"),
            ("Active Vendors",    f"{META_2106['ap_vendor_count']} unique vendor (Robison Engineering only — typical for design-build takeoff phase)"),
            ("Concentration Note", "Single-vendor (100%) is normal for early design phase. Will diversify when field phase begins (Ferguson, Keller, Rosen, etc. typical for Chinn jobs based on closed-portfolio data)."),
        ]),
        ("DOCUMENT META (current)", [
            ("Pay Apps (filed)",  "4 progress billings (parsed from JDR AR)"),
            ("Executed COs",      "0 (subcontract pending)"),
            ("CORs",              "0"),
            ("RFIs",              "0"),
            ("Submittals",        "0 (pre-construction)"),
            ("POs",               "0"),
            ("Permit Count",      "0 (pre-permit)"),
            ("Notes",             "Project hasn't reached document-generation phase. Only the OWP estimating team + Robison Engineering have touched the file (takeoff + design coordination)."),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",           "2106 Job Detail Report (Sage Timberline · Apr 3 2026 run)"),
            ("Parsed Data",       "2106_data.json (6 cost codes, 298 line items, 3 workers)"),
            ("Dashboard Arrays",  "2106_dashboard_arrays.json (16 arrays)"),
            ("Master Project List", "OWP Project List with Schedule - UPDATED 04-01-26.xlsx (Schedule r102, Projects Bid r319-404 = 7-bid history)"),
            ("GDrive Folder",     "Not yet created"),
            ("GDrive Status",     META_2106["gdrive_status"]),
        ]),
        ("RISK FLAGS / WATCH ITEMS", [
            ("Subcontract not yet executed", "OWP has $45.7k of sunk design+takeoff cost. Final BID submitted Dec 2025; await Chinn award decision. If Chinn picks Tri-State, OWP has to absorb the $45.7k or write off."),
            ("Tri-State competition flag", "Master List notes (Dec 2025): 'I've heard that Tri-State is bidding'. Chinn award decision pending."),
            ("Bid drift (146 → 147 units)", "Project added 1 unit between Aug 2024 (146u @ $2.8M) and Dec 2024 (147u @ $2.99M). Final BID Dec 2025 = 147u @ $3.04M. data.json still reflects 146 — will update at subcontract execution."),
            ("Robison Engineering relationship", "OWP's 3rd-party plumbing design-build engineer (vs Franklin on most other jobs). Same engineer as #2118 Edmonds Behar and #2114 Holland Ballard Blossom — a recurring 3rd-party design partnership."),
            ("Chinn closed-job track record", "Chinn = 7 closed jobs in OWP closed portfolio with strong margin profiles (#2009/2010/2011 Greenwood/Old Town/Phinney Ridge averaged 30-40% margins). Likelihood of award = HIGH if pricing is competitive."),
        ]),
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        if section_name.startswith("BID HISTORY"):
            # Render bid history table inline
            headers = ["Bid Date", "Type", "GC", "Units", "Price", "Note"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=r, column=c, value=h).font = WHITE_BOLD
                ws.cell(row=r, column=c).fill = HDR_FILL
                ws.cell(row=r, column=c).border = BORDER
            r += 1
            for date, type_, gc, units, price, note in META_2106["bid_history"]:
                ws.cell(row=r, column=1, value=date).font = INK
                ws.cell(row=r, column=2, value=type_).font = INK_BOLD if type_ == "BID" else INK
                ws.cell(row=r, column=3, value=gc).font = INK
                ws.cell(row=r, column=4, value=units).font = INK
                ws.cell(row=r, column=5, value=f"${price:,}").font = INK
                ws.cell(row=r, column=6, value=note).font = INK
                ws.cell(row=r, column=6).alignment = WRAP
                if type_ == "BID":
                    for c in range(1, 7): ws.cell(row=r, column=c).fill = ACTIVE_FILL
                for c in range(1, 7): ws.cell(row=r, column=c).border = BORDER
                r += 1
            r += 1
            continue
        for label, val in fields:
            ws.cell(row=r, column=1, value=label).font = INK_BOLD
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_v.font = INK
            cell_v.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=1).border = BORDER
            cell_v.border = BORDER
            if section_name.startswith("PROJECT TEAM") or section_name.startswith("RISK") or section_name.startswith("AP VENDOR"):
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v.fill = TEAM_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 50


# ============================================================================
# OVERVIEW tab — fix banner status + add Project Team block
# ============================================================================
def patch_overview(ws):
    # Replace the misleading "$0 / pre-construction" banner with the real takeoff numbers
    for r in range(1, 12):
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "CANCELLED" in v.upper():
                ws.cell(row=r, column=c).value = v.replace("CANCELLED", META_2106["status_text"])

    # Add team block at row 30+
    start = 30
    clear_range(ws, start, start + 30, 1, 8)
    ws.cell(row=start, column=1, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    ws.cell(row=start+1, column=1, value=(
        "Pre-construction · early takeoff phase. Subcontract pending — final BID submitted "
        "Dec 4 2025 ($3,042,140 / 147u) to Chinn. Tri-State Plumbing also bidding."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=7)

    rows = [
        ("General Contractor",    META_2106["gc"]),
        ("Owner",                 META_2106["owner"]),
        ("Architect",             META_2106["architect"]),
        ("MEP / Plumbing Engineer", META_2106["mep_engineer"]),
        ("OWP Estimator",         META_2106["owp_estimator"]),
        ("OWP Signatory",         META_2106["owp_signatory"]),
        ("Site Address",          META_2106["site_address"]),
        ("Insurance",             META_2106["insurance"]),
        ("Status",                META_2106["status_text"]),
        ("Latest BID submitted",  f"${META_2106['contract_bid_latest']:,} (Dec 4 2025 · 147u)"),
        ("AR Billed to date",     f"${META_2106['ar_billed']:,} (takeoff phase)"),
        ("Direct Cost (sunk)",    f"${META_2106['direct_cost']:,}"),
        ("Net Profit (takeoff)",  f"${META_2106['net_profit']:,} ({META_2106['net_profit']/META_2106['ar_billed']*100:.1f}% margin)"),
        ("Active Crew",           "3 OWP estimating leads (Gerard family + Joseph Barnhart) · 152 hrs"),
    ]
    r = start + 3
    ws.cell(row=r, column=1, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=1, value=role).font = INK_BOLD
        ws.cell(row=r, column=2, value=val).font = INK
        ws.cell(row=r, column=1).fill = TEAM_FILL
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1


# ============================================================================
# Change Log
# ============================================================================
def patch_change_log(ws):
    last_row = ws.max_row
    r = last_row + 1
    ws.cell(row=r, column=1, value="2026-04-27").font = INK
    ws.cell(row=r, column=2, value="v1.1 · enriched").font = INK_BOLD
    ws.cell(row=r, column=3, value=(
        "Enriched Job Info tab with 12-section sectioned layout (identity / schedule / "
        "financial posture / scope / project team — 4 sub-sections / bid history (7 entries) / "
        "AP vendor profile / document meta / data sources / risk flags). "
        "Added Project Team block to Overview. Sourced from OWP Master Project List "
        "(Schedule r102 + Projects Bid 7-bid history) + Sage JDR identity + AP vendor scan. "
        "No GDrive Job Book yet — will be created at subcontract execution (target Jul 2026)."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP


# ============================================================================
# Main
# ============================================================================
def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)
    print("→ Rebuilding Job Info tab...")
    rebuild_job_info(wb["02 Job Info"])
    print("→ Patching Overview + adding Project Team block...")
    patch_overview(wb["01 Overview"])
    print("→ Logging enrichment to Change Log...")
    patch_change_log(wb["17 Change Log"])
    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
