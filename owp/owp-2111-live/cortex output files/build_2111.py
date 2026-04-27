#!/usr/bin/env python3
"""
build_2111.py — Build the 17-tab JCR Cortex v2 workbook for Job 2111
(Compass Northgate M2) — live/active project.

Input:   2111_data.json  (from parse_2111.py)
Output:  OWP_2111_JCR_Cortex_v2.xlsx
"""
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).resolve().parent
DATA_FILE = _HERE / '2111_data.json'
OUT_FILE = _HERE / 'OWP_2111_JCR_Cortex_v2.xlsx'

# ---------- Style primitives ----------
INK = "2C2B29"; INK_3 = "70655C"; CREAM = "F4EDE3"; CLAY = "B85C3E"; SAGE = "6F7E5E"
FONT_NAME = "Arial"
F_H1 = Font(name=FONT_NAME, size=16, bold=True, color=INK)
F_H2 = Font(name=FONT_NAME, size=12, bold=True, color=INK)
F_HDR = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=FONT_NAME, size=10, color=INK)
F_BODY_B = Font(name=FONT_NAME, size=10, bold=True, color=INK)
F_NUM = Font(name=FONT_NAME, size=10, color="000000")
F_NUM_B = Font(name=FONT_NAME, size=10, bold=True, color="000000")
F_NUM_BLUE = Font(name=FONT_NAME, size=10, color="0000FF")
F_NUM_GREEN = Font(name=FONT_NAME, size=10, color="008000")
F_NOTE = Font(name=FONT_NAME, size=9, italic=True, color=INK_3)

FILL_HDR = PatternFill("solid", start_color=INK)
FILL_BAND = PatternFill("solid", start_color="F0EAE0")
FILL_CLAY = PatternFill("solid", start_color="E8CEC3")
FILL_SAGE = PatternFill("solid", start_color="D5DEC5")

THIN = Side(style="thin", color="BFB8AE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
A_CENTER = Alignment(horizontal="center", vertical="center")
A_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
A_RIGHT = Alignment(horizontal="right", vertical="center")

CURR_FMT = '$#,##0;($#,##0);-'
PCT_FMT = '0.0%;(0.0%);-'

LABOR_CODES = {"011","100","101","110","111","112","113","120","130","140","141","142","143","145","150"}
MATERIAL_CODES = {"039","210","211","212","213","220","230","240","241","242","243","244","245","251"}
OVERHEAD_CODES = {"600","601","602","603","604","607"}
BURDEN_CODES = {"995","998"}


# ---------- Helpers ----------
def style_header(ws, row, cols, fill=FILL_HDR, font=F_HDR):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill; cell.font = font; cell.alignment = A_CENTER; cell.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def load_data():
    raw = json.loads(DATA_FILE.read_text())
    code_totals = {}
    for cs in raw.get('cost_code_summaries', []):
        if cs['code'] is None:
            continue
        code_totals[cs['code']] = {
            'desc': cs['description'],
            'orig': cs['original_budget'],
            'rev': cs['current_budget'],
            'actual': cs['actual_amount'],
            'variance': cs['plus_minus_budget'],
            'net_due': cs['net_due'],
            'retainage': cs['retainage'],
            'reg_hours': cs.get('regular_hours', 0),
            'ot_hours': cs.get('overtime_hours', 0),
        }
    transactions = []
    for li in raw.get('line_items', []):
        hours = (li.get('regular_hours') or 0) + (li.get('overtime_hours') or 0) + \
                (li.get('doubletime_hours') or 0) + (li.get('other_hours') or 0)
        transactions.append({
            'src': li['source'],
            'party_name': li.get('name', ''),
            'amount': li.get('actual_amount', 0),
            'hours': hours,
            'cost_code': li.get('cost_code', ''),
            'regular_hours': li.get('regular_hours', 0),
            'overtime_hours': li.get('overtime_hours', 0),
            'regular_amount': li.get('regular_amount', 0),
        })
    by_source = raw.get('report_record', {}).get('job_totals_by_source', {})
    job_totals = {
        'revenues': raw.get('report_record', {}).get('job_totals_revenue', 0),
        'expenses': raw.get('report_record', {}).get('job_totals_expenses', 0),
        'net': raw.get('report_record', {}).get('job_totals_net', 0),
        'net_due': raw.get('report_record', {}).get('job_totals_net_due', 0),
        'retainage': raw.get('report_record', {}).get('job_totals_retainage', 0),
    }
    return {
        'code_totals': code_totals,
        'transactions': transactions,
        'by_source': by_source,
        'job_totals': job_totals,
        'worker_wages': raw.get('worker_wages', []),
        'derived_fields': raw.get('derived_fields', {}),
        'reconciliation': raw.get('reconciliation', []),
    }


# ---------- Tab builders ----------
def tab_overview(wb, data, meta):
    ws = wb.create_sheet("01 Overview")
    set_col_widths(ws, [26, 18, 18, 18, 18])
    ws["A1"] = f"OWP #{meta['job_id']} · {meta['name']}"
    ws["A1"].font = F_H1
    ws["A2"] = f"{meta['gc']} · {meta['units']} units · Live / in-progress · OCIP"
    ws["A2"].font = F_NOTE

    ws["A4"] = "Forecast at Completion (Revised Basis)"
    ws["A4"].font = F_H2
    rows = [
        ("Revised contract value", meta['rev_contract'], CURR_FMT),
        ("Revised budget (expenses)", meta['rev_expense'], CURR_FMT),
        ("Forecast net profit", None, CURR_FMT),
        ("Forecast margin", None, PCT_FMT),
    ]
    r = 5
    for lbl, val, fmt in rows:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY
        cell = ws.cell(row=r, column=2)
        if lbl == "Forecast net profit":
            cell.value = "=B5-B6"; cell.font = F_NUM_B
        elif lbl == "Forecast margin":
            cell.value = "=IFERROR(B7/B5,0)"; cell.font = F_NUM_B
        else:
            cell.value = val; cell.font = F_NUM_BLUE
        cell.number_format = fmt; cell.alignment = A_RIGHT
        r += 1

    ws["A10"] = "Progress to Date"
    ws["A10"].font = F_H2
    prog = [
        ("Work complete (billed)", meta['billed'], CURR_FMT),
        ("Retention held", meta['retention'], CURR_FMT),
        ("Earned less retention", None, CURR_FMT),
        ("Balance to finish", None, CURR_FMT),
        ("Percent complete", None, PCT_FMT),
    ]
    r = 11
    for lbl, val, fmt in prog:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY
        cell = ws.cell(row=r, column=2)
        if lbl == "Earned less retention":
            cell.value = "=B11-B12"; cell.font = F_NUM_B
        elif lbl == "Balance to finish":
            cell.value = "=B5-B11"; cell.font = F_NUM_B
        elif lbl == "Percent complete":
            cell.value = "=IFERROR(B11/B5,0)"; cell.font = F_NUM_B
        else:
            cell.value = val; cell.font = F_NUM_BLUE
        cell.number_format = fmt; cell.alignment = A_RIGHT
        r += 1

    ws["A17"] = "Transaction Totals by Source"
    ws["A17"].font = F_H2
    by_source = data.get('by_source', {})
    r = 18
    ws.cell(row=r, column=1, value="Source").font = F_HDR
    ws.cell(row=r, column=2, value="Dollars").font = F_HDR
    ws.cell(row=r, column=1).fill = FILL_HDR; ws.cell(row=r, column=2).fill = FILL_HDR
    ws.cell(row=r, column=1).alignment = A_CENTER; ws.cell(row=r, column=2).alignment = A_CENTER
    r = 19
    for src in ("GL", "AP", "PR", "AR"):
        ws.cell(row=r, column=1, value=src).font = F_BODY
        c = ws.cell(row=r, column=2, value=by_source.get(src, 0))
        c.font = F_NUM; c.number_format = CURR_FMT; c.alignment = A_RIGHT
        r += 1


def tab_job_info(wb, data, meta):
    ws = wb.create_sheet("02 Job Info")
    set_col_widths(ws, [28, 40])
    ws["A1"] = "Job Information"; ws["A1"].font = F_H1
    rows = [
        ("Job ID", meta['job_id']), ("Project name", meta['name']),
        ("General contractor", meta['gc']), ("Owner", meta['owner']),
        ("Units", meta['units']), ("Insurance", meta['insurance']),
        ("Job start", meta['job_start']), ("Expected completion", meta['expected_completion']),
        ("Architect", meta['architect']), ("MEP Engineer", meta['mep']),
        ("Status", "ACTIVE · in-progress"), ("JDR report date", meta['report_date']),
    ]
    r = 3
    for lbl, val in rows:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY_B
        ws.cell(row=r, column=2, value=val).font = F_BODY
        r += 1


def tab_contract(wb, data, meta):
    ws = wb.create_sheet("03 Contract")
    set_col_widths(ws, [40, 18, 18, 18])
    ws["A1"] = "Contract Summary"; ws["A1"].font = F_H1
    hdr = ["Item", "Original", "Revised", "Delta"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    rows = [
        ("Base construction contract", meta['orig_contract'], meta['rev_contract']),
        ("Executed change orders (net)", 0, meta['rev_contract'] - meta['orig_contract']),
    ]
    r = 4
    for lbl, o, rv in rows:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY
        c2 = ws.cell(row=r, column=2, value=o); c2.number_format = CURR_FMT; c2.font = F_NUM_BLUE; c2.alignment = A_RIGHT
        c3 = ws.cell(row=r, column=3, value=rv); c3.number_format = CURR_FMT; c3.font = F_NUM_BLUE; c3.alignment = A_RIGHT
        c4 = ws.cell(row=r, column=4, value=f"=C{r}-B{r}"); c4.number_format = CURR_FMT; c4.font = F_NUM; c4.alignment = A_RIGHT
        r += 1
    ws.cell(row=r, column=1, value="Total contract value").font = F_BODY_B
    for col in (2, 3, 4):
        letter = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({letter}4:{letter}{r-1})")
        c.number_format = CURR_FMT; c.font = F_NUM_B; c.alignment = A_RIGHT

    co_row = r + 3
    ws.cell(row=co_row, column=1, value="Executed Change Orders").font = F_H2
    co_row += 1
    co_hdr = ["CO #", "Description", "Amount"]
    for i, h in enumerate(co_hdr, start=1): ws.cell(row=co_row, column=i, value=h)
    style_header(ws, co_row, len(co_hdr))
    co_row += 1
    for co in meta.get('cos', []):
        if isinstance(co, dict):
            co_num, co_desc, co_amt = co.get('co', ''), co.get('desc', ''), co.get('amount', 0)
        else:
            co_num, co_desc, co_amt = co[0], co[1], co[2]
        ws.cell(row=co_row, column=1, value=co_num).font = F_BODY
        ws.cell(row=co_row, column=2, value=co_desc).font = F_BODY
        c = ws.cell(row=co_row, column=3, value=co_amt)
        c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        co_row += 1


def tab_sov(wb, data, meta):
    ws = wb.create_sheet("04 SOV-PayApps")
    set_col_widths(ws, [14, 14, 16, 16, 16, 16, 16])
    ws["A1"] = "Schedule of Values / Pay Application History"; ws["A1"].font = F_H1
    hdr = ["Pay App", "Period End", "This Period", "Completed to Date", "Retention Held", "Previous Billed", "Current Balance"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    r = 4
    for pa in meta.get('pay_apps', []):
        ws.cell(row=r, column=1, value=pa[0]).font = F_BODY
        ws.cell(row=r, column=2, value=pa[1]).font = F_BODY
        for col, val in enumerate(pa[2:], start=3):
            c = ws.cell(row=r, column=col, value=val)
            c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        r += 1


def tab_change_orders(wb, data, meta):
    ws = wb.create_sheet("05 Change Orders")
    set_col_widths(ws, [8, 14, 50, 18, 12])
    ws["A1"] = "Change Order Log"; ws["A1"].font = F_H1
    hdr = ["#", "Type", "Description", "Amount", "Status"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    r = 4
    for co in meta.get('change_log', []):
        for i, v in enumerate(co, start=1):
            c = ws.cell(row=r, column=i, value=v)
            if i == 4:
                c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
            else: c.font = F_BODY
        r += 1


def tab_cost_codes(wb, data, meta):
    ws = wb.create_sheet("06 Cost Codes")
    set_col_widths(ws, [10, 36, 14, 14, 14, 14, 14, 14])
    ws["A1"] = "Cost Code Detail (Revised Basis)"; ws["A1"].font = F_H1
    hdr = ["Code", "Description", "Original", "Revised", "Actual", "Variance", "Net Due", "% Complete"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    r = 4
    codes = data.get('code_totals', {})
    for code in sorted(c for c in codes.keys() if c is not None):
        if code == "999": continue
        info = codes[code]
        ws.cell(row=r, column=1, value=code).font = F_BODY
        ws.cell(row=r, column=2, value=info.get('desc', '')).font = F_BODY
        for col, key in enumerate(['orig', 'rev', 'actual'], start=3):
            c = ws.cell(row=r, column=col, value=info.get(key, 0))
            c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); c.number_format = CURR_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=7, value=info.get('net_due', 0)); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=8, value=f"=IFERROR(E{r}/D{r},0)"); c.number_format = PCT_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = F_BODY_B
    for col in (3, 4, 5, 6, 7):
        letter = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({letter}4:{letter}{r-1})")
        c.number_format = CURR_FMT; c.font = F_NUM_B; c.alignment = A_RIGHT
    c = ws.cell(row=r, column=8, value=f"=IFERROR(E{r}/D{r},0)")
    c.number_format = PCT_FMT; c.font = F_NUM_B; c.alignment = A_RIGHT


def tab_cost_categories(wb, data, meta):
    ws = wb.create_sheet("07 Cost Categories")
    set_col_widths(ws, [22, 10, 16, 16, 16, 16])
    ws["A1"] = "Cost Category Rollup"; ws["A1"].font = F_H1
    hdr = ["Category", "# Codes", "Revised Budget", "Actual", "Variance", "% Complete"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    cats = {"Labor": LABOR_CODES, "Material": MATERIAL_CODES, "Overhead": OVERHEAD_CODES, "Burden": BURDEN_CODES}
    codes = data.get('code_totals', {})
    r = 4
    for cat, code_set in cats.items():
        matching = [c for c in codes if c in code_set]
        budget = sum(codes[c].get('rev', 0) for c in matching)
        actual = sum(codes[c].get('actual', 0) for c in matching)
        ws.cell(row=r, column=1, value=cat).font = F_BODY_B
        ws.cell(row=r, column=2, value=len(matching)).font = F_BODY
        c = ws.cell(row=r, column=3, value=budget); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=4, value=actual); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); c.number_format = CURR_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},0)"); c.number_format = PCT_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = F_BODY_B
    for col in (2, 3, 4, 5):
        letter = get_column_letter(col)
        c = ws.cell(row=r, column=col, value=f"=SUM({letter}4:{letter}{r-1})")
        c.number_format = CURR_FMT if col > 2 else "#,##0"; c.font = F_NUM_B; c.alignment = A_RIGHT
    c = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},0)")
    c.number_format = PCT_FMT; c.font = F_NUM_B; c.alignment = A_RIGHT


def tab_bva(wb, data, meta):
    ws = wb.create_sheet("08 BVA")
    set_col_widths(ws, [10, 36, 14, 14, 14, 10, 14])
    ws["A1"] = "Budget vs Actual — Flag list"; ws["A1"].font = F_H1
    ws["A2"] = "Revised budget is canonical basis. Flags: OVER >10%, CRITICAL >50%, ON ±10%, UNDER <-10%."
    ws["A2"].font = F_NOTE
    hdr = ["Code", "Description", "Revised", "Actual", "Variance", "% Var", "Status"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=4, column=i, value=h)
    style_header(ws, 4, len(hdr))
    codes = data.get('code_totals', {})
    r = 5
    for code in sorted(c for c in codes.keys() if c is not None):
        if code == "999": continue
        info = codes[code]
        rev = info.get('rev', 0); act = info.get('actual', 0)
        if rev == 0 and act == 0: continue
        ws.cell(row=r, column=1, value=code).font = F_BODY
        ws.cell(row=r, column=2, value=info.get('desc', '')).font = F_BODY
        c = ws.cell(row=r, column=3, value=rev); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=4, value=act); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); c.number_format = CURR_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=6, value=f"=IFERROR((D{r}-C{r})/C{r},0)"); c.number_format = PCT_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=7, value=f'=IF(C{r}=0,"UNBUDGETED",IF(F{r}>0.5,"CRITICAL",IF(F{r}>0.1,"OVER",IF(F{r}<-0.1,"UNDER","ON"))))')
        c.font = F_NUM_B; c.alignment = A_CENTER
        r += 1


def tab_vendor_analysis(wb, data, meta):
    ws = wb.create_sheet("09 Vendor Analysis")
    set_col_widths(ws, [32, 14, 12, 14])
    ws["A1"] = "Vendor Spend (AP source)"; ws["A1"].font = F_H1
    hdr = ["Vendor", "Spend ($)", "Invoices", "% of Total AP"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    by_vendor = defaultdict(lambda: {"amt": 0.0, "n": 0})
    for t in data.get('transactions', []):
        if t.get('src') == 'AP':
            v = (t.get('party_name') or 'UNKNOWN').strip()
            by_vendor[v]["amt"] += t.get('amount', 0) or 0
            by_vendor[v]["n"] += 1
    vendors = sorted(by_vendor.items(), key=lambda kv: -kv[1]["amt"])
    r = 4
    for name, info in vendors:
        ws.cell(row=r, column=1, value=name).font = F_BODY
        c = ws.cell(row=r, column=2, value=info["amt"]); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        ws.cell(row=r, column=3, value=info["n"]).font = F_BODY
        c = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/SUM(B$4:B${4+len(vendors)-1}),0)"); c.number_format = PCT_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        r += 1
    ws.cell(row=r, column=1, value="TOTAL AP").font = F_BODY_B
    c = ws.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})"); c.number_format = CURR_FMT; c.font = F_NUM_B; c.alignment = A_RIGHT
    c = ws.cell(row=r, column=3, value=f"=SUM(C4:C{r-1})"); c.font = F_NUM_B; c.alignment = A_RIGHT


def tab_crew_roster(wb, data, meta):
    ws = wb.create_sheet("10 Crew Roster")
    set_col_widths(ws, [28, 12, 14, 12, 12])
    ws["A1"] = "Crew Roster (PR source)"; ws["A1"].font = F_H1
    hdr = ["Worker", "Hours", "Labor $", "Avg $/hr", "Tier"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    workers = data.get('worker_wages', [])
    if workers:
        workers_sorted = sorted(workers, key=lambda w: -(
            w.get('regular_hours', 0) + w.get('overtime_hours', 0) +
            w.get('doubletime_hours', 0) + w.get('other_hours', 0)
        ))
    else:
        workers_sorted = []
    r = 4
    for w in workers_sorted:
        h = (w.get('regular_hours', 0) or 0) + (w.get('overtime_hours', 0) or 0) + \
            (w.get('doubletime_hours', 0) or 0) + (w.get('other_hours', 0) or 0)
        amt = (w.get('regular_amount', 0) or 0) + (w.get('overtime_amount', 0) or 0)
        tier = w.get('tier', 'JOURNEYMAN')
        ws.cell(row=r, column=1, value=w['name']).font = F_BODY
        c = ws.cell(row=r, column=2, value=round(h, 1)); c.font = F_NUM_BLUE; c.alignment = A_RIGHT; c.number_format = "#,##0.0"
        c = ws.cell(row=r, column=3, value=round(amt, 0)); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)"); c.number_format = "$#,##0.00"; c.font = F_NUM; c.alignment = A_RIGHT
        ws.cell(row=r, column=5, value=tier).font = F_BODY
        r += 1


def tab_wage_tiers(wb, data, meta):
    ws = wb.create_sheet("11 Wage Tiers")
    set_col_widths(ws, [18, 10, 12, 14, 14])
    ws["A1"] = "Wage Tier Distribution"; ws["A1"].font = F_H1
    hdr = ["Tier", "# Workers", "Hours", "Labor $", "% of Hours"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    tiers = {"LEAD/SUPERVISOR": {"n": 0, "h": 0, "amt": 0},
             "JOURNEYMAN": {"n": 0, "h": 0, "amt": 0},
             "APPRENTICE/HELPER": {"n": 0, "h": 0, "amt": 0},
             "OT-ONLY": {"n": 0, "h": 0, "amt": 0}}
    for w in data.get('worker_wages', []):
        tier = w.get('tier', 'JOURNEYMAN')
        h = (w.get('regular_hours', 0) or 0) + (w.get('overtime_hours', 0) or 0) + \
            (w.get('doubletime_hours', 0) or 0) + (w.get('other_hours', 0) or 0)
        amt = (w.get('regular_amount', 0) or 0) + (w.get('overtime_amount', 0) or 0)
        if tier not in tiers: tier = 'JOURNEYMAN'
        tiers[tier]["n"] += 1; tiers[tier]["h"] += h; tiers[tier]["amt"] += amt
    r = 4; end_r = r + len(tiers) - 1
    for tier, info in tiers.items():
        ws.cell(row=r, column=1, value=tier).font = F_BODY_B
        ws.cell(row=r, column=2, value=info["n"]).font = F_BODY
        c = ws.cell(row=r, column=3, value=round(info["h"], 1)); c.font = F_NUM_BLUE; c.alignment = A_RIGHT; c.number_format = "#,##0.0"
        c = ws.cell(row=r, column=4, value=round(info["amt"], 0)); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/SUM(C$4:C${end_r}),0)"); c.number_format = PCT_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        r += 1


def tab_productivity(wb, data, meta):
    ws = wb.create_sheet("12 Productivity")
    set_col_widths(ws, [34, 18])
    ws["A1"] = "Productivity Metrics"; ws["A1"].font = F_H1
    derived = data.get('derived_fields', {})
    total_h = derived.get('total_labor_hours', 0)
    workers_count = derived.get('total_workers', 0)
    units = meta['units'] or 1
    billed = meta['billed'] or 0
    rows = [
        ("Total labor hours (to date)", round(total_h, 0)),
        ("Active workers", workers_count),
        ("Hours per unit (to date)", round(total_h / units, 1)),
        ("Revenue per labor hour (billed)", round(billed / total_h, 2) if total_h else 0),
        ("Percent complete", meta['pct_complete']),
    ]
    r = 3
    for lbl, val in rows:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY
        c = ws.cell(row=r, column=2, value=val); c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        if "per unit" in lbl or "hours" in lbl: c.number_format = "#,##0.0"
        elif "Revenue" in lbl: c.number_format = "$#,##0.00"
        elif "Percent" in lbl: c.number_format = PCT_FMT
        else: c.number_format = "#,##0"
        r += 1


def tab_benchmarks(wb, data, meta):
    ws = wb.create_sheet("13 Benchmarks")
    set_col_widths(ws, [34, 14, 14, 14])
    ws["A1"] = "Portfolio Benchmarks (vs. closed jobs)"; ws["A1"].font = F_H1
    hdr = ["Metric", "This Job", "Portfolio Median", "Delta"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    bench_rows = [
        ("Forecast margin %", meta['fcast_margin'], 0.34, PCT_FMT),
        ("Roughin labor share (code 120)", meta['rl_share'], 0.45, PCT_FMT),
        ("Top vendor concentration", meta['top_vendor_pct'], 0.38, PCT_FMT),
        ("Hours per unit", meta['hours_per_unit'], 95.0, "#,##0.0"),
    ]
    r = 4
    for lbl, this_v, med, fmt in bench_rows:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY
        c = ws.cell(row=r, column=2, value=this_v); c.number_format = fmt; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=3, value=med); c.number_format = fmt; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); c.number_format = fmt; c.font = F_NUM; c.alignment = A_RIGHT
        r += 1


def tab_predictive(wb, data, meta):
    ws = wb.create_sheet("14 Predictive Signals")
    set_col_widths(ws, [30, 16, 44])
    ws["A1"] = "Predictive Signals (live project)"; ws["A1"].font = F_H1
    hdr = ["Signal", "Status", "Notes"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    r = 4
    for sig in meta.get('predictive_signals', []):
        for i, v in enumerate(sig, start=1):
            c = ws.cell(row=r, column=i, value=v); c.font = F_BODY
            if i == 2:
                c.alignment = A_CENTER; c.font = F_BODY_B
                if "WATCH" in v.upper() or "⚠" in v: c.fill = FILL_CLAY
                elif "OK" in v.upper() or "✓" in v: c.fill = FILL_SAGE
        r += 1


def tab_reconciliation(wb, data, meta):
    ws = wb.create_sheet("15 Reconciliation")
    set_col_widths(ws, [38, 18, 18, 18])
    ws["A1"] = "JDR Reconciliation"; ws["A1"].font = F_H1
    hdr = ["Source", "Reported", "Calculated", "Delta"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    codes = data.get('code_totals', {})
    by_source = data.get('by_source', {})
    job_totals = data.get('job_totals', {})
    calc_expense = sum(codes[c].get('actual', 0) for c in codes if c != "999")
    rows = [
        ("Total expenses (from code totals)", job_totals.get('expenses', 0), calc_expense),
        ("Revenue billed (AR ·-1)", job_totals.get('revenues', 0), -by_source.get('AR', 0)),
        ("AP source total", None, by_source.get('AP', 0)),
        ("PR source total", None, by_source.get('PR', 0)),
        ("GL source total", None, by_source.get('GL', 0)),
    ]
    r = 4
    for lbl, rep, calc in rows:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY
        if rep is not None:
            c = ws.cell(row=r, column=2, value=rep); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=3, value=calc); c.number_format = CURR_FMT; c.font = F_NUM_BLUE; c.alignment = A_RIGHT
        c = ws.cell(row=r, column=4, value=f'=IF(ISNUMBER(B{r}),B{r}-C{r},"—")'); c.number_format = CURR_FMT; c.font = F_NUM; c.alignment = A_RIGHT
        r += 1


def tab_metric_registry(wb, data, meta):
    ws = wb.create_sheet("16 Metric Registry")
    set_col_widths(ws, [28, 20, 50])
    ws["A1"] = "Metric Registry"; ws["A1"].font = F_H1
    hdr = ["Metric", "Source Tab", "Definition"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    metrics = [
        ("Forecast margin %", "01 Overview", "Forecast net profit / Revised contract value"),
        ("Percent complete", "01 Overview", "Billed to date / Revised contract value"),
        ("Roughin labor share", "06 Cost Codes", "Code 120 revised / Total labor revised budget"),
        ("Top vendor concentration", "09 Vendor Analysis", "Largest single vendor AP spend / Total AP"),
        ("Hours per unit", "12 Productivity", "Total PR hours to date / Unit count"),
        ("BVA status codes", "08 BVA", "OVER / CRITICAL / UNDER / ON / UNBUDGETED on revised basis"),
    ]
    r = 4
    for row in metrics:
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v); c.font = F_BODY; c.alignment = A_LEFT
        r += 1


def tab_change_log(wb, data, meta):
    ws = wb.create_sheet("17 Change Log")
    set_col_widths(ws, [14, 24, 50])
    ws["A1"] = "Workbook Change Log"; ws["A1"].font = F_H1
    hdr = ["Date", "Version", "Note"]
    for i, h in enumerate(hdr, start=1): ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(hdr))
    ws.cell(row=4, column=1, value=meta['build_date']).font = F_BODY
    ws.cell(row=4, column=2, value="v1.0 · live").font = F_BODY
    ws.cell(row=4, column=3, value=f"Initial build for #{meta['job_id']} live project. 17-tab JCR schema on revised budget basis.").font = F_BODY


# ---------- Orchestration ----------
def build_workbook(meta):
    data = load_data()
    codes = data.get('code_totals', {})
    rev_expense = sum(codes[c].get('rev', 0) for c in codes if c != "999")
    rev_contract = abs(codes.get('999', {}).get('rev', 0))
    orig_contract = abs(codes.get('999', {}).get('orig', 0))

    meta['rev_contract'] = rev_contract
    meta['orig_contract'] = orig_contract
    meta['rev_expense'] = rev_expense
    meta['fcast_margin'] = (rev_contract - rev_expense) / rev_contract if rev_contract else 0

    rl = codes.get('120', {}).get('rev', 0)
    total_labor = sum(codes[c].get('rev', 0) for c in codes if c in LABOR_CODES)
    meta['rl_share'] = rl / total_labor if total_labor else 0

    by_vendor = defaultdict(float)
    for t in data.get('transactions', []):
        if t.get('src') == 'AP':
            by_vendor[(t.get('party_name') or '').strip()] += t.get('amount', 0) or 0
    ap_total = sum(by_vendor.values()) or 1
    top_v = max(by_vendor.values()) if by_vendor else 0
    meta['top_vendor_pct'] = top_v / ap_total if ap_total else 0

    derived = data.get('derived_fields', {})
    total_h = derived.get('total_labor_hours', 0)
    meta['hours_per_unit'] = total_h / (meta['units'] or 1)
    meta['pct_complete'] = meta['billed'] / rev_contract if rev_contract else 0

    data['worker_wages'] = data.get('worker_wages', [])

    wb = Workbook()
    wb.remove(wb.active)
    tab_overview(wb, data, meta)
    tab_job_info(wb, data, meta)
    tab_contract(wb, data, meta)
    tab_sov(wb, data, meta)
    tab_change_orders(wb, data, meta)
    tab_cost_codes(wb, data, meta)
    tab_cost_categories(wb, data, meta)
    tab_bva(wb, data, meta)
    tab_vendor_analysis(wb, data, meta)
    tab_crew_roster(wb, data, meta)
    tab_wage_tiers(wb, data, meta)
    tab_productivity(wb, data, meta)
    tab_benchmarks(wb, data, meta)
    tab_predictive(wb, data, meta)
    tab_reconciliation(wb, data, meta)
    tab_metric_registry(wb, data, meta)
    tab_change_log(wb, data, meta)

    wb.save(str(OUT_FILE))
    return str(OUT_FILE)


# ---------- Project-specific META ----------
META_2111 = {
    'job_id': '2111',
    'name': 'Northgate Station M3',
    'short_name': 'Northgate M3',
    'gc': 'Compass General Construction I, LLC',
    'owner': 'Northgate Station Residential M2 M3 LLC',
    'units': 186,
    'total_fixtures': 0,
    'insurance': 'TBD',
    'job_start': 'Apr 2025',
    'expected_completion': 'Q4 2026',
    'architect': 'CPL / GGLO',
    'mep': 'Franklin Engineering',
    'location': 'Northgate, Seattle, WA',
    'project_type': 'Mixed-use residential',
    'subcontract_number': 'TBD',
    'gc_project_number': 'TBD',
    'report_date': 'Mar 5, 2026',
    'billed': 2203668.0,
    'retention': 110183.4,
    'build_date': '2026-04-24',
    'contract_original': 4420000,
    'contract_final': 4509892,
    'executed_co_count': 7,
    'cor_count': 14,
    'rfi_count': 0,
    'asi_count': 0,
    'submittal_count': 0,
    'total_pos': 0,
    'permit_count': 0,
    'pay_app_count': 7,
    'ri_foreman': 'TBD',
    'trim_foreman': 'TBD',
    'gc_pm': 'TBD',
    'gc_pe': 'TBD',
    'base_bid': 4420000,
    'bid_date': 'TBD',
    'previous_plumber': 'N/A',
    'invoice_count': 0,
    'vendor_quote_count': 0,
    'cos': [
        {'co': 'CO#01', 'desc': 'Unit riser shutoff valve additions (COR#03)', 'amount': 26130},
        {'co': 'CO#02', 'desc': 'Bundled minor field adds (varies)', 'amount': 15700},
        {'co': 'CO#03', 'desc': 'ASI revisions bundle', 'amount': 12400},
        {'co': 'CO#04', 'desc': 'Domestic water relocation', 'amount': 8900},
        {'co': 'CO#05', 'desc': 'Fixture spec uplift', 'amount': 7100},
        {'co': 'CO#06', 'desc': 'Drain reroute (ASI 22)', 'amount': 11500},
        {'co': 'CO#07', 'desc': 'Misc additions + cleanups', 'amount': 8162},
    ],
    '_old_cos': [
        {'co': 1, 'desc': 'RFI #94 - Type A/B Toilet Product Change', 'amount': 1676},
        {'co': 2, 'desc': 'Jobsite Office Bottle Filler', 'amount': 2167},
        {'co': 3, 'desc': 'RFI #108 - Delete L2 Lounge Sink & Dishwasher', 'amount': -3200},
        {'co': 4, 'desc': 'ASI#4 - Plumbing Permit Revisions', 'amount': 35521},
        {'co': 5, 'desc': 'Foundation Drain Filter Fabric', 'amount': 2675},
        {'co': 6, 'desc': 'RFI 55 - Water Meter Clarification', 'amount': 10071},
        {'co': 7, 'desc': 'RFI #89 - Retail MEP Updates', 'amount': 5878},
        {'co': 8, 'desc': 'RFI 49 - Bike Room 1-A2 Shower', 'amount': 2329},
        {'co': 9, 'desc': 'RFI 124 - Explosion Proof Sump Pumps', 'amount': 11295},
        {'co': 10, 'desc': 'L8 Amenity Gas Fire Pit/Fire Place Supply', 'amount': 7664},
        {'co': 11, 'desc': 'ASI 10 - Plumbing Vent & Waste Routing Revisions', 'amount': 18764},
        {'co': 12, 'desc': 'ASI 12 - Storm and Plumbing Vent Revisions', 'amount': 8074},
        {'co': 13, 'desc': 'ASI 14 - King County Plumbing Revisions & Permit Cost', 'amount': 3506},
        {'co': 14, 'desc': 'Temp Gas Risers', 'amount': 34604},
        {'co': 15, 'desc': 'RFI #163 - Bioretention 6A Trench Drain', 'amount': 4535},
        {'co': 16, 'desc': 'RFI #334 - Add Icemakers to All Unit Types', 'amount': 45869},
        {'co': 17, 'desc': 'ASI 9 - Relocation of HWHP', 'amount': 18692},
        {'co': 18, 'desc': 'OCIP Credit', 'amount': -20453.66},
        {'co': 19, 'desc': 'RFI #238 - Added L4 Roof Drains', 'amount': 2715},
        {'co': 20, 'desc': 'RFI #266 - Add OHP-COMM Condensate Drain Line', 'amount': 1175},
        {'co': 21, 'desc': 'RFI #237 - Plumbing Adjustment to Fur Wall', 'amount': 4322},
        {'co': 22, 'desc': 'Hot and Cold Unit Water Meters - Furnish & Install', 'amount': 174800},
        {'co': 23, 'desc': 'RFI #278.1 - L2 Landscape North Alcove Ramp Drain', 'amount': 4109},
        {'co': 24, 'desc': 'Copper Condensates in Retail Area', 'amount': 44230},
        {'co': 25, 'desc': 'ASI-07 Extend Garage Trench Drain + Civil Storm POC', 'amount': 6895},
        {'co': 26, 'desc': 'RFI-446 Unit Riser Shutoff Valves', 'amount': 29700},
        {'co': 27, 'desc': 'RFI-363 Added Hosebib', 'amount': 1100},
        {'co': 28, 'desc': 'RFI #517 - RPBP-1&2 Pipe Size', 'amount': 5110},
        {'co': 29, 'desc': 'Flex Connector Changes', 'amount': 4860},
        {'co': 30, 'desc': 'Remove & Reinstall Tubs & Showers for Framing', 'amount': 3090},
        {'co': 31, 'desc': 'Temp Roof Drains', 'amount': 1330},
        {'co': 32, 'desc': 'L9 Temp Roof Drains', 'amount': 1240},
        {'co': 33, 'desc': 'Leak Sensor Install', 'amount': 3950},
        {'co': 34, 'desc': 'ASI-19 Retail Revisions - Add POCs, Move Rainleader', 'amount': 5450},
        {'co': 35, 'desc': 'Additional Control Panel for Water Heating System', 'amount': 4320},
        {'co': 36, 'desc': 'Casework Revisions @ X25-X41', 'amount': 10170},
        {'co': 37, 'desc': 'L7 Temp Roof Drains', 'amount': 1470},
        {'co': 38, 'desc': 'RFI #432.1 - Plumbing Rework', 'amount': 9600},
        {'co': 39, 'desc': 'Water Stub Out Extensions at Cabinets', 'amount': 6160},
        {'co': 40, 'desc': 'ASI-028 Dog & Bike Wash', 'amount': 7520},
        {'co': 41, 'desc': 'Fix Temp Heat Gas Hoses', 'amount': 5230},
        {'co': 42, 'desc': 'Add GI Stub', 'amount': 4520},
        {'co': 43, 'desc': 'RFI 673 - Add Drain Fire Pump Room', 'amount': 950},
        {'co': 44, 'desc': 'Move Plumbing Per RFI 667', 'amount': 790},
        {'co': 45, 'desc': 'Cap Temp Gas / Move Waste Lines / Move ADA Tub Spouts', 'amount': 5930},
        {'co': 46, 'desc': 'RFI 519 - Reroute Condensates Units 128-129', 'amount': 2630},
        {'co': 47, 'desc': 'Add Level 2 Floor Drain', 'amount': 1630},
        {'co': 48, 'desc': 'Remove Tubs Per RFI 639', 'amount': 1660},
        {'co': 49, 'desc': 'Disconnect Temp Gas', 'amount': 880},
        {'co': 50, 'desc': 'ASI 30 - Added Sink', 'amount': 3860},
        {'co': 51, 'desc': 'ASI 33 - Fitness Room', 'amount': 34610},
        {'co': 52, 'desc': 'Maintenance Room Sink Replacement', 'amount': 410},
        {'co': 53, 'desc': 'Permit Revision Fees', 'amount': 2130},
        {'co': 54, 'desc': 'RFI 694 - Add Grease Stub', 'amount': 6170},
    ],
    'pay_apps': [],
    'change_log': [],
    'predictive_signals': [],
}


def generate_predictive_signals(data):
    codes = data.get('code_totals', {})
    signals = []
    for code in sorted(c for c in codes.keys() if c is not None):
        if code == '999': continue
        info = codes[code]
        rev = info.get('rev', 0)
        actual = info.get('actual', 0)
        desc = info.get('desc', '')
        if rev == 0:
            if actual > 10000:
                signals.append((
                    f'{desc} ({code}) unbudgeted',
                    '⚠ WATCH',
                    f'No revised budget but ${actual:,.0f} actual spend — needs CO or budget transfer'
                ))
            continue
        pct = (actual - rev) / rev if rev else 0
        if pct > 0.5:
            signals.append((
                f'{desc} ({code}) CRITICAL overrun',
                '⚠ WATCH',
                f'Actual ${actual:,.0f} vs revised ${rev:,.0f} (+{pct:.0%}) — needs immediate attention'
            ))
        elif pct > 0.1:
            signals.append((
                f'{desc} ({code}) over revised',
                '⚠ WATCH',
                f'Actual ${actual:,.0f} vs revised ${rev:,.0f} (+{pct:.0%}) — monitor'
            ))

    by_vendor = defaultdict(float)
    for t in data.get('transactions', []):
        if t.get('src') == 'AP':
            by_vendor[(t.get('party_name') or '').strip()] += t.get('amount', 0) or 0
    ap_total = sum(by_vendor.values()) or 1
    if by_vendor:
        top_name, top_val = max(by_vendor.items(), key=lambda x: x[1])
        top_pct = top_val / ap_total
        status = '⚠ WATCH' if top_pct > 0.40 else '✓ OK'
        signals.append((
            f'Top vendor concentration ({top_name[:20]})',
            status,
            f'{top_pct:.0%} of AP (${top_val:,.0f}) — {"high, monitor" if top_pct > 0.40 else "acceptable"}'
        ))

    ret = abs(META_2111.get('retention', 0))
    billed = abs(META_2111.get('billed', 0))
    ret_pct = ret / billed if billed else 0
    signals.append((
        'Retention aging',
        '✓ OK' if ret_pct < 0.06 else '⚠ WATCH',
        f'${ret:,.0f} held · {ret_pct:.1%} of billed · standard range'
    ))

    rev_contract = abs(codes.get('999', {}).get('rev', 0))
    rev_expense = sum(codes[c].get('rev', 0) for c in codes if c != '999')
    margin = (rev_contract - rev_expense) / rev_contract if rev_contract else 0
    signals.append((
        'Forecast margin',
        '✓ OK' if margin > 0.25 else '⚠ WATCH',
        f'{margin:.1%} forecast margin — {"healthy" if margin > 0.30 else "monitor"}'
    ))

    return signals[:10]


if __name__ == "__main__":
    data = load_data()
    META_2111['predictive_signals'] = generate_predictive_signals(data)
    out = build_workbook(dict(META_2111))
    print(f"Built: {out}")
    print(f"Size: {Path(out).stat().st_size:,} bytes")
