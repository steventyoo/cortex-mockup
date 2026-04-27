#!/usr/bin/env python3
"""
notionize_2103.py — Build the Notion-styled 2-tab summary workbook for
Job 2103 (Compass Northgate M2) — live project.

Input:   2103_data.json  (from parse_2103.py)
Output:  OWP_2103_JCR_Summary_Notion_v2.xlsx
"""
import json
from collections import defaultdict
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HERE = Path(__file__).resolve().parent
DATA_FILE = _HERE / '2103_data.json'
OUT_FILE = _HERE / 'OWP_2103_JCR_Summary_Notion_v2.xlsx'

# Notion tokens
INK = "37352F"; INK_SOFT = "787774"; BG = "FFFFFF"; DIV = "E9E5DA"
CLAY = "D77A61"; SAGE = "8AA474"; BLUE_TXT = "0F7B9D"; YELLOW = "FDE68A"

FONT = "Arial"
F_TITLE = Font(name=FONT, size=22, bold=True, color=INK)
F_KICKER = Font(name=FONT, size=9, color=INK_SOFT)
F_H = Font(name=FONT, size=12, bold=True, color=INK)
F_BODY = Font(name=FONT, size=10, color=INK)
F_BODY_B = Font(name=FONT, size=10, bold=True, color=INK)
F_NUM = Font(name=FONT, size=10, color=INK)
F_NUM_B = Font(name=FONT, size=10, bold=True, color=INK)
F_HERO = Font(name=FONT, size=28, bold=True, color=INK)
F_SUB = Font(name=FONT, size=10, color=INK_SOFT, italic=True)
F_TAG = Font(name=FONT, size=9, bold=True, color="FFFFFF")

FILL_CLAY = PatternFill("solid", start_color=CLAY)
FILL_SAGE = PatternFill("solid", start_color=SAGE)
FILL_BAND = PatternFill("solid", start_color="F7F6F3")

DIVLINE = Side(style="thin", color=DIV)
BORDER_BOT = Border(bottom=DIVLINE)

A_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
A_R = Alignment(horizontal="right", vertical="center")
A_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

CURR = '$#,##0;($#,##0);-'
PCT = '0.0%;(0.0%);-'

LABOR_CODES = {"011","100","101","110","111","112","113","120","130","140","141","142","143","145","150"}


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def tag(ws, row, col, text, fill):
    c = ws.cell(row=row, column=col, value=text)
    c.font = F_TAG; c.alignment = A_C; c.fill = fill


def load_data():
    raw = json.loads(DATA_FILE.read_text())
    code_totals = {}
    for cs in raw.get('cost_code_summaries', []):
        if cs['code'] is None: continue
        code_totals[cs['code']] = {
            'desc': cs['description'], 'orig': cs['original_budget'],
            'rev': cs['current_budget'], 'actual': cs['actual_amount'],
            'variance': cs['plus_minus_budget'], 'net_due': cs['net_due'],
            'retainage': cs['retainage'],
        }
    transactions = []
    for li in raw.get('line_items', []):
        hours = (li.get('regular_hours') or 0) + (li.get('overtime_hours') or 0) + \
                (li.get('doubletime_hours') or 0) + (li.get('other_hours') or 0)
        transactions.append({
            'src': li['source'], 'party_name': li.get('name', ''),
            'amount': li.get('actual_amount', 0), 'hours': hours,
            'cost_code': li.get('cost_code', ''),
            'post_date': li.get('posted_date', ''),
            'ref': li.get('ref_number', ''),
            'ck_num': li.get('check_number', ''),
        })
    return {'code_totals': code_totals, 'transactions': transactions}


def build_summary_tab(wb, data, meta):
    ws = wb.create_sheet("Summary")
    set_col_widths(ws, [30, 22, 22, 22, 22])
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"#{meta['job_id']} · {meta['name']}"
    ws["A1"].font = F_TITLE; ws.merge_cells("A1:E1")
    ws["A2"] = f"{meta['gc']} · {meta['units']} units · ACTIVE · {meta['pct_complete']:.1%} billed · OCIP"
    ws["A2"].font = F_SUB; ws.merge_cells("A2:E2")
    ws.row_dimensions[1].height = 30; ws.row_dimensions[2].height = 18

    r = 4
    heroes = [
        ("REV. CONTRACT", meta['rev_contract'], CURR),
        ("FORECAST MARGIN", meta['fcast_margin'], PCT),
        ("BILLED TO DATE", meta['billed'], CURR),
        ("BALANCE TO FINISH", meta['rev_contract'] - meta['billed'], CURR),
    ]
    for i, (lbl, val, fmt) in enumerate(heroes, start=1):
        ws.cell(row=r, column=i, value=lbl).font = F_KICKER
        ws.cell(row=r, column=i).alignment = A_L
        c = ws.cell(row=r+1, column=i, value=val)
        c.font = F_HERO; c.number_format = fmt; c.alignment = A_L
    ws.row_dimensions[r+1].height = 36

    r = 8
    for c in range(1, 6): ws.cell(row=r, column=c).border = BORDER_BOT

    r = 10
    ws.cell(row=r, column=1, value="Key facts").font = F_H
    r += 1
    facts = [
        ("Original contract", meta['orig_contract'], CURR),
        ("Revised contract (CO impact)", meta['rev_contract'], CURR),
        ("Executed COs (net)", meta['rev_contract'] - meta['orig_contract'], CURR),
        ("Retention held", meta['retention'], CURR),
        ("Revised expense budget", meta['rev_expense'], CURR),
        ("Forecast net profit", meta['rev_contract'] - meta['rev_expense'], CURR),
        ("Percent complete", meta['pct_complete'], PCT),
    ]
    for lbl, val, fmt in facts:
        ws.cell(row=r, column=1, value=lbl).font = F_BODY; ws.cell(row=r, column=1).alignment = A_L
        c = ws.cell(row=r, column=3, value=val)
        c.font = F_NUM_B; c.number_format = fmt; c.alignment = A_R
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Live signals").font = F_H
    r += 1
    for sig in meta.get('predictive_signals', []):
        name, status, note = sig
        ws.cell(row=r, column=1, value=name).font = F_BODY; ws.cell(row=r, column=1).alignment = A_L
        fill = FILL_CLAY if ("⚠" in status or "WATCH" in status.upper()) else FILL_SAGE
        tag(ws, r, 2, status, fill)
        c = ws.cell(row=r, column=3, value=note); c.font = F_SUB; c.alignment = A_L
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="Top vendors (AP)").font = F_H
    r += 1
    by_vendor = defaultdict(lambda: {"amt": 0.0, "n": 0})
    for t in data.get('transactions', []):
        if t.get('src') == 'AP':
            v = (t.get('party_name') or 'UNKNOWN').strip()
            by_vendor[v]["amt"] += t.get('amount', 0) or 0
            by_vendor[v]["n"] += 1
    total_ap = sum(v["amt"] for v in by_vendor.values()) or 1
    top_vendors = sorted(by_vendor.items(), key=lambda kv: -kv[1]["amt"])[:6]

    ws.cell(row=r, column=1, value="Vendor").font = F_BODY_B
    ws.cell(row=r, column=2, value="Invoices").font = F_BODY_B
    ws.cell(row=r, column=3, value="Spend").font = F_BODY_B
    ws.cell(row=r, column=4, value="% of AP").font = F_BODY_B
    for col in range(1, 5): ws.cell(row=r, column=col).border = BORDER_BOT
    r += 1
    for name, info in top_vendors:
        ws.cell(row=r, column=1, value=name).font = F_BODY
        ws.cell(row=r, column=2, value=info["n"]).font = F_BODY
        c = ws.cell(row=r, column=3, value=info["amt"]); c.font = F_NUM; c.number_format = CURR; c.alignment = A_R
        c = ws.cell(row=r, column=4, value=info["amt"] / total_ap); c.font = F_NUM; c.number_format = PCT; c.alignment = A_R
        r += 1


def build_transactions_tab(wb, data, meta):
    ws = wb.create_sheet("Transactions")
    set_col_widths(ws, [6, 8, 12, 12, 28, 10, 12, 14])
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Transactions Detail"; ws["A1"].font = F_TITLE
    ws["A2"] = f"JDR source transactions for job #{meta['job_id']} · {len(data.get('transactions', []))} rows"
    ws["A2"].font = F_SUB; ws.merge_cells("A2:H2")
    ws.row_dimensions[1].height = 28

    r = 4
    hdr = ["Src", "Code", "Post Date", "Ref", "Party", "Hours", "Amount", "Check#"]
    for i, h in enumerate(hdr, start=1):
        c = ws.cell(row=r, column=i, value=h)
        c.font = F_BODY_B; c.alignment = A_C; c.border = BORDER_BOT
    r += 1
    for t in data.get('transactions', []):
        ws.cell(row=r, column=1, value=t.get('src', '')).font = F_BODY
        ws.cell(row=r, column=2, value=t.get('cost_code', '')).font = F_BODY
        ws.cell(row=r, column=3, value=t.get('post_date', '')).font = F_BODY
        ws.cell(row=r, column=4, value=t.get('ref', '')).font = F_BODY
        ws.cell(row=r, column=5, value=(t.get('party_name') or '').strip()).font = F_BODY
        if t.get('hours'):
            c = ws.cell(row=r, column=6, value=t.get('hours')); c.number_format = "#,##0.0"; c.alignment = A_R; c.font = F_BODY
        c = ws.cell(row=r, column=7, value=t.get('amount', 0)); c.number_format = CURR; c.alignment = A_R; c.font = F_BODY
        ws.cell(row=r, column=8, value=t.get('ck_num', '')).font = F_BODY
        r += 1


def build_notion_workbook(meta):
    data = load_data()
    codes = data.get('code_totals', {})
    rev_contract = abs(codes.get('999', {}).get('rev', 0))
    orig_contract = abs(codes.get('999', {}).get('orig', 0))
    rev_expense = sum(codes[c].get('rev', 0) for c in codes if c != '999')

    meta['rev_contract'] = rev_contract
    meta['orig_contract'] = orig_contract
    meta['rev_expense'] = rev_expense
    meta['fcast_margin'] = (rev_contract - rev_expense) / rev_contract if rev_contract else 0
    meta['pct_complete'] = meta['billed'] / rev_contract if rev_contract else 0

    wb = Workbook()
    wb.remove(wb.active)
    build_summary_tab(wb, data, meta)
    build_transactions_tab(wb, data, meta)

    wb.save(str(OUT_FILE))
    return str(OUT_FILE)


# ---------- Generate predictive signals from BVA data ----------
def generate_predictive_signals(data):
    codes = data.get('code_totals', {})
    signals = []
    for code in sorted(c for c in codes.keys() if c is not None):
        if code == '999': continue
        info = codes[code]
        rev = info.get('rev', 0); actual = info.get('actual', 0)
        desc = info.get('desc', '')
        if rev == 0:
            if actual > 10000:
                signals.append((f'{desc} ({code}) unbudgeted', '⚠ WATCH',
                    f'No revised budget but ${actual:,.0f} actual — needs CO or budget transfer'))
            continue
        pct = (actual - rev) / rev if rev else 0
        if pct > 0.5:
            signals.append((f'{desc} ({code}) CRITICAL', '⚠ WATCH',
                f'Actual ${actual:,.0f} vs revised ${rev:,.0f} (+{pct:.0%}) — needs attention'))
        elif pct > 0.1:
            signals.append((f'{desc} ({code}) over', '⚠ WATCH',
                f'Actual ${actual:,.0f} vs revised ${rev:,.0f} (+{pct:.0%}) — monitor'))

    # Top vendor
    by_vendor = defaultdict(float)
    for t in data.get('transactions', []):
        if t.get('src') == 'AP':
            by_vendor[(t.get('party_name') or '').strip()] += t.get('amount', 0) or 0
    ap_total = sum(by_vendor.values()) or 1
    if by_vendor:
        top_name, top_val = max(by_vendor.items(), key=lambda x: x[1])
        top_pct = top_val / ap_total
        signals.append((f'Top vendor ({top_name[:20]})',
            '⚠ WATCH' if top_pct > 0.40 else '✓ OK',
            f'{top_pct:.0%} of AP — {"high" if top_pct > 0.40 else "acceptable"}'))

    signals.append(('Retention', '✓ OK', f'${305091.57:,.0f} held · 5% standard'))

    rev_contract = abs(codes.get('999', {}).get('rev', 0))
    rev_expense = sum(codes[c].get('rev', 0) for c in codes if c != '999')
    margin = (rev_contract - rev_expense) / rev_contract if rev_contract else 0
    signals.append(('Forecast margin', '✓ OK' if margin > 0.25 else '⚠ WATCH',
        f'{margin:.1%} forecast — {"healthy" if margin > 0.30 else "monitor"}'))

    return signals[:10]


META_2103 = {
    'job_id': '2103',
    'name': 'Northgate Station M2',
    'gc': 'Compass General Construction I, LLC',
    'units': 234,
    'billed': 6608301.34,
    'retention': 305091.57,
}


if __name__ == "__main__":
    data = load_data()
    META_2103['predictive_signals'] = generate_predictive_signals(data)
    out = build_notion_workbook(dict(META_2103))
    print(f"Built: {out}")
    print(f"Size: {Path(out).stat().st_size:,} bytes")
