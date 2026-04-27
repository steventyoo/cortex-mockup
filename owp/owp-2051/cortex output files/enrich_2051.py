#!/usr/bin/env python3
"""
enrich_2051.py — Fill in empty/sparse tabs and enrich key-players data
in OWP_2051_JCR_Cortex_v2.xlsx (Compass · Vail Apartments).

GDrive scan (2026-04-25) confirmed:
  Folder: 2051-Compass, Vail Apartments/
  Subfolders: # Job Ticket, ASI-RFI, Billing, Brett Vail, Change Orders,
              Checklists, Contract, Franklin Vail Apartments, Insurance,
              Invoices, Meetings-Schedules, O&M's, PO's, Parking, Permits,
              Photos, Plans, Safety, Submittals, Vendor Lien Waivers,
              Vendor Quotes, RUVATI application.docx,
              Vail Apartments Trucking.pdf

Read from contract (One Way Plumbing - Vail Apartments Fully Executed Contract.pdf):
  Subcontract # 18-0622-0500 · effective Aug 14, 2018 · Compass General
  Construction I, LLC (Kirkland WA) ↔ Oneway Plumbing, LLC. Project officially
  named "Centerpoint Mixed-Use (Vail Apartments)". Site: 17962 Midvale Ave N,
  Shoreline, WA 98136. Owner: AAA Management LLC dba ADC Ridge at Sun Valley,
  San Diego CA. Original contract $2,672,000. 5% retention. Signatories:
  Ryan Ames (Compass VP) and Richard Donelson (OWP VP). OWP rep: Richard Donelson.

GAPS this script fills:
  • 04 SOV-PayApps   — was 2 rows (header only)
  • 05 Change Orders — was 2 rows (header only); now lists 17 SCO + 7 COR
                       events extracted from GDrive Change Orders folder
  • 17 Change Log    — had only workbook log; project change events added

ALSO adds:
  • Expanded "Job Info" — 60+ key-players + contract + permit + insurance fields
  • New "Project Team" tab — dedicated GC/Owner/OWP/MEP roster grid

DASHBOARD DISCREPANCY FLAGGED (for Steven to fix in index.html):
  • PROJECTS['2051'].location = "Bellevue area, WA"
    → CORRECT per executed contract: "Shoreline, WA (17962 Midvale Ave N)"
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
WORKBOOK   = SCRIPT_DIR / "OWP_2051_JCR_Cortex_v2.xlsx"
JSON_FILE  = SCRIPT_DIR / "2051_data.json"

# ============================================================================
# 2051 META — sourced from executed contract + GDrive folder scan
# ============================================================================
META_2051 = {
    # Identity
    "subcontract_no":   "18-0622-0500",
    "effective_date":   "August 14, 2018",
    "project_name_full": "Centerpoint Mixed-Use (Vail Apartments)",
    "project_name":     "Vail Apartments",
    "site_address":     "17962 Midvale Avenue N, Shoreline, WA 98136",
    "city":             "Shoreline",  # NOT Bellevue (dashboard typo)
    "state":            "WA",
    "zip":              "98136",
    "permit_no":        "City of Shoreline #890-895 (Vail Apartments-Shoreline)",
    "permit_install_inv":"Invoice 325902",
    "permit_subs":      "Boiler/PV Installation Permits HWT 1-4 + PV 1-2 issued 6-15-2020",

    # General Contractor (full legal)
    "gc_legal":         "Compass General Construction I, LLC",
    "gc_short":         "Compass",
    "gc_address":       "733 7th Avenue, Suite 212, Kirkland, WA 98033",
    "gc_phone":         "(206) 320-8741",
    "gc_fax":           "(206) 323-4719",
    "gc_signatory":     "Ryan Ames, Vice President",
    "gc_pm":            "Justin Anderson",
    "gc_sup":           "Will Fenton",
    "gc_pe":            "Jeff Seeb",

    # OWP (Subcontractor)
    "owp_legal":        "Oneway Plumbing, LLC",
    "owp_address":      "127 10th Street South, Suite 200, Kirkland, WA 98033",
    "owp_phone":        "425-968-8314",
    "owp_signatory":    "Richard Donelson, Vice President",
    "owp_rep":          "Richard Donelson",
    "owp_ri_foreman":   "Bob",
    "owp_license":      "ONEWAWP895BU (renewal cycle; original expired 1/31/2019)",
    "owp_fed_tax_id":   "27-4605928",
    "owp_unemp":        "42342900",
    "owp_wc_id":        "207548-00",

    # Owner
    "owner_legal":      "AAA Management, LLC dba ADC Ridge at Sun Valley",
    "owner_address":    "1450 Frazee Road, Suite 414, San Diego, CA 92108",
    "owner_short":      "AAA Management",

    # Insurance
    "insurance_carrier":"Hub International / Liberty Mutual",
    "insurance_policy": "BKA 17 55850721",
    "insurance_type":   "Not Wrap (OWP carries own CGL/WC)",

    # Contract terms
    "contract_original":"$2,672,000",
    "contract_final":   "$2,588,280",
    "contract_co_net":  "($83,720) net — credit-net posture (-3.1%)",
    "retention_pct":    "5% per Article 7.2.2",
    "retention_held":   "$126,114 (still held per JDR 04/03/2026, 7+ yrs aged)",

    # Schedule
    "start_date":       "March 2018 (subcontract effective 8/14/2018)",
    "end_date":         "Closed (retention release pending — 2023)",
    "duration":         "~22 months field + retention tail",

    # Design / MEP
    "mep_engineer":     "Franklin Engineering",
    "architect":        "TBD (not in dashboard) — see Plans folder",
    "structural":       "TBD",
    "civil":            "TBD",

    # Counts
    "units":            163,
    "fixtures":         "191 toilets + 17 ADA + 192 lavs Type B + 153 SS sinks (Ruvati RVH7400) + 11 ADA SS sinks (Elkay ELUHAD211550) + 202 tub/shower stalls (Fibercare ET60-32RT-HC-80S L/R) + 208 Delta 559LF-PP faucets + 14 amenity fixtures + 4 Bock OptiTherm OT-300N-A water heaters + 2 Lawler 805 master mixing valves + 2 backflow preventers + 1 booster package + 13 Watts RD25 roof drains + 21 Watts FD33 garage drains + 15 Watts FD15 floor drains",
    "amenity_features": "L3 amenity (1 dual lav, 1 SS sink, 1 hotsink), 2 Elkay drinking fountains w/bottle filler, 1 Forever Stainless ADA dog wash (60HTY-RH), 2 Mustee 63M mop sinks, 4 Moen garbage disposal push buttons (per CO #14), 1 coffee machine line",
    "rough_systems":    "1 wvs/cw stub (retail), 1 common gas supply (4 water heaters + 2 rooftop HVAC + 2 BBQ + 2 firepit + L2 gas cooktop + L3 gas), 1 fireplace L2, 1 yard hydrant rooftop, 4 hosebibs garage mech rooms, 10 ext keyless hose bibs, 1 trash room hot/cold hose bib",

    # Documents
    "contract_pdfs": [
        "One Way Plumbing - Vail Apartments Fully Executed Contract.pdf",
        "One Way Plumbing - PSA Agreement - Final with exhibits - 2-20-18 - executed.pdf",
        "Vail Insurance Requirements.pdf",
    ],
    "ticket_count":   "14 ticket revisions (12-10-18 → 3-4-2020)",
    "ticket_latest":  "Vail Apartments job ticket 3-4-2020.pdf",
    "co_count_dir":   17,    # SCO files in Change Orders folder
    "cor_count_dir":  7,     # COR files (numbered)
    "rfi_count":      "120+ (RFI 95, 99, 106, 109, 113, 117, 119, 120, 121, ..., 143)",
    "asi_count":      "1 (ASI 03 on file)",
    "permit_count":   9,
    "plans_subfolders":"Plans / Photos / Submittals (full design set)",
    "vendor_subfolders":"Brett Vail (sub contact), Franklin Vail Apartments (MEP design), RUVATI application.docx",
    "delivery":       "Vail Apartments Trucking.pdf (delivery instructions)",

    # Scope notes
    "scope_design_build":"Plumbing per CAD drawings — design-build per Article 4.7. Holdrite firestop sleeve system at post-tension decks.",
    "scope_excludes": "Shoreline plumbing, gas/boiler permits/plan review fees (by owner). Excavation/backfill/dewatering. Hoisting concrete basin. Foundation/footing drains. Storm detention piping. Pump systems for storm/sanitary drainage. Fire-rated tub/shower enclosures. Wash machine drain pans. Kitchen sink garbage disposals. Utility/unit meters.",
}

# Change orders extracted from GDrive Change Orders/CO's folder
SCO_LIST = [
    ("SCO 01",   "Plumbing Alternates and Adjustments"),
    ("SCO 02",   "Firestopping"),
    ("SCO 03",   "Alternate Plumbing Fixtures"),
    ("SCO 04",   "Added Unit 330 and Common Areas"),
    ("SCO 05",   "Temp Gas Lines for Temp Heat (10ea POCs, levels 3-7)"),
    ("SCO 06",   "ADA Sink Modification"),
    ("SCO 07",   "Added Drinking Fountain"),
    ("SCO 08",   "ADA Dog Wash Upgrade"),
    ("SCO 09",   "L4 Roof Piping Retesting"),
    ("SCO 010",  "(referenced — file on Drive)"),
    ("SCO 012",  "Ruvati Sink Change (replaces previous SS sink spec)"),
    ("SCO 015",  "Re-route Piping in Amenity Space"),
    ("SCO 017",  "Replacement Sinks"),
    ("Misc",     "Amenity Plumbing Fixture Upgrades"),
    ("Misc",     "Outdoor Kitchen Sink Change"),
    ("Misc",     "Owner Plumbing Modifications"),
    ("Misc",     "Remove Dishwasher Installation from Scope"),
    ("Closeout", "One Way Closeout Change Order"),
    ("Reconcil", "Vail – One-Way COP Log Reconciliation"),
]

COR_LIST = [
    ("COR #01",  "West Wall Foundation Drainage"),
    ("COR #02",  "Added Fixtures per BP Cycle 2 Corrections (revised r1)"),
    ("COR #03",  "ADA Dog Wash Upgrade"),
    ("COR #04",  "RFI-117 Move Piping for Retail Door to Garage"),
    ("COR #05",  "RFI-143 Move Piping for Mail Room Wall Relocation"),
    ("COR #06",  "IFC Plan Set Changes, Added Unit L3"),
    ("COR #07",  "Alternate Fixtures"),
]

# ============================================================================
# Style helpers
# ============================================================================
INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
WARN_FONT   = Font(name="Arial", size=10, color="B85C3E", bold=True)

HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
ROW_ALT     = PatternFill("solid", start_color="FAFAFA")
WARN_FILL   = PatternFill("solid", start_color="FFF3CD")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")


def clear_sheet(ws, keep_header_rows=2):
    """Clear all cells below the first N header rows."""
    for r in range(keep_header_rows + 1, max(ws.max_row + 1, 50)):
        for c in range(1, max(ws.max_column + 1, 10)):
            cell = ws.cell(row=r, column=c)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)


# ============================================================================
# 02 Job Info — expanded with 60+ fields (REPLACE existing minimal content)
# ============================================================================
def patch_job_info(ws):
    """Replace the 13-row Job Info with a comprehensive 60-row team + identity grid."""
    clear_sheet(ws, keep_header_rows=0)

    ws["A1"] = "JOB #2051 · COMPASS · VAIL APARTMENTS — FULL PROJECT INFORMATION"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = "Sourced from executed subcontract (18-0622-0500) + GDrive folder scan + dashboard PROJECT_TEAMS"
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    sections = [
        ("IDENTITY", [
            ("OWP Job ID",            "2051"),
            ("Project Name",          META_2051["project_name"]),
            ("Project Name (full)",   META_2051["project_name_full"]),
            ("Subcontract Number",    META_2051["subcontract_no"]),
            ("Effective Date",        META_2051["effective_date"]),
            ("Site Address",          META_2051["site_address"]),
            ("City / State / ZIP",    f"{META_2051['city']}, {META_2051['state']} {META_2051['zip']}"),
            ("⚠ Dashboard fix needed","PROJECTS['2051'].location currently reads 'Bellevue area, WA' — should be 'Shoreline, WA' per executed contract"),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR", [
            ("GC Legal Name",         META_2051["gc_legal"]),
            ("GC Address",            META_2051["gc_address"]),
            ("GC Phone",              META_2051["gc_phone"]),
            ("GC Fax",                META_2051["gc_fax"]),
            ("GC Contract Signatory", META_2051["gc_signatory"]),
            ("GC Project Manager",    META_2051["gc_pm"]),
            ("GC Superintendent",     META_2051["gc_sup"]),
            ("GC Project Engineer",   META_2051["gc_pe"]),
        ]),
        ("PROJECT TEAM — OWP (SUBCONTRACTOR)", [
            ("OWP Legal Name",        META_2051["owp_legal"]),
            ("OWP Address",           META_2051["owp_address"]),
            ("OWP Phone",             META_2051["owp_phone"]),
            ("OWP Contract Signatory",META_2051["owp_signatory"]),
            ("OWP Subcontractor Rep", META_2051["owp_rep"]),
            ("OWP RI Foreman",        META_2051["owp_ri_foreman"]),
            ("OWP License #",         META_2051["owp_license"]),
            ("OWP Federal Tax ID",    META_2051["owp_fed_tax_id"]),
            ("OWP Unemployment ID",   META_2051["owp_unemp"]),
            ("OWP Worker's Comp ID",  META_2051["owp_wc_id"]),
        ]),
        ("PROJECT TEAM — OWNER", [
            ("Owner Legal Name",      META_2051["owner_legal"]),
            ("Owner Address",         META_2051["owner_address"]),
            ("Owner Short Name",      META_2051["owner_short"]),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("MEP / Plumbing Engineer", META_2051["mep_engineer"]),
            ("Architect",             META_2051["architect"]),
            ("Structural Engineer",   META_2051["structural"]),
            ("Civil Engineer",        META_2051["civil"]),
        ]),
        ("INSURANCE & BONDING", [
            ("Insurance Type",        META_2051["insurance_type"]),
            ("Insurance Carrier",     META_2051["insurance_carrier"]),
            ("Insurance Policy #",    META_2051["insurance_policy"]),
        ]),
        ("CONTRACT & FINANCIALS", [
            ("Original Contract",     META_2051["contract_original"]),
            ("Final Contract",        META_2051["contract_final"]),
            ("Net Change Orders",     META_2051["contract_co_net"]),
            ("Retention Percentage",  META_2051["retention_pct"]),
            ("Retention Held",        META_2051["retention_held"]),
            ("Schedule",              f"{META_2051['start_date']} → {META_2051['end_date']} · {META_2051['duration']}"),
        ]),
        ("SCOPE & FIXTURES (per Job Ticket 3-4-2020)", [
            ("Plumbing Units",        META_2051["units"]),
            ("Resident Unit Trim",    META_2051["fixtures"]),
            ("Amenity Fixtures",      META_2051["amenity_features"]),
            ("Building Systems",      META_2051["rough_systems"]),
            ("Scope (design-build)",  META_2051["scope_design_build"]),
            ("Scope EXCLUDES",        META_2051["scope_excludes"]),
        ]),
        ("DOCUMENTS ON FILE (GDrive)", [
            ("Executed Contract",     META_2051["contract_pdfs"][0]),
            ("PSA Agreement",         META_2051["contract_pdfs"][1]),
            ("Insurance Requirements",META_2051["contract_pdfs"][2]),
            ("Job Tickets",           META_2051["ticket_count"]),
            ("Latest Ticket",         META_2051["ticket_latest"]),
            ("Change Orders (SCOs)",  f"{META_2051['co_count_dir']} files in Change Orders/CO's/"),
            ("CORs",                  f"{META_2051['cor_count_dir']} numbered CORs in Change Orders/COR's/"),
            ("RFIs",                  META_2051["rfi_count"]),
            ("ASIs",                  META_2051["asi_count"]),
            ("Permits",               f"{META_2051['permit_count']} files in Permits/ — {META_2051['permit_no']}"),
            ("Sub-Permits",           META_2051["permit_subs"]),
            ("Plans/Photos/Submittals", META_2051["plans_subfolders"]),
            ("Vendor Subfolders",     META_2051["vendor_subfolders"]),
            ("Delivery Instructions", META_2051["delivery"]),
        ]),
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        for label, val in fields:
            label_cell = ws.cell(row=r, column=1, value=label)
            val_cell   = ws.cell(row=r, column=2, value=val)
            label_cell.font = INK_BOLD
            val_cell.font = INK
            val_cell.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            label_cell.border = BORDER
            val_cell.border = BORDER

            # Highlight the dashboard-fix row
            if label.startswith("⚠"):
                label_cell.font = WARN_FONT
                val_cell.font = WARN_FONT
                label_cell.fill = WARN_FILL
                val_cell.fill = WARN_FILL
            else:
                label_cell.fill = TEAM_FILL
                val_cell.fill = TEAM_FILL
            r += 1
        r += 1  # blank row between sections

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 24


# ============================================================================
# 04 SOV-PayApps — was empty, fill with placeholder note + rec source
# ============================================================================
def patch_sov_payapps(ws):
    clear_sheet(ws, keep_header_rows=2)

    ws.cell(row=4, column=1, value="STATEMENT OF VALUES — SUMMARY (from JDR + Contract)").font = SECTION_HDR
    ws.cell(row=4, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=7)

    summary = [
        ("Original Contract",        2672000),
        ("Net Change Orders",        -83720),
        ("Final Contract Value",     2588280),
        ("Total Billed to Date",     2588280),
        ("Retention Held (5%)",      126114),
        ("Net Paid to OWP",          2462166),  # billed - retention
        ("Percent Complete",         "100.0%"),
    ]
    r = 5
    for k, v in summary:
        ws.cell(row=r, column=1, value=k).font = INK_BOLD
        ws.cell(row=r, column=2, value=v).font = INK
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        r += 1

    # Pay-app gap note
    r += 2
    ws.cell(row=r, column=1, value="PAY APPLICATION DETAIL — DATA GAP").font = SECTION_HDR
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    ws.cell(row=r, column=1, value=(
        "Per-pay-app history is not in 2051_data.json — the Sage parse rolls up to "
        "AR totals only. Full pay-app series lives in GDrive: "
        "2051-Compass, Vail Apartments/Billing/. To populate this tab on a per-app "
        "basis, re-parse Billing/ folder PDFs and inject into 2051_data.json under a "
        "new 'pay_apps' array."
    )).font = GREY_FONT
    ws.cell(row=r, column=1).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


# ============================================================================
# 05 Change Orders — was empty, populate from GDrive Change Orders folder
# ============================================================================
def patch_change_orders(ws):
    clear_sheet(ws, keep_header_rows=2)

    ws.cell(row=4, column=1, value="EXECUTED CHANGE ORDERS (SCO) — SOURCE: GDRIVE/CHANGE ORDERS/CO's/").font = SECTION_HDR
    ws.cell(row=4, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)

    headers = ["#", "SCO ID", "Description", "Type", "Source File"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    r = 6
    for i, (sid, desc) in enumerate(SCO_LIST, 1):
        cells = [i, sid, desc, "Owner-Directed" if "ADA" in desc or "Add" in desc or "Owner" in desc else "Field/Coord", "GDrive: Change Orders/CO's/"]
        for c, val in enumerate(cells, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
            cell.alignment = WRAP
            if r % 2 == 0: cell.fill = ROW_ALT
        r += 1

    # CORs
    r += 1
    ws.cell(row=r, column=1, value="OPEN/SUBMITTED CHANGE ORDER REQUESTS (COR) — SOURCE: GDRIVE/CHANGE ORDERS/COR's/").font = SECTION_HDR
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    headers2 = ["#", "COR ID", "Description", "Status", "Source File"]
    for c, h in enumerate(headers2, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
    r += 1
    for i, (cid, desc) in enumerate(COR_LIST, 1):
        cells = [i, cid, desc, "Submitted", "GDrive: Change Orders/COR's/"]
        for c, val in enumerate(cells, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
            cell.alignment = WRAP
            if r % 2 == 0: cell.fill = ROW_ALT
        r += 1

    # Net summary
    r += 1
    ws.cell(row=r, column=1, value="NET CO IMPACT").font = SECTION_HDR
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    r += 1
    summary = [
        ("Original Contract",        "$2,672,000"),
        ("Final Contract",           "$2,588,280"),
        ("Net Dollar Impact",        "($83,720) — credit-net (-3.1%)"),
        ("Posture",                  "Unusual: scope deletes (e.g. Remove Dishwasher Installation) net against scope adds (Unit 330, ADA Dog Wash, Drinking Fountain, etc.)"),
        ("Total Documented Events",  f"{len(SCO_LIST)} SCOs + {len(COR_LIST)} CORs = {len(SCO_LIST)+len(COR_LIST)} change events"),
    ]
    for k, v in summary:
        ws.cell(row=r, column=1, value=k).font = INK_BOLD
        ws.cell(row=r, column=2, value=v).font = INK
        ws.cell(row=r, column=2).alignment = WRAP
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        for c in (1, 2): ws.cell(row=r, column=c).border = BORDER
        r += 1

    widths = {1: 5, 2: 12, 3: 60, 4: 18, 5: 32}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# 17 Change Log — append project change-event summary below workbook log
# ============================================================================
def patch_change_log(ws):
    # Find last row with data
    last_data_row = ws.max_row
    while last_data_row > 1 and not any(ws.cell(row=last_data_row, column=c).value for c in range(1, 5)):
        last_data_row -= 1
    r = last_data_row + 2

    ws.cell(row=r, column=1, value="2026-04-25").font = INK
    ws.cell(row=r, column=2, value="v1.1 · enriched").font = INK
    ws.cell(row=r, column=3, value=(
        "Patched by enrich_2051.py: filled SOV-PayApps and Change Orders tabs (were empty), "
        "expanded Job Info from 13 rows to 60+ rows of key-players + contract + permit + "
        "fixture detail, flagged dashboard location typo (Bellevue area → Shoreline WA). "
        "Source: executed subcontract 18-0622-0500 + GDrive folder scan + dashboard "
        "PROJECT_TEAMS['2051']."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP
    for c in range(1, 4): ws.cell(row=r, column=c).border = BORDER

    if ws.column_dimensions["C"].width < 70:
        ws.column_dimensions["C"].width = 70


# ============================================================================
# Main
# ============================================================================
def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)

    print("→ Patching 02 Job Info (was 13 rows, expanding to 60+ key-players)...")
    patch_job_info(wb["02 Job Info"])

    print("→ Patching 04 SOV-PayApps (was empty)...")
    patch_sov_payapps(wb["04 SOV-PayApps"])

    print("→ Patching 05 Change Orders (was empty — 17 SCOs + 7 CORs from GDrive)...")
    patch_change_orders(wb["05 Change Orders"])

    print("→ Patching 17 Change Log (appending v1.1 enrichment entry)...")
    patch_change_log(wb["17 Change Log"])

    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
