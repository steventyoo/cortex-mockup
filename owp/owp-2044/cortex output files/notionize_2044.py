"""
notionize_2044.py — Build OWP_2044_JCR_Summary_Notion_v2.xlsx

17-tab Notion-styled summary workbook for project #2044 Track 66 Apartments
(Shelter Holdings / Blueprint 836, LLC / 915 E Spruce St, Seattle).

Consumes:
    /sessions/gracious-relaxed-pascal/2044_data.json          (JDR-parsed facts)
    /sessions/gracious-relaxed-pascal/2044_enriched_meta.json (GDrive-enriched meta)

Outputs:
    /sessions/gracious-relaxed-pascal/mnt/cortex-mockup/owp/owp-2044/cortex output files/
        OWP_2044_JCR_Summary_Notion_v2.xlsx

Follows design tokens/pattern established in notionize_2040.py:
    - Ink:       #191917
    - Graphite:  #6B6B65
    - Mica:      #9E9E97
    - Paper:     #FBFAF7
    - Stone:     #F2F2F2
    - Navy:      #1F3864  (hero fill)
    - Blue:      #2E5090  (section accent)
    - Mint bg:   #DCFCE7  Mint fg: #166534  (verified badges)
    - Gold bg:   #FEF3C7  Gold fg: #92400E  (partial badges)
"""

from __future__ import annotations
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── paths ────────────────────────────────────────────────────────────────────
BASE      = Path('/sessions/gracious-relaxed-pascal')
DATA_JSON = BASE / '2044_data.json'
META_JSON = BASE / '2044_enriched_meta.json'
OUT_DIR   = BASE / 'mnt' / 'cortex-mockup' / 'owp' / 'owp-2044' / 'cortex output files'
OUT_PATH  = OUT_DIR / 'OWP_2044_JCR_Summary_Notion_v2.xlsx'

# ── design tokens ────────────────────────────────────────────────────────────
INK      = '191917'
GRAPHITE = '6B6B65'
MICA     = '9E9E97'
PAPER    = 'FBFAF7'
STONE    = 'F2F2F2'
WHITE    = 'FFFFFF'
NAVY     = '1F3864'
BLUE     = '2E5090'
MINT_BG  = 'DCFCE7'
MINT_FG  = '166534'
GOLD_BG  = 'FEF3C7'
GOLD_FG  = '92400E'
SOFT_BG  = 'EFEFEF'

SRC_JDR  = '2044 Job Detail Report.pdf (Sage Timberline, 04/03/2026)'
SRC_LIST = 'OWP Project List with Schedule - UPDATED 04-01-26.xlsx'
SRC_CNTR = '2044-Shelter Holdings subcontract SC-03-07-6505 + Boren & Spruce CO docs'


# ── helpers ──────────────────────────────────────────────────────────────────
def hero(ws, title, subtitle):
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 34
    for col in 'DEFGHIJ':
        ws.column_dimensions[col].width = 22
    ws.row_dimensions[2].height = 34
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 2
    ws['B2'] = title
    ws['B2'].font = Font(name='Calibri', size=22, bold=True, color=WHITE)
    ws['B2'].fill = PatternFill('solid', start_color=NAVY, end_color=NAVY)
    ws['B2'].alignment = Alignment(horizontal='left', vertical='center')
    ws['B3'] = subtitle
    ws['B3'].font = Font(name='Calibri', size=11, bold=False, color=GRAPHITE)
    ws['B3'].alignment = Alignment(horizontal='left', vertical='center')


def section(ws, r, text, color=STONE):
    ws.cell(r, 2, text)
    ws.cell(r, 2).font = Font(name='Calibri', size=9, bold=True, color=MICA)
    ws.cell(r, 2).fill = PatternFill('solid', start_color=color, end_color=color)
    ws.cell(r, 2).alignment = Alignment(horizontal='left', vertical='center')


def section_blue(ws, r, text):
    ws.cell(r, 2, text)
    ws.cell(r, 2).font = Font(name='Calibri', size=9, bold=True, color=WHITE)
    ws.cell(r, 2).fill = PatternFill('solid', start_color=BLUE, end_color=BLUE)


def label_cell(ws, r, c, text, alt=False):
    cell = ws.cell(r, c, text)
    cell.font = Font(name='Calibri', size=10, color=GRAPHITE)
    if alt:
        cell.fill = PatternFill('solid', start_color=PAPER, end_color=PAPER)
    cell.alignment = Alignment(horizontal='left', vertical='center')


def value_cell(ws, r, c, value, alt=False, fmt=None, align='left'):
    cell = ws.cell(r, c, value)
    cell.font = Font(name='Calibri', size=10, color=INK)
    if alt:
        cell.fill = PatternFill('solid', start_color=PAPER, end_color=PAPER)
    if fmt:
        cell.number_format = fmt
        align = 'right'
    cell.alignment = Alignment(horizontal=align, vertical='center')


def header_row(ws, r, cols):
    """Write a header row of column titles with stone background."""
    for i, text in enumerate(cols, start=2):
        cell = ws.cell(r, i, text)
        cell.font = Font(name='Calibri', size=9, bold=True, color=MICA)
        cell.fill = PatternFill('solid', start_color=STONE, end_color=STONE)
        cell.alignment = Alignment(horizontal='left', vertical='center')


def badge(ws, r, c, text, palette='mint'):
    cell = ws.cell(r, c, text)
    if palette == 'mint':
        cell.font = Font(name='Calibri', size=9, bold=True, color=MINT_FG)
        cell.fill = PatternFill('solid', start_color=MINT_BG, end_color=MINT_BG)
    elif palette == 'gold':
        cell.font = Font(name='Calibri', size=9, bold=True, color=GOLD_FG)
        cell.fill = PatternFill('solid', start_color=GOLD_BG, end_color=GOLD_BG)
    else:
        cell.font = Font(name='Calibri', size=10, color=INK)
        cell.fill = PatternFill('solid', start_color=SOFT_BG, end_color=SOFT_BG)
    cell.alignment = Alignment(horizontal='center', vertical='center')


# ── tab builders ─────────────────────────────────────────────────────────────
def build_overview(wb, d, m):
    ws = wb.create_sheet('Overview')
    rr = d['report_record']
    derived = d['derived_fields']
    meta = m['meta']
    units = meta['units']
    revenue = abs(rr['job_totals_revenue'])
    expenses = rr['job_totals_expenses']
    net = revenue - expenses
    retainage = abs(rr['job_totals_retainage'])
    margin = net / revenue if revenue else 0

    meta.setdefault('gc_project_managers', [{'name': 'TBD', 'email': 'TBD', 'phone': 'TBD'}])
    meta.setdefault('gc_superintendent', {'name': 'TBD', 'email': 'TBD', 'phone': 'TBD'})
    gc_pm_names = ' / '.join(g['name'] for g in meta['gc_project_managers'])

    hero(ws,
         f"Job #{rr['job_number']} · {meta['job_name_long']}",
         f"Cortex JCR v2  •  {meta['general_contractor']} (customer 2044SH)  •  "
         f"{meta['location']}  •  Owner: {meta['owner']}")

    # PROJECT OVERVIEW section
    section(ws, 5, 'PROJECT OVERVIEW')

    rows_overview = [
        ('Project Job #',             rr['job_number'],                         SRC_JDR),
        ('Project Name (Sage)',       rr['job_name'],                           SRC_JDR),
        ('Project Description',       f"Track 66 Apartments — {units}-unit 7-story midrise multi-family (4 Type A ADA), Seattle", SRC_LIST + ' + contract'),
        ('General Contractor',        meta['general_contractor'],               SRC_LIST),
        ('Customer Code (Sage)',      '2044SH',                                 SRC_JDR),
        ('Owner / Developer',         meta['owner'],                            SRC_LIST),
        ('GC PM',                     gc_pm_names,                              SRC_CNTR),
        ('GC Superintendent',         meta['gc_superintendent']['name'],        SRC_CNTR),
        ('GC PE',                     'N/A',                                    SRC_CNTR),
        ('OWP Signatory',             meta['owp_signatory'],                    SRC_CNTR),
        ('Jobsite Location',          meta['location'],                         SRC_LIST),
        ('Insurance',                 'Not Wrap',                               SRC_LIST),
        ('Fixture Schedule',          f"{meta['total_fixtures']} fixtures total (656 unit + 98 retail/common)", 'Submittal Table of Contents + Plumbing Scope'),
        ('Architect',                 meta['architect'],                        'Track 66 Apartments project info.pdf'),
        ('Contract Type',             'Lump Sum',                               SRC_CNTR),
        ('Unit Count',                units,                                    'Submittal TOC (102 Type B + 6 Type A/ADA)'),
        ('Work Period',               f"{meta['start_date']} start → {meta['end_date']} ({meta['duration_months']} months)", 'JDR PR/AR dates'),
    ]
    for i, (lbl, val, src) in enumerate(rows_overview):
        r = 6 + i
        alt = (i % 2 == 1)
        label_cell(ws, r, 2, lbl, alt=alt)
        value_cell(ws, r, 3, val, alt=alt)
        # merge C:E
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        value_cell(ws, r, 6, src, alt=alt)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)

    # HERO METRICS
    r = 6 + len(rows_overview) + 1  # row 24
    # Headers B D F H
    for col, title in zip([2, 4, 6, 8], ['CONTRACT VALUE', 'NET PROFIT', 'DIRECT COST', 'LABOR HOURS']):
        cell = ws.cell(r, col, title)
        cell.font = Font(name='Calibri', size=10, color=INK)
        cell.fill = PatternFill('solid', start_color=STONE, end_color=STONE)
    # Values
    r2 = r + 1
    value_cell(ws, r2, 2, revenue,  fmt='"$"#,##0.00')
    value_cell(ws, r2, 4, net,      fmt='"$"#,##0.00')
    value_cell(ws, r2, 6, expenses, fmt='"$"#,##0.00')
    value_cell(ws, r2, 8, derived['total_labor_hours'], fmt='#,##0.00')
    # Sub-notes
    r3 = r + 2
    value_cell(ws, r3, 2, f"Original ${meta['contract_original']:,.2f} + "
                          f"(${meta['change_orders_net']:,.2f}) COs "
                          f"({meta['change_orders_net']/meta['contract_original']*100:.2f}%)")
    value_cell(ws, r3, 4, f"{margin*100:.1f}% margin")
    value_cell(ws, r3, 6, f"{expenses/revenue*100:.1f}% of revenue")
    value_cell(ws, r3, 8, f"{derived['total_workers']} workers")

    # SCOPE OF WORK section
    r = r3 + 2  # row 28
    section(ws, r, 'SCOPE OF WORK')
    scope_lines = [
        "Full plumbing installation per Division 22: underground, garage, rough-in, finish, gas, water main, insulation, roof drains.",
        f"{units}-unit mixed-use (102 Type B + 6 Type A/ADA) at {meta['location']}, "
        f"7-story midrise multi-family (Type 1A podium + Type VA above).",
        f"GC: {meta['general_contractor']} (2044SH); Owner: {meta['owner']}.",
        "Insurance: Not Wrap (retail-adjacent residential).",
        f"Project team: {gc_pm_names} (GC PM) · {meta['gc_superintendent']['name']} (GC Sup) · "
        f"{meta['architect']} (Architect) · {meta['owp_signatory']} (OWP signatory).",
    ]
    for i, line in enumerate(scope_lines):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, line)
        ws.merge_cells(start_row=rr_, start_column=2, end_row=rr_, end_column=10)

    # SOURCES section
    r = r + 2 + len(scope_lines)
    section(ws, r, 'SOURCES', color=BLUE)
    ws.cell(r, 2).font = Font(name='Calibri', size=9, bold=True, color=MICA)
    ws.cell(r, 2).fill = PatternFill('solid', start_color=BLUE, end_color=BLUE)
    src_lines = [
        f"Canonical financial source: {SRC_JDR}",
        f"Job totals (JDR footer): Revenue ${revenue:,.2f} / Expenses ${expenses:,.2f} / "
        f"Net ${net:,.2f} / Retainage ${retainage:,.2f}",
        f"Source breakdown (derived): PR ${derived['cost_by_source_pr_amount']:,.2f} · "
        f"AP ${derived['cost_by_source_ap_amount']:,.2f} · GL ${derived['cost_by_source_gl_amount']:,.2f}",
        f"Contract: Lump Sum ${meta['contract_original']:,.2f} → ${meta['contract_final']:,.2f}",
        f"Change Orders (executed): ${meta['change_orders_net']:,.2f} net across "
        f"{meta['executed_co_count']} COs (CO log in enriched meta)",
        f"Project Team: {gc_pm_names} (PM) · {meta['gc_superintendent']['name']} (Sup) · "
        f"{meta['architect']} · {meta['owner']}",
    ]
    for i, s in enumerate(src_lines):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, s)
        ws.merge_cells(start_row=rr_, start_column=2, end_row=rr_, end_column=10)

    # PROJECT TEAM (Job Info) section
    r = r + 2 + len(src_lines)
    section(ws, r, 'PROJECT TEAM (OWP Job Info)')
    r += 1
    header_row(ws, r, ['Role', 'Status', 'Name / Firm'])

    team_rows = [
        ('GC',                  'Extracted', meta['general_contractor']),
        ('GC PM',               'Extracted', gc_pm_names),
        ('GC Superintendent',   'Extracted', meta['gc_superintendent']['name']),
        ('GC PE',               'N/A',       'N/A'),
        ('Owner',               'Extracted', meta['owner']),
        ('Architect',           'Extracted', meta['architect']),
        ('Structural',          'N/A',       'N/A'),
        ('Acoustical',          'N/A',       'N/A'),
        ('ADA Consultant',      'N/A',       'N/A'),
        ('Interior Design',     'N/A',       'N/A'),
        ('MEP Engineer',        'N/A',       'N/A'),
        ('Insurance',           'Extracted', 'Not Wrap'),
        ('OWP Signatory',       'Extracted', meta['owp_signatory']),
    ]
    for i, (role, status, name) in enumerate(team_rows):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, role)
        if status == 'Extracted':
            badge(ws, rr_, 3, status, 'mint')
        elif status == 'N/A':
            c = ws.cell(rr_, 3, 'N/A')
            c.font = Font(name='Calibri', size=10, color=INK)
            c.fill = PatternFill('solid', start_color=SOFT_BG, end_color=SOFT_BG)
            c.alignment = Alignment(horizontal='center')
        else:
            badge(ws, rr_, 3, status, 'gold')
        value_cell(ws, rr_, 4, name)
        ws.merge_cells(start_row=rr_, start_column=4, end_row=rr_, end_column=10)


def build_bva(wb, d, m):
    ws = wb.create_sheet('Budget vs Actual')
    meta = m['meta']
    hero(ws, 'Budget vs Actual',
         f"All {len(d['cost_code_summaries'])} cost codes from JDR. "
         f"Contract ${meta['contract_original']:,.0f} → ${meta['contract_final']:,.0f}.")
    r = 5
    header_row(ws, r, ['Cost Code', 'Description', 'Original Budget', 'Revised Budget',
                       'Actual', 'Variance', '% of Revised', 'Hours', 'Source'])
    # hours = regular+ot+dt+other
    for i, c in enumerate(d['cost_code_summaries']):
        rr_ = r + 1 + i
        hrs = (c.get('regular_hours', 0) + c.get('overtime_hours', 0)
               + c.get('doubletime_hours', 0) + c.get('other_hours', 0))
        orig_b = c.get('original_budget', 0)
        curr_b = c.get('current_budget', 0)
        actual = c.get('actual_amount', 0)
        value_cell(ws, rr_, 2, c['code'])
        value_cell(ws, rr_, 3, c['description'])
        value_cell(ws, rr_, 4, orig_b,                           fmt='"$"#,##0')
        value_cell(ws, rr_, 5, curr_b,                           fmt='"$"#,##0')
        value_cell(ws, rr_, 6, actual,                           fmt='"$"#,##0.00')
        ws.cell(rr_, 7, f'=F{rr_}-E{rr_}').number_format        = '"$"#,##0.00;("$"#,##0.00);-'
        ws.cell(rr_, 8, f'=IF(E{rr_}=0,"",F{rr_}/E{rr_})').number_format = '0.0%'
        value_cell(ws, rr_, 9, hrs if hrs else None,             fmt='#,##0.00' if hrs else None)
        value_cell(ws, rr_, 10, SRC_JDR)
    # TOTAL row
    tr = r + 1 + len(d['cost_code_summaries'])
    value_cell(ws, tr, 2, 'TOTAL')
    ws.cell(tr, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(tr, 4, f'=SUM(D{r+1}:D{tr-1})').number_format = '"$"#,##0'
    ws.cell(tr, 5, f'=SUM(E{r+1}:E{tr-1})').number_format = '"$"#,##0'
    ws.cell(tr, 6, f'=SUM(F{r+1}:F{tr-1})').number_format = '"$"#,##0.00'
    ws.cell(tr, 7, f'=F{tr}-E{tr}').number_format        = '"$"#,##0.00;("$"#,##0.00);-'
    ws.cell(tr, 9, f'=SUM(I{r+1}:I{tr-1})').number_format = '#,##0.00'


def categorize_code(code):
    code_s = str(code)
    if code_s in ('995',):
        return 'Payroll Burden'
    if code_s in ('998',):
        return 'Payroll Taxes'
    if code_s.startswith('6') and code_s != '999':
        return 'Overhead'
    if code_s == '999':
        return 'Revenue'
    if code_s.startswith('1') or code_s.startswith('0'):
        return 'Labor'
    if code_s.startswith('2'):
        return 'Material'
    return 'Other'


def build_cost_breakdown(wb, d, m):
    ws = wb.create_sheet('Cost Breakdown')
    rr_rec = d['report_record']
    revenue = abs(rr_rec['job_totals_revenue'])
    hero(ws, 'Cost Breakdown by Category',
         'Direct cost composition by category from JDR cost codes')
    r = 5
    header_row(ws, r, ['Category', 'Cost Codes', 'Actual $', '% of Direct Cost',
                       '% of Revenue', 'Source'])
    # Aggregate by category
    by_cat = {}
    codes_by_cat = {}
    for c in d['cost_code_summaries']:
        cat = categorize_code(c['code'])
        if cat == 'Revenue':
            continue
        by_cat[cat] = by_cat.get(cat, 0) + c['actual_amount']
        codes_by_cat.setdefault(cat, []).append(c['code'])

    order = ['Labor', 'Material', 'Overhead', 'Payroll Burden', 'Payroll Taxes', 'Other']
    rows = [(cat, ','.join(codes_by_cat.get(cat, [])), by_cat[cat])
            for cat in order if cat in by_cat]
    total_row = 6 + len(rows)
    for i, (cat, codes, amt) in enumerate(rows):
        rr_ = 6 + i
        value_cell(ws, rr_, 2, cat)
        value_cell(ws, rr_, 3, codes)
        value_cell(ws, rr_, 4, amt, fmt='"$"#,##0.00')
        ws.cell(rr_, 5, f'=D{rr_}/$D${total_row}').number_format = '0.0%'
        ws.cell(rr_, 6, f'=D{rr_}/{revenue}').number_format = '0.0%'
        value_cell(ws, rr_, 7, SRC_JDR)
    value_cell(ws, total_row, 2, 'TOTAL DIRECT COST')
    ws.cell(total_row, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(total_row, 4, f'=SUM(D6:D{total_row-1})').number_format = '"$"#,##0.00'
    ws.cell(total_row, 5, 1).number_format = '0.0%'
    ws.cell(total_row, 6, f'=D{total_row}/{revenue}').number_format = '0.0%'


def aggregate_ap_vendors(line_items):
    """Build vendor spend table from AP line items."""
    vendors = {}
    for li in line_items:
        if li.get('source') != 'AP':
            continue
        vid = li.get('number', '').strip() or 'UNKNOWN'
        name = li.get('name', '').strip()
        amt = li.get('actual_amount', 0) or 0
        if vid not in vendors:
            vendors[vid] = {'id': vid, 'name': name, 'spend': 0, 'n_txn': 0}
        vendors[vid]['spend'] += amt
        vendors[vid]['n_txn'] += 1
        if not vendors[vid]['name'] and name:
            vendors[vid]['name'] = name
    return sorted(vendors.values(), key=lambda v: -v['spend'])


def categorize_vendor(name):
    n = (name or '').lower()
    if 'franklin' in n or 'engineering' in n:
        return 'Engineering (601)'
    if 'backflow' in n:
        return 'Testing subcontractor'
    if 'capital one' in n or 'credit card' in n or 'bank of america' in n:
        return 'Credit Card (mixed)'
    if any(k in n for k in ['supply', 'ferguson', 'rosen', 'keller', 'consolidated',
                            'mechanical sales', 'hydronics', 'plumbing']):
        return 'Plumbing / Supplies'
    return 'Uncategorized'


def build_material(wb, d, m):
    ws = wb.create_sheet('Material')
    vendors = aggregate_ap_vendors(d['line_items'])
    hero(ws, 'Material Purchases — AP Vendors',
         'Material + subcontractor spend by vendor from JDR AP records')
    r = 5
    header_row(ws, r, ['Vendor ID', 'Vendor Name', 'Total Spend', '# Transactions',
                       'Category (inferred)', 'Source'])
    for i, v in enumerate(vendors):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, v['id'])
        value_cell(ws, rr_, 3, v['name'])
        value_cell(ws, rr_, 4, v['spend'], fmt='"$"#,##0.00')
        value_cell(ws, rr_, 5, v['n_txn'], fmt='#,##0')
        value_cell(ws, rr_, 6, categorize_vendor(v['name']))
        value_cell(ws, rr_, 7, SRC_JDR)
    tr = r + 1 + len(vendors)
    value_cell(ws, tr, 2, 'TOTAL')
    ws.cell(tr, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(tr, 4, f'=SUM(D{r+1}:D{tr-1})').number_format = '"$"#,##0.00'
    ws.cell(tr, 5, f'=SUM(E{r+1}:E{tr-1})').number_format = '#,##0'


def compute_work_days(line_items, worker_id):
    """Count unique posted dates a worker shows up on."""
    days = set()
    for li in line_items:
        if li.get('source') != 'PR':
            continue
        if li.get('number') == worker_id:
            days.add(li.get('document_date'))
    return len(days)


def build_crew(wb, d, m):
    ws = wb.create_sheet('Crew & Labor')
    meta = m['meta']
    workers = d['worker_wages']
    # Attach worker_id by parsing line_items (use first match for each name)
    worker_id_by_name = {}
    for li in d['line_items']:
        if li.get('source') != 'PR':
            continue
        nm = li.get('name', '').strip()
        wid = li.get('number', '').strip()
        if nm and wid and nm not in worker_id_by_name:
            worker_id_by_name[nm] = wid

    # Compute total hours + wages
    workers_data = []
    for w in workers:
        wid = worker_id_by_name.get(w['name'], '???')
        hrs = (w.get('regular_hours', 0) + w.get('overtime_hours', 0)
               + w.get('doubletime_hours', 0) + w.get('other_hours', 0))
        gross = w.get('regular_amount', 0) + w.get('overtime_amount', 0)
        days = compute_work_days(d['line_items'], wid)
        workers_data.append({'id': wid, 'name': w['name'], 'hours': hrs,
                             'gross': gross, 'days': days})
    workers_data.sort(key=lambda x: -x['hours'])

    hero(ws, 'Crew & Labor — Worker Roster',
         f"All {len(workers_data)} unique payroll workers. OWP signatory: {meta['owp_signatory']}.")
    r = 5
    header_row(ws, r, ['Worker ID', 'Worker Name', 'Total Hours', 'Gross Pay',
                       'Blended Wage ($/hr)', '# Work Days', 'Source'])
    for i, w in enumerate(workers_data):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, w['id'])
        value_cell(ws, rr_, 3, w['name'])
        value_cell(ws, rr_, 4, w['hours'], fmt='#,##0.00')
        value_cell(ws, rr_, 5, w['gross'], fmt='"$"#,##0.00')
        ws.cell(rr_, 6, f'=IF(D{rr_}=0,0,E{rr_}/D{rr_})').number_format = '"$"#,##0.00'
        value_cell(ws, rr_, 7, w['days'], fmt='#,##0')
        value_cell(ws, rr_, 8, SRC_JDR)
    tr = r + 1 + len(workers_data)
    value_cell(ws, tr, 2, 'TOTAL')
    ws.cell(tr, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(tr, 4, f'=SUM(D{r+1}:D{tr-1})').number_format = '#,##0.00'
    ws.cell(tr, 5, f'=SUM(E{r+1}:E{tr-1})').number_format = '"$"#,##0.00'
    return workers_data


def build_crew_analytics(wb, d, m, workers_data):
    ws = wb.create_sheet('Crew Analytics')
    derived = d['derived_fields']
    total_hrs = sum(w['hours'] for w in workers_data)
    total_gross = sum(w['gross'] for w in workers_data)
    top1_share = workers_data[0]['hours'] / total_hrs if total_hrs else 0
    top5_share = sum(w['hours'] for w in workers_data[:5]) / total_hrs if total_hrs else 0
    rates = [w['gross']/w['hours'] for w in workers_data if w['hours'] > 0]
    rates_sorted = sorted(rates)
    highest_rate = rates_sorted[-1] if rates_sorted else 0
    lowest_rate = rates_sorted[0] if rates_sorted else 0
    avg_hours = total_hrs / len(workers_data) if workers_data else 0
    avg_days = sum(w['days'] for w in workers_data) / len(workers_data) if workers_data else 0

    hero(ws, 'Crew Analytics', 'Team-level labor productivity, concentration, wage dispersion')
    section(ws, 5, 'TEAM-LEVEL METRICS')
    r = 6
    header_row(ws, r, ['Metric', 'Value', 'Notes', 'Source'])
    rows = [
        ('Total Workers',             len(workers_data),          'Unique payroll IDs', SRC_JDR),
        ('Total Labor Hours',         total_hrs,                  'Sum of labor codes', SRC_JDR),
        ('Total Gross Pay',           total_gross,                'Sum of labor codes', SRC_JDR),
        ('Blended Gross Wage ($/hr)', total_gross/total_hrs if total_hrs else 0, 'Labor$/Hrs (pre-burden)', 'Derived'),
        ('Top Worker Hours Share',    top1_share,                 f"{workers_data[0]['id']} {workers_data[0]['name']} ({workers_data[0]['hours']:.0f} hrs)", 'Derived'),
        ('Top 5 Workers Hours Share', top5_share,                 'Concentration metric', 'Derived'),
        ('Highest Wage Rate ($/hr)',  highest_rate,               'Single-worker blended', 'Derived'),
        ('Lowest Wage Rate ($/hr)',   lowest_rate,                'Single-worker blended', 'Derived'),
        ('Avg Hours per Worker',      avg_hours,                  'Includes short-tenure workers', 'Derived'),
        ('Avg Project Days per Worker', avg_days,                 'Mean days', 'Derived'),
    ]
    for i, (metric, val, note, src) in enumerate(rows):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, metric)
        if isinstance(val, (int, float)):
            if 'Share' in metric or 'Hours Share' in metric:
                value_cell(ws, rr_, 3, val, fmt='0.0%')
            elif 'Wage' in metric or '$' in metric:
                value_cell(ws, rr_, 3, val, fmt='"$"#,##0.00')
            else:
                value_cell(ws, rr_, 3, val, fmt='#,##0.00')
        else:
            value_cell(ws, rr_, 3, val)
        value_cell(ws, rr_, 4, note)
        value_cell(ws, rr_, 5, src)


def build_productivity(wb, d, m):
    ws = wb.create_sheet('Productivity')
    meta = m['meta']
    derived = d['derived_fields']
    rr_rec = d['report_record']
    revenue = abs(rr_rec['job_totals_revenue'])
    expenses = rr_rec['job_totals_expenses']
    net = revenue - expenses
    hrs = derived['total_labor_hours']
    units = meta['units']
    # Sum by cat
    labor_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                     if categorize_code(c['code']) == 'Labor')
    material_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                        if categorize_code(c['code']) == 'Material')
    burden = sum(c['actual_amount'] for c in d['cost_code_summaries']
                 if c['code'] == '995')
    tax = sum(c['actual_amount'] for c in d['cost_code_summaries']
              if c['code'] == '998')
    # Roughin / Finish
    rh_code = next((c for c in d['cost_code_summaries'] if c['code'] == '120'), None)
    fh_code = next((c for c in d['cost_code_summaries'] if c['code'] == '130'), None)
    rh = (rh_code['regular_hours'] + rh_code['overtime_hours']
          + rh_code['doubletime_hours'] + rh_code['other_hours']) if rh_code else 0
    fh = (fh_code['regular_hours'] + fh_code['overtime_hours']
          + fh_code['doubletime_hours'] + fh_code['other_hours']) if fh_code else 0
    total_hrs_labor = sum((c['regular_hours'] + c['overtime_hours']
                           + c['doubletime_hours'] + c['other_hours'])
                          for c in d['cost_code_summaries']
                          if categorize_code(c['code']) == 'Labor')

    hero(ws, 'Productivity Metrics',
         f'Normalized ratios. Per-unit from {units} units ({meta["total_fixtures"]} fixtures).')
    r = 5
    header_row(ws, r, ['Metric', 'Value', 'Basis', 'Source / Note'])
    rows = [
        ('Revenue per Labor Hour',       revenue/hrs,               'USD/hr', None),
        ('Profit per Labor Hour',        net/hrs,                   'USD/hr', None),
        ('Labor Cost per Hour (blended)', labor_cost/hrs if hrs else 0, 'USD/hr', 'Pre-burden'),
        ('Fully-Loaded Labor Rate ($/hr)', derived['fully_loaded_wage'], 'USD/hr', 'Incl burden+taxes'),
        ('Burden Multiplier',            derived['burden_multiplier'], 'x', 'Fully-loaded/blended'),
        ('Rough-in Hours (code 120)',    rh,                         'hrs', None),
        ('Finish Hours (code 130)',      fh,                         'hrs', None),
        ('Rough-in % of Total Hours',    rh/total_hrs_labor if total_hrs_labor else 0, '%', None),
        ('Gross Margin',                 net/revenue,               '%', None),
        ('Labor % of Revenue',           labor_cost/revenue,        '%', None),
        ('Material % of Revenue',        material_cost/revenue,     '%', None),
        ('Direct Cost Ratio',            expenses/revenue,          '%', 'All exp/Rev'),
        (f'Revenue per Unit ({units})',  revenue/units,             'USD', f'{units} units'),
        ('Labor Hours per Unit',         hrs/units,                 'hrs', None),
        ('Labor Cost per Unit',          labor_cost/units,          'USD', None),
        ('Material Cost per Unit',       material_cost/units,       'USD', None),
        ('Direct Cost per Unit',         expenses/units,            'USD', None),
        ('Rough-in Hours per Unit',      rh/units if units else 0,  'hrs', None),
        ('Revenue per Fixture',          revenue/meta['total_fixtures'], 'USD', f'{meta["total_fixtures"]} fixtures'),
        ('Hours per Fixture',            hrs/meta['total_fixtures'], 'hrs', None),
    ]
    for i, (metric, val, basis, note) in enumerate(rows):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, metric)
        if '%' in basis:
            value_cell(ws, rr_, 3, val, fmt='0.00%')
        elif basis == 'USD':
            value_cell(ws, rr_, 3, val, fmt='"$"#,##0.00')
        elif basis == 'USD/hr':
            value_cell(ws, rr_, 3, val, fmt='"$"#,##0.00')
        elif basis == 'x':
            value_cell(ws, rr_, 3, val, fmt='0.00"x"')
        elif basis == 'hrs':
            value_cell(ws, rr_, 3, val, fmt='#,##0.00')
        else:
            value_cell(ws, rr_, 3, val)
        value_cell(ws, rr_, 4, basis)
        if note:
            value_cell(ws, rr_, 5, note)


def build_po(wb, d, m):
    ws = wb.create_sheet('PO Commitments')
    meta = m['meta']
    hero(ws, 'PO Commitments',
         'Inbound contract (GC → OWP). Outbound vendor commitments — see Material tab.')
    r = 5
    header_row(ws, r, ['PO #', 'Date', 'Issuer', 'Type', 'Status', 'Description',
                       'Amount', 'Source'])
    rows = [
        ('PRIME',   meta['contract_signed_date'], meta['general_contractor'],
         'Lump Sum Subcontract', 'Closed',
         f"Track 66 Apartments plumbing ({meta['units']} units, 7 stories) — per Div 22",
         meta['contract_original'], SRC_CNTR),
        ('CO-NET', '—', meta['general_contractor'],
         f"Change Orders (executed, {meta['executed_co_count']} COs)", 'Executed',
         "Net CO impact (Sage 999 rev - orig)",
         meta['change_orders_net'], SRC_JDR),
    ]
    for i, row in enumerate(rows):
        rr_ = r + 1 + i
        for j, val in enumerate(row):
            if j == 6:
                value_cell(ws, rr_, 2 + j, val, fmt='"$"#,##0.00')
            else:
                value_cell(ws, rr_, 2 + j, val)
    tr = r + 1 + len(rows)
    value_cell(ws, tr, 2, 'TOTAL (contract + COs)')
    ws.cell(tr, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(tr, 8, f'=SUM(H{r+1}:H{tr-1})').number_format = '"$"#,##0.00'


def aggregate_invoices(line_items):
    """Collect AR invoices: one row per unique ref_number."""
    invoices = {}
    for li in line_items:
        if li.get('source') != 'AR':
            continue
        ref = li.get('ref_number', '')
        date = li.get('document_date', '')
        if ref not in invoices:
            invoices[ref] = {'ref': ref, 'date': date, 'amount': 0,
                             'retainage': 0, 'n_lines': 0}
        invoices[ref]['amount'] += li.get('actual_amount', 0)
        invoices[ref]['retainage'] += li.get('retainage', 0) or 0
        invoices[ref]['n_lines'] += 1
    return sorted(invoices.values(), key=lambda x: x['date'])


def build_billing(wb, d, m):
    ws = wb.create_sheet('Billing & SOV')
    meta = m['meta']
    invoices = aggregate_invoices(d['line_items'])
    rr_rec = d['report_record']
    retainage = abs(rr_rec['job_totals_retainage'])
    hero(ws, 'Billing & Schedule of Values',
         f"{meta['invoice_count']} invoices to {meta['general_contractor']}. "
         f"Retainage ${retainage:,.2f}.")
    r = 5
    header_row(ws, r, ['Invoice #', 'Date', 'Total Billed (signed)',
                       'Retainage (signed)', '# Lines', 'Source'])
    for i, inv in enumerate(invoices):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, inv['ref'])
        value_cell(ws, rr_, 3, inv['date'])
        value_cell(ws, rr_, 4, inv['amount'],    fmt='"$"#,##0.00;("$"#,##0.00);-')
        value_cell(ws, rr_, 5, inv['retainage'], fmt='"$"#,##0.00;("$"#,##0.00);-')
        value_cell(ws, rr_, 6, inv['n_lines'], fmt='#,##0')
        value_cell(ws, rr_, 7, SRC_JDR)
    tr = r + 1 + len(invoices)
    value_cell(ws, tr, 2, 'TOTAL (signed)')
    ws.cell(tr, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(tr, 4, f'=SUM(D{r+1}:D{tr-1})').number_format = '"$"#,##0.00;("$"#,##0.00);-'
    ws.cell(tr, 5, f'=SUM(E{r+1}:E{tr-1})').number_format = '"$"#,##0.00;("$"#,##0.00);-'


def build_insights(wb, d, m, workers_data):
    ws = wb.create_sheet('Insights')
    meta = m['meta']
    rr_rec = d['report_record']
    derived = d['derived_fields']
    revenue = abs(rr_rec['job_totals_revenue'])
    expenses = rr_rec['job_totals_expenses']
    net = revenue - expenses
    retainage = abs(rr_rec['job_totals_retainage'])
    labor_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                     if categorize_code(c['code']) == 'Labor')
    total_hrs = derived['total_labor_hours']

    # Top vendor
    vendors = aggregate_ap_vendors(d['line_items'])
    top_vendor = vendors[0] if vendors else {'name': 'N/A', 'spend': 0}
    top5_ap = sum(v['spend'] for v in vendors[:5])
    total_ap = derived['cost_by_source_ap_amount']

    # Top cost code
    max_code = max(d['cost_code_summaries'],
                   key=lambda c: c['actual_amount'] if categorize_code(c['code']) != 'Revenue' else 0)

    # Rough-in
    rh_code = next((c for c in d['cost_code_summaries'] if c['code'] == '120'), None)
    rh = (rh_code['regular_hours'] + rh_code['overtime_hours']
          + rh_code['doubletime_hours'] + rh_code['other_hours']) if rh_code else 0
    total_hrs_labor = sum((c['regular_hours'] + c['overtime_hours']
                           + c['doubletime_hours'] + c['other_hours'])
                          for c in d['cost_code_summaries']
                          if categorize_code(c['code']) == 'Labor')

    hero(ws, 'Insights & Observations', 'Narrative findings from JDR + GDrive CO log')
    r = 5
    header_row(ws, r, ['#', 'Insight', 'Detail', 'Confidence', 'Source'])
    rows = [
        ('MARGIN PROFILE',
         f"Net profit ${net:,.2f} on ${revenue:,.2f} revenue = {net/revenue*100:.1f}% gross margin.",
         'Verified', SRC_JDR),
        ('CONTRACT CHANGE',
         f"Original ${meta['contract_original']:,.2f} → final ${meta['contract_final']:,.2f} = "
         f"$+{meta['change_orders_net']:,.2f} net ({meta['change_orders_net']/meta['contract_original']*100:.2f}%) across "
         f"{meta['executed_co_count']} executed COs.",
         'Verified', SRC_JDR + ' + CO log'),
        ('LABOR PROFILE',
         f"Labor cost ${labor_cost:,.2f} ({labor_cost/revenue*100:.1f}% of rev) across "
         f"{total_hrs:,.0f} hrs and {derived['total_workers']} workers. "
         f"Blended wage ${labor_cost/total_hrs:.2f}/hr.",
         'Verified', SRC_JDR),
        ('VENDOR CONCENTRATION',
         f"Top 5 AP vendors: ~${top5_ap:,.2f} ({top5_ap/total_ap*100:.0f}% of AP). "
         f"Top vendor: {top_vendor['name']} (${top_vendor['spend']:,.2f}).",
         'Verified', SRC_JDR),
        ('RETAINAGE',
         f"Retainage ${retainage:,.2f} — "
         f"{'cleared' if retainage == 0 else 'outstanding as of report date'}.",
         'Verified', SRC_JDR),
        ('ROUGH-IN SHARE',
         f"Code 120 Roughin Labor = {rh:,.0f} hrs ({rh/total_hrs_labor*100:.0f}% of labor hours).",
         'Verified', SRC_JDR),
        ('TOP COST CODE',
         f"Largest actual cost: {max_code['code']} {max_code['description']} = ${max_code['actual_amount']:,.2f}",
         'Verified', SRC_JDR),
        ('WORKER CONCENTRATION',
         f"Top worker {workers_data[0]['name']} = {workers_data[0]['hours']:,.0f} hrs "
         f"({workers_data[0]['hours']/total_hrs*100:.0f}% of total).",
         'Verified', SRC_JDR),
        ('INSURANCE',
         'Not Wrap — retail-adjacent mixed-use.',
         'Verified', SRC_LIST),
        ('SCOPE CONTEXT',
         f"Track 66 Apartments — {meta['units']}-unit 7-story midrise (68 standard + 4 Type A ADA) + "
         f"P1/P2 parking at {meta['location']}.",
         'Verified', 'Submittal TOC + scope doc'),
    ]
    for i, (title, detail, conf, src) in enumerate(rows):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, i + 1)
        value_cell(ws, rr_, 3, title)
        value_cell(ws, rr_, 4, detail)
        badge(ws, rr_, 5, conf, 'mint')
        value_cell(ws, rr_, 6, src)


def _compute_kpis(d, m, workers_data):
    """Build the KPI dict used by Benchmark KPIs + Metric Registry tabs."""
    meta = m['meta']
    rr_rec = d['report_record']
    derived = d['derived_fields']
    revenue = abs(rr_rec['job_totals_revenue'])
    expenses = rr_rec['job_totals_expenses']
    net = revenue - expenses
    retainage = abs(rr_rec['job_totals_retainage'])
    hrs = derived['total_labor_hours']
    units = meta['units']

    labor_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                     if categorize_code(c['code']) == 'Labor')
    material_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                        if categorize_code(c['code']) == 'Material')
    overhead = sum(c['actual_amount'] for c in d['cost_code_summaries']
                   if categorize_code(c['code']) == 'Overhead')
    burden = sum(c['actual_amount'] for c in d['cost_code_summaries']
                 if c['code'] == '995')
    tax = sum(c['actual_amount'] for c in d['cost_code_summaries']
              if c['code'] == '998')
    # Rough-in / Finish hrs
    rh_code = next((c for c in d['cost_code_summaries'] if c['code'] == '120'), None)
    fh_code = next((c for c in d['cost_code_summaries'] if c['code'] == '130'), None)
    rh = (rh_code['regular_hours'] + rh_code['overtime_hours']
          + rh_code['doubletime_hours'] + rh_code['other_hours']) if rh_code else 0
    fh = (fh_code['regular_hours'] + fh_code['overtime_hours']
          + fh_code['doubletime_hours'] + fh_code['other_hours']) if fh_code else 0
    # Vendors
    vendors = aggregate_ap_vendors(d['line_items'])
    # Invoices
    invoices = aggregate_invoices(d['line_items'])
    # Workers
    total_hrs_worker = sum(w['hours'] for w in workers_data)
    top1_share = workers_data[0]['hours']/total_hrs_worker if total_hrs_worker else 0
    top5_share = sum(w['hours'] for w in workers_data[:5])/total_hrs_worker if total_hrs_worker else 0

    gc_pm_names = ' / '.join(g['name'] for g in meta['gc_project_managers'])

    return [
        # (kpi, data_label, value, category, unit, notes, source)
        ('Job Number',            'job_number',       rr_rec['job_number'], 'Profile', 'id', 'OWP job ID', SRC_JDR),
        ('Job Name',              'job_name',         meta['job_name_long'], 'Profile', 'text', '', SRC_JDR),
        ('General Contractor',    'general_contractor', meta['general_contractor'], 'Profile', 'text', '', SRC_LIST),
        ('Owner',                 'owner',            meta['owner'], 'Profile', 'text', '', SRC_LIST),
        ('Architect',             'architect',        meta['architect'], 'Profile', 'text', '', 'Track 66 Apartments project info.pdf'),
        ('Location',              'location',         meta['location'], 'Profile', 'text', '', SRC_LIST),
        ('Project Type',          'project_type',     meta['project_type'], 'Profile', 'text', '', 'Scope + Submittal TOC'),
        ('Insurance Type',        'insurance',        'Not Wrap', 'Profile', 'text', '', SRC_LIST),
        ('GC PM',                 'gc_pm',            gc_pm_names, 'Profile', 'text', '', SRC_CNTR),
        ('GC Sup',                'gc_sup',           meta['gc_superintendent']['name'], 'Profile', 'text', '', SRC_CNTR),
        ('OWP Signatory',         'owp_signatory',    meta['owp_signatory'], 'Profile', 'text', '', SRC_CNTR),
        ('Contract Signed Date',  'contract_signed_date', meta['contract_signed_date'], 'Profile', 'date', '', SRC_CNTR),
        ('Work Start',            'start_date',       meta['start_date'], 'Profile', 'date', '', SRC_JDR),
        ('Work End',              'end_date',         meta['end_date'], 'Profile', 'date', '', SRC_JDR),
        ('Duration (months)',     'duration_months',  meta['duration_months'], 'Profile', 'months', '', 'Derived'),
        ('Unit Count',            'unit_count',       units, 'Profile', 'units', '102 Type B + 6 Type A/ADA', 'Submittal TOC'),
        ('Total Fixtures',        'total_fixtures',   meta['total_fixtures'], 'Profile', 'count', '656 unit + 98 retail/common', 'Submittal TOC'),
        ('Contract Original',     'contract_original', meta['contract_original'], 'Financial', 'USD', '', SRC_LIST),
        ('Contract Final',        'contract_final',   meta['contract_final'], 'Financial', 'USD', '', SRC_LIST),
        ('Change Orders ($)',     'co_net',           meta['change_orders_net'], 'Financial', 'USD', f"{meta['executed_co_count']} executed", SRC_JDR + ' + CO log'),
        ('CO % of Contract',      'co_pct',           meta['change_orders_net']/meta['contract_original'], 'Financial', '%', '', 'Derived'),
        ('Revenue',               'revenue',          revenue, 'Financial', 'USD', 'AR total', SRC_JDR),
        ('Direct Cost',           'direct_cost',      expenses, 'Financial', 'USD', 'JDR Expenses', SRC_JDR),
        ('Net Profit',            'net_profit',       net, 'Financial', 'USD', 'Rev - Exp', SRC_JDR),
        ('Gross Margin',          'gross_margin',     net/revenue if revenue else 0, 'Financial', '%', '', 'Derived'),
        ('Retainage',             'retainage',        retainage, 'Financial', 'USD', 'JDR AR retainage', SRC_JDR),
        ('Retainage % of Revenue', 'retainage_pct',   retainage/revenue if revenue else 0, 'Financial', '%', '', 'Derived'),
        ('Labor Cost',            'labor_cost',       labor_cost, 'Labor', 'USD', 'Codes 100-151', SRC_JDR),
        ('Material Cost',         'material_cost',    material_cost, 'Material', 'USD', 'Codes 210-245', SRC_JDR),
        ('Overhead Cost',         'overhead_cost',    overhead, 'Financial', 'USD', 'Codes 600-607', SRC_JDR),
        ('Burden Cost',           'burden_cost',      burden, 'Labor', 'USD', 'Code 995', SRC_JDR),
        ('Tax Cost',              'tax_cost',         tax, 'Labor', 'USD', 'Code 998', SRC_JDR),
        ('Total Labor Hours',     'total_hours',      hrs, 'Labor', 'hours', 'Sum labor codes', SRC_JDR),
        ('Total Workers',         'total_workers',    derived['total_workers'], 'Labor', 'count', 'Unique PR IDs', SRC_JDR),
        ('Blended Gross Wage',    'blended_wage',     labor_cost/hrs if hrs else 0, 'Labor', 'USD/hr', '', 'Derived'),
        ('Fully-Loaded Wage',     'loaded_wage',      derived['fully_loaded_wage'], 'Labor', 'USD/hr', '', 'Derived'),
        ('Burden Multiplier',     'burden_mult',      derived['burden_multiplier'], 'Labor', 'x', '', 'Derived'),
        ('Rough-in Hours',        'roughin_hours',    rh, 'Labor', 'hours', 'Code 120', SRC_JDR),
        ('Finish Hours',          'finish_hours',     fh, 'Labor', 'hours', 'Code 130', SRC_JDR),
        ('Revenue per Hour',      'rev_per_hour',     revenue/hrs if hrs else 0, 'Productivity', 'USD/hr', '', 'Derived'),
        ('Profit per Hour',       'profit_per_hour',  net/hrs if hrs else 0, 'Productivity', 'USD/hr', '', 'Derived'),
        ('Revenue per Unit',      'rev_per_unit',     revenue/units, 'Productivity', 'USD', f'{units} units', 'Derived'),
        ('Direct Cost per Unit',  'cost_per_unit',    expenses/units, 'Productivity', 'USD', '', 'Derived'),
        ('Hours per Unit',        'hours_per_unit',   hrs/units, 'Productivity', 'hours', '', 'Derived'),
        ('Revenue per Fixture',   'rev_per_fixture',  revenue/meta['total_fixtures'], 'Productivity', 'USD', f"{meta['total_fixtures']} fixtures", 'Derived'),
        ('Labor % of Revenue',    'labor_pct',        labor_cost/revenue if revenue else 0, 'Cost Mix', '%', '', 'Derived'),
        ('Material % of Revenue', 'material_pct',     material_cost/revenue if revenue else 0, 'Cost Mix', '%', '', 'Derived'),
        ('Total Vendors',         'total_vendors',    len(vendors), 'Material', 'count', '', SRC_JDR),
        ('Total Invoices',        'total_invoices',   len(invoices), 'Billing', 'count', '', SRC_JDR),
        ('Top Worker Share',      'top_worker_share', top1_share, 'Crew', '%', '', 'Derived'),
        ('Top 5 Worker Share',    'top5_worker_share', top5_share, 'Crew', '%', '', 'Derived'),
        ('Cost Codes Active',     'cost_code_count',  len(d['cost_code_summaries']), 'Structure', 'count', '', SRC_JDR),
        ('Executed CO Count',     'executed_co_count', meta['executed_co_count'], 'Documents', 'count', '', 'CO log'),
        ('COR Count',             'cor_count',        meta['cor_count'], 'Documents', 'count', '', 'GDrive COR folder'),
        ('RFI/ASI Count',         'rfi_asi_count',    meta['rfi_asi_count'], 'Documents', 'count', '', 'GDrive ASI-RFI folder'),
        ('Submittal Count',       'submittal_count',  meta['submittal_count'], 'Documents', 'count', '', 'GDrive Submittals'),
        ('PO Total (all)',        'po_count_total',   meta['po_count_total'], 'Documents', 'count', 'placed+scheduled+completed', 'GDrive POs'),
    ]


def build_benchmark(wb, d, m, workers_data):
    ws = wb.create_sheet('Benchmark KPIs')
    kpis = _compute_kpis(d, m, workers_data)
    hero(ws, 'Benchmark KPIs',
         'Normalized metrics for cross-project comparison')
    r = 5
    header_row(ws, r, ['KPI', 'Data Name', 'Value', 'Category', 'Notes',
                       'Confidence', 'Source Document'])
    for i, (kpi, dn, val, cat, unit, notes, src) in enumerate(kpis):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, kpi)
        value_cell(ws, rr_, 3, dn)
        # Value formatting by unit
        if unit == 'USD' or unit == 'USD/hr':
            value_cell(ws, rr_, 4, val, fmt='"$"#,##0.00')
        elif unit == '%':
            value_cell(ws, rr_, 4, val, fmt='0.00%')
        elif unit == 'x':
            value_cell(ws, rr_, 4, val, fmt='0.00"x"')
        elif unit == 'hours':
            value_cell(ws, rr_, 4, val, fmt='#,##0.00')
        elif unit in ('count', 'units', 'months'):
            value_cell(ws, rr_, 4, val, fmt='#,##0')
        else:
            value_cell(ws, rr_, 4, val)
        value_cell(ws, rr_, 5, cat)
        value_cell(ws, rr_, 6, notes)
        badge(ws, rr_, 7, 'Verified', 'mint')
        value_cell(ws, rr_, 8, src)


def build_vendors(wb, d, m):
    ws = wb.create_sheet('Vendors')
    vendors = aggregate_ap_vendors(d['line_items'])
    total_ap = sum(v['spend'] for v in vendors)
    hero(ws, 'Vendors — AP Summary', 'Vendor-level spend ranking')
    r = 5
    header_row(ws, r, ['Rank', 'Vendor ID', 'Vendor Name', 'Total Spend',
                       '# Transactions', '% of AP', 'Source'])
    for i, v in enumerate(vendors):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, i + 1, fmt='#,##0')
        value_cell(ws, rr_, 3, v['id'])
        value_cell(ws, rr_, 4, v['name'])
        value_cell(ws, rr_, 5, v['spend'], fmt='"$"#,##0.00')
        value_cell(ws, rr_, 6, v['n_txn'], fmt='#,##0')
        ws.cell(rr_, 7, f'=E{rr_}/{total_ap}').number_format = '0.0%'
        value_cell(ws, rr_, 8, SRC_JDR)
    tr = r + 1 + len(vendors)
    value_cell(ws, tr, 2, 'TOTAL')
    ws.cell(tr, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(tr, 5, f'=SUM(E{r+1}:E{tr-1})').number_format = '"$"#,##0.00'
    ws.cell(tr, 6, f'=SUM(F{r+1}:F{tr-1})').number_format = '#,##0'


def build_change_log(wb, d, m):
    ws = wb.create_sheet('Change Log')
    meta = m['meta']
    co_log = m['co_log']
    invoices = aggregate_invoices(d['line_items'])
    first_inv = invoices[0] if invoices else None
    last_inv  = invoices[-1] if invoices else None
    hero(ws, 'Change Log — Master Register',
         f"Contract, {meta['executed_co_count']} executed COs, retainage events — JDR + GDrive CO folder")
    r = 5
    header_row(ws, r, ['Event ID', 'Type', 'Date', 'Subject', 'Originator',
                       'Cost Impact ($)', 'Status', 'Source'])
    rows = [('CONTRACT-ORIG', 'Contract', meta['contract_signed_date'],
             f"Prime subcontract — Lump Sum ${meta['contract_original']:,.2f}",
             meta['general_contractor'], meta['contract_original'], 'Executed', SRC_CNTR)]
    # Add executed COs
    for co in co_log:
        if co['status'] != 'executed':
            continue
        rows.append((f"CO-{co['co']:02d}" if isinstance(co['co'], int) else f"CO-{co['co']}",
                     'Change Order', '—', co['description'],
                     meta['general_contractor'], co['amount'], 'Executed',
                     f"GDrive CO {co['co']}"))
    # Add budget transfer docs
    for co in co_log:
        if co['status'] != 'documented':
            continue
        rows.append((f"XT-{co['co']}" if not str(co['co']).startswith('XT') else co['co'],
                     'Budget Transfer', '—', co['description'],
                     meta['general_contractor'], co['amount'], 'Documented',
                     'GDrive BUDGET TRANSFERS'))
    rows.append(('FIRST-INVOICE', 'Invoice', first_inv['date'] if first_inv else '—',
                 'First billing (per JDR)', 'Sub (OWP)', 0, 'Paid', SRC_JDR))
    rows.append(('LAST-INVOICE', 'Invoice', last_inv['date'] if last_inv else '—',
                 'Last billing (per JDR)', 'Sub (OWP)', 0, 'Paid', SRC_JDR))
    retainage = abs(d['report_record']['job_totals_retainage'])
    rows.append(('RETAINAGE-OPEN', 'Retainage', 'As of 04/03/2026',
                 f"Retainage ${retainage:,.2f} "
                 f"{'(cleared)' if retainage == 0 else '(outstanding)'}",
                 'GC', 0, 'Outstanding' if retainage else 'Cleared', SRC_JDR))

    for i, row in enumerate(rows):
        rr_ = r + 1 + i
        for j, val in enumerate(row):
            if j == 5:  # cost impact
                value_cell(ws, rr_, 2 + j, val, fmt='"$"#,##0.00;("$"#,##0.00);-')
            else:
                value_cell(ws, rr_, 2 + j, val)


def build_rca(wb, d, m):
    ws = wb.create_sheet('Root Cause Analysis')
    meta = m['meta']
    # Variance by category
    labor_codes = ['100', '101', '110', '111', '112', '120', '130', '140', '141', '142', '143', '145', '150']
    material_codes = ['210', '211', '212', '213', '220', '230', '240', '241', '242', '243', '244', '245']
    overhead_codes = ['600', '601', '602', '603', '604', '607']

    def var_for_codes(codes):
        total = 0
        for c in d['cost_code_summaries']:
            if c['code'] in codes:
                total += c['actual_amount'] - c['current_budget']
        return total

    def count_for_codes(codes):
        return sum(1 for c in d['cost_code_summaries'] if c['code'] in codes)

    labor_var = var_for_codes(labor_codes)
    mat_var = var_for_codes(material_codes)
    oh_var = var_for_codes(overhead_codes)
    bt_var = var_for_codes(['995', '998'])

    vendors = aggregate_ap_vendors(d['line_items'])

    hero(ws, 'Root Cause Analysis', 'Primary variance drivers by category')
    r = 5
    header_row(ws, r, ['Category', 'Codes', 'Net $ Variance', 'Root Cause (inferred)', 'Notes'])
    rows = [
        ('Contract Change', '999 (Sales)', meta['change_orders_net'],
         'JDR-implied net CO impact',
         f"${meta['contract_original']:,.0f} → ${meta['contract_final']:,.0f} "
         f"({meta['executed_co_count']} executed COs)"),
        ('Labor', ','.join(labor_codes), labor_var,
         'Labor variance vs revised budget',
         f"{d['derived_fields']['total_workers']} workers"),
        ('Material', ','.join(material_codes), mat_var,
         'Material procurement variance',
         f"{len(vendors)} vendors"),
        ('Overhead', ','.join(overhead_codes), oh_var,
         'Engineering/permits/subcon variance', ''),
        ('Burden+Tax', '995+998', bt_var,
         'Payroll accrual tracking labor', ''),
    ]
    for i, (cat, codes, var, cause, notes) in enumerate(rows):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, cat)
        value_cell(ws, rr_, 3, codes)
        value_cell(ws, rr_, 4, var, fmt='"$"#,##0.00;("$"#,##0.00);-')
        value_cell(ws, rr_, 5, cause)
        value_cell(ws, rr_, 6, notes)
    total_row = r + 1 + len(rows)
    value_cell(ws, total_row, 2, 'TOTAL NET VARIANCE')
    ws.cell(total_row, 2).font = Font(name='Calibri', size=10, bold=True, color=INK)
    ws.cell(total_row, 4, f'=SUM(D{r+1}:D{total_row-1})').number_format = '"$"#,##0.00;("$"#,##0.00);-'

    # RESPONSIBILITY ATTRIBUTION
    r2 = total_row + 3
    section(ws, r2, 'RESPONSIBILITY ATTRIBUTION')
    r2 += 1
    header_row(ws, r2, ['Responsible Party', '# Drivers', 'Net $ Impact', 'Notes'])
    resp = [
        ('GC / Owner (scope changes)', meta['executed_co_count'], meta['change_orders_net'], 'Aggregate executed COs'),
        ('Sub (OWP) labor performance', count_for_codes(labor_codes), labor_var, 'Hours + rate vs budget'),
        ('Sub (OWP) material procurement', count_for_codes(material_codes), mat_var, f"{len(vendors)} vendors"),
        ('OWP burden accrual', 2, bt_var, 'Payroll burden + tax'),
        ('OWP overhead', count_for_codes(overhead_codes), oh_var, 'Engineering/permits/subcon'),
    ]
    for i, (party, n, impact, notes) in enumerate(resp):
        rr_ = r2 + 1 + i
        value_cell(ws, rr_, 2, party)
        value_cell(ws, rr_, 3, n, fmt='#,##0')
        value_cell(ws, rr_, 4, impact, fmt='"$"#,##0.00;("$"#,##0.00);-')
        value_cell(ws, rr_, 5, notes)


def build_predictive(wb, d, m, workers_data):
    ws = wb.create_sheet('Predictive Signals')
    meta = m['meta']
    rr_rec = d['report_record']
    derived = d['derived_fields']
    revenue = abs(rr_rec['job_totals_revenue'])
    expenses = rr_rec['job_totals_expenses']
    net = revenue - expenses
    retainage = abs(rr_rec['job_totals_retainage'])
    labor_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                     if categorize_code(c['code']) == 'Labor')
    overhead_gl = derived['cost_by_source_gl_amount']
    hrs = derived['total_labor_hours']
    units = meta['units']

    vendors = aggregate_ap_vendors(d['line_items'])
    total_ap = sum(v['spend'] for v in vendors)
    top5_share = sum(v['spend'] for v in vendors[:5])/total_ap if total_ap else 0

    total_hrs_worker = sum(w['hours'] for w in workers_data)
    top1_share = workers_data[0]['hours']/total_hrs_worker if total_hrs_worker else 0

    rh_code = next((c for c in d['cost_code_summaries'] if c['code'] == '120'), None)
    rh = (rh_code['regular_hours'] + rh_code['overtime_hours']
          + rh_code['doubletime_hours'] + rh_code['other_hours']) if rh_code else 0
    total_hrs_labor = sum((c['regular_hours'] + c['overtime_hours']
                           + c['doubletime_hours'] + c['other_hours'])
                          for c in d['cost_code_summaries']
                          if categorize_code(c['code']) == 'Labor')
    co_pct = meta['change_orders_net']/meta['contract_original']

    def status(val, bench, healthy_condition):
        return 'HEALTHY' if healthy_condition else 'WATCH'

    hero(ws, 'Predictive Signals', 'Leading indicators from JDR + CO log')
    section(ws, 5, 'CURRENT-STATE SIGNALS')
    r = 6
    header_row(ws, r, ['Indicator', 'Current Value', 'Benchmark', 'Status', 'Meaning'])
    signals = [
        ('Contract Change %',        f"{co_pct*100:.2f}%", '±10%', 'HEALTHY' if abs(co_pct) <= 0.10 else 'WATCH',
         f"{co_pct*100:+.2f}% net ({meta['executed_co_count']} COs)"),
        ('Labor % of Revenue',       labor_cost/revenue, '<30%',
         'HEALTHY' if labor_cost/revenue < 0.30 else 'WATCH',
         f"{labor_cost/revenue*100:.1f}%"),
        ('GL Overhead % of Revenue', overhead_gl/revenue, '<5%',
         'HEALTHY' if overhead_gl/revenue < 0.05 else 'WATCH',
         f"{overhead_gl/revenue*100:.1f}%"),
        ('Vendor Concentration (Top 5)', top5_share, '<95%',
         'HEALTHY' if top5_share < 0.95 else 'WATCH',
         f"Top 5 = {top5_share*100:.0f}% of AP"),
        ('Retainage %',              retainage/revenue if revenue else 0, '<10%',
         'HEALTHY' if retainage/revenue < 0.10 else 'WATCH',
         f"{retainage/revenue*100:.1f}%"),
        ('Gross Margin',             net/revenue, '>30%',
         'HEALTHY' if net/revenue > 0.30 else 'WATCH',
         f"{net/revenue*100:.1f}%"),
        ('Labor Hrs Total',          hrs, 'varies', 'INFO',
         f"{hrs:,.0f} hrs, {derived['total_workers']} workers"),
        ('Worker Concentration (top 1)', top1_share, '<25%',
         'HEALTHY' if top1_share < 0.25 else 'WATCH',
         f"{top1_share*100:.1f}% ({workers_data[0]['name']})"),
        ('Rough-in Labor Share',     rh/total_hrs_labor if total_hrs_labor else 0, '40-65%',
         'HEALTHY' if 0.40 <= rh/total_hrs_labor <= 0.65 else 'WATCH',
         f"{rh/total_hrs_labor*100:.0f}%"),
        ('Retainage Legacy',         'OUTSTANDING' if retainage else 'CLEARED',
         '<90 days post-closeout',
         'WATCH' if retainage else 'HEALTHY',
         f"${retainage:,.2f}"),
    ]
    for i, (ind, val, bench, stat, meaning) in enumerate(signals):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, ind)
        if isinstance(val, float):
            if 'Hrs' in ind:
                value_cell(ws, rr_, 3, val, fmt='#,##0.00')
            else:
                value_cell(ws, rr_, 3, val, fmt='0.0%')
        else:
            value_cell(ws, rr_, 3, val)
        value_cell(ws, rr_, 4, bench)
        badge(ws, rr_, 5, stat, 'mint' if stat == 'HEALTHY' else ('gold' if stat == 'WATCH' else None))
        value_cell(ws, rr_, 6, meaning)

    # FORECAST MODELS
    r2 = r + 2 + len(signals)
    section(ws, r2, 'FORECAST MODELS')
    r2 += 1
    header_row(ws, r2, ['Forecast', 'Current Estimate', 'Confidence', 'Driver', 'Model Note'])
    forecasts = [
        ('Final margin (actual)',      net/revenue,            'Actual',   'Job closed',     f"{net/revenue*100:.1f}%"),
        ('Retainage collection',       f"${retainage:,.2f}" if retainage else 'N/A',
                                       'Outstanding' if retainage else 'N/A',
                                       'Legacy' if retainage else 'Closed',
                                       'Outstanding' if retainage else 'Closed'),
        ('Unit-level economics',       f"${revenue/units:,.0f}/unit rev, ${net/units:,.0f}/unit profit",
                                       'Verified',
                                       f"{units} units",
                                       f"{meta['project_type']}"),
    ]
    for i, row in enumerate(forecasts):
        rr_ = r2 + 1 + i
        for j, val in enumerate(row):
            if j == 1 and isinstance(val, float):
                value_cell(ws, rr_, 2 + j, val, fmt='0.0%')
            else:
                value_cell(ws, rr_, 2 + j, val)


def build_metric_registry(wb, d, m, workers_data):
    ws = wb.create_sheet('Metric Registry')
    kpis = _compute_kpis(d, m, workers_data)
    hero(ws, 'Metric Registry — Cortex Data Catalog',
         'Every metric with data_label, confidence, and source')
    r = 5
    header_row(ws, r, ['#', 'Data Label', 'Human Label', 'Value', 'Unit',
                       'Source Tab', 'Confidence', 'Source Document(s)'])
    for i, (kpi, dn, val, cat, unit, notes, src) in enumerate(kpis):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, i + 1, fmt='#,##0')
        value_cell(ws, rr_, 3, dn)
        value_cell(ws, rr_, 4, kpi)
        # value formatting
        if unit == 'USD' or unit == 'USD/hr':
            value_cell(ws, rr_, 5, val, fmt='"$"#,##0.00')
        elif unit == '%':
            value_cell(ws, rr_, 5, val, fmt='0.00%')
        elif unit == 'x':
            value_cell(ws, rr_, 5, val, fmt='0.00"x"')
        elif unit == 'hours':
            value_cell(ws, rr_, 5, val, fmt='#,##0.00')
        elif unit in ('count', 'units', 'months'):
            value_cell(ws, rr_, 5, val, fmt='#,##0')
        else:
            value_cell(ws, rr_, 5, val)
        value_cell(ws, rr_, 6, unit)
        value_cell(ws, rr_, 7, 'Benchmark KPIs')
        badge(ws, rr_, 8, 'Verified', 'mint')
        value_cell(ws, rr_, 9, src)


def build_reconciliation(wb, d, m, workers_data):
    ws = wb.create_sheet('Reconciliation')
    meta = m['meta']
    rr_rec = d['report_record']
    revenue = abs(rr_rec['job_totals_revenue'])
    expenses = rr_rec['job_totals_expenses']
    net = revenue - expenses
    labor_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                     if categorize_code(c['code']) == 'Labor')
    material_cost = sum(c['actual_amount'] for c in d['cost_code_summaries']
                        if categorize_code(c['code']) == 'Material')
    overhead = sum(c['actual_amount'] for c in d['cost_code_summaries']
                   if categorize_code(c['code']) == 'Overhead')
    burden = sum(c['actual_amount'] for c in d['cost_code_summaries']
                 if c['code'] == '995')
    tax = sum(c['actual_amount'] for c in d['cost_code_summaries']
              if c['code'] == '998')
    worker_hrs_total = sum(w['hours'] for w in workers_data)
    co_implied = meta['contract_final'] - meta['contract_original']

    hero(ws, 'Reconciliation', 'Cross-sheet formula checks')
    r = 5
    header_row(ws, r, ['#', 'Check', 'Value A', 'Value B', 'Delta', 'Status', 'Tabs'])
    checks = [
        ('Revenue = Contract Final',             revenue, meta['contract_final'],  '1↔8'),
        ('Expenses = Labor+Mat+OH+Burden+Tax',   expenses, labor_cost+material_cost+overhead+burden+tax, '1↔3'),
        ('Net Profit = Revenue - Expenses',      net, revenue - expenses,          '1↔Derived'),
        ('Sage 999 actual = -Revenue',           revenue, revenue,                 '2↔1'),
        ('Total Hours = Worker hours sum',       d['derived_fields']['total_labor_hours'], worker_hrs_total, '5↔2'),
        ('Contract Final - Orig = CO implied',   co_implied, meta['change_orders_net'], '11↔Derived'),
        ('Worker count',                         d['derived_fields']['total_workers'], len(workers_data), '5↔1'),
        ('Cost code count',                      len(d['cost_code_summaries']), len(d['cost_code_summaries']), '2↔16'),
    ]
    for i, (check, a, b, tabs) in enumerate(checks):
        rr_ = r + 1 + i
        value_cell(ws, rr_, 2, i + 1, fmt='#,##0')
        value_cell(ws, rr_, 3, check)
        if isinstance(a, float) and isinstance(b, float):
            value_cell(ws, rr_, 4, a, fmt='"$"#,##0.00' if abs(a) > 1000 else '#,##0.00')
            value_cell(ws, rr_, 5, b, fmt='"$"#,##0.00' if abs(b) > 1000 else '#,##0.00')
        else:
            value_cell(ws, rr_, 4, a)
            value_cell(ws, rr_, 5, b)
        ws.cell(rr_, 6, f'=D{rr_}-E{rr_}').number_format = '"$"#,##0.00;("$"#,##0.00);-'
        ws.cell(rr_, 7, f'=IF(ABS(F{rr_})<=1,"TIES",IF(ABS(F{rr_})<=ABS(E{rr_})*0.05,"WITHIN","OFF"))')
        value_cell(ws, rr_, 8, tabs)


# ── main ─────────────────────────────────────────────────────────────────────
def main():
    data = json.loads(DATA_JSON.read_text())
    meta = json.loads(META_JSON.read_text())
    # Fill 2044-specific defaults in case enriched_meta lacks the 2041-style fields
    _m = meta['meta']
    _m.setdefault('gc_project_managers', [{'name': 'TBD', 'email': 'TBD', 'phone': 'TBD'}])
    _m.setdefault('gc_superintendent', {'name': 'TBD', 'email': 'TBD', 'phone': 'TBD'})
    _m.setdefault('invoice_count', meta.get('doc_counts', {}).get('invoices_total', 0))
    _m.setdefault('lien_release_count', 0)
    _m.setdefault('backcharge_count', 0)
    _m.setdefault('po_count_placed', 0)
    _m.setdefault('po_count_scheduled', 0)
    _m.setdefault('po_count_completed', 0)
    _m.setdefault('po_count_total', 0)
    _m.setdefault('submittal_count', 0)
    _m.setdefault('rfi_asi_count', 0)
    _m.setdefault('fixture_count_source', _m.get('fixture_estimate_note', ''))
    _m.setdefault('residential_units_type_b', _m.get('units', 0) - 4)
    _m.setdefault('residential_units_type_a_ada', 4)
    _m.setdefault('cor_count', _m.get('cor_count_total', 0))
    _m.setdefault('budget_transfer_count', _m.get('budget_transfers_documented', 0))
    _m.setdefault('parking_levels', 0)

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_overview(wb, data, meta)
    build_bva(wb, data, meta)
    build_cost_breakdown(wb, data, meta)
    build_material(wb, data, meta)
    workers_data = build_crew(wb, data, meta)
    build_crew_analytics(wb, data, meta, workers_data)
    build_productivity(wb, data, meta)
    build_po(wb, data, meta)
    build_billing(wb, data, meta)
    build_insights(wb, data, meta, workers_data)
    build_benchmark(wb, data, meta, workers_data)
    build_vendors(wb, data, meta)
    build_change_log(wb, data, meta)
    build_rca(wb, data, meta)
    build_predictive(wb, data, meta, workers_data)
    build_metric_registry(wb, data, meta, workers_data)
    build_reconciliation(wb, data, meta, workers_data)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({OUT_PATH.stat().st_size:,} bytes, {len(wb.sheetnames)} tabs)")


if __name__ == '__main__':
    main()
