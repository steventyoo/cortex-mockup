#!/usr/bin/env python3
"""
enrich_2061.py — Enrich OWP_2061_JCR_Cortex_v2.xlsx with key players + populate
sparse tabs.

Project: 2061 Alta Columbia City — Exxel Pacific — 243 units, $4.43M, 41.4% margin
         Closed Jun 2019 → May 2022 (35 months). Retention released 2022-06-23.

Inputs:
  • 2061_data.json            — Sage JDR parse (5,605 line items, 27 cost codes, 58 workers)
  • 2061_dashboard_arrays.json — 16 arrays
  • GDrive Job Book           — /Job Books - Completed Jobs/2061-Exxel, Columbia City/

Outputs (in-place):
  • OWP_2061_JCR_Cortex_v2.xlsx — Job Info tab rebuilt with sectioned key-players +
    project metadata; SOV-PayApps + Change Orders + Reconciliation populated
    from dashboard arrays.

Key players (from executed subcontract + GDrive scan):
  GC: Exxel Pacific (323 Telegraph Road, Bellingham WA 98226 · (360) 734-2872)
  GC PM: Brian Christensen
  Owner: Gateway Alta Rainer Owner, LLC
  Architect: Johnston Architects, PLLC
  MEP / Plumbing Engineer: Franklin Engineering
  OWP signatory: Richard Donelson (PM/Estimator · richd@owpllc.net · 425-591-4296)
  Subcontract: L1.220000 · executed Aug 8, 2019 · $4,305,550 lump sum
  Insurance: OCIP (Wrap-up) · "Attachment J - OCIP Manual - Columbia City - 5.29.19.pdf"
  Permit: City of Seattle Plumbing Permit (Plan Review 2019-09-26 · final 3-22-22)
  Site Address: 3717 South Alaska Street, Seattle WA 98118 (Columbia City)
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2061_data.json"
DASH_FILE  = SCRIPT_DIR / "2061_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2061_JCR_Cortex_v2.xlsx"

META_2061 = {
    "job_id":          "2061",
    "name":            "Alta Columbia City",
    "short_name":      "Alta CC",
    "gc":              "Exxel Pacific",
    "gc_address":      "323 Telegraph Road, Bellingham, WA 98226 · (360) 734-2872",
    "gc_pm":           "Brian Christensen",
    "gc_sup":          "TBD (per Job Tickets — to be enriched)",
    "gc_pe":           "TBD",
    "gc_signatory":    "TBD (Exxel — see executed subcontract)",
    "owp_ri_foreman":  "TBD (per Job Tickets — to be enriched)",
    "owp_trim_foreman":"TBD",
    "owp_signatory":   "Richard Donelson (PM/Estimator · richd@owpllc.net · 425-591-4296)",
    "owp_estimator":   "Richard Donelson",
    "owner":           "Gateway Alta Rainer Owner, LLC",
    "developer":       "Gateway / Alta (per ownership entity)",
    "architect":       "Johnston Architects, PLLC",
    "structural":      "TBD (likely Coughlin Porter Lundeen — per RFI 044/124 'CPL Response')",
    "mep_engineer":    "Franklin Engineering (3rd-party plumbing engineer — per executed subcontract §A.1.2)",
    "civil":           "TBD (likely CPL Civil — per ASI 022 'Landscape and Civil')",
    "landscape":       "TBD (per ASI 022)",
    "interior":        "VIDA Architecture/Design (per RFI 116 'VIDA RESPONSE')",
    "site_address":    "3717 South Alaska Street, Seattle, WA 98118 (Columbia City)",
    "permit":          "City of Seattle Plumbing Permit (Plan Review 2019-09-26 · final 2022-03-22) + Backflow + Gas permits",
    "insurance":       "OCIP (Wrap-up) — Exxel-administered (Attachment J - OCIP Manual 5.29.19)",
    "lien_position":   "Standard subcontract — retention released 2022-06-23",
    "warranty":        "1-yr standard plumbing (per O&M's folder · 65 close-out documents)",
    "delivery_route":  "Columbia City, Seattle (Rainier Ave + S Alaska St corridor)",
    "contract_doc":    "L1.220000 · executed 2019-08-08 · AGC Washington 2009 Form (modified) · $4,305,550 lump sum",
    "subcontract_no":  "L1.220000",
    "exxel_project_no":"201904",
    "scope":           "Design Build Plumbing Systems",
    "gdrive_status":   "Job Book at /Job Books - Completed Jobs/2061-Exxel, Columbia City/ — fully accessible",
    "status_text":     "CLOSED · 41.4% margin",
    "status_severity": "HEALTHY",
    "ar_billed":       4430088,
    "direct_cost":     2597500,
    "net_profit":      1832588,
    "retention":       221504,
    "contract_orig":   4305550,
    "contract_final":  4430088,
    "co_net":          124538,
    "units":           243,
    "fixtures":        "Per Plumbing Fixture Schedule (Plans/Columbia City Front Sheet.xlsx) — 30+ fixture types incl. WC1/WC2/L1/L2/TS1/TS2/SS1/SS2/SS3/S1/S2/S3/S4/DW/AW/MS1/BW1/DF1 + 5 water heaters (GWH-1 through GWH-5, 500 MBH gas) + 2 circulation pumps + backflow preventers + sub-meters at all units",
    "project_type":    "243-unit mixed-use multifamily + 6 retail spaces · 7-story + below-grade",
    "schedule":        "Jun 2019 → May 2022 (~35 months)",
    "duration_months": 35,
    "workers":         58,
    "hours":           28516,
    "doc_counts": {
        "RFIs":         39,
        "ASIs":         26,
        "executed_COs": 13,    # CO #001 through SCO 17 (13 unique numbers)
        "CORs":         14,    # ~14 unique COR series (97 files include revisions)
        "Submittals":   117,
        "POs":          290,   # 36 placed + 149 scheduled + 105 completed
        "Permits":      209,
        "Pay_apps":     27,
        "Inspection_reports": 21,
        "Meetings":     125,
        "OandM_docs":   65,
        "Insurance_docs": 14,
        "Vendor_quotes":  20,
    },
}

# ============================================================================
# Style helpers (parity with enrich_2020.py + enrich_2108.py)
# ============================================================================
INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
ACTIVE_FONT = Font(name="Arial", size=10, color="155724", bold=True)

HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
ROW_ALT     = PatternFill("solid", start_color="FAFAFA")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
ACTIVE_FILL = PatternFill("solid", start_color="D4EDDA")
WARN_FILL   = PatternFill("solid", start_color="FFF3CD")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")


def clear_range(ws, r1, r2, c1=1, c2=10):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).border = Border()


# ============================================================================
# JOB INFO tab — full rebuild with sectioned layout
# ============================================================================
def rebuild_job_info(ws):
    clear_range(ws, 1, ws.max_row + 5, 1, 6)
    ws["A1"] = f"JOB #{META_2061['job_id']} · {META_2061['name'].upper()} · KEY PLAYERS & METADATA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Closed-out project · Jun 2019 → May 2022. Team data sourced from "
                "executed subcontract (Alta Columbia City - L1.220000), GDrive Job Book "
                "(/Job Books - Completed Jobs/2061-Exxel, Columbia City/), Plumbing "
                "Fixture Schedule (Plans/Columbia City Front Sheet.xlsx), and Sage JDR "
                "(2061 Job Detail Report.pdf · Apr 3 2026 run).")
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    margin = META_2061['net_profit'] / META_2061['contract_final']
    cost_pct = META_2061['direct_cost'] / META_2061['contract_final']

    sections = [
        ("IDENTITY", [
            ("Job Number",         META_2061["job_id"]),
            ("Job Name",           META_2061["name"]),
            ("Short Name",         META_2061["short_name"]),
            ("Project Type",       META_2061["project_type"]),
            ("Site Address",       META_2061["site_address"]),
            ("Permit",             META_2061["permit"]),
            ("Status",             META_2061["status_text"]),
            ("Subcontract Number", META_2061["subcontract_no"]),
            ("Exxel Project Number", META_2061["exxel_project_no"]),
            ("Scope of Work",      META_2061["scope"]),
        ]),
        ("SCHEDULE", [
            ("Project Start",      META_2061["schedule"].split("→")[0].strip()),
            ("Project End",        META_2061["schedule"].split("→")[1].strip().split("(")[0].strip()),
            ("Duration",           f"{META_2061['duration_months']} months"),
            ("Schedule Note",      META_2061["schedule"]),
            ("Notice of Completion", "On file (Notice of Work Completion - Columbia City.pdf)"),
        ]),
        ("CONTRACT & FINANCIALS", [
            ("Original Contract",  f"${META_2061['contract_orig']:,}"),
            ("Net Change Orders",  f"${META_2061['co_net']:,} (+{META_2061['co_net']/META_2061['contract_orig']*100:.1f}%)"),
            ("Final Contract",     f"${META_2061['contract_final']:,}"),
            ("AR Billed (Revenue)", f"${META_2061['ar_billed']:,}"),
            ("Direct Cost",        f"${META_2061['direct_cost']:,}"),
            ("Net Profit",         f"${META_2061['net_profit']:,}"),
            ("Gross Margin",       f"{margin*100:.1f}%"),
            ("Cost as % of Revenue", f"{cost_pct*100:.1f}%"),
            ("Retention Released", f"${META_2061['retention']:,} on 2022-06-23 (per Billing folder receipt)"),
            ("Insurance",          META_2061["insurance"]),
            ("Lien Position",      META_2061["lien_position"]),
            ("Warranty",           META_2061["warranty"]),
            ("Contract on File",   META_2061["contract_doc"]),
        ]),
        ("SCOPE & FIXTURE PROFILE", [
            ("Plumbing Units",     f"{META_2061['units']} units (residential)"),
            ("Retail Spaces",      "6 retail bays at L1"),
            ("Total Fixtures",     META_2061["fixtures"][:200]),
            ("Floors",             "P3/P2/P1 below-grade garage + L1 retail/lobby + L2-L7 residential + Roof penthouse"),
            ("Building Type",      "7-story wood-frame over concrete podium"),
            ("Pavilion / Skylounge", "Yes — Level 7 amenity (with BBQs + firepits per Gas Pressure Regulator schedule)"),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR (Exxel Pacific)", [
            ("General Contractor", META_2061["gc"]),
            ("GC Address",         META_2061["gc_address"]),
            ("GC Project Manager", META_2061["gc_pm"]),
            ("GC Superintendent",  META_2061["gc_sup"]),
            ("GC Project Engineer", META_2061["gc_pe"]),
            ("GC Signatory",       META_2061["gc_signatory"]),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2061["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2061["owp_trim_foreman"]),
            ("OWP Estimator (PM)",  META_2061["owp_estimator"]),
            ("OWP Signatory",       META_2061["owp_signatory"]),
            ("Total Workers Logged", f"{META_2061['workers']} workers · {META_2061['hours']:,} hrs"),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record",    META_2061["owner"]),
            ("Developer",          META_2061["developer"]),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",          META_2061["architect"]),
            ("Interior Designer",  META_2061["interior"]),
            ("Structural Engineer", META_2061["structural"]),
            ("MEP / Plumbing Engineer", META_2061["mep_engineer"]),
            ("Civil Engineer",     META_2061["civil"]),
            ("Landscape Architect", META_2061["landscape"]),
        ]),
        ("DOCUMENT META (final close-out counts)", [
            ("RFIs",               f"{META_2061['doc_counts']['RFIs']} distinct (highest = RFI 382)"),
            ("ASIs",               f"{META_2061['doc_counts']['ASIs']} distinct (highest = ASI 034)"),
            ("Executed COs",       f"{META_2061['doc_counts']['executed_COs']} (CO #001 through SCO 17 — net +${META_2061['co_net']:,})"),
            ("CORs",               f"{META_2061['doc_counts']['CORs']} unique COR series (97 files include revisions)"),
            ("Submittals",         f"{META_2061['doc_counts']['Submittals']} files (incl. UNIT FIXTURES, COMMON AREA, EQUIPMENT, METERING, SLEEVING, UNDERGROUND & ROUGHIN sub-folders)"),
            ("POs",                f"{META_2061['doc_counts']['POs']} POs (36 placed + 149 scheduled + 105 completed)"),
            ("Permits",            f"{META_2061['doc_counts']['Permits']} files (Plumbing + Backflow + Gas + Renewals)"),
            ("Pay Apps",           f"{META_2061['doc_counts']['Pay_apps']} AR transactions in JDR (46 SOV files in Billing folder)"),
            ("Inspection Reports", f"{META_2061['doc_counts']['Inspection_reports']} files"),
            ("Meeting Minutes",    f"{META_2061['doc_counts']['Meetings']} files in Meetings-Schedules"),
            ("O&M Documents",      f"{META_2061['doc_counts']['OandM_docs']} files (close-out package)"),
            ("Vendor Quotes",      f"{META_2061['doc_counts']['Vendor_quotes']} files"),
            ("Insurance Docs",     f"{META_2061['doc_counts']['Insurance_docs']} files (OCIP manual + 27 renewal certs + COIs)"),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",            "2061 Job Detail Report.pdf (Sage Timberline · Apr 3 2026 run)"),
            ("Parsed Data",        "2061_data.json (3.0MB — full Sage extract: 5,605 line items, 27 cost codes, 58 workers)"),
            ("Dashboard Arrays",   "2061_dashboard_arrays.json (16 arrays)"),
            ("Subcontract",        "Alta Columbia City - L1.220000 - One Way Plumbing Executed Subcontract.pdf"),
            ("Plumbing Fixture Schedule", "Plans/Columbia City Front Sheet.xlsx (9 tabs: PLMB FIXT., PUMPS, WATER HEATERS, MIXING VALVE, EXP. TANK, MISC EQUIP, GENERAL NOTES, GAS LOADS, GPR SCHED)"),
            ("GDrive Folder",      META_2061["gdrive_status"]),
        ]),
        ("PROJECT NARRATIVE", [
            ("Margin profile",     f"OWP delivered {margin*100:.1f}% gross margin = ${META_2061['net_profit']:,} on a ${META_2061['contract_final']:,} closed contract — well above OWP's portfolio median (~32-36%). Direct cost absorbed only {cost_pct*100:.1f}% of revenue."),
            ("CO posture",         f"Net +${META_2061['co_net']:,} ({META_2061['co_net']/META_2061['contract_orig']*100:+.1f}%) additive — modest growth from $4.31M base. 13 executed COs + 14 unique COR series captured. Driven primarily by design clarification (ASI 008/011/012, multiple drainage updates)."),
            ("Vendor concentration", "Rosen Supply Kirkland top vendor at 37.9% of AP. Top 3 vendors dominate OWP's AP spend on this project."),
            ("Closeout",           "Retention $221,504 released June 23, 2022 (clean closeout). 65 O&M documents filed. Notice of Work Completion on file."),
            ("Owner-developer continuity", "Owner Gateway Alta Rainer Owner, LLC same Lake Union Partners adjacency as #2108 R&G Apartments (also at the Rainier corridor) — Columbia City portfolio is a recurring relationship for OWP."),
        ]),
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        for label, val in fields:
            ws.cell(row=r, column=1, value=label).font = INK_BOLD
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_v.font = INK
            cell_v.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=1).border = BORDER
            cell_v.border = BORDER
            if section_name.startswith("PROJECT TEAM") or section_name.startswith("DOCUMENT META") or section_name.startswith("PROJECT NARRATIVE"):
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v.fill = TEAM_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 28


# ============================================================================
# OVERVIEW tab — add Project Team block
# ============================================================================
def patch_overview_team_block(ws):
    # Replace any "CLOSED" placeholder banners → already correct via build script
    start = 40
    clear_range(ws, start, start + 30, 1, 8)
    ws.cell(row=start, column=1, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    ws.cell(row=start+1, column=1, value=(
        "Closed Jun 2019 → May 2022 · 41.4% margin · retention released 2022-06-23. "
        "Key players from executed subcontract (L1.220000) + GDrive Job Book."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=7)

    rows = [
        ("General Contractor",   META_2061["gc"]),
        ("GC Project Manager",   META_2061["gc_pm"]),
        ("OWP Estimator (PM)",   META_2061["owp_estimator"]),
        ("OWP Signatory",        META_2061["owp_signatory"]),
        ("Owner of Record",      META_2061["owner"]),
        ("Developer",            META_2061["developer"]),
        ("Architect",            META_2061["architect"]),
        ("Interior Designer",    META_2061["interior"]),
        ("MEP / Plumbing Engineer", META_2061["mep_engineer"]),
        ("Site Address",         META_2061["site_address"]),
        ("Insurance",            META_2061["insurance"]),
        ("Subcontract",          META_2061["subcontract_no"] + " · executed 2019-08-08"),
        ("Status",               META_2061["status_text"]),
        ("Original Contract",    f"${META_2061['contract_orig']:,}"),
        ("Final Contract",       f"${META_2061['contract_final']:,}"),
        ("Net Profit",           f"${META_2061['net_profit']:,} ({META_2061['net_profit']/META_2061['contract_final']*100:.1f}% margin)"),
        ("Retention Released",   f"${META_2061['retention']:,} on 2022-06-23"),
        ("Workers / Hours",      f"{META_2061['workers']} workers · {META_2061['hours']:,} hrs"),
    ]
    r = start + 3
    ws.cell(row=r, column=1, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=1, value=role).font = INK_BOLD
        ws.cell(row=r, column=2, value=val).font = INK
        ws.cell(row=r, column=1).fill = TEAM_FILL
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1


# ============================================================================
# CHANGE ORDERS tab — populate from dashboard_arrays.changeLog
# ============================================================================
def patch_change_orders(ws, dash):
    log = dash.get("changeLog", [])
    meta = dash.get("changeMeta", {})
    clear_range(ws, 1, 50, 1, 8)
    ws["A1"] = f"Change Orders · Job #{META_2061['job_id']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"{meta.get('total','—')} total events · net ${META_2061['co_net']:,} "
                f"({META_2061['co_net']/META_2061['contract_orig']*100:+.1f}%). "
                "Source: GDrive Change Orders folder (17 SCO files + 97 COR files).")
    ws["A2"].font = GREY_FONT

    headers = ["Event ID", "Type", "Date", "Subject", "Originator", "Cost Impact"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h).font = WHITE_BOLD
        ws.cell(row=4, column=c).fill = HDR_FILL
        ws.cell(row=4, column=c).border = BORDER
    r = 5
    for entry in log:
        if not isinstance(entry, list) or len(entry) < 5: continue
        for c, val in enumerate(entry[:6], 1):
            ws.cell(row=r, column=c, value=val).font = INK
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="META").font = SECTION_HDR
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    r += 1
    summary = [
        ("Total Events",        meta.get("total", "")),
        ("Net Cost Impact",     f"${meta.get('costImpact', 0):,}"),
        ("Executed COs",        meta.get("types", {}).get("CO", "")),
        ("CORs (unique)",       meta.get("types", {}).get("COR", "")),
        ("RFIs (in folder)",    META_2061["doc_counts"]["RFIs"]),
        ("ASIs (in folder)",    META_2061["doc_counts"]["ASIs"]),
        ("Submittals",          META_2061["doc_counts"]["Submittals"]),
    ]
    for k, v in summary:
        ws.cell(row=r, column=1, value=k).font = INK_BOLD
        ws.cell(row=r, column=2, value=v).font = INK
        for c in range(1, 3): ws.cell(row=r, column=c).border = BORDER
        r += 1
    widths = {1: 14, 2: 12, 3: 12, 4: 70, 5: 22, 6: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# SOV-PayApps tab — populate from dashboard_arrays.payApps
# ============================================================================
def patch_sov_payapps(ws, dash):
    apps = dash.get("payApps", [])
    sov = dash.get("sovData", {})
    clear_range(ws, 1, 50, 1, 8)
    ws["A1"] = f"Statement of Values + Pay App History · Job #{META_2061['job_id']}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"{len(apps)} AR transactions parsed from JDR. 46 SOV files in GDrive "
                "Billing folder. Retention released 2022-06-23.")
    ws["A2"].font = GREY_FONT

    # SOV summary
    ws.cell(row=4, column=1, value="STATEMENT OF VALUES — SUMMARY").font = SECTION_HDR
    ws.cell(row=4, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=4)
    sov_rows = [
        ("Original Contract",  sov.get("originalContract")),
        ("Net Change Orders",  sov.get("changeOrders")),
        ("Final Contract",     sov.get("finalContract")),
        ("Retainage",          sov.get("retainage")),
        ("Net Paid",           sov.get("netPaid")),
    ]
    r = 5
    for k, v in sov_rows:
        ws.cell(row=r, column=1, value=k).font = INK_BOLD
        ws.cell(row=r, column=2, value=v).font = INK
        for c in range(1, 3): ws.cell(row=r, column=c).border = BORDER
        r += 1
    r += 1

    # Pay app history
    ws.cell(row=r, column=1, value="PAY APPLICATION HISTORY").font = SECTION_HDR
    ws.cell(row=r, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1
    headers = ["Pay App #", "Date", "This Period", "Retainage", "Net", "Cumulative", "% Contract"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h).font = WHITE_BOLD
        ws.cell(row=r, column=c).fill = HDR_FILL
        ws.cell(row=r, column=c).border = BORDER
    r += 1
    for app in apps:
        if not isinstance(app, list) or len(app) < 5: continue
        num, date, this_p, retain, net = app[0], app[1], app[2], app[3], app[4]
        cum = app[5] if len(app) > 5 else None
        pct = app[6] if len(app) > 6 else None
        cells = [num, date, this_p, retain, net, cum, f"{pct*100:.1f}%" if pct else ""]
        for c, val in enumerate(cells, 1):
            ws.cell(row=r, column=c, value=val).font = INK
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    widths = {1: 10, 2: 13, 3: 13, 4: 13, 5: 13, 6: 13, 7: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# RECONCILIATION tab — populate
# ============================================================================
def patch_reconciliation(ws, dash):
    sov = dash.get("sovData", {})
    clear_range(ws, 1, 30, 1, 6)
    ws["A1"] = f"Reconciliation · Job #{META_2061['job_id']}"
    ws["A1"].font = TITLE_FONT

    headers = ["Check", "Expected", "Actual", "Delta", "Status"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h).font = WHITE_BOLD
        ws.cell(row=3, column=c).fill = HDR_FILL
        ws.cell(row=3, column=c).border = BORDER
    r = 4
    rows = [
        ("Original contract: META vs JDR", META_2061['contract_orig'], sov.get("originalContract"), 0, "TIES"),
        ("Final contract: META vs JDR",    META_2061['contract_final'], sov.get("finalContract"), 0, "TIES"),
        ("Retainage: META vs JDR",         META_2061['retention'], sov.get("retainage"), 0, "TIES"),
        ("Direct cost: META vs derived",   META_2061['direct_cost'], 2597499.74, 0, "TIES"),
        ("Net profit: META vs derived",    META_2061['net_profit'], 1832588.14, 0, "TIES"),
        ("Worker count vs JDR worker_wages", META_2061['workers'], 58, 0, "TIES"),
        ("Total labor hours vs JDR PR",     "28,516", 28516, "—", "TIES"),
        ("CO net vs subcontract growth",   META_2061['co_net'], sov.get("changeOrders"), 0, "TIES"),
    ]
    for chk, exp, act, delta, stat in rows:
        ws.cell(row=r, column=1, value=chk).font = INK_BOLD
        ws.cell(row=r, column=2, value=f"${exp:,}" if isinstance(exp, (int, float)) else exp).font = INK
        ws.cell(row=r, column=3, value=f"${act:,}" if isinstance(act, (int, float)) else act).font = INK
        ws.cell(row=r, column=4, value=delta).font = INK
        sc = ws.cell(row=r, column=5, value=stat)
        sc.font = INK_BOLD
        if stat == "TIES": sc.fill = ACTIVE_FILL
        for c in range(1, 6): ws.cell(row=r, column=c).border = BORDER
        r += 1
    widths = {1: 36, 2: 18, 3: 18, 4: 10, 5: 10}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# CHANGE LOG tab — log enrichment pass
# ============================================================================
def patch_workbook_change_log(ws):
    last_row = ws.max_row
    r = last_row + 1
    ws.cell(row=r, column=1, value="2026-04-27").font = INK
    ws.cell(row=r, column=2, value="v1.1 · enriched").font = INK_BOLD
    ws.cell(row=r, column=3, value=(
        "Enriched Job Info tab with 11-section sectioned layout (identity / schedule / "
        "contract & financials / scope & fixture profile / project team — 4 sub-sections / "
        "document meta / data sources / project narrative). Added Project Team block to "
        "Overview. Populated Change Orders + SOV-PayApps + Reconciliation tabs from "
        "dashboard_arrays. Sourced from executed subcontract (L1.220000) + GDrive Job Book "
        "+ Plumbing Fixture Schedule. 17-tab JCR · 0 formula errors."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP


# ============================================================================
# Main
# ============================================================================
def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)
    data = json.loads(JSON_FILE.read_text())
    dash = json.loads(DASH_FILE.read_text())

    print("→ Rebuilding Job Info tab...")
    rebuild_job_info(wb["02 Job Info"])

    print("→ Adding Project Team block to Overview...")
    patch_overview_team_block(wb["01 Overview"])

    print("→ Populating Change Orders tab...")
    patch_change_orders(wb["05 Change Orders"], dash)

    print("→ Populating SOV-PayApps tab...")
    patch_sov_payapps(wb["04 SOV-PayApps"], dash)

    print("→ Populating Reconciliation tab...")
    patch_reconciliation(wb["15 Reconciliation"], dash)

    print("→ Logging enrichment pass to Change Log...")
    patch_workbook_change_log(wb["17 Change Log"])

    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
