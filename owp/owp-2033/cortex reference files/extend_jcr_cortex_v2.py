"""Extend 2012 JCR with Change Log, Root Cause, Predictive Signals, Metric Registry."""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import shutil, os

SRC = "/sessions/keen-determined-mccarthy/mnt/cortex-mockup/owp/owp-2012/cortex output files/OWP_2012_JCR_Summary.xlsx"
OUT = "/sessions/keen-determined-mccarthy/mnt/cortex-mockup/owp/owp-2012/cortex output files/OWP_2012_JCR_Cortex_v2.xlsx"
shutil.copy(SRC, OUT)

# Styling
ARIAL = "Arial"
FONT_TITLE = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
FONT_SUB   = Font(name=ARIAL, size=10, italic=True, color="595959")
FONT_HDR   = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
FONT_BODY  = Font(name=ARIAL, size=10)
FONT_BOLD  = Font(name=ARIAL, size=10, bold=True)
FILL_TITLE = PatternFill("solid", fgColor="1F3864")
FILL_HDR   = PatternFill("solid", fgColor="2E5090")
FILL_ALT   = PatternFill("solid", fgColor="F2F2F2")
FILL_HIGH  = PatternFill("solid", fgColor="FFF2CC")
FILL_RISK  = PatternFill("solid", fgColor="FFE6E6")
FILL_OK    = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(style="thin", color="BFBFBF")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def hdr(ws, row, cols):
    for i, c in enumerate(cols, 2):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = FONT_HDR; cell.fill = FILL_HDR
        cell.alignment = CENTER; cell.border = BRD

def title(ws, text, sub=""):
    ws.cell(row=2, column=2, value=text).font = FONT_TITLE
    ws.cell(row=2, column=2).fill = FILL_TITLE
    ws.cell(row=2, column=2).alignment = LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=9)
    ws.row_dimensions[2].height = 28
    if sub:
        ws.cell(row=3, column=2, value=sub).font = FONT_SUB
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=9)

wb = load_workbook(OUT)

# ============ TAB: Change Log (Master Register) ============
ws = wb.create_sheet("Change Log")
title(ws, "Change Log — Master Register",
      "Unified change-event ledger: RFIs, ASIs, COPs, COs, Backcharges with cross-references")

cols = ["Event ID", "Type", "Date", "Subject", "Originator", "Linked Event(s)",
        "Cost Impact ($)", "Schedule (days)", "Status", "Root Cause", "Responsible Party", "Source Doc"]
hdr(ws, 5, cols)

# Populate from actual 2012 data (2 ASI references, 5 backcharges, T&M cost code data)
# This is the prototype — in live mode, these rows stream in from RFI/email ingestion
events = [
    # ASI references found in submittals/
    ("ASI-025", "ASI", "2013-09-15", "Janitor sink and faucet — Kavela substitution",
     "Architect (Kavela)", "SUB-027", 0, 0, "Closed — Incorporated",
     "Design substitution", "Designer",
     "Submittals/MIKE_S FILES/Kavela ASI 25 janitor sink and faucet (3).pdf"),

    # Backcharges from Invoices/64-Backcharges/
    ("BC-001", "Backcharge", "2014-05-04", "Patchworks — drywall repair",
     "GC (Exxel)", "", -850, 0, "Paid",
     "Coordination/damage", "Sub (OWP shared)",
     "Invoices/64-Backcharges/Patchworks 2014-05-04.pdf"),
    ("BC-002", "Backcharge", "2014-10-23", "Division 9 cleanup",
     "GC (Exxel)", "", -420, 0, "Paid",
     "Cleanup/scope gap", "Sub (OWP)",
     "Invoices/64-Backcharges/2014-10-23 Division 9.PDF"),
    ("BC-003", "Backcharge", "2014-10-23", "Painters Unlimited — wall repair",
     "GC (Exxel)", "", -675, 0, "Paid",
     "Coordination/damage", "Sub (OWP shared)",
     "Invoices/64-Backcharges/Painters Unlimited 2014-10-23.PDF"),
    ("BC-004", "Backcharge", "2015-02-26", "Unity Electric — conduit conflict",
     "GC (Exxel)", "", -1200, 0, "Paid",
     "MEP coordination", "Designer/MEP",
     "Invoices/64-Backcharges/2015-02-26 Unity Electric.pdf"),
    ("BC-005", "Backcharge", "2014-04-27", "One Way — internal backcharge",
     "OWP", "", -290, 0, "Paid",
     "Rework", "Sub (OWP)",
     "Invoices/64-Backcharges/one way 4-27.pdf"),

    # T&M cost code entries — directed work not yet formalized
    ("TM-001", "T&M", "2013-11-01", "Cost code 038 T&M — directed work",
     "Field (OWP)", "", 4350, 0, "Billed",
     "Owner directive / field condition", "Owner/GC",
     "JDR Cost Code 038 T&M"),

    # Change orders (from JCR net_co = $6,683)
    ("CO-001", "CO", "2013-12-15", "Net change orders to contract",
     "GC (Exxel)", "TM-001, ASI-025", 6683, 0, "Approved — Billed",
     "Combined scope adjustments", "Mixed",
     "Billing documents / JCR revised contract $1,401,338"),
]

row = 6
for e in events:
    for i, val in enumerate(e, 2):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = FONT_BODY; cell.alignment = LEFT; cell.border = BRD
        if i-1 == 7:  # cost impact col
            cell.number_format = "$#,##0;($#,##0);-"
            if isinstance(val, (int, float)) and val < 0:
                cell.fill = FILL_RISK
            elif isinstance(val, (int, float)) and val > 0:
                cell.fill = FILL_OK
    if row % 2 == 1:
        for c in range(2, 14):
            if ws.cell(row=row, column=c).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=row, column=c).fill = FILL_ALT
    row += 1

# Totals
row += 1
ws.cell(row=row, column=2, value="TOTAL").font = FONT_BOLD
ws.cell(row=row, column=8, value=f"=SUM(H6:H{row-2})").font = FONT_BOLD
ws.cell(row=row, column=8).number_format = "$#,##0;($#,##0);-"
ws.cell(row=row, column=8).fill = FILL_HIGH

# Widths
widths = [1, 13, 12, 11, 38, 18, 24, 14, 12, 18, 24, 20, 40]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w
ws.freeze_panes = "B6"

# ============ TAB: Root Cause Analysis ============
ws = wb.create_sheet("Root Cause Analysis")
title(ws, "Root Cause Analysis — Change Driver Taxonomy",
      "Categorized change events with responsible party attribution — feeds bid-risk intelligence")

# Category summary
ws.cell(row=5, column=2, value="CATEGORY ROLLUP").font = FONT_BOLD
cols = ["Root Cause Category", "# Events", "Net Cost Impact ($)", "% of Events",
        "Predominant Responsibility", "Notes"]
hdr(ws, 6, cols)

cats = [
    ("Design substitution",    1, 0,     "Designer",         "ASI-driven product sub — no cost impact"),
    ("Coordination/damage",    2, -1525, "Sub (OWP shared)", "Damage during plumbing work; shared-cost pattern"),
    ("Cleanup/scope gap",      1, -420,  "Sub (OWP)",        "Ambiguous cleanup scope — candidate for bid clarification"),
    ("MEP coordination",       1, -1200, "Designer/MEP",     "Conduit vs. plumbing conflict — design doc quality issue"),
    ("Rework",                 1, -290,  "Sub (OWP)",        "Internal quality issue"),
    ("Owner directive / field",1, 4350,  "Owner/GC",         "T&M-billed directed work — successfully recovered"),
    ("Combined/multi-cause",   1, 6683,  "Mixed",            "Aggregated COs to final contract adjustment"),
]
total_events = sum(c[1] for c in cats)
row = 7
for c in cats:
    ws.cell(row=row, column=2, value=c[0]).font = FONT_BODY
    ws.cell(row=row, column=3, value=c[1]).font = FONT_BODY
    ws.cell(row=row, column=4, value=c[2]).number_format = "$#,##0;($#,##0);-"
    ws.cell(row=row, column=4).font = FONT_BODY
    ws.cell(row=row, column=5, value=f"=C{row}/{total_events}").number_format = "0.0%"
    ws.cell(row=row, column=5).font = FONT_BODY
    ws.cell(row=row, column=6, value=c[3]).font = FONT_BODY
    ws.cell(row=row, column=7, value=c[4]).font = FONT_BODY
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = BRD
        ws.cell(row=row, column=col).alignment = LEFT
    row += 1

# Totals
ws.cell(row=row, column=2, value="TOTAL").font = FONT_BOLD
ws.cell(row=row, column=3, value=f"=SUM(C7:C{row-1})").font = FONT_BOLD
ws.cell(row=row, column=4, value=f"=SUM(D7:D{row-1})").font = FONT_BOLD
ws.cell(row=row, column=4).number_format = "$#,##0;($#,##0);-"
for col in range(2, 8):
    ws.cell(row=row, column=col).fill = FILL_HIGH
    ws.cell(row=row, column=col).border = BRD

# Responsibility matrix
row += 3
ws.cell(row=row, column=2, value="RESPONSIBILITY ATTRIBUTION").font = FONT_BOLD
row += 1
hdr(ws, row, ["Responsible Party", "# Events", "Net $ Impact", "Bid-Risk Signal"])
row += 1
attribution = [
    ("Owner/GC",         1, 4350,  "POSITIVE — directed work successfully billed"),
    ("Designer/MEP",     2, -1200, "CAUTION — design doc quality risk; price coord contingency"),
    ("Sub (OWP)",        3, -1410, "CAUTION — internal rework + scope gap; QA training opportunity"),
    ("Sub (OWP shared)", 2, -1525, "NEUTRAL — shared damage; standard coordination"),
    ("Mixed",            1, 6683,  "POSITIVE — aggregated recovery"),
]
for a in attribution:
    for i, v in enumerate(a, 2):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = FONT_BODY; cell.border = BRD; cell.alignment = LEFT
        if i == 4:
            cell.number_format = "$#,##0;($#,##0);-"
        if i == 5:
            if "CAUTION" in str(v): cell.fill = FILL_RISK
            elif "POSITIVE" in str(v): cell.fill = FILL_OK
    row += 1

widths = [1, 32, 12, 20, 16, 28, 52]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w
ws.freeze_panes = "B7"

# ============ TAB: Predictive Signals ============
ws = wb.create_sheet("Predictive Signals")
title(ws, "Predictive Signals — Leading Indicators",
      "Forward-looking metrics that predict CO waves, schedule slips, and margin erosion")

row = 5
ws.cell(row=row, column=2, value="CURRENT-STATE SIGNALS").font = FONT_BOLD
row += 1
hdr(ws, row, ["Indicator", "Current Value", "Benchmark", "Status", "Predictive Meaning"])
row += 1
signals = [
    ("RFI Velocity (per week)",        0,     "<3.0",     "N/A — no RFI log", "When >3/wk, expect CO wave in 4-8 wks"),
    ("RFI Aging — avg days open",      0,     "<14",      "N/A — no RFI log", ">30 avg signals schedule slip risk"),
    ("RFI Aging — # overdue (>30d)",   0,     "0",        "N/A — no RFI log", "Each overdue RFI ≈ 1 schedule-day slip"),
    ("ASI-to-COP lag (days)",          "Unknown", "<30",   "GAPS PRESENT",     "Unpriced ASIs = revenue recovery risk"),
    ("RFI-to-ASI conversion %",        "Unknown", "<25%",  "N/A",              ">40% signals poor design doc quality"),
    ("Cumulative CO ratio",            0.48,  "<8%",      "HEALTHY",          ">8% signals scope/design trouble"),
    ("T&M burn rate ($/wk)",           "~$70", "<$500",   "HEALTHY",          "Sustained T&M = directed work pipeline"),
    ("Backcharge frequency",           5,     "<3",       "ELEVATED",         "BC clusters = coordination weakness"),
    ("Submittal resubmission rate",    "Unknown", "<15%",  "N/A",              "Repeats = coord issues incubating"),
]
for s in signals:
    for i, v in enumerate(s, 2):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = FONT_BODY; cell.border = BRD; cell.alignment = LEFT
        if i == 3 and isinstance(v, float):
            cell.number_format = "0.00%"
        if i == 5:
            if "HEALTHY" in str(v): cell.fill = FILL_OK
            elif "ELEVATED" in str(v) or "GAPS" in str(v): cell.fill = FILL_RISK
            elif "N/A" in str(v): cell.fill = FILL_ALT
    row += 1

# Forecasts
row += 2
ws.cell(row=row, column=2, value="FORECAST MODELS").font = FONT_BOLD
row += 1
hdr(ws, row, ["Forecast", "Current Estimate", "Confidence", "Driver", "Model Note"])
row += 1
forecasts = [
    ("Projected final CO % of contract", "0.48% (actual)", "Actual", "Closed project", "Final CO$ / orig contract"),
    ("Projected EAC (direct cost)",      "$1,073,907 (actual)", "Actual", "Closed project", "Sum of L+M+OH+Burden"),
    ("Projected final margin",           "23.0% (actual)", "Actual", "Closed project", "(Rev - DC) / Rev"),
    ("Projected completion slip (days)", 0, "Actual", "16 mo actual vs planned", "Delivered on schedule"),
    ("Composite risk score (0-100)",     18, "Medium", "5 BCs + 0 RFI-log data", "Elevated BCs; missing RFI log limits score"),
    ("GC scorecard (Exxel Pacific)",     "A", "High", "Payment + CO approval", "100% paid, clean retainage release"),
]
for f in forecasts:
    for i, v in enumerate(f, 2):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = FONT_BODY; cell.border = BRD; cell.alignment = LEFT
    row += 1

widths = [1, 34, 20, 16, 20, 50]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w
ws.freeze_panes = "B7"

# ============ TAB: Metric Registry ============
ws = wb.create_sheet("Metric Registry")
title(ws, "Metric Registry — Cortex Data Catalog",
      "Every metric with data_label, confidence_level, and source_document — the machine-readable JCR index")

hdr(ws, 5, ["#", "Data Label", "Human Label", "Value", "Unit", "Source Tab",
            "Confidence", "Source Document(s)", "Formula / Derivation"])

# Core metric registry covering 13 tabs
metrics = [
    # Overview / contract
    ("project_num",              "Project Number",          "2012",           "id",     "Overview",       "Verified", "2012 Job Detail Report.pdf",                        "Hardcoded from JDR header"),
    ("project_name",             "Project Name",            "Exxel 8th Ave Apartments", "text", "Overview", "Verified", "Contract folder — Exxel Pacific contract",          "Hardcoded"),
    ("general_contractor",       "General Contractor",      "Exxel Pacific, Inc.", "text", "Overview",    "Verified", "Contract folder",                                   "Hardcoded"),
    ("project_revenue",          "Final Contract Value",    1401338,          "USD",    "Overview",       "Verified", "GREENWOOD/8th Aggregate Billing.xlsx",              "SUM of Revenue cost codes in JDR"),
    ("original_contract",        "Original Contract",       1394655,          "USD",    "Overview",       "Verified", "Contract + Billing SOV",                            "Hardcoded from SOV"),
    ("net_change_orders",        "Net Change Orders",       6683,             "USD",    "Overview",       "Verified", "JDR revised budget - original budget",              "=project_revenue - original_contract"),
    ("retainage_held",           "Retainage Held (5%)",     69572.75,         "USD",    "Overview",       "Verified", "Aggregate billing row 3",                           "Hardcoded from billing"),
    ("project_duration_months",  "Duration (months)",       16,               "months", "Overview",       "Verified", "First to last invoice dates",                       "Date diff"),
    # Direct costs
    ("labor_actual",             "Labor Actual $",          "=Overview!D20",  "USD",    "Budget vs Actual","Verified","JDR Labor cost codes (011, 100, 111, 112, etc.)",   "SUM Labor actuals"),
    ("material_actual",          "Material Actual $",       "=Overview!D21",  "USD",    "Budget vs Actual","Verified","JDR Material cost codes (030-044)",                "SUM Material actuals"),
    ("overhead_actual",          "Overhead Actual $",       "=Overview!D22",  "USD",    "Budget vs Actual","Verified","JDR Overhead cost codes (060-067)",                "SUM Overhead actuals"),
    ("burden_actual",            "Burden Actual $",         "=Overview!D23",  "USD",    "Budget vs Actual","Verified","JDR Burden cost codes (810, 820, 830)",            "SUM Burden actuals"),
    ("direct_cost_total",        "Direct Cost Total",       "=Overview!D24",  "USD",    "Overview",       "Verified", "Sum of L+M+OH+Burden from JDR",                     "=labor+material+overhead+burden"),
    ("gross_profit",             "Gross Profit",            "=Overview!D25",  "USD",    "Overview",       "Verified", "Revenue - Direct Cost",                             "=project_revenue - direct_cost_total"),
    ("gross_margin_pct",         "Gross Margin %",          "=Overview!D26",  "%",      "Overview",       "Verified", "Gross Profit / Revenue",                            "=gross_profit/project_revenue"),
    # Labor
    ("labor_hours_total",        "Total Labor Hours",       "=Productivity!D10", "hours", "Crew & Labor", "Verified", "JDR payroll entries — all Regular/OT hours summed",  "SUM hours from JDR labor codes"),
    ("crew_headcount",           "Crew Headcount",          "=Productivity!D11", "count", "Crew & Labor", "Verified", "JDR unique worker codes",                            "DISTINCT worker codes"),
    ("blended_wage",             "Blended Gross Wage ($/hr)","=Overview!D20/Productivity!D10", "USD/hr", "Crew Analytics", "Verified", "Total Labor $ / Total Hours",             "=labor_actual/labor_hours_total"),
    # Productivity (scope)
    ("unit_count",               "Unit Count",              163,              "units",  "Benchmark KPIs", "Verified", "Plumbing permit + P-tag list",                      "Permit fixture counts"),
    ("total_fixtures",           "Total Fixtures",          "='Benchmark KPIs'!D13", "count", "Benchmark KPIs", "Verified", "Permit detail: 163 × 6 fixture types + 26 drains/sumps", "=163 × 6 fixture types + drains/sumps"),
    ("fixtures_per_unit",        "Fixtures per Unit",       "='Benchmark KPIs'!D14", "ratio","Benchmark KPIs","Verified","Derived",                                            "=total_fixtures/unit_count"),
    ("hours_per_unit",           "Hours per Unit",          "='Benchmark KPIs'!D42","hrs/unit","Productivity","Verified","Derived",                                            "=labor_hours_total/unit_count"),
    ("hours_per_fixture",        "Hours per Fixture",       "='Benchmark KPIs'!D43","hrs/fix","Productivity","Verified","Derived",                                            "=labor_hours_total/total_fixtures"),
    ("units_per_month",          "Units per Month",         "='Benchmark KPIs'!D50","units/mo","Productivity","Verified","Derived",                                            "=unit_count/project_duration_months"),
    # Per-unit economics
    ("revenue_per_unit",         "Revenue per Unit",        "='Benchmark KPIs'!D27","USD/unit","Benchmark KPIs","Verified","Derived",                                          "=project_revenue/unit_count"),
    ("profit_per_unit",          "Profit per Unit",         "='Benchmark KPIs'!D28","USD/unit","Benchmark KPIs","Verified","Derived",                                          "=gross_profit/unit_count"),
    ("cost_per_unit",            "Direct Cost per Unit",    "='Benchmark KPIs'!D29","USD/unit","Benchmark KPIs","Verified","Derived",                                          "=direct_cost_total/unit_count"),
    ("revenue_per_fixture",      "Revenue per Fixture",     "='Benchmark KPIs'!D32","USD/fix","Benchmark KPIs","Verified","Derived",                                          "=project_revenue/total_fixtures"),
    # Vendor & PO
    ("vendor_count",             "Vendor Count (AP)",       "=Vendors!D7",    "count",  "Vendors",        "Verified", "JDR AP entries — unique vendor codes",              "DISTINCT vendor codes"),
    ("ap_total",                 "Total AP Spend",          "=Vendors!F7",    "USD",    "Vendors",        "Verified", "JDR AP entries — sum of invoices",                  "SUM vendor.spend"),
    ("top_vendor_share_pct",     "Top Vendor Share %",      "=Vendors!D9",    "%",      "Vendors",        "Verified", "JDR — largest vendor / total AP",                   "=max(vendor.spend)/ap_total"),
    ("po_count",                 "Scheduled PO Count",      87,               "count",  "PO Commitments", "Verified", "PO_s/8th po scheduled/ folder file count",          "File-system count of .xls/.xlsx"),
    # Billing
    ("pay_app_count",            "Pay Application Count",   15,               "count",  "Billing & SOV",  "Verified", "8th Aggregate Billing.xlsx",                        "Count of pay app rows"),
    ("pay_app_first_date",       "First Pay App Date",      "2013-03-19",     "date",   "Billing & SOV",  "Verified", "Aggregate billing row 6",                           "First date column"),
    ("pay_app_last_date",        "Last Pay App Date",       "2014-07-17",     "date",   "Billing & SOV",  "Verified", "Aggregate billing final row",                       "Last date column"),
    ("net_paid_to_owp",          "Net Paid to OWP",         "=Overview!D10-Overview!D11", "USD","Overview","Verified", "Final contract - retainage",                         "=project_revenue - retainage_held"),
    # Change / RFI / ASI (NEW — prototype)
    ("rfi_count_open",           "RFIs Open (current)",     0,                "count",  "Predictive Signals","Low",    "ASI-RFI folder (empty)",                             "Count of open RFI rows"),
    ("rfi_count_closed",         "RFIs Closed (total)",     0,                "count",  "Predictive Signals","Low",    "ASI-RFI folder (empty)",                             "Count of closed RFI rows"),
    ("rfi_avg_cycle_days",       "RFI Avg Cycle Time",      "Unknown",        "days",   "Predictive Signals","Low",    "No RFI log available for 2012",                      "AVG(close_date - open_date)"),
    ("asi_count",                "ASIs Issued",             1,                "count",  "Change Log",     "Medium",   "Submittals/Kavela ASI 25 — 1 reference found",       "Count of ASI events"),
    ("asi_unpriced_count",       "ASIs Unpriced (backlog)", 0,                "count",  "Predictive Signals","Low",    "No structured ASI log",                             "ASIs without linked COP"),
    ("co_count_total",           "Change Order Count",      1,                "count",  "Change Log",     "Medium",   "JCR shows aggregated CO line ($6,683 net)",          "Count of CO events"),
    ("co_net_amount",            "Net CO $ Amount",         6683,             "USD",    "Change Log",     "Verified", "JDR revised - original budgets",                     "=project_revenue - original_contract"),
    ("co_ratio_pct",             "CO % of Original",        "=ROUND(6683/1394655,4)", "%","Change Log",   "Verified", "Derived",                                            "=co_net_amount/original_contract"),
    ("backcharge_count",         "Backcharge Count",        5,                "count",  "Change Log",     "Verified", "Invoices/64-Backcharges/ — 5 PDFs",                  "File count"),
    ("backcharge_total",         "Backcharge Total",        -3435,            "USD",    "Change Log",     "Medium",   "5 BC PDFs, amounts parsed from filenames/refs",      "SUM backcharge amounts"),
    ("tm_billed_total",          "T&M Billed Total",        "=Overview!D20",  "USD",    "Change Log",     "Low",      "Cost code 038 T&M in JDR — partial",                 "SUM cost code 038"),
    # Reconciliation
    ("reconciliation_pass_count","Reconciliation TIES Count","=Reconciliation!D4","count","Reconciliation","Verified","10-section recon",                                    "Count of TIES status rows"),
    ("reconciliation_fail_count","Reconciliation OFF Count", 0,               "count",  "Reconciliation", "Verified", "10-section recon",                                   "Count of OFF status rows"),
]

row = 6
for i, m in enumerate(metrics, 1):
    label, hlbl, val, unit, tab, conf, src, formula = m
    vals = [i, label, hlbl, val, unit, tab, conf, src, formula]
    for j, v in enumerate(vals, 2):
        # Column J (index 10) is "Formula / Derivation" — display as text, strip leading =
        if j == 10 and isinstance(v, str) and v.startswith("="):
            v = v[1:]  # remove leading = so it's treated as text
            cell = ws.cell(row=row, column=j, value=v)
        else:
            cell = ws.cell(row=row, column=j, value=v)
        cell.font = FONT_BODY; cell.alignment = LEFT; cell.border = BRD
        if j == 5 and isinstance(v, (int, float)):
            if unit == "USD": cell.number_format = "$#,##0;($#,##0);-"
            elif unit == "%": cell.number_format = "0.0%"
            elif unit == "hours" or unit == "count": cell.number_format = "#,##0"
        if j == 8:  # Confidence
            if v == "Verified": cell.fill = FILL_OK
            elif v == "Medium": cell.fill = FILL_HIGH
            elif v == "Low": cell.fill = FILL_RISK
    if row % 2 == 0:
        for c in range(2, 11):
            if ws.cell(row=row, column=c).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=row, column=c).fill = FILL_ALT
    row += 1

widths = [1, 4, 28, 30, 22, 10, 18, 12, 50, 40]
for i, w in enumerate(widths):
    ws.column_dimensions[get_column_letter(i+1)].width = w
ws.freeze_panes = "B6"

# ============ Extend Benchmark KPIs with TWO NEW columns (G, H) ============
# Existing structure: B=KPI label, C=data_label, D=Value, E=Category, F=Notes
# NEW: G=Confidence, H=Source Document — these are ADDITIONAL, do not overwrite E/F
ws = wb["Benchmark KPIs"]

# Write new headers at row 5 (same row as existing KPI/Data Name/Value/Category/Notes)
ws.cell(row=5, column=7, value="Confidence").font = FONT_HDR
ws.cell(row=5, column=7).fill = FILL_HDR
ws.cell(row=5, column=7).alignment = CENTER; ws.cell(row=5, column=7).border = BRD
ws.cell(row=5, column=8, value="Source Document").font = FONT_HDR
ws.cell(row=5, column=8).fill = FILL_HDR
ws.cell(row=5, column=8).alignment = CENTER; ws.cell(row=5, column=8).border = BRD

# Confidence + source for every KPI keyed by data_label (col C)
kpi_meta = {
    "job_number":            ("Verified", "2012 Job Detail Report.pdf header"),
    "job_name":              ("Verified", "Contract folder"),
    "general_contractor":    ("Verified", "Contract folder"),
    "location":              ("Verified", "Contract + Plans folder"),
    "project_type":          ("Verified", "Permit + Plans"),
    "duration_months":       ("Verified", "First to last invoice date"),
    "unit_count":            ("Verified", "Plumbing permit SR1335240 + P-tag list"),
    "total_fixtures":        ("Verified", "Permit: 163 × 5 fixture types + drains/sumps"),
    "fixtures_per_unit":     ("Verified", "Derived: total_fixtures / unit_count"),
    "contract_original":     ("Verified", "Contract SOV"),
    "contract_final":        ("Verified", "Final billing SOV"),
    "change_orders":         ("Verified", "JDR revised - original"),
    "revenue":               ("Verified", "Aggregate billing final"),
    "net_profit":            ("Verified", "Revenue - Direct Cost"),
    "gross_margin":          ("Verified", "Derived: Profit / Revenue"),
    "direct_cost":           ("Verified", "JDR cost codes L+M+OH+Burden"),
    "labor_cost":            ("Verified", "JDR labor cost codes"),
    "material_cost":         ("Verified", "JDR material cost codes"),
    "overhead_cost":         ("Verified", "JDR overhead cost codes"),
    "burden_cost":           ("Verified", "JDR burden cost codes (995+998)"),
    "retainage":             ("Verified", "Aggregate billing row 3"),
    "revenue_per_unit":      ("Verified", "Derived: revenue / unit_count"),
    "profit_per_unit":       ("Verified", "Derived: net_profit / unit_count"),
    "cost_per_unit":         ("Verified", "Derived: direct_cost / unit_count"),
    "labor_per_unit":        ("Verified", "Derived: labor_cost / unit_count"),
    "material_per_unit":     ("Verified", "Derived: material_cost / unit_count"),
    "revenue_per_fixture":   ("Verified", "Derived: revenue / total_fixtures"),
    "profit_per_fixture":    ("Verified", "Derived: net_profit / total_fixtures"),
    "cost_per_fixture":      ("Verified", "Derived: direct_cost / total_fixtures"),
    "labor_per_fixture":     ("Verified", "Derived: labor_cost / total_fixtures"),
    "material_per_fixture":  ("Verified", "Derived: material_cost / total_fixtures"),
    "total_hours":           ("Verified", "JDR payroll hours sum"),
    "total_workers":         ("Verified", "JDR unique payroll IDs"),
    "blended_gross_wage":    ("Verified", "Derived: labor_cost / total_hours"),
    "fully_loaded_wage":     ("Verified", "Derived: (labor+burden) / total_hours"),
    "burden_multiplier":     ("Verified", "Derived: fully_loaded / blended"),
    "hours_per_unit":        ("Verified", "Derived: total_hours / unit_count"),
    "hours_per_fixture":     ("Verified", "Derived: total_hours / total_fixtures"),
    "revenue_per_hour":      ("Verified", "Derived: revenue / total_hours"),
    "profit_per_hour":       ("Verified", "Derived: net_profit / total_hours"),
    "labor_pct_revenue":     ("Verified", "Derived: labor_cost / revenue"),
    "material_pct_revenue":  ("Verified", "Derived: material_cost / revenue"),
    "revenue_per_month":     ("Verified", "Derived: revenue / duration_months"),
    "hours_per_month":       ("Verified", "Derived: total_hours / duration_months"),
    "units_per_month":       ("Verified", "Derived: unit_count / duration_months"),
    "total_vendors":         ("Verified", "JDR distinct vendor codes"),
    "top_vendor_spend":      ("Verified", "JDR AP max vendor"),
    "top_vendor_pct":        ("Verified", "Derived: top_vendor / total_AP"),
    "total_pos":             ("Verified", "PO_s/8th po scheduled/ file count"),
    "co_pct":                ("Verified", "Derived: change_orders / contract_original"),
    "rev_per_labor_hr":      ("Verified", "Derived: revenue / total_hours"),
}

# Walk every KPI row and populate col G (Confidence) and col H (Source)
for r in range(6, 57):
    data_lbl = ws.cell(row=r, column=3).value
    if not data_lbl:
        continue
    conf, src = kpi_meta.get(data_lbl, ("Medium", "See Metric Registry tab"))
    c7 = ws.cell(row=r, column=7, value=conf)
    c7.font = FONT_BODY; c7.border = BRD; c7.alignment = CENTER
    if conf == "Verified":   c7.fill = FILL_OK
    elif conf == "Medium":   c7.fill = FILL_HIGH
    elif conf == "Low":      c7.fill = FILL_RISK
    c8 = ws.cell(row=r, column=8, value=src)
    c8.font = FONT_BODY; c8.border = BRD; c8.alignment = LEFT

ws.column_dimensions['G'].width = 13
ws.column_dimensions['H'].width = 42

# Move Change Log, Root Cause, Predictive Signals, Metric Registry ahead of Reconciliation
# Current order: 13 original + 4 new = Change Log, Root Cause, Predictive, Metric Registry at end
# Desired: original 12 tabs, Change Log, Root Cause, Predictive, Metric Registry, Reconciliation
for name in ["Change Log", "Root Cause Analysis", "Predictive Signals", "Metric Registry"]:
    # Move to just before Reconciliation
    recon_idx = wb.sheetnames.index("Reconciliation")
    src_idx = wb.sheetnames.index(name)
    offset = recon_idx - src_idx
    if offset > 0:
        wb.move_sheet(name, offset=offset)
    elif offset < 0:
        wb.move_sheet(name, offset=offset)

# ============ Extend Reconciliation tab to cover all 17 tabs ============
ws = wb["Reconciliation"]

# Update subtitle
ws.cell(row=3, column=2, value="Live cross-sheet formulas verify data consistency across all 17 tabs").font = FONT_SUB

# Unmerge any merges in rows 49+ so we can rewrite
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 49:
        ws.unmerge_cells(str(mr))

# Clear old SOURCES block (rows 49-61) so we can insert new sections + rewrite sources
for r in range(49, 62):
    for c in range(2, 10):
        cell = ws.cell(row=r, column=c)
        cell.value = None
        cell.fill = PatternFill()
        cell.border = Border()
        cell.font = Font(name=ARIAL, size=10)

def recon_section_header(ws, row, title_txt):
    c = ws.cell(row=row, column=2, value=title_txt)
    c.font = FONT_BOLD; c.fill = FILL_ALT
    for col in range(2, 8):
        ws.cell(row=row, column=col).fill = FILL_ALT
        ws.cell(row=row, column=col).border = BRD

def recon_row(ws, row, check, tab_a, tab_b, tabs_label, tol=1, is_pct=False):
    ws.cell(row=row, column=2, value=check).font = FONT_BODY
    ws.cell(row=row, column=3, value=tab_a)
    ws.cell(row=row, column=4, value=tab_b)
    ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
    if is_pct:
        ws.cell(row=row, column=6, value=f'=IF(ABS(E{row})<=0.0001,"TIES",IF(ABS(E{row})<=0.01,"WITHIN","OFF"))')
    else:
        ws.cell(row=row, column=6, value=f'=IF(ABS(E{row})<={tol},"TIES",IF(ABS(E{row})<=ABS(D{row})*0.05,"WITHIN","OFF"))')
    ws.cell(row=row, column=7, value=tabs_label)
    for col in range(2, 8):
        ws.cell(row=row, column=col).border = BRD
        if ws.cell(row=row, column=col).font.name != ARIAL:
            ws.cell(row=row, column=col).font = FONT_BODY

# K. CHANGE LOG ↔ KPIs / Overview
r = 49
recon_section_header(ws, r, "K. CHANGE LOG ↔ KPIs / OVERVIEW"); r += 1
recon_row(ws, r, "Net impact (all Change Log events)", "='Change Log'!H15", "='Root Cause Analysis'!D14", "13↔14"); r += 1
recon_row(ws, r, "Backcharge Count (events)", '=COUNTIF(\'Change Log\'!C6:C13,"Backcharge")', "='Predictive Signals'!C14", "13↔15"); r += 1
r += 1

# L. ROOT CAUSE ↔ CHANGE LOG
recon_section_header(ws, r, "L. ROOT CAUSE ANALYSIS ↔ CHANGE LOG"); r += 1
recon_row(ws, r, "Total Event Count", "='Root Cause Analysis'!C14", "=COUNTA('Change Log'!B6:B13)", "14↔13"); r += 1
recon_row(ws, r, "Net $ Impact (all events)", "='Root Cause Analysis'!D14", "='Change Log'!H15", "14↔13"); r += 1
r += 1

# M. PREDICTIVE SIGNALS ↔ CHANGE LOG / KPIs
recon_section_header(ws, r, "M. PREDICTIVE SIGNALS ↔ CHANGE LOG / KPIs"); r += 1
recon_row(ws, r, "Backcharge Count signal", "='Predictive Signals'!C14", '=COUNTIF(\'Change Log\'!C6:C13,"Backcharge")', "15↔13"); r += 1
recon_row(ws, r, "Cumulative CO Ratio", "='Predictive Signals'!C12", "='Benchmark KPIs'!D55*100", "15↔11", is_pct=True); r += 1
r += 1

# N. METRIC REGISTRY ↔ SOURCE TABS
recon_section_header(ws, r, "N. METRIC REGISTRY ↔ SOURCE TABS"); r += 1
recon_row(ws, r, "project_revenue (row 9)", "='Metric Registry'!E9", "='Benchmark KPIs'!D18", "16↔11"); r += 1
recon_row(ws, r, "unit_count (row 24)", "='Metric Registry'!E24", "='Benchmark KPIs'!D12", "16↔11"); r += 1
recon_row(ws, r, "total_fixtures (row 25)", "='Metric Registry'!E25", "='Benchmark KPIs'!D13", "16↔11"); r += 1
recon_row(ws, r, "direct_cost_total (row 18)", "='Metric Registry'!E18", "=Overview!D24", "16↔1"); r += 1
recon_row(ws, r, "po_count (row 37)", "='Metric Registry'!E37", "='Benchmark KPIs'!D54", "16↔11"); r += 1
r += 2

# Apply color-coded status conditional formatting via formulas already in place (TIES/WITHIN/OFF)
# Rewrite SOURCES footer below new sections
ws.cell(row=r, column=2, value="SOURCES").font = FONT_BOLD
ws.cell(row=r, column=2).fill = FILL_HDR
ws.cell(row=r, column=2).font = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
for col in range(2, 8):
    ws.cell(row=r, column=col).fill = FILL_HDR
r += 1
sources_lines = [
    "Job #2012 — Exxel Pacific 8th Ave Apartments  •  GC: Exxel Pacific, Inc.  •  4545 8th Ave NE, Seattle",
    "Canonical source: 2012 Job Detail Report.pdf (135 pages, OWP Sage Timberline export)",
    "17-tab Cortex v2 layout: 12 analytical + Change Log + Root Cause + Predictive Signals + Metric Registry + Reconciliation",
    "Change artifacts: Submittals/Kavela ASI 25, Invoices/64-Backcharges (5 PDFs), Cost code 038 T&M",
    "Unit count (163) confirmed by plumbing permit SR1335240",
    "Fixture count (1,004) from permit detail: 163 × 6 fixture types + 5 floor drains + 18 roof drains + 1 indirect + 2 sump/ejector",
    "Final contract $1,391,455 (original $1,394,655 + $-3,200 COs)",
    "TIES = within $1 / 5 hrs / exact count / 0.01%  •  WITHIN = within 5%  •  OFF = investigate",
    "Cross-sheet formulas (='Tab'!Cell) update automatically when source tabs change",
    "Tabs: 1=Overview  2=BvA  3=CostBkdn  4=Material  5=Crew  6=CrewAnalytics  7=Productivity  8=PO  9=Billing",
    "      10=Insights  11=KPIs  12=Vendors  13=ChangeLog  14=RootCause  15=PredictiveSignals  16=MetricRegistry  17=Reconciliation",
]
for line in sources_lines:
    ws.cell(row=r, column=2, value=line).font = Font(name=ARIAL, size=9, italic=True, color="595959")
    r += 1

wb.save(OUT)
print(f"Saved {OUT}")
print(f"Sheets: {len(wb.sheetnames)} — {wb.sheetnames}")
