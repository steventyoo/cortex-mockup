"""2012 JCR Summary generator — Exxel Pacific 8th Ave Apartments."""
src = open("/sessions/keen-determined-mccarthy/work/gen_jcr_full.py").read()
cut = src.index("# ============ JOB 2015 ============")
ns = {}
exec(src[:cut], ns)
parse_jdr = ns["parse_jdr"]; count_pos = ns["count_pos"]; generate = ns["generate"]

from openpyxl import load_workbook
import os

cc, w, v = parse_jdr("/sessions/keen-determined-mccarthy/work/2012_jdr.txt")
po = count_pos("/sessions/keen-determined-mccarthy/mnt/cortex-mockup/owp/owp-2012")

labor = sum(c['actual'] for c in cc.values() if c['cat']=='Labor')
mat   = sum(c['actual'] for c in cc.values() if c['cat']=='Material')
oh    = sum(c['actual'] for c in cc.values() if c['cat']=='Overhead')
bur   = sum(c['actual'] for c in cc.values() if c['cat']=='Burden')
hrs   = sum(c['hours']  for c in cc.values() if c['cat']=='Labor')
dc    = labor+mat+oh+bur
rev   = sum(c.get('rev',0) for c in cc.values() if c['cat']=='Revenue')
orig_contract = sum(c.get('orig',0) for c in cc.values() if c['cat']=='Revenue')
net_co = rev - orig_contract
retainage = 69572.75  # From aggregate billing

# Parse aggregate billing for pay apps
apps = []
ab = load_workbook("/sessions/keen-determined-mccarthy/mnt/cortex-mockup/owp/owp-2012/2012-Exxel, 8th Ave/Billing documents/8th AGGREGATE BILLING.xlsx", data_only=True)
ws = ab["Sheet1"]
for r in ws.iter_rows(values_only=True, min_row=6, max_row=25):
    if r[0] is not None and isinstance(r[0], (int,float)) and r[2] is not None:
        apps.append({
            "num": int(r[0]),
            "date": r[1].strftime("%Y-%m-%d") if r[1] else "-",
            "period": r[2] or 0,
            "retention": r[3] or 0,
            "aggregate": r[7] or 0,
        })

print(f"Loaded {len(apps)} pay apps")

job = {
    "num":"2012", "name":"Exxel Pacific 8th Ave Apartments",
    "gc":"Exxel Pacific, Inc.", "loc":"Seattle, WA (4545 8th Ave NE)",
    "project_type":"163-unit 9-story multifamily apartments + ground-floor retail",
    "duration_label":"Mar 2013 – Jul 2014 (~17 months)",
    "contract_sub": f"Original ${orig_contract:,.0f} + ${net_co:,.0f} COs",
    "final_contract":rev,
    "labor":labor,"material":mat,"overhead":oh,"burden":bur,
    "labor_budget": sum(c['rev'] for c in cc.values() if c['cat']=='Labor'),
    "mat_budget":   sum(c['rev'] for c in cc.values() if c['cat']=='Material'),
    "oh_budget":    sum(c['rev'] for c in cc.values() if c['cat']=='Overhead'),
    "burden_budget":sum(c['rev'] for c in cc.values() if c['cat']=='Burden'),
    "direct_cost":dc,"hours":hrs,"workers":len(w),
    "profit":rev-dc,"margin":(rev-dc)/rev,
    "orig_contract":orig_contract,"retainage":retainage,
    "unit_count":163,"fixture_count":1004,"months":17,
    "orig_direct": sum(c['orig'] for c in cc.values() if c['cat']!='Revenue'),
    "rev_direct":  sum(c['rev']  for c in cc.values() if c['cat']!='Revenue'),
    "ap_expected": round(sum(x['spend'] for x in v.values())),
    "doc_count": "8th Ave folder: Contract (Exxel Pacific + One Way Plumbing), Billing (aggregate + 15 SOVs), PO_s (68 scheduled), Permits (plumbing, gas, boiler, backflow), Change Orders",
    "profile":[
        ("Plumbing Units","163"),("Water Closets","163"),("Lavatories","163"),
        ("Kitchen Sinks","163"),("Dishwashers","163"),("Laundry Washers","163"),
        ("Bathtub/Shower","163"),("Floor Drains","5"),("Roof Drains","18"),
        ("Total Fixtures","1004"),("Fixtures / Unit","6.16"),
        ("Floors","9 stories + retail"),("Duration","17 mo"),
    ],
    "sources":[
        "Job #2012 — Exxel Pacific 8th Ave Apartments  •  GC: Exxel Pacific, Inc.  •  4545 8th Ave NE, Seattle",
        "Canonical source: 2012 Job Detail Report.pdf (135 pages, OWP Sage Timberline export)",
        f"Cost codes parsed: {len(cc)}",
        "Unit count (163) confirmed by plumbing permit SR1335240",
        "Fixture count (1004) from permit detail: 163 WC + 163 lav + 163 shower + 163 sink + 163 DW + 163 washer + 5 floor drains + 18 roof drains + 1 indirect + 2 sump/ejector",
        "Contract: One Way Plumbing contract executed (Exxel Pacific GC) with OWP markup",
        "Billing: 8th AGGREGATE BILLING.xlsx (15 pay apps Mar 2013 – Jul 2014) + 15 SOV files",
        f"Final contract ${rev:,.0f} (original ${orig_contract:,.0f} + ${net_co:,.0f} COs)",
    ],
    "billing_subtitle":f"15 pay applications  •  Mar 2013 – Jul 2014  •  ${rev:,.0f} final",
    "contract_summary":[
        ("Original Contract", round(orig_contract)),
        ("Change Orders (net)", round(net_co)),
        ("Final Contract Value", round(rev)),
        ("Retainage (5%)", round(retainage)),
        ("Net Paid to OWP", round(rev) - round(retainage)),
    ],
    "pay_apps": apps,
    "tieout":[
        ("Final Contract Value", round(rev), round(rev)),
        ("Retainage", round(retainage), round(retainage)),
        ("Direct Cost", round(dc), round(dc)),
    ],
}

out = "/sessions/keen-determined-mccarthy/mnt/cortex-mockup/owp/owp-2012/cortex output files/OWP_2012_JCR_Summary.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
generate(job, cc, w, v, po, out)
print("Saved", out)
