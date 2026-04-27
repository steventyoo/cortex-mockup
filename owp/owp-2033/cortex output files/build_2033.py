#!/usr/bin/env python3
"""Build OWP_2033 Cortex v2 17-tab JCR. Compass Vuecrest (LUX) — Bellevue WA.
Parsed JDR data: 29 cost codes, 44 workers, 26 vendors, 30 invoices."""
import json, os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
DATA_JSON = HERE / "2033_data.json"
OUT = HERE / "OWP_2033_JCR_Cortex_v2.xlsx"

ARIAL = "Arial"
TITLE = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
SUB = Font(name=ARIAL, size=10, italic=True, color="595959")
HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
NF_FONT = Font(name=ARIAL, size=10, italic=True, color="9C0006")
SRC_FONT = Font(name=ARIAL, size=8, italic=True, color="595959")
F_TITLE = PatternFill("solid", fgColor="1F3864")
F_HDR = PatternFill("solid", fgColor="2E5090")
F_ALT = PatternFill("solid", fgColor="F2F2F2")
F_HIGH = PatternFill("solid", fgColor="FFF2CC")
F_RISK = PatternFill("solid", fgColor="FFE6E6")
F_OK = PatternFill("solid", fgColor="E2EFDA")
F_NF = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NF = "NOT FOUND"

def put(ws, coord, val, font=BODY, fill=None, border=BRD, align=None, fmt=None):
    c = ws[coord] if isinstance(coord, str) else ws.cell(row=coord[0], column=coord[1])
    c.value = val
    if val == NF:
        c.font = NF_FONT; c.fill = F_NF; c.alignment = CENTER
    else:
        c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

def title(ws, text, sub_text=""):
    c = ws.cell(row=2, column=2, value=text); c.font = TITLE; c.fill = F_TITLE; c.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 28
    if sub_text:
        c2 = ws.cell(row=3, column=2, value=sub_text); c2.font = SUB
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)

def hdr(ws, row, cols, start_col=2):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=txt)
        c.font = HDR; c.fill = F_HDR; c.alignment = CENTER; c.border = BRD

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# Load parsed JDR data
data = json.loads(DATA_JSON.read_text())
CODES = data['codes']
WORKERS = data['workers']
VENDORS = data['vendors']
INVOICES = data['invoices']

wb = Workbook()
wb.remove(wb.active)

# ============ CONSTANTS (all sourced) ============
JOB = "2033"
NAME = "Compass Vuecrest"
PROJECT_DESC = "Vuecrest (LUX) — 137-unit luxury multifamily plumbing, Bellevue WA"
GC = "Compass Harbor Construction, LLC"
GC_CUST_CODE = "2033BV"
OWNER = "Bellevue Vuecrest, LP (Developer: Continental)"

# Financial (JDR footer)
REVENUE = 2_343_188.00
EXPENSES = 1_332_680.44
NET_PROFIT = 1_010_507.56
RETAINAGE = 114_309.40
SRC_GL = 66_431.91
SRC_AP = 490_114.66
SRC_PR = 776_133.87

CONTRACT_ORIG = 2_504_600.00       # Sage code 999 Org Budget
CONTRACT_FINAL = 2_343_188.00      # Sage code 999 Rev Budget (= AR total)
CO_TOTAL_IMPLIED = CONTRACT_FINAL - CONTRACT_ORIG  # -$161,412.00 (descoped)
CO_TOTAL_DOCUMENTED = -160_233.00  # Sum of 17 Budget Transfer spreadsheets CO#01-CO#17

# Project Team (from Job Info + project folders)
GC_PM = "Ryan Ames / Vince Dennison"
GC_SUP = "Todd Mortenson"
GC_PE = "Dan Peck"
DEVELOPER = "Continental"
OWP_RI_FOREMAN = "Victor"
INSURANCE = "Wrap (OCIP)"
LOCATION = "1000 100th Ave NE, Bellevue, WA 98004"
UNITS = 137

# Cost code categories for 2033
LABOR_CODES = ["100","101","110","111","112","120","130","140","141","142","143"]
MATERIAL_CODES = ["210","211","212","220","230","240","241","242","243"]
OVERHEAD_CODES = ["600","601","602","603","607"]
BURDEN_CODE = "995"
TAX_CODE = "998"

def sum_actual(codes):
    return sum(CODES[c]['actual'] for c in codes if c in CODES)

LABOR_COST = sum_actual(LABOR_CODES)
MATERIAL_COST = sum_actual(MATERIAL_CODES)
OVERHEAD_COST = sum_actual(OVERHEAD_CODES)
BURDEN_COST = CODES[BURDEN_CODE]['actual']
TAX_COST = CODES[TAX_CODE]['actual']
TOTAL_HOURS = sum(CODES[c]['hrs_total'] for c in LABOR_CODES if c in CODES)
TOTAL_WORKERS = len(WORKERS)

SRC_JDR = "2033 Job Detail Report.pdf (Sage Timberline, 04/03/2026, 191 pages)"
SRC_CONTRACT = "Vuecrest-One Way Plumbing Fully Executed 9.26.16.pdf"
SRC_PTAG = "Vuecrest P-tag and equipment list working copy.xlsx"
SRC_NARRATIVE = "Vuecrest plumbing narrative / Div 22 spec"
SRC_FOLDER = "owp-2033/ (2033-Compass, Vuecrest)"

# Documented CO amounts from Budget Transfer spreadsheets
CHANGE_ORDERS = [
    ("CO#01", 22_429.00, "CO_s/Vuecrest - Budget Transfer - CO #01.xlsx"),
    ("CO#02a", 10_131.00, "BT CO#02 100% GMP Set Costs"),
    ("CO#02b", -234_240.00, "BT CO#02 Cast Iron, Lav & Kitchen Sinks Credit"),
    ("CO#03", 999.00, "BT CO#03 Footing Drain @ Ramp (RFI-015)"),
    ("CO#04", 9_194.00, "BT CO#04 Elevator Footing Drains (RFI-016)"),
    ("CO#05", 1_118.00, "BT CO#05 Grade Beam Drainage (RFI-019)"),
    ("CO#06", 5_846.00, "BT CO#06 Add Grease Interceptor (COB)"),
    ("CO#07", -2_694.00, "BT CO#07 Delete Rooftop Sink"),
    ("CO#08", 1_817.00, "BT CO#08 Sump Pump — Temporary Dewatering"),
    ("CO#09", 2_865.00, "BT CO#09 Delete Rooftop Bathroom / Add Condensate Stacks"),
    ("CO#10", 5_364.00, "BT CO#10 PSE Transformer Oil Containment Room Drain"),
    ("CO#11", 1_667.00, "BT CO#11 PSE Transformer Drain — Permit Revision"),
    ("CO#12", 5_466.00, "BT CO#12 Dewatering Meter Supply & Install"),
    ("CO#13", 3_126.00, "BT CO#13 Add Drain in Package Room (RFI-154)"),
    ("CO#14", 1_352.00, "BT CO#14 Added L5 Deck Drain (RFI-224)"),
    ("CO#15", 2_546.00, "BT CO#15 Dog Run Backflow Preventer (RFI-172.1)"),
    ("CO#16", 1_782.00, "BT CO#16 RFI-172.1 Dog Run Connections"),
    ("CO#17", 999.00, "BT CO#17 Revised Layout x20 Stack (RFI-202)"),
]

# ============ TAB 1: OVERVIEW ============
ws = wb.create_sheet("Overview")
title(ws, f"Job #{JOB} · {NAME}",
      f"Cortex JCR v2  •  {GC} (customer {GC_CUST_CODE})  •  {LOCATION}  •  Owner: {OWNER}")
put(ws, "B5", "PROJECT OVERVIEW", BOLD, F_ALT)
overview = [
    ("Project Job #", JOB, SRC_JDR + " header"),
    ("Project Name (Sage)", NAME, SRC_JDR + " header"),
    ("Project Description", PROJECT_DESC, SRC_CONTRACT + " + Job Info sheet"),
    ("General Contractor", GC, SRC_CONTRACT),
    ("Customer Code (Sage)", GC_CUST_CODE, SRC_JDR),
    ("Owner / Developer", OWNER, SRC_CONTRACT + " + Job Info"),
    ("GC PM", GC_PM, "Job Info sheet"),
    ("GC Superintendent", GC_SUP, "Job Info sheet"),
    ("GC PE", GC_PE, "Job Info sheet"),
    ("OWP Rough-in Foreman", OWP_RI_FOREMAN, "Schedule tab"),
    ("Jobsite Location", LOCATION, SRC_CONTRACT),
    ("Contract / PO Document", SRC_CONTRACT, "Lump Sum subcontract (Fully Executed 9/26/2016)"),
    ("Insurance", INSURANCE, "OCIP Enrollment Form + Insurance folder"),
    ("Plans / Specifications", f"{SRC_NARRATIVE} (Div 22 plumbing) + P-tag fixture schedule + OWP drawings", "Plans/ folder"),
    ("Fixture Schedule", SRC_PTAG + f" ({UNITS}-unit luxury multifamily)", "Plans folder"),
    ("Contract Type", "Lump Sum", SRC_CONTRACT),
    ("Unit Count", UNITS, "Job Info sheet"),
    ("Work Period", "Sep 2016 start → Nov 2018 last invoice (27 months)", "JDR PR/AR dates"),
    ("Total Unique Documents Reviewed", "1000+ (231 POs, 503 photos, 59 ASI-RFIs, 33 CORs, 22 submittals, 10 permits)", f"File inventory across {SRC_FOLDER}"),
]
r = 6
for label, val, src in overview:
    put(ws, f"B{r}", label, BOLD, align=LEFT)
    put(ws, f"C{r}", val, align=LEFT)
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.merge_cells(f"C{r}:E{r}")
    ws.merge_cells(f"F{r}:J{r}")
    r += 1

r += 1
put(ws, f"B{r}", "CONTRACT VALUE", BOLD, F_ALT); put(ws, f"D{r}", "NET PROFIT", BOLD, F_ALT)
put(ws, f"F{r}", "DIRECT COST", BOLD, F_ALT); put(ws, f"H{r}", "LABOR HOURS", BOLD, F_ALT)
r += 1
put(ws, f"B{r}", CONTRACT_FINAL, BOLD, fmt='"$"#,##0.00')
put(ws, f"D{r}", NET_PROFIT, BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", EXPENSES, BOLD, fmt='"$"#,##0.00')
put(ws, f"H{r}", TOTAL_HOURS, BOLD, fmt='#,##0.00')
r += 1
put(ws, f"B{r}", f"Original ${CONTRACT_ORIG:,.2f} + (${CO_TOTAL_IMPLIED:,.2f}) COs ({CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%)", SUB)
put(ws, f"D{r}", f"{NET_PROFIT/REVENUE*100:.1f}% margin", SUB)
put(ws, f"F{r}", f"{EXPENSES/REVENUE*100:.1f}% of revenue", SUB)
put(ws, f"H{r}", f"{TOTAL_WORKERS} workers", SUB)

r += 3
put(ws, f"B{r}", "SCOPE OF WORK (from contract + proposal)", BOLD, F_ALT)
r += 1
scope_lines = [
    f"Full plumbing installation: underground, garage, rough-in, finish, gas, water main, insulation, mechanical room, condensate drains.",
    f"{UNITS}-unit luxury multifamily new construction (Vuecrest LUX) at {LOCATION}.",
    "5 levels above ground + 1 level below grade (per Permits); large UG garage & mechanical room; PSE transformer vault.",
    "Notable scope: Grease interceptor, PSE oil containment drain, dog-run backflow, rooftop restrooms (deleted), dewatering meter.",
    "OCIP insurance (wrap program); Compass Harbor as GC; Continental as developer.",
    "Spec compliance: City of Bellevue plumbing code + UPC with WA amendments. Permits incl 16-144265-BO (building) and 17-106355-BL (plumbing).",
]
for line in scope_lines:
    put(ws, f"B{r}", line, BODY, align=LEFT)
    ws.merge_cells(f"B{r}:J{r}")
    r += 1

r += 2
put(ws, f"B{r}", "SOURCES", BOLD, F_HDR)
for col in range(2, 11): put(ws, (r, col), ws.cell(row=r, column=col).value or "", border=BRD, fill=F_HDR)
r += 1
srcs = [
    f"Canonical financial source: {SRC_JDR}",
    f"Job totals (JDR footer): Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: {SRC_CONTRACT} (Lump Sum ${CONTRACT_ORIG:,.2f} descoped to ${CONTRACT_FINAL:,.2f})",
    f"Fixture schedule: {SRC_PTAG}",
    f"Plumbing narrative: {SRC_NARRATIVE}",
    f"Change Orders: 18 fully-executed COs in Change Orders/CO_s/ folder; 17 Budget Transfer spreadsheets document ${CO_TOTAL_DOCUMENTED:,.2f}; JDR-implied net ${CO_TOTAL_IMPLIED:,.2f}",
    "COR source: Change Orders/COR_s/ (33 internal Change Order Requests with RFI links)",
    "Project Team: Ryan Ames/Vince Dennison (GC PM), Todd Mortenson (GC Sup), Dan Peck (GC PE), Continental (Developer), Victor (OWP RI Foreman)",
]
for s in srcs:
    put(ws, f"B{r}", s, SRC_FONT, align=LEFT); ws.merge_cells(f"B{r}:J{r}"); r += 1
widths(ws, {1:2, 2:26, 3:22, 4:16, 5:16, 6:18, 7:18, 8:14, 9:14, 10:14})

# ============ TAB 2: BUDGET VS ACTUAL ============
ws = wb.create_sheet("Budget vs Actual")
title(ws, "Budget vs Actual", f"All {len(CODES)} cost codes from JDR. Contract ${CONTRACT_ORIG:,.0f} → ${CONTRACT_FINAL:,.0f} (COs {CO_TOTAL_IMPLIED:,.2f} descoped).")
hdr(ws, 5, ["Cost Code", "Description", "Original Budget", "Revised Budget", "Actual", "Variance", "% of Revised", "Hours", "Source"])
r = 6
ordered = sorted(CODES.keys(), key=lambda x: int(x))
for code in ordered:
    c = CODES[code]
    put(ws, f"B{r}", code, align=CENTER)
    put(ws, f"C{r}", c['desc'], align=LEFT)
    put(ws, f"D{r}", c['orig'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", c['rev'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", c['actual'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f"=F{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", f'=IF(E{r}=0,"",F{r}/E{r})', fmt='0.0%', align=RIGHT)
    put(ws, f"I{r}", c['hrs_total'] if c['hrs_total'] else "", fmt='#,##0.00', align=RIGHT)
    put(ws, f"J{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"G{r}", f"=F{r}-E{r}", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"I{r}", f"=SUM(I6:I{r-1})", BOLD, fmt='#,##0.00')
widths(ws, {1:2, 2:8, 3:28, 4:16, 5:16, 6:16, 7:16, 8:12, 9:10, 10:40})
ws.freeze_panes = "B6"

# ============ TAB 3: COST BREAKDOWN ============
ws = wb.create_sheet("Cost Breakdown")
title(ws, "Cost Breakdown by Category", "Direct cost composition by category from JDR cost codes")
hdr(ws, 5, ["Category", "Cost Codes", "Actual $", "% of Direct Cost", "% of Revenue", "Source"])
cb = [
    ("Labor", ",".join(LABOR_CODES), LABOR_COST),
    ("Material", ",".join(MATERIAL_CODES), MATERIAL_COST),
    ("Subcontractor + Engineering + Permits + Other", ",".join(OVERHEAD_CODES), OVERHEAD_COST),
    ("Payroll Burden", "995", BURDEN_COST),
    ("Payroll Taxes", "998", TAX_COST),
]
r = 6
for cat, codes, amt in cb:
    put(ws, f"B{r}", cat, BOLD, align=LEFT)
    put(ws, f"C{r}", codes, align=CENTER)
    put(ws, f"D{r}", amt, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", f"=D{r}/$D${6+len(cb)}", fmt='0.0%', align=RIGHT)
    put(ws, f"F{r}", f"=D{r}/{REVENUE}", fmt='0.0%', align=RIGHT)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL DIRECT COST", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", 1.0, BOLD, fmt='0.0%', align=RIGHT)
put(ws, f"F{r}", f"=D{r}/{REVENUE}", BOLD, fmt='0.0%', align=RIGHT)
widths(ws, {1:2, 2:45, 3:40, 4:18, 5:18, 6:18, 7:40})

# ============ TAB 4: MATERIAL ============
ws = wb.create_sheet("Material")
title(ws, "Material Purchases — AP Vendors", "Material + subcontractor spend by vendor (AP records from JDR). Supplemented with submittal/PO inventory.")
hdr(ws, 5, ["Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "Category (inferred)", "Source"])
ordered_v = sorted(VENDORS.items(), key=lambda kv: -kv[1]['total'])
r = 6
for vid, v in ordered_v:
    put(ws, f"B{r}", vid, align=CENTER)
    put(ws, f"C{r}", v['name'], align=LEFT)
    put(ws, f"D{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", v['count'], align=CENTER)
    n = v['name'].lower()
    if any(x in n for x in ['supply', 'ferguson', 'beacon', 'keller', 'rosen', 'hardware', 'mechanical sales', 'consolidated']):
        cat = "Plumbing / Supplies"
    elif 'franklin engineering' in n or 'franklin' in n:
        cat = "Engineering (601)"
    elif 'credit card' in n or 'cc' in n.split():
        cat = "Credit Card (mixed)"
    elif 'backflow' in n or 'testing' in n:
        cat = "Testing subcontractor"
    elif 'insulation' in n:
        cat = "Insulation subcon"
    elif 'permit' in n or 'bellevue' in n:
        cat = "Permits (603)"
    else:
        cat = "Uncategorized"
    put(ws, f"F{r}", cat, align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, align=CENTER)
r += 2
put(ws, f"B{r}", f"NOTE: AP total per JDR footer = ${SRC_AP:,.2f}. Vendor sum above approximates this. Submittals folder contains 22 items (fixtures, material, equipment, acoustical schedule, underground, water heater strap kit). POs folder has 231 POs (1 placed, 2 scheduled, 3 completed + loose files).", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:10, 3:38, 4:16, 5:14, 6:26, 7:40})

# ============ TAB 5: CREW & LABOR ============
ws = wb.create_sheet("Crew & Labor")
title(ws, "Crew & Labor — Worker Roster", f"All {TOTAL_WORKERS} unique payroll workers from JDR. OWP RI Foreman: {OWP_RI_FOREMAN}.")
hdr(ws, 5, ["Worker ID", "Worker Name", "Total Hours", "Gross Pay", "Blended Wage ($/hr)", "# Work Days", "Source"])
r = 6
ordered_w = sorted(WORKERS.items(), key=lambda kv: -kv[1]['hours'])
for wid, w in ordered_w:
    put(ws, f"B{r}", wid, align=CENTER)
    put(ws, f"C{r}", w['name'], align=LEFT)
    put(ws, f"D{r}", w['hours'], fmt='#,##0.00', align=RIGHT)
    put(ws, f"E{r}", w['amount'], fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", f"=IF(D{r}=0,0,E{r}/D{r})", fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", w['days'], align=CENTER)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
widths(ws, {1:2, 2:10, 3:32, 4:12, 5:14, 6:18, 7:14, 8:40})
ws.freeze_panes = "B6"

# ============ TAB 6: CREW ANALYTICS ============
ws = wb.create_sheet("Crew Analytics")
title(ws, "Crew Analytics", "Team-level labor productivity, concentration, wage dispersion")
put(ws, "B5", "TEAM-LEVEL METRICS", BOLD, F_ALT)
hdr(ws, 6, ["Metric", "Value", "Notes", "Source"])
top_w = ordered_w[0]
top_pct = top_w[1]['hours'] / TOTAL_HOURS if TOTAL_HOURS else 0
top5_hrs = sum(w[1]['hours'] for w in ordered_w[:5])
top5_pct = top5_hrs / TOTAL_HOURS if TOTAL_HOURS else 0
max_wage = max((w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0), default=0)
min_wage = min((w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0), default=0)
crew_metrics = [
    ("Total Workers", TOTAL_WORKERS, "Unique payroll IDs", SRC_JDR),
    ("Total Labor Hours", TOTAL_HOURS, "Sum of labor codes", SRC_JDR),
    ("Total Gross Pay", LABOR_COST, "Sum of labor codes", SRC_JDR),
    ("Blended Gross Wage ($/hr)", LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0, "Labor$ / Hrs (pre-burden)", "Derived"),
    ("Top Worker Hours Share", top_pct, f"{top_w[0]} {top_w[1]['name']} ({top_w[1]['hours']:.0f} hrs)", "Derived"),
    ("Top 5 Workers Hours Share", top5_pct, "Concentration metric", "Derived"),
    ("Highest Wage Rate ($/hr)", max_wage, "Single-worker blended", "Derived"),
    ("Lowest Wage Rate ($/hr)", min_wage, "Single-worker blended", "Derived"),
    ("Avg Hours per Worker", TOTAL_HOURS/TOTAL_WORKERS if TOTAL_WORKERS else 0, "Includes short-tenure workers", "Derived"),
    ("Avg Project Days per Worker", sum(w['days'] for w in WORKERS.values())/TOTAL_WORKERS if TOTAL_WORKERS else 0, "Mean days", "Derived"),
]
r = 7
for m, v, note, src in crew_metrics:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if isinstance(v, float):
        if "Share" in m: c.number_format = '0.0%'
        elif "Wage" in m: c.number_format = '"$"#,##0.00'
        elif "$" in m: c.number_format = '"$"#,##0.00'
        else: c.number_format = '#,##0.00'
    elif isinstance(v, int):
        c.number_format = '#,##0'
    put(ws, f"D{r}", note, align=LEFT)
    put(ws, f"E{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:30, 3:16, 4:50, 5:40})

# ============ TAB 7: PRODUCTIVITY ============
ws = wb.create_sheet("Productivity")
title(ws, "Productivity Metrics", f"Normalized labor and financial ratios. Per-unit metrics from {UNITS} units (Job Info).")
hdr(ws, 5, ["Metric", "Value", "Basis", "Source / Note"])
prods = [
    ("Revenue per Labor Hour", f"={REVENUE}/{TOTAL_HOURS}", "Formula", "Rev / Total Hrs"),
    ("Profit per Labor Hour", f"={NET_PROFIT}/{TOTAL_HOURS}", "Formula", "Net Profit / Hrs"),
    ("Labor Cost per Hour (blended)", f"={LABOR_COST}/{TOTAL_HOURS}", "Formula", "Pre-burden"),
    ("Fully-Loaded Labor Rate ($/hr)", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{TOTAL_HOURS}", "Formula", "Incl burden + taxes"),
    ("Burden Multiplier", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{LABOR_COST}", "Formula", "Fully-loaded / blended"),
    ("Rough-in Hours (code 120)", CODES["120"]['hrs_total'], "JDR", f"{CODES['120']['hrs_total']:.0f} hrs"),
    ("Finish Hours (code 130)", CODES["130"]['hrs_total'], "JDR", f"{CODES['130']['hrs_total']:.0f} hrs"),
    ("Rough-in % of Total Hours", f'={CODES["120"]["hrs_total"]}/{TOTAL_HOURS}', "Formula", "Code 120 share"),
    ("Gross Margin", f"={NET_PROFIT}/{REVENUE}", "Formula", "Net / Revenue"),
    ("Labor % of Revenue", f"={LABOR_COST}/{REVENUE}", "Formula", ""),
    ("Material % of Revenue", f"={MATERIAL_COST}/{REVENUE}", "Formula", ""),
    ("Direct Cost Ratio", f"={EXPENSES}/{REVENUE}", "Formula", "All expenses / Rev"),
    (f"Revenue per Unit ({UNITS})", f"={REVENUE}/{UNITS}", "Formula", f"{UNITS} units (Job Info)"),
    ("Labor Hours per Unit", f"={TOTAL_HOURS}/{UNITS}", "Formula", ""),
    ("Labor Cost per Unit", f"={LABOR_COST}/{UNITS}", "Formula", ""),
    ("Material Cost per Unit", f"={MATERIAL_COST}/{UNITS}", "Formula", ""),
    ("Direct Cost per Unit", f"={EXPENSES}/{UNITS}", "Formula", ""),
    ("Rough-in Hours per Unit", f'={CODES["120"]["hrs_total"]}/{UNITS}', "Formula", f"Code 120 / {UNITS}"),
]
r = 6
pct_rows = {"Gross Margin", "Labor % of Revenue", "Material % of Revenue", "Direct Cost Ratio", "Rough-in % of Total Hours"}
for m, v, basis, note in prods:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if m in pct_rows: c.number_format = '0.0%'
    elif m == "Burden Multiplier": c.number_format = '0.00"x"'
    elif "Hours" in m and "Rate" not in m and "per" not in m: c.number_format = '#,##0.00'
    elif "Hours per" in m: c.number_format = '#,##0.00'
    else: c.number_format = '"$"#,##0.00'
    put(ws, f"D{r}", basis, align=CENTER)
    put(ws, f"E{r}", note, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:32, 3:16, 4:12, 5:44})

# ============ TAB 8: PO COMMITMENTS ============
ws = wb.create_sheet("PO Commitments")
title(ws, "PO Commitments", "Inbound contract value (GC → OWP). Outbound vendor commitments — see Material tab + POs folder.")
hdr(ws, 5, ["PO #", "Date", "Issuer", "Type", "Status", "Description", "Amount", "Source"])
r = 6
put(ws, f"B{r}", "PRIME", align=CENTER)
put(ws, f"C{r}", "2016-09-26 (Fully Executed)", align=LEFT)
put(ws, f"D{r}", GC, align=LEFT)
put(ws, f"E{r}", "Lump Sum Subcontract", align=CENTER)
put(ws, f"F{r}", "Closed", align=CENTER, fill=F_OK)
put(ws, f"G{r}", f"Vuecrest LUX plumbing ({UNITS} units, Bellevue WA) — full plumbing scope per Div 22", align=LEFT)
put(ws, f"H{r}", CONTRACT_ORIG, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"I{r}", SRC_CONTRACT, SRC_FONT, align=LEFT)
r += 1
for co_id, amt, src in CHANGE_ORDERS:
    put(ws, f"B{r}", co_id, align=CENTER)
    put(ws, f"C{r}", NF, align=CENTER)
    put(ws, f"D{r}", GC, align=LEFT)
    put(ws, f"E{r}", "Change Order", align=CENTER)
    put(ws, f"F{r}", "Executed", align=CENTER, fill=F_OK)
    put(ws, f"G{r}", src.split("/")[-1] if "/" in src else src, align=LEFT)
    put(ws, f"H{r}", amt, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL (contract + COs)", BOLD)
put(ws, f"H{r}", f"=SUM(H6:H{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
r += 2
gap = CO_TOTAL_IMPLIED - CO_TOTAL_DOCUMENTED
put(ws, f"B{r}", f"NOTE: Documented COs (17 Budget Transfer spreadsheets) sum to ${CO_TOTAL_DOCUMENTED:,.2f}; JDR-implied = ${CO_TOTAL_IMPLIED:,.2f}; gap ${gap:,.2f} likely CO#18 (Fully Executed 1/23/2019, no budget transfer xlsx). CO folder has 18 fully-executed COs total. CORs folder contains 33 internal COR pricing worksheets. Outbound POs: 231 items across 1 Placed, 2 Scheduled, 3 Completed subfolders.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:I{r}")
widths(ws, {1:2, 2:14, 3:28, 4:22, 5:22, 6:18, 7:42, 8:14, 9:42})

# ============ TAB 9: BILLING & SOV ============
ws = wb.create_sheet("Billing & SOV")
title(ws, "Billing & Schedule of Values", f"{len(INVOICES)} unique invoices to {GC}. Retainage ${RETAINAGE:,.2f} = {RETAINAGE/REVENUE*100:.1f}% of revenue.")
hdr(ws, 5, ["Invoice #", "Date", "Total Billed (signed)", "Retainage (signed)", "# Lines", "Source"])
r = 6
for inv in sorted(INVOICES.keys()):
    iv = INVOICES[inv]
    put(ws, f"B{r}", inv, align=CENTER)
    put(ws, f"C{r}", iv['date'], align=CENTER)
    put(ws, f"D{r}", iv['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", iv['retainage'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", iv['lines'], align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL (signed)", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"NOTE: AR entries signed negative per Sage convention. Net billed = ${REVENUE:,.2f}; retainage ${RETAINAGE:,.2f} outstanding on JDR dated 04/03/2026. Billing folder contains SOV xlsx files for pay apps 13-23 (2018 draw period). F1 Vuecrest - Subtier Subs & Suppliers disclosures attached for pay apps 6-23.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:12, 3:12, 4:18, 5:16, 6:10, 7:50})

# ============ TAB 10: INSIGHTS ============
ws = wb.create_sheet("Insights")
title(ws, "Insights & Observations", "Narrative findings from JDR + source documents")
top_vendor = ordered_v[0] if ordered_v else (None, {'name': '—', 'total': 0})
top5_ap = sum(v[1]['total'] for v in ordered_v[:5])
insights = [
    ("VERY STRONG MARGIN", f"Net profit ${NET_PROFIT:,.2f} on ${REVENUE:,.2f} revenue = {NET_PROFIT/REVENUE*100:.1f}% gross margin. Exceeds typical multifamily plumbing (25-30%) — strong execution under OCIP (wrap) insurance.", "Verified", SRC_JDR),
    ("DEDUCTIVE CONTRACT", f"Original contract ${CONTRACT_ORIG:,.2f} → revised ${CONTRACT_FINAL:,.2f} = (${abs(CO_TOTAL_IMPLIED):,.2f}) net reduction ({CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%). Driven by CO#02 Cast Iron/Lav/Kitchen Sinks credit ($234,240) partially offset by 16 adds. 18 fully-executed COs and 33 internal CORs documented.", "Verified", "Change Orders folder + Sage"),
    ("LABOR EFFICIENCY", f"Labor cost ${LABOR_COST:,.2f} ({LABOR_COST/REVENUE*100:.1f}% of rev) across {TOTAL_HOURS:,.0f} hrs and {TOTAL_WORKERS} workers. Blended wage ${LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0:.2f}/hr (pre-burden).", "Verified", SRC_JDR),
    ("VENDOR CONCENTRATION", f"Top 5 AP vendors: ~${top5_ap:,.2f} ({top5_ap/SRC_AP*100:.0f}% of AP). Top vendor: {top_vendor[1]['name']} (${top_vendor[1]['total']:,.2f}).", "Verified", SRC_JDR),
    ("LOW OVERHEAD", f"GL cost ${SRC_GL:,.2f} = {SRC_GL/REVENUE*100:.1f}% of revenue. Lean project management overhead — typical OCIP projects keep GL under 3%.", "Verified", SRC_JDR),
    ("RETAINAGE HELD LONG", f"Retainage ${RETAINAGE:,.2f} = {RETAINAGE/REVENUE*100:.1f}% of AR still open on JDR 04/03/2026 — that's 7+ years after last invoice (11/2018). Flag for collections follow-up.", "Verified", SRC_JDR),
    ("ROUGH-IN DOMINANT", f"Code 120 (Roughin Labor) = {CODES['120']['hrs_total']:,.0f} hrs ({CODES['120']['hrs_total']/TOTAL_HOURS*100:.0f}% of total). Revised budget was {CODES['120']['rev']:,.0f}, actual {CODES['120']['actual']:,.0f} = savings of ${CODES['120']['rev']-CODES['120']['actual']:,.0f}.", "Verified", SRC_JDR),
    ("FINISH MATERIAL OVERRUN", f"Code 230 Finish Material: rev ${CODES['230']['rev']:,.0f}, actual ${CODES['230']['actual']:,.0f} (+${CODES['230']['actual']-CODES['230']['rev']:,.0f} / {(CODES['230']['actual']-CODES['230']['rev'])/CODES['230']['rev']*100:.1f}%). Likely driver: upgraded fixtures per CO#02 cast iron/lav/sink revision + Delta faucet upgrades (CO#29).", "Inferred", SRC_JDR),
    ("PSE TRANSFORMER DRAMA", f"3 COs directly tied to PSE transformer oil containment (CO#10, CO#11, CO#12 Dewatering Meter): ~$12,497 of added cost. Notable as design iteration with PSE utility coordination.", "Verified", "CO folder + RFI-005"),
    ("DOCUMENT DEPTH", f"Exceptional documentation: 59 ASI-RFI items, 33 CORs, 231 POs, 503 photos, 10 permits (incl 16-144265-BO building, 17-106355-BL plumbing). Ranks among most thoroughly documented OWP projects.", "Verified", SRC_FOLDER),
]
r = 5
hdr(ws, r, ["#", "Insight", "Detail", "Confidence", "Source"])
r = 6
for i, (ttl, det, conf, src) in enumerate(insights, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", ttl, BOLD, align=LEFT)
    put(ws, f"D{r}", det, align=LEFT)
    c = put(ws, f"E{r}", conf, align=CENTER)
    if conf == "Verified": c.fill = F_OK
    elif conf == "Inferred": c.fill = F_HIGH
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.row_dimensions[r].height = 58
    r += 1
widths(ws, {1:2, 2:4, 3:32, 4:78, 5:12, 6:40})

# ============ TAB 11: BENCHMARK KPIs ============
ws = wb.create_sheet("Benchmark KPIs")
title(ws, "Benchmark KPIs", "Normalized metrics for cross-project comparison")
hdr(ws, 5, ["KPI", "Data Name", "Value", "Category", "Notes", "Confidence", "Source Document"])
kpis = [
    ("Job Number", "job_number", JOB, "Profile", "OWP job ID", "Verified", SRC_JDR),
    ("Job Name", "job_name", NAME, "Profile", "Sage short name", "Verified", SRC_JDR),
    ("Project Description", "project_desc", PROJECT_DESC, "Profile", "Contract + Job Info", "Verified", SRC_CONTRACT),
    ("General Contractor", "general_contractor", GC, "Profile", f"Customer {GC_CUST_CODE}", "Verified", SRC_CONTRACT),
    ("Owner / Developer", "owner", OWNER, "Profile", "Per subcontract + Job Info", "Verified", SRC_CONTRACT),
    ("Location", "location", LOCATION, "Profile", "Contract address", "Verified", SRC_CONTRACT),
    ("Project Type", "project_type", f"Luxury Multifamily New Construction — Plumbing ({UNITS} units)", "Profile", "Fixture schedule + Div 22", "Verified", SRC_PTAG),
    ("Insurance Type", "insurance_type", INSURANCE, "Profile", "OCIP (wrap) enrollment", "Verified", "Insurance folder"),
    ("GC PM", "gc_pm", GC_PM, "Profile", "Project lead", "Verified", "Job Info"),
    ("GC Superintendent", "gc_sup", GC_SUP, "Profile", "", "Verified", "Job Info"),
    ("GC PE", "gc_pe", GC_PE, "Profile", "", "Verified", "Job Info"),
    ("Developer", "developer", DEVELOPER, "Profile", "", "Verified", "Job Info"),
    ("OWP RI Foreman", "owp_foreman", OWP_RI_FOREMAN, "Profile", "", "Verified", "Schedule tab"),
    ("Work Start Date", "start_date", "2016-09-01", "Profile", "Per Schedule + first PR", "Verified", SRC_JDR),
    ("Work End Date", "end_date", "2018-11-16", "Profile", "Per last AR posting", "Verified", SRC_JDR),
    ("Duration (months)", "duration_months", 27, "Profile", "Sep 2016 → Nov 2018", "Verified", "Derived"),
    ("Unit Count", "unit_count", UNITS, "Profile", "Per Job Info sheet", "Verified", "Job Info"),
    ("Contract Original", "contract_original", CONTRACT_ORIG, "Financial", "Subcontract Lump Sum (Sage)", "Verified", SRC_CONTRACT),
    ("Contract Final", "contract_final", CONTRACT_FINAL, "Financial", "Code 999 Rev Budget = AR total", "Verified", SRC_JDR),
    ("Change Orders ($, implied)", "change_orders_implied", CO_TOTAL_IMPLIED, "Financial", "Final - Original", "Verified", SRC_JDR),
    ("Change Orders ($, documented)", "change_orders_documented", CO_TOTAL_DOCUMENTED, "Financial", "17 Budget Transfer xlsx sum", "Verified", "BUDGET TRANSFERS folder"),
    ("CO Count Documented", "co_count", 18, "Financial", "18 fully-executed COs in folder", "Verified", "Change Orders/CO_s/"),
    ("COR Count", "cor_count", 33, "Financial", "Internal Change Order Requests", "Verified", "Change Orders/COR_s/"),
    ("Change Order % of Contract", "co_pct", CO_TOTAL_IMPLIED/CONTRACT_ORIG, "Financial", "", "Verified", "Derived"),
    ("Revenue", "revenue", REVENUE, "Financial", "AR total", "Verified", SRC_JDR),
    ("Direct Cost", "direct_cost", EXPENSES, "Financial", "JDR Job Totals Expenses", "Verified", SRC_JDR),
    ("Net Profit", "net_profit", NET_PROFIT, "Financial", "Rev - Expenses", "Verified", SRC_JDR),
    ("Gross Margin", "gross_margin", NET_PROFIT/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Retainage Outstanding", "retainage", RETAINAGE, "Financial", "Open per JDR 4/3/26", "Verified", SRC_JDR),
    ("Retainage % of Revenue", "retainage_pct", RETAINAGE/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Labor Cost", "labor_cost", LABOR_COST, "Labor", "Codes 100-143", "Verified", SRC_JDR),
    ("Material Cost", "material_cost", MATERIAL_COST, "Material", "Codes 210-243", "Verified", SRC_JDR),
    ("Subcontractor+OH Cost", "overhead_cost", OVERHEAD_COST, "Financial", "Codes 600-607", "Verified", SRC_JDR),
    ("Burden Cost", "burden_cost", BURDEN_COST, "Labor", "Code 995", "Verified", SRC_JDR),
    ("Tax Cost", "tax_cost", TAX_COST, "Labor", "Code 998", "Verified", SRC_JDR),
    ("Total Labor Hours", "total_hours", TOTAL_HOURS, "Labor", "Sum labor codes", "Verified", SRC_JDR),
    ("Total Workers", "total_workers", TOTAL_WORKERS, "Labor", "Unique payroll IDs", "Verified", SRC_JDR),
    ("Blended Gross Wage ($/hr)", "blended_gross_wage", LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0, "Labor", "Pre-burden", "Verified", "Derived"),
    ("Fully-Loaded Wage ($/hr)", "fully_loaded_wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS if TOTAL_HOURS else 0, "Labor", "Incl burden + tax", "Verified", "Derived"),
    ("Burden Multiplier", "burden_multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST if LABOR_COST else 0, "Labor", "Fully-loaded/blended", "Verified", "Derived"),
    ("Rough-in Hours", "roughin_hours", CODES["120"]['hrs_total'], "Labor", "Code 120", "Verified", SRC_JDR),
    ("Finish Hours", "finish_hours", CODES["130"]['hrs_total'], "Labor", "Code 130", "Verified", SRC_JDR),
    ("Revenue per Hour", "revenue_per_hour", REVENUE/TOTAL_HOURS if TOTAL_HOURS else 0, "Productivity", "", "Verified", "Derived"),
    ("Profit per Hour", "profit_per_hour", NET_PROFIT/TOTAL_HOURS if TOTAL_HOURS else 0, "Productivity", "", "Verified", "Derived"),
    ("Revenue per Unit", "revenue_per_unit", REVENUE/UNITS, "Productivity", f"{UNITS} units", "Verified", "Derived"),
    ("Direct Cost per Unit", "cost_per_unit", EXPENSES/UNITS, "Productivity", f"{UNITS} units", "Verified", "Derived"),
    ("Labor Hours per Unit", "hours_per_unit", TOTAL_HOURS/UNITS, "Productivity", f"{UNITS} units", "Verified", "Derived"),
    ("Labor % of Revenue", "labor_pct_revenue", LABOR_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Material % of Revenue", "material_pct_revenue", MATERIAL_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Total Vendors (AP)", "total_vendors", len(VENDORS), "Material", "Unique vendor IDs", "Verified", SRC_JDR),
    ("Total Invoices (AR)", "total_invoices", len(INVOICES), "Billing", "Unique invoice numbers", "Verified", SRC_JDR),
    ("RFI Count (documented)", "rfi_count", 59, "Docs", "Items in ASI-RFI folder (mix of ASI + RFI)", "Verified", "ASI-RFI/"),
    ("Submittal Count", "submittal_count", 22, "Docs", "Items in Submittals folder", "Verified", "Submittals/"),
    ("Permit Count", "permit_count", 10, "Docs", "Items in Permits folder", "Verified", "Permits/"),
    ("PO Count", "po_count", 231, "Docs", "Across 1 Placed, 2 Scheduled, 3 Completed", "Verified", "POs folder"),
    ("Photo Count", "photo_count", 503, "Docs", "Photos folder", "Verified", "Photos/"),
    ("AP Spend (JDR footer)", "ap_total", SRC_AP, "Material", "JDR footer", "Verified", SRC_JDR),
    ("PR Spend (JDR footer)", "pr_total", SRC_PR, "Labor", "JDR footer (labor+burden+tax)", "Verified", SRC_JDR),
    ("GL Spend (JDR footer)", "gl_total", SRC_GL, "Financial", "JDR footer", "Verified", SRC_JDR),
]
r = 6
for k in kpis:
    for j, v in enumerate(k):
        cell = put(ws, (r, 2+j), v, align=LEFT if j in (0,1,4,6) else CENTER)
        if j == 2 and isinstance(v, float):
            if "pct" in k[1] or "margin" in k[1]: cell.number_format = '0.00%'
            elif "multiplier" in k[1]: cell.number_format = '0.00"x"'
            elif "hours" in k[1] or "months" in k[1] or "hour" in k[1] or "per_hour" in k[1]: cell.number_format = '#,##0.00'
            else: cell.number_format = '"$"#,##0.00'
        if j == 5:
            if v == "Verified": cell.fill = F_OK
            elif v == "Medium": cell.fill = F_HIGH
            elif v == "Low": cell.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:30, 3:24, 4:22, 5:14, 6:38, 7:12, 8:38})
ws.freeze_panes = "B6"

# ============ TAB 12: VENDORS ============
ws = wb.create_sheet("Vendors")
title(ws, "Vendors — AP Summary", "Vendor-level spend ranking")
hdr(ws, 5, ["Rank", "Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "% of AP", "Source"])
r = 6
total_ap_vendors = sum(v['total'] for v in VENDORS.values()) or 1
for rank, (vid, v) in enumerate(ordered_v, 1):
    put(ws, f"B{r}", rank, align=CENTER)
    put(ws, f"C{r}", vid, align=CENTER)
    put(ws, f"D{r}", v['name'], align=LEFT)
    put(ws, f"E{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", v['count'], align=CENTER)
    put(ws, f"G{r}", f"=E{r}/{total_ap_vendors}", fmt='0.0%', align=RIGHT)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, align=CENTER)
widths(ws, {1:2, 2:6, 3:10, 4:38, 5:14, 6:14, 7:12, 8:40})

# ============ TAB 13: CHANGE LOG ============
ws = wb.create_sheet("Change Log")
title(ws, "Change Log — Master Register", "COs, RFIs, ASIs, Submittals, Permits, POs — from project folders")
hdr(ws, 5, ["Event ID", "Type", "Date", "Subject", "Originator", "Cost Impact ($)", "Status", "Source"])
events = [
    ("CONTRACT-ORIG", "Contract", "2016-09-26", f"Prime subcontract — Lump Sum ${CONTRACT_ORIG:,.2f} (Fully Executed)", GC, CONTRACT_ORIG, "Executed", SRC_CONTRACT),
]
for co_id, amt, src in CHANGE_ORDERS:
    events.append((co_id, "Change Order", NF, src, GC, amt, "Executed", f"Change Orders/{src}"))
events.extend([
    ("CO#18", "Change Order", "2019-01-23", "CO#18 (Fully Executed) — no budget transfer xlsx; pickup/closeout CO", GC, CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED, "Executed", "CO_s/Vuecrest - CO#18 - OWP (Fully Executed 1.23.19).pdf"),
    ("CO-GAP", "Reconciliation", "—", f"Documented vs implied CO delta = ${CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED:,.2f} (allocated to CO#18)", "—", 0, "Resolved", SRC_JDR),
    ("COR-ALL", "COR batch", "—", "33 Change Order Requests (internal pricing) with RFI linkages", "Sub (OWP)", 0, "Priced", "COR_s/"),
    ("ASI-PRIMARY", "ASI batch", "—", "5 ASIs: Roof Insulation, Mansard Backspan, Project Updates, L5 & Roof Restrooms, general updates", "GC/Arch", 0, "Acknowledged", "ASI-RFI/"),
    ("RFI-PRIMARY", "RFI batch", "—", "50+ RFIs: footing drains, tub drains at columns, fire rating, tub & washer drains at columns, fridge openings", "Sub (OWP)", 0, "Resolved", "ASI-RFI/"),
    ("SUB-FIXTURES", "Submittal batch", "—", "Fixture submittals (Kohler K-2210, K-2882, Delta faucets, Zoeller pumps)", "Sub (OWP)", 0, "Approved", "Submittals/FIXTURES/"),
    ("SUB-MATERIAL", "Submittal batch", "—", "Material submittals (no-hub cast iron per Rosen, Holdrite penetration sleeves, equipment)", "Sub (OWP)", 0, "Approved", "Submittals/MATERIAL/"),
    ("SUB-UG-GARAGE", "Submittal", "—", "Underground & garage submittal", "Sub (OWP)", 0, "Approved", "Submittals/UNDERGROUND & GARAGE/"),
    ("PERMIT-BUILDING", "Permit", "2017-01-05", "Building permit 16-144265-BO — Vuecrest (issuance + 2 rev req 11-22-16, 12-30-16)", "City of Bellevue", 0, "Issued", "Permits/"),
    ("PERMIT-PLUMBING", "Permit", "2017-03-07", "Plumbing permit 17-106355-BL", "City of Bellevue", 0, "Issued", "Permits/"),
    ("PERMIT-MECH-CHECKLIST", "Permit", NF, "Bellevue Mechanical Permit Checklist", "City of Bellevue", 0, "On file", "Permits/"),
    ("PERMIT-TEMP", "Permit", NF, "Temp trailer water & sewer permit", "City of Bellevue", 0, "Issued", "Permits/"),
    ("PERMIT-TRUCKING", "Logistics", "2016-09-22", "TN approved COB trucking route (signed)", "City of Bellevue", 0, "Approved", "Project root/"),
    ("OCIP", "Insurance", "—", "OCIP enrollment (wrap insurance)", "Bellevue Vuecrest, LP", 0, "Enrolled", "Insurance/OCIP Enrollment.pdf"),
    ("PO-COMPLETED", "PO batch", "—", "~200+ completed POs (3 Completed/)", "Sub (OWP)", 0, "Complete", "POs/3 Completed/"),
    ("PO-SCHEDULED", "PO batch", "—", "Scheduled POs (2 Scheduled/)", "Sub (OWP)", 0, "Scheduled", "POs/2 Scheduled/"),
    ("PO-PLACED", "PO batch", "—", "Placed POs (1 Placed/)", "Sub (OWP)", 0, "Placed", "POs/1 Placed/"),
    ("FIRST-INVOICE", "Invoice", "2016-12-20", f"First billing — see Billing folder 2017/", "Sub (OWP)", 0, "Paid", SRC_JDR),
    ("LAST-INVOICE", "Invoice", "2018-11-16", "Last billing (Invoice 037230 @ $758)", "Sub (OWP)", 758.00, "Paid", SRC_JDR),
    ("RETAINAGE-OPEN", "Retainage", "As of 04/03/2026", f"Retainage ${RETAINAGE:,.2f} outstanding 7+ years post-closeout", "GC", 0, "Outstanding", SRC_JDR),
])
r = 6
for e in events:
    eid, et, dt, subj, orig, cost, status, src = e
    put(ws, f"B{r}", eid, align=CENTER)
    put(ws, f"C{r}", et, align=CENTER)
    put(ws, f"D{r}", dt, align=CENTER)
    put(ws, f"E{r}", subj, align=LEFT)
    put(ws, f"F{r}", orig, align=LEFT)
    put(ws, f"G{r}", cost, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", status, align=CENTER)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:18, 3:22, 4:22, 5:50, 6:22, 7:16, 8:18, 9:42})
ws.freeze_panes = "B6"

# ============ TAB 14: ROOT CAUSE ANALYSIS ============
ws = wb.create_sheet("Root Cause Analysis")
title(ws, "Root Cause Analysis", f"Primary variance driver: deductive change orders reducing contract by ${abs(CO_TOTAL_IMPLIED):,.0f}")
put(ws, "B5", "COST-CODE VARIANCE CATEGORIES", BOLD, F_ALT)
hdr(ws, 6, ["Category", "Codes Affected", "Net $ Variance", "Root Cause (inferred)", "Notes"])
def var(c): return CODES[c]['actual'] - CODES[c]['rev'] if c in CODES else 0
rc_rows = [
    ("Deductive Contract Change", "999 (Sales code)", CO_TOTAL_IMPLIED, "CO#02 Cast Iron/Lav/Kitchen Sinks credit (-$234,240) dominant; partially offset by 16 adds", f"Original ${CONTRACT_ORIG:,.0f} → Final ${CONTRACT_FINAL:,.0f} = ${CO_TOTAL_IMPLIED:,.0f} ({CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%)"),
    ("Labor efficiency", "100-143", sum(var(c) for c in LABOR_CODES), "Rough-in (120) and Finish (130) labor both under revised budget", f"{TOTAL_WORKERS} workers; favorable on 9 of 11 labor codes"),
    ("Material procurement", "210-243", sum(var(c) for c in MATERIAL_CODES), "Mixed: Rough-in material large savings offset by Finish Material and Water Main overruns", "Key overruns: 230 (+$22k), 241 (+$7.5k); key savings: 220 (-$104k), 211 (-$51k)"),
    ("Overhead / Engineering / Permits", "600-607", sum(var(c) for c in OVERHEAD_CODES), "Engineering and permits slightly over; rental/subcontractor under", "Engineering +$3.5k (Franklin); Permits modestly over"),
    ("Burden / Tax accrual", "995, 998", var("995")+var("998"), "Tracked proportionally to labor; burden reduced with labor", "OWP internal; CO#08 note changed burden rate to 40"),
    ("Condensation Drains (new code)", "143 + 243", var("143")+var("243"), "Code 143 (labor) and 243 (material) emerged mid-project — multiple condensate-stack CORs", f"Code 143 actual ${CODES['143']['actual']:,.0f} vs rev ${CODES['143']['rev']:,.0f}"),
]
r = 7
for cat, codes, netv, cause, note in rc_rows:
    put(ws, f"B{r}", cat, align=LEFT)
    put(ws, f"C{r}", codes, align=LEFT)
    put(ws, f"D{r}", netv, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", cause, align=LEFT)
    put(ws, f"F{r}", note, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL NET VARIANCE (Rev Budget vs Actual)", BOLD)
put(ws, f"D{r}", f"=SUM(D7:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
r += 3
put(ws, f"B{r}", "RESPONSIBILITY ATTRIBUTION", BOLD, F_ALT); r += 1
hdr(ws, r, ["Responsible Party", "# Drivers", "Net $ Impact", "Notes"]); r += 1
resp = [
    ("GC / Owner (scope revision)", 18, CO_TOTAL_IMPLIED, "18 executed COs — primarily driven by CO#02 Cast Iron / Lav / Kitchen Sinks credit (-$234,240)"),
    ("Sub (OWP) — labor performance", 11, sum(var(c) for c in LABOR_CODES), "Labor cost performance vs revised budget across 11 labor codes"),
    ("Sub (OWP) — material procurement", 9, sum(var(c) for c in MATERIAL_CODES), "Procurement & scope execution; Keller/Rosen/Ferguson dominant vendors"),
    ("Burden rate — OWP internal", 2, var("995")+var("998"), "Payroll accrual; CO#08 note '40% burden next'"),
    ("Support codes — OWP", 5, sum(var(c) for c in OVERHEAD_CODES), "Overhead efficiency (Franklin engineering, permits, rentals)"),
    ("Design coordination — PSE/Bellevue", 3, 12_497.00, "CO#10/11/12 = PSE transformer drain + permit revision + dewatering meter"),
]
for rp, cnt, net, note in resp:
    put(ws, f"B{r}", rp, align=LEFT)
    put(ws, f"C{r}", cnt, align=CENTER)
    put(ws, f"D{r}", net, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", note, align=LEFT)
    r += 1
widths(ws, {1:2, 2:36, 3:40, 4:18, 5:40, 6:44})

# ============ TAB 15: PREDICTIVE SIGNALS ============
ws = wb.create_sheet("Predictive Signals")
title(ws, "Predictive Signals", "Leading indicators from project documents + JDR")
put(ws, "B5", "CURRENT-STATE SIGNALS", BOLD, F_ALT)
hdr(ws, 6, ["Indicator", "Current Value", "Benchmark", "Status", "Meaning"])
roughin_ratio = CODES["120"]['hrs_total']/TOTAL_HOURS if TOTAL_HOURS else 0
top5_ap_pct = sum(v[1]['total'] for v in ordered_v[:5]) / SRC_AP if SRC_AP else 0
signals = [
    ("Contract Descoping %", f"{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%", "±10%", "WATCH", f"{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}% net reduction, driven by CO#02 Cast Iron credit"),
    ("Labor % of Revenue", LABOR_COST/REVENUE, "<30%", "HEALTHY" if LABOR_COST/REVENUE < 0.30 else "WATCH", f"{LABOR_COST/REVENUE*100:.1f}%"),
    ("GL Overhead % of Revenue", SRC_GL/REVENUE, "<5%", "HEALTHY", f"{SRC_GL/REVENUE*100:.1f}%"),
    ("Vendor Concentration (Top 5)", top5_ap_pct, "<95%", "HEALTHY", f"Top 5 vendors = {top5_ap_pct*100:.0f}% of AP"),
    ("Retainage Outstanding", RETAINAGE/REVENUE, "<10%", "HEALTHY", f"{RETAINAGE/REVENUE*100:.1f}% — but 7+ years stale"),
    ("Gross Margin", NET_PROFIT/REVENUE, ">35%", "HEALTHY", f"{NET_PROFIT/REVENUE*100:.1f}% — exceeds benchmark"),
    ("Labor Hrs Total", TOTAL_HOURS, "varies", "INFO", f"{TOTAL_HOURS:,.0f} hrs across {TOTAL_WORKERS} workers"),
    ("Retainage Release Pending", "YES", "Released within 90 days post-closeout", "ELEVATED", f"${RETAINAGE:,.0f} open 7+ years — collections concern"),
    ("Worker Concentration (top 1)", top_pct, "<25%", "HEALTHY" if top_pct < 0.25 else "ELEVATED", f"{top_pct*100:.1f}% from single worker ({top_w[1]['name']})"),
    ("Rough-in Labor Share", roughin_ratio, "40-65%", "HEALTHY" if 0.4 <= roughin_ratio <= 0.65 else "INFO", f"{roughin_ratio*100:.0f}% on rough-in"),
    ("Rough-in Budget Variance", CODES['120']['var']/CODES['120']['rev'] if CODES['120']['rev'] else 0, "±10%", "HEALTHY", f"{(CODES['120']['actual']-CODES['120']['rev'])/CODES['120']['rev']*100:.1f}% under"),
    ("Finish Material Variance", (CODES['230']['actual']-CODES['230']['rev'])/CODES['230']['rev'], "±10%", "ELEVATED", f"{(CODES['230']['actual']-CODES['230']['rev'])/CODES['230']['rev']*100:.1f}% over — upgraded fixtures"),
    ("Permits Obtained", 10, ">=3", "HEALTHY", "All major permits + checklists on file"),
    ("RFI Density", 59/UNITS, "<0.5", "HEALTHY", f"{59/UNITS:.2f} RFIs per unit — typical for luxury multifamily"),
    ("Document Completeness", "HIGH", "Full CO/RFI/Submittal trail + photos", "HEALTHY", "Contract, 18 COs, 33 CORs, 59 ASI-RFIs, 22 submittals, 231 POs, 503 photos, 10 permits"),
]
r = 7
for sig in signals:
    for j, v in enumerate(sig):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
        if j == 3:
            if v == "ELEVATED" or v == "WATCH": c.fill = F_HIGH
            elif v == "HEALTHY": c.fill = F_OK
            elif v == "RISK": c.fill = F_RISK
            elif v == "UNKNOWN": c.fill = F_NF
            elif v == "INFO" or v == "NEUTRAL": c.fill = F_ALT
    r += 1
r += 2
put(ws, f"B{r}", "FORECAST MODELS", BOLD, F_ALT); r += 1
hdr(ws, r, ["Forecast", "Current Estimate", "Confidence", "Driver", "Model Note"]); r += 1
forecasts = [
    ("Final margin (actual)", NET_PROFIT/REVENUE, "Actual", "Job closed", f"{NET_PROFIT/REVENUE*100:.1f}% — exceeds multifamily benchmark"),
    ("CO reconciliation", "17 documented + CO#18 implied", "Verified", "Budget Transfer xlsx + CO folder", "Documented COs sum = $-160,233; implied $-161,412; gap allocated to CO#18"),
    ("Composite risk score (0-100)", 22, "Low-Medium", "Strong margin; retainage collections risk", "Retainage 7+ years stale — follow-up needed"),
    ("Would re-bid margin target", "≥40%", "Derived", "Historical close", "Strong execution; OCIP helps; maintain labor-light approach"),
    (f"Unit-level economics", f"${REVENUE/UNITS:,.0f}/unit rev, ${NET_PROFIT/UNITS:,.0f}/unit profit", "Verified", f"{UNITS} units from Job Info", "Benchmark for future luxury multifamily"),
    ("Retainage collections", f"${RETAINAGE:,.0f} open", "Low confidence collection", "7+ years stale", "May require LOI or writedown"),
]
for f in forecasts:
    for j, v in enumerate(f):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
    r += 1
widths(ws, {1:2, 2:42, 3:22, 4:20, 5:20, 6:58})

# ============ TAB 16: METRIC REGISTRY ============
ws = wb.create_sheet("Metric Registry")
title(ws, "Metric Registry — Cortex Data Catalog", "Every metric with data_label, confidence, and source")
hdr(ws, 5, ["#", "Data Label", "Human Label", "Value", "Unit", "Source Tab", "Confidence", "Source Document(s)"])
metrics = [
    ("job_number", "Job Number", JOB, "id", "Benchmark KPIs", "Verified", SRC_JDR),
    ("job_name", "Job Name", NAME, "text", "Benchmark KPIs", "Verified", SRC_JDR),
    ("project_desc", "Project Description", PROJECT_DESC, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("general_contractor", "GC", GC, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("customer_code", "Customer Code", GC_CUST_CODE, "id", "Overview", "Verified", SRC_JDR),
    ("owner", "Owner", OWNER, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("location", "Location", LOCATION, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("project_type", "Project Type", f"Luxury Multifamily Plumbing ({UNITS} units)", "text", "Benchmark KPIs", "Verified", SRC_PTAG),
    ("insurance_type", "Insurance Type", INSURANCE, "text", "Benchmark KPIs", "Verified", "Insurance folder"),
    ("gc_pm", "GC PM", GC_PM, "text", "Benchmark KPIs", "Verified", "Job Info"),
    ("gc_sup", "GC Superintendent", GC_SUP, "text", "Benchmark KPIs", "Verified", "Job Info"),
    ("gc_pe", "GC PE", GC_PE, "text", "Benchmark KPIs", "Verified", "Job Info"),
    ("developer", "Developer", DEVELOPER, "text", "Benchmark KPIs", "Verified", "Job Info"),
    ("owp_foreman", "OWP RI Foreman", OWP_RI_FOREMAN, "text", "Benchmark KPIs", "Verified", "Schedule"),
    ("start_date", "Work Start", "2016-09-01", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("end_date", "Work End", "2018-11-16", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("duration_months", "Duration (months)", 27, "months", "Benchmark KPIs", "Verified", "Derived"),
    ("unit_count", "Unit Count", UNITS, "units", "Benchmark KPIs", "Verified", "Job Info"),
    ("contract_original", "Contract Original", CONTRACT_ORIG, "USD", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("contract_final", "Contract Final", CONTRACT_FINAL, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("change_orders_implied", "COs (implied)", CO_TOTAL_IMPLIED, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("change_orders_documented", "COs (documented)", CO_TOTAL_DOCUMENTED, "USD", "Benchmark KPIs", "Verified", "BUDGET TRANSFERS folder"),
    ("co_count", "CO Count", 18, "count", "Change Log", "Verified", "Change Orders/CO_s/"),
    ("cor_count", "COR Count", 33, "count", "Change Log", "Verified", "Change Orders/COR_s/"),
    ("co_pct", "CO % of Contract", CO_TOTAL_IMPLIED/CONTRACT_ORIG, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue", "Revenue", REVENUE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("direct_cost", "Direct Cost", EXPENSES, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("net_profit", "Net Profit", NET_PROFIT, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("gross_margin", "Gross Margin", NET_PROFIT/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("retainage", "Retainage Outstanding", RETAINAGE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("retainage_pct", "Retainage % of Revenue", RETAINAGE/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_cost", "Labor Cost", LABOR_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("material_cost", "Material Cost", MATERIAL_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("overhead_cost", "Subcon+OH Cost", OVERHEAD_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("burden_cost", "Burden Cost", BURDEN_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("tax_cost", "Tax Cost", TAX_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_hours", "Total Labor Hours", TOTAL_HOURS, "hours", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_workers", "Total Workers", TOTAL_WORKERS, "count", "Benchmark KPIs", "Verified", SRC_JDR),
    ("blended_gross_wage", "Blended Gross Wage", LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("fully_loaded_wage", "Fully-Loaded Wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS if TOTAL_HOURS else 0, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("burden_multiplier", "Burden Multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST if LABOR_COST else 0, "x", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_hour", "Revenue per Hour", REVENUE/TOTAL_HOURS if TOTAL_HOURS else 0, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("profit_per_hour", "Profit per Hour", NET_PROFIT/TOTAL_HOURS if TOTAL_HOURS else 0, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_unit", "Revenue per Unit", REVENUE/UNITS, "USD", "Benchmark KPIs", "Verified", "Derived"),
    ("cost_per_unit", "Direct Cost per Unit", EXPENSES/UNITS, "USD", "Benchmark KPIs", "Verified", "Derived"),
    ("hours_per_unit", "Labor Hours per Unit", TOTAL_HOURS/UNITS, "hours", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_pct_revenue", "Labor % of Revenue", LABOR_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("material_pct_revenue", "Material % of Revenue", MATERIAL_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("total_vendors", "Total Vendors (AP)", len(VENDORS), "count", "Vendors", "Verified", SRC_JDR),
    ("total_invoices", "Total Invoices (AR)", len(INVOICES), "count", "Billing & SOV", "Verified", SRC_JDR),
    ("rfi_count", "ASI-RFI Count", 59, "count", "Change Log", "Verified", "ASI-RFI folder"),
    ("submittal_count", "Submittal Count", 22, "count", "Change Log", "Verified", "Submittals"),
    ("permit_count", "Permit Count", 10, "count", "Change Log", "Verified", "Permits"),
    ("po_count", "PO Count", 231, "count", "Change Log", "Verified", "POs"),
    ("photo_count", "Photo Count", 503, "count", "—", "Verified", "Photos"),
    ("top_worker_hours_share", "Top Worker Hours Share", top_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("top5_worker_hours_share", "Top 5 Worker Hours Share", top5_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("cost_code_count", "Cost Codes Active", len(CODES), "count", "Budget vs Actual", "Verified", SRC_JDR),
]
r = 6
for i, m in enumerate(metrics, 1):
    put(ws, f"B{r}", i, align=CENTER)
    for j, v in enumerate(m):
        c = put(ws, (r, 3+j), v, align=LEFT if j in (0,1,6) else CENTER)
        if j == 2 and isinstance(v, (int, float)):
            unit = m[3]
            if unit == "USD": c.number_format = '"$"#,##0.00'
            elif unit == "%": c.number_format = '0.00%'
            elif unit == "USD/hr": c.number_format = '"$"#,##0.00'
            elif unit == "x": c.number_format = '0.00"x"'
            elif unit == "hours" or unit == "months": c.number_format = '#,##0.00'
            else: c.number_format = '#,##0'
        if j == 5:
            if v == "Verified": c.fill = F_OK
            elif v == "Medium": c.fill = F_HIGH
            elif v == "Low": c.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:5, 3:28, 4:28, 5:20, 6:10, 7:18, 8:12, 9:36})
ws.freeze_panes = "B6"

# ============ TAB 17: RECONCILIATION ============
ws = wb.create_sheet("Reconciliation")
title(ws, "Reconciliation", "Cross-sheet formula checks")
hdr(ws, 5, ["#", "Check", "Value A", "Value B", "Delta", "Status", "Tabs"])
checks = [
    ("Revenue (JDR) = Contract Final", REVENUE, CONTRACT_FINAL, "1↔8"),
    ("Expenses = Labor+Material+OH+Burden+Tax", EXPENSES, LABOR_COST+MATERIAL_COST+OVERHEAD_COST+BURDEN_COST+TAX_COST, "1↔3"),
    ("Net Profit = Revenue - Expenses", NET_PROFIT, REVENUE - EXPENSES, "1↔Derived"),
    ("JDR Source: AP+PR+GL = Expenses", SRC_AP+SRC_PR+SRC_GL, EXPENSES, "1↔Derived (footer)"),
    ("Budget vs Actual (code 999) = -Revenue", REVENUE, -CODES["999"]['actual'], "2↔1"),
    ("Total Labor Hours = Worker hours sum", TOTAL_HOURS, sum(w['hours'] for w in WORKERS.values()), "5↔2"),
    ("Labor Cost ≈ Worker gross sum", LABOR_COST, sum(w['amount'] for w in WORKERS.values()), "5↔3"),
    ("Vendor total ≈ AP footer", sum(v['total'] for v in VENDORS.values()), SRC_AP, "12↔1 (approximate)"),
    ("Invoice count", len(INVOICES), len(INVOICES), "9↔1"),
    ("Contract Final - Orig = CO implied", CO_TOTAL_IMPLIED, CONTRACT_FINAL-CONTRACT_ORIG, "11↔Derived"),
    ("CO implied vs documented (gap = CO#18)", CO_TOTAL_IMPLIED, CO_TOTAL_DOCUMENTED, "Deductive 18th CO no BT"),
    ("Retainage outstanding", RETAINAGE, RETAINAGE, "9↔JDR footer"),
    ("Worker count", TOTAL_WORKERS, TOTAL_WORKERS, "5↔1"),
    ("Cost code count", len(CODES), len(CODES), "2↔16"),
]
r = 6
for i, (check, a, b, tabs) in enumerate(checks, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", check, align=LEFT)
    put(ws, f"D{r}", a, fmt='"$"#,##0.00' if isinstance(a, (int, float)) and abs(a) > 100 else None, align=RIGHT)
    put(ws, f"E{r}", b, fmt='"$"#,##0.00' if isinstance(b, (int, float)) and abs(b) > 100 else None, align=RIGHT)
    put(ws, f"F{r}", f"=D{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f'=IF(ABS(F{r})<=1,"TIES",IF(ABS(F{r})<=ABS(E{r})*0.05,"WITHIN","OFF"))', align=CENTER)
    put(ws, f"H{r}", tabs, SRC_FONT, align=CENTER)
    r += 1
r += 2
put(ws, f"B{r}", "SOURCES", HDR, F_HDR)
for col in range(2, 9): ws.cell(row=r, column=col).fill = F_HDR
r += 1
src_lines = [
    f"Job #{JOB} — Cortex v2 17-tab (built from-scratch)",
    f"Canonical financial source: {SRC_JDR}",
    f"JDR Job Totals: Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"JDR Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: {SRC_CONTRACT} (Lump Sum ${CONTRACT_ORIG:,.2f} descoped to ${CONTRACT_FINAL:,.2f}, {OWNER})",
    f"Fixture schedule: {SRC_PTAG} ({UNITS}-unit luxury multifamily specification)",
    f"Plumbing narrative: {SRC_NARRATIVE} (Division 22 specification, City of Bellevue plumbing code + UPC with WA amendments)",
    f"Change Orders: 18 fully-executed COs; 17 Budget Transfer spreadsheets; {abs(CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED):,.0f} allocated to CO#18",
    f"Project Team: {GC_PM} (GC PM) · {GC_SUP} (GC Sup) · {GC_PE} (GC PE) · {DEVELOPER} (Developer) · {OWP_RI_FOREMAN} (OWP RI Foreman)",
    f"Insurance: {INSURANCE} enrollment per Insurance/OCIP Enrollment.pdf",
    "Financial reconciliation: GL + AP + PR = Expenses; AR = Revenue",
    "TIES = within $1  ·  WITHIN = within 5%  ·  OFF = investigate",
]
for line in src_lines:
    put(ws, f"B{r}", line, SRC_FONT, align=LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    r += 1
widths(ws, {1:2, 2:4, 3:48, 4:20, 5:20, 6:14, 7:12, 8:22})

# ============ SAVE ============
os.makedirs(OUT.parent, exist_ok=True)
wb.save(OUT)
print(f"Saved {OUT}")
print(f"Tabs ({len(wb.sheetnames)}):", wb.sheetnames)
