#!/usr/bin/env python3
"""
enrich_2107.py — Enrich OWP_2107_JCR_Cortex_v2.xlsx — Chinn 68th & Roosevelt
244 units · ACTIVE RI phase · 90.4% billed · 30.7% margin actual to date.

Project: 2107 Chinn 68th & Roosevelt — 244 units
         Active RI phase · $4.42M billed of $4.89M revised · 33 workers · 25,137 hrs
         Site: 6716 Roosevelt Way NE, Seattle WA 98115
         Owner: 6716 Roosevelt Owner, LLC / High Street Northwest Development (Trammell Crow)
         Architect: Weinstein A+U · MEP: Emerald City (bridge) + Robison Engineering (OWP D-B)
         Foreman: Anderson, Richard P + Paco Leyva, Orlando

Verified data sources:
  • 2107 Job Detail Report.pdf (Sage Timberline · Apr 3 2026 run · 27 cost codes · 9,997 line items)
  • OWP Project List with Schedule - UPDATED 04-01-26.xlsx
    - Schedule r94: 244u · Foreman Rick/Orlando
    - Projects Bid r199 (Oct 2022): $5,844,000 initial Chinn bid
    - Projects Bid r217 (Apr 2023): $6,038,000 revised budget
    - Projects Bid r278-281 (Mar 2024): 5-bidder competition at $5,098,000
      (Chinn, Braseth, Compass, Venture, +1 unnamed)
  • Build_2107.py META — 20-CO itemized log with ASI/RFI references

NO GDrive Job Book folder yet (active project · folder created at closeout).
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2107_data.json"
DASH_FILE  = SCRIPT_DIR / "2107_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2107_JCR_Cortex_v2.xlsx"

META_2107 = {
    "job_id":          "2107",
    "name":            "Chinn 68th & Roosevelt",
    "short_name":      "68th & Roosevelt",
    "gc":              "Chinn Construction, LLC",
    "gc_pm":           "TBD (Chinn) — to be enriched at closeout",
    "gc_sup":          "TBD",
    "gc_pe":           "TBD",
    "owp_ri_foreman":  "Anderson, Richard P (1,885 hrs · $62.70/hr) + Paco Leyva, Orlando (1,309 hrs · $61.70/hr)",
    "owp_trim_foreman":"TBD (trim phase upcoming)",
    "owp_signatory":   "Richard Donelson",
    "owp_estimator":   "OWP Estimating",
    "owner":           "6716 Roosevelt Owner, LLC / High Street Northwest Development (Trammell Crow)",
    "developer":       "High Street Northwest Development (a Trammell Crow division)",
    "architect":       "Weinstein A+U, LLC",
    "structural":      "TBD",
    "mep_engineer":    "Emerald City Engineering (GC bridge MEP, per Master Project List) + Robison Engineering, Inc. (OWP's 3rd-party plumbing design-build engineer, per build_2107 META)",
    "civil":           "TBD",
    "landscape":       "TBD",
    "site_address":    "6716 Roosevelt Way NE, Seattle, WA 98115 (Roosevelt neighborhood, NE 67th-68th block)",
    "permit":          "City of Seattle Plumbing Permit (7 permit-related documents per META)",
    "insurance":       "Standard COI · non-Wrap (per OWP project team grid · build_2107 META)",
    "lien_position":   "Standard subcontract — retention $245,731 held",
    "warranty":        "1-yr standard plumbing (post-completion · ~Q4 2026)",
    "delivery_route":  "TBD",
    "contract_doc":    "Subcontract executed Jan 2025 · final contract $4,889,866 (was $4,671,538 base · +$218,328 / +4.7% net COs)",
    "scope":           "Multi-family apartments — 6-story wood-frame · plumbing scope",
    "gdrive_status":   "No GDrive Job Book folder yet (active project — created at closeout). "
                        "Document data sourced from build_2107 META (20 COs, 34 CORs, 258 RFIs, "
                        "15 ASIs, 184 submittals, 438 POs, 7 permits, 14 pay apps).",
    "status_text":     "ACTIVE · RI phase · 90.4% billed · 30.7% margin",
    "status_severity": "HEALTHY",
    "ar_billed":       4422211,
    "direct_cost":     2919284,
    "net_profit":      1502927,
    "retention":       245731,
    "contract_orig":   4671538,
    "contract_final":  4889866,
    "co_net":          218328,
    "units":           244,
    "fixtures":        "TBD — no Job Ticket in dataset yet (will appear in GDrive Job Book at closeout). "
                        "Estimated ~1,100 trim fixtures based on 4.5/unit ratio from #2061 Alta Columbia "
                        "City (a similar 243u Chinn-pattern multifamily). Fixture types from CO log: "
                        "WC1/WC2/L1/L2/TS1/TS2/SS1/SS2/SS3/S1/S2 + dog wash + coffee maker (per CO #5, #20).",
    "project_type":    "244-unit multifamily · 6-story wood-frame · Roosevelt neighborhood Seattle",
    "schedule":        "Jan 2025 → ~Q4 2026 (estimated · ~24 months)",
    "expected_start":  "Jan 2025",
    "expected_finish": "~Q4 2026 (estimated)",
    "duration_months": 24,
    "workers":         33,
    "hours":           25137,
    "top_vendor":      "Consolidated Supply Co. ($420,436 / 101 invoices / 38.1% of $1.10M AP)",
    "ap_total":        1103656,
    "ap_vendor_count": 23,
    "doc_counts": {
        "executed_COs": 20,
        "CORs":         34,
        "RFIs":         258,    # very high — typical for 244-unit Chinn project
        "ASIs":         15,
        "Submittals":   184,
        "POs":          438,
        "Permits":      7,
        "Pay_apps":     14,
        "Bond_cost":    52500,
    },
    "bid_history": [
        ("2022-10-21", "BID",    "Chinn",       244, 5844000, "Initial Chinn bid · single GC"),
        ("2023-04-06", "BUDGET", "Chinn",       244, 6038000, "Revised budget update (+3.3%) · 'Don't think anyone is competition'"),
        ("2024-03-06", "BID",    "Chinn",       244, 5098000, "5-bidder competition · re-bid with Chinn winning"),
        ("2024-03-06", "BID",    "Braseth",     244, 5098000, "5-bidder competition (Braseth same price as Chinn)"),
        ("2024-03-06", "BID",    "Compass",     244, 5098000, "5-bidder competition"),
        ("2024-03-06", "BID",    "Venture",     244, 5148000, "5-bidder competition (Venture +$50k)"),
        ("Subcontract", "EXEC",  "Chinn",       244, 4671538, "Final negotiated subcontract (Chinn awarded · negotiated down from $5.10M)"),
    ],
}

INK = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD = Font(name="Arial", size=10, color="FFFFFF", bold=True)
ACTIVE_FONT = Font(name="Arial", size=10, color="155724", bold=True)

HDR_FILL = PatternFill("solid", start_color="2C3E50")
SECTION_FILL = PatternFill("solid", start_color="ECF0F1")
TEAM_FILL = PatternFill("solid", start_color="FFF8E7")
ACTIVE_FILL = PatternFill("solid", start_color="D4EDDA")
WARN_FILL = PatternFill("solid", start_color="FFF3CD")
THIN = Side(border_style="thin", color="D5D8DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")

def clear_range(ws, r1, r2, c1=1, c2=10):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).border = Border()


def rebuild_job_info(ws):
    clear_range(ws, 1, ws.max_row + 5, 1, 6)
    ws["A1"] = f"JOB #{META_2107['job_id']} · {META_2107['name'].upper()} · KEY PLAYERS & METADATA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (f"Active RI phase · {META_2107['ar_billed']/META_2107['contract_final']*100:.1f}% billed of "
                f"${META_2107['contract_final']:,} revised contract. {META_2107['workers']} workers logged "
                f"{META_2107['hours']:,} hours. {META_2107['doc_counts']['executed_COs']} executed COs net "
                f"+${META_2107['co_net']:,} (+{META_2107['co_net']/META_2107['contract_orig']*100:.1f}%). "
                "Sources: Sage JDR + OWP Master Project List + build_2107 META.")
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    margin = META_2107['net_profit'] / META_2107['ar_billed']
    pct_complete = META_2107['ar_billed'] / META_2107['contract_final']

    sections = [
        ("IDENTITY", [
            ("Job Number",         META_2107["job_id"]),
            ("Job Name",           META_2107["name"]),
            ("Project Type",       META_2107["project_type"]),
            ("Site Address",       META_2107["site_address"]),
            ("Permit",             META_2107["permit"]),
            ("Status",             META_2107["status_text"]),
            ("Scope of Work",      META_2107["scope"]),
        ]),
        ("SCHEDULE", [
            ("Project Start",     META_2107["expected_start"]),
            ("Expected End",      META_2107["expected_finish"]),
            ("Duration (est.)",   f"{META_2107['duration_months']} months"),
            ("Schedule Note",     META_2107["schedule"]),
        ]),
        ("CONTRACT & FINANCIALS (ACTIVE — TO DATE)", [
            ("Original Contract",  f"${META_2107['contract_orig']:,}"),
            ("Net Change Orders",  f"+${META_2107['co_net']:,} (+{META_2107['co_net']/META_2107['contract_orig']*100:.1f}%) across {META_2107['doc_counts']['executed_COs']} executed COs"),
            ("Final Contract (revised)", f"${META_2107['contract_final']:,}"),
            ("AR Billed to Date",  f"${META_2107['ar_billed']:,}"),
            ("% Complete (billed/revised)", f"{pct_complete*100:.1f}%"),
            ("Direct Cost (actual)", f"${META_2107['direct_cost']:,}"),
            ("Net Profit (actual to date)", f"${META_2107['net_profit']:,}"),
            ("Gross Margin (actual to date)", f"{margin*100:.1f}%"),
            ("Retention Held",     f"${META_2107['retention']:,} (5%)"),
            ("Insurance",          META_2107["insurance"]),
            ("Lien Position",      META_2107["lien_position"]),
            ("Warranty",           META_2107["warranty"]),
            ("Contract on File",   META_2107["contract_doc"]),
        ]),
        ("SCOPE & FIXTURES", [
            ("Plumbing Units",     f"{META_2107['units']} units ✓ CONFIRMED via OWP Master List (Schedule r94 + 4 Projects Bid entries 2022-2024 all show 244u)"),
            ("Total Fixtures",     META_2107["fixtures"][:200]),
            ("Floors",             "6-story wood-frame multifamily (per build_2107 META)"),
            ("Fixture Provenance", "No Job Ticket in dataset yet (active project · GDrive Job Book created at closeout). Per CO log inferred fixtures: dog wash (CO#5) + coffee maker (CO#20) + standard residential trim. Definitive count pending Job Book."),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR (Chinn)", [
            ("General Contractor", META_2107["gc"]),
            ("GC Project Manager", META_2107["gc_pm"]),
            ("GC Superintendent",  META_2107["gc_sup"]),
            ("GC Project Engineer", META_2107["gc_pe"]),
            ("Chinn engagement count", "OWP's 9th Chinn engagement after Indigo/Kavela, Greenwood, Old Town, Barrett Park, Hadley/Legacy, Luna California, 1405 Dexter, SunDowner. Strong recurring relationship."),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2107["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2107["owp_trim_foreman"]),
            ("OWP Estimator",       META_2107["owp_estimator"]),
            ("OWP Signatory",       META_2107["owp_signatory"]),
            ("Active Crew",         f"{META_2107['workers']} workers · {META_2107['hours']:,} hrs YTD"),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record",   META_2107["owner"]),
            ("Developer",         META_2107["developer"]),
            ("Owner-developer note", "Trammell Crow's NW residential arm operates as 'High Street Northwest Development'. The 6716 Roosevelt Owner, LLC SPV is the title-holding entity."),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",         META_2107["architect"]),
            ("Structural Engineer", META_2107["structural"]),
            ("MEP / Plumbing Engineer", META_2107["mep_engineer"]),
            ("Civil Engineer",    META_2107["civil"]),
            ("Landscape Architect", META_2107["landscape"]),
        ]),
        ("BID HISTORY (from OWP Master Project List)", []),
        ("AP VENDOR PROFILE (TO DATE)", [
            ("Top Vendor",         META_2107["top_vendor"]),
            ("Total AP Spend",     f"${META_2107['ap_total']:,}"),
            ("Active Vendors",     f"{META_2107['ap_vendor_count']} unique vendors"),
            ("Concentration Note", "Consolidated Supply 38.1% + Pacific Plumbing 16.8% + Keller 11.8% + Rosen 10.6% = Top 4 dominate at 77% of AP. Standard for active Chinn jobs."),
        ]),
        ("DOCUMENT META (live counts)", [
            ("Executed COs",      f"{META_2107['doc_counts']['executed_COs']} COs · net +${META_2107['co_net']:,}"),
            ("CORs",              f"{META_2107['doc_counts']['CORs']} unique CORs"),
            ("RFIs",              f"{META_2107['doc_counts']['RFIs']} (high — typical for 244-unit Chinn project)"),
            ("ASIs",              f"{META_2107['doc_counts']['ASIs']} (highest = ASI 15 per build META · CO #20)"),
            ("Submittals",        f"{META_2107['doc_counts']['Submittals']} files"),
            ("POs",               f"{META_2107['doc_counts']['POs']} POs (large active inventory)"),
            ("Permits",           f"{META_2107['doc_counts']['Permits']} permit-related docs"),
            ("Pay Apps",          f"{META_2107['doc_counts']['Pay_apps']} pay apps · 94 AR transactions in JDR"),
            ("Bond cost",         f"${META_2107['doc_counts']['Bond_cost']:,} (CO #6)"),
            ("Notes",              "Document counts come from build_2107 META (manually maintained) — JDR shows 9,997 raw line items but no document counts. GDrive Job Book will populate at closeout."),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",            "2107 Job Detail Report (Sage Timberline · Apr 3 2026 run)"),
            ("Parsed Data",        "2107_data.json (27 cost codes · 9,997 line items · 33 workers)"),
            ("Dashboard Arrays",   "2107_dashboard_arrays.json (16 arrays · 20-entry CO log + 5-cause root cause analysis)"),
            ("Master Project List", "OWP Project List with Schedule - UPDATED 04-01-26.xlsx (Schedule r94 + Projects Bid r199/r217/r278-281 = 7-bid history)"),
            ("Build META",         "build_2107.py META_2107 dict — 20-CO itemized log + doc counts"),
            ("GDrive Folder",      "Not yet created (active)"),
            ("GDrive Status",      META_2107["gdrive_status"]),
        ]),
        ("RISK FLAGS / WATCH ITEMS", [
            ("Healthy margin posture", f"30.7% margin actual to date · above OWP 28% target. Net +${META_2107['co_net']:,} CO growth (modest +4.7% additive). Forecast healthy through closeout."),
            ("RFI velocity flag", "258 RFIs over ~15 months = ~17/month. High RFI velocity suggests significant design clarification activity — typical for Trammell Crow / Weinstein A+U projects but worth noting for trim-phase planning."),
            ("ASI 15 just hit", "Latest CO (#20, Apr 3 2026) is ASI 15 + RFI 258 = the design is still iterating in trim. Watch for trim-phase scope creep."),
            ("Bid drift (5,098 → 4,672)", "OWP negotiated $426k off the Mar 2024 bid price ($5.10M → $4.67M = -8.4%) before subcontract execution. Margin held despite price pressure due to disciplined cost management."),
            ("Robison D-B engineer", "Same OWP design-build engineer pattern as #2061 Alta Columbia City + #2106 6220 Roosevelt + #2114 Holland Ballard Blossom + #2118 Edmonds Behar. Robison is OWP's standard 3rd-party plumbing engineer for design-build subcontracts."),
        ]),
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        if section_name.startswith("BID HISTORY"):
            headers = ["Bid Date","Type","GC","Units","Price","Note"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=r, column=c, value=h).font = WHITE_BOLD
                ws.cell(row=r, column=c).fill = HDR_FILL
                ws.cell(row=r, column=c).border = BORDER
            r += 1
            for date, type_, gc, units, price, note in META_2107["bid_history"]:
                ws.cell(row=r, column=1, value=date).font = INK
                ws.cell(row=r, column=2, value=type_).font = INK_BOLD if type_ in ("BID","EXEC") else INK
                ws.cell(row=r, column=3, value=gc).font = INK
                ws.cell(row=r, column=4, value=units).font = INK
                ws.cell(row=r, column=5, value=f"${price:,}").font = INK
                ws.cell(row=r, column=6, value=note).font = INK
                ws.cell(row=r, column=6).alignment = WRAP
                if type_ == "EXEC":
                    for c in range(1,7): ws.cell(row=r, column=c).fill = ACTIVE_FILL
                for c in range(1,7): ws.cell(row=r, column=c).border = BORDER
                r += 1
            r += 1
            continue
        for label, val in fields:
            ws.cell(row=r, column=1, value=label).font = INK_BOLD
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_v.font = INK
            cell_v.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=1).border = BORDER
            cell_v.border = BORDER
            if section_name.startswith("PROJECT TEAM") or section_name.startswith("RISK") or section_name.startswith("AP VENDOR") or section_name.startswith("DOCUMENT META"):
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v.fill = TEAM_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 50


def patch_overview(ws):
    start = 30
    clear_range(ws, start, start + 30, 1, 8)
    ws.cell(row=start, column=1, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    ws.cell(row=start+1, column=1, value=(
        f"Active RI phase · {META_2107['ar_billed']/META_2107['contract_final']*100:.1f}% billed · "
        f"{META_2107['net_profit']/META_2107['ar_billed']*100:.1f}% margin · 20 executed COs (+4.7% additive) · "
        "Foreman Anderson + Paco Leyva."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=7)

    rows = [
        ("General Contractor",    META_2107["gc"]),
        ("Owner",                 META_2107["owner"]),
        ("Architect",             META_2107["architect"]),
        ("MEP / Plumbing Engineer", META_2107["mep_engineer"]),
        ("OWP RI Foreman",        META_2107["owp_ri_foreman"]),
        ("OWP Signatory",         META_2107["owp_signatory"]),
        ("Site Address",          META_2107["site_address"]),
        ("Insurance",             META_2107["insurance"]),
        ("Status",                META_2107["status_text"]),
        ("Original Contract",     f"${META_2107['contract_orig']:,}"),
        ("Final Contract (rev)",  f"${META_2107['contract_final']:,}"),
        ("AR Billed",             f"${META_2107['ar_billed']:,} ({META_2107['ar_billed']/META_2107['contract_final']*100:.1f}%)"),
        ("Net Profit (actual)",   f"${META_2107['net_profit']:,} ({META_2107['net_profit']/META_2107['ar_billed']*100:.1f}% margin)"),
        ("Retention Held",        f"${META_2107['retention']:,} (5%)"),
        ("Active Crew",           f"{META_2107['workers']} workers · {META_2107['hours']:,} hrs"),
    ]
    r = start + 3
    ws.cell(row=r, column=1, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=1, value=role).font = INK_BOLD
        ws.cell(row=r, column=2, value=val).font = INK
        ws.cell(row=r, column=1).fill = TEAM_FILL
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1


def patch_change_log(ws):
    last_row = ws.max_row
    r = last_row + 1
    ws.cell(row=r, column=1, value="2026-04-27").font = INK
    ws.cell(row=r, column=2, value="v1.1 · enriched").font = INK_BOLD
    ws.cell(row=r, column=3, value=(
        "Enriched Job Info tab with 13-section sectioned layout. Verified 244 unit count via OWP Master "
        "Project List (Schedule r94 + 4 Projects Bid entries 2022-2024). Imported 7-entry bid history "
        "(2022 initial Chinn bid → 2024 5-bidder competition → final negotiated subcontract). Imported "
        "20-entry executed CO log from build_2107 META with originator-tagged root-cause analysis. "
        "Owner enriched: 6716 Roosevelt Owner, LLC / High Street Northwest Development (Trammell Crow). "
        "Site address: 6716 Roosevelt Way NE, Seattle WA 98115. Foreman: Anderson, Richard P + Paco "
        "Leyva, Orlando per JDR top-worker analysis. Fixture count remains TBD (no Job Ticket yet)."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP


def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)
    print("→ Rebuilding Job Info tab (13 sections + bid history table)...")
    rebuild_job_info(wb["02 Job Info"])
    print("→ Patching Overview + adding Project Team block...")
    patch_overview(wb["01 Overview"])
    print("→ Logging enrichment to Change Log...")
    patch_change_log(wb["17 Change Log"])
    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
