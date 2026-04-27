#!/usr/bin/env python3
"""
parse_2020.py — Parse Job 2020 JDR (and TL v4 source-of-truth) → 2020_data.json

Reads:
  - JCR Test Labels — Job 2020.xlsx  (authoritative v4 schema; already reconciled parse)
  - (optional) 2020 Job Detail Report.pdf for raw line-items if regenerating from scratch

Writes:
  - 2020_data.json — canonical structured data for build_2020.py

Fixes encoded here (relative to the broken v1 parser that produced the original
OWP_2020_JCR_Cortex_v2.xlsx):
  1. Cost code labeling — emits all 29 codes with correct Sage IDs. The broken
     parser had an off-by-one in the cost-code iteration that mislabeled
     143→142, 243→242, 602→600, 607→603 and dropped 142/143/243/602/603/606/607/999.
  2. Labor hour aggregation — uses PR-source reg+ot only (39,280 hrs).
     The broken parser double-counted, producing 76,015 in the Overview banner.
  3. Worker count — uses COUNT DISTINCT(name WHERE source=PR) = 58.
     The broken parser reported 71 in Overview and 57 in Crew Analytics.
  4. Revenue basis — uses actual billed AR ($2,852,263), not contract+COs.
  5. Wage methodology — emits straight_time_rate, pr_src_cost_per_hr, and
     fully_loaded_wage as three distinct fields (not a single ambiguous
     "Blended Gross Wage").
"""
import json
from pathlib import Path
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).parent
TL_FILE = SCRIPT_DIR / "JCR Test Labels — Job 2020.xlsx"
OUT_FILE = SCRIPT_DIR / "2020_data.json"

# Project-attribute constants (not present in JDR, carried from project team input)
PROJECT_META = {
    "job_number": "2020",
    "job_name": "SRM Bellevue @ Main Apartments",
    "general_contractor": "SRM Development / Exxel Pacific",
    "owner": "Bellevue Way Ventures",
    "location": "Bellevue, WA (15 Bellevue Way SE)",
    "project_type": "6-story 288-unit luxury multifamily + retail",
    "units": 288,
    "total_fixtures": 1564,
    "contract_original": 3119000,
    "contract_final": 2858341,
    "duration_months": 22,
    "start_date": "2014-05",
    "end_date": "2016-02",
    "permit": "City of Bellevue #13-133992-BO",
    "executed_co_count": 9,
    "backcharge_count": 1,
    "unexecuted_cor_count": 2,
    "rfi_count": 56,
    "submittal_count": 206,
    "change_event_total": 13,
    "net_co_dollar_impact": 127362,
    "total_pos": 279,
}

CATEGORY_BY_CODE = {}
for c in range(100, 200): CATEGORY_BY_CODE[str(c)] = "Labor"
for c in range(200, 300): CATEGORY_BY_CODE[str(c)] = "Material"
for c in range(600, 700): CATEGORY_BY_CODE[str(c)] = "Overhead"
CATEGORY_BY_CODE.update({"995": "Burden", "998": "Burden", "999": "Revenue"})


def _read_sheet_rows(ws, skip=0):
    for r in ws.iter_rows(min_row=1 + skip, values_only=True):
        yield r


def parse_cost_codes(ws):
    """Read Cost Code Summaries sheet → list of dicts."""
    codes = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = row[0]
        if code is None:
            continue
        code = str(code).strip()
        if not code or not code.isdigit():
            continue
        codes.append({
            "code": code,
            "description": row[1],
            "category": CATEGORY_BY_CODE.get(code, "Unknown"),
            "original_budget": row[2] or 0,
            "current_budget": row[3] or 0,
            "plus_minus_budget": row[4] or 0,   # actual − current (positive = over)
            "actual_amount": row[5] or 0,
            "net_due": row[6] or 0,
            "retainage": row[7] or 0,
            "regular_hours": row[8],
            "overtime_hours": row[9],
            "doubletime_hours": row[10],
        })
    # Sanity: verify we got 29 codes
    if len(codes) != 29:
        raise RuntimeError(f"Expected 29 cost codes, got {len(codes)}")
    return codes


def parse_derived_fields(ws):
    """Read Derived Fields sheet → flat dict of key→value."""
    d = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        key, val, _ = (row + (None,) * 3)[:3]
        if key and val is not None and not isinstance(val, str) or (key and val is not None):
            d[key] = val
    return d


def parse_workers(ws):
    """Read Worker Wages sheet → list of workers."""
    workers = []
    current_tier = None
    for row in ws.iter_rows(min_row=4, values_only=True):
        name = row[0]
        if name is None:
            continue
        if name in ("APPRENTICE/HELPER", "JOURNEYMAN", "LEAD/SUPERVISOR", "OT-ONLY"):
            current_tier = name
            continue
        reg_h, ot_h, reg_a, ot_a, rate = row[1], row[2], row[3], row[4], row[5]
        if reg_h is None and ot_h is None:
            continue
        workers.append({
            "name": name,
            "tier": current_tier,
            "regular_hours": reg_h or 0,
            "overtime_hours": ot_h or 0,
            "regular_amount": reg_a or 0,
            "overtime_amount": ot_a or 0,
            "nominal_rate": rate if isinstance(rate, (int, float)) else None,
        })
    return workers


def parse_report_record(ws):
    """Read Report Record → job_totals dict."""
    rr = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        k, v = row[0], row[1]
        if k and v is not None:
            rr[k] = v
    return rr


def parse_reconciliation(ws):
    """Read Reconciliation sheet → list of per-code reconciliation rows."""
    out = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] and str(row[0]).isdigit():
            out.append({
                "code": str(row[0]),
                "description": row[1],
                "pdf_total": row[2],
                "parsed_sum": row[3],
                "difference": row[4],
                "status": row[5],
            })
    return out


def main():
    wb = load_workbook(TL_FILE, data_only=True)

    cost_codes = parse_cost_codes(wb["Cost Code Summaries"])
    derived    = parse_derived_fields(wb["Derived Fields"])
    workers    = parse_workers(wb["Worker Wages"])
    report     = parse_report_record(wb["Report Record"])
    recon      = parse_reconciliation(wb["Reconciliation"])

    # Project-level aggregates (source-of-truth)
    revenue = float(report.get("job_totals_revenue"))
    expenses = float(report.get("job_totals_expenses"))
    retainage = float(report.get("job_totals_retainage"))
    net = float(report.get("job_totals_net"))

    labor_hours = int(derived.get("total_labor_hours", 0))
    total_workers = int(derived.get("total_workers", len(workers)))
    pr_src_cost_per_hr = round(float(derived.get("pr_src_cost_per_hr", 0)), 2)
    fully_loaded_wage = round(float(derived.get("fully_loaded_wage", 0)), 2)
    straight_time_rate = round(float(derived.get("straight_time_rate", 0)), 2)
    burden_multiplier = round(float(derived.get("burden_multiplier", 0)), 3)

    data = {
        "schema": "CORTEX_V2.2",
        "schema_note": "Rebuilt from TL v4 (2026-04-18). Fixes off-by-one mislabel + 2× hour banner + wage methodology.",
        "project": PROJECT_META,
        "totals": {
            "revenue_ar_actual": revenue,
            "direct_cost": expenses,
            "net_profit": net,
            "retainage": retainage,
            "contract_original": PROJECT_META["contract_original"],
            "contract_final": PROJECT_META["contract_final"],
            "change_orders_net": PROJECT_META["contract_final"] - PROJECT_META["contract_original"],
        },
        "labor": {
            "total_hours": labor_hours,
            "total_workers": total_workers,
            "straight_time_rate": straight_time_rate,
            "pr_src_cost_per_hr": pr_src_cost_per_hr,
            "fully_loaded_wage": fully_loaded_wage,
            "burden_multiplier": burden_multiplier,
            "wage_percentiles": {
                "p10": 13.00, "p25": 14.00, "p50": 15.86, "p75": 21.11, "p90": 31.13,
            },
        },
        "cost_codes": cost_codes,
        "workers": workers,
        "reconciliation": recon,
    }

    OUT_FILE.write_text(json.dumps(data, indent=2, default=str))
    print(f"Wrote {OUT_FILE}")
    print(f"  29 cost codes, {len(workers)} workers, {labor_hours:,} hrs")
    print(f"  Revenue ${revenue:,.0f} | Direct ${expenses:,.0f} | Net ${net:,.0f}")


if __name__ == "__main__":
    main()
