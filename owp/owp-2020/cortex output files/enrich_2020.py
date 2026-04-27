#!/usr/bin/env python3
"""
enrich_2020.py — Fill in empty/sparse tabs in OWP_2020_JCR_Cortex_v2.xlsx.

Audit (2026-04-25) flagged these gaps in the v2 workbook:
  • Vendors tab        — completely empty (header only)
  • Predictive Signals — single placeholder row
  • PO Commitments     — PO numbers without details
  • Billing & SOV      — only 2 duplicate rows (vs 10+ pay apps available)
  • Insights           — 2 of 10 insights populated
  • Change Log         — partial (no change-meta summary)
  • Root Cause Analysis — placeholder text only
  • Cost Breakdown     — formatting glitches in column B
  • Material           — vendor section starts at row 16 with no header

Plus: NO project-team / key-players info exists anywhere in the workbook.

This script:
  1. Loads the existing OWP_2020_JCR_Cortex_v2.xlsx
  2. Reads 2020_data.json (parsed JDR) + 2020_dashboard_arrays.json (full enriched
     dashboard data: vendors, COs, pay apps, insights, predictive signals, etc.)
  3. Adds a comprehensive "Project Team" block to the top of the Overview tab
  4. Patches every empty/sparse tab with full content
  5. Writes back to OWP_2020_JCR_Cortex_v2.xlsx (idempotent — re-run safe)

Source for project-team data: index.html PROJECT_TEAMS['2020'] (gold-standard
team grid maintained on the dashboard). GDrive contract PDFs were locked
("Resource deadlock avoided") at parse time — note recorded for future re-run.
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2020_data.json"
DASH_FILE  = SCRIPT_DIR / "2020_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2020_JCR_Cortex_v2.xlsx"

# ============================================================================
# Project-team metadata (key players) — sourced from index.html PROJECT_TEAMS['2020']
# Cross-checked against CLAUDE.md and the dashboard hero block.
# GDrive contract PDFs were inaccessible at this run (deadlock); team data here
# matches the gold-standard team grid maintained on the dashboard.
# ============================================================================
META_2020 = {
    "gc": "SRM Development",
    "gc_full": "SRM Development / Exxel Pacific (CM-at-risk)",
    "gc_pm": "Todd Rankin / Erik Benzel",
    "gc_pm_phase": "Todd Rankin (early) → Erik Benzel (closeout)",
    "gc_sup": "Bill Robinson / Jared Osborn",
    "gc_pe": "Renay Luzama",
    "owp_ri_foreman": "Nate (Veley)",
    "owp_trim_foreman": "Al",
    "owp_signatory": "Richard Donelson",
    "owner": "Bellevue Way Ventures, LLC",
    "developer": "Pillar Properties / SRM",
    "architect": "Benson & Bohl Architects",
    "structural": "DCI Engineers",
    "acoustical": "SSA Acoustics",
    "mep_engineer": "Pressler Engineers",
    "civil": "TBD (not in dashboard)",
    "landscape": "TBD",
    "ada_consultant": "n/a (not flagged in 2020)",
    "interior": "n/a (not flagged in 2020)",
    "permit": "City of Bellevue #13-133992-BO",
    "insurance": "Not Wrap (OWP carries CGL — Cert of Insurance 2016 on file)",
    "contract_doc": "One Way Plumbing Executed Contract OCR_D.pdf · Contract folder",
    "lien_position": "Subordination + Priority agreements both on file (Bellevue Way Subordination 04/15/14)",
    "warranty": "Bellevue Way Ventures Warranty (LH) — 1 yr standard plumbing",
    "site_address": "15 Bellevue Way SE, Bellevue, WA",
    "delivery_route": "Bellevue at Main – Delivery Truck Route.pdf (on file)",
    "gdrive_status": "Folder accessible; PDFs deadlocked at this run — team data sourced from dashboard PROJECT_TEAMS['2020']",
}

# ============================================================================
# Style helpers
# ============================================================================
INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)

HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
ROW_ALT     = PatternFill("solid", start_color="FAFAFA")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")


def clear_range(ws, r1, r2, c1=2, c2=11):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)


# ============================================================================
# TEAM BLOCK (added to Overview tab below the existing identity rows)
# ============================================================================
def patch_overview_team_block(ws):
    """
    Insert a 'PROJECT TEAM / KEY PLAYERS' block on the Overview tab.
    Placed at row 40+ (well below the existing identity / contract / financial blocks).
    """
    start = 40
    ws.cell(row=start, column=2, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=2, end_row=start, end_column=8)

    ws.cell(row=start+1, column=2, value=(
        "Sourced from index.html PROJECT_TEAMS['2020'] (gold-standard team grid). "
        "GDrive contract PDFs were locked at extraction time — team data mirrors "
        "the dashboard's verified roster."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=2, end_row=start+1, end_column=8)

    # Team grid — 2 columns of role/value pairs per row
    rows = [
        ("General Contractor",  META_2020["gc_full"]),
        ("GC Project Manager",  META_2020["gc_pm_phase"]),
        ("GC Superintendent",   META_2020["gc_sup"]),
        ("GC Project Engineer", META_2020["gc_pe"]),
        ("OWP Roughin Foreman", META_2020["owp_ri_foreman"]),
        ("OWP Trim Foreman",    META_2020["owp_trim_foreman"]),
        ("OWP Signatory",       META_2020["owp_signatory"]),
        ("Owner of Record",     META_2020["owner"]),
        ("Developer",           META_2020["developer"]),
        ("Architect",           META_2020["architect"]),
        ("Structural Engineer", META_2020["structural"]),
        ("MEP / Plumbing Engineer", META_2020["mep_engineer"]),
        ("Acoustical",          META_2020["acoustical"]),
        ("Civil Engineer",      META_2020["civil"]),
        ("Landscape",           META_2020["landscape"]),
        ("Permit Number",       META_2020["permit"]),
        ("Insurance",           META_2020["insurance"]),
        ("Lien / Priority Position", META_2020["lien_position"]),
        ("Warranty",            META_2020["warranty"]),
        ("Site Address",        META_2020["site_address"]),
        ("Delivery Route",      META_2020["delivery_route"]),
        ("Contract on File",    META_2020["contract_doc"]),
        ("GDrive Notes",        META_2020["gdrive_status"]),
    ]
    r = start + 3
    ws.cell(row=r, column=2, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=3, value="NAME / VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=3).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=3).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=2, value=role).font = INK_BOLD
        ws.cell(row=r, column=3, value=val).font = INK
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=3).fill = TEAM_FILL
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=3).border = BORDER
        ws.cell(row=r, column=3).alignment = WRAP
        r += 1

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 75


# ============================================================================
# VENDORS TAB
# ============================================================================
def patch_vendors(ws, dash):
    vendors = dash.get("allVendors", [])
    # Clear existing
    clear_range(ws, 5, 80)

    # Header row
    headers = ["#", "Vendor Name", "# Invoices", "Total Spend ($)", "% of AP", "Tier", "Notes"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    total_spend = sum(v[2] for v in vendors)
    r = 6
    for i, v in enumerate(vendors, 1):
        name, invoices, spend, pct = v[0], v[1], v[2], v[3]
        tier = "Top 5" if i <= 5 else ("Top 10" if i <= 10 else "Long tail")
        notes = ""
        if i == 1: notes = "Largest AP — concentration flag"
        elif name.lower().startswith("rosen"): notes = "Primary trim/finish supplier"
        elif name.lower().startswith("keller"): notes = "Roughin material supplier"
        elif name.lower().startswith("ferguson"): notes = "Roughin + finish dual-source"
        elif "rentals" in name.lower(): notes = "Equipment rental"
        elif "concrete" in name.lower(): notes = "Concrete cutting/coring"

        row = [i, name, invoices, spend, f"{pct:.1f}%", tier, notes]
        for c, val in enumerate(row, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
            if r % 2 == 0: cell.fill = ROW_ALT
        r += 1

    # Totals row
    ws.cell(row=r, column=2, value="TOTAL").font = INK_BOLD
    ws.cell(row=r, column=3, value=f"{len(vendors)} vendors").font = INK_BOLD
    ws.cell(row=r, column=4, value=sum(v[1] for v in vendors)).font = INK_BOLD
    ws.cell(row=r, column=5, value=total_spend).font = INK_BOLD
    ws.cell(row=r, column=6, value="100.0%").font = INK_BOLD
    for c in range(2, 8):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).fill = SECTION_FILL

    # Column widths
    widths = {2: 5, 3: 35, 4: 12, 5: 16, 6: 10, 7: 12, 8: 35}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# PREDICTIVE SIGNALS TAB
# ============================================================================
def patch_predictive_signals(ws, dash):
    signals = dash.get("predictiveSignals", [])
    clear_range(ws, 5, 30)

    headers = ["#", "Signal", "Current Value", "Threshold", "Severity"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    r = 6
    for i, s in enumerate(signals, 1):
        if not isinstance(s, list) or len(s) < 4: continue
        signal, val, thr, sev = s[0], s[1], s[2], s[3]
        ws.cell(row=r, column=2, value=i).font = INK
        ws.cell(row=r, column=3, value=signal).font = INK
        ws.cell(row=r, column=4, value=val).font = INK
        ws.cell(row=r, column=5, value=thr).font = INK
        sev_cell = ws.cell(row=r, column=6, value=sev)
        sev_cell.font = INK_BOLD
        if sev == "WARN": sev_cell.fill = PatternFill("solid", start_color="FFF3CD")
        elif sev == "CRIT": sev_cell.fill = PatternFill("solid", start_color="F8D7DA")
        elif sev == "INFO": sev_cell.fill = PatternFill("solid", start_color="D1ECF1")
        for c in range(2, 7): ws.cell(row=r, column=c).border = BORDER
        r += 1

    widths = {2: 5, 3: 38, 4: 18, 5: 22, 6: 12}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# INSIGHTS TAB
# ============================================================================
def patch_insights(ws, dash):
    insights = dash.get("insights", [])
    # Clear from row 11 (preserve existing header rows 1-10)
    clear_range(ws, 11, 30)

    # Header
    headers = ["#", "Theme", "Insight"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=10, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    r = 11
    for i, ins in enumerate(insights, 1):
        if not isinstance(ins, list) or len(ins) < 2: continue
        theme, body = ins[0], ins[1]
        ws.cell(row=r, column=2, value=i).font = INK_BOLD
        ws.cell(row=r, column=3, value=theme).font = INK_BOLD
        ws.cell(row=r, column=4, value=body).font = INK
        ws.cell(row=r, column=4).alignment = WRAP
        for c in range(2, 5): ws.cell(row=r, column=c).border = BORDER
        if r % 2 == 0:
            for c in range(2, 5): ws.cell(row=r, column=c).fill = ROW_ALT
        r += 1

    widths = {2: 5, 3: 22, 4: 90}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# CHANGE LOG TAB (full rebuild + change-meta summary)
# ============================================================================
def patch_change_log(ws, dash):
    log = dash.get("changeLog", [])
    meta = dash.get("changeMeta", {})
    clear_range(ws, 5, 50)

    headers = ["Event ID", "Type", "Date", "Subject", "Originator", "Cost Impact ($)"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    r = 6
    for entry in log:
        if not isinstance(entry, list) or len(entry) < 5: continue
        eid, etype, edate, subj, orig = entry[0], entry[1], entry[2], entry[3], entry[4]
        cost = entry[5] if len(entry) > 5 else None
        cells = [eid, etype, edate, subj, orig, cost]
        for c, val in enumerate(cells, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
            cell.alignment = WRAP
        r += 1

    # Summary block
    r += 1
    ws.cell(row=r, column=2, value="CHANGE-EVENT META").font = SECTION_HDR
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1
    summary = [
        ("Total events",        meta.get("total", "")),
        ("Net cost impact ($)", meta.get("costImpact", "")),
        ("CO count",            meta.get("types", {}).get("CO", "")),
        ("COR count",           meta.get("types", {}).get("COR", "")),
        ("Backcharge count",    meta.get("types", {}).get("Backcharge", "")),
        ("Memo / gap entries",  meta.get("types", {}).get("Memo", "")),
    ]
    for k, v in summary:
        ws.cell(row=r, column=2, value=k).font = INK_BOLD
        ws.cell(row=r, column=3, value=v).font = INK
        for c in range(2, 4): ws.cell(row=r, column=c).border = BORDER
        r += 1

    widths = {2: 18, 3: 12, 4: 12, 5: 60, 6: 22, 7: 16}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# ROOT CAUSE ANALYSIS TAB
# ============================================================================
def patch_rca(ws, dash):
    rca = dash.get("rootCauses", [])
    resp = dash.get("responsibility", [])
    clear_range(ws, 5, 30)

    # Top half: root cause categories
    ws.cell(row=5, column=2, value="ROOT CAUSE CATEGORIES").font = SECTION_HDR
    ws.cell(row=5, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=6)

    headers = ["Cause", "Event Count", "Cost Impact ($)", "Primary Owner"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    r = 7
    for entry in rca:
        if not isinstance(entry, list) or len(entry) < 4: continue
        for c, val in enumerate(entry, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
        r += 1

    # Bottom: responsibility matrix
    r += 2
    ws.cell(row=r, column=2, value="RESPONSIBILITY MATRIX").font = SECTION_HDR
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    headers2 = ["Party", "Event Count", "Net Cost Impact ($)"]
    for c, h in enumerate(headers2, 2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
    r += 1
    for entry in resp:
        if not isinstance(entry, list) or len(entry) < 3: continue
        for c, val in enumerate(entry, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
        r += 1

    widths = {2: 24, 3: 16, 4: 18, 5: 22}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# BILLING & SOV TAB (rebuild from payApps array)
# ============================================================================
def patch_billing_sov(ws, dash):
    apps = dash.get("payApps", [])
    sov = dash.get("sovData", {})
    clear_range(ws, 5, 30)

    headers = ["Pay App #", "Date", "This Period", "Retainage Held", "Net Pay", "Cumulative", "% of Contract"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    r = 6
    for app in apps:
        if not isinstance(app, list) or len(app) < 5: continue
        # Format from index.html: [num, date, this_period, retain, net, cumulative, pct]
        num, date, this_period, retain, net = app[0], app[1], app[2], app[3], app[4]
        cumulative = app[5] if len(app) > 5 else None
        pct = app[6] if len(app) > 6 else None
        cells = [num, date, this_period, retain, net, cumulative, f"{pct*100:.1f}%" if pct else ""]
        for c, val in enumerate(cells, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
        r += 1

    # SOV summary
    r += 1
    ws.cell(row=r, column=2, value="STATEMENT OF VALUES — SUMMARY").font = SECTION_HDR
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    r += 1
    summary = [
        ("Original Contract",  sov.get("originalContract", "")),
        ("Net Change Orders",  sov.get("changeOrders", "")),
        ("Final Contract",     sov.get("finalContract", "")),
        ("Retainage Held",     sov.get("retainage", "")),
        ("Net Paid to OWP",    sov.get("netPaid", "")),
    ]
    for k, v in summary:
        ws.cell(row=r, column=2, value=k).font = INK_BOLD
        ws.cell(row=r, column=3, value=v).font = INK
        for c in range(2, 4): ws.cell(row=r, column=c).border = BORDER
        r += 1

    widths = {2: 11, 3: 12, 4: 14, 5: 16, 6: 14, 7: 14, 8: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# COST BREAKDOWN TAB (rebuild from costCats)
# ============================================================================
def patch_cost_breakdown(ws, dash):
    cats = dash.get("costCats", [])
    clear_range(ws, 5, 50)

    headers = ["Category", "# Codes", "Budget", "Actual", "% of Direct Cost", "$ per Unit"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    r = 6
    total_budget = total_actual = 0
    for cat in cats:
        if not isinstance(cat, list) or len(cat) < 6: continue
        category, n_codes, budget, actual, pct, per_unit = cat[0], cat[1], cat[2], cat[3], cat[4], cat[5]
        ws.cell(row=r, column=2, value=category).font = INK_BOLD
        ws.cell(row=r, column=3, value=n_codes).font = INK
        ws.cell(row=r, column=4, value=budget).font = INK
        ws.cell(row=r, column=5, value=actual).font = INK
        ws.cell(row=r, column=6, value=f"{pct:.1f}%").font = INK
        ws.cell(row=r, column=7, value=per_unit).font = INK
        for c in range(2, 8): ws.cell(row=r, column=c).border = BORDER
        if r % 2 == 0:
            for c in range(2, 8): ws.cell(row=r, column=c).fill = ROW_ALT
        total_budget += budget
        total_actual += actual
        r += 1

    # Total row
    ws.cell(row=r, column=2, value="DIRECT COST TOTAL").font = INK_BOLD
    ws.cell(row=r, column=4, value=total_budget).font = INK_BOLD
    ws.cell(row=r, column=5, value=total_actual).font = INK_BOLD
    ws.cell(row=r, column=6, value="100.0%").font = INK_BOLD
    for c in range(2, 8):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).fill = SECTION_FILL

    widths = {2: 18, 3: 9, 4: 16, 5: 16, 6: 16, 7: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# PO COMMITMENTS TAB (rebuild)
# ============================================================================
def patch_po_commitments(ws, dash, data):
    """PO data isn't fully captured in dashboard, but we have total_pos count.
    Build a clean placeholder structure that mirrors the format used for other projects."""
    clear_range(ws, 5, 50)
    p = data["project"]
    total_pos = p.get("total_pos", 0)
    vendors = dash.get("allVendors", [])

    headers = ["PO Type", "Count Estimate", "% of Total POs", "Top Vendors", "Notes"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    # Approximate distribution based on category share
    total_ap = sum(v[2] for v in vendors)
    rosen = next((v for v in vendors if "rosen" in v[0].lower()), None)
    fei   = next((v for v in vendors if "ferguson" in v[0].lower()), None)
    keller= next((v for v in vendors if "keller" in v[0].lower()), None)
    other_pct = 100 - (rosen[3] if rosen else 0) - (fei[3] if fei else 0) - (keller[3] if keller else 0)

    rows = [
        ("Trim/Finish (Rosen)",  round(total_pos * (rosen[3] if rosen else 0) / 100), f"{rosen[3] if rosen else 0:.1f}%", "Rosen Supply Company", "Trim phase POs"),
        ("Roughin (Ferguson)",   round(total_pos * (fei[3] if fei else 0) / 100),    f"{fei[3] if fei else 0:.1f}%",   "Ferguson Enterprises",  "Roughin phase POs"),
        ("Roughin (Keller)",     round(total_pos * (keller[3] if keller else 0) / 100), f"{keller[3] if keller else 0:.1f}%", "Keller Supply",         "Roughin phase POs"),
        ("Mech Sales / Specialty", "—", "—",  "Mechanical Sales, Inc.", "Specialty mech room components"),
        ("Equipment Rentals",    "—", "—", "United Rentals",         "Lifts, scaffolding"),
        ("Concrete Cutting",     "—", "—", "MAD Concrete / Quality Concrete", "Slab penetrations + sleeves"),
        ("Other",                "—", f"{other_pct:.1f}%", "Long tail (~20 vendors)", ""),
    ]
    r = 6
    for row in rows:
        for c, val in enumerate(row, 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = INK
            cell.border = BORDER
            cell.alignment = WRAP
        r += 1

    # Summary
    r += 1
    ws.cell(row=r, column=2, value=f"TOTAL PURCHASE ORDERS: {total_pos}").font = INK_BOLD
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    r += 1
    ws.cell(row=r, column=2, value=(
        "Note: Detailed PO-by-PO records live in 2020-SRM, Bellevue @ Main/PO_s/ on GDrive (~280 individual POs). "
        "This tab summarizes commitments by vendor concentration. Per-PO detail can be re-extracted on demand."
    )).font = GREY_FONT
    ws.cell(row=r, column=2).alignment = WRAP
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)

    widths = {2: 24, 3: 14, 4: 14, 5: 32, 6: 32}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# MATERIAL TAB (rebuild — fix vendor section header)
# ============================================================================
def patch_material(ws, dash, data):
    """Rebuild Material tab with proper vendor section header."""
    vendors = dash.get("allVendors", [])
    clear_range(ws, 5, 60)

    # A. Material cost-code spend
    ws.cell(row=5, column=2, value="A. MATERIAL SPEND BY COST CODE").font = SECTION_HDR
    ws.cell(row=5, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=7)

    headers = ["Code", "Description", "Budget", "Actual", "Variance", "% of Material"]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(row=6, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER

    material_codes = [c for c in data["cost_codes"] if c.get("category") == "Material"]
    total_actual = sum(c["actual_amount"] for c in material_codes)
    r = 7
    for c in material_codes:
        cells = [c["code"], c["description"], c["current_budget"], c["actual_amount"],
                 c["current_budget"] - c["actual_amount"],
                 f"{c['actual_amount']/total_actual*100:.1f}%" if total_actual else ""]
        for ci, val in enumerate(cells, 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = INK
            cell.border = BORDER
        r += 1

    # B. Vendors by spend (proper header)
    r += 1
    ws.cell(row=r, column=2, value="B. VENDORS BY SPEND (AP)").font = SECTION_HDR
    ws.cell(row=r, column=2).fill = SECTION_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1
    headers2 = ["#", "Vendor", "Invoices", "Spend", "% of AP"]
    for c, h in enumerate(headers2, 2):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HDR_FILL
        cell.border = BORDER
        cell.alignment = CENTER
    r += 1
    for i, v in enumerate(vendors, 1):
        cells = [i, v[0], v[1], v[2], f"{v[3]:.1f}%"]
        for ci, val in enumerate(cells, 2):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.font = INK
            cell.border = BORDER
        r += 1

    widths = {2: 6, 3: 32, 4: 14, 5: 14, 6: 14, 7: 14}
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w


# ============================================================================
# JOB INFO tab (new — full project metadata + key players)
# Insert after Overview if not already present
# ============================================================================
def add_job_info_tab(wb):
    """Add a clean Job Info tab dedicated to all project metadata + team."""
    if "Job Info" in wb.sheetnames:
        del wb["Job Info"]

    ws = wb.create_sheet("Job Info", index=1)  # Insert after Overview (index 0)
    ws["B2"] = "JOB #2020 · PROJECT INFORMATION & KEY PLAYERS"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:F2")
    ws["B3"] = "Comprehensive project metadata. Sourced from JDR + index.html PROJECT_TEAMS + GDrive scan (PDF deadlock noted)."
    ws["B3"].font = GREY_FONT
    ws.merge_cells("B3:F3")

    sections = [
        ("IDENTITY", [
            ("Job Number", "2020"),
            ("Job Name", "SRM Bellevue @ Main Apartments"),
            ("Project Type", "6-story 288-unit luxury multifamily + retail"),
            ("Location", "15 Bellevue Way SE, Bellevue, WA"),
            ("Permit", META_2020["permit"]),
            ("Site Address", META_2020["site_address"]),
            ("Delivery Route", META_2020["delivery_route"]),
        ]),
        ("SCHEDULE", [
            ("Project Start", "May 2014"),
            ("Project End",   "Feb 2016"),
            ("Duration",      "~22 months"),
        ]),
        ("CONTRACT & FINANCIALS", [
            ("Original Contract", "$3,119,000"),
            ("Net Change Orders", "($260,659)"),
            ("Final Contract",    "$2,858,341"),
            ("Revenue (AR)",      "$2,852,263"),
            ("Direct Cost",       "$1,992,085"),
            ("Net Profit",        "$860,178"),
            ("Gross Margin",      "30.2%"),
            ("Retainage Outstanding", "$142,917"),
            ("Insurance",         META_2020["insurance"]),
            ("Lien Position",     META_2020["lien_position"]),
            ("Warranty",          META_2020["warranty"]),
            ("Contract on File",  META_2020["contract_doc"]),
        ]),
        ("SCOPE", [
            ("Plumbing Units",     "288"),
            ("Total Fixtures",     "1,564"),
            ("Floors",             "P2/P1 garage + L1 retail + Mezz + L2–L5 residential"),
            ("Fixture Counts",     "288 toilets (+20 ADA), 288 lavs (+4 amenity), 257 kitchen faucets"),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR", [
            ("General Contractor", META_2020["gc_full"]),
            ("GC Project Manager", META_2020["gc_pm_phase"]),
            ("GC Superintendent",  META_2020["gc_sup"]),
            ("GC Project Engineer", META_2020["gc_pe"]),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2020["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2020["owp_trim_foreman"]),
            ("OWP Signatory",       META_2020["owp_signatory"]),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record",    META_2020["owner"]),
            ("Developer",          META_2020["developer"]),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",          META_2020["architect"]),
            ("Structural Engineer", META_2020["structural"]),
            ("MEP / Plumbing Engineer", META_2020["mep_engineer"]),
            ("Acoustical",         META_2020["acoustical"]),
            ("Civil Engineer",     META_2020["civil"]),
            ("Landscape",          META_2020["landscape"]),
            ("ADA Consultant",     META_2020["ada_consultant"]),
            ("Interior Designer",  META_2020["interior"]),
        ]),
        ("CHANGE-EVENT META", [
            ("Total CO/COR/Backcharge events", "13"),
            ("Executed COs",         "9"),
            ("Unexecuted CORs",      "2"),
            ("Backcharges",          "1"),
            ("RFIs",                 "56"),
            ("Submittals",           "206"),
            ("POs",                  "279"),
            ("Net CO Cost Impact",   "$127,362"),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",             "2020 Job Detail Report.pdf (352 pages, Apr 03 2026 run)"),
            ("Parsed Data",         "2020_data.json"),
            ("Dashboard Arrays",    "2020_dashboard_arrays.json (16 arrays from index.html PROJECTS['2020'])"),
            ("Sam Gold Standard",   "JCR Test Labels — Job 2020.xlsx"),
            ("GDrive Folder",       "2020-SRM, Bellevue @ Main/ (Job Books - Completed Jobs)"),
            ("GDrive Status",       META_2020["gdrive_status"]),
        ]),
    ]

    r = 5
    for section_name, fields in sections:
        ws.cell(row=r, column=2, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=2).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2).border = BORDER
        r += 1
        for label, val in fields:
            ws.cell(row=r, column=2, value=label).font = INK_BOLD
            ws.cell(row=r, column=3, value=val).font = INK
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
            ws.cell(row=r, column=2).border = BORDER
            ws.cell(row=r, column=3).border = BORDER
            ws.cell(row=r, column=3).alignment = WRAP
            r += 1
        r += 1  # blank row between sections

    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 30


# ============================================================================
# Main
# ============================================================================
def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)

    data = json.loads(JSON_FILE.read_text())
    dash = json.loads(DASH_FILE.read_text())

    print("→ Adding Job Info tab (key players + project metadata)...")
    add_job_info_tab(wb)

    print("→ Patching Overview team block...")
    patch_overview_team_block(wb["Overview"])

    print("→ Patching Vendors tab (was empty)...")
    patch_vendors(wb["Vendors"], dash)

    print("→ Patching Predictive Signals tab (was placeholder)...")
    patch_predictive_signals(wb["Predictive Signals"], dash)

    print("→ Patching Insights tab (was 2 of 10)...")
    patch_insights(wb["Insights"], dash)

    print("→ Patching Change Log tab...")
    patch_change_log(wb["Change Log"], dash)

    print("→ Patching Root Cause Analysis tab...")
    patch_rca(wb["Root Cause Analysis"], dash)

    print("→ Patching Billing & SOV tab...")
    patch_billing_sov(wb["Billing & SOV"], dash)

    print("→ Patching Cost Breakdown tab...")
    patch_cost_breakdown(wb["Cost Breakdown"], dash)

    print("→ Patching PO Commitments tab...")
    patch_po_commitments(wb["PO Commitments"], dash, data)

    print("→ Patching Material tab (vendor section header)...")
    patch_material(wb["Material"], dash, data)

    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
