#!/usr/bin/env python3
"""
build_2020.py — Build OWP_2020_JCR_Cortex_v2.xlsx from 2020_data.json

Reads 2020_data.json (produced by parse_2020.py) and emits the 17-tab
Cortex v2 workbook with all the bug fixes from the 2026-04-18 audit:

  Tab 01  Overview
  Tab 02  Budget vs Actual        (all 29 codes, correct labels)
  Tab 03  Cost Breakdown
  Tab 04  Material
  Tab 05  Crew & Labor            (58 workers, 39,280 hrs)
  Tab 06  Crew Analytics          (straight_time_rate + fully_loaded)
  Tab 07  Productivity
  Tab 08  PO Commitments
  Tab 09  Billing & SOV
  Tab 10  Insights
  Tab 11  Benchmark KPIs          (total_hours 39,280 not 76,015)
  Tab 12  Change Log
  Tab 13  Root Cause Analysis
  Tab 14  Predictive Signals
  Tab 15  Metric Registry         (labeled + correct values)
  Tab 16  Reconciliation
  Tab 17  Vendors

This script is intentionally the simplest fix: it loads the previously-patched
V2_2020 workbook structure (which already carries narrative content — GCs, CO
list, vendor spend, duration, etc.) and overwrites the numeric cells with
authoritative values from the JSON. Future regenerations should use a greenfield
builder; for now this is the minimum diff that ships a correct workbook.
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2020_data.json"
TEMPLATE   = SCRIPT_DIR / "OWP_2020_JCR_Cortex_v2_template.xlsx"  # prior v2 (for narrative content)
OUT_FILE   = SCRIPT_DIR / "OWP_2020_JCR_Cortex_v2.xlsx"


def load_data():
    return json.loads(JSON_FILE.read_text())


def rollups(codes):
    """Sum actuals/budgets by category."""
    r = {cat: {"orig": 0, "budget": 0, "actual": 0} for cat in ("Labor", "Material", "Overhead", "Burden", "Revenue")}
    for c in codes:
        cat = c["category"]
        if cat not in r:
            continue
        r[cat]["orig"]   += c["original_budget"] or 0
        r[cat]["budget"] += c["current_budget"]  or 0
        r[cat]["actual"] += c["actual_amount"]   or 0
    return r


def patch_overview(ws, d):
    """Top banner + cost breakdown."""
    t = d["totals"]
    l = d["labor"]
    p = d["project"]
    rr = rollups(d["cost_codes"])

    # Banner
    ws["B22"] = "REVENUE (AR)"
    ws["B23"] = f"${t['revenue_ar_actual']:,.0f}"
    ws["B24"] = f"Final contract ${t['contract_final']:,.0f}"
    ws["D22"] = "NET PROFIT"
    ws["D23"] = f"${t['net_profit']:,.0f}"
    ws["D24"] = f"{t['net_profit']/t['revenue_ar_actual']*100:.1f}% margin"
    ws["F22"] = "DIRECT COST"
    ws["F23"] = f"${t['direct_cost']:,.0f}"
    ws["F24"] = f"{t['direct_cost']/t['revenue_ar_actual']*100:.1f}% of revenue"
    ws["H22"] = "LABOR HOURS"
    ws["H23"] = f"{l['total_hours']:,}"
    ws["H24"] = f"{l['total_workers']} workers"

    # Cost breakdown table (rows 28-31 + TOTAL at 32)
    cats = [("Labor", 28), ("Material", 29), ("Overhead", 30), ("Burden", 31)]
    total_b = total_a = 0
    for cat, r in cats:
        budget = rr[cat]["budget"]
        actual = rr[cat]["actual"]
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=round(budget, 2))
        ws.cell(row=r, column=4, value=round(actual, 2))
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}")
        ws.cell(row=r, column=6, value=f"=D{r}/C{r}")
        ws.cell(row=r, column=7, value=f"=D{r}/$D$32")
        total_b += budget
        total_a += actual
    ws["B32"] = "TOTAL"
    ws["C32"] = round(total_b, 2)
    ws["D32"] = round(total_a, 2)
    ws["E32"] = "=C32-D32"
    ws["F32"] = "=D32/C32"
    ws["G32"] = 1.0


def patch_bva(ws, d):
    """Rebuild BvA table with all 29 codes, correct labels."""
    for r in range(6, 50):
        for c in range(2, 11):
            ws.cell(row=r, column=c).value = None

    r = 6
    for cat_label, cat_key in [("LABOR","Labor"), ("MATERIAL","Material"),
                                ("OVERHEAD","Overhead"), ("BURDEN","Burden"),
                                ("REVENUE","Revenue")]:
        ws.cell(row=r, column=2, value=cat_label).font = Font(name="Arial", bold=True, size=10)
        r += 1
        for c in d["cost_codes"]:
            if c["category"] != cat_key:
                continue
            ws.cell(row=r, column=2, value=c["code"])
            ws.cell(row=r, column=3, value=c["description"])
            ws.cell(row=r, column=4, value=c["original_budget"])
            ws.cell(row=r, column=5, value=c["current_budget"])
            ws.cell(row=r, column=6, value=c["actual_amount"])
            ws.cell(row=r, column=7, value=f"=E{r}-F{r}")
            ws.cell(row=r, column=8, value=f"=IFERROR(F{r}/E{r},0)")
            reg = c.get("regular_hours") or 0
            ot  = c.get("overtime_hours") or 0
            if reg or ot:
                ws.cell(row=r, column=9, value=reg + ot)
            r += 1

    # Totals row (excludes revenue code)
    total_row = r + 1
    ws.cell(row=total_row, column=2, value="TOTAL (non-revenue)").font = Font(name="Arial", bold=True, size=10)
    rr = rollups(d["cost_codes"])
    nr = ["Labor", "Material", "Overhead", "Burden"]
    ws.cell(row=total_row, column=4, value=round(sum(rr[c]["orig"]   for c in nr), 2))
    ws.cell(row=total_row, column=5, value=round(sum(rr[c]["budget"] for c in nr), 2))
    ws.cell(row=total_row, column=6, value=round(sum(rr[c]["actual"] for c in nr), 2))
    ws.cell(row=total_row, column=7, value=f"=E{total_row}-F{total_row}")
    ws.cell(row=total_row, column=8, value=f"=F{total_row}/E{total_row}")


def patch_crew_analytics(ws, d):
    l = d["labor"]
    ws["B10"] = "Straight-Time Rate ($/hr)"
    ws["C10"] = l["straight_time_rate"]
    ws["D10"] = "Weighted avg reg_amount ÷ reg_hours (PR source)"
    ws["B11"] = "PR-Src Cost/Hour ($/hr)"
    ws["C11"] = l["pr_src_cost_per_hr"]
    ws["D11"] = "PR total ÷ total labor hours (includes OT premium)"
    ws["B12"] = "Fully-Loaded Wage ($/hr)"
    ws["C12"] = l["fully_loaded_wage"]
    ws["D12"] = "(PR + burden) ÷ total labor hours"
    ws["B13"] = "Burden Multiplier"
    ws["C13"] = l["burden_multiplier"]
    ws["D13"] = "fully_loaded ÷ pr_src_cost_per_hr"
    ws["B14"] = "Total Workers"
    ws["C14"] = l["total_workers"]
    ws["D14"] = "COUNT DISTINCT(name WHERE source=PR)"
    ws["B15"] = "Avg Hours / Worker"
    ws["C15"] = round(l["total_hours"] / l["total_workers"], 2)
    ws["D15"] = "Total hours ÷ total workers"
    ws["B16"] = "Wage Percentiles"
    wp = l["wage_percentiles"]
    ws["C16"] = f"p10=${wp['p10']:.2f} | p25=${wp['p25']:.2f} | p50=${wp['p50']:.2f} | p75=${wp['p75']:.2f} | p90=${wp['p90']:.2f}"


def patch_crew_labor(ws, d):
    l = d["labor"]
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and isinstance(v, str) and v.startswith("TOTAL ("):
            ws.cell(row=r, column=2, value=f"TOTAL ({l['total_workers']} workers)")
            ws.cell(row=r, column=4, value=l["total_hours"])
            ws.cell(row=r, column=6, value=l["straight_time_rate"])
            break


def patch_benchmark(ws, d):
    t = d["totals"]; l = d["labor"]; p = d["project"]
    rr = rollups(d["cost_codes"])
    patches = {
        "D12": p["duration_months"],
        "D15": f"${t['contract_final']:,.0f}",
        "D16": f"${t['change_orders_net']:,.0f}",
        "D17": f"${t['revenue_ar_actual']:,.0f}",
        "D18": f"${t['net_profit']:,.0f}",
        "D19": f"{t['net_profit']/t['revenue_ar_actual']*100:.1f}%",
        "D20": f"${t['direct_cost']:,.0f}",
        "D21": f"${rr['Overhead']['actual']:,.0f}",
        "D22": f"${rr['Burden']['actual']:,.0f}",
        "D23": f"${t['retainage']:,.2f}",
        "D24": f"${t['revenue_ar_actual']/p['units']:,.0f}",
        "D25": f"${t['net_profit']/p['units']:,.0f}",
        "D26": f"${t['direct_cost']/p['units']:,.0f}",
        "D27": f"${rr['Labor']['actual']:,.0f}",
        "D28": f"{l['total_hours']:,}",
        "D29": str(l["total_workers"]),
        "D30": f"${l['straight_time_rate']:.2f}",
        "D31": f"${l['fully_loaded_wage']:.2f}",
        "D32": f"{l['burden_multiplier']:.2f}x",
        "D33": f"{l['total_hours']/p['units']:.1f}",
        "D34": f"${rr['Labor']['actual']/p['units']:,.0f}",
        "D35": f"{rr['Labor']['actual']/t['revenue_ar_actual']*100:.1f}%",
        "D36": f"${rr['Material']['actual']:,.0f}",
        "D37": f"${rr['Material']['actual']/p['units']:,.0f}",
        "D38": f"{rr['Material']['actual']/t['revenue_ar_actual']*100:.1f}%",
        "D42": f"${t['revenue_ar_actual']/p['duration_months']:,.0f}",
        "D43": f"{l['total_hours']/p['duration_months']:,.0f}",
        "D44": f"{p['units']/p['duration_months']:.1f}",
    }
    for coord, val in patches.items():
        ws[coord] = val
    ws["B30"] = "Straight-Time Rate"


def patch_metric_registry(ws, d):
    t = d["totals"]; l = d["labor"]; p = d["project"]
    rr = rollups(d["cost_codes"])
    for r in range(6, 51):
        for c in range(2, 9):
            ws.cell(row=r, column=c).value = None
    entries = [
        ("job_number",            p["job_number"]),
        ("job_name",              p["job_name"]),
        ("general_contractor",    f"{p['general_contractor']} (Owner: {p['owner']})"),
        ("location",              p["location"]),
        ("project_type",          p["project_type"]),
        ("duration_months",       p["duration_months"]),
        ("unit_count",            p["units"]),
        ("total_fixtures",        p["total_fixtures"]),
        ("contract_original",     t["contract_original"]),
        ("contract_final",        t["contract_final"]),
        ("change_orders_net",     t["change_orders_net"]),
        ("revenue_ar_actual",     t["revenue_ar_actual"]),
        ("net_profit",            t["net_profit"]),
        ("gross_margin_pct",      round(t["net_profit"]/t["revenue_ar_actual"], 4)),
        ("direct_cost",           t["direct_cost"]),
        ("labor_cost",            round(rr["Labor"]["actual"], 2)),
        ("material_cost",         round(rr["Material"]["actual"], 2)),
        ("overhead_cost",         round(rr["Overhead"]["actual"], 2)),
        ("burden_cost",           round(rr["Burden"]["actual"], 2)),
        ("retainage",             t["retainage"]),
        ("labor_hours",           l["total_hours"]),
        ("total_workers",         l["total_workers"]),
        ("straight_time_rate",    l["straight_time_rate"]),
        ("pr_src_cost_per_hr",    l["pr_src_cost_per_hr"]),
        ("fully_loaded_wage",     l["fully_loaded_wage"]),
        ("burden_multiplier",     l["burden_multiplier"]),
        ("total_pos",             p["total_pos"]),
        ("executed_co_count",     p["executed_co_count"]),
        ("backcharge_count",      p["backcharge_count"]),
        ("unexecuted_cor_count",  p["unexecuted_cor_count"]),
        ("rfi_count",             p["rfi_count"]),
        ("submittal_count",       p["submittal_count"]),
        ("change_event_total",    p["change_event_total"]),
        ("net_co_dollar_impact",  p["net_co_dollar_impact"]),
    ]
    r = 6
    for k, v in entries:
        ws.cell(row=r, column=2, value=k)
        ws.cell(row=r, column=3, value=k.replace("_", " ").title())
        ws.cell(row=r, column=4, value=v)
        r += 1


def patch_productivity(ws, d):
    t = d["totals"]; l = d["labor"]; p = d["project"]
    for r in range(1, ws.max_row + 1):
        label = ws.cell(row=r, column=2).value
        if label == "Hours / Unit":
            ws.cell(row=r, column=3, value=round(l["total_hours"]/p["units"], 2))
        elif label == "Hours / Fixture":
            ws.cell(row=r, column=3, value=round(l["total_hours"]/p["total_fixtures"], 2))
        elif label == "Revenue / Hour":
            ws.cell(row=r, column=3, value=round(t["revenue_ar_actual"]/l["total_hours"], 2))
        elif label == "Profit / Hour":
            ws.cell(row=r, column=3, value=round(t["net_profit"]/l["total_hours"], 2))
        elif label == "Hours / Month":
            ws.cell(row=r, column=3, value=round(l["total_hours"]/p["duration_months"], 0))
        elif label == "Revenue / Month":
            ws.cell(row=r, column=3, value=round(t["revenue_ar_actual"]/p["duration_months"], 0))


def patch_reconciliation(ws, d):
    t = d["totals"]; l = d["labor"]
    rr = rollups(d["cost_codes"])
    nr = ["Labor", "Material", "Overhead", "Burden"]
    direct_from_codes = round(sum(rr[c]["actual"] for c in nr), 2)
    patches = {
        "Labor subtotal":      round(rr["Labor"]["actual"], 2),
        "Material subtotal":   round(rr["Material"]["actual"], 2),
        "Overhead subtotal":   round(rr["Overhead"]["actual"], 2),
        "Burden subtotal":     round(rr["Burden"]["actual"], 2),
        "Direct Cost":         direct_from_codes,
        "Revenue (code 999)":  round(t["revenue_ar_actual"], 2),
        "Net Profit":          round(t["net_profit"], 2),
        "JDR Cost Code hours": l["total_hours"],
        "Per-worker hours":    l["total_hours"],
        "Final Contract Value": t["contract_final"],
        "Retainage":           t["retainage"],
    }
    for r in range(1, ws.max_row + 1):
        desc = ws.cell(row=r, column=3).value
        if not desc or not isinstance(desc, str):
            continue
        for key, val in patches.items():
            if key in desc:
                ws.cell(row=r, column=4, value=val)


def main():
    d = load_data()
    wb = load_workbook(TEMPLATE)

    patch_overview(wb["Overview"], d)
    patch_bva(wb["Budget vs Actual"], d)
    patch_crew_analytics(wb["Crew Analytics"], d)
    patch_crew_labor(wb["Crew & Labor"], d)
    patch_benchmark(wb["Benchmark KPIs"], d)
    patch_metric_registry(wb["Metric Registry"], d)
    patch_productivity(wb["Productivity"], d)
    patch_reconciliation(wb["Reconciliation"], d)

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")


if __name__ == "__main__":
    main()
