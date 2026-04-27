"""Single canonical builder for Cortex v2 JCR workbooks.

Inputs:
  - schema.yaml  (cortex_v2_schema.yaml — the contract)
  - job_data.json  (all data for one project, structured per schema)

Output:
  - OWP_<JOB>_JCR_Cortex_v2.xlsx (schema-compliant, version-stamped)

Usage:
  python cortex_builder.py <job_data.json> [--out <path>]
"""
import json as _json
import json, sys, argparse, re
from pathlib import Path
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCHEMA_PATH = Path(__file__).parent.parent / "schema" / "cortex_v2_schema.yaml"

# ----- styles -----
TITLE = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBTITLE = Font(name="Calibri", size=10, italic=True, color="595959")
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD = Font(name="Calibri", size=10, bold=True)
BODY = Font(name="Calibri", size=10)
SRC = Font(name="Calibri", size=9, italic=True, color="808080")
F_TITLE = PatternFill("solid", fgColor="1F4E78")
F_HDR = PatternFill("solid", fgColor="2E75B6")
F_ALT = PatternFill("solid", fgColor="DDEBF7")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
_b = Side(style="thin", color="BFBFBF")
BRD = Border(left=_b, right=_b, top=_b, bottom=_b)


def load_schema(path=SCHEMA_PATH):
    return yaml.safe_load(Path(path).read_text())


def put(ws, addr, val, font=BODY, fill=None, align=None, fmt=None, border=None):
    cell = ws[addr] if isinstance(addr, str) else ws.cell(row=addr[0], column=addr[1])
    cell.value = val
    if font: cell.font = font
    if fill: cell.fill = fill
    if align: cell.alignment = align
    if fmt: cell.number_format = fmt
    if border: cell.border = border
    return cell


def fmt_for(dtype, schema):
    g = schema["global"]
    return {
        "currency": g["currency_nf"],
        "percent": g["percent_nf"],
        "date_iso": "yyyy-mm-dd",
    }.get(dtype)


def align_for(dtype):
    return {
        "currency": RIGHT, "number": RIGHT, "percent": RIGHT,
        "date_iso": CENTER, "string": LEFT, "any": LEFT, "list": LEFT,
    }.get(dtype, LEFT)


def write_title(ws, title, subtitle, schema_version, cols=10):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    put(ws, "A1", f"[{schema_version}] {title}", TITLE, F_TITLE, CENTER)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    put(ws, "A2", subtitle, SUBTITLE, align=CENTER)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def write_kv_sheet(ws, spec, data, schema):
    """Write a label/value layout sheet. `data` is a dict of label→value."""
    key_col = ord(spec["key_column"]) - ord("A") + 1
    val_col = ord(spec["value_column"]) - ord("A") + 1
    src_col = ord(spec.get("source_column", "F")) - ord("A") + 1

    header_row = spec.get("header_row", 5)
    put(ws, (header_row, key_col), spec["name"].upper(), BOLD, F_ALT, LEFT)
    put(ws, (header_row, val_col), "", fill=F_ALT)
    put(ws, (header_row, src_col), "SOURCE", BOLD, F_ALT, LEFT)

    r = header_row + 1
    for lbl in spec.get("required_labels") or []:
        label = lbl["label"]
        entry = data.get(label) if isinstance(data, dict) else None
        value = entry.get("value") if isinstance(entry, dict) else entry
        source = entry.get("source") if isinstance(entry, dict) else None
        put(ws, (r, key_col), label, BOLD, align=LEFT)
        put(ws, (r, val_col), value if value is not None else "", align=LEFT)
        # merge value across a few cols
        try:
            ws.merge_cells(start_row=r, start_column=val_col,
                           end_row=r, end_column=src_col - 1)
        except Exception:
            pass
        put(ws, (r, src_col), source or "", SRC, align=LEFT)
        r += 1
    # also emit any extra keys provided in data but not in required_labels
    extras = [k for k in (data or {}) if k not in {l["label"] for l in spec.get("required_labels") or []}]
    for label in extras:
        entry = data[label]
        value = entry.get("value") if isinstance(entry, dict) else entry
        source = entry.get("source") if isinstance(entry, dict) else None
        put(ws, (r, key_col), label, BOLD, align=LEFT)
        put(ws, (r, val_col), value if value is not None else "", align=LEFT)
        put(ws, (r, src_col), source or "", SRC, align=LEFT)
        r += 1

    set_widths(ws, {1: 2, 2: 28, 3: 22, 4: 18, 5: 18, 6: 32, 7: 14, 8: 14, 9: 14, 10: 14})


def write_table_sheet(ws, spec, rows, schema):
    """Write a columnar table. `rows` is a list of dicts keyed by canonical column name."""
    header_row = spec.get("header_row", schema["global"]["header_row_default"])
    cols = spec["columns"]

    # Header row
    for i, col in enumerate(cols, start=2):  # start at column B
        put(ws, (header_row, i), col["name"], HDR, F_HDR, CENTER, border=BRD)

    r = header_row + 1
    totals = {c["name"]: 0.0 for c in cols if c["dtype"] in ("currency", "number")}
    totals_active = spec.get("totals_row", False)

    for row in rows or []:
        for i, col in enumerate(cols, start=2):
            val = row.get(col["name"])
            # Coerce lists to joined strings (Excel can't store lists in a cell)
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            elif isinstance(val, dict):
                val = json.dumps(val, default=str)
            f = fmt_for(col["dtype"], schema)
            a = align_for(col["dtype"])
            put(ws, (r, i), val if val is not None else "",
                fmt=f, align=a, border=BRD)
            if totals_active and col["dtype"] in ("currency", "number") and isinstance(val, (int, float)):
                totals[col["name"]] += val
        r += 1

    if totals_active and rows:
        put(ws, (r, 2), "TOTAL", BOLD, F_ALT, LEFT, border=BRD)
        for i, col in enumerate(cols[1:], start=3):
            if col["dtype"] in ("currency", "number"):
                f = fmt_for(col["dtype"], schema)
                put(ws, (r, i), totals[col["name"]], BOLD, F_ALT, RIGHT, fmt=f, border=BRD)
            else:
                put(ws, (r, i), "", fill=F_ALT, border=BRD)

    # Column widths
    widths = {1: 2}
    for i, col in enumerate(cols, start=2):
        w = {"currency": 16, "number": 12, "percent": 12, "date_iso": 14,
             "string": 24, "list": 24, "any": 18}.get(col["dtype"], 16)
        widths[i] = w
    set_widths(ws, widths)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=2).coordinate


def build(job_data, schema, out_path):
    wb = Workbook()
    # Remove default
    wb.remove(wb.active)

    version = schema["schema_version"]
    job = job_data.get("job", "????")
    project_name = (job_data.get("overview", {}).get("project_name_sage") or {}).get("value", "")
    if not project_name:
        project_name = job_data.get("project_name", "")

    for tab_spec in schema["tabs"]:
        tname = tab_spec["name"]
        ws = wb.create_sheet(tname)
        write_title(ws, f"Job #{job} — {project_name} · {tname}",
                    f"Schema {version} · canonical v2 builder · single source of truth",
                    version)
        key = tname.lower().replace(" & ", "_").replace(" ", "_").replace("&", "and")
        # Normalize section key
        section_key = {
            "Overview": "overview",
            "Budget vs Actual": "budget_vs_actual",
            "Cost Breakdown": "cost_breakdown",
            "Material": "material",
            "Crew & Labor": "crew_labor",
            "Crew Analytics": "crew_analytics",
            "Productivity": "productivity",
            "PO Commitments": "po_commitments",
            "Billing & SOV": "billing_sov",
            "Insights": "insights",
            "Benchmark KPIs": "benchmark_kpis",
            "Change Log": "change_log",
            "Root Cause Analysis": "root_cause_analysis",
            "Predictive Signals": "predictive_signals",
            "Metric Registry": "metric_registry",
            "Reconciliation": "reconciliation",
            "Vendors": "vendors",
        }[tname]
        data = job_data.get(section_key)

        if tab_spec["kind"] == "kv_sheet":
            write_kv_sheet(ws, tab_spec, data or {}, schema)
        elif tab_spec["kind"] == "table":
            write_table_sheet(ws, tab_spec, data or [], schema)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("job_json", help="path to job_data_XXXX.json")
    ap.add_argument("--out", help="output xlsx path")
    ap.add_argument("--schema", default=str(SCHEMA_PATH))
    args = ap.parse_args()

    schema = load_schema(args.schema)
    job_data = json.loads(Path(args.job_json).read_text())
    job = job_data.get("job", Path(args.job_json).stem.split("_")[-1])
    out = Path(args.out) if args.out else Path(f"OWP_{job}_JCR_Cortex_v2.xlsx")
    build(job_data, schema, out)
    print(f"Built {out}")


if __name__ == "__main__":
    main()
