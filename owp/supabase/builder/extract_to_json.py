"""Extract existing Cortex v2 XLSX -> job_data.json, applying schema-driven
column-name normalization and value sanitization.

This is a one-way bootstrap: read the current (inconsistent) files, emit
canonical JSON conforming to cortex_v2_schema.yaml. cortex_builder.py
then reads that JSON and re-emits structurally-identical XLSX.

Usage:
  python extract_to_json.py <src.xlsx> [--out <path>]
  python extract_to_json.py --all <folder>
"""
import json, re, sys, argparse
from pathlib import Path
from collections import defaultdict
import yaml
from openpyxl import load_workbook

SCHEMA_PATH = Path(__file__).parent.parent / "schema" / "cortex_v2_schema.yaml"


def _norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""


def _scan_header(ws, expected_row, max_scan=12):
    def tc(r):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 20) + 1)]
        return sum(1 for v in vals if isinstance(v, str) and v.strip() and any(ch.isalpha() for ch in v))
    if expected_row and tc(expected_row) >= 3:
        return expected_row
    best_r, best_n = None, 0
    for r in range(1, min(ws.max_row, max_scan) + 1):
        n = tc(r)
        if n > best_n: best_r, best_n = r, n
    return best_r


def _get_row(ws, r, n=20):
    return [ws.cell(row=r, column=c).value for c in range(1, n + 1)]


def _build_alias_index(cols_spec):
    """name -> set of normalized aliases (incl. canonical name)."""
    idx = {}
    for col in cols_spec:
        names = [col["name"]] + (col.get("aliases") or [])
        idx[col["name"]] = {_norm(n) for n in names}
    return idx


def _map_headers_to_columns(headers, alias_idx):
    """Given row of headers, return dict canonical_col_name -> col_index (0-based)."""
    out = {}
    for i, h in enumerate(headers):
        hn = _norm(h)
        if not hn: continue
        for canonical, aliases in alias_idx.items():
            if hn in aliases and canonical not in out:
                out[canonical] = i
                break
    return out


# ----- sanitizers -----
def sanitize_val(val, col_spec, sanitizers_cfg):
    if val is None: return None
    s = val
    for s_name in col_spec.get("sanitize") or []:
        cfg = sanitizers_cfg.get(s_name, {})
        kind = cfg.get("kind")
        if isinstance(s, str):
            if kind == "regex_sub":
                s = re.sub(cfg["pattern"], cfg.get("repl", ""), s).strip()
            elif kind == "map":
                s = cfg.get("mapping", {}).get(s.strip(), s.strip())
            elif kind == "regex_extract":
                m = re.search(cfg["pattern"], s)
                if m: s = m.group(0)
                elif cfg.get("on_miss") is None: s = None
            elif kind == "regex_split":
                parts = re.split(cfg["pattern"], s)
                if cfg.get("strip"): parts = [p.strip() for p in parts]
                if cfg.get("drop_empty"): parts = [p for p in parts if p]
                s = parts
    return s


def _coerce(val, dtype):
    if val is None or val == "": return None
    if dtype == "currency" or dtype == "number":
        if isinstance(val, (int, float)): return val
        if isinstance(val, str):
            t = re.sub(r"[^\d\.\-]", "", val)
            try: return float(t) if t else None
            except: return None
    if dtype == "percent":
        if isinstance(val, (int, float)): return val if val <= 1 else val / 100
        if isinstance(val, str):
            t = val.replace("%", "").strip()
            try: return float(t) / 100
            except: return None
    if dtype == "date_iso":
        if hasattr(val, "isoformat"): return val.isoformat()[:10]
        if isinstance(val, str):
            m = re.search(r"\d{4}-\d{2}-\d{2}", val)
            return m.group(0) if m else val
    return val


def extract_table(ws, tab_spec, cohort_cfg, sanitizers_cfg):
    """Extract a table-kind tab into list[dict] keyed by canonical col name."""
    exp_hdr = tab_spec.get("header_row", 5)
    ovr = (cohort_cfg.get("header_row_offsets") or {}).get(tab_spec["name"])
    if ovr: exp_hdr = ovr
    hdr_row = _scan_header(ws, exp_hdr)
    if hdr_row is None: return []

    headers = _get_row(ws, hdr_row, 20)
    alias_idx = _build_alias_index(tab_spec["columns"])
    col_map = _map_headers_to_columns(headers, alias_idx)

    rows = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        raw = _get_row(ws, r, 20)
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in raw):
            continue
        # skip obvious total rows (first non-empty cell == "TOTAL")
        first_nonempty = next((v for v in raw if v is not None and str(v).strip()), None)
        if isinstance(first_nonempty, str) and first_nonempty.strip().upper() in ("TOTAL", "GRAND TOTAL"):
            continue

        row = {}
        for col in tab_spec["columns"]:
            cidx = col_map.get(col["name"])
            val = raw[cidx] if cidx is not None and cidx < len(raw) else None
            val = sanitize_val(val, col, sanitizers_cfg)
            val = _coerce(val, col["dtype"])
            row[col["name"]] = val
        # drop fully-null rows
        if any(v not in (None, "", []) for v in row.values()):
            rows.append(row)
    return rows


def extract_kv(ws, tab_spec):
    """Extract a label/value sheet. Returns dict label -> {value, source}."""
    key_col = ord(tab_spec.get("key_column", "B")) - ord("A") + 1
    val_col = ord(tab_spec.get("value_column", "C")) - ord("A") + 1
    src_col = ord(tab_spec.get("source_column", "F")) - ord("A") + 1
    label_aliases = tab_spec.get("label_aliases", {})
    # Build reverse alias map: normalized alias -> canonical label
    rev = {}
    for canonical, aliases in label_aliases.items():
        for a in [canonical] + (aliases or []):
            rev[_norm(a)] = canonical

    out = {}
    for r in range(1, ws.max_row + 1):
        raw_label = ws.cell(row=r, column=key_col).value
        if not raw_label or not str(raw_label).strip(): continue
        label_norm = _norm(raw_label)
        canonical = rev.get(label_norm, str(raw_label).strip())
        # only include if in required_labels OR it's a pass-through extra
        required = {l["label"] for l in tab_spec.get("required_labels") or []}
        # skip cells that are visually bands/titles
        if not canonical or canonical.upper() == canonical and len(canonical) > 30:
            continue
        value = ws.cell(row=r, column=val_col).value
        # coalesce merged cells: if C is blank but D has value
        if value is None:
            for extra in range(val_col + 1, src_col):
                v = ws.cell(row=r, column=extra).value
                if v is not None: value = v; break
        source = ws.cell(row=r, column=src_col).value
        if value is None and source is None:
            continue
        out[canonical] = {"value": value, "source": source}
    return out


def extract_file(path, schema):
    job_m = re.search(r"OWP_(\d+)", Path(path).name)
    job = job_m.group(1) if job_m else Path(path).stem

    legacy = set(str(x) for x in schema["cohorts"]["legacy_v1"]["files"])
    current = set(str(x) for x in schema["cohorts"]["current_v2"]["files"])
    cohort = "legacy_v1" if job in legacy else ("current_v2" if job in current else "unknown")
    cohort_cfg = schema["cohorts"].get(cohort, {})

    wb = load_workbook(path, data_only=True, read_only=False)
    sanitizers_cfg = schema.get("sanitizers", {})

    data = {
        "schema_version": schema["schema_version"],
        "job": job,
        "source_file": str(path),
        "cohort": cohort,
    }

    section_map = {
        "Overview": "overview",
        "Budget vs Actual": "budget_vs_actual",
        "Cost Breakdown": "cost_breakdown",
        "Material": "material",
        "Crew & Labor": "crew_labor",
        "Crew Analytics": "crew_analytics",
        "Productivity": "productivity",
        "PO Commitments": "po_commitments",
        "Billing & SOV": "billing_sov",
        "Insights": "insights",
        "Benchmark KPIs": "benchmark_kpis",
        "Change Log": "change_log",
        "Root Cause Analysis": "root_cause_analysis",
        "Predictive Signals": "predictive_signals",
        "Metric Registry": "metric_registry",
        "Reconciliation": "reconciliation",
        "Vendors": "vendors",
    }
    for tab_spec in schema["tabs"]:
        tname = tab_spec["name"]
        if tname not in wb.sheetnames:
            data[section_map[tname]] = [] if tab_spec["kind"] == "table" else {}
            continue
        ws = wb[tname]
        if tab_spec["kind"] == "table":
            data[section_map[tname]] = extract_table(ws, tab_spec, cohort_cfg, sanitizers_cfg)
        elif tab_spec["kind"] == "kv_sheet":
            data[section_map[tname]] = extract_kv(ws, tab_spec)
    wb.close()

    # Pull project_name for convenience
    ov = data.get("overview") or {}
    pn = ov.get("project_name_sage")
    if isinstance(pn, dict): pn = pn.get("value")
    data["project_name"] = pn or ""

    # Compute fallbacks for Benchmark KPIs
    bk = data.get("benchmark_kpis") or {}
    billing = data.get("billing_sov") or []
    cost_breakdown = data.get("cost_breakdown") or []
    crew = data.get("crew_labor") or []
    pos = data.get("po_commitments") or []
    vendors_ = data.get("vendors") or []
    cc_rows = data.get("budget_vs_actual") or []

    def _num(v):
        return v if isinstance(v, (int, float)) else None

    def _fb(label, val):
        if bk.get(label) is None and val is not None:
            bk[label] = {"value": val, "source": "computed fallback"}

    revenue = sum([r.get("this_period") or 0 for r in billing]) or None
    _fb("revenue_total", revenue)
    if billing:
        ct = [r.get("contract_total") for r in billing if r.get("contract_total")]
        if ct:
            _fb("contract_original", ct[0])
            _fb("contract_final", ct[-1])
    direct = sum([r.get("actual") or 0 for r in cost_breakdown]) or None
    _fb("direct_cost_total", direct)
    if revenue and direct:
        _fb("net_profit", revenue - direct)
        _fb("gross_margin_pct", (revenue - direct) / revenue)
    _fb("labor_hours_total", sum([r.get("total_hours") or 0 for r in crew]) or None)
    _fb("workers_count", len(crew) or None)
    _fb("vendors_count", len(vendors_) or None)
    _fb("pos_count", len(pos) or None)
    _fb("cost_codes_count", len(cc_rows) or None)

    data["benchmark_kpis"] = bk
    return data


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", help="src xlsx file OR folder (with --all)")
    ap.add_argument("--all", action="store_true", help="treat src as folder of xlsx")
    ap.add_argument("--out-dir", default=".", help="where to write json")
    args = ap.parse_args()

    schema = yaml.safe_load(SCHEMA_PATH.read_text())
    out_dir = Path(args.out_dir); out_dir.mkdir(parents=True, exist_ok=True)

    if args.all:
        files = sorted(Path(args.src).glob("*.xlsx"))
    else:
        files = [Path(args.src)]

    for f in files:
        data = extract_file(f, schema)
        job = data["job"]
        out = out_dir / f"job_data_{job}.json"
        out.write_text(json.dumps(data, indent=2, default=str))
        print(f"Extracted {f.name} -> {out.name}")


if __name__ == "__main__":
    main()
