"""Audit every XLSX in /cortex v2 files/ against cortex_v2_schema.yaml.

Emits audit_report.md grouped by severity:
  CRITICAL — missing required columns / required labels
  SCHEMA   — wrong header-row position
  LABEL    — missing canonical data_labels
  VALUE    — garbage suffixes, non-ISO dates, enum violations
  RECON    — Reconciliation rows with status OFF/FAIL

Also writes audit_summary.json for programmatic consumption.
"""
import json, re, yaml
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook

ROOT = Path("/sessions/keen-determined-mccarthy/work/supabase")
SRC_DIR = Path("/sessions/keen-determined-mccarthy/mnt/cortex-mockup/owp/cortex v2 files")
SCHEMA_PATH = ROOT / "schema/cortex_v2_schema.yaml"
OUT_MD = ROOT / "audit_report.md"
OUT_JSON = ROOT / "audit_summary.json"

SCHEMA = yaml.safe_load(SCHEMA_PATH.read_text())
LEGACY = set(str(x) for x in SCHEMA["cohorts"]["legacy_v1"]["files"])
CURRENT = set(str(x) for x in SCHEMA["cohorts"]["current_v2"]["files"])

def norm(s):
    return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""

def match_alias(cell_val, canonical_name, aliases):
    if cell_val is None:
        return False
    cv = norm(cell_val)
    candidates = [canonical_name] + (aliases or [])
    return any(norm(c) == cv for c in candidates)

def scan_header_row(ws, expected_row, max_scan=12):
    """Return the row index whose cells best match header text (non-numeric).
    If expected_row has >=3 text cells, trust it; else scan."""
    def row_text_count(r):
        vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 20) + 1)]
        return sum(1 for v in vals if isinstance(v, str) and v.strip() and any(ch.isalpha() for ch in v))
    if expected_row and row_text_count(expected_row) >= 3:
        return expected_row
    best_r, best_n = None, 0
    for r in range(1, min(ws.max_row, max_scan) + 1):
        n = row_text_count(r)
        if n > best_n:
            best_r, best_n = r, n
    return best_r

def get_row(ws, r, n_cols=None):
    n = n_cols or ws.max_column
    return [ws.cell(row=r, column=c).value for c in range(1, n + 1)]

def audit_file(path):
    job = re.search(r"OWP_(\d+)", path.name).group(1)
    cohort = "legacy_v1" if job in LEGACY else ("current_v2" if job in CURRENT else "unknown")
    findings = []

    wb = load_workbook(path, data_only=True, read_only=False)
    sheet_names = wb.sheetnames
    schema_tabs = [t["name"] for t in SCHEMA["tabs"]]

    # STRUCTURAL: 17 tabs in order
    if sheet_names != schema_tabs:
        findings.append(("CRITICAL", "TABS", f"Tab order/set mismatch. Expected {schema_tabs}, got {sheet_names}"))

    for spec in SCHEMA["tabs"]:
        tname = spec["name"]
        if tname not in sheet_names:
            findings.append(("CRITICAL", tname, "Tab missing"))
            continue
        ws = wb[tname]
        # Resolve expected header row (apply cohort overrides)
        exp_hdr = spec.get("header_row", SCHEMA["global"]["header_row_default"])
        ovr = SCHEMA["cohorts"][cohort]["header_row_offsets"].get(tname) if cohort in SCHEMA["cohorts"] else None
        if ovr: exp_hdr = ovr

        actual_hdr = scan_header_row(ws, exp_hdr)
        if actual_hdr != exp_hdr:
            findings.append(("SCHEMA", tname, f"Header row at {actual_hdr}, expected {exp_hdr}"))

        if spec["kind"] == "table":
            if actual_hdr is None:
                findings.append(("CRITICAL", tname, "No header row detected"))
                continue
            headers = get_row(ws, actual_hdr, 20)
            hdr_texts = [norm(h) for h in headers]

            # Required columns
            for col in spec.get("columns", []):
                if col.get("required"):
                    candidates = [col["name"]] + (col.get("aliases") or [])
                    if not any(norm(c) in hdr_texts for c in candidates):
                        findings.append(("CRITICAL", tname, f"Required column missing: {col['name']} (aliases: {col.get('aliases')})"))

            # Data-value scans on known columns
            data_rows = []
            for r in range(actual_hdr + 1, ws.max_row + 1):
                row_vals = get_row(ws, r, 20)
                if all(v is None or (isinstance(v, str) and not v.strip()) for v in row_vals):
                    continue
                data_rows.append((r, row_vals))

            # Value sanitize scans
            for col in spec.get("columns", []):
                sanitizers = col.get("sanitize") or []
                if not sanitizers: continue
                # find column index by alias match
                cidx = None
                candidates = [col["name"]] + (col.get("aliases") or [])
                for i, h in enumerate(hdr_texts):
                    if h and any(norm(c) == h for c in candidates):
                        cidx = i
                        break
                if cidx is None: continue
                for r, rv in data_rows[:200]:
                    val = rv[cidx] if cidx < len(rv) else None
                    if val is None: continue
                    sval = str(val)
                    if "strip_location_suffix" in sanitizers:
                        if re.search(r'\s*\$[A-Z][A-Za-z0-9\s&]+$', sval):
                            findings.append(("VALUE", tname, f"Row {r} col '{col['name']}' has $-suffix: {sval!r}"))
                    if "extract_iso_date" in sanitizers:
                        if not re.search(r'\d{4}-\d{2}-\d{2}', sval):
                            findings.append(("VALUE", tname, f"Row {r} col '{col['name']}' non-ISO date: {sval!r}"))
                    if "normalize_tier" in sanitizers:
                        if sval.strip() in ("Sr.", "Sr"):
                            findings.append(("VALUE", tname, f"Row {r} tier '{sval}' — normalize to 'Senior'"))

            # Enum checks
            for col in spec.get("columns", []):
                enum = col.get("enum")
                if not enum: continue
                cidx = None
                candidates = [col["name"]] + (col.get("aliases") or [])
                for i, h in enumerate(hdr_texts):
                    if h and any(norm(c) == h for c in candidates):
                        cidx = i; break
                if cidx is None: continue
                bad = set()
                for r, rv in data_rows:
                    val = rv[cidx] if cidx < len(rv) else None
                    if val is None or str(val).strip() == "": continue
                    if str(val).strip() not in enum:
                        bad.add(str(val).strip())
                for b in sorted(bad):
                    findings.append(("VALUE", tname, f"Column '{col['name']}' has non-enum value: {b!r} (allowed: {enum})"))

        elif spec["kind"] == "kv_sheet":
            # Collect labels present in B column
            labels_seen = set()
            key_col_idx = ord(spec.get("key_column", "B")) - ord("A") + 1
            for r in range(1, ws.max_row + 1):
                v = ws.cell(row=r, column=key_col_idx).value
                if v: labels_seen.add(norm(v))
            for lbl in spec.get("required_labels") or []:
                if lbl.get("required"):
                    aliases = SCHEMA["tabs"][schema_tabs.index(tname)].get("label_aliases", {}).get(lbl["label"], [])
                    candidates = [lbl["label"]] + aliases
                    if not any(norm(c) in labels_seen for c in candidates):
                        fb = lbl.get("fallback")
                        sev = "LABEL" if fb else "CRITICAL"
                        findings.append((sev, tname, f"Missing required label '{lbl['label']}'" + (f" (fallback: {fb})" if fb else "")))

        # Reconciliation gate
        if spec.get("load_gate") and actual_hdr:
            headers = get_row(ws, actual_hdr, 20)
            hdr_texts = [norm(h) for h in headers]
            status_idx = None
            for i, h in enumerate(hdr_texts):
                if h == "status": status_idx = i; break
            if status_idx is not None:
                for r in range(actual_hdr + 1, ws.max_row + 1):
                    v = ws.cell(row=r, column=status_idx + 1).value
                    if v and str(v).strip().upper() in ("OFF", "FAIL"):
                        check = ws.cell(row=r, column=3).value
                        findings.append(("RECON", tname, f"Row {r} FAILED check: {check} ({v})"))

    wb.close()
    return {"job": job, "cohort": cohort, "findings": findings}

# ---------- run ----------
results = {}
for f in sorted(SRC_DIR.glob("*.xlsx")):
    print(f"Auditing {f.name}")
    results[f.name] = audit_file(f)

# ---------- write JSON ----------
OUT_JSON.write_text(json.dumps(results, indent=2))

# ---------- write Markdown ----------
sev_order = ["CRITICAL", "RECON", "SCHEMA", "LABEL", "VALUE"]
sev_counts = defaultdict(int)
by_sev = defaultdict(list)

for fname, res in results.items():
    for sev, tab, msg in res["findings"]:
        sev_counts[sev] += 1
        by_sev[sev].append((res["job"], res["cohort"], tab, msg))

lines = [
    "# Cortex v2 Audit Report",
    f"**Generated:** 2026-04-14  ·  **Files scanned:** {len(results)}  ·  **Schema:** {SCHEMA['schema_version']}",
    "",
    "## Severity summary",
    "",
    "| Severity | Count | Meaning |",
    "|---|---|---|",
    f"| CRITICAL | {sev_counts['CRITICAL']} | Required column/label missing — blocks load |",
    f"| RECON    | {sev_counts['RECON']} | Reconciliation row FAIL — blocks that project's load |",
    f"| SCHEMA   | {sev_counts['SCHEMA']} | Wrong header-row position — normalizer fixes |",
    f"| LABEL    | {sev_counts['LABEL']} | Missing canonical data_label (fallback available) |",
    f"| VALUE    | {sev_counts['VALUE']} | Garbage suffix / non-ISO date / enum violation |",
    "",
    "## Per-file summary",
    "",
    "| Job | Cohort | CRITICAL | RECON | SCHEMA | LABEL | VALUE | Total |",
    "|---|---|---|---|---|---|---|---|",
]
for fname, res in sorted(results.items(), key=lambda kv: kv[1]["job"]):
    c = defaultdict(int)
    for sev, _, _ in res["findings"]:
        c[sev] += 1
    lines.append(f"| {res['job']} | {res['cohort']} | {c['CRITICAL']} | {c['RECON']} | {c['SCHEMA']} | {c['LABEL']} | {c['VALUE']} | {sum(c.values())} |")

lines += ["", "---", ""]

for sev in sev_order:
    if sev_counts[sev] == 0: continue
    lines.append(f"## {sev} findings ({sev_counts[sev]})")
    lines.append("")
    # Group by tab for readability
    by_tab = defaultdict(list)
    for job, cohort, tab, msg in by_sev[sev]:
        by_tab[tab].append((job, cohort, msg))
    for tab in sorted(by_tab.keys()):
        lines.append(f"### `{tab}`")
        lines.append("")
        for job, cohort, msg in sorted(by_tab[tab]):
            lines.append(f"- **{job}** ({cohort}): {msg}")
        lines.append("")

OUT_MD.write_text("\n".join(lines))
print(f"\nWrote {OUT_MD}")
print(f"Wrote {OUT_JSON}")
print(f"\nSeverity totals: {dict(sev_counts)}")
