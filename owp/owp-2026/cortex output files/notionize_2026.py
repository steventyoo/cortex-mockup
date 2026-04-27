#!/usr/bin/env python3
"""Apply Notion styling to freshly-built OWP_2026 Cortex v2 workbook (17 tabs)
and inject Project Team from Job Info. Save final to portfolio folder."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = Path('/sessions/keen-determined-mccarthy/work/OWP_2026_JCR_Cortex_v2.xlsx')
OUT = Path('/sessions/keen-determined-mccarthy/work/OWP_2026_Notion_v2.xlsx')

# ── Notion tokens ─────────────────────────────────────
FONT   = "Calibri"
INK    = "191917"
INK_3  = "6B6B65"
INK_4  = "9E9E97"
BG_PG  = "FBFAF7"
BG_MUT = "F3F0E8"
BG_DIV = "E8E3D6"
CLAY   = "B5553C"
SAGE   = "6B8563"
FOREST = "3D5A3D"

PILL = {
    "ON":("DCFCE7","166534"), "UNDER":("DBEAFE","1E40AF"),
    "OVER":("FEF3C7","92400E"), "CRITICAL":("FECACA","991B1B"),
    "TIES":("DCFCE7","166534"), "OFF":("FECACA","991B1B"),
    "Mapped":("DCFCE7","166534"), "Derived":("DBEAFE","1E40AF"),
    "Computed":("E9D5FF","6B21A8"), "Extracted":("DCFCE7","166534"),
    "Target":("FEF3C7","92400E"), "HEALTHY":("DCFCE7","166534"),
    "WATCH":("FEF3C7","92400E"), "ALERT":("FECACA","991B1B"),
    "Verified":("DCFCE7","166534"), "Inferred":("DBEAFE","1E40AF"),
    "High":("DCFCE7","166534"), "Medium":("FEF3C7","92400E"), "Low":("FECACA","991B1B"),
}

THIN = Side(style="thin", color=BG_DIV)
BORDER_BOTTOM = Border(bottom=THIN)

def f_title():    return Font(name=FONT, size=22, bold=True, color=INK)
def f_subtitle(): return Font(name=FONT, size=11, italic=True, color=INK_3)
def f_eyebrow():  return Font(name=FONT, size=9,  bold=True, color=INK_4)
def f_section():  return Font(name=FONT, size=12, bold=True, color=INK)
def f_header():   return Font(name=FONT, size=9,  bold=True, color=INK_4)
def f_body():     return Font(name=FONT, size=10, color=INK)
def f_mid():      return Font(name=FONT, size=10, color=INK_3)
def f_mono():     return Font(name="Consolas", size=9, color=INK_3)
def f_pill(c):    return Font(name=FONT, size=9,  bold=True, color=c)

def fill(color): return PatternFill("solid", fgColor=color)

# ── Job Info team for 2026 (Exxel Project 2026 / 2026) ─────
TEAM = [
    ("GC",             "Exxel Pacific, Inc."),
    ("GC PM",          "Justin Gotcher / Chris Austin"),
    ("GC Superintendent","Brian Proctor"),
    ("GC PE",          "Chris Austin"),
    ("Developer",      "Intra-Corp"),
    ("Architect",      "Nicholson Kovalchick"),
    ("Structural",     "N/A"),
    ("Acoustical",     "N/A"),
    ("ADA Consultant", "N/A"),
    ("Interior Design","N/A"),
    ("MEP Engineer",   "Pressler"),
]

wb = load_workbook(SRC)

def style_tab(ws):
    ws.sheet_view.showGridLines = False
    # Page background A:J up to max rows
    max_r = max(ws.max_row, 60)
    max_c = max(ws.max_column, 10)
    # Title (R2) & subtitle (R3)
    t = ws.cell(row=2, column=2)
    if t.value:
        t.font = f_title()
        t.alignment = Alignment(vertical="center", horizontal="left")
        ws.row_dimensions[2].height = 34
    s = ws.cell(row=3, column=2)
    if s.value:
        s.font = f_subtitle()
        s.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        ws.row_dimensions[3].height = 22
    # Divider row 4
    for c in range(2, max_c+1):
        ws.cell(row=4, column=c).fill = fill(BG_DIV)
    ws.row_dimensions[4].height = 2
    # Column widths — sensible defaults, col B wide
    widths = {1:2, 2:28, 3:34, 4:22, 5:22, 6:22, 7:22, 8:22, 9:22, 10:22}
    for col, w in widths.items():
        if col <= max_c:
            ws.column_dimensions[get_column_letter(col)].width = w
    # Style all rows 5+
    for r in range(5, ws.max_row+1):
        for c in range(2, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None:
                continue
            # Detect header row (all-caps short label in col B with no value in C..)
            if c == 2 and isinstance(v, str) and v.isupper() and len(v) < 60 and ws.cell(row=r, column=3).value in (None, ""):
                cell.font = f_eyebrow()
                cell.alignment = Alignment(vertical="center", horizontal="left")
                continue
            # Detect table header rows (bold-ish column labels)
            # Heuristic: if row has multiple string cells and next row has numeric values
            # Simpler: just apply body font uniformly, bold the first column labels
            if isinstance(v, str):
                # Pill detection
                sv = v.strip()
                if sv in PILL:
                    bg, fg = PILL[sv]
                    cell.fill = fill(bg)
                    cell.font = f_pill(fg)
                    cell.alignment = Alignment(vertical="center", horizontal="center")
                    continue
                if c == 2:
                    cell.font = f_mid()
                else:
                    cell.font = f_body()
                cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            elif isinstance(v, (int, float)):
                cell.font = f_body()
                cell.alignment = Alignment(vertical="center", horizontal="right")
                # Auto-format currency-ish values
                if not cell.number_format or cell.number_format == 'General':
                    if abs(v) >= 1000 or (isinstance(v, float) and v != int(v)):
                        cell.number_format = '"$"#,##0.00' if isinstance(v, float) and v != int(v) and abs(v) < 100000 else '"$"#,##0'
                    else:
                        cell.number_format = '#,##0'
    # Style table-header rows: any row where row r col B is string AND row r+1 col C is numeric OR
    # simple: row 5 is typically the header row
    hdr_r = 5
    for c in range(2, ws.max_column+1):
        cell = ws.cell(row=hdr_r, column=c)
        if cell.value and isinstance(cell.value, str):
            # If col 3 has a value in this row → it's a header row
            if ws.cell(row=hdr_r, column=3).value:
                cell.font = f_header()
                cell.fill = fill(BG_MUT)
                cell.border = BORDER_BOTTOM
                cell.alignment = Alignment(vertical="center", horizontal="left" if c == 2 else "left")
    # Zebra: light fill every other data row from row 6
    for r in range(6, ws.max_row+1):
        if (r - 6) % 2 == 1:
            for c in range(2, ws.max_column+1):
                cell = ws.cell(row=r, column=c)
                if not cell.fill or cell.fill.fgColor is None or (cell.fill.fgColor.rgb in (None, "00000000")):
                    cell.fill = fill(BG_PG)
    ws.row_dimensions[1].height = 6

# Inject Project Team on Overview — find last used row, append section
ov = wb['Overview']
last_r = ov.max_row
start = last_r + 3
ov.cell(row=start, column=2, value="PROJECT TEAM (OWP Job Info)")
start += 1
ov.cell(row=start, column=2, value="Role")
ov.cell(row=start, column=3, value="Status")
ov.cell(row=start, column=4, value="Name / Firm")
start += 1
for lbl, val in TEAM:
    ov.cell(row=start, column=2, value=lbl)
    pill = ov.cell(row=start, column=3, value=("Extracted" if val != "N/A" else "N/A"))
    if val == "N/A":
        pill.fill = fill("EFEFEF"); pill.font = f_pill(INK_4)
        pill.alignment = Alignment(vertical="center", horizontal="center")
    ov.merge_cells(start_row=start, start_column=4, end_row=start, end_column=10)
    ov.cell(row=start, column=4, value=val)
    start += 1

# Apply styling to all tabs
for name in wb.sheetnames:
    style_tab(wb[name])

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Saved {OUT} ({OUT.stat().st_size:,} bytes)")
print(f"Tabs: {wb.sheetnames}")
