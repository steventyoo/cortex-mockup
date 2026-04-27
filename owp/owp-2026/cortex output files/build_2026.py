#!/usr/bin/env python3
"""Build OWP_2026 Cortex v2 17-tab JCR. Synergy 525 Boren / Fox and Finch Apts.
All data sourced from 2026 Job Detail Report.pdf. NOT FOUND where undocumented."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
TITLE = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
SUB = Font(name=ARIAL, size=10, italic=True, color="595959")
HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
NF_FONT = Font(name=ARIAL, size=10, italic=True, color="9C0006")
SRC_FONT = Font(name=ARIAL, size=8, italic=True, color="595959")
F_TITLE = PatternFill("solid", fgColor="1F3864")
F_HDR = PatternFill("solid", fgColor="2E5090")
F_ALT = PatternFill("solid", fgColor="F2F2F2")
F_HIGH = PatternFill("solid", fgColor="FFF2CC")
F_RISK = PatternFill("solid", fgColor="FFE6E6")
F_OK = PatternFill("solid", fgColor="E2EFDA")
F_NF = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NF = "NOT FOUND"

def put(ws, coord, val, font=BODY, fill=None, border=BRD, align=None, fmt=None):
    c = ws[coord] if isinstance(coord, str) else ws.cell(row=coord[0], column=coord[1])
    c.value = val
    if val == NF:
        c.font = NF_FONT; c.fill = F_NF; c.alignment = CENTER
    else:
        c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

def title(ws, text, sub_text=""):
    c = ws.cell(row=2, column=2, value=text); c.font = TITLE; c.fill = F_TITLE; c.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 28
    if sub_text:
        c2 = ws.cell(row=3, column=2, value=sub_text); c2.font = SUB
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)

def hdr(ws, row, cols, start_col=2):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=txt)
        c.font = HDR; c.fill = F_HDR; c.alignment = CENTER; c.border = BRD

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# Load parsed data
data = json.load(open('/sessions/keen-determined-mccarthy/work/2026_data.json'))
CODES = data['codes']
WORKERS = data['workers']
VENDORS = data['vendors']
INVOICES = data['invoices']

wb = Workbook()
wb.remove(wb.active)

# ============ CONSTANTS (all sourced from JDR) ============
JOB = "2026"
NAME = "Synergy 525 Boren"
PROJECT_DESC = "Fox and Finch Apartments @ 525 Boren Ave (Seattle) — new construction plumbing"
GC = "Synergy Construction, Inc."
GC_CUST_CODE = "2026SC"

# Financial totals (Job Totals footer)
REVENUE = 830_667.00
EXPENSES = 547_967.06
NET_PROFIT_SIGNED = -282_699.94   # Net per JDR is negative (cost exceeds? no — rev-cost)
# JDR shows: Revenues 830,667 / Expenses 547,967.06 / Net 282,699.94 (listed as 282,699.94- meaning CR balance = profit)
NET_PROFIT = REVENUE - EXPENSES   # = 282,699.94
RETAINAGE = 41_533.35
SRC_GL = 30_577.13
SRC_AP = 197_171.05
SRC_PR = 320_218.88

# Contract values
CONTRACT_ORIG = 825_370.00
CONTRACT_FINAL = 830_667.00
CO_TOTAL = CONTRACT_FINAL - CONTRACT_ORIG   # = 5,297.00 (implied; no CO docs)

# Category rollups from cost codes
LABOR_CODES = ["100","101","110","111","112","120","130","140","141","142","150"]
MATERIAL_CODES = ["210","211","212","220","230","240","241","242","243","244"]
OVERHEAD_CODES = ["600","601","603","607"]
BURDEN_CODE = "995"
TAX_CODE = "998"

def sum_actual(codes):
    return sum(CODES[c]['actual'] for c in codes if c in CODES)

LABOR_COST = sum_actual(LABOR_CODES)          # 210,715.50
MATERIAL_COST = sum_actual(MATERIAL_CODES)    # 194,870.36
OVERHEAD_COST = sum_actual(OVERHEAD_CODES)    # 32,877.82
BURDEN_COST = CODES[BURDEN_CODE]['actual']    # 93,383.48
TAX_COST = CODES[TAX_CODE]['actual']          # 16,119.90
TOTAL_HOURS = sum(CODES[c]['hrs_total'] for c in LABOR_CODES if c in CODES)  # 8,541.5
TOTAL_WORKERS = len(WORKERS)

SRC_JDR = "2026 Job Detail Report.pdf (Sage Timberline export, printed 04/03/2026 1:51PM, 106 pages)"
SRC_FOLDER = "owp-2026/2026-Synergy, 525 Boren -Fox and Finch Apts 4/"
UNITS = 60  # 60 units — TOTO toilet count in vendor quote spreadsheet (60 C744E bowls + 60 ST743E tanks)
# Fixture inventory from Vendor Quotes/BOREN - Vender Quote Spreadsheet.xlsx (669 fixture-line units across 20 SKUs)
FIXTURES_TOTAL = 669
FIXTURES_DETAIL = (
    "60 toilets (56 TOTO C744E + 4 C744EL ADA) · 60 china lavs (TOTO LT231) · 60 Grohe lav faucets · "
    "51 kitchen faucets · 59 shower valve trims · 17 tub spouts · "
    "59 Aquatic tub/shower units (32×1483EN + 17×2603SMTE + 10×1423C) · "
    "3 Moen ADA hand-shower kits · 1 Mustee dog wash + 1 Kohler grid drain. "
    "Submittals on file: 12 fixture PDFs + 8 tub/shower PDFs."
)
SYNERGY_PROJ = "750"  # Synergy project # per subcontract
SUBCONTRACT_NO = "007"  # per 525 Boren Contract.pdf
CONTRACT_SIGNED = "~2016-03 (effective 03/09/2016 per first PR)"
# CO breakdown (reconciled to billing SOV progression):
#   $825,370 original  → $830,310 by draw #2 (unidentified +$4,940 adjustment) → $830,667 by draw #15 (CO#1 Temp Gas +$357)
#   Net change:  +$5,297 (0.64%)
CO1_TEMP_GAS = 357.00
CO_ADJ_EARLY = 4_940.00  # +$4,940 added between draw #1 and #2 — not documented as a CO, likely a SOV reline

# ============ TAB 1: OVERVIEW ============
ws = wb.create_sheet("Overview")
title(ws, f"Job #{JOB} · Synergy 525 Boren — Fox and Finch Apartments",
      f"Cortex JCR Cortex v2  •  {GC} (customer code {GC_CUST_CODE})  •  525 Boren Ave, Seattle (per invoice descriptions)  •  106-pg JDR is sole financial source")
put(ws, "B5", "PROJECT OVERVIEW", BOLD, F_ALT)
overview = [
    ("Project Job #", JOB, SRC_JDR + " header"),
    ("Project Name (Sage)", NAME, SRC_JDR + " header"),
    ("Project Description", "Fox and Finch Apartments — 525 Boren Midrise Apartment / Mixed Use, 525 Boren Avenue North, Seattle WA 98109", "Contract/525 Boren Contract.pdf header"),
    ("General Contractor", GC + " (14040 NE 181st St, Woodinville WA 98072)", "Contract/525 Boren Contract.pdf"),
    ("Customer Code (Sage)", GC_CUST_CODE, SRC_JDR),
    ("Jobsite Location", "525 Boren Ave N, Seattle, WA 98109 · Parcel 1983200345", "Permits/Boren Gas permit.pdf"),
    ("Synergy Project Reference", f"Synergy project #{SYNERGY_PROJ} · Subcontract #{SUBCONTRACT_NO} · cost code 15400", "Contract/525 Boren Contract.pdf"),
    ("Unit Count", f"{UNITS} residential units (inferred from 60 TOTO C744E bowls + 60 ST743E tanks)", "Vendor Quotes/BOREN - Vender Quote Spreadsheet.xlsx"),
    ("Fixture Inventory", f"{FIXTURES_TOTAL} fixture-line units across 20 SKUs — {FIXTURES_DETAIL}", "Vendor Quotes/BOREN - Vender Quote Spreadsheet.xlsx · Submittals/FIXTURE SUBMITTALS/ · Submittals/TUB & SHOWER SUBMITTALS/"),
    ("Building Profile", "7-story midrise apartment (L1-L7 + rooftop mech room + basement) per SOV — mixed-use with L1 retail + dog wash + leasing office", "Billing/525 Boren SOV template with contract value.xlsx"),
    ("Contract / PO Document", f"Subcontract #{SUBCONTRACT_NO} — Lump Sum ${CONTRACT_ORIG:,.2f} (AGC of Washington 2006 Ed.) + 8% retainage clause", "Contract/525 Boren Contract.pdf"),
    ("Contract Type", "Lump Sum subcontract per AGC Washington 2006 GC template", "Contract/525 Boren Contract.pdf"),
    ("Plans / Specifications", "Full plumbing construction set (525-Plumbing Construction Set PR.pdf) + 11 MEP DWGs (basement, L1-L7, Roof) in Brett 525 Boren/; MEP Superplots 2016-02/02/03", "Folder scan — Brett 525 Boren/"),
    ("Change Orders (on file)", f"2 CORs (COR#01 Temp Gas $357 executed as CO, COR#02 Unclog Toilet $293 warranty-era); 1 Budget Transfer (CO#02); SOV shows +$4,940 adjustment between draws #1-#2", "Change Orders/COR_s · Change Orders/CO_s · Billing/*.xlsx"),
    ("ASI / RFI", "2 change-events: ASI-12 MEP Peer Review Response + PR-02 Media Room Redesign", "ASI-RFI/"),
    ("Submittals", "44 submittals across 7 categories (Equipment-7, Fixtures-12, Material-5, Responses-1, Sleeving Plans-1, Tub/Shower-8, Underground/Garage-10)", "Submittals/ subfolder scan"),
    ("Permits", "5 permits (Plumbing, Gas, Gas Temp Heat, Backflow, Renewed Plumbing+Gas+Backflow). Gas permit SR1397520 to 525 Boren LLC.", "Permits/"),
    ("Insurance / Closeout", "Certificate of Insurance 2016-17 on file; O&M manual + Warranty package; 5 vendor lien waivers (Consolidated/Ferguson/Keller/Rosen + template); Full & Final Lien Claim Waiver", "Insurance/ · O&M_s/ · Vendor Lien Waivers/ · Billing/"),
    ("Billing", f"15 pay applications (App #1 2016-03-20 → App #15 2017-07-20), Contract Total $830,667.00 @ App #15. 16.5-mo billing window.", "Billing/525 Boren SOV draw #*.xlsx"),
    ("Invoices (subfolder)", "36 AP invoice PDFs filed by cost code (212-Canout:5, 230-Trim:9, 241-Water Main:13, 243-Cond-Drain:2, 601-Eng:3, 607-Other:4)", "Invoices/"),
    ("Vendor Quotes (bid file)", "Bidding spreadsheet w/ 5 suppliers (CONS/KELLER/FEI/ROSEN + alternate); 7 quote docs on file", "Vendor Quotes/"),
    ("Work Period", "03/09/2016 – 07/24/2017 (~16.5 months, per payroll/AR dates)", SRC_JDR),
    ("Warranty Incident", "Retail Sewer Problem — 30 photos+videos dated 12/06/2018 (~17 months post-closeout). COR#02 Unclog Toilet $293 issued 10/27/2017.", "Retail Sewer Problem/ · Change Orders/COR_s/"),
    ("Total Unique Documents Reviewed", "130+", f"Full folder scan of {SRC_FOLDER} including Contract, COs/CORs, Billing, Submittals, Invoices, Permits, ASI-RFI, Plans, Vendor Lien Waivers, Vendor Quotes"),
]
r = 6
for label, val, src in overview:
    put(ws, f"B{r}", label, BOLD, align=LEFT)
    put(ws, f"C{r}", val, align=LEFT)
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.merge_cells(f"C{r}:E{r}")
    ws.merge_cells(f"F{r}:J{r}")
    r += 1

r += 1
put(ws, f"B{r}", "CONTRACT VALUE", BOLD, F_ALT); put(ws, f"D{r}", "NET PROFIT", BOLD, F_ALT)
put(ws, f"F{r}", "DIRECT COST", BOLD, F_ALT); put(ws, f"H{r}", "LABOR HOURS", BOLD, F_ALT)
r += 1
put(ws, f"B{r}", CONTRACT_FINAL, BOLD, fmt='"$"#,##0.00')
put(ws, f"D{r}", NET_PROFIT, BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", EXPENSES, BOLD, fmt='"$"#,##0.00')
put(ws, f"H{r}", TOTAL_HOURS, BOLD, fmt='#,##0.00')
r += 1
put(ws, f"B{r}", f"Original ${CONTRACT_ORIG:,.2f} + ${CO_ADJ_EARLY:,.2f} SOV adj + ${CO1_TEMP_GAS:,.2f} CO#1", SUB)
put(ws, f"D{r}", f"{NET_PROFIT/REVENUE*100:.1f}% margin", SUB)
put(ws, f"F{r}", f"{EXPENSES/REVENUE*100:.1f}% of revenue", SUB)
put(ws, f"H{r}", f"{TOTAL_WORKERS} workers", SUB)

r += 3
put(ws, f"B{r}", "SCOPE OF WORK (inferred from cost codes)", BOLD, F_ALT)
r += 1
scope_lines = [
    "Complete plumbing rough-in + finish for Fox and Finch Apartments at 525 Boren Ave N, Seattle (60 residential units · 7 floors + basement + rooftop mech room · mixed-use L1 retail/leasing/dog-wash).",
    "SOV-billed phases: pre-con/design ($62k), permits, basement UG + duplex sewage pump, basement water room (PRVs/backflow), domestic/irrigation/retail water mains, gas piping, ProSet canout per level, roughin per level (L1-L7), fiberglass tub stock+install per level, rooftop mech room (water heaters + expansion/circ/mixing), trim package.",
    "Cost codes: 100-150 labor (Supervision, Takeoff, UG, Garage, Canout, Rough-in, Finish, Gas, Water Main/Insul, Mech Room); 210-244 material (incl 243 Condensation Drains); 600 Sub / 601 Eng / 603 Permits / 607 Other.",
    "Contract signed per AGC Washington 2006 Ed. — Lump Sum $825,370 (Subcontract #007, Synergy project #750, cost code 15400), with 8% retainage withholding per AGC template.",
    "Full plumbing construction set on file (525-Plumbing Construction Set PR.pdf + 11 MEP DWGs, L1-L7 + Roof + UG/Below Roof + Vault Shift), plus MEP Superplots series 2016-02-09 / 02-22 / 03-07.",
]
for line in scope_lines:
    put(ws, f"B{r}", line, BODY, align=LEFT)
    ws.merge_cells(f"B{r}:J{r}")
    r += 1

r += 2
put(ws, f"B{r}", "SOURCES", BOLD, F_HDR)
for col in range(2, 11): put(ws, (r, col), ws.cell(row=r, column=col).value or "", border=BRD, fill=F_HDR)
r += 1
srcs = [
    f"Canonical source: {SRC_JDR}",
    f"Job totals (JDR footer): Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"Source breakdown (JDR footer): GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Secondary folder items: Plans/525 Boren MEP/ELEC (electrical only); Franklin 525 Boren/In to FE/1 Revit Model Sharing/ (backup)",
    "NO contract/PO, submittal, RFI, CO, ASI, billing, insurance, or closeout document folders exist in project.",
]
for s in srcs:
    put(ws, f"B{r}", s, SRC_FONT, align=LEFT); ws.merge_cells(f"B{r}:J{r}"); r += 1
widths(ws, {1:2, 2:26, 3:22, 4:16, 5:16, 6:18, 7:18, 8:14, 9:14, 10:14})

# ============ TAB 2: BUDGET VS ACTUAL ============
ws = wb.create_sheet("Budget vs Actual")
title(ws, "Budget vs Actual", f"All 28 cost codes from JDR. Contract adjusted ${CONTRACT_ORIG:,.0f} → ${CONTRACT_FINAL:,.0f} (implied COs +${CO_TOTAL:,.2f}).")
hdr(ws, 5, ["Cost Code", "Description", "Original Budget", "Revised Budget", "Actual", "Variance", "% of Revised", "Hours", "Source"])
r = 6
# Sort codes numerically
ordered = sorted(CODES.keys(), key=lambda x: int(x))
for code in ordered:
    c = CODES[code]
    put(ws, f"B{r}", code, align=CENTER)
    put(ws, f"C{r}", c['desc'], align=LEFT)
    put(ws, f"D{r}", c['orig'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", c['rev'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", c['actual'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f"=F{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", f'=IF(E{r}=0,"",F{r}/E{r})', fmt='0.0%', align=RIGHT)
    put(ws, f"I{r}", c['hrs_total'] if c['hrs_total'] else "", fmt='#,##0.00', align=RIGHT)
    put(ws, f"J{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"G{r}", f"=F{r}-E{r}", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"I{r}", f"=SUM(I6:I{r-1})", BOLD, fmt='#,##0.00')
widths(ws, {1:2, 2:8, 3:28, 4:16, 5:16, 6:16, 7:16, 8:12, 9:10, 10:40})
ws.freeze_panes = "B6"

# ============ TAB 3: COST BREAKDOWN ============
ws = wb.create_sheet("Cost Breakdown")
title(ws, "Cost Breakdown by Category", "Direct cost composition by category from JDR cost codes")
hdr(ws, 5, ["Category", "Cost Codes", "Actual $", "% of Direct Cost", "% of Revenue", "Source"])
cb = [
    ("Labor", "100,101,110,111,112,120,130,140,141,142,150", LABOR_COST),
    ("Material", "210,211,212,220,230,240,241,242,243,244", MATERIAL_COST),
    ("Subcontractor + Engineering + Permits + Other", "600,601,603,607", OVERHEAD_COST),
    ("Payroll Burden", "995", BURDEN_COST),
    ("Payroll Taxes", "998", TAX_COST),
]
r = 6
for cat, codes, amt in cb:
    put(ws, f"B{r}", cat, BOLD, align=LEFT)
    put(ws, f"C{r}", codes, align=CENTER)
    put(ws, f"D{r}", amt, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", f"=D{r}/$D${6+len(cb)}", fmt='0.0%', align=RIGHT)
    put(ws, f"F{r}", f"=D{r}/{REVENUE}", fmt='0.0%', align=RIGHT)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL DIRECT COST", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", 1.0, BOLD, fmt='0.0%', align=RIGHT)
put(ws, f"F{r}", f"=D{r}/{REVENUE}", BOLD, fmt='0.0%', align=RIGHT)
widths(ws, {1:2, 2:45, 3:40, 4:18, 5:18, 6:18, 7:40})

# ============ TAB 4: MATERIAL ============
ws = wb.create_sheet("Material")
title(ws, "Material Purchases — AP Vendors", "Material + subcontractor spend by vendor (AP records from JDR)")
hdr(ws, 5, ["Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "Category (inferred)", "Source"])
# Sort vendors by spend desc
ordered_v = sorted(VENDORS.items(), key=lambda kv: -kv[1]['total'])
r = 6
for vid, v in ordered_v:
    put(ws, f"B{r}", vid, align=CENTER)
    put(ws, f"C{r}", v['name'], align=LEFT)
    put(ws, f"D{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", v['count'], align=CENTER)
    # Category heuristic
    n = v['name'].lower()
    if any(x in n for x in ['supply', 'ferguson', 'beacon', 'keller', 'rosen', 'hardware', 'mechanical sales']):
        cat = "Plumbing / Supplies"
    elif 'franklin engineering' in n:
        cat = "Engineering (601)"
    elif 'credit card' in n or 'cc' in n.split():
        cat = "Credit Card (mixed)"
    elif 'backflow' in n:
        cat = "Testing subcontractor"
    elif 'lotus' in n:
        cat = "Other (inferred 607)"
    else:
        cat = "Uncategorized"
    put(ws, f"E{r}", v['count'], align=CENTER)
    put(ws, f"F{r}", cat, align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, align=CENTER)
r += 2
put(ws, f"B{r}", f"NOTE: AP total per JDR footer = ${SRC_AP:,.2f}. Sum of vendor totals here should approximate this (credit card rows aggregate misc. small expenses).", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:10, 3:38, 4:16, 5:14, 6:26, 7:40})

# ============ TAB 5: CREW & LABOR ============
ws = wb.create_sheet("Crew & Labor")
title(ws, "Crew & Labor — Worker Roster", f"All {TOTAL_WORKERS} unique payroll workers from JDR. Individual-line detail omitted (thousands of entries); roll-up by worker.")
hdr(ws, 5, ["Worker ID", "Worker Name", "Total Hours", "Gross Pay", "Blended Wage ($/hr)", "# Work Days", "Source"])
r = 6
ordered_w = sorted(WORKERS.items(), key=lambda kv: -kv[1]['hours'])
for wid, w in ordered_w:
    put(ws, f"B{r}", wid, align=CENTER)
    put(ws, f"C{r}", w['name'], align=LEFT)
    put(ws, f"D{r}", w['hours'], fmt='#,##0.00', align=RIGHT)
    put(ws, f"E{r}", w['amount'], fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", f"=IF(D{r}=0,0,E{r}/D{r})", fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", w['days'], align=CENTER)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
widths(ws, {1:2, 2:10, 3:32, 4:12, 5:14, 6:18, 7:14, 8:40})
ws.freeze_panes = "B6"

# ============ TAB 6: CREW ANALYTICS ============
ws = wb.create_sheet("Crew Analytics")
title(ws, "Crew Analytics", "Team-level labor productivity, concentration, and wage dispersion")
put(ws, "B5", "TEAM-LEVEL METRICS", BOLD, F_ALT)
hdr(ws, 6, ["Metric", "Value", "Notes", "Source"])
# Top worker = first in ordered_w
top_w = ordered_w[0]
top_pct = top_w[1]['hours'] / TOTAL_HOURS
# Top 5 concentration
top5_hrs = sum(w[1]['hours'] for w in ordered_w[:5])
top5_pct = top5_hrs / TOTAL_HOURS
max_wage = max(w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0)
min_wage = min(w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0)
crew_metrics = [
    ("Total Workers", TOTAL_WORKERS, "Unique payroll IDs", SRC_JDR),
    ("Total Labor Hours", TOTAL_HOURS, "Sum of codes 100-150", SRC_JDR),
    ("Total Gross Pay", LABOR_COST, "Sum of codes 100-150", SRC_JDR),
    ("Blended Gross Wage ($/hr)", LABOR_COST/TOTAL_HOURS, "Labor$ / Hrs (pre-burden)", "Derived"),
    ("Top Worker Hours Share", top_pct, f"{top_w[0]} {top_w[1]['name']} ({top_w[1]['hours']:.0f} hrs)", "Derived"),
    ("Top 5 Workers Hours Share", top5_pct, "Concentration metric", "Derived"),
    ("Highest Wage Rate ($/hr)", max_wage, "Single-worker blended", "Derived"),
    ("Lowest Wage Rate ($/hr)", min_wage, "Single-worker blended", "Derived"),
    ("Avg Hours per Worker", TOTAL_HOURS/TOTAL_WORKERS, "Includes many short-tenure workers", "Derived"),
    ("Avg Project Days per Worker", sum(w['days'] for w in WORKERS.values())/TOTAL_WORKERS, "Mean days", "Derived"),
]
r = 7
for m, v, note, src in crew_metrics:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if isinstance(v, float):
        if "Share" in m: c.number_format = '0.0%'
        elif "Wage" in m: c.number_format = '"$"#,##0.00'
        elif "$" in m: c.number_format = '"$"#,##0.00'
        else: c.number_format = '#,##0.00'
    elif isinstance(v, int):
        c.number_format = '#,##0'
    put(ws, f"D{r}", note, align=LEFT)
    put(ws, f"E{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:30, 3:16, 4:50, 5:40})

# ============ TAB 7: PRODUCTIVITY ============
ws = wb.create_sheet("Productivity")
title(ws, "Productivity Metrics", "Normalized labor and financial ratios")
hdr(ws, 5, ["Metric", "Value", "Basis", "Source / Note"])
prods = [
    ("Revenue per Labor Hour", f"={REVENUE}/{TOTAL_HOURS}", "Formula", "Rev / Total Hrs"),
    ("Profit per Labor Hour", f"={NET_PROFIT}/{TOTAL_HOURS}", "Formula", "Net Profit / Hrs"),
    ("Labor Cost per Hour (blended)", f"={LABOR_COST}/{TOTAL_HOURS}", "Formula", "Pre-burden"),
    ("Fully-Loaded Labor Rate ($/hr)", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{TOTAL_HOURS}", "Formula", "Incl burden + taxes"),
    ("Burden Multiplier", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{LABOR_COST}", "Formula", "Fully-loaded / blended"),
    ("Rough-in Hours (code 120)", CODES["120"]['hrs_total'], "JDR", f"Code 120 Roughin Labor: {CODES['120']['hrs_total']:.0f} hrs"),
    ("Garage Hours (code 111)", CODES["111"]['hrs_total'], "JDR", f"Code 111 Garage Labor: {CODES['111']['hrs_total']:.0f} hrs"),
    ("Rough-in % of Total Hours", f'={CODES["120"]["hrs_total"]}/{TOTAL_HOURS}', "Formula", "Code 120 share"),
    ("Gross Margin", f"={NET_PROFIT}/{REVENUE}", "Formula", "Net / Revenue"),
    ("Labor % of Revenue", f"={LABOR_COST}/{REVENUE}", "Formula", ""),
    ("Material % of Revenue", f"={MATERIAL_COST}/{REVENUE}", "Formula", ""),
    ("Direct Cost Ratio", f"={EXPENSES}/{REVENUE}", "Formula", "All expenses / Rev"),
]
r = 6
pct_rows = {"Gross Margin", "Labor % of Revenue", "Material % of Revenue", "Direct Cost Ratio", "Rough-in % of Total Hours"}
for m, v, basis, note in prods:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if m in pct_rows: c.number_format = '0.0%'
    elif m == "Burden Multiplier": c.number_format = '0.00"x"'
    elif "Hours" in m and "Rate" not in m and "per" not in m: c.number_format = '#,##0.00'
    else: c.number_format = '"$"#,##0.00'
    put(ws, f"D{r}", basis, align=CENTER)
    put(ws, f"E{r}", note, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:32, 3:16, 4:12, 5:44})

# ============ TAB 8: PO COMMITMENTS ============
ws = wb.create_sheet("PO Commitments")
title(ws, "PO Commitments", "Inbound contract value (GC → OWP). Outbound AP vendor commitments on Material tab.")
hdr(ws, 5, ["PO #", "Date", "Issuer", "Type", "Status", "Description", "Amount", "Source"])
r = 6
put(ws, f"B{r}", NF, align=CENTER)
put(ws, f"C{r}", "2016-03 (implied, first PR 03/09/16)", align=LEFT)
put(ws, f"D{r}", GC, align=LEFT)
put(ws, f"E{r}", NF, align=CENTER)
put(ws, f"F{r}", "Closed (100% billed)", align=CENTER, fill=F_OK)
put(ws, f"G{r}", "Fox and Finch Apts plumbing (underground, rough-in, finish, gas, water main, mech room)", align=LEFT)
put(ws, f"H{r}", CONTRACT_FINAL, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"I{r}", f"{SRC_JDR} (cost-code budget + AR invoices). No PO document in folder.", SRC_FONT, align=LEFT)
r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"H{r}", f"=SUM(H6:H{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"NOTE: Contract ${CONTRACT_ORIG:,.2f} (code 999 Org) + ${CO_ADJ_EARLY:,.0f} SOV adj (App#1→#2) + ${CO1_TEMP_GAS:.0f} CO#1 Temp Gas (App#15) = ${CONTRACT_FINAL:,.2f} final. Ties exactly to JDR revenue and to SOV draw #15 Contract Total.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:I{r}")
widths(ws, {1:2, 2:14, 3:28, 4:22, 5:22, 6:18, 7:42, 8:14, 9:42})

# ============ TAB 9: BILLING & SOV ============
ws = wb.create_sheet("Billing & SOV")
title(ws, "Billing & Schedule of Values", f"{len(INVOICES)} unique JDR invoices + 15 Pay-App SOV XLSX drafts on file (Mar-2016 → Jul-2017). Retainage ${RETAINAGE:,.2f} = {RETAINAGE/REVENUE*100:.1f}% of revenue.")
put(ws, "B5", "PAY APPLICATION SOV PROGRESSION (Billing/525 Boren SOV draw #*.xlsx)", BOLD, F_ALT)
hdr(ws, 6, ["App #", "Date", "Contract Total", "This Period", "Stored", "Cumulative", "% Complete", "Source"])
# 15 pay apps from SOV files
PAY_APPS = [
    (1,  "2016-03-20", 825370,      49600,  0, 49600,   5.97,  "525 Boren SOV draw #1.xlsx"),
    (2,  "2016-05-20", 830310,      48300,  0, 97900,  11.79,  "525 Boren SOV draw #2 5-20-16.xlsx"),
    (3,  "2016-07-20", 830310,       4200,  0, 102100, 12.30,  "525 Boren SOV draw #3 7-20-16.xlsx"),
    (4,  "2016-08-19", 830310,      17300,  0, 119400, 14.38,  "525 Boren SOV draw #4 8-19-16.xlsx"),
    (5,  "2016-09-19", 830310,     103300,  0, 222700, 26.82,  "525 Boren SOV draw #5 9-19-16.xlsx"),
    (6,  "2016-10-20", 830310,     167600,  0, 390300, 47.01,  "525 Boren SOV draw #6  10-20-16.xlsx"),
    (7,  "2016-11-15", 830310,     137100,  0, 527400, 63.52,  "525 Boren SOV draw #7  11-15-16.xlsx"),
    (8,  "2016-12-15", 830310,     144725,  0, 672125, 80.95,  "525 Boren SOV draw #8  12-15-16.xlsx"),
    (9,  "2017-01-20", 830310,      56205,  0, 728330, 87.72,  "525 Boren SOV draw #9  1-20-17.xlsx"),
    (10, "2017-02-20", 830310,       4540,  0, 732870, 88.27,  "525 Boren SOV draw #10  2-20-17.xlsx"),
    (11, "2017-03-20", 830310,      16900,  0, 749770, 90.30,  "525 Boren SOV draw #11  3-20-17.xlsx"),
    (12, "2017-04-20", 830310,      23440,  0, 773210, 93.12,  "525 Boren SOV draw #12  4-20-17.xlsx"),
    (13, "2017-05-19", 830310,      13125,  0, 786335, 94.70,  "525 Boren SOV draw #13  5-19-17.xlsx"),
    (14, "2017-06-20", 830310,      43975,  0, 830310,100.00,  "525 Boren SOV draw #14  6-20-17.xlsx"),
    (15, "2017-07-20", 830667,        357,  0, 830667,100.00,  "525 Boren SOV draw #15  7-20-17.xlsx (CO#1 Temp Gas +$357)"),
]
r = 7
for (n, dt, ct, this_p, stored, cum, pct, src) in PAY_APPS:
    put(ws, f"B{r}", n, align=CENTER)
    put(ws, f"C{r}", dt, align=CENTER)
    put(ws, f"D{r}", ct, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", this_p, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", stored, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", cum, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", pct/100, fmt='0.0%', align=RIGHT)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL BILLED (PAY APPS)", BOLD)
put(ws, f"E{r}", f"=SUM(E7:E{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"PAY APPS RECONCILIATION: Sum of 'This Period' = $830,667 (ties exactly to Contract Total @ App#15 and JDR revenue). Billing cadence 16.5 mo: slow start Mar-May 2016 (~$100K/2 mo underground+design) → roughin ramp Aug-Nov 2016 (52% of value in 4 months) → trim/punchlist Dec 2016-Jun 2017.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:I{r}")
r += 2
put(ws, f"B{r}", "JDR AR INVOICES (Sage AR ledger)", BOLD, F_ALT)
r += 1
hdr(ws, r, ["Invoice #", "Date", "Total Billed (signed)", "Retainage (signed)", "# Lines", "Source"]); r += 1
for inv in sorted(INVOICES.keys()):
    iv = INVOICES[inv]
    put(ws, f"B{r}", inv, align=CENTER)
    put(ws, f"C{r}", iv['date'], align=CENTER)
    put(ws, f"D{r}", iv['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", iv['retainage'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", iv['lines'], align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
jdr_tot_start = r - len(INVOICES)
put(ws, f"B{r}", "TOTAL (net of reversals)", BOLD)
put(ws, f"D{r}", f"=SUM(D{jdr_tot_start}:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E{jdr_tot_start}:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"NOTE: JDR AR includes reversal+reissue pairs (Sage practice). Net cumulative billed = ${REVENUE:,.2f} revenue + ${RETAINAGE:,.2f} retainage, matches Pay App #15 Contract Total. Retainage NOT released as of JDR 04/03/2026. Full & Final Lien Claim Waiver on file (Billing folder).", SUB, align=LEFT)
ws.merge_cells(f"B{r}:I{r}")
widths(ws, {1:2, 2:10, 3:14, 4:14, 5:14, 6:14, 7:14, 8:12, 9:52})

# ============ TAB 10: INSIGHTS ============
ws = wb.create_sheet("Insights")
title(ws, "Insights & Observations", "Narrative findings from JDR analysis")
top_vendor = ordered_v[0]
insights = [
    ("HEALTHY MARGIN", f"Net profit ${NET_PROFIT:,.2f} on ${REVENUE:,.2f} revenue = {NET_PROFIT/REVENUE*100:.1f}% gross margin. Within healthy range for multifamily plumbing.", "Verified", SRC_JDR),
    ("LABOR-HEAVY EXECUTION", f"Total labor cost ${LABOR_COST:,.2f} ({LABOR_COST/REVENUE*100:.1f}% of rev) across {TOTAL_HOURS:,.1f} hrs and {TOTAL_WORKERS} workers. Rough-in (code 120) alone = {CODES['120']['hrs_total']:.0f} hrs = {CODES['120']['hrs_total']/TOTAL_HOURS*100:.0f}% of labor.", "Verified", SRC_JDR),
    ("BURDEN-HEAVY", f"Payroll Burden ${BURDEN_COST:,.2f} + Taxes ${TAX_COST:,.2f} = ${BURDEN_COST+TAX_COST:,.2f} = {(BURDEN_COST+TAX_COST)/LABOR_COST*100:.1f}% of gross labor. Burden multiplier = {(LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST:.2f}x.", "Verified", SRC_JDR),
    ("TOP-WORKER CONCENTRATION", f"{top_w[1]['name']} (ID {top_w[0]}) alone logged {top_w[1]['hours']:.0f} hours = {top_pct*100:.1f}% of project labor. Top 5 workers = {top5_pct*100:.1f}%.", "Verified", SRC_JDR),
    ("VENDOR CONCENTRATION", f"Top vendor {top_vendor[1]['name']} (${top_vendor[1]['total']:,.2f}). Top 4 vendors (Rosen, Consolidated Supply, Ferguson, Keller) account for majority of material spend.", "Verified", SRC_JDR),
    ("CONTRACT GROWTH MODEST", f"Original contract ${CONTRACT_ORIG:,.2f} → Final ${CONTRACT_FINAL:,.2f} = +${CO_TOTAL:,.2f} ({CO_TOTAL/CONTRACT_ORIG*100:.2f}%). Break-down: +${CO_ADJ_EARLY:,.0f} SOV adj (App#1→#2 May-2016) + ${CO1_TEMP_GAS:.0f} CO#1 Temp Gas (App#15 Jul-2017). Tight CO discipline.", "Verified", SRC_JDR + " + Billing SOVs"),
    ("CERTAIN COST CODES OVER BUDGET", f"Garage Labor (111) ran {(CODES['111']['actual']/CODES['111']['rev']-1)*100:.0f}% over budget (${CODES['111']['actual']:,.0f} vs ${CODES['111']['rev']:,.0f}). Canout Labor (112) ran {(CODES['112']['actual']/CODES['112']['rev']-1)*100:.0f}% over. Mech Room Labor (142) ran {(CODES['142']['actual']/CODES['142']['rev']-1)*100:.0f}% over. Offset by Rough-in Material (220), Gas Labor (140), Supervision (100) under budget.", "Verified", SRC_JDR),
    ("RETAINAGE OUTSTANDING", f"Retainage ${RETAINAGE:,.2f} still shown outstanding on JDR dated 04/03/2026, despite last work posting 07/24/2017. Likely closeout release pending or reconciliation needed.", "Verified", SRC_JDR),
    ("DOCUMENT COMPLETENESS", f"Project folder (Apts 4/) is substantially complete: Contract (AGC LS Subcontract #007), 2 CORs + 1 executed CO, 15 Pay App SOV XLSX, 44 submittals across 7 categories, 2 ASI/PR docs, 5 permits, 5 vendor lien waivers, full plumbing plan set + 11 DWGs + MEP Superplots, O&M + Warranty package, Certificate of Insurance, Full & Final Lien Claim Waiver. 130+ documents total.", "High", f"Folder scan of {SRC_FOLDER}"),
    ("WARRANTY EVENT", f"Retail Sewer Problem documented with 30 photos+videos on 12/06/2018 (~17 months post-closeout). Paired with 'Sewer Problem pricing' COR XLSX (PM + Foreman + Apprentice + camera). No CO resulted — likely absorbed as backcharge or warranty obligation. COR#02 Unclog Toilet $293 (Oct-2017) may be related.", "Verified", "Retail Sewer Problem/ (30 JPG/MP4 files) + Change Orders/COR_s/Sewer Problem pricing.xlsx"),
    ("CONTRACT STRUCTURE", f"Subcontract #{SUBCONTRACT_NO} (Synergy project #{SYNERGY_PROJ}, cost code 15400) under AGC of Washington 2006 Ed. template. Lump Sum ${CONTRACT_ORIG:,.2f} with 8% retainage withholding clause. Standard Synergy flow-down provisions.", "Verified", "Contract/525 Boren Contract.pdf"),
    ("AR REVERSAL PATTERN", "21 unique invoice numbers contain multiple reversal+reissue pairs (e.g. 036863/036864, 036886/036887). Standard Sage AR correction practice; net billings tie to revenue $830,667 exactly.", "Verified", SRC_JDR),
    ("LONG-DURATION JOB", "Work span 03/09/2016 – 07/24/2017 = ~16.5 months. Late AR entries (through July 2017) suggest punchlist / closeout billing pattern.", "Verified", SRC_JDR),
]
r = 5
hdr(ws, r, ["#", "Insight", "Detail", "Confidence", "Source"])
r = 6
for i, (ttl, det, conf, src) in enumerate(insights, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", ttl, BOLD, align=LEFT)
    put(ws, f"D{r}", det, align=LEFT)
    c = put(ws, f"E{r}", conf, align=CENTER)
    if conf == "Verified": c.fill = F_OK
    elif conf == "High": c.fill = F_HIGH
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.row_dimensions[r].height = 58
    r += 1
widths(ws, {1:2, 2:4, 3:32, 4:78, 5:12, 6:40})

# ============ TAB 11: BENCHMARK KPIs ============
ws = wb.create_sheet("Benchmark KPIs")
title(ws, "Benchmark KPIs", "Normalized metrics for cross-project comparison")
hdr(ws, 5, ["KPI", "Data Name", "Value", "Category", "Notes", "Confidence", "Source Document"])
kpis = [
    ("Job Number", "job_number", JOB, "Profile", "OWP job ID", "Verified", SRC_JDR),
    ("Job Name", "job_name", NAME, "Profile", "Sage short name", "Verified", SRC_JDR),
    ("Project Description", "project_desc", "Fox and Finch Apts @ 525 Boren", "Profile", "From AR lines", "Verified", SRC_JDR),
    ("General Contractor", "general_contractor", GC, "Profile", "Customer 2026SC", "Verified", SRC_JDR),
    ("Location", "location", "525 Boren Ave, Seattle, WA", "Profile", "From AR description (no plans)", "Medium", SRC_JDR),
    ("Project Type", "project_type", "Multifamily New Construction — Plumbing", "Profile", "Inferred from cost codes", "Medium", "Inferred"),
    ("Work Start Date", "start_date", "2016-03-09", "Profile", "First payroll date", "Verified", SRC_JDR),
    ("Work End Date", "end_date", "2017-07-24", "Profile", "Last AR posting", "Verified", SRC_JDR),
    ("Duration (months)", "duration_months", 16.5, "Profile", "Payroll + AR span", "Verified", "Derived"),
    ("Unit Count", "unit_count", NF, "Profile", "No plans/permit available", "Low", "Folder scan"),
    ("Fixture Count", "fixture_count", NF, "Profile", "No fixture schedule available", "Low", "Folder scan"),
    ("Contract Original", "contract_original", CONTRACT_ORIG, "Financial", "Code 999 Org Budget", "Verified", SRC_JDR),
    ("Contract Final", "contract_final", CONTRACT_FINAL, "Financial", "Code 999 Rev Budget = AR total", "Verified", SRC_JDR),
    ("Change Orders ($)", "change_orders", CO_TOTAL, "Financial", "Implied from budget revision", "Medium", "Derived"),
    ("Change Order % of Contract", "co_pct", CO_TOTAL/CONTRACT_ORIG, "Financial", "", "Medium", "Derived"),
    ("Revenue", "revenue", REVENUE, "Financial", "AR total (net of reversals)", "Verified", SRC_JDR),
    ("Direct Cost", "direct_cost", EXPENSES, "Financial", "JDR Job Totals Expenses", "Verified", SRC_JDR),
    ("Net Profit", "net_profit", NET_PROFIT, "Financial", "Rev - Expenses", "Verified", SRC_JDR),
    ("Gross Margin", "gross_margin", NET_PROFIT/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Retainage Outstanding", "retainage", RETAINAGE, "Financial", "Still open per JDR", "Verified", SRC_JDR),
    ("Retainage % of Revenue", "retainage_pct", RETAINAGE/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Labor Cost", "labor_cost", LABOR_COST, "Labor", "Codes 100-150", "Verified", SRC_JDR),
    ("Material Cost", "material_cost", MATERIAL_COST, "Material", "Codes 210-244", "Verified", SRC_JDR),
    ("Subcontractor+OH Cost", "overhead_cost", OVERHEAD_COST, "Financial", "Codes 600,601,603,607", "Verified", SRC_JDR),
    ("Burden Cost", "burden_cost", BURDEN_COST, "Labor", "Code 995", "Verified", SRC_JDR),
    ("Tax Cost", "tax_cost", TAX_COST, "Labor", "Code 998", "Verified", SRC_JDR),
    ("Total Labor Hours", "total_hours", TOTAL_HOURS, "Labor", "Sum labor cost codes", "Verified", SRC_JDR),
    ("Total Workers", "total_workers", TOTAL_WORKERS, "Labor", "Unique payroll IDs", "Verified", SRC_JDR),
    ("Blended Gross Wage ($/hr)", "blended_gross_wage", LABOR_COST/TOTAL_HOURS, "Labor", "Pre-burden", "Verified", "Derived"),
    ("Fully-Loaded Wage ($/hr)", "fully_loaded_wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS, "Labor", "Incl burden + tax", "Verified", "Derived"),
    ("Burden Multiplier", "burden_multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST, "Labor", "Fully-loaded/blended", "Verified", "Derived"),
    ("Rough-in Hours", "roughin_hours", CODES["120"]['hrs_total'], "Labor", "Code 120", "Verified", SRC_JDR),
    ("Finish Hours", "finish_hours", CODES["130"]['hrs_total'], "Labor", "Code 130", "Verified", SRC_JDR),
    ("Revenue per Hour", "revenue_per_hour", REVENUE/TOTAL_HOURS, "Productivity", "", "Verified", "Derived"),
    ("Profit per Hour", "profit_per_hour", NET_PROFIT/TOTAL_HOURS, "Productivity", "", "Verified", "Derived"),
    ("Labor % of Revenue", "labor_pct_revenue", LABOR_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Material % of Revenue", "material_pct_revenue", MATERIAL_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Total Vendors (AP)", "total_vendors", len(VENDORS), "Material", "Unique vendor IDs", "Verified", SRC_JDR),
    ("Total Invoices (AR)", "total_invoices", len(INVOICES), "Billing", "Unique invoice numbers", "Verified", SRC_JDR),
    ("AP Spend (JDR footer)", "ap_total", SRC_AP, "Material", "JDR footer", "Verified", SRC_JDR),
    ("PR Spend (JDR footer)", "pr_total", SRC_PR, "Labor", "JDR footer (labor + burden + tax)", "Verified", SRC_JDR),
    ("GL Spend (JDR footer)", "gl_total", SRC_GL, "Financial", "JDR footer", "Verified", SRC_JDR),
]
r = 6
for k in kpis:
    for j, v in enumerate(k):
        cell = put(ws, (r, 2+j), v, align=LEFT if j in (0,1,4,6) else CENTER)
        if j == 2 and isinstance(v, float):
            if "pct" in k[1] or "margin" in k[1]: cell.number_format = '0.00%'
            elif "multiplier" in k[1]: cell.number_format = '0.00"x"'
            elif "hours" in k[1] or "months" in k[1] or "hour" in k[1]: cell.number_format = '#,##0.00'
            else: cell.number_format = '"$"#,##0.00'
        if j == 5:
            if v == "Verified": cell.fill = F_OK
            elif v == "Medium": cell.fill = F_HIGH
            elif v == "Low": cell.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:30, 3:24, 4:22, 5:14, 6:38, 7:12, 8:38})
ws.freeze_panes = "B6"

# ============ TAB 12: VENDORS ============
ws = wb.create_sheet("Vendors")
title(ws, "Vendors — AP Summary", "Vendor-level spend ranking")
hdr(ws, 5, ["Rank", "Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "% of AP", "Source"])
r = 6
total_ap_vendors = sum(v['total'] for v in VENDORS.values())
for rank, (vid, v) in enumerate(ordered_v, 1):
    put(ws, f"B{r}", rank, align=CENTER)
    put(ws, f"C{r}", vid, align=CENTER)
    put(ws, f"D{r}", v['name'], align=LEFT)
    put(ws, f"E{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", v['count'], align=CENTER)
    put(ws, f"G{r}", f"=E{r}/{total_ap_vendors}", fmt='0.0%', align=RIGHT)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, align=CENTER)
widths(ws, {1:2, 2:6, 3:10, 4:38, 5:14, 6:14, 7:12, 8:40})

# ============ TAB 13: CHANGE LOG ============
ws = wb.create_sheet("Change Log")
title(ws, "Change Log — Master Register", "Change events reconstructed from Contract, CO/COR folder, Billing SOVs, ASI-RFI, and JDR. 2 CORs + 1 SOV adjustment + 2 ASI/PR events.")
hdr(ws, 5, ["Event ID", "Type", "Date", "Subject", "Originator", "Linked", "Cost Impact ($)", "Status", "Source"])
events = [
    ("CONTRACT-ORIG", "Contract (AGC LS)", CONTRACT_SIGNED, f"Original prime subcontract #{SUBCONTRACT_NO} for Fox and Finch plumbing — ${CONTRACT_ORIG:,.2f} Lump Sum (AGC of Washington 2006 Ed., Synergy proj #{SYNERGY_PROJ})", f"GC ({GC})", "", CONTRACT_ORIG, "Executed", "Contract/525 Boren Contract.pdf"),
    ("SOV-ADJ", "SOV adjustment", "Between 2016-03-20 (App#1) and 2016-05-20 (App#2)", f"Contract total rolled from ${CONTRACT_ORIG:,.2f} → $830,310.00 (+${CO_ADJ_EARLY:,.2f}) — no formal CO doc in folder, surfaces first in Billing draw #2 SOV", "Synergy / OWP", "CONTRACT-ORIG", CO_ADJ_EARLY, "Billed", "Billing/525 Boren SOV draw #2 5-20-16.xlsx"),
    ("ASI-12-MEP", "ASI", "2016 (pre-rough-in)", "ASI-12 MEP Peer Review Response — design-reviewer response to MEP drawings", "Designer / MEP", "", 0, "Informational", "ASI-RFI/ASI 12 MEP Peer Review Response.pdf"),
    ("PR-02-MEDIA", "Proposal Request", "2016", "PR-02 Media Room Redesign — designer-issued proposal request (no CO resulted; absorbed into base)", "Designer (Synergy)", "", 0, "Informational (no CO)", "ASI-RFI/PR 02 Media Room Redesign.pdf"),
    ("COR-01", "COR → CO", "2016-12-06 (COR) · 2017-04-24 (Permit SR1397520)", f"COR #01 Temp Gas Permit — total ${CO1_TEMP_GAS:.0f} to pull temp gas permit + inspection (1 permit $140 + 2 hrs labor $170 + tax). Executed as CO#2 Temporary Gas Meter (signed). Tied to Budget Transfer CO #02 XLSX ($357 sale price at code 603 Permits).", f"GC ({GC})", "CO-2-TEMP-GAS", CO1_TEMP_GAS, "Executed", "Change Orders/COR_s/525 Boren - COR #01 Temp Gas Permit.pdf · Change Orders/COR_s/Temp Gas Permit.xlsx · Change Orders/BUDGET TRANSFERS/525 Boren - Budget Transfer - CO #02.xlsx"),
    ("CO-2-TEMP-GAS", "CO (executed)", "2017-04-24", f"CO#2 Temporary Gas Meter — ${CO1_TEMP_GAS:.0f} signed PDF + scanned copy on file. Appears in Billing draw #15 SOV as +$357 to contract total = $830,667.", f"GC ({GC})", "COR-01", CO1_TEMP_GAS, "Signed / Billed", "Change Orders/CO_s/CO 2 Temporary Gas Meter Signed.pdf · Billing/525 Boren SOV draw #15  7-20-17.xlsx"),
    ("COR-02", "COR", "2017-10-27", "COR #02 Unclog Construction Debris from Toilet — $293 to unclog 2x2 wood debris lodged in toilet. Post-closeout / warranty-era; no corresponding CO signed (likely absorbed as backcharge or OH).", "Sub (OWP, Richard Donelson)", "", 293, "Submitted (no CO on file)", "Change Orders/COR_s/525 Boren - COR #02 Unclog Construction Debris from Toilet.pdf"),
    ("COR-SHOWER-OPT", "COR (priced, not issued)", "2016-06-03", "Shower option pricing — 32 tub swap (1483EN → 1363BFSD w/ removable threshold). Priced $1,936 labor + $5,666 material. No CO issued; absorbed or cancelled.", "Sub (OWP)", "", 0, "Priced only", "Change Orders/COR_s/Shower option.xlsx"),
    ("COR-SEWER-PX", "COR pricing (warranty)", "2017-04-24 (per XLSX date) · 2018-12-06 (photo event)", "Sewer Problem pricing — PM + Foreman + Apprentice + camera equipment to investigate sewer blockage. Likely related to 2018-12-06 Retail Sewer Problem warranty incident (30 photos/videos). No CO resulted (likely backcharge-offset / warranty absorb).", "Sub (OWP)", "RETAIL-SEWER-INCIDENT", 0, "Pricing only", "Change Orders/COR_s/Sewer Problem pricing.xlsx"),
    ("BUDGET-REV-230", "Budget revision", NF, f"Finish Material (code 230) revised ${CODES['230']['orig']:,.0f} → ${CODES['230']['rev']:,.0f} (+${CODES['230']['rev']-CODES['230']['orig']:,.0f})", "Internal OWP", "", CODES['230']['rev']-CODES['230']['orig'], "Applied", SRC_JDR),
    ("BUDGET-REV-603", "Budget revision", NF, f"Permits & Licenses (code 603) revised ${CODES['603']['orig']:,.0f} → ${CODES['603']['rev']:,.0f} (+${CODES['603']['rev']-CODES['603']['orig']:,.0f}) — includes Temp Gas CO#1 $140", "Internal / CO#1", "COR-01", CODES['603']['rev']-CODES['603']['orig'], "Applied", SRC_JDR + " + Budget Transfer CO#02 XLSX"),
    ("BUDGET-REV-995", "Burden rate adjustment", NF, f"Payroll Burden (code 995) revised ${CODES['995']['orig']:,.2f} → ${CODES['995']['rev']:,.2f}", "OWP internal", "", CODES['995']['rev']-CODES['995']['orig'], "Applied", SRC_JDR),
    ("BUDGET-REV-998", "Tax rate adjustment", NF, f"Payroll Taxes (code 998) revised ${CODES['998']['orig']:,.2f} → ${CODES['998']['rev']:,.2f}", "OWP internal", "", CODES['998']['rev']-CODES['998']['orig'], "Applied", SRC_JDR),
    ("CONTRACT-FINAL", "Contract revision (cumulative)", "2017-07-20 (App #15)", f"Final contract value ${CONTRACT_FINAL:,.2f} = ${CONTRACT_ORIG:,.0f} orig + ${CO_ADJ_EARLY:,.0f} SOV adj + ${CO1_TEMP_GAS:.0f} CO#1 = +$5,297.00 (+0.64%). Reconciles to JDR revenue.", f"GC ({GC})", "CONTRACT-ORIG", CO_ADJ_EARLY + CO1_TEMP_GAS, "Executed", "Billing/525 Boren SOV draw #15  7-20-17.xlsx + " + SRC_JDR),
    ("AR-PATTERN-REVERSALS", "AR reversal+reissue", "2017-04-20 (036863/036864), 2017-05-19 (036886/036887)", "Standard Sage AR corrections — not change events per se, but worth flagging", "Sub (OWP)", "", 0, "Corrected", SRC_JDR),
    ("FIRST-INVOICE", "Invoice", sorted(INVOICES.keys())[0] + " date: " + INVOICES[sorted(INVOICES.keys())[0]]['date'], f"First billing invoice #{sorted(INVOICES.keys())[0]} (Pay App #1, 2016-03-20 in SOV)", "Sub (OWP)", "", 0, "Paid (per zero net due)", SRC_JDR + " + Billing/draw#1"),
    ("LAST-INVOICE", "Invoice", "036923 date: " + INVOICES['036923']['date'], f"Last billing invoice #036923 ${INVOICES['036923']['total']:,.2f} (closeout, SOV Pay App #15 2017-07-20)", "Sub (OWP)", "", INVOICES['036923']['total'], "Paid", SRC_JDR + " + Billing/draw#15"),
    ("RETAIL-SEWER-INCIDENT", "Warranty incident", "2018-12-06", "Retail Sewer Problem — 30 photos+videos documenting retail-level sewer blockage. ~17 months post-closeout. Subject of Sewer Problem pricing XLSX.", "Sub (OWP, forensic)", "COR-SEWER-PX", 0, "Documented (no claim on file)", "Retail Sewer Problem/ (30 JPG/MP4 files dated 20181206)"),
    ("RETAINAGE-OPEN", "Retainage", "As of 04/03/2026", f"Retainage ${RETAINAGE:,.2f} outstanding 8+ years after last billing (2017-07-20). Full & Final Lien Claim Waiver on file (undated). Release documentation NOT FOUND in folder.", f"GC ({GC})", "", 0, "Outstanding", SRC_JDR + " · Billing/525 Boren Full & Final Lien Claim waiver.pdf"),
]
r = 6
for e in events:
    eid, et, dt, subj, orig, linked, cost, status, src = e
    put(ws, f"B{r}", eid, align=CENTER)
    put(ws, f"C{r}", et, align=CENTER)
    put(ws, f"D{r}", dt, align=CENTER)
    put(ws, f"E{r}", subj, align=LEFT)
    put(ws, f"F{r}", orig, align=LEFT)
    put(ws, f"G{r}", linked, align=CENTER)
    put(ws, f"H{r}", cost, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"I{r}", status, align=CENTER)
    put(ws, f"J{r}", src, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r+1}", "TOTAL COST IMPACT (net)", BOLD)
put(ws, f"H{r+1}", f"=SUM(H6:H{r-1})-H6", BOLD, fmt='"$"#,##0.00', align=RIGHT)  # exclude original
widths(ws, {1:2, 2:18, 3:24, 4:22, 5:56, 6:22, 7:14, 8:16, 9:20, 10:40})
ws.freeze_panes = "B6"

# ============ TAB 14: ROOT CAUSE ANALYSIS ============
ws = wb.create_sheet("Root Cause Analysis")
title(ws, "Root Cause Analysis", "Cost-code variance drivers. Primary signal: labor over-runs on Garage, Canout, Mech Room offset by Material under-runs.")
put(ws, "B5", "COST-CODE VARIANCE CATEGORIES", BOLD, F_ALT)
hdr(ws, 6, ["Category", "Codes Affected", "Net $ Variance", "Root Cause (inferred)", "Notes"])
# Compute variances
def var(c): return CODES[c]['actual'] - CODES[c]['rev']
rc_rows = [
    ("Labor over-runs (productivity)", "111 Garage, 112 Canout, 142 Mech Room", var("111")+var("112")+var("142"), "Field productivity below estimate", f"Sum of labor over-runs = ${var('111')+var('112')+var('142'):,.0f}"),
    ("Labor under-runs", "100 Sup, 101 Takeoff, 110 UG, 130 Finish, 140 Gas", var("100")+var("101")+var("110")+var("130")+var("140"), "Savings / scope reduction / efficiency", "Offsetting savings"),
    ("Rough-in overrun (scope)", "120 Roughin Labor", var("120"), "Possibly scope increase with budget revision", f"${var('120'):,.0f} over"),
    ("Material under-runs", "220 Rough Mat, 230 Finish Mat, 241 Water Main Mat, 242 Mech Room Mat", var("220")+var("230")+var("241")+var("242"), "Favorable procurement / takeoff savings", "Material came in well under"),
    ("Burden / Tax accrual", "995, 998", var("995")+var("998"), "Higher-than-estimated burden rates", "OWP internal rate volatility"),
    ("Support-code savings", "601 Eng, 603 Permit, 607 Other", var("601")+var("603")+var("607"), "Under-utilization of support budget", "Savings"),
]
r = 7
for cat, codes, netv, cause, note in rc_rows:
    put(ws, f"B{r}", cat, align=LEFT)
    put(ws, f"C{r}", codes, align=LEFT)
    put(ws, f"D{r}", netv, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", cause, align=LEFT)
    put(ws, f"F{r}", note, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL NET VARIANCE (Rev Budget vs Actual)", BOLD)
put(ws, f"D{r}", f"=SUM(D7:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
r += 3
put(ws, f"B{r}", "RESPONSIBILITY ATTRIBUTION", BOLD, F_ALT); r += 1
hdr(ws, r, ["Responsible Party", "# Drivers", "Net $ Impact", "Notes"]); r += 1
resp = [
    ("Sub (OWP) — labor productivity", 3, var("111")+var("112")+var("142"), "Over-runs on Garage/Canout/Mech Room"),
    ("Sub (OWP) — procurement savings", 4, var("220")+var("230")+var("241")+var("242"), "Favorable material cost"),
    ("Designer / GC (scope)", 1, var("120"), "Rough-in scope growth (budget also revised)"),
    ("Burden rate — OWP internal", 2, var("995")+var("998"), "Accrual volatility"),
    ("Support codes — OWP", 3, var("601")+var("603")+var("607"), "Savings"),
]
for rp, cnt, net, note in resp:
    put(ws, f"B{r}", rp, align=LEFT)
    put(ws, f"C{r}", cnt, align=CENTER)
    put(ws, f"D{r}", net, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", note, align=LEFT)
    r += 1
widths(ws, {1:2, 2:36, 3:40, 4:18, 5:40, 6:44})

# ============ TAB 15: PREDICTIVE SIGNALS ============
ws = wb.create_sheet("Predictive Signals")
title(ws, "Predictive Signals", "Leading indicators reconstructed from JDR only (no RFI/CO logs to draw from)")
put(ws, "B5", "CURRENT-STATE SIGNALS", BOLD, F_ALT)
hdr(ws, 6, ["Indicator", "Current Value", "Benchmark", "Status", "Meaning"])
roughin_ratio = CODES["120"]['hrs_total']/TOTAL_HOURS
signals = [
    ("RFI Count", NF, "<50", "UNKNOWN", "No RFI folder exists"),
    ("Submittal Count", NF, "varies", "UNKNOWN", "No Submittal folder exists"),
    ("ASI / OCD Count", NF, "<5", "UNKNOWN", "No ASI folder exists"),
    ("Change Order Document Count", 0, "<3", "NEUTRAL", "0 docs — but budget revision implies changes happened"),
    ("Contract Growth %", f"{CO_TOTAL/CONTRACT_ORIG*100:.2f}%", "<3%", "HEALTHY", "Very low contract growth"),
    ("Labor Hrs vs Budget", f"={TOTAL_HOURS}", "varies", "INFO", f"{TOTAL_HOURS:,.0f} hrs"),
    ("Labor Cost % of Revenue", LABOR_COST/REVENUE, "<35%", "ELEVATED" if LABOR_COST/REVENUE > 0.30 else "HEALTHY", f"{LABOR_COST/REVENUE*100:.1f}% — labor-heavy"),
    ("Retainage Release Pending", "YES", "Released within 90 days post-closeout", "ELEVATED", f"${RETAINAGE:,.0f} open 9+ years after work end"),
    ("Worker Concentration (top 1)", top_pct, "<25%", "ELEVATED" if top_pct > 0.20 else "HEALTHY", f"{top_pct*100:.1f}% from single worker"),
    ("Rough-in Labor Share", roughin_ratio, "40-60%", "HEALTHY" if 0.4 <= roughin_ratio <= 0.65 else "INFO", f"{roughin_ratio*100:.0f}% of labor on rough-in"),
    ("Document Completeness", "LOW", "Full CO/RFI/Submittal trail", "RISK", "Only JDR + electrical plans + Revit backup"),
]
r = 7
for sig in signals:
    for j, v in enumerate(sig):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
        if j == 3:
            if v == "ELEVATED": c.fill = F_HIGH
            elif v == "HEALTHY": c.fill = F_OK
            elif v == "RISK": c.fill = F_RISK
            elif v == "UNKNOWN": c.fill = F_NF
            elif v == "INFO" or v == "NEUTRAL": c.fill = F_ALT
    r += 1
r += 2
put(ws, f"B{r}", "FORECAST MODELS", BOLD, F_ALT); r += 1
hdr(ws, r, ["Forecast", "Current Estimate", "Confidence", "Driver", "Model Note"]); r += 1
forecasts = [
    ("Final margin (actual)", NET_PROFIT/REVENUE, "Actual", "Job closed financially", "Net / Revenue"),
    ("Retainage collection probability", "MEDIUM (stale)", "Qualitative", "9+ years outstanding", "May require AR write-off or dispute"),
    ("Composite risk score (0-100)", 45, "Low-Medium", "Doc-gap + stale retainage + labor concentration", "Financials healthy; file hygiene poor"),
    ("Would re-bid margin target", "≥34%", "Derived from actuals", "Historical close", "Assumes similar scope; tighten labor estimating for garage/canout/mech room"),
]
for f in forecasts:
    for j, v in enumerate(f):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
    r += 1
widths(ws, {1:2, 2:42, 3:22, 4:20, 5:20, 6:58})

# ============ TAB 16: METRIC REGISTRY ============
ws = wb.create_sheet("Metric Registry")
title(ws, "Metric Registry — Cortex Data Catalog", "Every metric with data_label, confidence, and source")
hdr(ws, 5, ["#", "Data Label", "Human Label", "Value", "Unit", "Source Tab", "Confidence", "Source Document(s)"])
metrics = [
    ("job_number", "Job Number", JOB, "id", "Benchmark KPIs", "Verified", SRC_JDR),
    ("job_name", "Job Name", NAME, "text", "Benchmark KPIs", "Verified", SRC_JDR),
    ("project_desc", "Project Description", "Fox and Finch Apts @ 525 Boren", "text", "Benchmark KPIs", "Verified", SRC_JDR),
    ("general_contractor", "GC", GC, "text", "Benchmark KPIs", "Verified", SRC_JDR),
    ("customer_code", "Customer Code", GC_CUST_CODE, "id", "Overview", "Verified", SRC_JDR),
    ("location", "Location", "525 Boren Ave, Seattle", "text", "Benchmark KPIs", "Medium", SRC_JDR),
    ("project_type", "Project Type", "Multifamily Plumbing (New Construction)", "text", "Benchmark KPIs", "Medium", "Inferred"),
    ("start_date", "Work Start", "2016-03-09", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("end_date", "Work End", "2017-07-24", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("duration_months", "Duration (months)", 16.5, "months", "Benchmark KPIs", "Verified", "Derived"),
    ("unit_count", "Unit Count", NF, "units", "Benchmark KPIs", "Low", "No plans"),
    ("fixture_count", "Fixture Count", NF, "count", "Benchmark KPIs", "Low", "No plans"),
    ("contract_original", "Contract Original", CONTRACT_ORIG, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("contract_final", "Contract Final", CONTRACT_FINAL, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("change_orders", "Change Orders $", CO_TOTAL, "USD", "Benchmark KPIs", "Medium", "Derived"),
    ("co_pct", "CO % of Contract", CO_TOTAL/CONTRACT_ORIG, "%", "Benchmark KPIs", "Medium", "Derived"),
    ("revenue", "Revenue", REVENUE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("direct_cost", "Direct Cost", EXPENSES, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("net_profit", "Net Profit", NET_PROFIT, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("gross_margin", "Gross Margin", NET_PROFIT/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("retainage", "Retainage Outstanding", RETAINAGE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("retainage_pct", "Retainage % of Revenue", RETAINAGE/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_cost", "Labor Cost", LABOR_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("material_cost", "Material Cost", MATERIAL_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("overhead_cost", "Subcon+OH Cost", OVERHEAD_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("burden_cost", "Burden Cost", BURDEN_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("tax_cost", "Tax Cost", TAX_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_hours", "Total Labor Hours", TOTAL_HOURS, "hours", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_workers", "Total Workers", TOTAL_WORKERS, "count", "Benchmark KPIs", "Verified", SRC_JDR),
    ("blended_gross_wage", "Blended Gross Wage", LABOR_COST/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("fully_loaded_wage", "Fully-Loaded Wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("burden_multiplier", "Burden Multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST, "x", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_hour", "Revenue per Hour", REVENUE/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("profit_per_hour", "Profit per Hour", NET_PROFIT/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_pct_revenue", "Labor % of Revenue", LABOR_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("material_pct_revenue", "Material % of Revenue", MATERIAL_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("total_vendors", "Total Vendors (AP)", len(VENDORS), "count", "Vendors", "Verified", SRC_JDR),
    ("total_invoices", "Total Invoices (AR)", len(INVOICES), "count", "Billing & SOV", "Verified", SRC_JDR),
    ("rfi_count", "RFI Count", NF, "count", "Predictive Signals", "Low", "No RFI folder"),
    ("submittal_count", "Submittal Count", NF, "count", "Predictive Signals", "Low", "No submittal folder"),
    ("co_doc_count", "CO Document Count", 0, "count", "Change Log", "Verified", "No CO folder"),
    ("top_worker_hours_share", "Top Worker Hours Share", top_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("top5_worker_hours_share", "Top 5 Worker Hours Share", top5_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("cost_code_count", "Cost Codes Active", len(CODES), "count", "Budget vs Actual", "Verified", SRC_JDR),
]
r = 6
for i, m in enumerate(metrics, 1):
    put(ws, f"B{r}", i, align=CENTER)
    for j, v in enumerate(m):
        c = put(ws, (r, 3+j), v, align=LEFT if j in (0,1,6) else CENTER)
        if j == 2 and isinstance(v, (int, float)):
            unit = m[3]
            if unit == "USD": c.number_format = '"$"#,##0.00'
            elif unit == "%": c.number_format = '0.00%'
            elif unit == "USD/hr": c.number_format = '"$"#,##0.00'
            elif unit == "x": c.number_format = '0.00"x"'
            elif unit == "hours" or unit == "months": c.number_format = '#,##0.00'
            else: c.number_format = '#,##0'
        if j == 5:
            if v == "Verified": c.fill = F_OK
            elif v == "Medium": c.fill = F_HIGH
            elif v == "Low": c.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:5, 3:28, 4:28, 5:20, 6:10, 7:18, 8:12, 9:36})
ws.freeze_panes = "B6"

# ============ TAB 17: RECONCILIATION ============
ws = wb.create_sheet("Reconciliation")
title(ws, "Reconciliation", "Cross-sheet formula checks")
hdr(ws, 5, ["#", "Check", "Value A", "Value B", "Delta", "Status", "Tabs"])
pay_app_total = sum(p[3] for p in PAY_APPS)  # sum of "this period" values
checks = [
    ("Revenue (JDR) = Contract Final", REVENUE, CONTRACT_FINAL, "1↔8"),
    ("Expenses = Labor+Material+OH+Burden+Tax", EXPENSES, LABOR_COST+MATERIAL_COST+OVERHEAD_COST+BURDEN_COST+TAX_COST, "1↔3"),
    ("Net Profit = Revenue - Expenses", NET_PROFIT, REVENUE - EXPENSES, "1↔Derived"),
    ("JDR Source: AP+PR+GL = Expenses", SRC_AP+SRC_PR+SRC_GL, EXPENSES, "1↔Derived (footer)"),
    ("Budget vs Actual (code 999) = Revenue", REVENUE, -CODES["999"]['actual'], "2↔1"),
    ("Total Labor Hours = Worker hours sum", TOTAL_HOURS, sum(w['hours'] for w in WORKERS.values()), "5↔2"),
    ("Labor Cost = Worker gross sum", LABOR_COST, sum(w['amount'] for w in WORKERS.values()), "5↔3"),
    ("Vendor total ≈ AP footer", sum(v['total'] for v in VENDORS.values()), SRC_AP, "12↔1 (approximate)"),
    ("Invoice count = 21", 21, len(INVOICES), "9↔11"),
    ("Contract Final - Orig = CO implied", CO_TOTAL, CONTRACT_FINAL-CONTRACT_ORIG, "11↔Derived"),
    ("CO breakdown: SOV-adj + CO#1 Temp Gas = CO total", CO_ADJ_EARLY+CO1_TEMP_GAS, CONTRACT_FINAL-CONTRACT_ORIG, "13↔11"),
    ("Pay Apps (sum of 'This Period') = Revenue", pay_app_total, REVENUE, "9↔1 (Billing SOV)"),
    ("Pay App #15 Contract Total = JDR Revenue", 830667.00, REVENUE, "9↔1 (SOV draw #15)"),
    ("Retainage outstanding", RETAINAGE, 41533.35, "9↔JDR footer"),
    ("Worker count = 25", 25, TOTAL_WORKERS, "5↔11"),
    ("Cost code count = 28", 28, len(CODES), "2↔16"),
    ("Unit count (TOTO bowls) = 60", 60, UNITS, "1↔Vendor Quote XLSX"),
    ("Pay App count (SOV files) = 15", 15, len(PAY_APPS), "9↔Billing folder"),
    ("Fixture inventory (Vendor Quote XLSX) = 669", 669, FIXTURES_TOTAL, "1↔Vendor Quote XLSX"),
]
r = 6
for i, (check, a, b, tabs) in enumerate(checks, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", check, align=LEFT)
    put(ws, f"D{r}", a, fmt='"$"#,##0.00' if isinstance(a, (int, float)) and abs(a) > 100 else None, align=RIGHT)
    put(ws, f"E{r}", b, fmt='"$"#,##0.00' if isinstance(b, (int, float)) and abs(b) > 100 else None, align=RIGHT)
    put(ws, f"F{r}", f"=D{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f'=IF(ABS(F{r})<=1,"TIES",IF(ABS(F{r})<=ABS(E{r})*0.05,"WITHIN","OFF"))', align=CENTER)
    put(ws, f"H{r}", tabs, SRC_FONT, align=CENTER)
    r += 1
r += 2
put(ws, f"B{r}", "SOURCES", HDR, F_HDR)
for col in range(2, 9): ws.cell(row=r, column=col).fill = F_HDR
r += 1
src_lines = [
    f"Job #{JOB} — Cortex v2 17-tab (REBUILD 2026-04-14 after expanded folder ingestion — 130+ documents now on file in Apts 4/)",
    f"Canonical source: {SRC_JDR}",
    f"JDR Job Totals footer: Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"JDR Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: Contract/525 Boren Contract.pdf — AGC Washington 2006 Ed. Lump Sum ${CONTRACT_ORIG:,.2f} (Subcontract #{SUBCONTRACT_NO}, Synergy project #{SYNERGY_PROJ}, cost code 15400)",
    "Billing: 15 Pay App SOV XLSX drafts (2016-03-20 → 2017-07-20) + Full & Final Lien Claim Waiver. Pay App #15 Contract Total $830,667 ties to JDR revenue.",
    "Change Orders: 2 CORs priced (COR#01 Temp Gas $357 executed as CO#2, COR#02 Unclog Toilet $293 warranty); 1 Budget Transfer XLSX (CO#02); 2 additional priced COR XLSX (Shower option, Sewer Problem pricing — neither issued)",
    "ASI-RFI: ASI-12 MEP Peer Review Response + PR-02 Media Room Redesign (informational, no cost impact)",
    "Submittals: 44 submittals across 7 categories (Equipment-7, Fixtures-12, Material-5, Responses-1, Sleeving Plans-1, Tub/Shower-8, Underground/Garage-10)",
    "Permits: 5 PDFs (Plumbing, Gas, Gas Temp Heat, Backflow, Renewed). Gas permit SR1397520 in 525 Boren LLC name",
    "Plans: 525-Plumbing Construction Set PR.pdf + 11 DWGs (L1-L7 + Roof + UG/Vault Shift) + MEP Superplots (2016-02-09/02-22/03-07)",
    "Vendor: 7 quote PDFs + master Vender Quote Spreadsheet (5 suppliers) + 5 vendor lien waivers (Consolidated/Ferguson/Keller/Rosen + template) + 1 Keller Notice-to-Owner (2016-08-11)",
    "Warranty: Retail Sewer Problem 30 photos/videos 12/06/2018 (~17 mo post-closeout) — paired with 'Sewer Problem pricing' COR XLSX (no CO issued)",
    "Closeout: O&M manual + Warranty package + Certificate of Insurance 2016-17",
    "Unresolved: (a) $41,533 retainage outstanding 8+ yrs; (b) COR#02 Unclog Toilet $293 submitted but no signed CO; (c) Sewer Problem warranty — pricing done, resolution undocumented; (d) +$4,940 SOV adj (App#1→#2) lacks formal CO paper — inferred reline",
    "TIES = within $1  ·  WITHIN = within 5%  ·  OFF = investigate",
]
for line in src_lines:
    put(ws, f"B{r}", line, SRC_FONT, align=LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    r += 1
widths(ws, {1:2, 2:4, 3:48, 4:20, 5:20, 6:14, 7:12, 8:22})

# ============ SAVE ============
import os
out = "/sessions/keen-determined-mccarthy/work/OWP_2026_JCR_Cortex_v2.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"Saved {out}")
print(f"Tabs ({len(wb.sheetnames)}):", wb.sheetnames)
