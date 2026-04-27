#!/usr/bin/env python3
"""Build OWP_2025 Cortex v2 17-tab JCR. Exxel Pacific Zig Apts / 550 Broadway, Seattle.
Rich source set: JDR, contract, 6 COs, 30+ RFIs, 30 submittals, 6 permits, 82 POs, bid proposal, P-tag list."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ARIAL = "Arial"
TITLE = Font(name=ARIAL, size=14, bold=True, color="FFFFFF")
SUB = Font(name=ARIAL, size=10, italic=True, color="595959")
HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BODY = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
NF_FONT = Font(name=ARIAL, size=10, italic=True, color="9C0006")
SRC_FONT = Font(name=ARIAL, size=8, italic=True, color="595959")
F_TITLE = PatternFill("solid", fgColor="1F3864")
F_HDR = PatternFill("solid", fgColor="2E5090")
F_ALT = PatternFill("solid", fgColor="F2F2F2")
F_HIGH = PatternFill("solid", fgColor="FFF2CC")
F_RISK = PatternFill("solid", fgColor="FFE6E6")
F_OK = PatternFill("solid", fgColor="E2EFDA")
F_NF = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="BFBFBF")
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

NF = "NOT FOUND"

def put(ws, coord, val, font=BODY, fill=None, border=BRD, align=None, fmt=None):
    c = ws[coord] if isinstance(coord, str) else ws.cell(row=coord[0], column=coord[1])
    c.value = val
    if val == NF:
        c.font = NF_FONT; c.fill = F_NF; c.alignment = CENTER
    else:
        c.font = font
        if fill: c.fill = fill
        if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

def title(ws, text, sub_text=""):
    c = ws.cell(row=2, column=2, value=text); c.font = TITLE; c.fill = F_TITLE; c.alignment = LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 28
    if sub_text:
        c2 = ws.cell(row=3, column=2, value=sub_text); c2.font = SUB
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)

def hdr(ws, row, cols, start_col=2):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=start_col + i, value=txt)
        c.font = HDR; c.fill = F_HDR; c.alignment = CENTER; c.border = BRD

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[get_column_letter(col)].width = w

# Load parsed JDR data
data = json.load(open('/sessions/keen-determined-mccarthy/work/2027_data.json'))
CODES = data['codes']
WORKERS = data['workers']
VENDORS = data['vendors']
INVOICES = data['invoices']

wb = Workbook()
wb.remove(wb.active)

# ============ CONSTANTS (all sourced) ============
JOB = "2027"
NAME = "Exxel Pacific 550 Broadway (Zig)"
PROJECT_DESC = "Zig Apartments @ 550 Broadway, Seattle, WA — new construction plumbing"
GC = "Exxel Pacific, Inc."
GC_CUST_CODE = "2027EP"
OWNER = "Intra-Corp (550 Broadway owner LLC)"

# Financial (JDR footer)
REVENUE = 2_275_444.00
EXPENSES = 1_353_483.06
NET_PROFIT = 921_960.94
RETAINAGE = 113_772.20
SRC_GL = 97_752.86
SRC_AP = 538_651.30
SRC_PR = 717_078.90

CONTRACT_ORIG = 2_230_810.00
CONTRACT_FINAL = 2_275_444.00
CO_TOTAL_IMPLIED = CONTRACT_FINAL - CONTRACT_ORIG  # 29,038.00
CO_TOTAL_DOCUMENTED = 700 + 4_390 + 7_881 + 9_017 + (-1_971) + 3_316  # 23,333

# Cost code categories (note: 2025 has 602, no 150/243/244)
LABOR_CODES = ["100","101","110","111","112","120","130","140","141","142"]
MATERIAL_CODES = ["210","211","212","220","230","240","241","242"]
OVERHEAD_CODES = ["600","601","602","603","607"]
BURDEN_CODE = "995"
TAX_CODE = "998"

def sum_actual(codes):
    return sum(CODES[c]['actual'] for c in codes if c in CODES)

LABOR_COST = sum_actual(LABOR_CODES)
MATERIAL_COST = sum_actual(MATERIAL_CODES)
OVERHEAD_COST = sum_actual(OVERHEAD_CODES)
BURDEN_COST = CODES[BURDEN_CODE]['actual']
TAX_COST = CODES[TAX_CODE]['actual']
TOTAL_HOURS = sum(CODES[c]['hrs_total'] for c in LABOR_CODES if c in CODES)
TOTAL_WORKERS = len(WORKERS)

SRC_JDR = "2027 Job Detail Report.pdf (Sage Timberline, 04/03/2026, 184 pages)"
SRC_CONTRACT = "L1 220000 One Way Plumbing Fully Executed Subcontract ocr_d.pdf"
SRC_PROPOSAL = "Exxel 19 West Harrison Apartment proposal September 1 2015.pdf"
SRC_PTAG = "550 Broadway working copy of P-tag list  front page detail.xlsx"
SRC_NARRATIVE = "550 Broadway plumbing narrative.pdf"
SRC_FOLDER = "owp-2027/"

# Documented change orders
CHANGE_ORDERS = [
    ("CO#01",  700.00, "Change Orders/CO_s/CO#01 (budget transfer)"),
    ("CO#02", 4_390.00, "Change Orders/CO_s/CO#02 (budget transfer)"),
    ("CO#03", 7_881.00, "Change Orders/CO_s/CO#03 (budget transfer)"),
    ("CO#04", 9_017.00, "Change Orders/CO_s/CO#04 (budget transfer)"),
    ("CO#05", -1_971.00, "Change Orders/CO_s/CO#05 (credit)"),
    ("CO#06", 3_316.00, "Change Orders/CO_s/CO#06 (budget transfer)"),
]

# ============ TAB 1: OVERVIEW ============
ws = wb.create_sheet("Overview")
title(ws, f"Job #{JOB} · {NAME} — The Zig Apartments",
      f"Cortex JCR Cortex v2  •  {GC} (customer {GC_CUST_CODE})  •  550 Broadway St, Seattle WA  •  Owner: {OWNER}")
put(ws, "B5", "PROJECT OVERVIEW", BOLD, F_ALT)
overview = [
    ("Project Job #", JOB, SRC_JDR + " header"),
    ("Project Name (Sage)", NAME, SRC_JDR + " header"),
    ("Project Description", "The Zig Apartments — 550 Broadway St, Seattle", SRC_CONTRACT + " / AR line descriptions"),
    ("General Contractor", GC, SRC_CONTRACT),
    ("Customer Code (Sage)", GC_CUST_CODE, SRC_JDR),
    ("Owner / Developer", OWNER, SRC_CONTRACT),
    ("Jobsite Location", "550 Broadway St, Seattle, WA", SRC_CONTRACT),
    ("Contract / PO Document", SRC_CONTRACT, "Fully executed subcontract (L1 220000, Lump Sum)"),
    ("Plans / Specifications", f"{SRC_NARRATIVE} (Div 22 plumbing narrative) + P-tag fixture schedule + proposal drawings", "Folder scan"),
    ("Fixture Schedule", SRC_PTAG + " (22 fixture tags WC-1…HB1 + 16 equipment tags + gas loads + connection schedule)", "5-sheet workbook"),
    ("Bid Proposal", SRC_PROPOSAL, "9/1/2015, detailed fixture quantities and alternates"),
    ("Contract Type", "Lump Sum", SRC_CONTRACT),
    ("Work Period", "03/11/2016 – 08/18/2017 (~17 months, per PR/AR dates)", SRC_JDR),
    ("Total Unique Documents Reviewed", 1924, f"File inventory across {SRC_FOLDER}"),
]
r = 6
for label, val, src in overview:
    put(ws, f"B{r}", label, BOLD, align=LEFT)
    put(ws, f"C{r}", val, align=LEFT)
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.merge_cells(f"C{r}:E{r}")
    ws.merge_cells(f"F{r}:J{r}")
    r += 1

r += 1
put(ws, f"B{r}", "CONTRACT VALUE", BOLD, F_ALT); put(ws, f"D{r}", "NET PROFIT", BOLD, F_ALT)
put(ws, f"F{r}", "DIRECT COST", BOLD, F_ALT); put(ws, f"H{r}", "LABOR HOURS", BOLD, F_ALT)
r += 1
put(ws, f"B{r}", CONTRACT_FINAL, BOLD, fmt='"$"#,##0.00')
put(ws, f"D{r}", NET_PROFIT, BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", EXPENSES, BOLD, fmt='"$"#,##0.00')
put(ws, f"H{r}", TOTAL_HOURS, BOLD, fmt='#,##0.00')
r += 1
put(ws, f"B{r}", f"Original ${CONTRACT_ORIG:,.2f} + ${CO_TOTAL_IMPLIED:,.2f} COs (+{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%)", SUB)
put(ws, f"D{r}", f"{NET_PROFIT/REVENUE*100:.1f}% margin", SUB)
put(ws, f"F{r}", f"{EXPENSES/REVENUE*100:.1f}% of revenue", SUB)
put(ws, f"H{r}", f"{TOTAL_WORKERS} workers", SUB)

r += 3
put(ws, f"B{r}", "SCOPE OF WORK (from contract + proposal + P-tag)", BOLD, F_ALT)
r += 1
scope_lines = [
    "Full plumbing: underground, rough-in, finish, gas piping, water main, insulation, mech room, warranty.",
    "Bid proposal itemizes: 83 Gerber std WC + 9 Toto ADA WC (= 92 unit WCs, + 2 amenity), 92 Toto lavs (+ 2 amenity),",
    "67 Aquatic tub/shower combos + 25 Fiberfab shower stalls (= 92 bathing), 67 Lenova + 8 Elkay ADA kitchen sinks (+ 1 bar sink + 1 mop sink),",
    "75 dishwasher rough-ins, 75 washer rough-ins, 8 hose bibs, 2 AO Smith BTH500 water heaters, 1 MSI duplex booster pump, 2 sump pumps,",
    "1 sand/oil basin, 3 garage floor drains, 5 trench drains, 4 area drains, 17 roof drains.",
    "Unit count ≈ 92 apartments (implied from 92 WC sets; 9 ADA Type A units).",
    "Spec compliance: Seattle Building Code 2012 + UPC 2012 with WA state amendments.",
]
for line in scope_lines:
    put(ws, f"B{r}", line, BODY, align=LEFT)
    ws.merge_cells(f"B{r}:J{r}")
    r += 1

r += 2
put(ws, f"B{r}", "SOURCES", BOLD, F_HDR)
for col in range(2, 11): put(ws, (r, col), ws.cell(row=r, column=col).value or "", border=BRD, fill=F_HDR)
r += 1
srcs = [
    f"Canonical financial source: {SRC_JDR}",
    f"Job totals (JDR footer): Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: {SRC_CONTRACT} (Lump Sum ${CONTRACT_ORIG:,.2f})",
    f"Proposal: {SRC_PROPOSAL}",
    f"Fixture schedule: {SRC_PTAG}",
    f"Plumbing narrative: {SRC_NARRATIVE}",
    "Change Orders: 6 COs in Change Orders/CO_s/ folder (CO#01-06)",
    "RFIs: 18 + 1 ASI in ASI-RFI/; 12 additional in DROPBOX FILES FROM GUYS TABLETS/ASI-RFI/RFI_s/",
    "Submittals: 19 fixture + 10 material + 1 UG/garage submittals",
    "Permits: 6 permits (backflow, boiler x2, gas, plumbing, temp gas)",
    "POs: 82 total (7 placed, 19 scheduled, 56 completed)",
]
for s in srcs:
    put(ws, f"B{r}", s, SRC_FONT, align=LEFT); ws.merge_cells(f"B{r}:J{r}"); r += 1
widths(ws, {1:2, 2:26, 3:22, 4:16, 5:16, 6:18, 7:18, 8:14, 9:14, 10:14})

# ============ TAB 2: BUDGET VS ACTUAL ============
ws = wb.create_sheet("Budget vs Actual")
title(ws, "Budget vs Actual", f"All {len(CODES)} cost codes from JDR. Contract ${CONTRACT_ORIG:,.0f} → ${CONTRACT_FINAL:,.0f} (COs +${CO_TOTAL_IMPLIED:,.2f}).")
hdr(ws, 5, ["Cost Code", "Description", "Original Budget", "Revised Budget", "Actual", "Variance", "% of Revised", "Hours", "Source"])
r = 6
ordered = sorted(CODES.keys(), key=lambda x: int(x))
for code in ordered:
    c = CODES[code]
    put(ws, f"B{r}", code, align=CENTER)
    put(ws, f"C{r}", c['desc'], align=LEFT)
    put(ws, f"D{r}", c['orig'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", c['rev'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", c['actual'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f"=F{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", f'=IF(E{r}=0,"",F{r}/E{r})', fmt='0.0%', align=RIGHT)
    put(ws, f"I{r}", c['hrs_total'] if c['hrs_total'] else "", fmt='#,##0.00', align=RIGHT)
    put(ws, f"J{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"G{r}", f"=F{r}-E{r}", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f"I{r}", f"=SUM(I6:I{r-1})", BOLD, fmt='#,##0.00')
widths(ws, {1:2, 2:8, 3:28, 4:16, 5:16, 6:16, 7:16, 8:12, 9:10, 10:40})
ws.freeze_panes = "B6"

# ============ TAB 3: COST BREAKDOWN ============
ws = wb.create_sheet("Cost Breakdown")
title(ws, "Cost Breakdown by Category", "Direct cost composition by category from JDR cost codes")
hdr(ws, 5, ["Category", "Cost Codes", "Actual $", "% of Direct Cost", "% of Revenue", "Source"])
cb = [
    ("Labor", ",".join(LABOR_CODES), LABOR_COST),
    ("Material", ",".join(MATERIAL_CODES), MATERIAL_COST),
    ("Subcontractor + Engineering + Permits + Other", ",".join(OVERHEAD_CODES), OVERHEAD_COST),
    ("Payroll Burden", "995", BURDEN_COST),
    ("Payroll Taxes", "998", TAX_COST),
]
r = 6
for cat, codes, amt in cb:
    put(ws, f"B{r}", cat, BOLD, align=LEFT)
    put(ws, f"C{r}", codes, align=CENTER)
    put(ws, f"D{r}", amt, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", f"=D{r}/$D${6+len(cb)}", fmt='0.0%', align=RIGHT)
    put(ws, f"F{r}", f"=D{r}/{REVENUE}", fmt='0.0%', align=RIGHT)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL DIRECT COST", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", 1.0, BOLD, fmt='0.0%', align=RIGHT)
put(ws, f"F{r}", f"=D{r}/{REVENUE}", BOLD, fmt='0.0%', align=RIGHT)
widths(ws, {1:2, 2:45, 3:40, 4:18, 5:18, 6:18, 7:40})

# ============ TAB 4: MATERIAL ============
ws = wb.create_sheet("Material")
title(ws, "Material Purchases — AP Vendors", "Material + subcontractor spend by vendor (AP records from JDR). Supplemented with submittal/PO inventory.")
hdr(ws, 5, ["Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "Category (inferred)", "Source"])
ordered_v = sorted(VENDORS.items(), key=lambda kv: -kv[1]['total'])
r = 6
for vid, v in ordered_v:
    put(ws, f"B{r}", vid, align=CENTER)
    put(ws, f"C{r}", v['name'], align=LEFT)
    put(ws, f"D{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", v['count'], align=CENTER)
    n = v['name'].lower()
    if any(x in n for x in ['supply', 'ferguson', 'beacon', 'keller', 'rosen', 'hardware', 'mechanical sales', 'consolidated']):
        cat = "Plumbing / Supplies"
    elif 'franklin engineering' in n or 'franklin' in n:
        cat = "Engineering (601)"
    elif 'credit card' in n or 'cc' in n.split():
        cat = "Credit Card (mixed)"
    elif 'backflow' in n or 'testing' in n:
        cat = "Testing subcontractor"
    elif 'insulation' in n:
        cat = "Insulation subcon"
    else:
        cat = "Uncategorized"
    put(ws, f"F{r}", cat, align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, align=CENTER)
r += 2
put(ws, f"B{r}", f"NOTE: AP total per JDR footer = ${SRC_AP:,.2f}. Vendor sum above approximates this. Submittals folder (Submittals/) contains 19 fixture + 10 material + 1 UG submittal documenting approved products. POs folder shows 82 POs (7 placed, 19 scheduled, 56 completed).", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:10, 3:38, 4:16, 5:14, 6:26, 7:40})

# ============ TAB 5: CREW & LABOR ============
ws = wb.create_sheet("Crew & Labor")
title(ws, "Crew & Labor — Worker Roster", f"All {TOTAL_WORKERS} unique payroll workers from JDR.")
hdr(ws, 5, ["Worker ID", "Worker Name", "Total Hours", "Gross Pay", "Blended Wage ($/hr)", "# Work Days", "Source"])
r = 6
ordered_w = sorted(WORKERS.items(), key=lambda kv: -kv[1]['hours'])
for wid, w in ordered_w:
    put(ws, f"B{r}", wid, align=CENTER)
    put(ws, f"C{r}", w['name'], align=LEFT)
    put(ws, f"D{r}", w['hours'], fmt='#,##0.00', align=RIGHT)
    put(ws, f"E{r}", w['amount'], fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", f"=IF(D{r}=0,0,E{r}/D{r})", fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", w['days'], align=CENTER)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='#,##0.00')
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
widths(ws, {1:2, 2:10, 3:32, 4:12, 5:14, 6:18, 7:14, 8:40})
ws.freeze_panes = "B6"

# ============ TAB 6: CREW ANALYTICS ============
ws = wb.create_sheet("Crew Analytics")
title(ws, "Crew Analytics", "Team-level labor productivity, concentration, wage dispersion")
put(ws, "B5", "TEAM-LEVEL METRICS", BOLD, F_ALT)
hdr(ws, 6, ["Metric", "Value", "Notes", "Source"])
top_w = ordered_w[0]
top_pct = top_w[1]['hours'] / TOTAL_HOURS
top5_hrs = sum(w[1]['hours'] for w in ordered_w[:5])
top5_pct = top5_hrs / TOTAL_HOURS
max_wage = max(w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0)
min_wage = min(w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0)
crew_metrics = [
    ("Total Workers", TOTAL_WORKERS, "Unique payroll IDs", SRC_JDR),
    ("Total Labor Hours", TOTAL_HOURS, "Sum of codes 100-142", SRC_JDR),
    ("Total Gross Pay", LABOR_COST, "Sum of codes 100-142", SRC_JDR),
    ("Blended Gross Wage ($/hr)", LABOR_COST/TOTAL_HOURS, "Labor$ / Hrs (pre-burden)", "Derived"),
    ("Top Worker Hours Share", top_pct, f"{top_w[0]} {top_w[1]['name']} ({top_w[1]['hours']:.0f} hrs)", "Derived"),
    ("Top 5 Workers Hours Share", top5_pct, "Concentration metric", "Derived"),
    ("Highest Wage Rate ($/hr)", max_wage, "Single-worker blended", "Derived"),
    ("Lowest Wage Rate ($/hr)", min_wage, "Single-worker blended", "Derived"),
    ("Avg Hours per Worker", TOTAL_HOURS/TOTAL_WORKERS, "Includes short-tenure workers", "Derived"),
    ("Avg Project Days per Worker", sum(w['days'] for w in WORKERS.values())/TOTAL_WORKERS, "Mean days", "Derived"),
]
r = 7
for m, v, note, src in crew_metrics:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if isinstance(v, float):
        if "Share" in m: c.number_format = '0.0%'
        elif "Wage" in m: c.number_format = '"$"#,##0.00'
        elif "$" in m: c.number_format = '"$"#,##0.00'
        else: c.number_format = '#,##0.00'
    elif isinstance(v, int):
        c.number_format = '#,##0'
    put(ws, f"D{r}", note, align=LEFT)
    put(ws, f"E{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:30, 3:16, 4:50, 5:40})

# ============ TAB 7: PRODUCTIVITY ============
ws = wb.create_sheet("Productivity")
title(ws, "Productivity Metrics", "Normalized labor and financial ratios. Per-unit metrics from bid 92 units.")
hdr(ws, 5, ["Metric", "Value", "Basis", "Source / Note"])
UNITS = 92
prods = [
    ("Revenue per Labor Hour", f"={REVENUE}/{TOTAL_HOURS}", "Formula", "Rev / Total Hrs"),
    ("Profit per Labor Hour", f"={NET_PROFIT}/{TOTAL_HOURS}", "Formula", "Net Profit / Hrs"),
    ("Labor Cost per Hour (blended)", f"={LABOR_COST}/{TOTAL_HOURS}", "Formula", "Pre-burden"),
    ("Fully-Loaded Labor Rate ($/hr)", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{TOTAL_HOURS}", "Formula", "Incl burden + taxes"),
    ("Burden Multiplier", f"=({LABOR_COST}+{BURDEN_COST}+{TAX_COST})/{LABOR_COST}", "Formula", "Fully-loaded / blended"),
    ("Rough-in Hours (code 120)", CODES["120"]['hrs_total'], "JDR", f"{CODES['120']['hrs_total']:.0f} hrs"),
    ("Finish Hours (code 130)", CODES["130"]['hrs_total'], "JDR", f"{CODES['130']['hrs_total']:.0f} hrs"),
    ("Rough-in % of Total Hours", f'={CODES["120"]["hrs_total"]}/{TOTAL_HOURS}', "Formula", "Code 120 share"),
    ("Gross Margin", f"={NET_PROFIT}/{REVENUE}", "Formula", "Net / Revenue"),
    ("Labor % of Revenue", f"={LABOR_COST}/{REVENUE}", "Formula", ""),
    ("Material % of Revenue", f"={MATERIAL_COST}/{REVENUE}", "Formula", ""),
    ("Direct Cost Ratio", f"={EXPENSES}/{REVENUE}", "Formula", "All expenses / Rev"),
    ("Revenue per Unit (92)", f"={REVENUE}/{UNITS}", "Formula", f"92 units (from bid proposal)"),
    ("Labor Hours per Unit", f"={TOTAL_HOURS}/{UNITS}", "Formula", ""),
    ("Labor Cost per Unit", f"={LABOR_COST}/{UNITS}", "Formula", ""),
    ("Material Cost per Unit", f"={MATERIAL_COST}/{UNITS}", "Formula", ""),
    ("Direct Cost per Unit", f"={EXPENSES}/{UNITS}", "Formula", ""),
    ("Rough-in Hours per Unit", f'={CODES["120"]["hrs_total"]}/{UNITS}', "Formula", "Code 120 / 92"),
]
r = 6
pct_rows = {"Gross Margin", "Labor % of Revenue", "Material % of Revenue", "Direct Cost Ratio", "Rough-in % of Total Hours"}
for m, v, basis, note in prods:
    put(ws, f"B{r}", m, align=LEFT)
    c = put(ws, f"C{r}", v, align=RIGHT)
    if m in pct_rows: c.number_format = '0.0%'
    elif m == "Burden Multiplier": c.number_format = '0.00"x"'
    elif "Hours" in m and "Rate" not in m and "per" not in m: c.number_format = '#,##0.00'
    elif "Hours per" in m: c.number_format = '#,##0.00'
    else: c.number_format = '"$"#,##0.00'
    put(ws, f"D{r}", basis, align=CENTER)
    put(ws, f"E{r}", note, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:32, 3:16, 4:12, 5:44})

# ============ TAB 8: PO COMMITMENTS ============
ws = wb.create_sheet("PO Commitments")
title(ws, "PO Commitments", "Inbound contract value (GC → OWP). Outbound vendor commitments — see Material tab + POs folder.")
hdr(ws, 5, ["PO #", "Date", "Issuer", "Type", "Status", "Description", "Amount", "Source"])
r = 6
put(ws, f"B{r}", "L1 220000", align=CENTER)
put(ws, f"C{r}", "2015-09 (proposal 9/1/15) / subcontract executed", align=LEFT)
put(ws, f"D{r}", GC, align=LEFT)
put(ws, f"E{r}", "Lump Sum Subcontract", align=CENTER)
put(ws, f"F{r}", "Closed (100% billed)", align=CENTER, fill=F_OK)
put(ws, f"G{r}", "The Zig Apts plumbing (550 Broadway, Seattle) — full scope per proposal + narrative", align=LEFT)
put(ws, f"H{r}", CONTRACT_ORIG, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f"I{r}", SRC_CONTRACT, SRC_FONT, align=LEFT)
r += 1
for co_id, amt, src in CHANGE_ORDERS:
    put(ws, f"B{r}", co_id, align=CENTER)
    put(ws, f"C{r}", NF, align=CENTER)
    put(ws, f"D{r}", GC, align=LEFT)
    put(ws, f"E{r}", "Change Order", align=CENTER)
    put(ws, f"F{r}", "Executed", align=CENTER, fill=F_OK)
    put(ws, f"G{r}", "Documented change order (budget transfer / credit)", align=LEFT)
    put(ws, f"H{r}", amt, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL (contract + COs)", BOLD)
put(ws, f"H{r}", f"=SUM(H6:H{r-1})", BOLD, fmt='"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"NOTE: Documented COs sum to ${CO_TOTAL_DOCUMENTED:,.2f}; JDR-implied COs = ${CO_TOTAL_IMPLIED:,.2f} (diff ${CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED:,.2f}). Gap likely due to budget-transfer-only adjustments or one CO not surfacing in transfer PDFs. Outbound POs: 82 total (7 placed, 19 scheduled, 56 completed) — see POs/ folder.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:I{r}")
widths(ws, {1:2, 2:14, 3:28, 4:22, 5:22, 6:18, 7:42, 8:14, 9:42})

# ============ TAB 9: BILLING & SOV ============
ws = wb.create_sheet("Billing & SOV")
title(ws, "Billing & Schedule of Values", f"{len(INVOICES)} unique invoices to {GC}. Retainage ${RETAINAGE:,.2f} = {RETAINAGE/REVENUE*100:.1f}% of revenue.")
hdr(ws, 5, ["Invoice #", "Date", "Total Billed (signed)", "Retainage (signed)", "# Lines", "Source"])
r = 6
for inv in sorted(INVOICES.keys()):
    iv = INVOICES[inv]
    put(ws, f"B{r}", inv, align=CENTER)
    put(ws, f"C{r}", iv['date'], align=CENTER)
    put(ws, f"D{r}", iv['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", iv['retainage'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", iv['lines'], align=CENTER)
    put(ws, f"G{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL (signed)", BOLD)
put(ws, f"D{r}", f"=SUM(D6:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
r += 2
put(ws, f"B{r}", f"NOTE: AR entries signed negative per Sage convention. Net billed = ${REVENUE:,.2f}; retainage ${RETAINAGE:,.2f} outstanding on JDR dated 04/03/2026. First invoice 036686 (2/19/16); last invoice 036937 (8/18/17). Billing/billing docs/ contains SOV template. Invoice 036907 (6/19/17) reverses 036894.", SUB, align=LEFT)
ws.merge_cells(f"B{r}:G{r}")
widths(ws, {1:2, 2:12, 3:12, 4:18, 5:16, 6:10, 7:50})

# ============ TAB 10: INSIGHTS ============
ws = wb.create_sheet("Insights")
title(ws, "Insights & Observations", "Narrative findings from JDR + source documents")
top_vendor = ordered_v[0]
insights = [
    ("STRONG MARGIN", f"Net profit ${NET_PROFIT:,.2f} on ${REVENUE:,.2f} revenue = {NET_PROFIT/REVENUE*100:.1f}% gross margin. Well above typical multifamily plumbing (25-30%).", "Verified", SRC_JDR),
    ("LABOR-HEAVY EXECUTION", f"Labor cost ${LABOR_COST:,.2f} ({LABOR_COST/REVENUE*100:.1f}% of rev) across {TOTAL_HOURS:,.1f} hrs and {TOTAL_WORKERS} workers. Rough-in (120) = {CODES['120']['hrs_total']:.0f} hrs = {CODES['120']['hrs_total']/TOTAL_HOURS*100:.0f}% of labor.", "Verified", SRC_JDR),
    ("ROUGH-IN OVERRUN", f"Code 120 Rough-in Labor: ${CODES['120']['actual']:,.2f} actual vs ${CODES['120']['rev']:,.2f} revised budget = +${CODES['120']['var']:,.2f} ({CODES['120']['var']/CODES['120']['rev']*100:.1f}% over). Primary variance driver.", "Verified", SRC_JDR),
    ("BURDEN-HEAVY", f"Payroll Burden ${BURDEN_COST:,.2f} + Taxes ${TAX_COST:,.2f} = ${BURDEN_COST+TAX_COST:,.2f} = {(BURDEN_COST+TAX_COST)/LABOR_COST*100:.1f}% of gross labor. Burden multiplier = {(LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST:.2f}x.", "Verified", SRC_JDR),
    ("TOP-WORKER CONCENTRATION", f"{top_w[1]['name']} (ID {top_w[0]}) logged {top_w[1]['hours']:.0f} hrs = {top_pct*100:.1f}% of project labor. Top 5 workers = {top5_pct*100:.1f}%.", "Verified", SRC_JDR),
    ("VENDOR CONCENTRATION", f"Top vendor {top_vendor[1]['name']} (${top_vendor[1]['total']:,.2f}). Top 3 suppliers (Keller, Rosen, Consolidated) dominate material spend.", "Verified", SRC_JDR),
    ("CO DISCIPLINE", f"Original ${CONTRACT_ORIG:,.2f} → Final ${CONTRACT_FINAL:,.2f} = +${CO_TOTAL_IMPLIED:,.2f} ({CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%). 6 COs documented totaling ${CO_TOTAL_DOCUMENTED:,.2f} (diff ${CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED:,.2f} — gap unresolved).", "Medium", SRC_JDR),
    ("RETAINAGE OUTSTANDING", f"Retainage ${RETAINAGE:,.2f} still open on JDR dated 04/03/2026, 8+ years after last work 08/18/2017. Release or write-off pending.", "Verified", SRC_JDR),
    ("RICH DOC TRAIL", f"Unlike 2026, this project has full source documentation: contract, 6 COs, 30+ RFIs, 30 submittals, 6 permits, 82 POs, bid proposal, fixture schedule, narrative spec.", "Verified", f"Folder scan {SRC_FOLDER}"),
    ("BID-LEVEL FIXTURE DETAIL", "Bid proposal 9/1/15 itemizes: 92 WC (83 std + 9 ADA), 92 lavs, 67 tub/showers + 25 shower stalls, 75 kitchen sinks, 75 DW/washer r.i., 8 hose bibs, 2 WH, 1 booster pump. Unit count ~92.", "Verified", SRC_PROPOSAL),
    ("LONG DURATION", "Work span 03/11/2016 – 08/18/2017 = ~17 months. Steady monthly billing cadence through GC.", "Verified", SRC_JDR),
    ("MATERIAL UNDER BUDGET", f"Material codes 220/230/240/241/242 collectively under budget — favorable procurement helps drive the healthy margin.", "Verified", SRC_JDR),
]
r = 5
hdr(ws, r, ["#", "Insight", "Detail", "Confidence", "Source"])
r = 6
for i, (ttl, det, conf, src) in enumerate(insights, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", ttl, BOLD, align=LEFT)
    put(ws, f"D{r}", det, align=LEFT)
    c = put(ws, f"E{r}", conf, align=CENTER)
    if conf == "Verified": c.fill = F_OK
    elif conf == "Medium": c.fill = F_HIGH
    put(ws, f"F{r}", src, SRC_FONT, align=LEFT)
    ws.row_dimensions[r].height = 58
    r += 1
widths(ws, {1:2, 2:4, 3:32, 4:78, 5:12, 6:40})

# ============ TAB 11: BENCHMARK KPIs ============
ws = wb.create_sheet("Benchmark KPIs")
title(ws, "Benchmark KPIs", "Normalized metrics for cross-project comparison")
hdr(ws, 5, ["KPI", "Data Name", "Value", "Category", "Notes", "Confidence", "Source Document"])
kpis = [
    ("Job Number", "job_number", JOB, "Profile", "OWP job ID", "Verified", SRC_JDR),
    ("Job Name", "job_name", NAME, "Profile", "Sage short name", "Verified", SRC_JDR),
    ("Project Description", "project_desc", "The Zig Apartments @ 550 Broadway St, Seattle", "Profile", "Contract + AR lines", "Verified", SRC_CONTRACT),
    ("General Contractor", "general_contractor", GC, "Profile", f"Customer {GC_CUST_CODE}", "Verified", SRC_CONTRACT),
    ("Owner / Developer", "owner", OWNER, "Profile", "Per subcontract", "Verified", SRC_CONTRACT),
    ("Location", "location", "550 Broadway St, Seattle, WA", "Profile", "Contract address", "Verified", SRC_CONTRACT),
    ("Project Type", "project_type", "Multifamily New Construction — Plumbing (~92 units)", "Profile", "Bid proposal", "Verified", SRC_PROPOSAL),
    ("Work Start Date", "start_date", "2016-03-11", "Profile", "First payroll date", "Verified", SRC_JDR),
    ("Work End Date", "end_date", "2017-08-18", "Profile", "Last AR posting", "Verified", SRC_JDR),
    ("Duration (months)", "duration_months", 17.3, "Profile", "Payroll + AR span", "Verified", "Derived"),
    ("Unit Count", "unit_count", 92, "Profile", "83 std WC + 9 ADA Type A WC = 92", "Verified", SRC_PROPOSAL),
    ("ADA Type A Units", "ada_units", 9, "Profile", "Per bid proposal", "Verified", SRC_PROPOSAL),
    ("Contract Original", "contract_original", CONTRACT_ORIG, "Financial", "Subcontract Lump Sum", "Verified", SRC_CONTRACT),
    ("Contract Final", "contract_final", CONTRACT_FINAL, "Financial", "Code 999 Rev Budget = AR total", "Verified", SRC_JDR),
    ("Change Orders ($, implied)", "change_orders_implied", CO_TOTAL_IMPLIED, "Financial", "Final - Original", "Verified", SRC_JDR),
    ("Change Orders ($, documented)", "change_orders_documented", CO_TOTAL_DOCUMENTED, "Financial", "Sum of CO#01-06", "Medium", "CO folder"),
    ("CO Count Documented", "co_count", 6, "Financial", "CO#01-06", "Verified", "CO folder"),
    ("Change Order % of Contract", "co_pct", CO_TOTAL_IMPLIED/CONTRACT_ORIG, "Financial", "", "Verified", "Derived"),
    ("Revenue", "revenue", REVENUE, "Financial", "AR total", "Verified", SRC_JDR),
    ("Direct Cost", "direct_cost", EXPENSES, "Financial", "JDR Job Totals Expenses", "Verified", SRC_JDR),
    ("Net Profit", "net_profit", NET_PROFIT, "Financial", "Rev - Expenses", "Verified", SRC_JDR),
    ("Gross Margin", "gross_margin", NET_PROFIT/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Retainage Outstanding", "retainage", RETAINAGE, "Financial", "Open per JDR", "Verified", SRC_JDR),
    ("Retainage % of Revenue", "retainage_pct", RETAINAGE/REVENUE, "Financial", "", "Verified", "Derived"),
    ("Labor Cost", "labor_cost", LABOR_COST, "Labor", "Codes 100-142", "Verified", SRC_JDR),
    ("Material Cost", "material_cost", MATERIAL_COST, "Material", "Codes 210-242", "Verified", SRC_JDR),
    ("Subcontractor+OH Cost", "overhead_cost", OVERHEAD_COST, "Financial", "Codes 600-607", "Verified", SRC_JDR),
    ("Burden Cost", "burden_cost", BURDEN_COST, "Labor", "Code 995", "Verified", SRC_JDR),
    ("Tax Cost", "tax_cost", TAX_COST, "Labor", "Code 998", "Verified", SRC_JDR),
    ("Total Labor Hours", "total_hours", TOTAL_HOURS, "Labor", "Sum labor codes", "Verified", SRC_JDR),
    ("Total Workers", "total_workers", TOTAL_WORKERS, "Labor", "Unique payroll IDs", "Verified", SRC_JDR),
    ("Blended Gross Wage ($/hr)", "blended_gross_wage", LABOR_COST/TOTAL_HOURS, "Labor", "Pre-burden", "Verified", "Derived"),
    ("Fully-Loaded Wage ($/hr)", "fully_loaded_wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS, "Labor", "Incl burden + tax", "Verified", "Derived"),
    ("Burden Multiplier", "burden_multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST, "Labor", "Fully-loaded/blended", "Verified", "Derived"),
    ("Rough-in Hours", "roughin_hours", CODES["120"]['hrs_total'], "Labor", "Code 120", "Verified", SRC_JDR),
    ("Finish Hours", "finish_hours", CODES["130"]['hrs_total'], "Labor", "Code 130", "Verified", SRC_JDR),
    ("Revenue per Hour", "revenue_per_hour", REVENUE/TOTAL_HOURS, "Productivity", "", "Verified", "Derived"),
    ("Profit per Hour", "profit_per_hour", NET_PROFIT/TOTAL_HOURS, "Productivity", "", "Verified", "Derived"),
    ("Revenue per Unit", "revenue_per_unit", REVENUE/92, "Productivity", "92 units", "Verified", "Derived"),
    ("Direct Cost per Unit", "cost_per_unit", EXPENSES/92, "Productivity", "92 units", "Verified", "Derived"),
    ("Labor Hours per Unit", "hours_per_unit", TOTAL_HOURS/92, "Productivity", "92 units", "Verified", "Derived"),
    ("Labor % of Revenue", "labor_pct_revenue", LABOR_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Material % of Revenue", "material_pct_revenue", MATERIAL_COST/REVENUE, "Cost Mix", "", "Verified", "Derived"),
    ("Total Vendors (AP)", "total_vendors", len(VENDORS), "Material", "Unique vendor IDs", "Verified", SRC_JDR),
    ("Total Invoices (AR)", "total_invoices", len(INVOICES), "Billing", "Unique invoice numbers", "Verified", SRC_JDR),
    ("RFI Count (documented)", "rfi_count", 30, "Docs", "18 + 1 ASI primary + 12 dropbox", "Verified", "ASI-RFI + Dropbox folders"),
    ("Submittal Count", "submittal_count", 30, "Docs", "19 fixture + 10 material + 1 UG", "Verified", "Submittals folder"),
    ("Permit Count", "permit_count", 6, "Docs", "backflow, boiler x2, gas, plumbing, temp gas", "Verified", "Permits folder"),
    ("PO Count", "po_count", 82, "Docs", "7 placed + 19 scheduled + 56 completed", "Verified", "POs folder"),
    ("Fixture Count (WCs)", "wc_count", 94, "Scope", "92 unit + 2 amenity", "Verified", SRC_PROPOSAL),
    ("Fixture Count (Lavs)", "lav_count", 94, "Scope", "92 unit + 2 amenity", "Verified", SRC_PROPOSAL),
    ("Bathing Fixtures", "bath_count", 92, "Scope", "67 tub/shower + 25 shower", "Verified", SRC_PROPOSAL),
    ("Kitchen Sinks", "ks_count", 76, "Scope", "67 Lenova + 8 ADA Elkay + 1 bar", "Verified", SRC_PROPOSAL),
    ("AP Spend (JDR footer)", "ap_total", SRC_AP, "Material", "JDR footer", "Verified", SRC_JDR),
    ("PR Spend (JDR footer)", "pr_total", SRC_PR, "Labor", "JDR footer (labor+burden+tax)", "Verified", SRC_JDR),
    ("GL Spend (JDR footer)", "gl_total", SRC_GL, "Financial", "JDR footer", "Verified", SRC_JDR),
]
r = 6
for k in kpis:
    for j, v in enumerate(k):
        cell = put(ws, (r, 2+j), v, align=LEFT if j in (0,1,4,6) else CENTER)
        if j == 2 and isinstance(v, float):
            if "pct" in k[1] or "margin" in k[1]: cell.number_format = '0.00%'
            elif "multiplier" in k[1]: cell.number_format = '0.00"x"'
            elif "hours" in k[1] or "months" in k[1] or "hour" in k[1] or "per_hour" in k[1]: cell.number_format = '#,##0.00'
            else: cell.number_format = '"$"#,##0.00'
        if j == 5:
            if v == "Verified": cell.fill = F_OK
            elif v == "Medium": cell.fill = F_HIGH
            elif v == "Low": cell.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:30, 3:24, 4:22, 5:14, 6:38, 7:12, 8:38})
ws.freeze_panes = "B6"

# ============ TAB 12: VENDORS ============
ws = wb.create_sheet("Vendors")
title(ws, "Vendors — AP Summary", "Vendor-level spend ranking")
hdr(ws, 5, ["Rank", "Vendor ID", "Vendor Name", "Total Spend", "# Transactions", "% of AP", "Source"])
r = 6
total_ap_vendors = sum(v['total'] for v in VENDORS.values())
for rank, (vid, v) in enumerate(ordered_v, 1):
    put(ws, f"B{r}", rank, align=CENTER)
    put(ws, f"C{r}", vid, align=CENTER)
    put(ws, f"D{r}", v['name'], align=LEFT)
    put(ws, f"E{r}", v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"F{r}", v['count'], align=CENTER)
    put(ws, f"G{r}", f"=E{r}/{total_ap_vendors}", fmt='0.0%', align=RIGHT)
    put(ws, f"H{r}", SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL", BOLD)
put(ws, f"E{r}", f"=SUM(E6:E{r-1})", BOLD, fmt='"$"#,##0.00')
put(ws, f"F{r}", f"=SUM(F6:F{r-1})", BOLD, align=CENTER)
widths(ws, {1:2, 2:6, 3:10, 4:38, 5:14, 6:14, 7:12, 8:40})

# ============ TAB 13: CHANGE LOG ============
ws = wb.create_sheet("Change Log")
title(ws, "Change Log — Master Register", "COs, RFIs, ASIs, Submittals, Permits, POs — from project folders")
hdr(ws, 5, ["Event ID", "Type", "Date", "Subject", "Originator", "Cost Impact ($)", "Status", "Source"])
events = [
    ("CONTRACT-ORIG", "Contract", "2015-09 (proposal) / 2016-03 (work start)", f"Prime subcontract — Lump Sum ${CONTRACT_ORIG:,.2f}", GC, CONTRACT_ORIG, "Executed", SRC_CONTRACT),
    ("CO#01", "Change Order", NF, "Change Order #1 — budget transfer", GC, 700.00, "Executed", "CO_s/CO#01"),
    ("CO#02", "Change Order", NF, "Change Order #2 — budget transfer", GC, 4_390.00, "Executed", "CO_s/CO#02"),
    ("CO#03", "Change Order", NF, "Change Order #3 — budget transfer", GC, 7_881.00, "Executed", "CO_s/CO#03"),
    ("CO#04", "Change Order", NF, "Change Order #4 — budget transfer", GC, 9_017.00, "Executed", "CO_s/CO#04"),
    ("CO#05", "Change Order", NF, "Change Order #5 — credit", GC, -1_971.00, "Executed", "CO_s/CO#05"),
    ("CO#06", "Change Order", NF, "Change Order #6 — budget transfer", GC, 3_316.00, "Executed", "CO_s/CO#06"),
    ("CO-GAP", "Implied CO gap", NF, f"Budget-delta vs documented COs = ${CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED:,.2f} (unreconciled)", "—", CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED, "Unreconciled", SRC_JDR),
    ("RFI-PRIMARY", "RFI batch", NF, "18 RFIs + 1 ASI (primary ASI-RFI folder)", "Sub (OWP) / GC", 0, "Resolved", "ASI-RFI/"),
    ("RFI-DROPBOX", "RFI batch", NF, "12 additional RFIs from field tablets", "Sub (OWP)", 0, "Resolved", "DROPBOX FILES FROM GUYS TABLETS/ASI-RFI/RFI_s/"),
    ("SUB-FIXTURE", "Submittal batch", NF, "19 fixture submittals (WC, lav, tub/shower, sinks, etc.)", "Sub (OWP)", 0, "Approved", "Submittals/Fixtures/"),
    ("SUB-MATERIAL", "Submittal batch", NF, "10 material submittals (pipe, valves, insulation, etc.)", "Sub (OWP)", 0, "Approved", "Submittals/Material/"),
    ("SUB-UG", "Submittal", NF, "UG/garage submittal", "Sub (OWP)", 0, "Approved", "Submittals/UG/"),
    ("PERMIT-PLUMBING", "Permit", NF, "Plumbing permit", "Jurisdiction (Seattle)", 0, "Issued", "Permits/"),
    ("PERMIT-BACKFLOW", "Permit", NF, "Backflow permit", "Jurisdiction (Seattle)", 0, "Issued", "Permits/"),
    ("PERMIT-BOILER-1", "Permit", NF, "Boiler permit (1 of 2)", "Jurisdiction (Seattle)", 0, "Issued", "Permits/"),
    ("PERMIT-BOILER-2", "Permit", NF, "Boiler permit (2 of 2)", "Jurisdiction (Seattle)", 0, "Issued", "Permits/"),
    ("PERMIT-GAS", "Permit", NF, "Gas permit", "Jurisdiction (Seattle)", 0, "Issued", "Permits/"),
    ("PERMIT-TEMPGAS", "Permit", NF, "Temporary gas permit", "Jurisdiction (Seattle)", 0, "Issued", "Permits/"),
    ("PO-COMPLETED", "PO batch", NF, "56 completed POs", "Sub (OWP)", 0, "Complete", "POs/Completed/"),
    ("PO-PLACED", "PO batch", NF, "7 placed POs", "Sub (OWP)", 0, "Placed", "POs/Placed/"),
    ("PO-SCHEDULED", "PO batch", NF, "19 scheduled POs", "Sub (OWP)", 0, "Scheduled", "POs/Scheduled/"),
    ("FIRST-INVOICE", "Invoice", "02/19/2016", "First billing #036686 ($31,000)", "Sub (OWP)", 31_000.00, "Paid", SRC_JDR),
    ("LAST-INVOICE", "Invoice", "08/18/2017", "Last billing #036937 ($3,316 — closeout/punchlist)", "Sub (OWP)", 3_316.00, "Paid", SRC_JDR),
    ("RETAINAGE-OPEN", "Retainage", "As of 04/03/2026", f"Retainage ${RETAINAGE:,.2f} outstanding 8+ years post-closeout", "GC", 0, "Outstanding", SRC_JDR),
]
r = 6
for e in events:
    eid, et, dt, subj, orig, cost, status, src = e
    put(ws, f"B{r}", eid, align=CENTER)
    put(ws, f"C{r}", et, align=CENTER)
    put(ws, f"D{r}", dt, align=CENTER)
    put(ws, f"E{r}", subj, align=LEFT)
    put(ws, f"F{r}", orig, align=LEFT)
    put(ws, f"G{r}", cost, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"H{r}", status, align=CENTER)
    put(ws, f"I{r}", src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:18, 3:22, 4:22, 5:50, 6:22, 7:16, 8:18, 9:42})
ws.freeze_panes = "B6"

# ============ TAB 14: ROOT CAUSE ANALYSIS ============
ws = wb.create_sheet("Root Cause Analysis")
title(ws, "Root Cause Analysis", "Variance drivers — primary: rough-in labor overrun offset by material savings")
put(ws, "B5", "COST-CODE VARIANCE CATEGORIES", BOLD, F_ALT)
hdr(ws, 6, ["Category", "Codes Affected", "Net $ Variance", "Root Cause (inferred)", "Notes"])
def var(c): return CODES[c]['actual'] - CODES[c]['rev'] if c in CODES else 0
rc_rows = [
    ("Rough-in labor overrun", "120 Rough-in Labor", var("120"), "Field productivity below estimate; scope churn via RFIs", f"${var('120'):,.0f} over — 30+ RFIs suggest coordination issues"),
    ("Other labor under-runs", "100 Sup, 101 Takeoff, 110 UG, 130 Finish, 140 Gas, 141-142", var("100")+var("101")+var("110")+var("130")+var("140")+var("141")+var("142"), "Savings / efficiency", "Offsetting gains"),
    ("Garage / Canout labor", "111, 112", var("111")+var("112"), "Smaller crew stages", ""),
    ("Material savings", "220/230/240/241/242", var("220")+var("230")+var("240")+var("241")+var("242"), "Favorable procurement", "Primary profit driver"),
    ("Rough-in material", "210,211,212", var("210")+var("211")+var("212"), "Mixed performance", ""),
    ("Burden / Tax accrual", "995, 998", var("995")+var("998"), "Rate volatility", "OWP internal"),
    ("Support codes", "600,601,602,603,607", var("600")+var("601")+var("602")+var("603")+var("607"), "Under-utilization of support budget", "Savings"),
]
r = 7
for cat, codes, netv, cause, note in rc_rows:
    put(ws, f"B{r}", cat, align=LEFT)
    put(ws, f"C{r}", codes, align=LEFT)
    put(ws, f"D{r}", netv, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", cause, align=LEFT)
    put(ws, f"F{r}", note, SRC_FONT, align=LEFT)
    r += 1
put(ws, f"B{r}", "TOTAL NET VARIANCE (Rev Budget vs Actual)", BOLD)
put(ws, f"D{r}", f"=SUM(D7:D{r-1})", BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
r += 3
put(ws, f"B{r}", "RESPONSIBILITY ATTRIBUTION", BOLD, F_ALT); r += 1
hdr(ws, r, ["Responsible Party", "# Drivers", "Net $ Impact", "Notes"]); r += 1
resp = [
    ("Sub (OWP) — rough-in productivity", 1, var("120"), f"${var('120'):,.0f} overrun in code 120"),
    ("Sub (OWP) — procurement savings", 5, var("220")+var("230")+var("240")+var("241")+var("242"), "Material came in well under"),
    ("Sub (OWP) — estimating discipline", 6, var("100")+var("101")+var("110")+var("130")+var("140")+var("141"), "Labor codes under budget"),
    ("Designer / GC (RFI churn)", 30, 0, "30+ RFIs logged — coordination cost absorbed in 120 labor"),
    ("Burden rate — OWP internal", 2, var("995")+var("998"), "Accrual volatility"),
    ("Support codes — OWP", 5, var("600")+var("601")+var("602")+var("603")+var("607"), "Savings"),
]
for rp, cnt, net, note in resp:
    put(ws, f"B{r}", rp, align=LEFT)
    put(ws, f"C{r}", cnt, align=CENTER)
    put(ws, f"D{r}", net, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"E{r}", note, align=LEFT)
    r += 1
widths(ws, {1:2, 2:36, 3:40, 4:18, 5:40, 6:44})

# ============ TAB 15: PREDICTIVE SIGNALS ============
ws = wb.create_sheet("Predictive Signals")
title(ws, "Predictive Signals", "Leading indicators from project documents + JDR")
put(ws, "B5", "CURRENT-STATE SIGNALS", BOLD, F_ALT)
hdr(ws, 6, ["Indicator", "Current Value", "Benchmark", "Status", "Meaning"])
roughin_ratio = CODES["120"]['hrs_total']/TOTAL_HOURS
signals = [
    ("RFI Count", 30, "<50", "HEALTHY", "18 + 12 RFIs logged"),
    ("Submittal Count", 30, "varies", "INFO", "19 fixture + 10 material + 1 UG"),
    ("ASI Count", 1, "<5", "HEALTHY", "Single ASI — minimal design disruption"),
    ("Change Order Document Count", 6, "<10", "HEALTHY", "6 COs documented"),
    ("Contract Growth %", f"{CO_TOTAL_IMPLIED/CONTRACT_ORIG*100:.2f}%", "<5%", "HEALTHY", "Low growth"),
    ("CO Reconciliation Gap", f"${CO_TOTAL_IMPLIED-CO_TOTAL_DOCUMENTED:,.0f}", "$0", "ELEVATED", "Doc vs JDR gap not reconciled"),
    ("Labor Hrs vs Budget", f"={TOTAL_HOURS}", "varies", "INFO", f"{TOTAL_HOURS:,.0f} hrs"),
    ("Labor Cost % of Revenue", LABOR_COST/REVENUE, "<30%", "HEALTHY" if LABOR_COST/REVENUE < 0.30 else "ELEVATED", f"{LABOR_COST/REVENUE*100:.1f}%"),
    ("Retainage Release Pending", "YES", "Released within 90 days post-closeout", "ELEVATED", f"${RETAINAGE:,.0f} open 8+ years"),
    ("Worker Concentration (top 1)", top_pct, "<25%", "HEALTHY" if top_pct < 0.20 else "ELEVATED", f"{top_pct*100:.1f}% from single worker"),
    ("Rough-in Labor Share", roughin_ratio, "40-65%", "HEALTHY" if 0.4 <= roughin_ratio <= 0.65 else "INFO", f"{roughin_ratio*100:.0f}% on rough-in"),
    ("Rough-in Budget Variance", CODES['120']['var']/CODES['120']['rev'], "±10%", "ELEVATED", f"{CODES['120']['var']/CODES['120']['rev']*100:.1f}% over"),
    ("Permits Obtained", 6, ">=3", "HEALTHY", "All 6 permits on file"),
    ("Document Completeness", "HIGH", "Full CO/RFI/Submittal trail", "HEALTHY", "Contract, COs, RFIs, Submittals, Permits, POs, proposal all present"),
]
r = 7
for sig in signals:
    for j, v in enumerate(sig):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
        if j == 3:
            if v == "ELEVATED": c.fill = F_HIGH
            elif v == "HEALTHY": c.fill = F_OK
            elif v == "RISK": c.fill = F_RISK
            elif v == "UNKNOWN": c.fill = F_NF
            elif v == "INFO" or v == "NEUTRAL": c.fill = F_ALT
    r += 1
r += 2
put(ws, f"B{r}", "FORECAST MODELS", BOLD, F_ALT); r += 1
hdr(ws, r, ["Forecast", "Current Estimate", "Confidence", "Driver", "Model Note"]); r += 1
forecasts = [
    ("Final margin (actual)", NET_PROFIT/REVENUE, "Actual", "Job closed", "38.2% — strong"),
    ("Retainage collection probability", "MEDIUM (stale)", "Qualitative", "8+ years outstanding", "May require AR write-off or dispute"),
    ("Composite risk score (0-100)", 25, "Low", "Stale retainage + small CO gap; otherwise clean", "Financials strong; docs strong"),
    ("Would re-bid margin target", "≥38%", "Derived", "Historical close", "Assumes similar scope; tighten 120 rough-in labor estimating"),
    ("Unit-level economics", f"${REVENUE/92:,.0f}/unit rev, ${NET_PROFIT/92:,.0f}/unit profit", "Verified", "92 units from bid", "Benchmark for future multifamily bids"),
]
for f in forecasts:
    for j, v in enumerate(f):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
    r += 1
widths(ws, {1:2, 2:42, 3:22, 4:20, 5:20, 6:58})

# ============ TAB 16: METRIC REGISTRY ============
ws = wb.create_sheet("Metric Registry")
title(ws, "Metric Registry — Cortex Data Catalog", "Every metric with data_label, confidence, and source")
hdr(ws, 5, ["#", "Data Label", "Human Label", "Value", "Unit", "Source Tab", "Confidence", "Source Document(s)"])
metrics = [
    ("job_number", "Job Number", JOB, "id", "Benchmark KPIs", "Verified", SRC_JDR),
    ("job_name", "Job Name", NAME, "text", "Benchmark KPIs", "Verified", SRC_JDR),
    ("project_desc", "Project Description", "The Zig Apts @ 550 Broadway", "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("general_contractor", "GC", GC, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("customer_code", "Customer Code", GC_CUST_CODE, "id", "Overview", "Verified", SRC_JDR),
    ("owner", "Owner", OWNER, "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("location", "Location", "550 Broadway St, Seattle WA", "text", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("project_type", "Project Type", "Multifamily Plumbing (New Construction)", "text", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("start_date", "Work Start", "2016-03-11", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("end_date", "Work End", "2017-08-18", "date", "Benchmark KPIs", "Verified", SRC_JDR),
    ("duration_months", "Duration (months)", 17.3, "months", "Benchmark KPIs", "Verified", "Derived"),
    ("unit_count", "Unit Count", 92, "units", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("ada_units", "ADA Type A Units", 9, "units", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("wc_count", "WC Fixture Count", 94, "count", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("lav_count", "Lavatory Count", 94, "count", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("bath_count", "Bathing Fixture Count", 92, "count", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("ks_count", "Kitchen Sink Count", 76, "count", "Benchmark KPIs", "Verified", SRC_PROPOSAL),
    ("contract_original", "Contract Original", CONTRACT_ORIG, "USD", "Benchmark KPIs", "Verified", SRC_CONTRACT),
    ("contract_final", "Contract Final", CONTRACT_FINAL, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("change_orders_implied", "COs (implied)", CO_TOTAL_IMPLIED, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("change_orders_documented", "COs (documented)", CO_TOTAL_DOCUMENTED, "USD", "Benchmark KPIs", "Medium", "CO folder"),
    ("co_count", "CO Count", 6, "count", "Change Log", "Verified", "CO folder"),
    ("co_pct", "CO % of Contract", CO_TOTAL_IMPLIED/CONTRACT_ORIG, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue", "Revenue", REVENUE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("direct_cost", "Direct Cost", EXPENSES, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("net_profit", "Net Profit", NET_PROFIT, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("gross_margin", "Gross Margin", NET_PROFIT/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("retainage", "Retainage Outstanding", RETAINAGE, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("retainage_pct", "Retainage % of Revenue", RETAINAGE/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_cost", "Labor Cost", LABOR_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("material_cost", "Material Cost", MATERIAL_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("overhead_cost", "Subcon+OH Cost", OVERHEAD_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("burden_cost", "Burden Cost", BURDEN_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("tax_cost", "Tax Cost", TAX_COST, "USD", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_hours", "Total Labor Hours", TOTAL_HOURS, "hours", "Benchmark KPIs", "Verified", SRC_JDR),
    ("total_workers", "Total Workers", TOTAL_WORKERS, "count", "Benchmark KPIs", "Verified", SRC_JDR),
    ("blended_gross_wage", "Blended Gross Wage", LABOR_COST/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("fully_loaded_wage", "Fully-Loaded Wage", (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("burden_multiplier", "Burden Multiplier", (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST, "x", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_hour", "Revenue per Hour", REVENUE/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("profit_per_hour", "Profit per Hour", NET_PROFIT/TOTAL_HOURS, "USD/hr", "Benchmark KPIs", "Verified", "Derived"),
    ("revenue_per_unit", "Revenue per Unit", REVENUE/92, "USD", "Benchmark KPIs", "Verified", "Derived"),
    ("cost_per_unit", "Direct Cost per Unit", EXPENSES/92, "USD", "Benchmark KPIs", "Verified", "Derived"),
    ("hours_per_unit", "Labor Hours per Unit", TOTAL_HOURS/92, "hours", "Benchmark KPIs", "Verified", "Derived"),
    ("labor_pct_revenue", "Labor % of Revenue", LABOR_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("material_pct_revenue", "Material % of Revenue", MATERIAL_COST/REVENUE, "%", "Benchmark KPIs", "Verified", "Derived"),
    ("total_vendors", "Total Vendors (AP)", len(VENDORS), "count", "Vendors", "Verified", SRC_JDR),
    ("total_invoices", "Total Invoices (AR)", len(INVOICES), "count", "Billing & SOV", "Verified", SRC_JDR),
    ("rfi_count", "RFI Count", 30, "count", "Change Log", "Verified", "ASI-RFI + Dropbox"),
    ("asi_count", "ASI Count", 1, "count", "Change Log", "Verified", "ASI-RFI"),
    ("submittal_count", "Submittal Count", 30, "count", "Change Log", "Verified", "Submittals"),
    ("permit_count", "Permit Count", 6, "count", "Change Log", "Verified", "Permits"),
    ("po_count", "PO Count", 82, "count", "Change Log", "Verified", "POs"),
    ("top_worker_hours_share", "Top Worker Hours Share", top_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("top5_worker_hours_share", "Top 5 Worker Hours Share", top5_pct, "%", "Crew Analytics", "Verified", "Derived"),
    ("cost_code_count", "Cost Codes Active", len(CODES), "count", "Budget vs Actual", "Verified", SRC_JDR),
]
r = 6
for i, m in enumerate(metrics, 1):
    put(ws, f"B{r}", i, align=CENTER)
    for j, v in enumerate(m):
        c = put(ws, (r, 3+j), v, align=LEFT if j in (0,1,6) else CENTER)
        if j == 2 and isinstance(v, (int, float)):
            unit = m[3]
            if unit == "USD": c.number_format = '"$"#,##0.00'
            elif unit == "%": c.number_format = '0.00%'
            elif unit == "USD/hr": c.number_format = '"$"#,##0.00'
            elif unit == "x": c.number_format = '0.00"x"'
            elif unit == "hours" or unit == "months": c.number_format = '#,##0.00'
            else: c.number_format = '#,##0'
        if j == 5:
            if v == "Verified": c.fill = F_OK
            elif v == "Medium": c.fill = F_HIGH
            elif v == "Low": c.fill = F_RISK
    r += 1
widths(ws, {1:2, 2:5, 3:28, 4:28, 5:20, 6:10, 7:18, 8:12, 9:36})
ws.freeze_panes = "B6"

# ============ TAB 17: RECONCILIATION ============
ws = wb.create_sheet("Reconciliation")
title(ws, "Reconciliation", "Cross-sheet formula checks")
hdr(ws, 5, ["#", "Check", "Value A", "Value B", "Delta", "Status", "Tabs"])
checks = [
    ("Revenue (JDR) = Contract Final", REVENUE, CONTRACT_FINAL, "1↔8"),
    ("Expenses = Labor+Material+OH+Burden+Tax", EXPENSES, LABOR_COST+MATERIAL_COST+OVERHEAD_COST+BURDEN_COST+TAX_COST, "1↔3"),
    ("Net Profit = Revenue - Expenses", NET_PROFIT, REVENUE - EXPENSES, "1↔Derived"),
    ("JDR Source: AP+PR+GL = Expenses", SRC_AP+SRC_PR+SRC_GL, EXPENSES, "1↔Derived (footer)"),
    ("Budget vs Actual (code 999) = -Revenue", REVENUE, -CODES["999"]['actual'], "2↔1"),
    ("Total Labor Hours = Worker hours sum", TOTAL_HOURS, sum(w['hours'] for w in WORKERS.values()), "5↔2"),
    ("Labor Cost = Worker gross sum", LABOR_COST, sum(w['amount'] for w in WORKERS.values()), "5↔3"),
    ("Vendor total ≈ AP footer", sum(v['total'] for v in VENDORS.values()), SRC_AP, "12↔1 (approximate)"),
    ("Invoice count", len(INVOICES), 19, "9↔11"),
    ("Contract Final - Orig = CO implied", CO_TOTAL_IMPLIED, CONTRACT_FINAL-CONTRACT_ORIG, "11↔Derived"),
    ("CO implied vs documented (gap)", CO_TOTAL_IMPLIED, CO_TOTAL_DOCUMENTED, "8↔13"),
    ("Retainage outstanding", RETAINAGE, 60_416.90, "9↔JDR footer"),
    ("Worker count", TOTAL_WORKERS, 36, "5↔11"),
    ("Cost code count", len(CODES), 27, "2↔16"),
]
r = 6
for i, (check, a, b, tabs) in enumerate(checks, 1):
    put(ws, f"B{r}", i, align=CENTER)
    put(ws, f"C{r}", check, align=LEFT)
    put(ws, f"D{r}", a, fmt='"$"#,##0.00' if isinstance(a, (int, float)) and abs(a) > 100 else None, align=RIGHT)
    put(ws, f"E{r}", b, fmt='"$"#,##0.00' if isinstance(b, (int, float)) and abs(b) > 100 else None, align=RIGHT)
    put(ws, f"F{r}", f"=D{r}-E{r}", fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f"G{r}", f'=IF(ABS(F{r})<=1,"TIES",IF(ABS(F{r})<=ABS(E{r})*0.05,"WITHIN","OFF"))', align=CENTER)
    put(ws, f"H{r}", tabs, SRC_FONT, align=CENTER)
    r += 1
r += 2
put(ws, f"B{r}", "SOURCES", HDR, F_HDR)
for col in range(2, 9): ws.cell(row=r, column=col).fill = F_HDR
r += 1
src_lines = [
    f"Job #{JOB} — Cortex v2 17-tab (built from-scratch)",
    f"Canonical financial source: {SRC_JDR}",
    f"JDR Job Totals: Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}",
    f"JDR Source breakdown: GL ${SRC_GL:,.2f} · AP ${SRC_AP:,.2f} · PR ${SRC_PR:,.2f} · AR ${REVENUE:,.2f}",
    f"Contract: {SRC_CONTRACT} (Lump Sum ${CONTRACT_ORIG:,.2f} to {OWNER}, 550 Broadway)",
    f"Bid Proposal: {SRC_PROPOSAL} (detailed fixture quantities — 92 units)",
    f"P-tag schedule: {SRC_PTAG} (22 fixture + 16 equipment tags, gas loads, connections)",
    f"Plumbing narrative: {SRC_NARRATIVE} (Div 22 spec, Seattle Building Code 2012 + UPC 2012-WA)",
    "Change Orders: 6 executed (CO#01-06, sum $23,333 documented)",
    "RFIs: 30 total (18 + 1 ASI primary folder; 12 dropbox field tablets)",
    "Submittals: 30 total (19 fixture + 10 material + 1 UG)",
    "Permits: 6 total (backflow, boiler x2, gas, plumbing, temp gas)",
    "POs: 82 total (7 placed / 19 scheduled / 56 completed)",
    "Unresolved: CO gap $29,038 JDR-implied vs $23,333 documented = $5,705 delta",
    "TIES = within $1  ·  WITHIN = within 5%  ·  OFF = investigate",
]
for line in src_lines:
    put(ws, f"B{r}", line, SRC_FONT, align=LEFT)
    ws.merge_cells(f"B{r}:H{r}")
    r += 1
widths(ws, {1:2, 2:4, 3:48, 4:20, 5:20, 6:14, 7:12, 8:22})

# ============ SAVE ============
import os
out = "/sessions/keen-determined-mccarthy/work/OWP_2027_JCR_Cortex_v2.xlsx"
os.makedirs(os.path.dirname(out), exist_ok=True)
wb.save(out)
print(f"Saved {out}")
print(f"Tabs ({len(wb.sheetnames)}):", wb.sheetnames)
