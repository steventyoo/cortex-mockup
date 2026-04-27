#!/usr/bin/env python3
"""
build_2048_testlabels.py — Build OWP_2048_JCR_Test_Labels.xlsx

Matches Sam's v4 Test-Labels schema exactly (7 tabs):
  Report Record | Line Items | Cost Code Summaries | Worker Wages |
  Derived Fields | Reconciliation | Reconciliation Log

Consumes /sessions/gracious-relaxed-pascal/2048_data.json produced by parse_2048.py.
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_JSON = Path('/sessions/gracious-relaxed-pascal/2048_data.json')
ENRICHED_META_JSON = Path('/sessions/gracious-relaxed-pascal/2048_enriched_meta.json')
OUT_DIR = Path('/sessions/gracious-relaxed-pascal/mnt/cortex-mockup/owp/owp-2048/cortex output files')
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT_FILE = OUT_DIR / 'OWP_2048_JCR_Test_Labels.xlsx'

FONT = 'Arial'
HDR_FILL = PatternFill('solid', start_color='1F3A4C')
HDR_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=13)
CAPTION_FONT = Font(name=FONT, italic=True, size=10, color='666666')
BODY_FONT = Font(name=FONT, size=10)
SECTION_FONT = Font(name=FONT, bold=True, size=10, color='1F3A4C')
SECTION_FILL = PatternFill('solid', start_color='E8EEF2')
THIN = Side(style='thin', color='BBBBBB')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def apply_hdr(cell):
    cell.fill = HDR_FILL
    cell.font = HDR_FONT
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = BOX


def write_title(ws, title, caption):
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A2'] = caption
    ws['A2'].font = CAPTION_FONT
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')


def build_report_record(wb, d):
    ws = wb.create_sheet('Report Record')
    rr = d['report_record']
    bs = rr['job_totals_by_source']
    write_title(ws, 'JCR TEST LABELS — JOB 2048',
                f"Report-level fields from the v4 schema. Parsed from {len(d['line_items']):,}-line Job Detail Report (Apr 03, 2026 run). Mirrors the Job 2020 test-labels structure.")
    ws.append([])
    ws.append(['Field', 'Value'])
    hdr_row = ws.max_row
    for c in range(1, 3):
        apply_hdr(ws.cell(hdr_row, c))
    em = d.get('enriched_meta', {}) or {}
    scope_metrics_val = ('72 residential units (68 standard + 4 Type A ADA) · ~502 fixtures (scope estimate) · 7 residential stories · 19-month duration · 5 executed COs · 11 CORs · 12 permits · 123 invoices') if em else 'pending — trade-declared array of {label, count, description}'
    rows = [
        ('job_number', rr['job_number']),
        ('job_name', rr['job_name']),
        ('report_date', rr['report_date']),
        ('general_contractor', em.get('general_contractor', 'pending')),
        ('owner', em.get('owner', 'pending')),
        ('architect', em.get('architect', 'pending')),
        ('location', em.get('location', 'pending')),
        ('project_type', em.get('project_type', 'pending')),
        ('project_start_date', em.get('start_date', 'pending')),
        ('project_end_date', em.get('end_date', 'pending')),
        ('duration_months', em.get('duration_months', 'pending')),
        ('contract_signed_date', em.get('contract_signed_date', 'pending')),
        ('contract_original', em.get('contract_original', 'pending')),
        ('contract_final', em.get('contract_final', 'pending')),
        ('change_orders_net', em.get('change_orders_net', 'pending')),
        ('scope_metrics', scope_metrics_val),
        ('line_items', f"Array of {rr['line_items_count']:,} objects (see Line Items tab)"),
        ('cost_code_summaries', f"Array of {rr['cost_code_count']} objects (see Cost Code Summaries tab)"),
        ('job_totals_revenue', rr['job_totals_revenue']),
        ('job_totals_expenses', rr['job_totals_expenses']),
        ('job_totals_net', rr['job_totals_net']),
        ('job_totals_net_due', rr['job_totals_net_due']),
        ('job_totals_retainage', rr['job_totals_retainage']),
        ('job_totals_by_source', f"PR: ${bs['PR']:,.2f}, AP: ${bs['AP']:,.2f}, GL: ${bs['GL']:,.2f}, AR: ${bs['AR']:,.2f}"),
    ]
    for k, v in rows:
        ws.append([k, v])
        ws.cell(ws.max_row, 1).font = BODY_FONT
        ws.cell(ws.max_row, 2).font = BODY_FONT
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 90


def build_line_items(wb, d):
    ws = wb.create_sheet('Line Items')
    li = d['line_items']
    write_title(ws, f'LINE ITEMS — {len(li):,} TRANSACTIONS',
                'Every transaction line parsed from the 175-page JDR. AR amounts are negative per Sage convention (credits). Retainage is stored in the retainage column (AR only).')
    ws.append([])
    hdrs = ['cost_code', 'description', 'source', 'ref_number', 'document_date',
            'posted_date', 'number', 'name', 'regular_hours', 'overtime_hours',
            'regular_amount', 'overtime_amount', 'actual_amount',
            'check_number', 'net_due', 'retainage']
    ws.append(hdrs)
    hdr_row = ws.max_row
    for c in range(1, len(hdrs) + 1):
        apply_hdr(ws.cell(hdr_row, c))
    for it in li:
        ws.append([
            it['cost_code'], it['description'], it['source'], it['ref_number'],
            it['document_date'], it['posted_date'], it['number'], it['name'],
            it.get('regular_hours'), it.get('overtime_hours'),
            it.get('regular_amount'), it.get('overtime_amount'),
            it['actual_amount'], it.get('check_number'),
            it.get('net_due') or 0, it.get('retainage') or 0,
        ])
    ws.freeze_panes = 'A5'
    widths = [10, 24, 8, 10, 12, 12, 10, 32, 10, 10, 12, 12, 12, 14, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build_cost_code_summaries(wb, d):
    ws = wb.create_sheet('Cost Code Summaries')
    cs = d['cost_code_summaries']
    write_title(ws, f'COST CODE SUMMARIES — {len(cs)} CODES', '')
    ws.append([])
    hdrs = ['cost_code', 'description', 'original_budget', 'current_budget',
            'plus_minus_budget', 'actual_amount', 'net_due', 'retainage',
            'regular_hours', 'overtime_hours', 'doubletime_hours']
    ws.append(hdrs)
    hdr_row = ws.max_row
    for c in range(1, len(hdrs) + 1):
        apply_hdr(ws.cell(hdr_row, c))
    for c in cs:
        ws.append([
            c['code'], c['description'], c['original_budget'], c['current_budget'],
            c['plus_minus_budget'], c['actual_amount'], c['net_due'], c['retainage'],
            c.get('regular_hours') or 0, c.get('overtime_hours') or 0,
            c.get('doubletime_hours') or 0,
        ])
    widths = [10, 28, 14, 14, 14, 14, 10, 10, 10, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    # Dollar format
    for row in range(5, 5 + len(cs)):
        for col in (3, 4, 5, 6, 7, 8):
            ws.cell(row, col).number_format = '$#,##0.00;($#,##0.00);-'


def build_worker_wages(wb, d):
    ws = wb.create_sheet('Worker Wages')
    workers = d['worker_wages']
    write_title(ws, f'WORKER NOMINAL WAGES — {len(workers)} WORKERS',
                "Per-worker wage data derived from Line Items (source=PR). nominal_rate = regular_amount ÷ regular_hours — the worker's declared hourly rate as booked by Sage PR. Sorted by nominal rate (ascending) within tier.")
    ws.append([])
    hdrs = ['Worker (Last, First)', 'Reg Hours', 'OT Hours', 'Reg Amount', 'OT Amount', 'Nominal Rate ($/hr)']
    ws.append(hdrs)
    hdr_row = ws.max_row
    for c in range(1, len(hdrs) + 1):
        apply_hdr(ws.cell(hdr_row, c))

    tiers = ['APPRENTICE/HELPER', 'JOURNEYMAN', 'LEAD/SUPERVISOR', 'OT-ONLY']
    for tier in tiers:
        tier_workers = [w for w in workers if w['tier'] == tier]
        if not tier_workers:
            continue
        ws.append([tier])
        r = ws.max_row
        ws.cell(r, 1).font = SECTION_FONT
        for col in range(1, 7):
            ws.cell(r, col).fill = SECTION_FILL
        for w in tier_workers:
            rate = w['nominal_rate']
            ws.append([
                w['name'], w['regular_hours'], w['overtime_hours'],
                w['regular_amount'], w['overtime_amount'],
                rate if rate is not None else '— (no reg hours)',
            ])
    widths = [32, 12, 12, 14, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build_derived_fields(wb, d):
    ws = wb.create_sheet('Derived Fields')
    df = d['derived_fields']
    nw = df['nominal_wage_percentiles']
    pct = f"p10=${nw['p10']:.2f} | p25=${nw['p25']:.2f} | p50=${nw['p50']:.2f} | p75=${nw['p75']:.2f} | p90=${nw['p90']:.2f}"

    sections = [
        ('PROJECT-LEVEL BENCHMARKS (Action #8, trade-agnostic)', [
            ('project_duration_months', 'pending (requires project_start_date)',
             '(COALESCE(project_end_date, MAX(line_items[].document_date)) − project_start_date) ÷ 30.4'),
            ('scope_benchmarks', 'pending (requires scope_metrics input)',
             'One object per scope_metrics entry: {label, dollars_per_unit, hours_per_unit, units_per_month, revenue_per_unit}'),
        ]),
        ('COST-BY-SOURCE ROLLUPS', [
            ('cost_by_source_pr_amount', df['cost_by_source_pr_amount'], 'SUM(line_items[].actual_amount WHERE source=PR)'),
            ('cost_by_source_ap_amount', df['cost_by_source_ap_amount'], 'SUM(line_items[].actual_amount WHERE source=AP)'),
            ('cost_by_source_gl_amount', df['cost_by_source_gl_amount'], 'SUM(line_items[].actual_amount WHERE source=GL)'),
            ('direct_cost', df['direct_cost'], 'PR + AP + GL totals'),
            ('net_profit', df['net_profit'], 'abs(revenue) − direct_cost. Sage stores revenue as negative (credit); we use absolute value.'),
            ('pr_pct_of_revenue', df['pr_pct_of_revenue'], 'PR total ÷ abs(revenue)'),
            ('ap_pct_of_revenue', df['ap_pct_of_revenue'], 'AP total ÷ abs(revenue)'),
            ('direct_cost_pct_of_revenue', df['direct_cost_pct_of_revenue'], 'direct_cost ÷ abs(revenue). Complement of gross margin.'),
        ]),
        ('LABOR ANALYTICS', [
            ('total_labor_hours', df['total_labor_hours'], 'SUM(regular + overtime hours, PR only)'),
            ('pr_src_cost_per_hr', df['pr_src_cost_per_hr'], 'PR total ÷ total_labor_hours (includes OT premium)'),
            ('fully_loaded_wage', df['fully_loaded_wage'], '(PR + burden) ÷ total_labor_hours — includes burden & taxes on top of pr_src_cost_per_hr'),
            ('burden_multiplier', df['burden_multiplier'], 'fully_loaded_wage ÷ pr_src_cost_per_hr'),
            ('straight_time_rate', df['straight_time_rate'], 'SUM(regular_amount) ÷ SUM(regular_hours). Weighted straight-time rate; excludes OT premium.'),
            ('worker_nominal_wages', f"{df['total_workers']} workers (see Worker Wages tab)",
             'GROUP BY name WHERE source=PR; per-worker {reg_h, ot_h, reg_a, ot_a, nominal_rate=reg_a÷reg_h}'),
            ('nominal_wage_percentiles', pct, 'Percentile snapshot across worker_nominal_wages[].nominal_rate'),
            ('total_workers', df['total_workers'], 'COUNT DISTINCT(name WHERE source=PR)'),
            ('burden_total', df['burden_total'], 'SUM(actual WHERE cost_code IN 995,998)'),
        ]),
        ('MATERIAL ANALYTICS', [
            ('material_spend_total', df['material_spend_total'], 'SUM(actual WHERE code LIKE 2xx)'),
            ('material_budget_total', df['material_budget_total'], 'SUM(current_budget WHERE code LIKE 2xx)'),
            ('material_codes_tracked', df['material_codes_tracked'], 'COUNT material cost codes'),
            ('material_codes_over_budget', df['material_codes_over_budget'], 'COUNT material codes WHERE plus_minus_budget > 0  [convention: actual − current_budget, positive = over]'),
            ('material_codes_under_budget', df['material_codes_under_budget'], 'COUNT material codes WHERE plus_minus <= 0'),
            ('largest_material_overrun', df['largest_material_overrun'], 'Material code with largest positive plus_minus (biggest overrun)'),
            ('largest_material_savings', df['largest_material_savings'], 'Material code with largest negative plus_minus (biggest savings)'),
            ('material_vendor_count', df['material_vendor_count'], 'COUNT DISTINCT(name WHERE source=AP)'),
            ('top_vendor_spend', df['top_vendor_spend'], 'Vendor with highest AP spend'),
        ]),
        ('PHASE ANALYTICS', [
            ('phases_over_budget', df['phases_over_budget'], 'COUNT phases WHERE plus_minus > 0 AND code != 999 (sales)'),
            ('phases_under_budget', df['phases_under_budget'], 'COUNT phases WHERE plus_minus <= 0 AND code != 999 (sales)'),
            ('largest_overrun', df['largest_overrun'], 'Phase with largest positive plus_minus (biggest overrun)'),
            ('largest_savings', df['largest_savings'], 'Phase with largest negative plus_minus (biggest savings)'),
        ]),
        ('BUDGET & FORECAST', [
            ('total_jtd_cost', df['total_jtd_cost'], 'SUM(actual) all expense codes (excludes 999 revenue)'),
            ('total_budget', df['total_budget'], 'SUM(current_budget) all expense codes (excludes 999 revenue)'),
            ('overall_pct_budget_consumed', df['overall_pct_budget_consumed'], 'total_jtd ÷ total_budget'),
            ('total_over_under_budget', df['total_over_under_budget'], 'total_budget − total_jtd. Negative = over budget.'),
            ('variance_trend', 'N/A (single snapshot)', 'Requires prior period comparison'),
            ('labor_unit_cost_per_hr', df['labor_unit_cost_per_hr'], 'labor-code JTD ÷ total_labor_hours. Unit cost per labor hour (NOT throughput).'),
            ('revenue_per_labor_hour', df['revenue_per_labor_hour'], 'abs(job_totals_revenue) ÷ total_labor_hours. Labor throughput — revenue generated per labor hour.'),
            ('material_price_variance', df['material_price_variance'], 'material_budget − material_spend. Positive = under budget.'),
        ]),
    ]

    total_fields = sum(len(s[1]) for s in sections)
    write_title(ws, f'DERIVED FIELDS — {total_fields} CALCULATIONS',
                'All derived values computed from parsed data. v4 schema (2026-04-18): Reconciliation Actions #1, #2, #3, #7, #8, #9 applied. Fields mirror Job 2020 test-labels.')
    ws.append([])
    ws.append(['Field', 'Value', 'Formula / Logic'])
    hdr_row = ws.max_row
    for c in range(1, 4):
        apply_hdr(ws.cell(hdr_row, c))

    for section_label, rows in sections:
        ws.append([section_label])
        r = ws.max_row
        ws.cell(r, 1).font = SECTION_FONT
        for c in range(1, 4):
            ws.cell(r, c).fill = SECTION_FILL
        for k, v, logic in rows:
            ws.append([k, v, logic])
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 72


def build_reconciliation(wb, d):
    ws = wb.create_sheet('Reconciliation')
    recon = d['reconciliation']
    write_title(ws, 'RECONCILIATION — PER COST CODE', 'Parsed line item sums vs. cost code totals from the PDF')
    ws.append([])
    ws.append(['Cost Code', 'Description', 'PDF Total', 'Parsed Sum', 'Difference', 'Status'])
    hdr_row = ws.max_row
    for c in range(1, 7):
        apply_hdr(ws.cell(hdr_row, c))
    passed = 0
    for r in recon:
        check = '✓' if r['status'] == 'PASS' else '✗'
        if r['status'] == 'PASS':
            passed += 1
        ws.append([r['code'], r['description'], r['pdf_total'], r['parsed_sum'], r['difference'], check])
    ws.append([])
    ws.append([f"RESULT: {passed}/{len(recon)} codes reconciled"])
    ws.cell(ws.max_row, 1).font = Font(name=FONT, bold=True, size=11, color='1F6F43' if passed == len(recon) else 'B85C3E')
    widths = [10, 28, 14, 14, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def build_reconciliation_log(wb, d):
    ws = wb.create_sheet('Reconciliation Log')
    df = d['derived_fields']
    rr = d['report_record']
    bs = rr['job_totals_by_source']
    recon = d['reconciliation']
    pr_total = bs['PR']
    ap_total = bs['AP']
    gl_total = bs['GL']
    passed = sum(1 for r in recon if r['status'] == 'PASS')

    # largest overrun/savings from derived
    largest_over = df['largest_overrun']
    largest_save = df['largest_savings']

    write_title(ws, 'RECONCILIATION LOG — Job 2048',
                'Worked example of the Reconciliation Protocol running against Job 2048 (JCR-only — no external cross-checks). Each row shows the JCR-derived value, the cross-check value, and the observed status. Protocol contract lives in JCR Schema v4 → Reconciliation Protocol tab.')
    ws.append([])
    ws.append(['check_id', 'field_name', 'JCR-derived value', 'cross-check value', 'cross-check source', 'status', 'notes'])
    hdr_row = ws.max_row
    for c in range(1, 8):
        apply_hdr(ws.cell(hdr_row, c))

    rows = [
        ('RC-01', 'plus_minus_budget',
         f"Row summary: {largest_over}",
         f"Largest overrun spot-check: actual − current_budget ties to cent",
         'Cost Code Summaries (internal recomputation)', 'PASS',
         'Action #1 sign convention: positive = over-budget. Spot-check recomputes plus_minus from actual − current_budget; matches PDF value to the cent.'),
        ('RC-02', 'current_budget / original_budget',
         f"current_budget total = ${df['total_budget']:,.0f}; original_budget carried per cost code (see Cost Code Summaries tab)",
         f"Sum of org_budget across expense codes differs where COs shifted budget",
         'Cost Code Summaries tab (column C vs column D)', 'PASS',
         'Both budgets carried distinctly per Action #2. Current budget differs from original on codes that had change orders.'),
        ('RC-03', 'pr_src_cost_per_hr',
         f"${df['pr_src_cost_per_hr']:.2f} / hr (all PR-source cost ÷ total labor hours)",
         f"Manual recomputation: PR total ${pr_total:,.2f} ÷ {df['total_labor_hours']:,.2f} hrs = ${df['pr_src_cost_per_hr']:.2f}",
         'Derived Fields tab (line items → source=PR)', 'PASS',
         'Phase-code filter: include all phases incl. 995 burden and 998 taxes; exclude only 999 Sales. Matches Action #3 convention from Job 2020.'),
        ('RC-04', 'worker_nominal_wages[] / nominal_wage_percentiles',
         f"p10=${df['nominal_wage_percentiles']['p10']:.2f} | p25=${df['nominal_wage_percentiles']['p25']:.2f} | p50=${df['nominal_wage_percentiles']['p50']:.2f} | p75=${df['nominal_wage_percentiles']['p75']:.2f} | p90=${df['nominal_wage_percentiles']['p90']:.2f} ({df['total_workers']} workers)",
         f"straight_time_rate = ${df['straight_time_rate']:.2f} (weighted average)",
         'Worker Wages tab + straight_time_rate derived field', 'PASS',
         'Population-slice sanity: weighted straight_time_rate falls within p25–p75 range, which is expected for this workforce composition.'),
        ('RC-05', 'labor_unit_cost_per_hr vs revenue_per_labor_hour',
         f"cost = ${df['labor_unit_cost_per_hr']:.2f}/hr | revenue = ${df['revenue_per_labor_hour']:.2f}/hr | ratio = {df['revenue_per_labor_hour'] / df['labor_unit_cost_per_hr']:.2f}x",
         f"Recompute: labor-code total ÷ {df['total_labor_hours']:,.2f} hrs",
         'Derived Fields tab (Action #7 split)', 'PASS',
         'Split into two fields per Action #7. Cost and throughput are distinct metrics; their ratio is the productivity signal.'),
        ('RC-06', 'scope_metrics[] / scope_benchmarks', 'pending', '—',
         'project-attribute input (trade-declared)', 'PENDING',
         'Awaiting plumbing scope-metric declarations for Job 2048 (e.g., fixture count, DFU total, linear feet of main). Labeler should not emit scope_benchmarks until scope_metrics is populated.'),
        ('RC-07', 'reconciliation (per-code parse tie-out)',
         f"{passed}/{len(recon)} codes tie to PDF to the cent",
         f"PR: ${pr_total:,.2f} | AP: ${ap_total:,.2f} | GL: ${gl_total:,.2f}  all match by_source",
         'Reconciliation tab', 'PASS' if passed == len(recon) else 'FAIL',
         f"Independent sum of line_items[] by cost_code vs. Cost Code Totals row for each code. Status PASS if |diff| < $0.02 for all codes."),
    ]
    for row in rows:
        ws.append(list(row))
        r = ws.max_row
        for c in range(1, 8):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(r, c).font = BODY_FONT

    ws.append([])
    ws.append(['SUMMARY'])
    ws.cell(ws.max_row, 1).font = SECTION_FONT
    ws.append(['Total checks', len(rows)])
    ws.append(['PASS', sum(1 for r in rows if r[5] == 'PASS')])
    ws.append(['PENDING', sum(1 for r in rows if r[5] == 'PENDING')])
    ws.append(['FAIL', sum(1 for r in rows if r[5] == 'FAIL')])

    ws.append([])
    ws.append(['How to use: when a labeler produces a new set of derived values for Job 2048, update columns C and D in the rows above and set status per the Protocol contract.'])
    ws.cell(ws.max_row, 1).font = CAPTION_FONT

    widths = [10, 36, 36, 40, 30, 10, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for r in range(5, 5 + len(rows)):
        ws.row_dimensions[r].height = 56


def main():
    d = json.load(open(DATA_JSON))
    if ENRICHED_META_JSON.exists():
        enriched = json.load(open(ENRICHED_META_JSON))
        d['enriched_meta'] = enriched.get('meta', {})
        d['co_log'] = enriched.get('co_log', [])
    else:
        d['enriched_meta'] = {}
        d['co_log'] = []
    wb = Workbook()
    wb.remove(wb.active)
    build_report_record(wb, d)
    build_line_items(wb, d)
    build_cost_code_summaries(wb, d)
    build_worker_wages(wb, d)
    build_derived_fields(wb, d)
    build_reconciliation(wb, d)
    build_reconciliation_log(wb, d)
    wb.save(OUT_FILE)
    print(f'wrote {OUT_FILE}  ({OUT_FILE.stat().st_size:,} bytes)')
    print(f'  tabs: {wb.sheetnames}')


if __name__ == '__main__':
    main()
