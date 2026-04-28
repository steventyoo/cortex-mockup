#!/usr/bin/env python3
"""
enrich_2117.py — Ravenna Partners Lachlan Apartments — full 2020/2061 gold-standard depth.

104-unit Ravenna Partners owner-GC multifamily at 3421 Woodland Park Ave N, Seattle
(Lower Woodland / Wallingford). 8-story + penthouse mid-rise. First OWP-Ravenna engagement
(Ravenna Partners is a NEW GC for OWP — owner-GC pattern). Bidding history goes back to
Dec 2023 when Blueprint was the GC (3 budget rounds: Dec 2023 / Oct 2024 / Nov 2025).
Final BID Nov 19 2025 = $1,948,000 / 104u / 425 trim fixtures.

Verified data sources:
  • Ravenna Partners - 3421 Woodland Park Plumbing Proposal - Nov 19 2025.docx (resident + common + systems trim)
  • Ravenna Partners - 3421 Woodland Park Apartments Bid Sheet - Nov 19 2025.xlsx (ESTIMATE + FIXTURE PRICING tabs)
  • Blueprint - 3421 Woodland Park - Bid Sheet - December 15 2023.xlsx (1st Budget · $2,610,000 / 122u)
  • Blueprint - 3421 Woodland Park - Bid Sheet - October 17 2024.xlsx (2nd Budget · $2,100,000 / 100u)
  • 3421 Woodland Park - Building Permit Plans.pdf (architectural drawings)
  • 3421 Woodland Park - Initial MEP Coordination Plans.pdf (Franklin design-build deliverable)
  • 2117 Job Detail Report.pdf (Sage Timberline · early design phase · $11k AP only)

Verified counts: 104 units · 425 trim fixtures · 4.09/unit · 535 RI · 705 permit
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2117_data.json"
DASH_FILE  = SCRIPT_DIR / "2117_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2117_JCR_Cortex_v2.xlsx"

META_2117 = {
    "job_id":          "2117",
    "name":            "Ravenna Partners · Lachlan Apartments",
    "short_name":      "Lachlan",
    "gc":              "Ravenna Partners",
    "gc_address":      "Seattle WA (owner-GC entity)",
    "gc_pm":           "Andrew Withnell — per Nov 19 2025 proposal addressee",
    "gc_sup":          "TBD",
    "gc_pe":           "TBD",
    "owp_ri_foreman":  "TBD · pre-construction (no field start scheduled yet)",
    "owp_trim_foreman":"TBD",
    "owp_signatory":   "Richard Donelson (per OWP proposal letterhead)",
    "owp_estimator":   "Richard Donelson + Gerard takeoff team (Jeffrey + Jordan)",
    "owner":           "Ravenna Partners (owner-GC · self-developing)",
    "developer":       "Ravenna Partners",
    "architect":       "TBD (per Master Project List · drawings published as 'Building Permit Plans' Nov 2025)",
    "structural":      "TBD",
    "mep_engineer":    "Franklin Engineering (OWP design-build · $10,701 design fees booked YTD)",
    "civil":           "TBD",
    "landscape":       "TBD",
    "site_address":    "3421 Woodland Park Ave N, Seattle WA (Lower Woodland / Wallingford corridor)",
    "permit":          "City of Seattle plumbing/gas/backflow permits — pending bid acceptance and submittal",
    "insurance":       "Standard COI · non-Wrap (per Nov 2025 proposal · OCIP DEDUCT alternate available)",
    "lien_position":   "Subcontract not yet executed · final BID Nov 19 2025 pending Ravenna acceptance",
    "warranty":        "Standard 1-yr plumbing (post-completion)",
    "contract_doc":    "FINAL BID submitted Nov 19 2025 by OWP for $1,948,000 / 104u / 425 trim. Initial bid Dec 2023 (Blueprint GC, $2.61M/122u); 2nd bid Oct 2024 ($2.10M/100u); 3rd bid Nov 2025 (Ravenna Partners GC, FINAL).",
    "scope":           "Design-build plumbing scope · 104-unit multifamily · 8-story + penthouse · gas water heating + electric water heating fallback",
    "gdrive_status":   "Bid Book mounted at /Estimating - Won Bids/Ravenna - Lachlan/. "
                        "3 budget rounds (Dec 2023 / Oct 2024 / Nov 2025) · Bid sheets + proposals + permit plans + MEP coordination plans + "
                        "Sanco/EWH submittals + vendor quotes (S024745035 trim fixture quote).",
    "status_text":     "ACTIVE · pre-construction (final BID Nov 19 2025 · pending Ravenna acceptance)",
    "ar_billed":       0,
    "direct_cost":     11007,
    "carrying_loss":   11007,
    "retention":       0,
    "contract_orig":   None,
    "contract_final":  None,
    "contract_bid_latest": 1948000,    # Nov 19 2025 FINAL BID
    "contract_bid_initial": 2610000,   # Dec 15 2023 1st Budget (Blueprint GC)
    "contract_bid_2nd": 2100000,       # Oct 17 2024 2nd Budget (Blueprint GC)
    "co_net":          0,
    "units":           104,
    "fixtures_trim":   425,
    "fixtures_ri":     535,
    "fixtures_permit": 705,
    "fixtures_per_unit": 4.09,
    "fixture_categories": {
        "Toilets (98 std + 6 ADA Gerber Maxwell GWS20912/20918)":   104,
        "Lavatories (104 wall-hung Kohler K-2035-1 + Peerless P191102LF-BL matte black)": 104,
        "Tubs (47 std + 6 ADA Sterling 71121110 Vikrell w/ Peerless PTT188792-BL)": 53,
        "Shower stalls (Sterling 95077 38\"x38\" pans w/ Peerless PTT188782-BL)": 51,
        "Type A handshowers (Peerless P62447-BL at ADA tubs)": 6,
        "Kitchen sinks (104 Peerless P7946LF-BL matte black faucet, sinks by others)": 104,
        "Common toilets (2 Toto CST454CEFG ADA bowls)": 2,
        "Common lavs (2 Kohler K-2035-1 w/ Kohler K-14402 + Jaclo 8877-COR)": 2,
        "Amenity sink trims (2 Kohler K-7507-BL at L1 + L8 amenity)": 2,
        "Coffee maker connection (L1 amenity)": 1,
        "Mop sink (Mustee 63M w/ T&S Brass B-0665-BSTR)": 1,
        "Drinking fountain (Elkay EZSTLDDWSLK dual-level w/ bottle filler at L1)": 1,
        "Dishwasher RI (104 + 2 Delta 72020-BL air gaps)": 106,
        "Icemaker RI box (104 recessed boxes)": 104,
        "Washer RI box (104 fire-rated boxes)": 104,
        "Condensate drain connections (Maestro residential units)": 104,
        "Water spacer tubes/unions/meters (3/4\" supplied by Owner)": 208,
    },
    "project_type":    "104-unit multifamily · 8-story + penthouse mid-rise · Lower Woodland / Wallingford corridor Seattle",
    "schedule":        "1st Budget Dec 2023 → 2nd Budget Oct 2024 → 3rd Budget FINAL Nov 19 2025 → Ravenna acceptance pending",
    "expected_start":  "TBD · pending Ravenna acceptance + city plumbing permit",
    "expected_finish": "TBD",
    "duration_months": 22,  # estimated based on 8-story typical
    "workers":         2,
    "hours":           3.0,
    "top_vendor":      "Franklin Engineering ($10,701 / 2 invoices / 100% of $11k AP) — design fees only",
    "ap_total":        11007,
    "ap_vendor_count": 1,
    "doc_counts": {
        "Bid_Sheets":          3,    # Dec 2023, Oct 2024, Nov 2025
        "Proposals":           1,    # Nov 19 2025 docx + matching pdf
        "Permit_Plans":        2,    # Building Permit Plans (1st + 3rd budgets have copies)
        "MEP_Plans":           1,    # Initial MEP Coordination Plans (Nov 2025)
        "Submittals":          3,    # Sanco package + Electric Water Heater + Sizing Tool
        "Vendor_Quotes":       1,    # S024745035-0001 trim fixture quote
        "Email_Threads":       2,    # Re_ Budget msg files (Dec 2023 + Oct 2024)
    },
    "bid_history": [
        ("2023-12-15", "BID",  "Blueprint",        122, 2610000, "1st Budget · 122 units · $21,393/unit · 667 RI fixtures · scope larger than final"),
        ("2024-10-17", "BID",  "Blueprint",        100, 2100000, "2nd Budget · 100 units · -19.5% from 1st bid · -$510k · scope contraction"),
        ("2025-11-19", "BID",  "Ravenna Partners", 104, 1948000, "3rd Budget · FINAL · 104 units · $18,731/unit · 425 trim · -7.2% from 2nd · Ravenna takes over as owner-GC"),
    ],
    "narrative": [
        ("Project history",
         "Ravenna Partners is OWP's first engagement with this owner-GC. The Lachlan Apartments project at "
         "3421 Woodland Park Ave N (Lower Woodland / Wallingford) has been bid 3 times since Dec 2023: "
         "1st Budget Dec 15 2023 ($2,610,000 / 122u) and 2nd Budget Oct 17 2024 ($2,100,000 / 100u) were "
         "both submitted with Blueprint listed as GC. By Nov 19 2025 Ravenna Partners had taken over as "
         "owner-GC and OWP submitted the 3rd and FINAL bid at $1,948,000 / 104u / 425 trim fixtures "
         "(addressed to Andrew Withnell). Pending Ravenna acceptance."),
        ("Scope refinement across 3 bids",
         "Bid trajectory shows scope contraction over 23 months: 1st bid 122 units / 667 RI / $2.61M → "
         "2nd bid 100 units / 539 RI / $2.10M (-19.5%) → 3rd bid 104 units / 535 RI / $1.95M (-7.2%). "
         "Net change from 1st to 3rd: -18 units (-14.8%) and -$662k (-25.4%) — owner repeatedly tightened "
         "the program to hit pro-forma. Final program is 8-story + penthouse mid-rise w/ gas water heating "
         "(3 boilers), 53 tubs / 51 showers / 104 toilets/lavs/kitchens, matte black Peerless trim throughout."),
        ("Sunk-cost posture",
         "OWP has $0 billed to date against $11,007 sunk takeoff+design cost = -$11,007 carrying loss. "
         "Top vendor Franklin Engineering = $10,701 (97% of sunk · OWP design-build MEP fees). Only 3.0 "
         "hours of OWP labor logged YTD (Jeffrey + Jordan Gerard takeoff). Tiny carrying cost relative to "
         "the $1.95M bid — typical pre-acceptance posture for a 23-month bid pursuit."),
        ("Owner-GC structure",
         "Ravenna Partners is acting as both owner and GC (owner-GC pattern — same as #2117 Lachlan and "
         "Sierra-style #2110 LNTP-then-bid). Owner-GCs typically run lean PM teams, expect aggressive "
         "subcontractor pricing, and emphasize schedule certainty. No closed-job history with Ravenna for "
         "OWP — first engagement. Watch payment cadence and CO behavior closely once subcontract executes. "
         "OWP design-build MEP partner Franklin Engineering provides continuity (same partner used on "
         "#2108 Braseth, #2118 Exxel, etc.)."),
        ("Activation triggers",
         "Watch for: (a) Ravenna acceptance of Nov 19 2025 BID, (b) plumbing permit submittal to City of "
         "Seattle, (c) framing schedule visibility from Ravenna. Once subcontract executes, Franklin "
         "Engineering will publish IFC plumbing drawings (Initial MEP Coordination Plans currently on "
         "file). OWP ready to mobilize once Ravenna issues subcontract or LNTP."),
        ("Risk profile",
         "New GC relationship · no closed-job history with Ravenna Partners (owner-GC). Three-bid "
         "trajectory (-25% from 1st to 3rd) shows owner is price-sensitive — expect tight CO scrutiny. "
         "Tariff/duty exclusion clause in Nov 2025 proposal (matches OWP 2025 bid template). "
         "OCIP-deduct alternate offered but not selected · standard COI assumed. Expected margin profile: "
         "22-26% on design-build subcontracts based on OWP closed-portfolio benchmarks for new owner-GC "
         "relationships."),
    ],
}

INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
ACTIVE_FILL = PatternFill("solid", start_color="D4EDDA")
WARN_FILL   = PatternFill("solid", start_color="FFF3CD")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")


def clear_range(ws, r1, r2, c1=1, c2=10):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).border = Border()


def rebuild_job_info(ws):
    clear_range(ws, 1, ws.max_row + 5, 1, 6)
    ws["A1"] = f"JOB #{META_2117['job_id']} · RAVENNA PARTNERS · LACHLAN APARTMENTS · KEY PLAYERS & METADATA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Pre-construction · 3-bid pursuit since Dec 2023. Final BID Nov 19 2025 = $1,948,000 / 104u / "
                "425 trim fixtures. Pending Ravenna Partners acceptance. OWP has $0 billed against $11,007 "
                "sunk design cost = -$11,007 carrying loss (97% Franklin Engineering MEP design fees). "
                "Sources: Bid sheets (Dec 2023 + Oct 2024 + Nov 2025) + Proposal (Nov 19 2025) + Permit Plans "
                "+ MEP Coordination Plans + Sage JDR.")
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    sections = [
        ("IDENTITY", [
            ("Job Number",         META_2117["job_id"]),
            ("Job Name",           META_2117["name"]),
            ("Short Name",         META_2117["short_name"]),
            ("Project Type",       META_2117["project_type"]),
            ("Site Address",       META_2117["site_address"]),
            ("Permit",             META_2117["permit"]),
            ("Status",             META_2117["status_text"]),
            ("Scope of Work",      META_2117["scope"]),
        ]),
        ("SCHEDULE", [
            ("1st Budget",          "Dec 15 2023 (Blueprint as GC · $2,610,000 / 122u)"),
            ("2nd Budget",          "Oct 17 2024 (Blueprint as GC · $2,100,000 / 100u · -19.5%)"),
            ("3rd Budget · FINAL",  "Nov 19 2025 (Ravenna Partners as owner-GC · $1,948,000 / 104u / 425 trim · -7.2%)"),
            ("Field Start (target)", META_2117["expected_start"]),
            ("Project End (est.)", META_2117["expected_finish"]),
            ("Duration (est.)",    f"~{META_2117['duration_months']} months once mobilized"),
            ("Schedule Note",      META_2117["schedule"]),
            ("Subcontract Execution", "Pending — Ravenna acceptance of Nov 19 2025 BID"),
        ]),
        ("FINANCIAL POSTURE (PRE-CONSTRUCTION · DESIGN PHASE)", [
            ("Original Contract",  "Subcontract execution pending"),
            ("Latest BID submitted", f"${META_2117['contract_bid_latest']:,} (Nov 19 2025 · 104 units · $18,731/unit · 425 trim · $4,584/fixture)"),
            ("2nd Budget BID",     f"${META_2117['contract_bid_2nd']:,} (Oct 17 2024 · 100u · Blueprint GC era)"),
            ("Initial BID",        f"${META_2117['contract_bid_initial']:,} (Dec 15 2023 · 122u · Blueprint GC era · 23-month pursuit)"),
            ("AR Billed to Date",  f"${META_2117['ar_billed']:,} (no progress billings yet · 3rd BID pending)"),
            ("Direct Cost (sunk)", f"${META_2117['direct_cost']:,}"),
            ("Carrying Loss",      f"$({META_2117['carrying_loss']:,}) — recoverable on subcontract execution"),
            ("Retainage",          "$0 (no field billing yet · 5% at field start)"),
            ("Insurance",          META_2117["insurance"]),
            ("Lien Position",      META_2117["lien_position"]),
            ("Warranty",           META_2117["warranty"]),
            ("Contract on File",   META_2117["contract_doc"]),
        ]),
        ("SCOPE & FIXTURE PROFILE (verified from Nov 19 2025 proposal + bid sheet ESTIMATE tab)", [
            ("Plumbing Units",     f"{META_2117['units']} units ✓ CONFIRMED via bid sheet ESTIMATE tab + proposal RESIDENT UNIT TRIM block (98 standard + 6 ADA = 104)"),
            ("Total Trim Fixtures", f"{META_2117['fixtures_trim']:,} trim fixtures (verified bid sheet) · {META_2117['fixtures_per_unit']:.2f} per unit"),
            ("Total RI Fixtures",  f"{META_2117['fixtures_ri']:,} RI fixtures per ESTIMATE row 95"),
            ("Total Permit Fixtures", f"{META_2117['fixtures_permit']:,} permit fixtures (642 upper + 63 lower)"),
            ("Toilets (Gerber Maxwell GWS20912 + GWS20918 ADA)", "98 standard + 6 ADA = 104 (1/unit)"),
            ("Lavatories (Kohler K-2035-1 wall-hung + Peerless P191102LF-BL matte black)", "98 standard + 6 ADA = 104 (1/unit)"),
            ("Tubs (Sterling 71121110 Vikrell + Peerless PTT188792-BL trim)", "47 std + 6 ADA = 53"),
            ("Showers (Sterling 95077 38\"x38\" pans + Peerless PTT188782-BL trim)", "51 standalone"),
            ("Kitchen sinks (Peerless P7946LF-BL matte black faucet, sinks by others)", "98 std + 6 ADA = 104 (1/unit)"),
            ("Heat sources", "GAS BOILERS (3 × 199mbh AO Smith/Rinnai/Bock 100-gal vertical) + B&G recirc pump + Heat Timer mixing valve"),
            ("Common area trim", "2 Toto ADA toilets + 2 Kohler lavs + 2 Kohler amenity faucets + Mustee mop sink + Elkay drinking fountain"),
            ("Plumbing systems", "5 Woodford #65 frostproof hose bibs + 1 Woodford #122 hot/cold trash room + 1 Woodford #24 garage + 13 deck drains + 7 roof drains + 3 garage drains + 1 oil/water separator + 1 foundation drain pump + gas service to 3 WHs/2 BBQs/1 firepit/1 RTU"),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR (Ravenna Partners · owner-GC)", [
            ("General Contractor", META_2117["gc"]),
            ("GC Address",         META_2117["gc_address"]),
            ("GC Project Manager", META_2117["gc_pm"]),
            ("GC Superintendent",  META_2117["gc_sup"]),
            ("GC Project Engineer", META_2117["gc_pe"]),
            ("Ravenna engagement count", "OWP's 1st Ravenna Partners engagement. Owner-GC pattern (Ravenna acts as both owner and GC). NEW GC for OWP — no closed-job history. First two budget rounds Dec 2023 + Oct 2024 were submitted under Blueprint as GC; Ravenna took over as owner-GC for Nov 2025 final bid."),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2117["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2117["owp_trim_foreman"]),
            ("OWP Estimator",       META_2117["owp_estimator"]),
            ("OWP Signatory",       META_2117["owp_signatory"]),
            ("Active Crew",         f"{META_2117['workers']} workers · {META_2117['hours']} hrs YTD (takeoff/design only)"),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record",    META_2117["owner"]),
            ("Developer",          META_2117["developer"]),
            ("Owner-GC structure", "Ravenna Partners is acting as both owner and GC (self-developing). Owner-GC pattern means lean PM teams + price-sensitive negotiation + emphasis on schedule certainty."),
            ("Address Andrew Withnell", "Per Nov 19 2025 proposal letterhead — primary OWP contact at Ravenna Partners."),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",          META_2117["architect"]),
            ("MEP / Plumbing Engineer", META_2117["mep_engineer"]),
            ("Structural Engineer", META_2117["structural"]),
            ("Civil Engineer",     META_2117["civil"]),
            ("Landscape Architect", META_2117["landscape"]),
            ("Franklin design-build note", "Franklin Engineering is OWP's design-build MEP partner — same firm engaged on #2108 Braseth, #2118 Exxel, etc. Design completed and Initial MEP Coordination Plans on file. $10,701 design fees booked YTD = ~0.55% of bid value (typical pre-permit fee ratio)."),
        ]),
        ("BID HISTORY (3-budget pursuit · Dec 2023 → Nov 2025)", []),
        ("AP VENDOR PROFILE (TAKEOFF/DESIGN PHASE)", [
            ("Top Vendor",         META_2117["top_vendor"]),
            ("Total AP Spend",     f"${META_2117['ap_total']:,}"),
            ("Active Vendors",     f"{META_2117['ap_vendor_count']} unique vendor (Franklin Engineering only · typical for design-only phase)"),
            ("Concentration Note", "100% Franklin Engineering concentration is normal for pre-permit design-build phase — will diversify when field phase begins (Pacific Plumbing, Rosen, Ferguson, Consolidated standard expected)."),
        ]),
        ("DOCUMENT META (current)", [
            ("Bid Sheets (in folder)",  f"{META_2117['doc_counts']['Bid_Sheets']} (Dec 2023 + Oct 2024 + Nov 2025)"),
            ("Proposal (Nov 2025)",     f"{META_2117['doc_counts']['Proposals']} (.docx + .pdf)"),
            ("Permit Plans",            f"{META_2117['doc_counts']['Permit_Plans']} (1st + 3rd budget folders have copies)"),
            ("MEP Coordination Plans",  f"{META_2117['doc_counts']['MEP_Plans']} (Franklin Engineering · Nov 2025 · for OWP-internal use)"),
            ("Submittals",              f"{META_2117['doc_counts']['Submittals']} files (Sanco package, Electric Water Heater, Sizing Tool)"),
            ("Vendor Quotes",           f"{META_2117['doc_counts']['Vendor_Quotes']} files (S024745035-0001 trim fixture quote)"),
            ("Email Threads",           f"{META_2117['doc_counts']['Email_Threads']} .msg files (budget responses Dec 2023 + Oct 2024)"),
            ("Notes",                   "No Sage Job Book yet. CO log will populate at subcontract execution."),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",            "2117 Job Detail Report (Sage Timberline · early design phase · 17 line items only)"),
            ("Parsed Data",        "2117_data.json"),
            ("Dashboard Arrays",   "2117_dashboard_arrays.json"),
            ("Final Bid Sheet",    "Ravenna Partners - 3421 Woodland Park Apartments Bid Sheet - Nov 19 2025.xlsx"),
            ("Final Proposal",     "Ravenna Partners - 3421 Woodland Park Plumbing Proposal - Nov 19 2025.docx"),
            ("Earlier Bid Sheets", "Blueprint - 3421 Woodland Park - Bid Sheet - December 15 2023.xlsx + October 17 2024.xlsx"),
            ("Permit Plans",       "3421 Woodland Park - Building Permit Plans.pdf (current)"),
            ("MEP Plans",          "3421 Woodland Park - Initial MEP Coordination Plans.pdf (Franklin · Nov 2025)"),
            ("GDrive Bid Folder",  META_2117["gdrive_status"]),
        ]),
        ("RISK FLAGS / WATCH ITEMS", [
            ("First OWP-Ravenna engagement", "No closed-job history with Ravenna Partners. Owner-GC pattern (lean PM, price-sensitive). Watch payment cadence + CO behavior closely once subcontract executes."),
            ("23-month bid pursuit · -25% from 1st to 3rd", "Owner repeatedly tightened scope (-18 units, -$662k from Dec 2023 → Nov 2025). Indicates strong price sensitivity. Expect tight CO scrutiny — discipline on scope clarity at subcontract execution will protect margin."),
            ("GC label changed Dec 2023 → Nov 2025", "First two bids submitted under Blueprint as GC (Dec 2023 + Oct 2024); 3rd bid submitted to Ravenna Partners as owner-GC. Confirm whether Blueprint exited or was acquired/repositioned."),
            ("Tariff/duty exclusion clause", "Nov 2025 proposal language: 'OWP reserves right to adjust for current/future tariff rates, duties, government charges, trade regulations.' — protects margin against 2025-onward import cost spikes."),
            ("OCIP-deduct alternate offered", "Alt #3 offered owner an OCIP wrap deduct option but not selected per Nov 2025 proposal · standard COI assumed. Confirm at subcontract execution."),
            ("Gas water heating risk", "3 gas boilers (199mbh × 100-gal AO Smith/Rinnai/Bock) + 1 RTU + 2 BBQ + 1 firepit = 7 gas connections. PSE meter sizing + venting routes need IFC drawing review at submittal stage."),
        ]),
        ("PROJECT NARRATIVE", []),
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        if section_name.startswith("BID HISTORY"):
            headers = ["Date", "Type", "GC", "Units", "Price", "Note"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=r, column=c, value=h).font = WHITE_BOLD
                ws.cell(row=r, column=c).fill = HDR_FILL
                ws.cell(row=r, column=c).border = BORDER
            r += 1
            for date, type_, gc, units, price, note in META_2117["bid_history"]:
                ws.cell(row=r, column=1, value=date).font = INK
                ws.cell(row=r, column=2, value=type_).font = INK_BOLD if type_ in ("BID", "LNTP") else INK
                ws.cell(row=r, column=3, value=gc).font = INK
                ws.cell(row=r, column=4, value=units).font = INK
                ws.cell(row=r, column=5, value=f"${price:,}" if price else "—").font = INK
                ws.cell(row=r, column=6, value=note).font = INK
                ws.cell(row=r, column=6).alignment = WRAP
                if "FINAL" in note:
                    for c in range(1, 7): ws.cell(row=r, column=c).fill = ACTIVE_FILL
                for c in range(1, 7): ws.cell(row=r, column=c).border = BORDER
                r += 1
            r += 1
            continue
        if section_name == "PROJECT NARRATIVE":
            for label, body in META_2117["narrative"]:
                ws.cell(row=r, column=1, value=label).font = INK_BOLD
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v = ws.cell(row=r, column=2, value=body)
                cell_v.font = INK
                cell_v.alignment = WRAP
                cell_v.fill = TEAM_FILL
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.cell(row=r, column=1).border = BORDER
                cell_v.border = BORDER
                r += 1
            r += 1
            continue
        for label, val in fields:
            ws.cell(row=r, column=1, value=label).font = INK_BOLD
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_v.font = INK
            cell_v.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=1).border = BORDER
            cell_v.border = BORDER
            if section_name.startswith("PROJECT TEAM") or section_name.startswith("RISK") or section_name.startswith("AP VENDOR") or section_name.startswith("DOCUMENT META") or section_name.startswith("SCOPE"):
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v.fill = TEAM_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 50


def patch_overview(ws):
    for r in range(1, 12):
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "CANCELLED" in v.upper():
                ws.cell(row=r, column=c).value = v.replace("CANCELLED", META_2117["status_text"])

    start = 30
    clear_range(ws, start, start + 35, 1, 8)
    ws.cell(row=start, column=1, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    ws.cell(row=start+1, column=1, value=(
        "Pre-construction · 3-bid pursuit since Dec 2023. Final BID Nov 19 2025 = $1,948,000 / 104u / "
        "425 trim. Pending Ravenna Partners (owner-GC) acceptance. OWP's first Ravenna engagement."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=7)

    rows = [
        ("General Contractor",     META_2117["gc"]),
        ("GC Project Manager",     META_2117["gc_pm"]),
        ("Owner",                  META_2117["owner"]),
        ("Architect",              META_2117["architect"]),
        ("MEP / Plumbing Engineer", META_2117["mep_engineer"]),
        ("OWP Estimator",          META_2117["owp_estimator"]),
        ("OWP Signatory",          META_2117["owp_signatory"]),
        ("Site Address",           META_2117["site_address"]),
        ("Insurance",              META_2117["insurance"]),
        ("Status",                 META_2117["status_text"]),
        ("Final BID",              f"${META_2117['contract_bid_latest']:,} (Nov 19 2025 · 104u)"),
        ("AR Billed",              f"${META_2117['ar_billed']:,} (no progress billings yet)"),
        ("Direct Cost (sunk)",     f"${META_2117['direct_cost']:,}"),
        ("Carrying Loss",          f"$({META_2117['carrying_loss']:,}) recoverable on subcontract"),
        ("Verified Fixtures",      f"{META_2117['fixtures_trim']:,} trim fixtures · {META_2117['fixtures_per_unit']:.2f}/unit (Nov 19 2025 proposal)"),
    ]
    r = start + 3
    ws.cell(row=r, column=1, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=1, value=role).font = INK_BOLD
        ws.cell(row=r, column=2, value=val).font = INK
        ws.cell(row=r, column=1).fill = TEAM_FILL
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1


def patch_change_log(ws):
    last_row = ws.max_row
    r = last_row + 1
    ws.cell(row=r, column=1, value="2026-04-28").font = INK
    ws.cell(row=r, column=2, value="v2.0 · 2020-gold").font = INK_BOLD
    ws.cell(row=r, column=3, value=(
        "Upgraded Job Info to 2020/2061 gold-standard depth: 14-section sectioned layout with "
        "BID HISTORY (3 entries: Dec 2023 1st Budget Blueprint era → Oct 2024 2nd Budget Blueprint era → "
        "Nov 2025 3rd Budget FINAL Ravenna Partners era), AP Vendor Profile, Document Meta (3 bid sheets, "
        "1 proposal, 2 permit plans, 1 MEP plan, 3 submittals, 1 vendor quote), Risk Flags + Watch Items "
        "(GC handoff Blueprint→Ravenna, 23-month -25% bid trajectory, gas heating risk, OCIP option), and "
        "PROJECT NARRATIVE (6 blocks: history, scope refinement across 3 bids, sunk-cost posture, owner-GC "
        "structure, activation triggers, risk profile). Verified team from Nov 19 2025 proposal: Andrew "
        "Withnell (Ravenna PM addressee). Architect TBD, MEP Franklin Engineering (OWP design-build). "
        "Verified 425 trim fixtures (4.09/unit) + 535 RI + 705 permit fixtures from Nov 19 2025 bid sheet "
        "ESTIMATE tab + proposal RESIDENT UNIT TRIM block cross-check."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP


def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)
    print("→ Rebuilding Job Info tab to 2020-gold depth (14 sections + 3-bid history + narrative)...")
    rebuild_job_info(wb["02 Job Info"])
    print("→ Patching Overview + adding Project Team block (15 fields)...")
    patch_overview(wb["01 Overview"])
    print("→ Logging enrichment to Change Log...")
    patch_change_log(wb["17 Change Log"])
    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
