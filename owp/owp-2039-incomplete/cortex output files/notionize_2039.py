#!/usr/bin/env python3
"""Generic Notion-style summary workbook. Usage: python3 notionize_generic.py <PID>"""
import sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PID = sys.argv[1]
BASE = Path(f'/sessions/gracious-relaxed-pascal/mnt/cortex-mockup/owp/owp-{PID}/cortex output files')
SRC = BASE / f'OWP_{PID}_JCR_Cortex_v2.xlsx'
OUT = BASE / f'OWP_{PID}_JCR_Summary_Notion_v2.xlsx'

TEAM_MAP = {
    '2035': [('GC','Natural & Built LLC'),('GC PM','Tyson Cornett'),('GC Superintendent','Tyson Cornett'),('GC PE','N/A'),('Developer','Natural & Built'),('Owner','162 Ten Apartments LLC'),('Architect','N/A'),('Structural','N/A'),('Acoustical','N/A'),('ADA Consultant','N/A'),('Interior Design','N/A'),('MEP Engineer','N/A'),('Insurance','Not Wrap'),('OWP RI Foreman','Reuben')],
    '2036': [('GC','Exxel Pacific, Inc.'),('GC PM','David Strid'),('GC Superintendent','Shon Geer'),('GC PE','Jose Tapia'),('Developer','Exxel Pacific'),('Owner','Exxel Pacific Westridge LLC'),('Architect','N/A'),('Structural','N/A'),('Acoustical','N/A'),('ADA Consultant','N/A'),('Interior Design','N/A'),('MEP Engineer','N/A'),('Insurance','Not Wrap'),('OWP RI Foreman','Gustavo')],
    '2037': [('GC','Marpac Construction'),('GC PM','Evan Chan'),('GC Superintendent','Doyle Gustafson'),('GC PE','N/A'),('Developer','MYSA'),('Owner','MYSA University Apartments LP'),('Architect','N/A'),('Structural','N/A'),('Acoustical','N/A'),('ADA Consultant','N/A'),('Interior Design','N/A'),('MEP Engineer','N/A'),('Insurance','Not Wrap'),('OWP RI Foreman','Rick / Joe')],
    '2038': [('GC','Compass Harbor Construction, LLC'),('GC PM','Justin Anderson'),('GC Superintendent','Kurt Weagant'),('GC PE','Karl Clocksin'),('Developer','Continental'),('Owner','2nd & John Apartments'),('Architect','N/A'),('Structural','N/A'),('Acoustical','N/A'),('ADA Consultant','N/A'),('Interior Design','N/A'),('MEP Engineer','N/A'),('Insurance','Wrap (OCIP)'),('OWP RI Foreman','Thaddeus / Joe')],
    '2039': [('GC','Shelter Holdings'),('GC PM','Renay Luzama'),('GC Superintendent','Bill Robinson'),('GC PE','N/A'),('Developer','Shelter Holdings'),('Owner','Ravello HOA'),('Architect','N/A'),('Structural','N/A'),('Acoustical','N/A'),('ADA Consultant','N/A'),('Interior Design','N/A'),('MEP Engineer','N/A'),('Insurance','Not Wrap'),('OWP RI Foreman','OWP RI')],
    '2040': [('GC','Blueprint Capital'),('GC PM','Andrew Withnell'),('GC Superintendent','Mike Sanderson'),('GC PE','Kyle Stenson'),('Developer','Blueprint'),('Owner','Brooklyn 65 LLC'),('Architect','N/A'),('Structural','N/A'),('Acoustical','N/A'),('ADA Consultant','N/A'),('Interior Design','N/A'),('MEP Engineer','N/A'),('Insurance','Not Wrap'),('OWP RI Foreman','Garrett / Joe')],
}
TEAM = TEAM_MAP[PID]

FONT='Calibri'; INK='191917'; INK_3='6B6B65'; INK_4='9E9E97'
BG_PG='FBFAF7'; BG_MUT='F3F0E8'; BG_DIV='E8E3D6'
PILL = {'ON':('DCFCE7','166534'),'UNDER':('DBEAFE','1E40AF'),'OVER':('FEF3C7','92400E'),
        'CRITICAL':('FECACA','991B1B'),'TIES':('DCFCE7','166534'),'OFF':('FECACA','991B1B'),
        'Mapped':('DCFCE7','166534'),'Derived':('DBEAFE','1E40AF'),'Computed':('E9D5FF','6B21A8'),
        'Extracted':('DCFCE7','166534'),'Target':('FEF3C7','92400E'),'HEALTHY':('DCFCE7','166534'),
        'WATCH':('FEF3C7','92400E'),'ALERT':('FECACA','991B1B'),'ELEVATED':('FEF3C7','92400E'),
        'Verified':('DCFCE7','166534'),'Inferred':('DBEAFE','1E40AF'),'INFO':('E8E3D6','6B6B65'),
        'High':('DCFCE7','166534'),'Medium':('FEF3C7','92400E'),'Low':('FECACA','991B1B')}

THIN = Side(style='thin', color=BG_DIV)
BORDER_BOTTOM = Border(bottom=THIN)
def f_title():return Font(name=FONT,size=22,bold=True,color=INK)
def f_subtitle():return Font(name=FONT,size=11,italic=True,color=INK_3)
def f_eyebrow():return Font(name=FONT,size=9,bold=True,color=INK_4)
def f_header():return Font(name=FONT,size=9,bold=True,color=INK_4)
def f_body():return Font(name=FONT,size=10,color=INK)
def f_mid():return Font(name=FONT,size=10,color=INK_3)
def f_pill(c):return Font(name=FONT,size=9,bold=True,color=c)
def fill(c):return PatternFill('solid', fgColor=c)

wb = load_workbook(SRC)

def style_tab(ws):
    ws.sheet_view.showGridLines = False
    max_c = max(ws.max_column, 10)
    t = ws.cell(row=2, column=2)
    if t.value:
        t.font = f_title()
        t.alignment = Alignment(vertical='center', horizontal='left')
        ws.row_dimensions[2].height = 34
    s = ws.cell(row=3, column=2)
    if s.value:
        s.font = f_subtitle()
        s.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
        ws.row_dimensions[3].height = 22
    for c in range(2, max_c+1):
        ws.cell(row=4, column=c).fill = fill(BG_DIV)
    ws.row_dimensions[4].height = 2
    widths = {1:2, 2:28, 3:34, 4:22, 5:22, 6:22, 7:22, 8:22, 9:22, 10:22}
    for col, w in widths.items():
        if col <= max_c:
            ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, ws.max_row+1):
        for c in range(2, ws.max_column+1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if v is None: continue
            if c == 2 and isinstance(v, str) and v.isupper() and len(v) < 60 and ws.cell(row=r, column=3).value in (None, ''):
                cell.font = f_eyebrow()
                cell.alignment = Alignment(vertical='center', horizontal='left')
                continue
            if isinstance(v, str):
                sv = v.strip()
                if sv in PILL:
                    bg, fg = PILL[sv]
                    cell.fill = fill(bg); cell.font = f_pill(fg)
                    cell.alignment = Alignment(vertical='center', horizontal='center')
                    continue
                cell.font = f_mid() if c == 2 else f_body()
                cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
            elif isinstance(v, (int, float)):
                cell.font = f_body()
                cell.alignment = Alignment(vertical='center', horizontal='right')
                if not cell.number_format or cell.number_format == 'General':
                    if abs(v) >= 1000: cell.number_format = '"$"#,##0'
                    else: cell.number_format = '#,##0'
    hdr_r = 5
    for c in range(2, ws.max_column+1):
        cell = ws.cell(row=hdr_r, column=c)
        if cell.value and isinstance(cell.value, str):
            if ws.cell(row=hdr_r, column=3).value:
                cell.font = f_header(); cell.fill = fill(BG_MUT); cell.border = BORDER_BOTTOM
                cell.alignment = Alignment(vertical='center', horizontal='left')
    for r in range(6, ws.max_row+1):
        if (r-6) % 2 == 1:
            for c in range(2, ws.max_column+1):
                cell = ws.cell(row=r, column=c)
                if not cell.fill or cell.fill.fgColor is None or (cell.fill.fgColor.rgb in (None, '00000000')):
                    cell.fill = fill(BG_PG)
    ws.row_dimensions[1].height = 6

# Inject team on Overview
ov = wb['Overview']
start = ov.max_row + 3
ov.cell(row=start, column=2, value='PROJECT TEAM (OWP Job Info)')
start += 1
ov.cell(row=start, column=2, value='Role')
ov.cell(row=start, column=3, value='Status')
ov.cell(row=start, column=4, value='Name / Firm')
start += 1
for lbl, val in TEAM:
    ov.cell(row=start, column=2, value=lbl)
    pill = ov.cell(row=start, column=3, value=('Extracted' if val != 'N/A' else 'N/A'))
    if val == 'N/A':
        pill.fill = fill('EFEFEF'); pill.font = f_pill(INK_4)
        pill.alignment = Alignment(vertical='center', horizontal='center')
    ov.merge_cells(start_row=start, start_column=4, end_row=start, end_column=10)
    ov.cell(row=start, column=4, value=val)
    start += 1

for name in wb.sheetnames:
    style_tab(wb[name])

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f'Saved {OUT} ({OUT.stat().st_size:,} bytes)')
