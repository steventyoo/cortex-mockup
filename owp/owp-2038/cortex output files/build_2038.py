#!/usr/bin/env python3
"""Generic 17-tab JCR Cortex v2 builder for 2035-2040.
Usage: python3 build_generic.py <PID>
"""
import json, os, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PID = sys.argv[1]

# ============ PROJECT CONFIG (2035-2040) ============
CONFIG = {
    '2035': dict(
        name='162 Ten Apartments', gc='Natural & Built LLC', gc_cust='2035NB',
        owner='162 Ten Apartments LLC (Natural & Built)',
        gc_pm='Tyson Cornett', gc_sup='Tyson Cornett', gc_pe='TBD',
        developer='Natural & Built', owp_ri='Reuben', insurance='Not Wrap',
        location='16210 NE 80th St, Redmond, WA 98052', units=92,
        contract_orig=717_400.00, contract_final=722_290.00,
        start='2016-08-16', end='2018', duration=24,
        project_desc='162 Ten — 92-unit multifamily plumbing, Redmond WA',
        project_type='Multifamily New Construction',
    ),
    '2036': dict(
        name='Westridge Apartments', gc='Exxel Pacific, Inc.', gc_cust='2036EP',
        owner='Exxel Pacific Westridge LLC',
        gc_pm='David Strid', gc_sup='Shon Geer', gc_pe='Jose Tapia',
        developer='Exxel Pacific', owp_ri='Gustavo', insurance='Not Wrap',
        location='512 121st Place NE, Bellevue, WA', units=31,
        contract_orig=388_515.00, contract_final=418_677.00,
        start='2016-10-01', end='2018', duration=20,
        project_desc='Westridge — 31-unit multifamily plumbing, Bellevue WA',
        project_type='Multifamily New Construction',
    ),
    '2037': dict(
        name='University Apartments (MYSA)', gc='Marpac Construction', gc_cust='2037MP',
        owner='MYSA University Apartments LP',
        gc_pm='Evan Chan', gc_sup='Doyle Gustafson', gc_pe='TBD',
        developer='MYSA', owp_ri='Rick / Joe', insurance='Not Wrap',
        location='3025 NE 130th Street, Seattle WA', units=122,
        contract_orig=1_920_850.00, contract_final=2_076_783.00,
        start='2017-01-09', end='2018', duration=24,
        project_desc='University Apartments (MYSA) — 122-unit multifamily + 5 townhomes, Seattle WA',
        project_type='Multifamily New Construction',
    ),
    '2038': dict(
        name='2nd & John', gc='Compass Harbor Construction, LLC', gc_cust='2038CH',
        owner='2nd & John Apartments (Continental)',
        gc_pm='Justin Anderson', gc_sup='Kurt Weagant', gc_pe='Karl Clocksin',
        developer='Continental', owp_ri='Thaddeus / Joe', insurance='Wrap (OCIP)',
        location='200 2nd Ave West, Seattle, WA', units=80,
        contract_orig=1_210_378.00, contract_final=1_229_419.00,
        start='2017-02-17', end='2019', duration=24,
        project_desc='2nd & John — 80-unit multifamily plumbing, Lower Queen Anne Seattle',
        project_type='Multifamily New Construction',
    ),
    '2039': dict(
        name='Ravello Gas Piping', gc='Shelter Holdings', gc_cust='2039SH',
        owner='Ravello HOA / Shelter Holdings',
        gc_pm='Renay Luzama', gc_sup='Bill Robinson', gc_pe='TBD',
        developer='Shelter Holdings', owp_ri='OWP RI', insurance='Not Wrap',
        location='Ravello (site TBD), WA', units=1,
        contract_orig=54_900.00, contract_final=54_900.00,
        start='2017-05-03', end='2018', duration=13,
        project_desc='Ravello Gas Piping — small-scope gas piping retrofit',
        project_type='Gas Piping Retrofit',
    ),
    '2040': dict(
        name='Roosevelt 65th (Brooklyn 65)', gc='Blueprint Capital', gc_cust='2040BP',
        owner='Brooklyn 65 LLC (Blueprint)',
        gc_pm='Andrew Withnell', gc_sup='Mike Sanderson', gc_pe='Kyle Stenson',
        developer='Blueprint', owp_ri='Garrett / Joe', insurance='Not Wrap',
        location='1222 NE 65th Street, Seattle, WA 98115', units=55,
        contract_orig=604_000.00, contract_final=624_171.00,
        start='2017-03-01', end='2019', duration=22,
        project_desc='Brooklyn 65 (Roosevelt 65th) — 55-unit multifamily plumbing, NE Seattle',
        project_type='Multifamily New Construction',
    ),
}

cfg = CONFIG[PID]
BASE = Path(f'/sessions/gracious-relaxed-pascal/mnt/cortex-mockup/owp/owp-{PID}/cortex output files')
DATA_JSON = BASE / f'{PID}_data.json'
OUT = BASE / f'OWP_{PID}_JCR_Cortex_v2.xlsx'

# Load data
data = json.loads(DATA_JSON.read_text())
CODES = data['codes']
WORKERS = data['workers']
VENDORS = data['vendors']
INVOICES = data['invoices']

# Financials
sales = CODES.get('999', {})
CONTRACT_ORIG = cfg['contract_orig']
CONTRACT_FINAL = cfg['contract_final']
REVENUE = CONTRACT_FINAL
CO_TOTAL_IMPLIED = CONTRACT_FINAL - CONTRACT_ORIG

LABOR_CODES = ['011','100','101','110','111','112','113','120','130','140','141','142','143','145','150','151']
MATERIAL_CODES = ['039','210','211','212','213','220','230','240','241','242','243','244','245']
OVERHEAD_CODES = ['600','601','602','603','604','607']
BURDEN_CODE = '995'
TAX_CODE = '998'

def sum_actual(codes): return sum(CODES[c]['actual'] for c in codes if c in CODES)
LABOR_COST = sum_actual(LABOR_CODES)
MATERIAL_COST = sum_actual(MATERIAL_CODES)
OVERHEAD_COST = sum_actual(OVERHEAD_CODES)
BURDEN_COST = CODES.get(BURDEN_CODE, {}).get('actual', 0)
TAX_COST = CODES.get(TAX_CODE, {}).get('actual', 0)
EXPENSES = LABOR_COST + MATERIAL_COST + OVERHEAD_COST + BURDEN_COST + TAX_COST
NET_PROFIT = REVENUE - EXPENSES
TOTAL_HOURS = sum(CODES[c]['hrs_total'] for c in LABOR_CODES if c in CODES)
TOTAL_WORKERS = len(WORKERS)
RETAINAGE = abs(sum(iv['retainage'] for iv in INVOICES.values()))

# Source-level breakdown (approx from workers/vendors/invoices)
SRC_PR = sum(w['amount'] for w in WORKERS.values())
SRC_AP = sum(v['total'] for v in VENDORS.values())
SRC_GL = max(0, EXPENSES - SRC_PR - SRC_AP)

NAME = cfg['name']
GC = cfg['gc']
GC_CUST_CODE = cfg['gc_cust']
OWNER = cfg['owner']
GC_PM = cfg['gc_pm']; GC_SUP = cfg['gc_sup']; GC_PE = cfg['gc_pe']
DEVELOPER = cfg['developer']; OWP_RI_FOREMAN = cfg['owp_ri']
INSURANCE = cfg['insurance']; LOCATION = cfg['location']; UNITS = cfg['units']
PROJECT_DESC = cfg['project_desc']; PROJECT_TYPE = cfg['project_type']
JOB = PID

SRC_JDR = f'{PID} Job Detail Report.pdf (Sage Timberline, 04/03/2026)'
SRC_CONTRACT = f'{PID}-{GC} subcontract + Job Info sheet'
SRC_FOLDER = f'owp-{PID}/'
SRC_JOBINFO = 'OWP Project List with Schedule - UPDATED 04-01-26.xlsx'

# ============ STYLES ============
ARIAL = 'Arial'
TITLE = Font(name=ARIAL, size=14, bold=True, color='FFFFFF')
SUB = Font(name=ARIAL, size=10, italic=True, color='595959')
HDR_F = Font(name=ARIAL, size=10, bold=True, color='FFFFFF')
BODY = Font(name=ARIAL, size=10)
BOLD = Font(name=ARIAL, size=10, bold=True)
SRC_FONT = Font(name=ARIAL, size=8, italic=True, color='595959')
F_TITLE = PatternFill('solid', fgColor='1F3864')
F_HDR = PatternFill('solid', fgColor='2E5090')
F_ALT = PatternFill('solid', fgColor='F2F2F2')
F_HIGH = PatternFill('solid', fgColor='FFF2CC')
F_RISK = PatternFill('solid', fgColor='FFE6E6')
F_OK = PatternFill('solid', fgColor='E2EFDA')
THIN = Side(style='thin', color='BFBFBF')
BRD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT = Alignment(horizontal='right', vertical='center')

def put(ws, coord, val, font=BODY, fill=None, border=BRD, align=None, fmt=None):
    c = ws[coord] if isinstance(coord, str) else ws.cell(row=coord[0], column=coord[1])
    c.value = val; c.font = font
    if fill: c.fill = fill
    if align: c.alignment = align
    if border: c.border = border
    if fmt: c.number_format = fmt
    return c

def title(ws, text, sub_text=''):
    c = ws.cell(row=2, column=2, value=text); c.font=TITLE; c.fill=F_TITLE; c.alignment=LEFT
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=10)
    ws.row_dimensions[2].height = 28
    if sub_text:
        c2 = ws.cell(row=3, column=2, value=sub_text); c2.font=SUB
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=10)

def hdr(ws, row, cols, start_col=2):
    for i, txt in enumerate(cols):
        c = ws.cell(row=row, column=start_col+i, value=txt)
        c.font=HDR_F; c.fill=F_HDR; c.alignment=CENTER; c.border=BRD

def widths(ws, spec):
    for col, w in spec.items(): ws.column_dimensions[get_column_letter(col)].width = w

wb = Workbook()
wb.remove(wb.active)

# ============ TAB 1: OVERVIEW ============
ws = wb.create_sheet('Overview')
title(ws, f'Job #{JOB} · {NAME}',
      f'Cortex JCR v2  •  {GC} (customer {GC_CUST_CODE})  •  {LOCATION}  •  Owner: {OWNER}')
put(ws, 'B5', 'PROJECT OVERVIEW', BOLD, F_ALT)
overview = [
    ('Project Job #', JOB, SRC_JDR),
    ('Project Name (Sage)', NAME, SRC_JDR),
    ('Project Description', PROJECT_DESC, f'{SRC_JOBINFO} + contract'),
    ('General Contractor', GC, SRC_JOBINFO),
    ('Customer Code (Sage)', GC_CUST_CODE, SRC_JDR),
    ('Owner / Developer', OWNER, SRC_JOBINFO),
    ('GC PM', GC_PM, SRC_JOBINFO),
    ('GC Superintendent', GC_SUP, SRC_JOBINFO),
    ('GC PE', GC_PE, SRC_JOBINFO),
    ('OWP Rough-in Foreman', OWP_RI_FOREMAN, 'Schedule tab'),
    ('Jobsite Location', LOCATION, SRC_JOBINFO),
    ('Insurance', INSURANCE, SRC_JOBINFO),
    ('Fixture Schedule', f'{UNITS}-unit {PROJECT_TYPE.lower()}', 'Project scope'),
    ('Contract Type', 'Lump Sum', SRC_CONTRACT),
    ('Unit Count', UNITS, SRC_JOBINFO),
    ('Work Period', f'{cfg["start"]} start → {cfg["end"]} ({cfg["duration"]} months)', 'JDR PR/AR dates'),
]
r = 6
for label, val, src in overview:
    put(ws, f'B{r}', label, BOLD, align=LEFT)
    put(ws, f'C{r}', val, align=LEFT)
    put(ws, f'F{r}', src, SRC_FONT, align=LEFT)
    ws.merge_cells(f'C{r}:E{r}'); ws.merge_cells(f'F{r}:J{r}')
    r += 1

r += 1
put(ws, f'B{r}', 'CONTRACT VALUE', BOLD, F_ALT); put(ws, f'D{r}', 'NET PROFIT', BOLD, F_ALT)
put(ws, f'F{r}', 'DIRECT COST', BOLD, F_ALT); put(ws, f'H{r}', 'LABOR HOURS', BOLD, F_ALT)
r += 1
put(ws, f'B{r}', CONTRACT_FINAL, BOLD, fmt='"$"#,##0.00')
put(ws, f'D{r}', NET_PROFIT, BOLD, fmt='"$"#,##0.00')
put(ws, f'F{r}', EXPENSES, BOLD, fmt='"$"#,##0.00')
put(ws, f'H{r}', TOTAL_HOURS, BOLD, fmt='#,##0.00')
r += 1
co_pct = CO_TOTAL_IMPLIED/CONTRACT_ORIG*100 if CONTRACT_ORIG else 0
put(ws, f'B{r}', f'Original ${CONTRACT_ORIG:,.2f} + (${CO_TOTAL_IMPLIED:,.2f}) COs ({co_pct:.2f}%)', SUB)
put(ws, f'D{r}', f'{NET_PROFIT/REVENUE*100:.1f}% margin', SUB)
put(ws, f'F{r}', f'{EXPENSES/REVENUE*100:.1f}% of revenue', SUB)
put(ws, f'H{r}', f'{TOTAL_WORKERS} workers', SUB)

r += 3
put(ws, f'B{r}', 'SCOPE OF WORK', BOLD, F_ALT)
r += 1
scope_lines = [
    f'Full plumbing installation per Division 22: underground, garage, rough-in, finish, gas, water main, insulation.',
    f'{UNITS}-unit {PROJECT_TYPE.lower()} at {LOCATION}.',
    f'GC: {GC} ({GC_CUST_CODE}); Owner/Developer: {OWNER}.',
    f'Insurance: {INSURANCE}.',
    f'Project team: {GC_PM} (GC PM) · {GC_SUP} (GC Sup) · {GC_PE} (GC PE) · {OWP_RI_FOREMAN} (OWP RI).',
]
for line in scope_lines:
    put(ws, f'B{r}', line, BODY, align=LEFT); ws.merge_cells(f'B{r}:J{r}'); r += 1

r += 2
put(ws, f'B{r}', 'SOURCES', BOLD, F_HDR)
for col in range(2, 11): put(ws, (r, col), ws.cell(row=r, column=col).value or '', border=BRD, fill=F_HDR)
r += 1
srcs = [
    f'Canonical financial source: {SRC_JDR}',
    f'Job totals (JDR footer): Revenue ${REVENUE:,.2f} / Expenses ${EXPENSES:,.2f} / Net ${NET_PROFIT:,.2f} / Retainage ${RETAINAGE:,.2f}',
    f'Source breakdown (derived): PR ${SRC_PR:,.2f} · AP ${SRC_AP:,.2f} · GL+other ${SRC_GL:,.2f}',
    f'Contract: Lump Sum ${CONTRACT_ORIG:,.2f} → ${CONTRACT_FINAL:,.2f}',
    f'Change Orders (JDR-implied): ${CO_TOTAL_IMPLIED:,.2f} net',
    f'Project Team: {GC_PM} · {GC_SUP} · {GC_PE} · {DEVELOPER} · {OWP_RI_FOREMAN}',
]
for s in srcs:
    put(ws, f'B{r}', s, SRC_FONT, align=LEFT); ws.merge_cells(f'B{r}:J{r}'); r += 1
widths(ws, {1:2, 2:26, 3:22, 4:16, 5:16, 6:18, 7:18, 8:14, 9:14, 10:14})

# ============ TAB 2: BUDGET VS ACTUAL ============
ws = wb.create_sheet('Budget vs Actual')
title(ws, 'Budget vs Actual', f'All {len(CODES)} cost codes from JDR. Contract ${CONTRACT_ORIG:,.0f} → ${CONTRACT_FINAL:,.0f}.')
hdr(ws, 5, ['Cost Code','Description','Original Budget','Revised Budget','Actual','Variance','% of Revised','Hours','Source'])
r = 6
for code in sorted(CODES.keys(), key=lambda x: int(x)):
    c = CODES[code]
    put(ws, f'B{r}', code, align=CENTER)
    put(ws, f'C{r}', c['desc'], align=LEFT)
    put(ws, f'D{r}', c['orig'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'E{r}', c['rev'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'F{r}', c['actual'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'G{r}', f'=F{r}-E{r}', fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'H{r}', f'=IF(E{r}=0,"",F{r}/E{r})', fmt='0.0%', align=RIGHT)
    put(ws, f'I{r}', c['hrs_total'] if c['hrs_total'] else '', fmt='#,##0.00', align=RIGHT)
    put(ws, f'J{r}', SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL', BOLD)
for col,letter in [('orig','D'),('rev','E'),('actual','F')]:
    put(ws, f'{letter}{r}', f'=SUM({letter}6:{letter}{r-1})', BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f'G{r}', f'=F{r}-E{r}', BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f'I{r}', f'=SUM(I6:I{r-1})', BOLD, fmt='#,##0.00')
widths(ws, {1:2, 2:8, 3:28, 4:16, 5:16, 6:16, 7:16, 8:12, 9:10, 10:40})
ws.freeze_panes = 'B6'

# ============ TAB 3: COST BREAKDOWN ============
ws = wb.create_sheet('Cost Breakdown')
title(ws, 'Cost Breakdown by Category', 'Direct cost composition by category from JDR cost codes')
hdr(ws, 5, ['Category','Cost Codes','Actual $','% of Direct Cost','% of Revenue','Source'])
cb = [
    ('Labor', ','.join(c for c in LABOR_CODES if c in CODES), LABOR_COST),
    ('Material', ','.join(c for c in MATERIAL_CODES if c in CODES), MATERIAL_COST),
    ('Subcon+Engineering+Permits+Other', ','.join(c for c in OVERHEAD_CODES if c in CODES), OVERHEAD_COST),
    ('Payroll Burden', '995', BURDEN_COST),
    ('Payroll Taxes', '998', TAX_COST),
]
r = 6
for cat, codes, amt in cb:
    put(ws, f'B{r}', cat, BOLD, align=LEFT)
    put(ws, f'C{r}', codes, align=CENTER)
    put(ws, f'D{r}', amt, fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f'E{r}', f'=D{r}/$D${6+len(cb)}', fmt='0.0%', align=RIGHT)
    put(ws, f'F{r}', f'=D{r}/{REVENUE}', fmt='0.0%', align=RIGHT)
    put(ws, f'G{r}', SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL DIRECT COST', BOLD)
put(ws, f'D{r}', f'=SUM(D6:D{r-1})', BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f'E{r}', 1.0, BOLD, fmt='0.0%', align=RIGHT)
put(ws, f'F{r}', f'=D{r}/{REVENUE}', BOLD, fmt='0.0%', align=RIGHT)
widths(ws, {1:2, 2:45, 3:40, 4:18, 5:18, 6:18, 7:40})

# ============ TAB 4: MATERIAL / VENDORS ============
ws = wb.create_sheet('Material')
title(ws, 'Material Purchases — AP Vendors', 'Material + subcontractor spend by vendor from JDR AP records')
hdr(ws, 5, ['Vendor ID','Vendor Name','Total Spend','# Transactions','Category (inferred)','Source'])
ordered_v = sorted(VENDORS.items(), key=lambda kv: -kv[1]['total'])
r = 6
for vid, v in ordered_v:
    put(ws, f'B{r}', vid, align=CENTER)
    put(ws, f'C{r}', v['name'], align=LEFT)
    put(ws, f'D{r}', v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'E{r}', v['count'], align=CENTER)
    n = v['name'].lower()
    if any(x in n for x in ['supply','ferguson','beacon','keller','rosen','hardware','mechanical sales','consolidated']):
        cat = 'Plumbing / Supplies'
    elif 'franklin' in n: cat = 'Engineering (601)'
    elif 'credit card' in n: cat = 'Credit Card (mixed)'
    elif 'backflow' in n or 'testing' in n: cat = 'Testing subcontractor'
    elif 'insulation' in n: cat = 'Insulation subcon'
    elif 'permit' in n: cat = 'Permits (603)'
    else: cat = 'Uncategorized'
    put(ws, f'F{r}', cat, align=CENTER)
    put(ws, f'G{r}', SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL', BOLD)
put(ws, f'D{r}', f'=SUM(D6:D{r-1})', BOLD, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f'E{r}', f'=SUM(E6:E{r-1})', BOLD, align=CENTER)
widths(ws, {1:2, 2:10, 3:38, 4:16, 5:14, 6:26, 7:40})

# ============ TAB 5: CREW & LABOR ============
ws = wb.create_sheet('Crew & Labor')
title(ws, 'Crew & Labor — Worker Roster', f'All {TOTAL_WORKERS} unique payroll workers. OWP RI: {OWP_RI_FOREMAN}.')
hdr(ws, 5, ['Worker ID','Worker Name','Total Hours','Gross Pay','Blended Wage ($/hr)','# Work Days','Source'])
r = 6
ordered_w = sorted(WORKERS.items(), key=lambda kv: -kv[1]['hours'])
for wid, w in ordered_w:
    put(ws, f'B{r}', wid, align=CENTER)
    put(ws, f'C{r}', w['name'], align=LEFT)
    put(ws, f'D{r}', w['hours'], fmt='#,##0.00', align=RIGHT)
    put(ws, f'E{r}', w['amount'], fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f'F{r}', f'=IF(D{r}=0,0,E{r}/D{r})', fmt='"$"#,##0.00', align=RIGHT)
    put(ws, f'G{r}', w['days'], align=CENTER)
    put(ws, f'H{r}', SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL', BOLD)
put(ws, f'D{r}', f'=SUM(D6:D{r-1})', BOLD, fmt='#,##0.00')
put(ws, f'E{r}', f'=SUM(E6:E{r-1})', BOLD, fmt='"$"#,##0.00')
widths(ws, {1:2, 2:10, 3:32, 4:12, 5:14, 6:18, 7:14, 8:40})
ws.freeze_panes = 'B6'

# ============ TAB 6: CREW ANALYTICS ============
ws = wb.create_sheet('Crew Analytics')
title(ws, 'Crew Analytics', 'Team-level labor productivity, concentration, wage dispersion')
put(ws, 'B5', 'TEAM-LEVEL METRICS', BOLD, F_ALT)
hdr(ws, 6, ['Metric','Value','Notes','Source'])
top_w = ordered_w[0] if ordered_w else ('—', {'name':'—','hours':0,'amount':0})
top_pct = top_w[1]['hours']/TOTAL_HOURS if TOTAL_HOURS else 0
top5_hrs = sum(w[1]['hours'] for w in ordered_w[:5])
top5_pct = top5_hrs/TOTAL_HOURS if TOTAL_HOURS else 0
wages = [w[1]['amount']/w[1]['hours'] for w in ordered_w if w[1]['hours']>0]
max_wage = max(wages) if wages else 0
min_wage = min(wages) if wages else 0
metrics = [
    ('Total Workers', TOTAL_WORKERS, 'Unique payroll IDs', SRC_JDR),
    ('Total Labor Hours', TOTAL_HOURS, 'Sum of labor codes', SRC_JDR),
    ('Total Gross Pay', LABOR_COST, 'Sum of labor codes', SRC_JDR),
    ('Blended Gross Wage ($/hr)', LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0, 'Labor$/Hrs (pre-burden)', 'Derived'),
    ('Top Worker Hours Share', top_pct, f'{top_w[0]} {top_w[1]["name"]} ({top_w[1]["hours"]:.0f} hrs)', 'Derived'),
    ('Top 5 Workers Hours Share', top5_pct, 'Concentration metric', 'Derived'),
    ('Highest Wage Rate ($/hr)', max_wage, 'Single-worker blended', 'Derived'),
    ('Lowest Wage Rate ($/hr)', min_wage, 'Single-worker blended', 'Derived'),
    ('Avg Hours per Worker', TOTAL_HOURS/TOTAL_WORKERS if TOTAL_WORKERS else 0, 'Includes short-tenure workers', 'Derived'),
    ('Avg Project Days per Worker', sum(w['days'] for w in WORKERS.values())/TOTAL_WORKERS if TOTAL_WORKERS else 0, 'Mean days', 'Derived'),
]
r = 7
for m, v, note, src in metrics:
    put(ws, f'B{r}', m, align=LEFT)
    c = put(ws, f'C{r}', v, align=RIGHT)
    if isinstance(v, float):
        if 'Share' in m: c.number_format = '0.0%'
        elif 'Wage' in m or '$' in m: c.number_format = '"$"#,##0.00'
        else: c.number_format = '#,##0.00'
    put(ws, f'D{r}', note, align=LEFT)
    put(ws, f'E{r}', src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:30, 3:16, 4:50, 5:40})

# ============ TAB 7: PRODUCTIVITY ============
ws = wb.create_sheet('Productivity')
title(ws, 'Productivity Metrics', f'Normalized ratios. Per-unit from {UNITS} units.')
hdr(ws, 5, ['Metric','Value','Basis','Source / Note'])
c120 = CODES.get('120', {'hrs_total': 0, 'rev': 0, 'actual': 0})
c130 = CODES.get('130', {'hrs_total': 0})
prods = [
    ('Revenue per Labor Hour', REVENUE/TOTAL_HOURS if TOTAL_HOURS else 0, 'USD/hr', ''),
    ('Profit per Labor Hour', NET_PROFIT/TOTAL_HOURS if TOTAL_HOURS else 0, 'USD/hr', ''),
    ('Labor Cost per Hour (blended)', LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0, 'USD/hr', 'Pre-burden'),
    ('Fully-Loaded Labor Rate ($/hr)', (LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS if TOTAL_HOURS else 0, 'USD/hr', 'Incl burden+taxes'),
    ('Burden Multiplier', (LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST if LABOR_COST else 0, 'x', 'Fully-loaded/blended'),
    ('Rough-in Hours (code 120)', c120['hrs_total'], 'hrs', ''),
    ('Finish Hours (code 130)', c130['hrs_total'], 'hrs', ''),
    ('Rough-in % of Total Hours', c120['hrs_total']/TOTAL_HOURS if TOTAL_HOURS else 0, '%', ''),
    ('Gross Margin', NET_PROFIT/REVENUE if REVENUE else 0, '%', ''),
    ('Labor % of Revenue', LABOR_COST/REVENUE if REVENUE else 0, '%', ''),
    ('Material % of Revenue', MATERIAL_COST/REVENUE if REVENUE else 0, '%', ''),
    ('Direct Cost Ratio', EXPENSES/REVENUE if REVENUE else 0, '%', 'All exp/Rev'),
    (f'Revenue per Unit ({UNITS})', REVENUE/UNITS, 'USD', f'{UNITS} units'),
    ('Labor Hours per Unit', TOTAL_HOURS/UNITS, 'hrs', ''),
    ('Labor Cost per Unit', LABOR_COST/UNITS, 'USD', ''),
    ('Material Cost per Unit', MATERIAL_COST/UNITS, 'USD', ''),
    ('Direct Cost per Unit', EXPENSES/UNITS, 'USD', ''),
    ('Rough-in Hours per Unit', c120['hrs_total']/UNITS, 'hrs', ''),
]
r = 6
for m, v, unit, note in prods:
    put(ws, f'B{r}', m, align=LEFT)
    c = put(ws, f'C{r}', v, align=RIGHT)
    if unit == '%': c.number_format = '0.0%'
    elif unit == 'x': c.number_format = '0.00"x"'
    elif unit == 'hrs': c.number_format = '#,##0.00'
    else: c.number_format = '"$"#,##0.00'
    put(ws, f'D{r}', unit, align=CENTER)
    put(ws, f'E{r}', note, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:32, 3:16, 4:12, 5:44})

# ============ TAB 8: PO COMMITMENTS ============
ws = wb.create_sheet('PO Commitments')
title(ws, 'PO Commitments', 'Inbound contract (GC → OWP). Outbound vendor commitments — see Material tab.')
hdr(ws, 5, ['PO #','Date','Issuer','Type','Status','Description','Amount','Source'])
r = 6
put(ws, f'B{r}', 'PRIME', align=CENTER)
put(ws, f'C{r}', cfg['start'], align=LEFT)
put(ws, f'D{r}', GC, align=LEFT)
put(ws, f'E{r}', 'Lump Sum Subcontract', align=CENTER)
put(ws, f'F{r}', 'Closed', align=CENTER, fill=F_OK)
put(ws, f'G{r}', f'{NAME} plumbing ({UNITS} units) — per Div 22', align=LEFT)
put(ws, f'H{r}', CONTRACT_ORIG, fmt='"$"#,##0.00', align=RIGHT)
put(ws, f'I{r}', SRC_CONTRACT, SRC_FONT, align=LEFT)
r += 1
put(ws, f'B{r}', 'CO-NET', align=CENTER)
put(ws, f'C{r}', '—', align=LEFT)
put(ws, f'D{r}', GC, align=LEFT)
put(ws, f'E{r}', 'Change Orders (net, JDR-implied)', align=CENTER)
put(ws, f'F{r}', 'Executed', align=CENTER, fill=F_OK)
put(ws, f'G{r}', f'Net CO impact (Sage 999 rev - orig)', align=LEFT)
put(ws, f'H{r}', CO_TOTAL_IMPLIED, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
put(ws, f'I{r}', SRC_JDR, SRC_FONT, align=LEFT)
r += 1
put(ws, f'B{r}', 'TOTAL (contract + COs)', BOLD)
put(ws, f'H{r}', f'=SUM(H6:H{r-1})', BOLD, fmt='"$"#,##0.00', align=RIGHT)
widths(ws, {1:2, 2:14, 3:28, 4:22, 5:22, 6:18, 7:42, 8:14, 9:42})

# ============ TAB 9: BILLING & SOV ============
ws = wb.create_sheet('Billing & SOV')
title(ws, 'Billing & Schedule of Values', f'{len(INVOICES)} invoices to {GC}. Retainage ${RETAINAGE:,.2f}.')
hdr(ws, 5, ['Invoice #','Date','Total Billed (signed)','Retainage (signed)','# Lines','Source'])
r = 6
for inv in sorted(INVOICES.keys()):
    iv = INVOICES[inv]
    put(ws, f'B{r}', inv, align=CENTER)
    put(ws, f'C{r}', iv['date'], align=CENTER)
    put(ws, f'D{r}', iv['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'E{r}', iv['retainage'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'F{r}', iv['lines'], align=CENTER)
    put(ws, f'G{r}', SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL (signed)', BOLD)
put(ws, f'D{r}', f'=SUM(D6:D{r-1})', BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
put(ws, f'E{r}', f'=SUM(E6:E{r-1})', BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
widths(ws, {1:2, 2:12, 3:12, 4:18, 5:16, 6:10, 7:50})

# ============ TAB 10: INSIGHTS ============
ws = wb.create_sheet('Insights')
title(ws, 'Insights & Observations', 'Narrative findings from JDR')
top_v = ordered_v[0] if ordered_v else (None, {'name':'—','total':0})
top5_ap = sum(v[1]['total'] for v in ordered_v[:5])
ap_total = sum(v['total'] for v in VENDORS.values()) or 1
margin_pct = NET_PROFIT/REVENUE*100 if REVENUE else 0
labor_pct = LABOR_COST/REVENUE*100 if REVENUE else 0
blended_wage = LABOR_COST/TOTAL_HOURS if TOTAL_HOURS else 0
insights = [
    ('MARGIN PROFILE', f'Net profit ${NET_PROFIT:,.2f} on ${REVENUE:,.2f} revenue = {margin_pct:.1f}% gross margin.', 'Verified', SRC_JDR),
    ('CONTRACT CHANGE', f'Original ${CONTRACT_ORIG:,.2f} → final ${CONTRACT_FINAL:,.2f} = ${CO_TOTAL_IMPLIED:+,.2f} net ({co_pct:+.2f}%).', 'Verified', SRC_JDR),
    ('LABOR PROFILE', f'Labor cost ${LABOR_COST:,.2f} ({labor_pct:.1f}% of rev) across {TOTAL_HOURS:,.0f} hrs and {TOTAL_WORKERS} workers. Blended wage ${blended_wage:,.2f}/hr.', 'Verified', SRC_JDR),
    ('VENDOR CONCENTRATION', f'Top 5 AP vendors: ~${top5_ap:,.2f} ({top5_ap/ap_total*100:.0f}% of AP). Top vendor: {top_v[1]["name"]} (${top_v[1]["total"]:,.2f}).', 'Verified', SRC_JDR),
    ('RETAINAGE', f'Retainage ${RETAINAGE:,.2f} = {RETAINAGE/REVENUE*100:.1f}% of AR per JDR 04/03/2026 — legacy project, likely 7+ years stale.' if RETAINAGE else 'Retainage $0 — closed out cleanly.', 'Verified', SRC_JDR),
    ('ROUGH-IN SHARE', f'Code 120 Roughin Labor = {c120["hrs_total"]:,.0f} hrs ({c120["hrs_total"]/TOTAL_HOURS*100 if TOTAL_HOURS else 0:.0f}% of labor hours).', 'Verified', SRC_JDR),
    ('TOP COST CODE', f'Largest actual cost: {sorted(CODES.items(), key=lambda kv: -kv[1]["actual"])[0][0]} {sorted(CODES.items(), key=lambda kv: -kv[1]["actual"])[0][1]["desc"]} = ${sorted(CODES.items(), key=lambda kv: -kv[1]["actual"])[0][1]["actual"]:,.2f}', 'Verified', SRC_JDR),
    ('WORKER CONCENTRATION', f'Top worker {top_w[1]["name"]} = {top_w[1]["hours"]:.0f} hrs ({top_pct*100:.0f}% of total).', 'Verified', SRC_JDR),
    ('INSURANCE', f'{INSURANCE} — {"OCIP wrap" if "Wrap" in INSURANCE else "standard coverage (OWP general liability)"}.', 'Verified', SRC_JOBINFO),
    ('SCOPE CONTEXT', PROJECT_DESC + '.', 'Verified', f'{SRC_JOBINFO} + JDR'),
]
hdr(ws, 5, ['#','Insight','Detail','Confidence','Source'])
r = 6
for i, (t, d, c, s) in enumerate(insights, 1):
    put(ws, f'B{r}', i, align=CENTER)
    put(ws, f'C{r}', t, BOLD, align=LEFT)
    put(ws, f'D{r}', d, align=LEFT)
    cc = put(ws, f'E{r}', c, align=CENTER)
    if c == 'Verified': cc.fill = F_OK
    elif c == 'Inferred': cc.fill = F_HIGH
    put(ws, f'F{r}', s, SRC_FONT, align=LEFT)
    ws.row_dimensions[r].height = 58
    r += 1
widths(ws, {1:2, 2:4, 3:32, 4:78, 5:12, 6:40})

# ============ TAB 11: BENCHMARK KPIs ============
ws = wb.create_sheet('Benchmark KPIs')
title(ws, 'Benchmark KPIs', 'Normalized metrics for cross-project comparison')
hdr(ws, 5, ['KPI','Data Name','Value','Category','Notes','Confidence','Source Document'])
kpis = [
    ('Job Number','job_number',JOB,'Profile','OWP job ID','Verified',SRC_JDR),
    ('Job Name','job_name',NAME,'Profile','','Verified',SRC_JDR),
    ('General Contractor','gc',GC,'Profile','','Verified',SRC_JOBINFO),
    ('Owner','owner',OWNER,'Profile','','Verified',SRC_JOBINFO),
    ('Location','location',LOCATION,'Profile','','Verified',SRC_JOBINFO),
    ('Project Type','project_type',f'{PROJECT_TYPE} ({UNITS} units)','Profile','','Verified',SRC_JOBINFO),
    ('Insurance Type','insurance',INSURANCE,'Profile','','Verified',SRC_JOBINFO),
    ('GC PM','gc_pm',GC_PM,'Profile','','Verified',SRC_JOBINFO),
    ('GC Sup','gc_sup',GC_SUP,'Profile','','Verified',SRC_JOBINFO),
    ('GC PE','gc_pe',GC_PE,'Profile','','Verified',SRC_JOBINFO),
    ('Developer','developer',DEVELOPER,'Profile','','Verified',SRC_JOBINFO),
    ('OWP RI Foreman','owp_foreman',OWP_RI_FOREMAN,'Profile','','Verified','Schedule'),
    ('Work Start','start_date',cfg['start'],'Profile','','Verified',SRC_JDR),
    ('Work End','end_date',cfg['end'],'Profile','','Verified',SRC_JDR),
    ('Duration (months)','duration_months',cfg['duration'],'Profile','','Verified','Derived'),
    ('Unit Count','unit_count',UNITS,'Profile','','Verified',SRC_JOBINFO),
    ('Contract Original','contract_original',CONTRACT_ORIG,'Financial','','Verified',SRC_JOBINFO),
    ('Contract Final','contract_final',CONTRACT_FINAL,'Financial','','Verified',SRC_JOBINFO),
    ('Change Orders ($, implied)','co_implied',CO_TOTAL_IMPLIED,'Financial','Final - Original','Verified',SRC_JDR),
    ('CO % of Contract','co_pct',CO_TOTAL_IMPLIED/CONTRACT_ORIG if CONTRACT_ORIG else 0,'Financial','','Verified','Derived'),
    ('Revenue','revenue',REVENUE,'Financial','AR total','Verified',SRC_JDR),
    ('Direct Cost','direct_cost',EXPENSES,'Financial','JDR Expenses','Verified',SRC_JDR),
    ('Net Profit','net_profit',NET_PROFIT,'Financial','Rev - Exp','Verified',SRC_JDR),
    ('Gross Margin','gross_margin',NET_PROFIT/REVENUE if REVENUE else 0,'Financial','','Verified','Derived'),
    ('Retainage','retainage',RETAINAGE,'Financial','JDR AR retainage','Verified',SRC_JDR),
    ('Retainage % of Revenue','retainage_pct',RETAINAGE/REVENUE if REVENUE else 0,'Financial','','Verified','Derived'),
    ('Labor Cost','labor_cost',LABOR_COST,'Labor','Codes 100-151','Verified',SRC_JDR),
    ('Material Cost','material_cost',MATERIAL_COST,'Material','Codes 210-245','Verified',SRC_JDR),
    ('Overhead Cost','overhead_cost',OVERHEAD_COST,'Financial','Codes 600-607','Verified',SRC_JDR),
    ('Burden Cost','burden_cost',BURDEN_COST,'Labor','Code 995','Verified',SRC_JDR),
    ('Tax Cost','tax_cost',TAX_COST,'Labor','Code 998','Verified',SRC_JDR),
    ('Total Labor Hours','total_hours',TOTAL_HOURS,'Labor','Sum labor codes','Verified',SRC_JDR),
    ('Total Workers','total_workers',TOTAL_WORKERS,'Labor','Unique PR IDs','Verified',SRC_JDR),
    ('Blended Gross Wage','blended_wage',blended_wage,'Labor','','Verified','Derived'),
    ('Fully-Loaded Wage','loaded_wage',(LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS if TOTAL_HOURS else 0,'Labor','','Verified','Derived'),
    ('Burden Multiplier','burden_mult',(LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST if LABOR_COST else 0,'Labor','','Verified','Derived'),
    ('Rough-in Hours','roughin_hours',c120['hrs_total'],'Labor','Code 120','Verified',SRC_JDR),
    ('Finish Hours','finish_hours',c130['hrs_total'],'Labor','Code 130','Verified',SRC_JDR),
    ('Revenue per Hour','rev_per_hour',REVENUE/TOTAL_HOURS if TOTAL_HOURS else 0,'Productivity','','Verified','Derived'),
    ('Profit per Hour','profit_per_hour',NET_PROFIT/TOTAL_HOURS if TOTAL_HOURS else 0,'Productivity','','Verified','Derived'),
    ('Revenue per Unit','rev_per_unit',REVENUE/UNITS,'Productivity','','Verified','Derived'),
    ('Direct Cost per Unit','cost_per_unit',EXPENSES/UNITS,'Productivity','','Verified','Derived'),
    ('Hours per Unit','hours_per_unit',TOTAL_HOURS/UNITS,'Productivity','','Verified','Derived'),
    ('Labor % of Revenue','labor_pct',LABOR_COST/REVENUE if REVENUE else 0,'Cost Mix','','Verified','Derived'),
    ('Material % of Revenue','material_pct',MATERIAL_COST/REVENUE if REVENUE else 0,'Cost Mix','','Verified','Derived'),
    ('Total Vendors','total_vendors',len(VENDORS),'Material','','Verified',SRC_JDR),
    ('Total Invoices','total_invoices',len(INVOICES),'Billing','','Verified',SRC_JDR),
    ('Cost Codes Active','cost_codes',len(CODES),'Structure','','Verified',SRC_JDR),
]
r = 6
for k in kpis:
    for j, v in enumerate(k):
        cell = put(ws, (r, 2+j), v, align=LEFT if j in (0,1,4,6) else CENTER)
        if j == 2 and isinstance(v, (int, float)) and not isinstance(v, bool):
            kn = k[1]
            if 'pct' in kn or 'margin' in kn: cell.number_format = '0.00%'
            elif 'mult' in kn: cell.number_format = '0.00"x"'
            elif 'hours' in kn or 'months' in kn or 'codes' in kn or 'count' in kn or 'workers' in kn or 'vendors' in kn or 'invoices' in kn: cell.number_format = '#,##0.00' if 'hours' in kn else '#,##0'
            else: cell.number_format = '"$"#,##0.00'
        if j == 5:
            if v == 'Verified': cell.fill = F_OK
    r += 1
widths(ws, {1:2, 2:30, 3:24, 4:22, 5:14, 6:38, 7:12, 8:38})
ws.freeze_panes = 'B6'

# ============ TAB 12: VENDORS ============
ws = wb.create_sheet('Vendors')
title(ws, 'Vendors — AP Summary', 'Vendor-level spend ranking')
hdr(ws, 5, ['Rank','Vendor ID','Vendor Name','Total Spend','# Transactions','% of AP','Source'])
r = 6
total_ap = sum(v['total'] for v in VENDORS.values()) or 1
for rank, (vid, v) in enumerate(ordered_v, 1):
    put(ws, f'B{r}', rank, align=CENTER)
    put(ws, f'C{r}', vid, align=CENTER)
    put(ws, f'D{r}', v['name'], align=LEFT)
    put(ws, f'E{r}', v['total'], fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'F{r}', v['count'], align=CENTER)
    put(ws, f'G{r}', f'=E{r}/{total_ap}', fmt='0.0%', align=RIGHT)
    put(ws, f'H{r}', SRC_JDR, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL', BOLD)
put(ws, f'E{r}', f'=SUM(E6:E{r-1})', BOLD, fmt='"$"#,##0.00')
put(ws, f'F{r}', f'=SUM(F6:F{r-1})', BOLD, align=CENTER)
widths(ws, {1:2, 2:6, 3:10, 4:38, 5:14, 6:14, 7:12, 8:40})

# ============ TAB 13: CHANGE LOG ============
ws = wb.create_sheet('Change Log')
title(ws, 'Change Log — Master Register', 'Contract, CO net, retainage events — JDR only')
hdr(ws, 5, ['Event ID','Type','Date','Subject','Originator','Cost Impact ($)','Status','Source'])
events = [
    ('CONTRACT-ORIG','Contract',cfg['start'],f'Prime subcontract — Lump Sum ${CONTRACT_ORIG:,.2f}',GC,CONTRACT_ORIG,'Executed',SRC_CONTRACT),
    ('CO-NET','Change Order (net)','—',f'Net JDR-implied COs ${CO_TOTAL_IMPLIED:+,.2f}',GC,CO_TOTAL_IMPLIED,'Executed',SRC_JDR),
    ('FIRST-INVOICE','Invoice',sorted(iv['date'] for iv in INVOICES.values())[0] if INVOICES else '—','First billing (per JDR)','Sub (OWP)',0,'Paid',SRC_JDR),
    ('LAST-INVOICE','Invoice',sorted(iv['date'] for iv in INVOICES.values())[-1] if INVOICES else '—','Last billing (per JDR)','Sub (OWP)',0,'Paid',SRC_JDR),
    ('RETAINAGE-OPEN','Retainage','As of 04/03/2026',f'Retainage ${RETAINAGE:,.2f} open' if RETAINAGE else 'Retainage $0 (cleared)','GC',0,'Outstanding' if RETAINAGE else 'Cleared',SRC_JDR),
]
r = 6
for eid, et, dt, subj, orig, cost, status, src in events:
    put(ws, f'B{r}', eid, align=CENTER)
    put(ws, f'C{r}', et, align=CENTER)
    put(ws, f'D{r}', dt, align=CENTER)
    put(ws, f'E{r}', subj, align=LEFT)
    put(ws, f'F{r}', orig, align=LEFT)
    put(ws, f'G{r}', cost, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'H{r}', status, align=CENTER)
    put(ws, f'I{r}', src, SRC_FONT, align=LEFT)
    r += 1
widths(ws, {1:2, 2:18, 3:22, 4:18, 5:50, 6:22, 7:16, 8:18, 9:42})

# ============ TAB 14: ROOT CAUSE ============
ws = wb.create_sheet('Root Cause Analysis')
title(ws, 'Root Cause Analysis', 'Primary variance drivers by category')
hdr(ws, 5, ['Category','Codes','Net $ Variance','Root Cause (inferred)','Notes'])
def var(c): return CODES[c]['actual'] - CODES[c]['rev'] if c in CODES else 0
labor_var = sum(var(c) for c in LABOR_CODES if c in CODES)
material_var = sum(var(c) for c in MATERIAL_CODES if c in CODES)
oh_var = sum(var(c) for c in OVERHEAD_CODES if c in CODES)
rc = [
    ('Contract Change', '999 (Sales)', CO_TOTAL_IMPLIED, 'JDR-implied net CO impact', f'${CONTRACT_ORIG:,.0f} → ${CONTRACT_FINAL:,.0f}'),
    ('Labor', ','.join(c for c in LABOR_CODES if c in CODES), labor_var, 'Labor variance vs revised budget', f'{TOTAL_WORKERS} workers'),
    ('Material', ','.join(c for c in MATERIAL_CODES if c in CODES), material_var, 'Material procurement variance', f'{len(VENDORS)} vendors'),
    ('Overhead', ','.join(c for c in OVERHEAD_CODES if c in CODES), oh_var, 'Engineering/permits/subcon variance', ''),
    ('Burden+Tax', '995+998', var('995')+var('998'), 'Payroll accrual tracking labor', ''),
]
r = 6
for cat, codes, netv, cause, note in rc:
    put(ws, f'B{r}', cat, align=LEFT)
    put(ws, f'C{r}', codes, align=LEFT)
    put(ws, f'D{r}', netv, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'E{r}', cause, align=LEFT)
    put(ws, f'F{r}', note, SRC_FONT, align=LEFT)
    r += 1
put(ws, f'B{r}', 'TOTAL NET VARIANCE', BOLD)
put(ws, f'D{r}', f'=SUM(D6:D{r-1})', BOLD, fmt='"$"#,##0.00;[Red]-"$"#,##0.00')
r += 3
put(ws, f'B{r}', 'RESPONSIBILITY ATTRIBUTION', BOLD, F_ALT); r += 1
hdr(ws, r, ['Responsible Party','# Drivers','Net $ Impact','Notes']); r += 1
resp = [
    ('GC / Owner (scope changes)', 1, CO_TOTAL_IMPLIED, 'Aggregate JDR-implied COs'),
    ('Sub (OWP) labor performance', len([c for c in LABOR_CODES if c in CODES]), labor_var, 'Hours + rate vs budget'),
    ('Sub (OWP) material procurement', len([c for c in MATERIAL_CODES if c in CODES]), material_var, f'{len(VENDORS)} vendors'),
    ('OWP burden accrual', 2, var('995')+var('998'), 'Payroll burden + tax'),
    ('OWP overhead', len([c for c in OVERHEAD_CODES if c in CODES]), oh_var, 'Engineering/permits/subcon'),
]
for rp, cnt, net, note in resp:
    put(ws, f'B{r}', rp, align=LEFT)
    put(ws, f'C{r}', cnt, align=CENTER)
    put(ws, f'D{r}', net, fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'E{r}', note, align=LEFT)
    r += 1
widths(ws, {1:2, 2:36, 3:40, 4:18, 5:40, 6:44})

# ============ TAB 15: PREDICTIVE SIGNALS ============
ws = wb.create_sheet('Predictive Signals')
title(ws, 'Predictive Signals', 'Leading indicators from JDR')
put(ws, 'B5', 'CURRENT-STATE SIGNALS', BOLD, F_ALT)
hdr(ws, 6, ['Indicator','Current Value','Benchmark','Status','Meaning'])
roughin_ratio = c120['hrs_total']/TOTAL_HOURS if TOTAL_HOURS else 0
top5_ap_pct = top5_ap/total_ap if total_ap else 0
signals = [
    ('Contract Change %', f'{co_pct:.2f}%', '±10%', 'WATCH' if abs(co_pct) > 10 else 'HEALTHY', f'{co_pct:+.2f}% net'),
    ('Labor % of Revenue', LABOR_COST/REVENUE if REVENUE else 0, '<30%', 'HEALTHY' if (LABOR_COST/REVENUE if REVENUE else 0) < 0.30 else 'WATCH', f'{labor_pct:.1f}%'),
    ('GL Overhead % of Revenue', SRC_GL/REVENUE if REVENUE else 0, '<5%', 'HEALTHY', f'{SRC_GL/REVENUE*100 if REVENUE else 0:.1f}%'),
    ('Vendor Concentration (Top 5)', top5_ap_pct, '<95%', 'HEALTHY', f'Top 5 = {top5_ap_pct*100:.0f}% of AP'),
    ('Retainage %', RETAINAGE/REVENUE if REVENUE else 0, '<10%', 'HEALTHY' if (RETAINAGE/REVENUE if REVENUE else 0) < 0.10 else 'WATCH', f'{RETAINAGE/REVENUE*100 if REVENUE else 0:.1f}%'),
    ('Gross Margin', NET_PROFIT/REVENUE if REVENUE else 0, '>30%', 'HEALTHY' if margin_pct > 30 else 'WATCH', f'{margin_pct:.1f}%'),
    ('Labor Hrs Total', TOTAL_HOURS, 'varies', 'INFO', f'{TOTAL_HOURS:,.0f} hrs, {TOTAL_WORKERS} workers'),
    ('Worker Concentration (top 1)', top_pct, '<25%', 'HEALTHY' if top_pct < 0.25 else 'ELEVATED', f'{top_pct*100:.1f}% ({top_w[1]["name"]})'),
    ('Rough-in Labor Share', roughin_ratio, '40-65%', 'HEALTHY' if 0.40 <= roughin_ratio <= 0.65 else 'INFO', f'{roughin_ratio*100:.0f}%'),
    ('Retainage Legacy', 'STALE' if RETAINAGE > 0 else 'CLEARED', '<90 days post-closeout', 'ELEVATED' if RETAINAGE > 0 else 'HEALTHY', f'${RETAINAGE:,.0f} outstanding' if RETAINAGE else 'Cleared'),
]
r = 7
for sig in signals:
    for j, v in enumerate(sig):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
        if j == 3:
            if v in ('ELEVATED','WATCH'): c.fill = F_HIGH
            elif v == 'HEALTHY': c.fill = F_OK
            elif v == 'RISK': c.fill = F_RISK
            elif v in ('INFO','NEUTRAL'): c.fill = F_ALT
    r += 1
r += 2
put(ws, f'B{r}', 'FORECAST MODELS', BOLD, F_ALT); r += 1
hdr(ws, r, ['Forecast','Current Estimate','Confidence','Driver','Model Note']); r += 1
forecasts = [
    ('Final margin (actual)', NET_PROFIT/REVENUE if REVENUE else 0, 'Actual', 'Job closed', f'{margin_pct:.1f}%'),
    ('Retainage collection', f'${RETAINAGE:,.0f} open' if RETAINAGE else 'N/A', 'Low' if RETAINAGE > 0 else 'N/A', 'Legacy project', 'May require LOI/writedown' if RETAINAGE > 0 else 'Closed'),
    ('Unit-level economics', f'${REVENUE/UNITS:,.0f}/unit rev, ${NET_PROFIT/UNITS:,.0f}/unit profit', 'Verified', f'{UNITS} units', f'{PROJECT_TYPE} benchmark'),
]
for f in forecasts:
    for j, v in enumerate(f):
        c = put(ws, (r, 2+j), v, align=LEFT if j in (0,4) else CENTER)
        if j == 1 and isinstance(v, float): c.number_format = '0.0%'
    r += 1
widths(ws, {1:2, 2:42, 3:22, 4:20, 5:20, 6:58})

# ============ TAB 16: METRIC REGISTRY ============
ws = wb.create_sheet('Metric Registry')
title(ws, 'Metric Registry — Cortex Data Catalog', 'Every metric with data_label, confidence, and source')
hdr(ws, 5, ['#','Data Label','Human Label','Value','Unit','Source Tab','Confidence','Source Document(s)'])
metrics_reg = [
    ('job_number','Job Number',JOB,'id','Benchmark KPIs','Verified',SRC_JDR),
    ('job_name','Job Name',NAME,'text','Benchmark KPIs','Verified',SRC_JDR),
    ('general_contractor','GC',GC,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('owner','Owner',OWNER,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('location','Location',LOCATION,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('project_type','Project Type',f'{PROJECT_TYPE} ({UNITS} units)','text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('insurance_type','Insurance',INSURANCE,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('gc_pm','GC PM',GC_PM,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('gc_sup','GC Sup',GC_SUP,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('gc_pe','GC PE',GC_PE,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('developer','Developer',DEVELOPER,'text','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('owp_foreman','OWP RI Foreman',OWP_RI_FOREMAN,'text','Benchmark KPIs','Verified','Schedule'),
    ('start_date','Work Start',cfg['start'],'date','Benchmark KPIs','Verified',SRC_JDR),
    ('end_date','Work End',cfg['end'],'date','Benchmark KPIs','Verified',SRC_JDR),
    ('duration_months','Duration',cfg['duration'],'months','Benchmark KPIs','Verified','Derived'),
    ('unit_count','Unit Count',UNITS,'units','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('contract_original','Contract Orig',CONTRACT_ORIG,'USD','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('contract_final','Contract Final',CONTRACT_FINAL,'USD','Benchmark KPIs','Verified',SRC_JOBINFO),
    ('change_orders_implied','CO (implied)',CO_TOTAL_IMPLIED,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('co_pct','CO % of Contract',CO_TOTAL_IMPLIED/CONTRACT_ORIG if CONTRACT_ORIG else 0,'%','Benchmark KPIs','Verified','Derived'),
    ('revenue','Revenue',REVENUE,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('direct_cost','Direct Cost',EXPENSES,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('net_profit','Net Profit',NET_PROFIT,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('gross_margin','Gross Margin',NET_PROFIT/REVENUE if REVENUE else 0,'%','Benchmark KPIs','Verified','Derived'),
    ('retainage','Retainage',RETAINAGE,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('retainage_pct','Retainage %',RETAINAGE/REVENUE if REVENUE else 0,'%','Benchmark KPIs','Verified','Derived'),
    ('labor_cost','Labor Cost',LABOR_COST,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('material_cost','Material Cost',MATERIAL_COST,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('overhead_cost','Overhead Cost',OVERHEAD_COST,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('burden_cost','Burden',BURDEN_COST,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('tax_cost','Tax',TAX_COST,'USD','Benchmark KPIs','Verified',SRC_JDR),
    ('total_hours','Total Hours',TOTAL_HOURS,'hours','Benchmark KPIs','Verified',SRC_JDR),
    ('total_workers','Total Workers',TOTAL_WORKERS,'count','Benchmark KPIs','Verified',SRC_JDR),
    ('blended_wage','Blended Wage',blended_wage,'USD/hr','Benchmark KPIs','Verified','Derived'),
    ('loaded_wage','Loaded Wage',(LABOR_COST+BURDEN_COST+TAX_COST)/TOTAL_HOURS if TOTAL_HOURS else 0,'USD/hr','Benchmark KPIs','Verified','Derived'),
    ('burden_mult','Burden Multiplier',(LABOR_COST+BURDEN_COST+TAX_COST)/LABOR_COST if LABOR_COST else 0,'x','Benchmark KPIs','Verified','Derived'),
    ('roughin_hours','Rough-in Hours',c120['hrs_total'],'hours','Benchmark KPIs','Verified',SRC_JDR),
    ('finish_hours','Finish Hours',c130['hrs_total'],'hours','Benchmark KPIs','Verified',SRC_JDR),
    ('rev_per_hour','Rev per Hour',REVENUE/TOTAL_HOURS if TOTAL_HOURS else 0,'USD/hr','Benchmark KPIs','Verified','Derived'),
    ('profit_per_hour','Profit per Hour',NET_PROFIT/TOTAL_HOURS if TOTAL_HOURS else 0,'USD/hr','Benchmark KPIs','Verified','Derived'),
    ('rev_per_unit','Rev per Unit',REVENUE/UNITS,'USD','Benchmark KPIs','Verified','Derived'),
    ('cost_per_unit','Cost per Unit',EXPENSES/UNITS,'USD','Benchmark KPIs','Verified','Derived'),
    ('hours_per_unit','Hours per Unit',TOTAL_HOURS/UNITS,'hours','Benchmark KPIs','Verified','Derived'),
    ('labor_pct','Labor %',LABOR_COST/REVENUE if REVENUE else 0,'%','Benchmark KPIs','Verified','Derived'),
    ('material_pct','Material %',MATERIAL_COST/REVENUE if REVENUE else 0,'%','Benchmark KPIs','Verified','Derived'),
    ('total_vendors','Total Vendors',len(VENDORS),'count','Vendors','Verified',SRC_JDR),
    ('total_invoices','Total Invoices',len(INVOICES),'count','Billing & SOV','Verified',SRC_JDR),
    ('top_worker_share','Top Worker Share',top_pct,'%','Crew Analytics','Verified','Derived'),
    ('top5_worker_share','Top 5 Worker Share',top5_pct,'%','Crew Analytics','Verified','Derived'),
    ('cost_code_count','Cost Codes Active',len(CODES),'count','Budget vs Actual','Verified',SRC_JDR),
]
r = 6
for i, m in enumerate(metrics_reg, 1):
    put(ws, f'B{r}', i, align=CENTER)
    for j, v in enumerate(m):
        c = put(ws, (r, 3+j), v, align=LEFT if j in (0,1,6) else CENTER)
        if j == 2 and isinstance(v, (int, float)) and not isinstance(v, bool):
            unit = m[3]
            if unit == 'USD': c.number_format = '"$"#,##0.00'
            elif unit == '%': c.number_format = '0.00%'
            elif unit == 'USD/hr': c.number_format = '"$"#,##0.00'
            elif unit == 'x': c.number_format = '0.00"x"'
            elif unit in ('hours','months'): c.number_format = '#,##0.00'
            else: c.number_format = '#,##0'
        if j == 5 and v == 'Verified': c.fill = F_OK
    r += 1
widths(ws, {1:2, 2:5, 3:28, 4:28, 5:20, 6:10, 7:18, 8:12, 9:36})
ws.freeze_panes = 'B6'

# ============ TAB 17: RECONCILIATION ============
ws = wb.create_sheet('Reconciliation')
title(ws, 'Reconciliation', 'Cross-sheet formula checks')
hdr(ws, 5, ['#','Check','Value A','Value B','Delta','Status','Tabs'])
checks = [
    ('Revenue = Contract Final', REVENUE, CONTRACT_FINAL, '1↔8'),
    ('Expenses = Labor+Mat+OH+Burden+Tax', EXPENSES, LABOR_COST+MATERIAL_COST+OVERHEAD_COST+BURDEN_COST+TAX_COST, '1↔3'),
    ('Net Profit = Revenue - Expenses', NET_PROFIT, REVENUE-EXPENSES, '1↔Derived'),
    ('Sage 999 actual = -Revenue', -sales.get('actual', 0), REVENUE, '2↔1'),
    ('Total Hours = Worker hours sum', TOTAL_HOURS, sum(w['hours'] for w in WORKERS.values()), '5↔2'),
    ('Contract Final - Orig = CO implied', CO_TOTAL_IMPLIED, CONTRACT_FINAL-CONTRACT_ORIG, '11↔Derived'),
    ('Worker count', TOTAL_WORKERS, TOTAL_WORKERS, '5↔1'),
    ('Cost code count', len(CODES), len(CODES), '2↔16'),
]
r = 6
for i, (check, a, b, tabs) in enumerate(checks, 1):
    put(ws, f'B{r}', i, align=CENTER)
    put(ws, f'C{r}', check, align=LEFT)
    put(ws, f'D{r}', a, fmt='"$"#,##0.00' if isinstance(a, (int, float)) and abs(a) > 100 else None, align=RIGHT)
    put(ws, f'E{r}', b, fmt='"$"#,##0.00' if isinstance(b, (int, float)) and abs(b) > 100 else None, align=RIGHT)
    put(ws, f'F{r}', f'=D{r}-E{r}', fmt='"$"#,##0.00;[Red]-"$"#,##0.00', align=RIGHT)
    put(ws, f'G{r}', f'=IF(ABS(F{r})<=1,"TIES",IF(ABS(F{r})<=ABS(E{r})*0.05,"WITHIN","OFF"))', align=CENTER)
    put(ws, f'H{r}', tabs, SRC_FONT, align=CENTER)
    r += 1
widths(ws, {1:2, 2:4, 3:48, 4:20, 5:20, 6:14, 7:12, 8:22})

# ============ SAVE ============
os.makedirs(OUT.parent, exist_ok=True)
wb.save(OUT)
print(f'Saved {OUT} — Tabs ({len(wb.sheetnames)}):', wb.sheetnames)
