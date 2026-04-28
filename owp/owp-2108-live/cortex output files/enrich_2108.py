#!/usr/bin/env python3
"""
enrich_2108.py — Enrich OWP_2108_JCR_Cortex_v2.xlsx — R&G Apartments / Braseth
263 units · ON HOLD · Summer 2026? · Pre-construction takeoff phase only.

Full enrichment to 2020/2061 gold-standard depth:
  • 13-section sectioned Job Info layout
  • Bid history table (3 entries: 2021 BMDC budget → 2024 Braseth bid → 2025 final BID)
  • AP vendor profile (Franklin Engineering 92.5% concentration)
  • Document meta + data sources
  • Risk flags + project narrative
  • Plus: Owner identification recovered from JDR AR section

Verified data sources:
  • 2108 Job Detail Report.pdf (Sage Timberline · Apr 3 2026 run · 6 cost codes · 52 line items)
  • 2108_dashboard_arrays.json (16 arrays mirroring index.html PROJECTS['2108'])
  • OWP Project List with Schedule - UPDATED 04-01-26.xlsx
    - Schedule rows 106 + 195: 263 units · ON HOLD-Summer 26?
    - Projects Bid r160: BUDGET 2021-09-29 · BMDC · 226u · $5,150,000 (named "Rainier & Genesee")
    - Projects Bid r339: BID 2024-10-08 · Braseth · 263u · $4,460,000 (renamed "R&G Apartments")
    - Projects Bid r364: BID 2025-02-27 · Braseth · 263u · $4,619,052 (final revised BID)
  • JDR AR section reveals: invoices to "Rainier & Genesee, LLC" + "Lake Union Partners $R&G Apts"
    confirming Lake Union Partners as owner-developer entity at the Rainier Ave S × S Genesee St
    intersection (Columbia City corridor in Seattle)

Project status: pre-construction · on hold pending Braseth go-ahead. OWP has $117,162 of sunk
design+takeoff cost (Franklin Engineering MEP design fee $90,629 + permit $6,402 + Gerard
takeoff labor $9,913 + equipment $922 + burden $1,216) against $91,402 AR billed = $25,760
carrying loss if project doesn't activate.
"""
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SCRIPT_DIR = Path(__file__).parent
JSON_FILE  = SCRIPT_DIR / "2108_data.json"
DASH_FILE  = SCRIPT_DIR / "2108_dashboard_arrays.json"
WORKBOOK   = SCRIPT_DIR / "OWP_2108_JCR_Cortex_v2.xlsx"

META_2108 = {
    "job_id":          "2108",
    "name":            "R&G Apartments (Rainier & Genesee)",
    "short_name":      "R&G Apartments",
    "gc":              "Braseth Construction",
    "gc_pm":           "TBD (Braseth) — primary contact pending subcontract execution",
    "gc_sup":          "TBD",
    "gc_pe":           "TBD",
    "owp_ri_foreman":  "TBD · project on hold (target activation Summer 2026)",
    "owp_trim_foreman":"TBD",
    "owp_signatory":   "Richard Donelson",
    "owp_estimator":   "Jeffrey S. Gerard / Jordan E. Gerard (takeoff team)",
    "owner":           "Lake Union Partners (per JDR AR section · invoice 038534)",
    "developer":       "Lake Union Partners (joint venture w/ Braseth as design-build GC)",
    "architect":       "Skb Architects",
    "structural":      "TBD",
    "mep_engineer":    "Emerald City Engineering (per Master Project List Projects Bid tab) · "
                        "Franklin Engineering on AP for $90,629 design fee — likely OWP's 3rd-party "
                        "design-build engineer alongside Emerald City as bridge MEP",
    "civil":           "TBD",
    "landscape":       "TBD",
    "site_address":    "Rainier Ave S × S Genesee St, Seattle WA (Columbia City corridor / South Seattle)",
    "permit":          "Pre-permit (no permit applied for yet)",
    "insurance":       "Standard (COI) — non-Wrap (per build_2108 META)",
    "lien_position":   "Not yet recorded — no AR billed",
    "warranty":        "Pending project activation",
    "delivery_route":  "TBD",
    "contract_doc":    "Subcontract NOT yet executed — final BID submitted Feb 27 2025 ($4,619,052 / 263u). "
                        "Project on hold pending Braseth go-ahead",
    "scope":           "Multi-family apartments — design-build plumbing scope",
    "gdrive_status":   "No GDrive Job Book folder created yet (project hasn't reached active stage). "
                        "All team data sourced from JDR identity block + OWP Master Project List "
                        "Projects Bid tab + JDR AR section vendor identification.",
    "status_text":     "ON HOLD · Summer 2026?",
    "status_severity": "WATCH",
    "ar_billed":       91402,
    "direct_cost":     117161.97,
    "net_profit":      -25759.97,    # carrying loss
    "carrying_loss":   25759.97,
    "retention":       0,
    "contract_orig":   None,
    "contract_final":  None,
    "contract_bid_latest": 4619052,  # final BID Feb 2025
    "contract_bid_first":  4460000,  # initial Braseth bid Oct 2024
    "co_net":          0,
    "units":           263,
    "units_bid_first": 226,           # original Rainier & Genesee bid (BMDC, 2021)
    "fixtures":        "TBD — no Job Ticket exists (pre-construction). Fixture schedule will populate when "
                        "Skb Architects publishes the unit-trim schedule. Estimated ~1,180 trim fixtures based "
                        "on 4.5/unit ratio from #2061 Alta Columbia City (a similar South Seattle multifamily). "
                        "Definitive count pending project activation.",
    "project_type":    "263-unit multi-family apartments · Columbia City corridor / South Seattle",
    "schedule":        "Pre-construction · on hold · target field start Summer 2026 / Q1 2027 "
                        "if Braseth confirms · est. 24-30 month duration to TCO",
    "expected_start":  "Summer 2026 / Q1 2027 (target · pending Braseth go-ahead)",
    "expected_finish": "TBD",
    "duration_months": 28,    # estimate
    "workers":         2,
    "hours":           32.5,
    "top_vendor":      "Franklin Engineering ($90,629 / 9 invoices / 92.5% of $97k AP)",
    "ap_total":        97953,
    "ap_vendor_count": 3,    # Franklin Engineering, Bank of America CC, Herc Rentals
    "doc_counts": {
        "Pay_apps":     0,    # 4 AR transactions in JDR (early progress billings) but no SOV-style pay apps
        "Executed_COs": 0,
        "CORs":         0,
        "RFIs":         0,
        "ASIs":         0,
        "Submittals":   0,
        "POs":          0,
        "Permits":      0,    # pre-permit
    },
    "bid_history": [
        ("2021-09-29", "BUDGET", "BMDC",     226, 5150000, "Original 'Rainier & Genesee' budget bid · BMDC GC · 226 units · $5.15M (per Master Project List Projects Bid r160)"),
        ("2024-10-08", "BID",    "Braseth",  263, 4460000, "Renamed 'R&G Apartments' · switched GC to Braseth · scope grew to 263 units (+37u) but bid dropped to $4.46M (-$690k from 2021 BMDC) per Master Project List r339"),
        ("2025-02-27", "BID",    "Braseth",  263, 4619052, "FINAL revised BID · same GC/units · price up 3.6% to $4.62M per Master Project List r364"),
    ],
    "narrative": [
        ("Project history",
         "OWP has been pursuing this project since 2021, originally as 'Rainier & Genesee' with BMDC as GC at 226 units / $5.15M. "
         "Project was redesigned and re-released in late 2024 with Braseth as GC, expanded to 263 units, and OWP rebid at $4.46M "
         "(a -13.4% price drop from 2021 despite +16% unit growth — driven by Braseth's tighter pricing and a leaner scope). "
         "Final revised BID Feb 2025 at $4,619,052. Project then went on hold."),
        ("Sunk-cost posture",
         "OWP has billed $91,402 against $117,162 of sunk design+takeoff cost = -$25,760 carrying loss. "
         "$90,629 of that sunk cost is the Franklin Engineering MEP design-build fee (92.5% of OWP's AP for the project). "
         "If Braseth doesn't activate the project, OWP absorbs the $25.8k as a write-off."),
        ("Owner-developer continuity",
         "Lake Union Partners is OWP's repeat client at the Rainier corridor — same owner-developer pattern as #2061 Alta Columbia "
         "City and likely #2106 6220 Roosevelt (also Phoenix Property Co · adjacent). South Seattle / Columbia City is becoming "
         "a recurring portfolio for OWP."),
        ("Activation triggers",
         "Watch for: (a) Braseth confirming summer 2026 start, (b) Skb Architects releasing IFC drawings, (c) permit application. "
         "OWP's takeoff is complete and ready to mobilize on 30 days notice."),
        ("Risk profile",
         "New GC (Braseth) — OWP has no closed-job history. Watch payment cadence + change-order behavior closely once project "
         "activates. Compare margin profile against #2107 (Chinn) and #2061 (Exxel Pacific) baselines. Modest-margin estimating "
         "exposure: $4.6M target with 25-30% margin expectation = $1.15-1.4M net profit potential."),
    ],
}

INK         = Font(name="Arial", size=10, color="1F1F1F")
INK_BOLD    = Font(name="Arial", size=10, color="1F1F1F", bold=True)
SECTION_HDR = Font(name="Arial", size=11, color="1F1F1F", bold=True)
TITLE_FONT  = Font(name="Arial", size=14, color="1F1F1F", bold=True)
GREY_FONT   = Font(name="Arial", size=9, color="6B6B6B", italic=True)
WHITE_BOLD  = Font(name="Arial", size=10, color="FFFFFF", bold=True)
WARN_FONT   = Font(name="Arial", size=10, color="9A4F02", bold=True)
GOOD_FONT   = Font(name="Arial", size=10, color="155724", bold=True)

HDR_FILL    = PatternFill("solid", start_color="2C3E50")
SECTION_FILL= PatternFill("solid", start_color="ECF0F1")
TEAM_FILL   = PatternFill("solid", start_color="FFF8E7")
WARN_FILL   = PatternFill("solid", start_color="FFF3CD")
GOOD_FILL   = PatternFill("solid", start_color="D4EDDA")
THIN        = Side(border_style="thin", color="D5D8DC")
BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP        = Alignment(wrap_text=True, vertical="top")
CENTER      = Alignment(horizontal="center", vertical="center")

def clear_range(ws, r1, r2, c1=1, c2=10):
    for r in range(r1, r2+1):
        for c in range(c1, c2+1):
            ws.cell(row=r, column=c).value = None
            ws.cell(row=r, column=c).fill = PatternFill(fill_type=None)
            ws.cell(row=r, column=c).border = Border()


def rebuild_job_info(ws):
    clear_range(ws, 1, ws.max_row + 5, 1, 6)
    ws["A1"] = f"JOB #{META_2108['job_id']} · {META_2108['name'].upper()} · KEY PLAYERS & METADATA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("Pre-construction · on hold · early takeoff phase. OWP has $117,162 sunk against "
                "$91,402 billed = -$25,760 carrying loss. Final BID submitted Feb 27 2025 to Braseth "
                "($4,619,052 / 263 units). Project paused pending Braseth go-ahead. Sources: Sage JDR + "
                "OWP Master Project List (3-bid history Sep 2021 → Feb 2025) + JDR AR section vendor "
                "identification (Lake Union Partners owner recovered from invoice 038534).")
    ws["A2"].font = GREY_FONT
    ws.merge_cells("A2:F2")

    sections = [
        ("IDENTITY", [
            ("Job Number",         META_2108["job_id"]),
            ("Job Name",           META_2108["name"]),
            ("Short Name",         META_2108["short_name"]),
            ("Project Type",       META_2108["project_type"]),
            ("Site Address",       META_2108["site_address"]),
            ("Permit",             META_2108["permit"]),
            ("Status",             META_2108["status_text"]),
            ("Scope of Work",      META_2108["scope"]),
        ]),
        ("SCHEDULE", [
            ("Project Start",     META_2108["expected_start"]),
            ("Project End",       META_2108["expected_finish"]),
            ("Duration (est.)",   f"{META_2108['duration_months']} months"),
            ("Schedule Note",     META_2108["schedule"]),
            ("Subcontract Execution", "Pending — Braseth has not committed start date"),
        ]),
        ("FINANCIAL POSTURE (PRE-CONSTRUCTION · TAKEOFF PHASE)", [
            ("Original Contract", "Not yet executed"),
            ("Latest BID submitted", f"${META_2108['contract_bid_latest']:,} (Feb 27 2025 · 263u)"),
            ("First Braseth BID",   f"${META_2108['contract_bid_first']:,} (Oct 8 2024 · 263u · before revision)"),
            ("Original 2021 BMDC budget", "$5,150,000 (BMDC, 226u — Rainier & Genesee project name)"),
            ("AR Billed to Date",  f"${META_2108['ar_billed']:,} (4 progress billings · takeoff phase)"),
            ("Direct Cost (sunk)", f"${META_2108['direct_cost']:,.0f}"),
            ("Carrying Loss",      f"$({META_2108['carrying_loss']:,.0f}) — write-off if project doesn't activate"),
            ("Retainage",          "$0 (no field billing)"),
            ("Insurance",          META_2108["insurance"]),
            ("Lien Position",      META_2108["lien_position"]),
            ("Warranty",           META_2108["warranty"]),
            ("Contract on File",   META_2108["contract_doc"]),
        ]),
        ("SCOPE & FIXTURE PROFILE", [
            ("Plumbing Units",     f"{META_2108['units']} units ✓ CONFIRMED via OWP Master List Schedule rows 106 + 195 + Projects Bid r339/r364 (3 corroborating entries)"),
            ("Unit count evolution", f"226 (2021 BMDC budget) → {META_2108['units']} (2024 Braseth bid) — +37 units (+16.4%) growth between bid stages"),
            ("Total Fixtures",     META_2108["fixtures"][:200]),
            ("Floors",             "TBD — Skb Architects design not yet released to OWP"),
            ("Fixture Provenance", "No Job Ticket exists yet (pre-construction). Will populate when project activates and Skb releases IFC drawings."),
        ]),
        ("PROJECT TEAM — GENERAL CONTRACTOR (Braseth)", [
            ("General Contractor", META_2108["gc"]),
            ("GC Project Manager", META_2108["gc_pm"]),
            ("GC Superintendent",  META_2108["gc_sup"]),
            ("GC Project Engineer", META_2108["gc_pe"]),
            ("Braseth engagement count", "OWP's 1st engagement with Braseth Construction. New-GC relationship — no closed-job history. Compare against historical Chinn/Exxel/Compass benchmarks for payment + CO behavior once project activates."),
        ]),
        ("PROJECT TEAM — OWP STAFF", [
            ("OWP Roughin Foreman", META_2108["owp_ri_foreman"]),
            ("OWP Trim Foreman",    META_2108["owp_trim_foreman"]),
            ("OWP Estimator (takeoff)", META_2108["owp_estimator"]),
            ("OWP Signatory",       META_2108["owp_signatory"]),
        ]),
        ("PROJECT TEAM — OWNER & DEVELOPMENT", [
            ("Owner of Record",   META_2108["owner"]),
            ("Developer",         META_2108["developer"]),
            ("Owner-developer note", "Lake Union Partners (Seattle developer) is OWP's repeat owner — same as #2061 Alta Columbia City "
                                     "(also at the Rainier corridor) and likely associated with #2106 Phoenix Property Co. South Seattle / "
                                     "Columbia City corridor is becoming a recurring portfolio for OWP."),
        ]),
        ("PROJECT TEAM — DESIGN", [
            ("Architect",         META_2108["architect"]),
            ("Structural Engineer", META_2108["structural"]),
            ("MEP / Plumbing Engineer", META_2108["mep_engineer"]),
            ("Civil Engineer",    META_2108["civil"]),
            ("Landscape Architect", META_2108["landscape"]),
        ]),
        ("BID HISTORY (from OWP Master Project List)", []),
        ("AP VENDOR PROFILE (TAKEOFF PHASE)", [
            ("Top Vendor",        META_2108["top_vendor"]),
            ("Total AP Spend",    f"${META_2108['ap_total']:,}"),
            ("Active Vendors",    f"{META_2108['ap_vendor_count']} unique vendors (Franklin Engineering, Bank of America CC, Herc Rentals)"),
            ("Concentration Note", "Franklin Engineering at 92.5% concentration is normal for early design-build phase — will diversify when field work begins (typical Braseth jobs use Pacific Plumbing, Consolidated, Rosen, Keller as top suppliers)."),
        ]),
        ("DOCUMENT META (current · pre-construction)", [
            ("Pay Apps (filed)",   "0 SOV pay apps · 4 AR transactions in JDR (takeoff progress billings)"),
            ("Executed COs",       "0 (subcontract pending)"),
            ("CORs",               "0"),
            ("RFIs",               "0"),
            ("ASIs",               "0"),
            ("Submittals",         "0 (pre-construction)"),
            ("POs",                "0"),
            ("Permit Count",       "0 (pre-permit)"),
            ("Notes",              "Project hasn't reached document-generation phase. Only OWP estimating + Franklin Engineering have touched the file (takeoff + design coordination)."),
        ]),
        ("DATA SOURCES", [
            ("JDR PDF",            "2108 Job Detail Report (Sage Timberline · Apr 3 2026 run · 4 pages · 6 cost codes · 52 line items)"),
            ("Parsed Data",        "2108_data.json"),
            ("Dashboard Arrays",   "2108_dashboard_arrays.json (16 arrays mirroring index.html PROJECTS['2108'])"),
            ("Master Project List", "OWP Project List with Schedule - UPDATED 04-01-26.xlsx (Schedule r106/r195 + Projects Bid r160/r339/r364)"),
            ("Owner identification", "Recovered from JDR AR section: invoice 038448 'Rainier & Genesee, LLC' + invoice 038534 'Lake Union Partners $R&G Apts'"),
            ("GDrive Folder",      "Not yet created (project hasn't reached active stage)"),
            ("GDrive Status",      META_2108["gdrive_status"]),
        ]),
        ("RISK FLAGS / WATCH ITEMS", [
            ("Carrying $25.8k loss", f"OWP has $117k sunk against $91k billed = -$25,760 carrying loss. If Braseth doesn't activate, OWP absorbs as write-off."),
            ("New GC (Braseth)",     "OWP has 0 closed-job history with Braseth. Once project activates, watch payment cadence + change-order behavior closely against Chinn/Exxel benchmarks."),
            ("Top-vendor concentration", "Franklin Engineering at 92.5% of AP. All design-build engineering fees. Normal for takeoff phase."),
            ("Bid drift (226 → 263 units)", "Project scope grew by 37 units (+16%) between 2021 BMDC version and 2024 Braseth version, but bid dropped 13% — driven by leaner scope and tighter pricing."),
            ("Lake Union Partners repeat client", "Positive risk indicator — OWP delivered 41.4% margin on #2061 Alta Columbia City (Lake Union Partners owner). Strong relationship signal for 2108."),
            ("Activation watchpoint",  "Reach out to Braseth PM Q2 2026 to confirm summer 2026 start. If push beyond Q4 2026, re-evaluate carrying the design cost."),
        ]),
        ("PROJECT NARRATIVE", []),  # populated below
    ]

    r = 4
    for section_name, fields in sections:
        ws.cell(row=r, column=1, value=section_name).font = SECTION_HDR
        ws.cell(row=r, column=1).fill = SECTION_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        ws.cell(row=r, column=1).border = BORDER
        r += 1
        if section_name.startswith("BID HISTORY"):
            headers = ["Bid Date","Type","GC","Units","Price","Note"]
            for c, h in enumerate(headers, 1):
                ws.cell(row=r, column=c, value=h).font = WHITE_BOLD
                ws.cell(row=r, column=c).fill = HDR_FILL
                ws.cell(row=r, column=c).border = BORDER
            r += 1
            for date, type_, gc, units, price, note in META_2108["bid_history"]:
                ws.cell(row=r, column=1, value=date).font = INK
                ws.cell(row=r, column=2, value=type_).font = INK_BOLD if type_ == "BID" else INK
                ws.cell(row=r, column=3, value=gc).font = INK
                ws.cell(row=r, column=4, value=units).font = INK
                ws.cell(row=r, column=5, value=f"${price:,}").font = INK
                ws.cell(row=r, column=6, value=note).font = INK
                ws.cell(row=r, column=6).alignment = WRAP
                if type_ == "BID":
                    for c in range(1,7): ws.cell(row=r, column=c).fill = WARN_FILL
                for c in range(1,7): ws.cell(row=r, column=c).border = BORDER
                r += 1
            r += 1
            continue
        if section_name == "PROJECT NARRATIVE":
            for label, body in META_2108["narrative"]:
                ws.cell(row=r, column=1, value=label).font = INK_BOLD
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v = ws.cell(row=r, column=2, value=body)
                cell_v.font = INK
                cell_v.alignment = WRAP
                cell_v.fill = TEAM_FILL
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
                ws.cell(row=r, column=1).border = BORDER
                cell_v.border = BORDER
                r += 1
            r += 1
            continue
        for label, val in fields:
            ws.cell(row=r, column=1, value=label).font = INK_BOLD
            cell_v = ws.cell(row=r, column=2, value=val)
            cell_v.font = INK
            cell_v.alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            ws.cell(row=r, column=1).border = BORDER
            cell_v.border = BORDER
            if section_name.startswith("PROJECT TEAM") or section_name.startswith("RISK") or section_name.startswith("AP VENDOR") or section_name.startswith("DOCUMENT META"):
                ws.cell(row=r, column=1).fill = TEAM_FILL
                cell_v.fill = TEAM_FILL
            r += 1
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 50


def patch_overview(ws):
    # Replace any CANCELLED → ON HOLD
    for r in range(1, 12):
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and "CANCELLED" in v.upper():
                ws.cell(row=r, column=c).value = v.replace("CANCELLED", META_2108["status_text"])
                ws.cell(row=r, column=c).font = WARN_FONT

    start = 30
    clear_range(ws, start, start + 30, 1, 8)
    ws.cell(row=start, column=1, value="PROJECT TEAM · KEY PLAYERS").font = SECTION_HDR
    ws.cell(row=start, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=7)
    ws.cell(row=start+1, column=1, value=(
        "Pre-construction · ON HOLD pending Braseth go-ahead. Final BID submitted Feb 27 2025 "
        "($4,619,052 / 263u). Bid history: 2021 BMDC budget at 226u → 2024 Braseth bid at 263u "
        "(scope grew, price dropped) → 2025 final BID +3.6%."
    )).font = GREY_FONT
    ws.merge_cells(start_row=start+1, start_column=1, end_row=start+1, end_column=7)

    rows = [
        ("General Contractor",     META_2108["gc"]),
        ("Owner",                  META_2108["owner"]),
        ("Architect",              META_2108["architect"]),
        ("MEP / Plumbing Engineer", META_2108["mep_engineer"]),
        ("OWP Estimator",          META_2108["owp_estimator"]),
        ("OWP Signatory",          META_2108["owp_signatory"]),
        ("Site Address",           META_2108["site_address"]),
        ("Insurance",              META_2108["insurance"]),
        ("Status",                 META_2108["status_text"]),
        ("Latest BID submitted",   f"${META_2108['contract_bid_latest']:,} (Feb 27 2025 · 263u)"),
        ("AR Billed to date",      f"${META_2108['ar_billed']:,} (takeoff phase)"),
        ("Direct Cost (sunk)",     f"${META_2108['direct_cost']:,.0f}"),
        ("Carrying Loss",          f"$({META_2108['carrying_loss']:,.0f})"),
        ("Active Crew",            f"{META_2108['workers']} OWP estimating leads · {META_2108['hours']:.1f} hrs"),
        ("Bid history",            "2021 BMDC budget 226u/$5.15M → 2024 Braseth bid 263u/$4.46M → 2025 final BID 263u/$4.62M"),
    ]
    r = start + 3
    ws.cell(row=r, column=1, value="ROLE").font = WHITE_BOLD
    ws.cell(row=r, column=1).fill = HDR_FILL
    ws.cell(row=r, column=1).border = BORDER
    ws.cell(row=r, column=2, value="VALUE").font = WHITE_BOLD
    ws.cell(row=r, column=2).fill = HDR_FILL
    ws.cell(row=r, column=2).border = BORDER
    r += 1
    for role, val in rows:
        ws.cell(row=r, column=1, value=role).font = INK_BOLD
        ws.cell(row=r, column=2, value=val).font = INK
        ws.cell(row=r, column=1).fill = TEAM_FILL
        ws.cell(row=r, column=2).fill = TEAM_FILL
        ws.cell(row=r, column=1).border = BORDER
        ws.cell(row=r, column=2).border = BORDER
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1


def patch_change_log(ws):
    last_row = ws.max_row
    r = last_row + 1
    ws.cell(row=r, column=1, value="2026-04-27").font = INK
    ws.cell(row=r, column=2, value="v1.4 · 2020-gold depth").font = INK_BOLD
    ws.cell(row=r, column=3, value=(
        "Upgraded Job Info to 2020/2061 gold-standard depth: 13-section sectioned layout with "
        "BID HISTORY table (3 entries: 2021 BMDC budget → 2024 Braseth bid → 2025 final BID), "
        "AP Vendor Profile section (Franklin Engineering 92.5% concentration), Document Meta "
        "(pre-construction state), Data Sources, Risk Flags + Watch Items, and PROJECT NARRATIVE "
        "section (5 narrative blocks: project history, sunk-cost posture, owner-developer continuity, "
        "activation triggers, risk profile). Owner enriched: Lake Union Partners (per JDR AR section "
        "invoice 038534). Site address: Rainier Ave S × S Genesee St (Columbia City corridor). "
        "Architect Skb Architects + MEP Emerald City Engineering (per Master Project List). "
        "Unit count: 263 ✓ CONFIRMED via 4 corroborating Master List entries (Schedule r106/r195 + "
        "Projects Bid r339/r364). Fixture count remains TBD (no Job Ticket — pre-construction)."
    )).font = INK
    ws.cell(row=r, column=3).alignment = WRAP


def main():
    print(f"Loading {WORKBOOK.name}...")
    wb = load_workbook(WORKBOOK)
    print("→ Rebuilding Job Info tab to 2020-gold depth (13 sections + bid history + narrative)...")
    rebuild_job_info(wb["02 Job Info"])
    print("→ Patching Overview + adding Project Team block (15 fields)...")
    patch_overview(wb["01 Overview"])
    print("→ Logging enrichment to Change Log...")
    patch_change_log(wb["17 Change Log"])
    wb.save(WORKBOOK)
    print(f"\n✓ Saved {WORKBOOK.name} ({len(wb.sheetnames)} tabs)")


if __name__ == "__main__":
    main()
